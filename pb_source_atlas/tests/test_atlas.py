# -*- coding: utf-8 -*-
"""NETROLE Phase 4 — the Source Atlas, over a miniature of a real pay run.

The scheme is deliberately the shape ABM's June run has: values arriving on a
spreadsheet header, values read off the contract, a constant nobody feeds, a
component nothing fed at all, and a formula chain that reaches NETPAY through an
aggregate. That is what makes the numbered cases mean something — a fixture
where every value came from the same place would pass an Atlas that only knew
one lane.

The provenance blobs are written DIRECTLY, exactly as the SOURCING resolver
writes them (``input_provenance.entry``), rather than by running a payroll: the
Atlas's contract is with the blob, and a test that recomputed a payslip would be
testing the resolver instead.
"""

import base64
import io
import json

from odoo.exceptions import AccessError
from odoo.tests import TransactionCase, tagged

from odoo.addons.pb_hr_payroll_formula.models import input_provenance


@tagged('post_install', '-at_install')
class TestSourceAtlas(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Atlas = cls.env['pb.source.atlas']

        Category = cls.env['hr.salary.rule.category']
        for code in ('BASIC', 'ALW', 'DED', 'NET', 'OTH'):
            if not Category.search([('code', '=', code)], limit=1):
                Category.create({'name': code.title(), 'code': code})
        other = Category.search([('code', '=', 'OTH')], limit=1)

        # ------------------------------------------------------- the scheme
        cls.config = cls.env['hr.formula.config'].create({
            'name': 'Atlas miniature', 'code': 'ATLASMINI',
            'country_code': 'VN',
        })
        spec = [
            ('A', 'BASESAL', 'Base Salary', 'input', ''),
            ('B', 'ALWMEAL', 'Meal Allowance', 'input', ''),
            ('C', 'OTHOURS', 'Overtime Hours', 'input', ''),
            ('D', 'SILEVY', 'Insurance Levy', 'constant', ''),
            ('E', 'GHOSTIN', 'Never Fed', 'input', ''),
            ('F', 'GROSSAG', 'Total Income', 'formula', '=A5+B5'),
            ('G', 'TAXAMT', 'Tax', 'formula', '=ROUND(F5*0.1,0)'),
            ('H', 'NETPAY', 'Net Pay', 'formula', '=F5-G5'),
        ]
        cls.rules = {}
        sequence = 10
        for letter, code, name, ctype, formula in spec:
            vals = {
                'config_id': cls.config.id, 'name': name, 'code': code,
                'column_type': ctype, 'sequence': sequence,
                'column_letter': letter, 'appears_on_payslip': True,
                'category_id': other.id,
            }
            if formula:
                vals['excel_formula'] = formula
            if ctype == 'constant':
                vals['constant_value'] = 105.0
            cls.rules[code] = cls.env['hr.formula.rule'].create(vals)
            sequence += 10
        cls.config.classify_net_roles()

        # ------------------------------------------------- people + the run
        cls.employees = cls.env['hr.employee']
        for index in range(4):
            cls.employees |= cls.env['hr.employee'].create({
                'name': 'Atlas Tester %s' % chr(ord('A') + index),
                'barcode': 'ATL%03d' % index,
            })
        cls.contracts = cls.env['hr.contract']
        for employee in cls.employees:
            cls.contracts |= cls.env['hr.contract'].create({
                'name': 'Contract %s' % employee.name,
                'employee_id': employee.id,
                'wage': 10000000.0,
                'state': 'open',
                'date_start': '2026-01-01',
            })

        cls.payrun = cls.env['hr.payslip.run'].create({
            'name': 'Atlas Run 2026-06',
            'date_start': '2026-06-01',
            'date_end': '2026-06-30',
        })

        # The spreadsheet the values came off, so the journey has a raw row to
        # name rather than a claim it cannot evidence.
        cls.batch = cls.env['hr.payroll.import.batch'].create({
            'name': 'Atlas June workbook',
            'source_type': 'excel',
            'formula_config_id': cls.config.id,
            'date_from': '2026-06-01',
            'date_to': '2026-06-30',
        })

        cls.slips = cls.env['hr.payslip']
        for index, employee in enumerate(cls.employees):
            contract = cls.contracts.filtered(lambda c: c.employee_id == employee)
            slip = cls.env['hr.payslip'].create({
                'name': 'Atlas slip %s' % employee.name,
                'employee_id': employee.id,
                'contract_id': contract.id,
                'payslip_run_id': cls.payrun.id,
                'date_from': '2026-06-01',
                'date_to': '2026-06-30',
                'formula_config_id': cls.config.id,
            })
            base = 10000000.0 + index * 1000000.0
            meal = 730000.0
            gross = base + meal
            tax = round(gross * 0.1)
            values = {
                'BASESAL': base, 'ALWMEAL': meal, 'OTHOURS': 8.0 + index,
                'SILEVY': 105.0, 'GHOSTIN': 0.0,
            }
            computed = dict(values, GROSSAG=gross, TAXAMT=tax, NETPAY=gross - tax)
            sources = {
                'BASESAL': input_provenance.entry(
                    'excel', key='Base_Salary', via='header'),
                'ALWMEAL': input_provenance.entry(
                    'contract_component', via='contract'),
                'OTHOURS': input_provenance.entry(
                    'feed', key='OTHRS150', via='connector_mapping'),
                'SILEVY': input_provenance.entry('constant', via='constant'),
                'GHOSTIN': input_provenance.entry('none', via='default'),
            }
            slip.write({
                'formula_input_values': json.dumps(values),
                'formula_computed_values': json.dumps(computed),
                'formula_input_sources': json.dumps(sources),
            })
            cls.env['hr.payroll.import.line'].create({
                'batch_id': cls.batch.id,
                'sequence': index + 1,
                'employee_id': employee.id,
                'payslip_id': slip.id,
                'employee_name': employee.name,
                'raw_data_json': json.dumps({
                    'Base_Salary': base, 'Meal': meal, 'OTHRS150': 8.0 + index,
                }),
            })
            cls.slips |= slip

        # The pre-SOURCING payslip: computed before origins were recorded.
        cls.legacy_slip = cls.env['hr.payslip'].create({
            'name': 'Atlas legacy slip',
            'employee_id': cls.employees[0].id,
            'contract_id': cls.contracts[0].id,
            'payslip_run_id': cls.payrun.id,
            'date_from': '2026-06-01',
            'date_to': '2026-06-30',
            'formula_config_id': cls.config.id,
            'formula_input_values': json.dumps({'BASESAL': 9000000.0}),
            'formula_computed_values': json.dumps({'BASESAL': 9000000.0}),
            'formula_input_sources': '',
        })

    # ==================================================================
    # helpers
    # ==================================================================
    def _row_counts(self):
        counts = {}
        for table in ('hr_payslip', 'hr_payslip_line', 'hr_api_data_store',
                      'hr_payroll_import_batch', 'hr_payroll_import_line',
                      'hr_formula_rule'):
            self.env.cr.execute('SELECT count(*) FROM %s' % table)
            counts[table] = self.env.cr.fetchone()[0]
        return counts

    def _attachment_from(self, action):
        self.assertEqual(action['type'], 'ir.actions.act_url')
        attachment_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        return self.env['ir.attachment'].browse(attachment_id)

    # ==================================================================
    # 1 — lanes count correctly, per the fixed vocabulary
    # ==================================================================
    def test_01_lanes_reflect_the_vocabulary(self):
        atlas = self.Atlas.get_run_atlas(self.payrun.id)
        lanes = {lane['key']: lane for lane in atlas['lanes']}

        # every lane in the vocabulary is on the map, used or not
        for key in input_provenance.SOURCES:
            self.assertIn(key, lanes, "lane %s is missing from the map" % key)

        self.assertEqual(lanes['excel']['components'], 1)
        self.assertEqual(lanes['excel']['employees'], 4)
        self.assertEqual(lanes['excel']['cells'], 4)
        self.assertEqual(lanes['contract_component']['cells'], 4)
        self.assertEqual(lanes['feed']['cells'], 4)
        self.assertEqual(lanes['none']['cells'], 4)

        # the constant lane holds BOTH the recorded entries and nothing else,
        # and the calculated lane is DECLARED by the rules' own column_type
        self.assertEqual(lanes['constant']['cells'], 4)
        self.assertEqual(lanes['calculated']['components'], 3)
        self.assertEqual(lanes['calculated']['employees'], 4)
        self.assertIsNone(lanes['calculated']['amount'],
                          "a calculated total would double-count the lanes it is made of")

        # a lane nobody used is muted, never hidden — absence is information
        self.assertTrue(lanes['rule']['muted'])
        self.assertEqual(lanes['rule']['cells'], 0)

        # money: the spreadsheet carried the four base salaries and nothing else
        self.assertAlmostEqual(
            lanes['excel']['amount'],
            sum(10000000.0 + i * 1000000.0 for i in range(4)), places=2)
        # …and overtime HOURS never enter a money total, even though the
        # classifier calls the component an earning
        self.assertEqual(lanes['feed']['amount'], 0.0)

        self.assertEqual(atlas['run']['slip_count'], 5)
        self.assertEqual(atlas['no_provenance_slips'], 1)

    # ==================================================================
    # 2 — grid windowing and search
    # ==================================================================
    def test_02_grid_windowing_and_search(self):
        first = self.Atlas.get_grid(self.payrun.id, offset=0, limit=2)
        self.assertEqual(first['total'], 5)
        self.assertEqual(len(first['rows']), 2)

        second = self.Atlas.get_grid(self.payrun.id, offset=2, limit=2)
        self.assertEqual(len(second['rows']), 2)
        self.assertFalse(
            {r['slip_id'] for r in first['rows']} & {r['slip_id'] for r in second['rows']},
            "a window must not repeat rows from the previous one")

        narrowed = self.Atlas.get_grid(self.payrun.id, search='Atlas Tester B')
        self.assertTrue(narrowed['total'] >= 1)
        for row in narrowed['rows']:
            self.assertIn('Tester B', row['employee'])

        # every cell knows its lane, and the declared lanes are present too
        cells = first['rows'][0]['cells']
        self.assertEqual(cells['BASESAL']['l'], 'excel')
        self.assertEqual(cells['BASESAL']['k'], 'Base_Salary')
        self.assertEqual(cells['ALWMEAL']['l'], 'contract_component')
        self.assertEqual(cells['NETPAY']['l'], 'calculated')
        self.assertEqual(cells['NETPAY']['kind'], 'declared')

        # a lane filter narrows the cells, never the truth of the ones left
        only_feed = self.Atlas.get_grid(self.payrun.id, limit=2, lane='feed')
        for row in only_feed['rows']:
            for cell in row['cells'].values():
                self.assertEqual(cell['l'], 'feed')

    # ==================================================================
    # 3 — a spreadsheet value: raw header key, then hops that end at NET
    # ==================================================================
    def test_03_journey_from_a_spreadsheet_header_to_net(self):
        slip = self.slips[0]
        journey = self.Atlas.get_journey(self.payrun.id, slip.id, 'BASESAL')

        self.assertEqual(journey['lane'], 'excel')
        self.assertEqual(journey['source']['key'], 'Base_Salary')
        self.assertEqual(journey['source']['raw_value'], 10000000.0)
        self.assertEqual(journey['source']['row']['batch'], self.batch.display_name)
        self.assertEqual(journey['source']['row']['row_no'], 1)

        codes = [hop['code'] for hop in journey['hops']]
        self.assertTrue(codes, "a base salary must reach net pay")
        self.assertEqual(codes[-1], 'NETPAY')
        self.assertIn('GROSSAG', codes)
        self.assertTrue(journey['hops'][-1]['is_net'])
        # gross adds, so the chain into net pay is additive the whole way
        self.assertEqual(journey['hops'][0]['sign'], 1)
        self.assertEqual(journey['hops'][-1]['cum_sign'], 1)

        # the tax hop is the same component read the other way round: it is
        # SUBTRACTED, and the sign badge must say so
        tax = self.Atlas.get_journey(self.payrun.id, slip.id, 'TAXAMT')
        self.assertEqual([h['code'] for h in tax['hops']], ['NETPAY'])
        self.assertEqual(tax['hops'][0]['sign'], -1)

    # ==================================================================
    # 4 — a fallback value invents nothing
    # ==================================================================
    def test_04_journey_for_a_fallback_is_honest(self):
        journey = self.Atlas.get_journey(self.payrun.id, self.slips[0].id, 'GHOSTIN')
        self.assertEqual(journey['lane'], 'none')
        self.assertEqual(journey['source']['via'], 'default')
        self.assertIn('default', journey['source']['via_label'])
        self.assertFalse(journey['source'].get('row'),
                         "a component nothing fed must not be given a source row")
        # it reaches no formula, so there are no hops and the screen says why
        self.assertEqual(journey['hops'], [])
        self.assertTrue(journey['warnings'])

    # ==================================================================
    # 5 — a payslip from before source tracking says so
    # ==================================================================
    def test_05_pre_sourcing_slip_says_so(self):
        journey = self.Atlas.get_journey(self.payrun.id, self.legacy_slip.id, 'BASESAL')
        self.assertTrue(journey['no_provenance'])
        self.assertEqual(journey['source']['kind'], 'untracked')
        self.assertIn('before', journey['source']['title'].lower())
        # the chain it CAN still read — the scheme's formulas — is still drawn
        self.assertEqual([h['code'] for h in journey['hops']][-1], 'NETPAY')

        grid = self.Atlas.get_grid(self.payrun.id, search='legacy')
        atlas = self.Atlas.get_run_atlas(self.payrun.id)
        self.assertEqual(atlas['no_provenance_slips'], 1)
        self.assertIsInstance(grid['rows'], list)

    # ==================================================================
    # 6 — every download opens, with the sheets it promised
    # ==================================================================
    def test_06_downloads_open_as_workbooks(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not available to verify the workbook")

        for lane in ('excel', 'feed', 'contract_component', 'constant',
                     'calculated', 'none', 'matrix'):
            with self.subTest(lane=lane):
                action = self.Atlas.download_lane(self.payrun.id, lane)
                attachment = self._attachment_from(action)
                self.assertTrue(attachment.name.endswith('.xlsx'))
                book = load_workbook(
                    io.BytesIO(base64.b64decode(attachment.datas)), read_only=True)
                self.assertIn('Values', book.sheetnames)
                self.assertIn('Sources', book.sheetnames)
                self.assertIn('How to read this', book.sheetnames)
                if lane in ('excel', 'matrix'):
                    self.assertIn('Atlas June workbook', book.sheetnames,
                                  "the spreadsheet lane must carry its raw rows")
                sheet = book['Values']
                headers = [c.value for c in next(sheet.iter_rows(max_row=1))]
                self.assertEqual(headers[:3],
                                 ['Employee code', 'Employee', 'Department'])
                book.close()

        # a lane nothing used refuses out loud rather than shipping an empty file
        from odoo.exceptions import UserError
        with self.assertRaises(UserError):
            self.Atlas.download_lane(self.payrun.id, 'rule')

    # ==================================================================
    # 7 — the Atlas writes nothing to payroll
    # ==================================================================
    def test_07_atlas_is_read_only(self):
        before = self._row_counts()
        self.Atlas.get_run_atlas(self.payrun.id)
        self.Atlas.get_grid(self.payrun.id, offset=0, limit=3, search='Atlas')
        self.Atlas.get_journey(self.payrun.id, self.slips[0].id, 'BASESAL')
        self.Atlas.get_journey(self.payrun.id, self.slips[0].id, 'NETPAY')
        self.Atlas.get_journey(self.payrun.id, self.slips[0].id, 'GHOSTIN')
        self.Atlas.download_lane(self.payrun.id, 'matrix')
        self.env.flush_all()
        self.assertEqual(self._row_counts(), before,
                         "the Atlas must not add, change or remove payroll rows")

        # and the values it reported are still the values on the payslip
        self.slips[0].invalidate_recordset()
        self.assertEqual(
            json.loads(self.slips[0].formula_input_values)['BASESAL'], 10000000.0)

    # ==================================================================
    # 8 — below the officer tier, the door is closed
    # ==================================================================
    def test_08_refuses_below_officer(self):
        # The outsider is BORROWED, never created. `payobook_template` has no
        # active member of base.group_system, so `res.users.create` there dies
        # on Odoo's own "you must have at least an administrator" constraint
        # long before it reaches anything of ours — a property of that database,
        # not of the Atlas. An existing user who holds none of the gate groups
        # proves exactly the same thing and works on every database.
        gate = [
            'pb_hr_payroll_base.group_payroll_base_officer',
            'pb_hr_payroll_base.group_payroll_base_manager',
            'pb_hr_payroll_base.group_payroll_super_admin',
            'om_hr_payroll.group_hr_payroll_manager',
        ]

        def gated(candidate):
            if candidate._is_superuser() or candidate._is_admin():
                return True
            return any(candidate.has_group(name) for name in gate)

        outsider = self.env['res.users'].search(
            [('active', '=', True)]).filtered(lambda u: not gated(u))[:1]
        if not outsider:
            outsider = self.env.ref('base.public_user')
        self.assertTrue(outsider, "no user below the officer tier to test with")
        atlas = self.env['pb.source.atlas'].with_user(outsider)
        with self.assertRaises(AccessError):
            atlas.get_run_atlas(self.payrun.id)
        with self.assertRaises(AccessError):
            atlas.get_grid(self.payrun.id)
        with self.assertRaises(AccessError):
            atlas.get_journey(self.payrun.id, self.slips[0].id, 'BASESAL')
        with self.assertRaises(AccessError):
            atlas.download_lane(self.payrun.id, 'matrix')
