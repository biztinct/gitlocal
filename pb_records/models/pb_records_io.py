# -*- coding: utf-8 -*-
"""RECORDS R3 — the desk as a file, and that file back as the desk.

Two methods, one idea: **the workbook is a detached view of the grid.** It
carries the same columns the desk shows, the same values in the same words a
person reads on screen (a selection's LABEL, `Yes`/`No`, a real date), and it
comes back through the same door — `preview_changes` to judge it and
`apply_changes` to write it, with the same whitelist, the same company scoping,
the same audit row and the same Undo.

Three rails that shape every line below:

  * **Import never writes.** `import_peek` reads a file and asks
    `preview_changes` what it would mean. Not one byte is written until the
    person presses Apply, and Apply is `apply_changes` — the R2 method,
    unchanged apart from being told the change arrived in a file.
  * **A file never creates anybody.** A row that matches no employee is LISTED
    with the reason, never turned into an employee, a contract, a batch or a
    payslip. Onboarding is the Import door's job; this is the Records Desk.
  * **A blank cell is not a blank value.** On the grid, emptying a cell is a
    gesture that means "clear this". In a file it means nothing at all — the
    column simply was not filled in — because the alternative is that dropping
    an empty template back wipes every mapped field of everybody in it. Blank
    cells are counted and reported, never staged.

Column identity is asked in three ways, best first, because a spreadsheet on a
person's laptop is a document they will edit: the header cell's COMMENT carries
`id: f:hr.contract:shuipart`; the hidden `_payobook` sheet carries the same map
by column position; and a heading that still reads like the card's label matches
by label. A column none of the three can place is IGNORED and named on screen,
never guessed at.
"""
import base64
import csv
import io
import logging
import re

from datetime import date, datetime

from odoo import _, api, models

from odoo.addons.pb_hr_payroll_formula.models.column_role_classifier import (
    EMPLOYEE_CODE_HEADER_CANDIDATES, EMPLOYEE_NAME_HEADER_CANDIDATES)

_logger = logging.getLogger(__name__)

EMP = 'hr.employee'
CON = 'hr.contract'

#: One export is at most this many people. The desk's own paging loop would
#: happily walk a 60,000-person roster into a 40 MB workbook nobody can open;
#: the cap is said out loud in the return and on screen, never silently.
MAX_ROWS = 10000
#: A file bigger than this is refused before it is parsed.
MAX_BYTES = 10 * 1024 * 1024

SHEET = 'Records'
#: The durable column map. Hidden, because it is machinery — but never the
#: PRIMARY key for anything: a person who deletes it still gets a working file
#: through the header comments and the labels.
META_SHEET = '_payobook'
XLSX_MIME = ('application/vnd.openxmlformats-officedocument'
             '.spreadsheetml.sheet')
#: The technical identity a header comment carries, e.g. `id: b:acc_number`.
ID_COMMENT = 'id: %s'
ID_RE = re.compile(r'^\s*id:\s*([^\s\r\n]+)', re.M)

#: The three identity columns, in the order somebody reads a name badge.
IDENTITY_COLS = (
    ('code', "Employee code"),
    ('name', "Name"),
    ('email', "Work email"),
)
EMAIL_HEADER_CANDIDATES = (
    'email', 'work email', 'work_email', 'e-mail', 'emp_email',
    'employee_email', 'employee email',
)
#: Employee fields that can hold "the code on the spreadsheet", in the order the
#: batch's own `_find_employee` ladder tries them. `identification_id` is on the
#: list because on ABM that is what a row actually matches (RD6) — the source
#: system's own code never reaches an employee from a plain spreadsheet.
CODE_FIELDS = ('barcode', 'employee_id', 'pb_source_ref', 'identification_id')


def _norm(text):
    """A heading, reduced to what two people would agree it says."""
    return re.sub(r'\s+', ' ', str(text if text is not None else '')).strip().lower()


class PbRecordsDeskIo(models.AbstractModel):
    """The export/import half of `pb.records.desk`.

    A separate FILE rather than more methods in `pb_records_desk.py` because the
    two halves answer different questions — that one is "what does this mapped
    field mean", this one is "what does this file say" — and the read/write
    contract between them is exactly the three public methods R2 already had.
    """
    _inherit = 'pb.records.desk'

    # =================================================================
    # Export
    # =================================================================
    @api.model
    def _io_hint(self, card):
        """The plain sentence under a column's technical id, in its comment.

        Every column says what it will accept BEFORE somebody types into it,
        because the alternative is finding out at the Apply step, forty rows
        later, in a list of refusals.
        """
        ttype = card.get('ttype')
        if ttype == 'selection':
            labels = [p['label'] for p in (card.get('selection') or [])]
            if labels:
                return _("One of: %s") % ", ".join(labels)
            return _("One of this field's choices.")
        if ttype == 'boolean':
            return _("Yes or No.")
        if ttype == 'many2one':
            m2o = card.get('m2o') or {}
            if m2o.get('creates_missing'):
                return _("A name. One that does not exist yet is created.")
            return _("A name that already exists — new ones are not created "
                     "from a file.")
        if ttype in ('date', 'datetime'):
            return _("A date, like 2026-08-29.")
        if ttype in ('integer', 'float', 'monetary', 'amount'):
            return _("A number, like 1500000.")
        if ttype == 'bank':
            return _("Part of the bank account. It is only saved when an "
                     "account number results.")
        return _("Text.")

    @api.model
    def _io_choices(self, card):
        """The dropdown a selection or a boolean column offers, or `None`.

        Excel's list validation is a FORMULA, and a formula is capped at 255
        characters. A longer list is not truncated (a dropdown missing half its
        choices is worse than no dropdown) — the comment carries them instead.
        """
        if card.get('ttype') == 'boolean':
            return [_("Yes"), _("No")]
        if card.get('ttype') != 'selection':
            return None
        labels = [p['label'] for p in (card.get('selection') or []) if p['label']]
        if not labels:
            return None
        joined = ",".join(labels)
        if len(joined) > 250 or any(',' in lb for lb in labels):
            return None
        return labels

    @api.model
    def _io_export_value(self, probe, card, employee, contract):
        """One cell, as a spreadsheet should hold it.

        Numbers stay numbers and dates stay DATES — a date written as text is a
        date Excel will not sort, filter or reformat, and it is the single most
        common way a round trip comes back as a string nobody can parse.
        Everything else is written exactly as the grid shows it, which is what
        makes the file readable and the round trip lossless at the same time.
        """
        cell = self._read_cell(probe, card, employee, contract)
        if cell.get('missing'):
            return None, None
        kind = card['id'][:1]
        ttype = card.get('ttype')
        if kind == 'f' and ttype in ('date', 'datetime'):
            record = employee if card['model'] == EMP else contract
            raw = record[card['field']] if record else None
            if not raw:
                return None, None
            if isinstance(raw, datetime):
                return raw.date(), 'yyyy-mm-dd'
            if isinstance(raw, date):
                return raw, 'yyyy-mm-dd'
            return str(raw), None
        if ttype == 'boolean':
            return (_("Yes") if cell.get('v') else _("No")), None
        if ttype in ('integer', 'float', 'monetary', 'amount'):
            raw = cell.get('v')
            if not isinstance(raw, (int, float)) or isinstance(raw, bool):
                return None, None
            if not cell.get('label'):
                # The grid shows a contract component that holds nothing as an
                # EMPTY cell rather than as `0`, and the file is a detached view
                # of the grid — so it shows the same thing. (A mapped numeric
                # FIELD at zero still reads `0` in both, because `_num(0)` is
                # `'0'`: zero is a value there, MJ15.)
                return None, None
            if ttype == 'integer':
                return int(raw), '#,##0'
            return float(raw), '#,##0.##'
        label = cell.get('label')
        return (label if label not in (None, '') else None), None

    @api.model
    def _io_contracts(self, employees):
        """`{employee_id: contract}` for the whole export — ONE query.

        `_get_latest_contract` is the definition of "this person's contract"
        and stays so everywhere else. It is not called PER PERSON here for two
        measured reasons, both RD11's rule (a) in a new costume: it sorts an
        employee's `contract_ids` in Python, and it writes an INFO line naming
        the candidates — 4,533 of each on payobook's roster, which is most of a
        four-minute export.

        The answer is identical. With no dates on the probe that method reduces
        to `contract_ids.sorted(key=date_start or date.min, reverse=True)[0]`,
        and the winner is picked here with that same key — in PYTHON, not in the
        `ORDER BY`, because PostgreSQL sorts NULLs LAST ascending while that
        method sorts a date-less contract FIRST, and the two disagree exactly
        for the person whose contract has no start date.
        """
        if not employees:
            return {}
        floor = date.min
        rows = self.env[CON].sudo().search_read(
            [('employee_id', 'in', employees.ids)],
            ['employee_id', 'date_start'], order='id asc')
        best = {}
        for row in rows:
            if not row.get('employee_id'):
                continue
            emp_id = row['employee_id'][0]
            key = (row.get('date_start') or floor, row['id'])
            if emp_id not in best or key > best[emp_id][0]:
                best[emp_id] = (key, row['id'])
        latest = {emp_id: entry[1] for emp_id, entry in best.items()}
        found = self.env[CON].sudo().browse(list(latest.values()))
        found.mapped('id')          # one prefetch pass for the whole set
        by_id = {c.id: c for c in found}
        return {emp_id: by_id[con_id] for emp_id, con_id in latest.items()
                if con_id in by_id}

    @api.model
    def _io_filename(self, config_id):
        configs = self._configs(config_id)
        code = ''
        if len(configs) == 1:
            code = (configs.code or configs.name or '').strip()
        stem = re.sub(r'[^A-Za-z0-9]+', '_', code).strip('_') or 'records'
        return '%s_records_%s.xlsx' % (stem, date.today().isoformat())

    @api.model
    def export_records(self, config_id=0, filters=None, field_ids=None,
                       mode='data'):
        """The current desk view as an `.xlsx` — with the values, or blank.

        Same scoping as `search_people`, on purpose and not by coincidence: the
        export reuses `_matching`, so a file can never contain a person the desk
        would not have shown.
        """
        self._check_read()
        try:
            import openpyxl
            from openpyxl.comments import Comment
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
        except Exception:       # noqa: BLE001 — a missing library, not a bug
            return {'ok': False, 'msg': _(
                "Spreadsheets cannot be written on this server yet. Ask an "
                "administrator to install the spreadsheet library.")}

        cards = self._cards(config_id)
        picked = [cards[fid] for fid in (field_ids or []) if fid in cards]
        if not picked:
            return {'ok': False, 'msg': _(
                "Pick at least one field first — a file needs a column to "
                "carry.")}

        employees = self._matching(filters or {}, ctx={})
        total = len(employees)
        truncated = total > MAX_ROWS
        employees = employees[:MAX_ROWS]
        contracts = self._io_contracts(employees)
        if any(c['id'][:1] == 'b' for c in picked):
            # The bank read walks `employee.bank_account_ids`; one read of the
            # whole o2m, not one per person.
            employees.mapped('bank_account_ids')
        probe = self._probe(config_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET
        bold = Font(bold=True, color='241F52')
        id_fill = PatternFill('solid', fgColor='EEF2F6')
        head_fill = PatternFill('solid', fgColor='EAF1FB')
        wrap = Alignment(vertical='center')

        for idx, (_key, label) in enumerate(IDENTITY_COLS, start=1):
            cell = ws.cell(row=1, column=idx, value=_(label))
            cell.font = bold
            cell.fill = id_fill
            cell.alignment = wrap
            cell.comment = Comment(
                text=_("Identity — used to match the row to a person. It is "
                       "never imported."), author='Payobook')
            ws.column_dimensions[get_column_letter(idx)].width = \
                max(16, min(38, len(label) + 6))

        first_data_col = len(IDENTITY_COLS) + 1
        for offset, card in enumerate(picked):
            col = first_data_col + offset
            cell = ws.cell(row=1, column=col, value=card['label'])
            cell.font = bold
            cell.fill = head_fill
            cell.alignment = wrap
            cell.comment = Comment(
                text='%s\n%s' % (ID_COMMENT % card['id'], self._io_hint(card)),
                author='Payobook')
            ws.column_dimensions[get_column_letter(col)].width = \
                max(14, min(40, len(card['label'] or '') + 6))
            choices = self._io_choices(card)
            if choices:
                dv = DataValidation(
                    type='list', formula1='"%s"' % ",".join(choices),
                    allow_blank=True, showDropDown=False)
                dv.error = _("Pick one of the listed values.")
                dv.errorTitle = _("Not one of the choices")
                ws.add_data_validation(dv)
                letter = get_column_letter(col)
                dv.add('%s2:%s%s' % (letter, letter,
                                     max(2, len(employees) + 1)))

        empty_contract = self.env[CON]
        for r, emp in enumerate(employees, start=2):
            contract = contracts.get(emp.id) or empty_contract
            ws.cell(row=r, column=1, value=emp.barcode or emp.employee_id or '')
            ws.cell(row=r, column=2, value=emp.display_name or emp.name or '')
            ws.cell(row=r, column=3, value=emp.work_email or '')
            if mode == 'template':
                continue
            for offset, card in enumerate(picked):
                value, numfmt = self._io_export_value(probe, card, emp, contract)
                if value is None:
                    continue
                cell = ws.cell(row=r, column=first_data_col + offset, value=value)
                if numfmt:
                    cell.number_format = numfmt

        ws.freeze_panes = 'D2'
        ws.auto_filter.ref = 'A1:%s%s' % (
            get_column_letter(first_data_col + len(picked) - 1),
            max(1, len(employees) + 1))

        meta = wb.create_sheet(title=META_SHEET)
        meta['A1'] = 'config_id'
        meta['B1'] = int(config_id or 0)
        meta['A2'] = 'column_index'
        meta['B2'] = 'field_id'
        meta['C2'] = 'label'
        for offset, card in enumerate(picked):
            meta.cell(row=3 + offset, column=1, value=first_data_col + offset)
            meta.cell(row=3 + offset, column=2, value=card['id'])
            meta.cell(row=3 + offset, column=3, value=card['label'])
        meta.sheet_state = 'hidden'

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': self._io_filename(config_id),
            'mimetype': XLSX_MIME,
            'rows': len(employees),
            'total': total,
            'columns': len(picked),
            'truncated': truncated,
            'mode': 'template' if mode == 'template' else 'data',
        }

    # =================================================================
    # Import — reading the file
    # =================================================================
    @api.model
    def _io_read_xlsx(self, raw):
        """`(rows, comments, meta)` from an `.xlsx`.

        NOT `read_only=True`. A read-only worksheet hands back `ReadOnlyCell`s,
        and a `ReadOnlyCell` has no `.comment` — which is where the column's
        technical identity lives, and therefore the whole reason a retyped
        heading still lands on the right field. The 10 MB size guard is what
        keeps a full load affordable.
        """
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        meta = {'config_id': 0, 'by_index': {}}
        if META_SHEET in wb.sheetnames:
            ms = wb[META_SHEET]
            if _norm(ms['A1'].value) == 'config_id':
                try:
                    meta['config_id'] = int(ms['B1'].value or 0)
                except (TypeError, ValueError):
                    meta['config_id'] = 0
            for row in ms.iter_rows(min_row=3, values_only=True):
                if not row or row[0] in (None, ''):
                    continue
                try:
                    meta['by_index'][int(row[0])] = str(row[1] or '')
                except (TypeError, ValueError):
                    continue
        sheet = None
        if SHEET in wb.sheetnames and wb[SHEET].sheet_state == 'visible':
            sheet = wb[SHEET]
        else:
            for name in wb.sheetnames:
                if name == META_SHEET:
                    continue
                if wb[name].sheet_state == 'visible':
                    sheet = wb[name]
                    break
        if sheet is None:
            return [], {}, meta
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        comments = {}
        if rows:
            for idx in range(1, len(rows[0]) + 1):
                cell = sheet.cell(row=1, column=idx)
                if cell.comment is not None and cell.comment.text:
                    comments[idx] = cell.comment.text
        return rows, comments, meta

    @api.model
    def _io_read_csv(self, raw):
        text = None
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return []
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except csv.Error:
            dialect = csv.excel
        return [list(r) for r in csv.reader(io.StringIO(text), dialect)]

    @api.model
    def _io_columns(self, header, comments, meta, cards):
        """Which field each column carries, asked five ways, best first.

            1. the header cell's COMMENT (`id: f:hr.contract:shuipart`)
            2. the hidden `_payobook` sheet, by column position
            3. one of the THREE identity headings this desk itself writes
            4. a heading that still reads like a card's label
            5. the batch's own looser identity spellings (`MSNV`, `emp code`…)

        A column none of the five can place is IGNORED, and named on screen.

        Why 3 comes before 4: a scheme may MAP the very field an identity
        column carries — ABM maps `work_email`, whose card is labelled "Work
        Email", and the identity column this desk writes is headed "Work
        email". With the label check first, that column was read as a
        destination, its comment ("Identity — … It is never imported") became
        untrue, and the file's own email column counted as a change. Identity
        is the safe reading: it can only ever MATCH a row, never write one.
        An exported data column is unaffected — its comment answers at step 1,
        whatever its heading says.
        """
        used, ignored, identity_cols = {}, [], {}
        by_label = {}
        for card in cards.values():
            by_label.setdefault(_norm(card['label']), card['id'])
        id_labels = {_norm(lb): key for key, lb in IDENTITY_COLS}
        code_headers = {_norm(h) for h in EMPLOYEE_CODE_HEADER_CANDIDATES}
        name_headers = {_norm(h) for h in EMPLOYEE_NAME_HEADER_CANDIDATES}
        email_headers = {_norm(h) for h in EMAIL_HEADER_CANDIDATES}

        for idx, raw_head in enumerate(header, start=1):
            head = _norm(raw_head)
            comment = comments.get(idx) or ''
            match = ID_RE.search(comment) if comment else None
            field_id = match.group(1).strip() if match else ''
            if field_id and field_id in cards:
                used[idx] = field_id
                continue
            from_meta = meta.get('by_index', {}).get(idx)
            if from_meta and from_meta in cards:
                used[idx] = from_meta
                continue
            if not head:
                continue
            key = id_labels.get(head)
            if key and key not in identity_cols:
                identity_cols[key] = idx
                continue
            if key:
                continue
            if head in by_label:
                used[idx] = by_label[head]
                continue
            if head in code_headers:
                key = 'code'
            elif head in name_headers:
                key = 'name'
            elif head in email_headers:
                key = 'email'
            if key and key not in identity_cols:
                identity_cols[key] = idx
                continue
            if key:
                continue
            ignored.append(str(raw_head).strip())
        return used, identity_cols, ignored

    @api.model
    def _io_cell(self, row, idx):
        if not idx or idx > len(row):
            return None
        return row[idx - 1]

    @api.model
    def _io_text(self, value):
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @api.model
    def _io_index(self, employees):
        """Every way a row could name a person, built once for the whole file.

        A per-row `search` over a 4,500-person roster is RD11 in a different
        costume; the whole scoped roster is read once into three dicts, and a
        key claimed by two people maps to a LIST so "which of these two?" is
        answerable rather than a coin toss.
        """
        by_code, by_email, by_name = {}, {}, {}
        fields_present = [f for f in CODE_FIELDS if f in employees._fields]
        rows = employees.read(['name', 'work_email'] + fields_present,
                              load=False) if employees else []
        def claim(bucket, key, emp_id):
            # One person carrying the same string in two of the code fields
            # (a badge id that IS the id-card number, which is ABM's own shape)
            # must not read as two people sharing a code.
            holders = bucket.setdefault(key, [])
            if emp_id not in holders:
                holders.append(emp_id)

        for row in rows:
            for fname in fields_present:
                code = self._normalise_code(row.get(fname))
                if code:
                    claim(by_code, code, row['id'])
            email = (row.get('work_email') or '').strip().lower()
            if email:
                claim(by_email, email, row['id'])
            name = _norm(row.get('name'))
            if name:
                claim(by_name, name, row['id'])
        return {'code': by_code, 'email': by_email, 'name': by_name}

    @api.model
    def _normalise_code(self, value):
        """The batch's `_normalize_code`, spelled without a probe.

        A code is compared case-insensitively here because a spreadsheet is
        typed by hand and `abc-1` and `ABC-1` are the same badge; the batch
        compares case-sensitively only because it is comparing two machine
        strings.
        """
        if value in (None, False, ''):
            return ''
        if isinstance(value, bool):
            return ''
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip().lower()

    @api.model
    def _io_match(self, index, code, name, email):
        """`(employee_id, method, why)` — the ladder, and its refusals in words."""
        if code:
            hits = index['code'].get(self._normalise_code(code)) or []
            if len(hits) == 1:
                return hits[0], 'code', ''
            if len(hits) > 1:
                return None, '', _(
                    "%(n)s people carry the code %(code)s — this row cannot "
                    "say which one.") % {'n': len(hits), 'code': code}
        if email:
            hits = index['email'].get(email.strip().lower()) or []
            if len(hits) == 1:
                return hits[0], 'email', ''
            if len(hits) > 1:
                return None, '', _(
                    "%(n)s people use the email %(email)s — add the employee "
                    "code.") % {'n': len(hits), 'email': email}
        if name:
            hits = index['name'].get(_norm(name)) or []
            if len(hits) == 1:
                return hits[0], 'name', ''
            if len(hits) > 1:
                return None, '', _(
                    "%(n)s people called %(name)s — add the employee code.") % {
                        'n': len(hits), 'name': name}
        if not (code or email or name):
            return None, '', _("This row names nobody — it has no employee "
                               "code, name or email.")
        return None, '', _("Nobody here matches %s.") % (
            code or email or name)

    @api.model
    def _io_probe_guard(self, file_b64, filename):
        """`(raw, is_csv, refusal)` — the three things both readers need first.

        Shared by the probe and the peek so a file refused by one is refused by
        the other in the same words. A sentence that changes between two calls
        about the same file is a sentence nobody trusts.
        """
        name = (filename or '').strip().lower()
        try:
            raw = base64.b64decode(file_b64 or '')
        except Exception:       # noqa: BLE001 — a malformed upload
            raw = b''
        if not raw:
            return b'', False, _("That file is empty.")
        if len(raw) > MAX_BYTES:
            return b'', False, _(
                "That file is larger than 10 MB. Export a smaller slice — "
                "filter to one department and try again.")
        if name.endswith('.csv'):
            return raw, True, ''
        if not name.endswith(('.xlsx', '.xlsm')):
            return b'', False, _(
                "This is not a spreadsheet. Drop an .xlsx or a .csv file.")
        return raw, False, ''

    @api.model
    def import_probe(self, file_b64='', filename=''):
        """How many rows this file holds. Nothing else, and nothing written.

        R4 D6. `import_peek` parses, indexes the whole roster, matches every
        row and previews every value — on a 4,500-row file that is a wait, and
        a busy veil that only spins does not say whether it is a wait of one
        second or thirty. This is the cheap half of that work, split out so the
        veil can say the size of the job: *"Matching 4,512 rows to people…"*.

        `read_only=True` here and NOT in `_io_read_xlsx` (RD20), because the
        difference between the two is exactly the cell COMMENTS — which a row
        count does not need and a column identity cannot live without.

        Advisory by construction: every failure returns `{'ok': False}` with a
        reason and the client carries on to the peek, which will refuse the
        same file in the same words. A count is a courtesy, never a gate.
        """
        self._check_read()
        raw, is_csv, refusal = self._io_probe_guard(file_b64, filename)
        if refusal:
            return {'ok': False, 'msg': refusal, 'rows': 0}
        cap = MAX_ROWS + 2
        try:
            if is_csv:
                rows = self._io_probe_csv(raw, cap)
            else:
                rows = self._io_probe_xlsx(raw, cap)
        except Exception as err:       # noqa: BLE001 — a file, not a bug
            _logger.info("Records Desk: could not count rows in %s (%s)",
                         filename, err)
            return {'ok': False, 'rows': 0, 'msg': _(
                "That file could not be opened as a spreadsheet.")}
        # The heading row is not a row of data, and the count on screen is a
        # count of people, not of lines in a file.
        return {'ok': True, 'rows': max(0, rows - 1),
                'truncated': rows >= cap}

    @api.model
    def _io_probe_xlsx(self, raw, cap):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True,
                                    data_only=True)
        try:
            sheet = None
            for name in wb.sheetnames:
                if name == META_SHEET:
                    continue
                if wb[name].sheet_state == 'visible':
                    sheet = wb[name]
                    break
            if sheet is None:
                return 0
            seen = 0
            for row in sheet.iter_rows(values_only=True):
                if row and any(v not in (None, '') for v in row):
                    seen += 1
                    if seen >= cap:
                        break
            return seen
        finally:
            wb.close()

    @api.model
    def _io_probe_csv(self, raw, cap):
        seen = 0
        for row in self._io_read_csv(raw):
            if row and any(v not in (None, '') for v in row):
                seen += 1
                if seen >= cap:
                    break
        return seen

    @api.model
    def import_peek(self, config_id=0, file_b64='', filename=''):
        """Read a file and say what it WOULD do. Writes nothing, ever.

        Every failure is a sentence a person can act on, returned as
        `{'ok': False, 'msg': …}` — the `import_test_samples` contract
        (`pb_formula_studio.py:12197`): a bad file is a thing that happens, not
        an exception for somebody else to catch.
        """
        self._check_read()
        name = (filename or '').strip()
        raw, is_csv, refusal = self._io_probe_guard(file_b64, name)
        if refusal:
            return {'ok': False, 'msg': refusal}

        meta = {'config_id': 0, 'by_index': {}}
        comments = {}
        try:
            if is_csv:
                rows = self._io_read_csv(raw)
            else:
                rows, comments, meta = self._io_read_xlsx(raw)
        except Exception as err:       # noqa: BLE001 — a file, not a bug
            _logger.info("Records Desk: unreadable import file %s (%s)",
                         name, err)
            return {'ok': False, 'msg': _(
                "That file could not be opened as a spreadsheet. If it was "
                "renamed, save it again as .xlsx or .csv.")}

        # Blank rows are dropped, but their POSITION is kept: "Row 7" on screen
        # has to be row 7 in the file the person is looking at, or the sentence
        # sends them to the wrong line.
        rows = [(i, r) for i, r in enumerate(rows, start=1)
                if r and any(v not in (None, '') for v in r)]
        if not rows:
            return {'ok': False, 'msg': _("This file has no heading row.")}
        header = rows[0][1]
        if not any(_norm(h) for h in header):
            return {'ok': False, 'msg': _("This file has no heading row.")}

        cards = self._cards(config_id)
        used, identity_cols, ignored = self._io_columns(
            header, comments, meta, cards)
        if not used:
            return {'ok': False, 'msg': _(
                "None of the headings match a field this pay scheme maps — "
                "export a file from this desk first, fill it in and drop it "
                "back."), 'columns_ignored': ignored}
        if not identity_cols:
            return {'ok': False, 'msg': _(
                "No employee code, name or email column was recognised, so no "
                "row can be matched to a person.")}

        wrong_scheme = bool(meta.get('config_id')) and bool(config_id) \
            and int(meta['config_id']) != int(config_id)

        body = rows[1:]
        over_cap = len(body) > MAX_ROWS
        body = body[:MAX_ROWS]
        if not body:
            return {'ok': True, 'empty': True, 'msg': _(
                "This file has headings only — fill it in and drop it again."),
                'summary': {'rows': 0, 'people_matched': 0,
                            'people_unmatched': 0, 'changes_ok': 0,
                            'changes_same': 0, 'changes_refused': 0,
                            'cells_blank': 0,
                            'columns_used': len(used),
                            'columns_ignored': ignored},
                'changes': [], 'unmatched': [], 'items': [],
                'identity': '', 'filename': name, 'truncated': False,
                'wrong_scheme': wrong_scheme}

        employees = self._matching({}, ctx={})
        index = self._io_index(employees)

        changes, unmatched = [], []
        matched_ids, blank = set(), 0
        methods = {}
        for offset, row in body:
            code = self._io_text(self._io_cell(row, identity_cols.get('code')))
            person = self._io_text(self._io_cell(row, identity_cols.get('name')))
            email = self._io_text(self._io_cell(row, identity_cols.get('email')))
            values = {}
            for idx, field_id in used.items():
                raw_value = self._io_cell(row, idx)
                text = raw_value if isinstance(raw_value, (int, float)) \
                    and not isinstance(raw_value, bool) else \
                    self._io_text(raw_value)
                if text in (None, ''):
                    blank += 1
                    continue
                values[field_id] = text
            emp_id, method, why = self._io_match(index, code, person, email)
            if not emp_id:
                unmatched.append({
                    'row': offset, 'code': code, 'name': person,
                    'email': email, 'why': why, 'values': values,
                })
                continue
            matched_ids.add(emp_id)
            methods[method] = methods.get(method, 0) + 1
            for field_id, value in values.items():
                changes.append({'emp_id': emp_id, 'field_id': field_id,
                                'value': value})

        preview = self.preview_changes(config_id, changes) if changes else \
            {'items': [], 'counts': {'ok': 0, 'same': 0, 'refused': 0,
                                     'people': 0}}
        counts = preview['counts']
        identity = max(methods, key=methods.get) if methods else ''
        return {
            'ok': True,
            'empty': False,
            'filename': name,
            'wrong_scheme': wrong_scheme,
            'identity': identity,
            'truncated': over_cap,
            'summary': {
                'rows': len(body),
                'people_matched': len(matched_ids),
                'people_unmatched': len(unmatched),
                'changes_ok': counts['ok'],
                'changes_same': counts['same'],
                'changes_refused': counts['refused'],
                'people_changed': counts['people'],
                'cells_blank': blank,
                'columns_used': len(used),
                'columns_ignored': ignored,
            },
            'changes': changes,
            'items': preview['items'],
            'unmatched': unmatched,
        }

    # =================================================================
    # Binding an unmatched row by hand
    # =================================================================
    @api.model
    def lookup_people(self, term='', limit=10):
        """The typeahead behind "Find person…" on an unmatched row.

        Scoped exactly like the grid — `_people_domain` — so a row can only be
        bound to somebody the person doing the binding can already see.
        """
        self._check_read()
        Employee = self.env[EMP].sudo()
        domain = self._people_domain({'q': (term or '').strip()})
        records = Employee.search(domain, limit=int(limit), order='name, id')
        out = []
        for emp in records:
            bits = [emp.barcode or emp.employee_id or '',
                    emp.department_id.display_name or '']
            tail = " · ".join([b for b in bits if b])
            out.append({'id': emp.id,
                        'label': "%s%s" % (emp.display_name or emp.name or '',
                                           (" — %s" % tail) if tail else '')})
        return out
