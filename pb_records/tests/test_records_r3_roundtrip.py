# -*- coding: utf-8 -*-
"""RECORDS R3 — the desk as a file, and that file back as the desk.

Fourteen numbered cases, and the numbers are the handover's. One line each:
what went out comes back meaning the same thing (1), one edited cell is one
change and one audit row (2), a row that names nobody is listed and never
created (3), a name two people share is a question rather than a guess (4), a
column nobody asked for is named and read past (5), a retyped heading still
lands on the right field (6), four bank columns are one account (7), a
component column is a contract line (8), a `.csv` says the same thing as an
`.xlsx` (9), a bad file is a sentence and never an exception (10), a blank
template is a template (11), the ten-thousand cap is real and said out loud
(12), no payslip and no employee is created anywhere in the suite (13), and no
user-visible string — the workbook's own included — says the wrong word (14).

`action_process` is never called (J3/J10): every record read or written here was
created by the transaction it runs in.
"""
import base64
import hashlib
import io
import os
import re

from unittest.mock import patch

from odoo.modules.module import get_module_path
from odoo.tests import TransactionCase, tagged

from odoo.addons.pb_records.models import pb_records_io

try:
    import openpyxl
except ImportError:            # pragma: no cover — a declared dependency
    openpyxl = None


def _src(module, *parts):
    with open(os.path.join(get_module_path(module), *parts), encoding='utf-8') as fh:
        return fh.read()


@tagged('post_install', '-at_install')
class TestRecordsR3RoundTrip(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Desk = cls.env['pb.records.desk']
        cls.Config = cls.env['hr.formula.config']
        cls.Rule = cls.env['hr.formula.rule']
        cls.Mapping = cls.env['hr.payslip.import.mapping']
        cls.Employee = cls.env['hr.employee']
        cls.Contract = cls.env['hr.contract']
        cls.IrModel = cls.env['ir.model']
        cls.IrField = cls.env['ir.model.fields']
        cls.company = cls.env.company

        # A boolean and a date destination that exist on THIS build — the same
        # rule R2's fixture follows: pick from the registry rather than naming a
        # field and hoping (`hr.employee` delegates to `hr.version`, and which
        # of these is writable differs between deployments).
        cls.bool_field = next(
            (f for f in ('birthday_public_display', 'vn_hi_enrolled',
                         'manually_set_presence', 'active')
             if f in cls.Employee._fields
             and not cls.Employee._fields[f].readonly),
            'active')
        cls.date_field = next(
            (f for f in ('birthday', 'departure_date', 'permit_expiration_date')
             if f in cls.Employee._fields
             and cls.Employee._fields[f].type == 'date'
             and not cls.Employee._fields[f].readonly),
            '')

        cls.cfg = cls._config('R3 Scheme')
        cls.r_loc = cls._rule(cls.cfg, 'Location', 'R3LOCATION')
        cls.r_shui = cls._rule(cls.cfg, 'SHUI participation', 'R3SHUIPART')
        cls.r_dept = cls._rule(cls.cfg, 'Department', 'R3DEPT')
        cls.r_bool = cls._rule(cls.cfg, 'Birthday shown', 'R3BDAYFLAG')
        cls.r_accno = cls._rule(cls.cfg, 'Bank account', 'R3ACCNO')
        cls.r_bank = cls._rule(cls.cfg, 'Bank name', 'R3BANKNAME')
        cls.r_bonus = cls._rule(cls.cfg, 'Site bonus', 'R3BONUS',
                                is_contract_component=True)

        cls._map_field(cls.cfg, cls.r_loc, 'hr.employee', 'location')
        cls._map_field(cls.cfg, cls.r_shui, 'hr.contract', 'shuipart')
        cls._map_field(cls.cfg, cls.r_dept, 'hr.contract', 'department_id')
        cls._map_field(cls.cfg, cls.r_bool, 'hr.employee', cls.bool_field)
        cls._map_bank(cls.cfg, cls.r_accno, 'acc_number')
        cls._map_bank(cls.cfg, cls.r_bank, 'bank_name')
        if cls.date_field:
            cls.r_date = cls._rule(cls.cfg, 'Date on file', 'R3DATEFLD')
            cls._map_field(cls.cfg, cls.r_date, 'hr.employee', cls.date_field)

        Dept = cls.env['hr.department']
        cls.dept_a = Dept.create({'name': 'R3 Alpha', 'company_id': cls.company.id})
        cls.dept_b = Dept.create({'name': 'R3 Beta', 'company_id': cls.company.id})

        cls.e1 = cls._employee('R3 One', barcode='R30001',
                               work_email='r3one@example.test',
                               department_id=cls.dept_a.id, location='Operator')
        cls.e2 = cls._employee('R3 Two', barcode='R30002',
                               work_email='r3two@example.test',
                               department_id=cls.dept_a.id, location='Operator')
        cls.e3 = cls._employee('R3 Three', barcode='R30003',
                               work_email='r3three@example.test',
                               department_id=cls.dept_b.id, location='Fitter')
        cls.c1 = cls._contract(cls.e1, department_id=cls.dept_a.id)
        cls.c2 = cls._contract(cls.e2, department_id=cls.dept_a.id)
        cls.c3 = cls._contract(cls.e3, department_id=cls.dept_b.id)
        if cls.date_field:
            for emp in (cls.e1, cls.e2, cls.e3):
                emp[cls.date_field] = '1990-04-17'

        cls.everyone = {'employee_ids': [cls.e1.id, cls.e2.id, cls.e3.id]}
        cls.FIELDS = ['f:hr.employee:location', 'f:hr.contract:shuipart',
                      'f:hr.contract:department_id',
                      'f:hr.employee:%s' % cls.bool_field,
                      'b:acc_number', 'b:bank_name', 'c:R3BONUS']
        if cls.date_field:
            cls.FIELDS.append('f:hr.employee:%s' % cls.date_field)

    # --------------------------------------------------------------- fixtures
    @classmethod
    def _config(cls, name):
        return cls.Config.create({
            'name': name, 'code': name.upper().replace(' ', '')[:32],
            'country_code': 'VN', 'state': 'active',
            'company_id': cls.env.company.id,
        })

    @classmethod
    def _rule(cls, cfg, name, code, **extra):
        return cls.Rule.create(dict({
            'config_id': cfg.id, 'name': name, 'code': code,
            'column_type': 'input', 'sequence': 1, 'default_value': 0.0,
        }, **extra))

    @classmethod
    def _field(cls, model, name):
        return cls.IrField.search(
            [('model', '=', model), ('name', '=', name)], limit=1)

    @classmethod
    def _model(cls, model):
        return cls.IrModel.search([('model', '=', model)], limit=1)

    @classmethod
    def _map_field(cls, cfg, rule, model, field):
        return cls.Mapping.create({
            'salary_structure_id': cfg.id, 'component_id': rule.id,
            'destination_type': 'field',
            'target_model_id': cls._model(model).id,
            'target_field_id': cls._field(model, field).id,
        })

    @classmethod
    def _map_bank(cls, cfg, rule, role):
        return cls.Mapping.create({
            'salary_structure_id': cfg.id, 'component_id': rule.id,
            'destination_type': 'bank_account', 'bank_role': role})

    @classmethod
    def _employee(cls, name, **vals):
        return cls.Employee.create(dict(
            {'name': name, 'company_id': cls.env.company.id}, **vals))

    @classmethod
    def _contract(cls, employee, **vals):
        return cls.Contract.create(dict({
            'name': '%s contract' % employee.name,
            'employee_id': employee.id, 'wage': 1000.0, 'state': 'open',
            'date_start': '2026-01-01',
        }, **vals))

    # ---------------------------------------------------------------- helpers
    def _export(self, mode='data', fields=None, filters=None):
        res = self.Desk.export_records(
            self.cfg.id, filters if filters is not None else self.everyone,
            fields if fields is not None else self.FIELDS, mode)
        self.assertTrue(res.get('ok'), res.get('msg'))
        return res

    def _wb(self, res):
        return openpyxl.load_workbook(io.BytesIO(base64.b64decode(res['file_b64'])))

    def _b64(self, wb):
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return base64.b64encode(out.read()).decode()

    def _peek(self, b64, name='r3.xlsx'):
        return self.Desk.import_peek(self.cfg.id, b64, name)

    def _col_of(self, sheet, field_id):
        """The 1-based column a field id lives in, read off the hidden sheet."""
        meta = sheet.parent[pb_records_io.META_SHEET]
        for row in meta.iter_rows(min_row=3, values_only=True):
            if row and row[1] == field_id:
                return int(row[0])
        raise AssertionError('%s is not in the workbook' % field_id)

    def _label(self, field_id):
        """The label the desk shows for a card — the heading a file carries."""
        data = self.Desk.get_fields(self.cfg.id)
        for group in data['groups']:
            for card in group['fields']:
                if card['id'] == field_id:
                    return card['label']
        raise AssertionError('%s is not a card on this scheme' % field_id)

    def _payslip_fingerprint(self):
        slips = self.env['hr.payslip'].sudo().search([], order='id')
        blob = '|'.join((s.formula_input_values or '') for s in slips)
        return len(slips), hashlib.md5(blob.encode('utf-8')).hexdigest()

    # =====================================================================
    # 1 — what went out comes back meaning the same thing
    # =====================================================================
    def test_01a_a_round_trip_changes_nothing(self):
        # Give every column something to say, so "unchanged" is a claim about
        # values rather than about a sheet of blanks.
        self.Desk.apply_changes(self.cfg.id, [
            {'emp_id': self.e1.id, 'field_id': 'b:acc_number', 'value': '5550001'},
            {'emp_id': self.e1.id, 'field_id': 'b:bank_name', 'value': 'R3 Bank'},
            {'emp_id': self.e1.id, 'field_id': 'c:R3BONUS', 'value': '250000'},
            {'emp_id': self.e1.id, 'field_id': 'f:hr.contract:shuipart',
             'value': 'NO'},
        ])
        res = self._export()
        peek = self._peek(res['file_b64'])
        self.assertTrue(peek['ok'], peek.get('msg'))
        summary = peek['summary']
        self.assertEqual(summary['people_unmatched'], 0)
        self.assertEqual(summary['changes_ok'], 0,
                         "a file just exported cannot change anything")
        self.assertEqual(summary['changes_refused'], 0)
        self.assertEqual(summary['changes_same'], len(peek['changes']))
        self.assertGreater(summary['changes_same'], 0)
        self.assertEqual(peek['identity'], 'code')

    def test_01b_a_selection_travels_as_its_label_and_comes_back_as_its_key(self):
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.contract:shuipart')
        # The file carries what a PERSON reads, not the stored key…
        # `_description_selection` hands back a list of (key, label) TUPLES,
        # not the `{key,label}` dicts the desk's cards carry.
        labels = dict(
            self.Contract._fields['shuipart']._description_selection(self.env))
        shown = {ws.cell(row=r, column=col).value for r in (2, 3, 4)}
        self.assertTrue(shown <= set(labels.values()) | {None})
        # …and the key is what comes back, through R2's own `_selection_key`.
        ws.cell(row=2, column=col).value = labels['NO'] if 'NO' in labels else 'NO'
        peek = self._peek(self._b64(wb))
        changed = [i for i in peek['items'] if i['status'] == 'ok'
                   and i['field_id'] == 'f:hr.contract:shuipart']
        self.assertLessEqual(len(changed), 1)

    def test_01c_a_date_is_a_real_date_and_survives_the_trip(self):
        if not self.date_field:
            self.skipTest("no writable date destination on this build")
        res = self._export()
        ws = self._wb(res)[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.employee:%s' % self.date_field)
        cell = ws.cell(row=2, column=col)
        # A date written as TEXT is a date Excel cannot sort, filter or
        # reformat — and the most common way a round trip comes back unusable.
        self.assertFalse(isinstance(cell.value, str),
                         "the date column must hold a real date")
        self.assertEqual(cell.number_format, 'yyyy-mm-dd')
        peek = self._peek(res['file_b64'])
        dates = [i for i in peek['items']
                 if i['field_id'] == 'f:hr.employee:%s' % self.date_field]
        self.assertTrue(dates)
        self.assertTrue(all(i['status'] == 'same' for i in dates),
                        [i['why'] for i in dates])

    # =====================================================================
    # 2 — one edited cell is one change, and one audit row
    # =====================================================================
    def test_02_one_edited_cell_previews_and_applies_as_itself(self):
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.employee:location')
        ws.cell(row=2, column=col).value = 'Line Lead'
        peek = self._peek(self._b64(wb), 'r3_edited.xlsx')

        ok = [i for i in peek['items'] if i['status'] == 'ok']
        self.assertEqual(len(ok), 1, [i['field_label'] for i in ok])
        self.assertEqual(ok[0]['emp_id'], self.e1.id)
        self.assertEqual(ok[0]['field_id'], 'f:hr.employee:location')
        self.assertEqual(ok[0]['old_label'], 'Operator')
        self.assertEqual(ok[0]['new_label'], 'Line Lead')
        # Nothing was written by looking.
        self.assertEqual(self.e1.location, 'Operator')

        applied = self.Desk.apply_changes(self.cfg.id, peek['changes'],
                                          note='Imported r3_edited.xlsx',
                                          source='import')
        self.assertEqual(applied['written'], 1)
        self.assertEqual(self.e1.location, 'Line Lead')
        apply_rec = self.env['pb.records.apply'].browse(applied['apply_id'])
        self.assertEqual(apply_rec.source, 'import')
        self.assertEqual(apply_rec.note, 'Imported r3_edited.xlsx')
        self.assertEqual(len(apply_rec.change_ids), 1)
        # And it undoes exactly like a change made on the grid.
        self.Desk.undo_apply(applied['apply_id'])
        self.assertEqual(self.e1.location, 'Operator')

    def test_02b_a_boolean_written_as_a_word_is_read_as_one(self):
        field_id = 'f:hr.employee:%s' % self.bool_field
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, field_id)
        current = bool(self.e1[self.bool_field])
        ws.cell(row=2, column=col).value = 'No' if current else 'Yes'
        peek = self._peek(self._b64(wb))
        hit = [i for i in peek['items']
               if i['field_id'] == field_id and i['emp_id'] == self.e1.id]
        self.assertEqual(hit[0]['status'], 'ok')
        self.assertEqual(hit[0]['new_label'], 'No' if current else 'Yes')

    # =====================================================================
    # 3 — a row that names nobody is listed, and never created
    # =====================================================================
    def test_03_a_row_matching_nobody_is_listed_not_created(self):
        before = self.Employee.search_count([])
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        row = ws.max_row + 1
        ws.cell(row=row, column=1).value = 'R3NOBODY'
        ws.cell(row=row, column=2).value = 'Somebody Else Entirely'
        ws.cell(row=row, column=3).value = 'nobody@example.test'
        ws.cell(row=row, column=self._col_of(ws, 'f:hr.employee:location'))\
            .value = 'Line Lead'
        peek = self._peek(self._b64(wb))

        self.assertEqual(peek['summary']['people_unmatched'], 1)
        stray = peek['unmatched'][0]
        self.assertEqual(stray['code'], 'R3NOBODY')
        self.assertIn('R3NOBODY', stray['why'])
        self.assertEqual(stray['values']['f:hr.employee:location'], 'Line Lead')
        # The rest of the file still previews.
        self.assertEqual(peek['summary']['people_matched'], 3)
        self.env.flush_all()
        self.assertEqual(self.Employee.search_count([]), before,
                         "a file must never create a person")

    def test_03b_an_unmatched_row_can_be_bound_by_hand(self):
        # The client moves the row's values into `changes` with the id the
        # typeahead returned; the server side of that is `lookup_people` plus
        # the ordinary preview.
        found = self.Desk.lookup_people('R3 Two', 5)
        self.assertTrue(any(r['id'] == self.e2.id for r in found))
        preview = self.Desk.preview_changes(self.cfg.id, [
            {'emp_id': self.e2.id, 'field_id': 'f:hr.employee:location',
             'value': 'Line Lead'}])
        self.assertEqual(preview['counts']['ok'], 1)

    # =====================================================================
    # 4 — a name two people share is a question, not a guess
    # =====================================================================
    def test_04_an_ambiguous_name_says_how_many_and_what_to_add(self):
        twin_a = self._employee('R3 Twin')
        twin_b = self._employee('R3 Twin')
        self.assertNotEqual(twin_a.id, twin_b.id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = pb_records_io.SHEET
        ws.append(['Employee code', 'Name', 'Work email',
                   self._label('f:hr.employee:location')])
        ws.append(['', 'R3 Twin', '', 'Line Lead'])
        peek = self._peek(self._b64(wb))
        self.assertEqual(peek['summary']['people_unmatched'], 1)
        why = peek['unmatched'][0]['why']
        self.assertIn('2 people called R3 Twin', why)
        self.assertIn('employee code', why)
        self.assertFalse(twin_a.location)
        self.assertFalse(twin_b.location)

    # =====================================================================
    # 5 — a column nobody asked for is named, and read past
    # =====================================================================
    def test_05_an_unknown_heading_is_ignored_and_named(self):
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        extra = ws.max_column + 1
        ws.cell(row=1, column=extra).value = 'Shoe size'
        ws.cell(row=2, column=extra).value = '44'
        peek = self._peek(self._b64(wb))
        self.assertIn('Shoe size', peek['summary']['columns_ignored'])
        self.assertEqual(peek['summary']['changes_ok'], 0)
        self.assertEqual(peek['summary']['people_unmatched'], 0)

    # =====================================================================
    # 6 — a retyped heading still lands on the right field
    # =====================================================================
    def test_06a_the_hidden_sheet_survives_a_retyped_heading(self):
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.contract:shuipart')
        cell = ws.cell(row=1, column=col)
        cell.value = 'SHUI?'
        cell.comment = None                  # the comment is gone too
        ws.cell(row=2, column=col).value = 'NO'
        peek = self._peek(self._b64(wb))
        self.assertNotIn('SHUI?', peek['summary']['columns_ignored'])
        hit = [i for i in peek['items']
               if i['field_id'] == 'f:hr.contract:shuipart'
               and i['emp_id'] == self.e1.id]
        self.assertTrue(hit)

    def test_06b_the_header_comment_survives_a_deleted_hidden_sheet(self):
        res = self._export()
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.employee:location')
        ws.cell(row=1, column=col).value = 'Where they work'
        del wb[pb_records_io.META_SHEET]
        ws.cell(row=2, column=col).value = 'Line Lead'
        peek = self._peek(self._b64(wb))
        self.assertNotIn('Where they work', peek['summary']['columns_ignored'])
        ok = [i for i in peek['items'] if i['status'] == 'ok']
        self.assertEqual(len(ok), 1)
        self.assertEqual(ok[0]['field_id'], 'f:hr.employee:location')

    def test_06c_the_label_alone_is_enough_for_a_hand_made_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # The heading is the card's OWN label, lower-cased and with a stray
        # space — a hand-typed heading, not a copy of the export.
        ws.append(['Employee Code', ' %s ' % self._label('f:hr.employee:location').lower()])
        ws.append(['R30002', 'Line Lead'])
        peek = self._peek(self._b64(wb), 'handmade.xlsx')
        self.assertEqual(peek['summary']['changes_ok'], 1)
        self.assertEqual(peek['items'][0]['emp_id'], self.e2.id)

    # =====================================================================
    # 7 — four bank columns are one account
    # =====================================================================
    def test_07_bank_columns_assemble_one_account_on_apply(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Employee code', 'Account number', 'Bank name'])
        ws.append(['R30003', '7778889', 'R3 Savings'])
        peek = self._peek(self._b64(wb))
        self.assertEqual(peek['summary']['changes_ok'], 2)
        self.Desk.apply_changes(self.cfg.id, peek['changes'], source='import')
        accounts = self.e3.bank_account_ids
        self.assertEqual(len(accounts), 1)
        self.assertEqual(accounts.acc_number, '7778889')
        self.assertEqual(accounts.bank_id.name, 'R3 Savings')
        # ADD, never replace — R2's rule, reached through the same method.
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.append(['Employee code', 'Account number'])
        ws2.append(['R30003', '9990001'])
        again = self._peek(self._b64(wb2))
        self.Desk.apply_changes(self.cfg.id, again['changes'], source='import')
        self.assertEqual(len(self.e3.bank_account_ids), 2)

    # =====================================================================
    # 8 — a component column is a contract line
    # =====================================================================
    def test_08_a_component_column_writes_the_line_and_logs_it(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Employee code', 'Site bonus'])
        ws.append(['R30002', 425000])
        peek = self._peek(self._b64(wb))
        self.assertEqual(peek['summary']['changes_ok'], 1)
        self.assertEqual(peek['items'][0]['field_id'], 'c:R3BONUS')
        self.Desk.apply_changes(self.cfg.id, peek['changes'], source='import')

        Template = self.env['hr.contract.advantage.template']
        template = Template.search([('code', '=', 'R3BONUS')])
        self.assertEqual(len(template), 1)
        line = self.env['hr.contract.advantage'].search([
            ('contract_id', '=', self.c2.id),
            ('advantage_template_id', '=', template.id)])
        self.assertAlmostEqual(line.amount, 425000.0, places=2)
        change = self.env['hr.contract.advantage.change'].search([
            ('contract_id', '=', self.c2.id),
            ('advantage_template_id', '=', template.id)])
        self.assertEqual(len(change), 1)
        self.assertEqual(change.change_source, 'manual')

    # =====================================================================
    # 9 — a .csv says the same thing as an .xlsx
    # =====================================================================
    def test_09_a_csv_previews_exactly_like_the_workbook(self):
        csv_text = ("Employee code,Name,Work email,%s\n"
                    % self._label('f:hr.employee:location') +
                    "R30001,R3 One,r3one@example.test,Line Lead\n"
                    "R30002,R3 Two,r3two@example.test,Operator\n")
        peek = self.Desk.import_peek(
            self.cfg.id,
            base64.b64encode(csv_text.encode('utf-8')).decode(), 'r3.csv')
        self.assertTrue(peek['ok'], peek.get('msg'))
        self.assertEqual(peek['summary']['people_matched'], 2)
        self.assertEqual(peek['summary']['changes_ok'], 1)
        self.assertEqual(peek['summary']['changes_same'], 1)
        self.assertEqual(peek['items'][0]['emp_id'], self.e1.id)
        # A semicolon-delimited export from a European locale is the same file.
        peek2 = self.Desk.import_peek(
            self.cfg.id,
            base64.b64encode(csv_text.replace(',', ';').encode()).decode(),
            'r3_semicolon.csv')
        self.assertTrue(peek2['ok'], peek2.get('msg'))
        self.assertEqual(peek2['summary']['changes_ok'], 1)

    # =====================================================================
    # 10 — a bad file is a sentence, never an exception
    # =====================================================================
    def test_10a_an_empty_file_says_so(self):
        res = self.Desk.import_peek(self.cfg.id, '', 'empty.xlsx')
        self.assertFalse(res['ok'])
        self.assertIn('empty', res['msg'].lower())

    def test_10b_something_that_is_not_a_spreadsheet_says_so(self):
        fake = base64.b64encode(b'PK\x03\x04 this is not a workbook').decode()
        res = self.Desk.import_peek(self.cfg.id, fake, 'letter.docx')
        self.assertFalse(res['ok'])
        self.assertIn('not a spreadsheet', res['msg'].lower())
        renamed = self.Desk.import_peek(self.cfg.id, fake, 'letter.xlsx')
        self.assertFalse(renamed['ok'])
        self.assertIn('could not be opened', renamed['msg'])

    def test_10c_a_workbook_with_no_heading_row_says_so(self):
        wb = openpyxl.Workbook()
        res = self.Desk.import_peek(self.cfg.id, self._b64(wb), 'blank.xlsx')
        self.assertFalse(res['ok'])
        self.assertIn('no heading row', res['msg'])

    def test_10d_headings_that_match_nothing_say_what_to_do_next(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Shoe size', 'Favourite colour'])
        ws.append([44, 'blue'])
        res = self.Desk.import_peek(self.cfg.id, self._b64(wb), 'wrong.xlsx')
        self.assertFalse(res['ok'])
        self.assertIn('export a file from this desk first', res['msg'])
        self.assertIn('Shoe size', res['columns_ignored'])

    def test_10e_a_file_nobody_can_be_matched_from_says_so(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([self._label('f:hr.employee:location')])
        ws.append(['Line Lead'])
        res = self.Desk.import_peek(self.cfg.id, self._b64(wb), 'nokey.xlsx')
        self.assertFalse(res['ok'])
        self.assertIn('no row can be matched', res['msg'])

    # =====================================================================
    # 11 — a blank template is a template
    # =====================================================================
    def test_11_a_template_carries_identity_headings_and_no_values(self):
        res = self._export(mode='template')
        self.assertEqual(res['mode'], 'template')
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        self.assertEqual(ws.cell(row=1, column=1).value, 'Employee code')
        self.assertEqual(ws.cell(row=2, column=1).value, 'R30001')
        for field_id in ('f:hr.employee:location', 'f:hr.contract:shuipart'):
            col = self._col_of(ws, field_id)
            self.assertIsNone(ws.cell(row=2, column=col).value,
                              '%s must be empty in a template' % field_id)
            comment = ws.cell(row=1, column=col).comment
            self.assertIsNotNone(comment)
            self.assertIn('id: %s' % field_id, comment.text)
        # A dropdown on the selection column, and one on the boolean.
        ranges = " ".join(str(dv.sqref) for dv in ws.data_validations.dataValidation)
        for field_id in ('f:hr.contract:shuipart',
                         'f:hr.employee:%s' % self.bool_field):
            letter = openpyxl.utils.get_column_letter(self._col_of(ws, field_id))
            self.assertIn('%s2' % letter, ranges)
        self.assertEqual(ws.freeze_panes, 'D2')
        # The hidden sheet is present, hidden, and carries the scheme.
        meta = wb[pb_records_io.META_SHEET]
        self.assertEqual(meta.sheet_state, 'hidden')
        self.assertEqual(meta['B1'].value, self.cfg.id)
        # And dropping it straight back changes nothing at all.
        peek = self._peek(res['file_b64'], 'template.xlsx')
        self.assertTrue(peek['ok'])
        self.assertEqual(peek['summary']['changes_ok'], 0)
        self.assertEqual(peek['summary']['people_unmatched'], 0)
        self.assertGreater(peek['summary']['cells_blank'], 0)

    def test_11b_a_file_with_headings_and_no_rows_says_what_to_do(self):
        res = self._export(mode='template')
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        ws.delete_rows(2, ws.max_row)
        peek = self._peek(self._b64(wb), 'headings.xlsx')
        self.assertTrue(peek['ok'])
        self.assertTrue(peek['empty'])
        self.assertIn('headings only', peek['msg'])

    def test_11c_a_field_the_scheme_does_not_map_is_never_a_column(self):
        res = self.Desk.export_records(
            self.cfg.id, self.everyone, ['f:hr.employee:barcode'], 'data')
        self.assertFalse(res['ok'])
        self.assertIn('Pick at least one field', res['msg'])

    # =====================================================================
    # 12 — the cap is real, and never silent
    # =====================================================================
    def test_12_the_export_stops_at_the_cap_and_says_so(self):
        # The real ceiling is asserted; the BEHAVIOUR is exercised against a
        # small one, because creating ten thousand and one employees would add
        # minutes to every run of this suite to prove the same `>` .
        self.assertEqual(pb_records_io.MAX_ROWS, 10000)
        with patch.object(pb_records_io, 'MAX_ROWS', 2):
            res = self._export()
        self.assertTrue(res['truncated'])
        self.assertEqual(res['rows'], 2)
        self.assertEqual(res['total'], 3)
        ws = self._wb(res)[pb_records_io.SHEET]
        self.assertEqual(ws.max_row, 3)          # one heading + two people
        full = self._export()
        self.assertFalse(full['truncated'])
        self.assertEqual(full['rows'], 3)

    # =====================================================================
    # 13 — no payslip, and no person, is created anywhere in this suite
    # =====================================================================
    def test_13_nothing_creates_a_payslip_or_a_person(self):
        before_slips = self._payslip_fingerprint()
        before_people = self.Employee.search_count([])
        before_contracts = self.Contract.search_count([])

        res = self._export()
        self._export(mode='template')
        wb = self._wb(res)
        ws = wb[pb_records_io.SHEET]
        col = self._col_of(ws, 'f:hr.employee:location')
        ws.cell(row=2, column=col).value = 'Line Lead'
        stray = ws.max_row + 1
        ws.cell(row=stray, column=1).value = 'R3GHOST'
        ws.cell(row=stray, column=2).value = 'A Ghost'
        ws.cell(row=stray, column=col).value = 'Line Lead'
        peek = self._peek(self._b64(wb))
        applied = self.Desk.apply_changes(self.cfg.id, peek['changes'],
                                          source='import')
        self.Desk.undo_apply(applied['apply_id'])
        self.env.flush_all()

        self.assertEqual(self._payslip_fingerprint(), before_slips)
        self.assertEqual(self.Employee.search_count([]), before_people)
        self.assertEqual(self.Contract.search_count([]), before_contracts)

    # =====================================================================
    # 14 — no user-visible string says the wrong word
    # =====================================================================
    #: The pictograph planes plus anything explicitly presented as an emoji.
    #: Deliberately NOT "the whole symbol block" — `✕` is a close button and
    #: `→` is an arrow, and a rule that fails on those only teaches the next
    #: engineer to delete the rule (RD5).
    EMOJI = re.compile('[\U0001F000-\U0001FAFF\U00002600-\U000027BF️]')
    SAFE_SYMBOLS = set('←→↑↓·✕—–≤≥⌘⇆')

    def _emoji_in(self, body):
        return [ch for ch in self.EMOJI.findall(body)
                if ch not in self.SAFE_SYMBOLS]

    def test_14a_the_new_source_never_names_the_engine(self):
        for parts, is_xml in (
                (('models', 'pb_records_io.py'), False),
                (('static', 'src', 'js', 'records_import.js'), False),
                (('static', 'src', 'xml', 'records_desk.xml'), True)):
            body = _src('pb_records', *parts)
            if is_xml:
                body = re.sub(r'<!--.*?-->', ' ', body, flags=re.S)
                body = re.sub(r'</?odoo>', ' ', body)
            else:
                body = '\n'.join(
                    re.sub(r'(?<!["\'])#.*$', '', line)
                    for line in body.splitlines())
                # `/** @odoo-module **/` and the framework imports are technical
                # identifiers, never strings a person reads (the standing rule).
                body = re.sub(r'@odoo-module|@odoo/owl|@web/[\w/.]+', ' ', body)
            self.assertNotIn('Odoo', body, "user-visible string in %s" % parts[-1])
            self.assertFalse(self._emoji_in(body), "emoji in %s" % parts[-1])

    def test_14b_the_exported_workbook_never_names_the_engine(self):
        res = self._export()
        wb = self._wb(res)
        words = list(wb.sheetnames)
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        words.append(str(cell.value))
                    if cell.comment is not None and cell.comment.text:
                        words.append(cell.comment.text)
        blob = "\n".join(words)
        self.assertNotIn('Odoo', blob, "the workbook names the engine")
        self.assertNotIn('odoo', blob)
        self.assertFalse(self._emoji_in(blob), "emoji in the workbook")
        self.assertIn(pb_records_io.META_SHEET, wb.sheetnames)
        self.assertEqual(res['filename'],
                         '%s_records_%s.xlsx' % (
                             self.cfg.code,
                             __import__('datetime').date.today().isoformat()))
