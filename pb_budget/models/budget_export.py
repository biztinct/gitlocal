# -*- coding: utf-8 -*-
"""`pb.budget.export` — the year as a workbook, and the year as a page.

Both exports are built from the SAME payload the screen draws (`pb.budget
.get_board`), never from a second query: a spreadsheet that disagrees with the
board it was exported from is worse than no spreadsheet at all.

The file comes back as base64 and the browser saves it without leaving the page —
the idiom `pb_records` uses (`records_desk.js:731`). No attachment is left
behind, because an export is a copy somebody takes away, not a record.

The PDF is a real `ir.actions.report` over this transient, so it renders through
the platform's own layout, honours the company's letterhead and can be printed
from anywhere a report can. Its narrative lines are written in plain English on
the server, once, so the page and the screen say exactly the same thing.
"""

import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .budget_common import type_label

_logger = logging.getLogger(__name__)


class PbBudgetExport(models.TransientModel):
    _name = 'pb.budget.export'
    _description = 'Budget export'

    fy = fields.Integer(string='Year')
    budget_type = fields.Char(string='Budget for')
    currency_mode = fields.Char(string='Reported in', default='report')

    # ---------------------------------------------------------------- entry
    @api.model
    def build(self, fy=None, budget_type='manpower', currency='report',
              kind='xlsx'):
        board = self.env['pb.budget'].get_board(fy, budget_type, currency)
        rec = self.create({
            'fy': board['fy'],
            'budget_type': board['budget_type'],
            'currency_mode': board['currency']['mode'],
        })
        if kind == 'pdf':
            return rec._pdf(board)
        return rec._xlsx(board)

    def board(self):
        """What the QWeb template reads. Rebuilt for the render, from the same
        facade — a report that took its numbers from a stashed blob would print
        whatever the screen looked like when the button was pressed."""
        self.ensure_one()
        return self.env['pb.budget'].get_board(
            self.fy, self.budget_type, self.currency_mode)

    # ------------------------------------------------------------- the sheet
    def _xlsx(self, board):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_(
                "This system cannot build spreadsheets at the moment. Use the "
                "PDF, or ask an administrator to look at it."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Budget %s') % board['fy_label']

        head = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='6355C7')
        bold = Font(bold=True)
        money = '#,##0'

        cur = board['currency']['code']
        ws.cell(row=1, column=1,
                value=_('%(type)s budget %(year)s — every figure in %(cur)s',
                        type=type_label(board['budget_type'], self.env),
                        year=board['fy_label'], cur=cur)).font = bold
        ws.cell(row=2, column=1, value=board['headline'])

        cols = [_('Function'), _('Department'), _('Budget'), _('Spent'),
                _('Left'), _('Used %'), _('Year gone %'), _('Reading')]
        cols += [m['label'] + ' ' + _('budget') for m in board['months']]
        cols += [m['label'] + ' ' + _('spent') for m in board['months']]
        for i, label in enumerate(cols, start=1):
            c = ws.cell(row=4, column=i, value=label)
            c.font = head
            c.fill = fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        row = 5
        for f in board['functions']:
            vals = [f['name'], _('All'), f['budget'], f['spent'], f['left'],
                    f['burn'], f['pace'], f['tone_label']]
            vals += [m['budget'] for m in f['months']]
            vals += [m['spent'] for m in f['months']]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                if i in (3, 4, 5) or i > 8:
                    cell.number_format = money
                if i == 1:
                    cell.font = bold
            row += 1
            for d in f['departments']:
                vals = ['', d['name'], d['budget'], d['spent'],
                        round(d['budget'] - d['spent'], 2), '', '',
                        _('No budget set') if d['unbudgeted'] else '']
                for i, v in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=i, value=v)
                    if i in (3, 4, 5):
                        cell.number_format = money
                row += 1

        k = board['kpis']
        row += 1
        ws.cell(row=row, column=1, value=_('Total')).font = bold
        for i, v in ((3, k['budget']), (4, k['spent']), (5, k['left']),
                     (6, k['burn']), (7, k['pace'])):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = bold
            if i in (3, 4, 5):
                cell.number_format = money

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 26
        for i in range(3, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14
        ws.freeze_panes = 'C5'

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': _('Budget %(year)s %(type)s.xlsx',
                          year=board['fy_label'],
                          type=type_label(board['budget_type'], self.env)),
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
            'rows': len(board['functions']),
        }

    # --------------------------------------------------------------- the page
    def _pdf(self, board):
        self.ensure_one()
        report = self.env.ref('pb_budget.action_report_budget_year',
                              raise_if_not_found=False)
        if not report:
            raise UserError(_("The budget report is not installed on this "
                              "system."))
        # NOT `sudo()`. The report re-reads the board while it renders, and a
        # superuser env sees every company — which put another company's
        # departments into a company-scoped person's PDF, with totals that then
        # disagreed with the spreadsheet beside it. The render runs as the
        # person who pressed the button, carrying their company set explicitly.
        pdf, _ext = report.with_context(
            allowed_company_ids=self.env.companies.ids,
        )._render_qweb_pdf('pb_budget.report_budget_year_document',
                           res_ids=self.ids)
        return {
            'ok': True,
            'file_b64': base64.b64encode(pdf).decode(),
            'filename': _('Budget %(year)s %(type)s.pdf',
                          year=board['fy_label'],
                          type=type_label(board['budget_type'], self.env)),
            'mimetype': 'application/pdf',
            'rows': len(board['functions']),
        }

    # ------------------------------------------------------------ the wording
    def month_bars(self, board):
        """Twelve rows: the month, what was spent in it, and that as a share of
        the biggest month — so a page with no chart engine still SHOWS the
        shape of the year rather than listing it."""
        self.ensure_one()
        totals = {}
        for f in board['functions']:
            for m in f['months']:
                totals[m['key']] = totals.get(m['key'], 0.0) + (m['spent'] or 0.0)
        peak = max(totals.values()) if totals else 0.0
        return [{
            'key': m['key'],
            'label': '%s %s' % (m['label'], m['year']),
            'spent': round(totals.get(m['key'], 0.0), 2),
            'pct': round(totals.get(m['key'], 0.0) / peak * 100, 1) if peak else 0.0,
        } for m in board['months']]

    def narrative(self, board=None):
        """The variance lines, in the words a person would use.

        One sentence per function, and each one says the same three things: how
        much of the budget has gone, where the year is, and what that means.
        """
        self.ensure_one()
        board = board or self.board()
        lines = []
        for f in board['functions']:
            if f['tone'] == 'none':
                lines.append(_(
                    "%(name)s has spent %(spent)s with no budget set against "
                    "it.", name=f['name'],
                    spent='{:,.0f}'.format(f['spent'])))
                continue
            if not f['budget']:
                continue
            gap = f['gap']
            if gap > 15:
                lines.append(_(
                    "%(name)s is %(gap)s%% ahead of budget pace — %(burn)s%% "
                    "spent, %(pace)s%% of the year gone.", name=f['name'],
                    gap='{:,.0f}'.format(gap), burn='{:,.0f}'.format(f['burn']),
                    pace='{:,.0f}'.format(f['pace'])))
            elif gap > 5:
                lines.append(_(
                    "%(name)s is running warm — %(burn)s%% spent against "
                    "%(pace)s%% of the year.", name=f['name'],
                    burn='{:,.0f}'.format(f['burn']),
                    pace='{:,.0f}'.format(f['pace'])))
            elif gap < -10:
                lines.append(_(
                    "%(name)s is %(gap)s%% behind budget pace — %(burn)s%% "
                    "spent, %(pace)s%% of the year gone.", name=f['name'],
                    gap='{:,.0f}'.format(abs(gap)),
                    burn='{:,.0f}'.format(f['burn']),
                    pace='{:,.0f}'.format(f['pace'])))
            else:
                lines.append(_(
                    "%(name)s is on pace — %(burn)s%% spent against %(pace)s%% "
                    "of the year.", name=f['name'],
                    burn='{:,.0f}'.format(f['burn']),
                    pace='{:,.0f}'.format(f['pace'])))
        return lines
