# -*- coding: utf-8 -*-
"""`pb.budget.upload.wizard` — a year's budget arrives as a spreadsheet.

THE THREE RAILS, WHICH ARE `pb_records`' RAILS AND FOR THE SAME REASONS

  1. **The upload never writes.** `peek()` reads the file and says what it would
     mean — how many rows would be created, how many updated, and every row it
     could not place. Not one byte is written until somebody presses Apply, and
     Apply re-reads the same file rather than trusting anything the preview
     stashed: a plan held between two clicks is a plan that can be out of date
     by the time it is used.
  2. **A file never creates a department.** A row naming a department this
     database has never heard of is LISTED with its name and skipped. Inventing
     an organisation from a spreadsheet is how a budget board ends up with two
     Marketings.
  3. **A blank cell is not a zero.** An empty month means "not filled in" and is
     left exactly as it was; a typed 0 means "nothing budgeted here" and is
     written. The alternative is that dropping a half-filled template back wipes
     the eleven months somebody else did.

WHAT IT WRITES: `forecast_cost` — the BUDGET — and the row's currency, its manual
rate and `pb_source='upload'`. It never touches `actual_cost`: what was spent is
the actuals job's column and typing over it would mean a budget file could
rewrite payroll history.

IDEMPOTENT BY (company, department, month, budget type). The same file twice
updates the same rows to the same numbers. Proven by test T2.
"""

import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from odoo.addons.pb_budget.models.budget_common import (BUDGET_TYPES, TYPE_KEYS,
                                                        counted, fold,
                                                        type_label)

_logger = logging.getLogger(__name__)

#: The columns the template ships, in the order it ships them. Everything before
#: the months is fixed; the months are twelve, named by the year they belong to.
FIXED = ['Department', 'Department id', 'Budget for', 'Money in',
         'Rate to reporting currency']
SHEET = 'Budget'
#: Rows a single file may carry. A cap that is reported, never silent.
FILE_ROW_CAP = 2000


class PbBudgetUploadWizard(models.TransientModel):
    _name = 'pb.budget.upload.wizard'
    _description = 'Budget upload'

    fy = fields.Integer(
        string='Year', required=True,
        default=lambda self: self.env['pb.budget']._current_fy())
    budget_type = fields.Selection(
        BUDGET_TYPES, string='Budget for', default='manpower', required=True)
    upload_file = fields.Binary(string='Filled-in file')
    upload_filename = fields.Char(string='File name')
    template_file = fields.Binary(string='Template', readonly=True)
    template_filename = fields.Char(string='Template name', readonly=True)
    summary = fields.Text(string='What would happen', readonly=True)

    # ============================================================== the template
    @api.model
    def template_xlsx(self, fy=None, budget_type='manpower'):
        """The blank file: every department down the side, the year across."""
        self.env['pb.budget']._require_edit()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_(
                "This system cannot build spreadsheets at the moment. Ask an "
                "administrator to look at it."))

        fy = int(fy or self.env['pb.budget']._current_fy())
        btype = budget_type if budget_type in TYPE_KEYS else 'manpower'
        months = self.env['pb.budget']._fy_months(fy)
        company = self.env.company

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET
        head = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='6355C7')

        ws.cell(row=1, column=1, value=_(
            "Budget for %(year)s — %(type)s. Put the amount for each month "
            "under that month. Leave a month empty to leave it as it is; type "
            "0 to say nothing is budgeted. Do not change the department id "
            "column.", year=self.env['pb.budget']._fy_label(fy),
            type=type_label(btype, self.env))).font = Font(italic=True)

        cols = list(FIXED) + [m.strftime('%Y-%m') for m in months]
        for i, label in enumerate(cols, start=1):
            c = ws.cell(row=3, column=i, value=label)
            c.font = head
            c.fill = fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        depts = self.env['hr.department'].search(
            [('company_id', 'in', self.env.companies.ids)],
            order='complete_name')
        row = 4
        for dept in depts:
            ws.cell(row=row, column=1, value=dept.complete_name or dept.name)
            ws.cell(row=row, column=2, value=dept.id)
            ws.cell(row=row, column=3, value=type_label(btype, self.env))
            ws.cell(row=row, column=4, value=company.currency_id.name)
            ws.cell(row=row, column=5, value=0)
            row += 1
        # One last line for money the whole company spends and no team owns.
        ws.cell(row=row, column=1, value=_('Whole company'))
        ws.cell(row=row, column=3, value=type_label(btype, self.env))
        ws.cell(row=row, column=4, value=company.currency_id.name)
        ws.cell(row=row, column=5, value=0)

        ws.column_dimensions['A'].width = 40
        for i in range(2, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14
        ws.freeze_panes = 'B4'

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': _('Budget template %(year)s %(type)s.xlsx',
                          year=self.env['pb.budget']._fy_label(fy),
                          type=type_label(btype, self.env)),
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
            'departments': len(depts),
        }

    # ================================================================ the read
    @api.model
    def peek(self, file_b64, fy=None, budget_type='manpower'):
        """What this file WOULD do. Writes nothing."""
        self.env['pb.budget']._require_edit()
        plan = self._plan(file_b64, fy, budget_type)
        plan['message'] = self._sentence(plan, applied=False)
        # The working list never crosses the wire — the browser gets the
        # COUNTS and a readable sample, and Apply re-reads the file itself.
        plan.pop('writes', None)
        return plan

    @api.model
    def apply(self, file_b64, fy=None, budget_type='manpower'):
        """Read it again, then write it. The preview is never trusted."""
        self.env['pb.budget']._require_edit()
        plan = self._plan(file_b64, fy, budget_type)
        Budget = self.env['wfp.budget.actual']
        created = updated = 0
        for row in plan['writes']:
            rec = Budget.search([
                ('company_id', '=', row['company_id']),
                ('department_id', '=', row['department_id'] or False),
                ('period_month', '=', row['month']),
                ('pb_budget_type', '=', row['budget_type']),
            ], limit=1)
            vals = {
                'forecast_cost': row['amount'],
                'pb_source': 'upload',
                'pb_currency_id': row['currency_id'],
                'pb_manual_rate': row['manual_rate'],
            }
            if rec:
                rec.write(vals)
                updated += 1
            else:
                vals.update({
                    'company_id': row['company_id'],
                    'department_id': row['department_id'] or False,
                    'period_month': row['month'],
                    'pb_budget_type': row['budget_type'],
                })
                Budget.create(vals)
                created += 1
        plan['created'] = created
        plan['updated'] = updated
        plan['applied'] = True
        plan['message'] = self._sentence(plan, applied=True)
        # A new budget row changes what "unbudgeted" means for that month, and
        # the spend already read is unaffected — so nothing is re-read here.
        plan.pop('writes', None)
        return plan

    # ================================================================ the plan
    @api.model
    def _plan(self, file_b64, fy=None, budget_type='manpower'):
        if not file_b64:
            raise UserError(_("Pick the filled-in file first."))
        fy = int(fy or self.env['pb.budget']._current_fy())
        default_type = budget_type if budget_type in TYPE_KEYS else 'manpower'
        months = self.env['pb.budget']._fy_months(fy)
        month_by_key = {m.strftime('%Y-%m'): m for m in months}

        try:
            import openpyxl
            raw = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:                 # noqa: BLE001
            _logger.info('pb_budget: unreadable upload: %s', e)
            raise UserError(_(
                "That file could not be opened. It needs to be the .xlsx "
                "template this screen gives you — a .csv or an older .xls "
                "will not do."))
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]

        header_row, headers = self._headers(ws)
        if not headers:
            raise UserError(_(
                "That file has no heading row this screen recognises. Download "
                "the template again and fill that in — the headings are how "
                "each column is understood."))

        depts = self._departments()
        writes, unknown, blanks, skipped = [], [], 0, []
        seen, over_cap = set(), 0
        company = self.env.company

        for r in range(header_row + 1, ws.max_row + 1):
            if len(writes) >= FILE_ROW_CAP:
                over_cap += 1
                continue
            cells = {h: ws.cell(row=r, column=c).value
                     for h, c in headers.items()}
            name = (cells.get('department') or '')
            dept_id = self._as_int(cells.get('department_id'))
            whole = (not dept_id and fold(name) in
                     (fold(_('Whole company')), 'whole company', ''))
            if dept_id and dept_id not in depts:
                dept_id = 0
            if not dept_id and not whole:
                if not str(name or '').strip():
                    continue                    # a genuinely empty line
                match = depts.get(fold(name)) or depts.get(
                    fold(str(name).split('/')[-1]))
                if not match:
                    unknown.append(str(name).strip())
                    continue
                dept_id = match
            row_type = self._as_type(cells.get('budget_for'), default_type)
            currency = self._as_currency(cells.get('money_in'), company)
            rate = self._as_float(cells.get('rate_to_reporting_currency'), 0.0)

            for key, month in month_by_key.items():
                raw_val = cells.get(key)
                if raw_val in (None, ''):
                    blanks += 1
                    continue
                amount = self._as_float(raw_val, None)
                if amount is None:
                    skipped.append(_("%(dept)s, %(month)s: “%(what)s” is not a "
                                     "number.",
                                     dept=name or _('Whole company'),
                                     month=key, what=raw_val))
                    continue
                key3 = (company.id, dept_id or 0, month, row_type)
                if key3 in seen:
                    continue                    # the file names it twice; first wins
                seen.add(key3)
                writes.append({
                    'company_id': company.id,
                    'department_id': dept_id or 0,
                    'department': (name or _('Whole company')),
                    'month': month,
                    'month_key': key,
                    'budget_type': row_type,
                    'amount': amount,
                    'currency_id': currency.id,
                    'currency': currency.name,
                    'manual_rate': rate,
                })

        existing = self._existing(writes)
        for row in writes:
            row['exists'] = (row['company_id'], row['department_id'] or 0,
                             row['month'], row['budget_type']) in existing
        return {
            'ok': True,
            'applied': False,
            'fy': fy,
            'fy_label': self.env['pb.budget']._fy_label(fy),
            'writes': writes,
            'to_create': len([w for w in writes if not w['exists']]),
            'to_update': len([w for w in writes if w['exists']]),
            'unknown': sorted(set(unknown)),
            'blanks': blanks,
            'skipped': skipped[:20],
            'over_cap': over_cap,
            'preview': [{k: (str(v) if k == 'month' else v)
                         for k, v in w.items()} for w in writes[:60]],
        }

    # ---------------------------------------------------------------- readers
    @api.model
    def _headers(self, ws):
        """Find the heading row and place every column by its heading.

        The template's own headings, folded, so an editor that re-cased or
        re-spaced them still lands. A column nothing recognises is ignored and
        never guessed at.
        """
        wanted = {fold(h): h.lower().replace(' ', '_') for h in FIXED}
        for r in range(1, min(ws.max_row, 12) + 1):
            found = {}
            for c in range(1, min(ws.max_column, 60) + 1):
                raw = ws.cell(row=r, column=c).value
                if raw in (None, ''):
                    continue
                text = str(raw).strip()
                key = wanted.get(fold(text))
                if key:
                    found[key] = c
                    continue
                # a month heading: 2026-01, or a real date the editor made
                token = text[:7]
                if len(token) == 7 and token[4] == '-' and token[:4].isdigit():
                    found[token] = c
                elif hasattr(raw, 'strftime'):
                    found[raw.strftime('%Y-%m')] = c
            if 'department' in found or 'department_id' in found:
                return r, found
        return 0, {}

    @api.model
    def _departments(self):
        recs = self.env['hr.department'].search_read(
            [('company_id', 'in', self.env.companies.ids)],
            ['id', 'name', 'complete_name'])
        out = {}
        for r in recs:
            out[r['id']] = r['id']
            for label in (r.get('complete_name'), r.get('name')):
                if label:
                    out.setdefault(fold(label), r['id'])
        return out

    @api.model
    def _existing(self, writes):
        if not writes:
            return set()
        months = sorted({w['month'] for w in writes})
        recs = self.env['wfp.budget.actual'].search_read([
            ('period_month', '>=', months[0]),
            ('period_month', '<=', months[-1]),
        ], ['company_id', 'department_id', 'period_month', 'pb_budget_type'])
        return {((r['company_id'] or [0])[0],
                 (r['department_id'] or [0])[0] or 0,
                 r['period_month'], r['pb_budget_type']) for r in recs}

    # ----------------------------------------------------------- the coercers
    @api.model
    def _as_int(self, raw):
        try:
            return int(float(str(raw).strip()))
        except (TypeError, ValueError):
            return 0

    @api.model
    def _as_float(self, raw, default=0.0):
        if raw in (None, ''):
            return default
        try:
            text = str(raw).strip().replace(',', '').replace(' ', '')
            return float(text)
        except (TypeError, ValueError):
            return default

    @api.model
    def _as_type(self, raw, fallback):
        needle = fold(raw or '')
        if not needle:
            return fallback
        for key, label in BUDGET_TYPES:
            if needle in (fold(key), fold(label), fold(type_label(key, self.env))):
                return key
        return fallback

    @api.model
    def _as_currency(self, raw, company):
        code = str(raw or '').strip().upper()
        if code:
            cur = self.env['res.currency'].sudo().search(
                [('name', '=', code)], limit=1)
            if cur:
                return cur
        return company.currency_id

    # ------------------------------------------------------------ the wording
    @api.model
    def _sentence(self, plan, applied=False):
        """ONE expression per sentence, so the spaces survive (R34)."""
        if applied:
            bits = [_("%(new)s added and %(upd)s updated.",
                      new=plan.get('created', 0), upd=plan.get('updated', 0))]
        elif not plan['writes']:
            bits = [_("There is nothing in that file to write — every month "
                      "cell is empty.")]
        else:
            bits = [_("%(new)s would be added and %(upd)s updated.",
                      new=plan['to_create'], upd=plan['to_update'])]
        if plan['unknown']:
            bits.append(_(
                "%(n)s named a department this system does not have, so they "
                "were left out: %(names)s.",
                n=counted(len(plan['unknown']), _("1 row"), _("%s rows")),
                names=', '.join(plan['unknown'][:6])))
        if plan['blanks']:
            bits.append(_("%s empty month cells were left exactly as they are.",
                          plan['blanks']))
        if plan['skipped']:
            bits.append(_("%s cells did not hold a number.",
                          len(plan['skipped'])))
        if plan['over_cap']:
            bits.append(_("The file is longer than %s rows; the rest was not "
                          "read.", FILE_ROW_CAP))
        return ' '.join(bits)

    # ============================================================ the native door
    def action_template(self):
        """The same template, for somebody using the plain form."""
        self.ensure_one()
        res = self.template_xlsx(self.fy, self.budget_type)
        self.write({'template_file': res['file_b64'],
                    'template_filename': res['filename']})
        return self._reopen()

    def action_peek(self):
        self.ensure_one()
        plan = self.peek(self.upload_file, self.fy, self.budget_type)
        self.summary = plan['message']
        return self._reopen()

    def action_apply(self):
        self.ensure_one()
        plan = self.apply(self.upload_file, self.fy, self.budget_type)
        self.summary = plan['message']
        return self._reopen()

    def _reopen(self):
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}
