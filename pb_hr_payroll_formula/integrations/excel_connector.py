# -*- coding: utf-8 -*-
"""
Excel Connector - Import payroll data from Excel/CSV files.
"""

import base64
import csv
import io
import json
import re
from collections import OrderedDict
from typing import Dict, List, Any, Optional, Tuple
import logging

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_connector import BaseHRConnector

_logger = logging.getLogger(__name__)

# The cell note on a generated template's first column. Not translated here on
# purpose — this module is plain python with no `_` in scope; the caller that
# has an environment passes a translated heading in (`pk_header`).
_TEMPLATE_PK_NOTE = (
    "The employee this row belongs to.\n"
    "Every row is matched on this column — leave it filled in."
)


class ExcelConnector(BaseHRConnector):
    """
    Excel/CSV file connector for importing payroll data.

    Supports:
    - Excel files (.xlsx, .xls)
    - CSV files
    - Auto-detection of headers
    - Column mapping wizard
    - Data preview
    """

    def __init__(self, connector_record=None):
        # ExcelConnector can work without a connector record for standalone usage
        # (e.g., parsing Excel files without Odoo integration)
        if connector_record is not None:
            super().__init__(connector_record)
        else:
            self.connector = None
            self.env = None
        self.workbook = None
        self.headers = []
        self.data_rows = []

    # ==========================================
    # AUTHENTICATION (Not required for Excel)
    # ==========================================

    def authenticate(self) -> bool:
        """
        Excel connector doesn't require authentication.
        """
        return True

    def test_connection(self) -> Tuple[bool, str]:
        """
        Test if Excel processing is available.
        """
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl library not installed. Install with: pip install openpyxl"

        return True, "Excel connector ready"

    # ==========================================
    # FILE PROCESSING
    # ==========================================

    def load_file(
        self,
        file_content: bytes,
        filename: str,
        header_row: int = None,
        data_start_row: int = None,
        sheet_name: str = None
    ) -> Dict[str, Any]:
        """
        Load Excel or CSV file content.

        Args:
            file_content: File content as bytes
            filename: Original filename
            header_row: Row number containing headers (1-based), overrides connector setting
            data_start_row: Row number where data starts (1-based), overrides connector setting
            sheet_name: Sheet name for Excel files, overrides connector setting

        Returns:
            Dictionary with 'headers' and 'rows' keys, or raises exception
        """
        # Store parameters for use in loading methods
        self._header_row = header_row
        self._data_start_row = data_start_row
        self._sheet_name = sheet_name

        try:
            if filename.lower().endswith('.csv'):
                success = self._load_csv(file_content)
            elif filename.lower().endswith(('.xlsx', '.xls')):
                success = self._load_excel(file_content)
            else:
                raise ValueError(f"Unsupported file format: {filename}")

            if success:
                return {
                    'headers': self.headers,
                    'rows': self.data_rows,
                    'total_rows': len(self.data_rows),
                    'total_columns': len(self.headers),
                }
            else:
                raise ValueError("Failed to load file")

        except Exception as e:
            _logger.exception(f"Failed to load file: {e}")
            raise

    def _load_csv(self, content: bytes) -> bool:
        """
        Load CSV file content.

        Args:
            content: CSV file content as bytes

        Returns:
            True if loaded successfully
        """
        try:
            # Try different encodings
            text = None
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if text is None:
                _logger.error("Failed to decode CSV file")
                return False

            # Parse CSV
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)

            if not rows:
                return False

            # Get header row - prefer parameter, then connector setting, then default
            header_row = self._header_row or (getattr(self.connector, 'file_header_row', None) if self.connector else None) or 1
            data_start_row = self._data_start_row or (getattr(self.connector, 'file_data_start_row', None) if self.connector else None) or 2

            self.headers = rows[header_row - 1] if len(rows) >= header_row else []
            self.data_rows = rows[data_start_row - 1:] if len(rows) >= data_start_row else []

            # Update connector with filename if available
            if self.connector and hasattr(self.connector, 'write'):
                try:
                    self.connector.write({
                        'last_import_filename': 'imported.csv',
                        'connection_status': 'connected',
                    })
                except Exception:
                    pass  # Ignore if connector can't be updated

            return True

        except Exception as e:
            _logger.exception(f"CSV loading failed: {e}")
            return False

    def _load_excel(self, content: bytes) -> bool:
        """
        Load Excel file content.

        Args:
            content: Excel file content as bytes

        Returns:
            True if loaded successfully
        """
        if not OPENPYXL_AVAILABLE:
            _logger.error("openpyxl not available")
            return False

        try:
            # Load workbook
            self.workbook = openpyxl.load_workbook(
                io.BytesIO(content),
                data_only=True  # Get calculated values
            )

            # Get sheet - prefer parameter, then connector setting
            sheet_name = self._sheet_name or (getattr(self.connector, 'file_sheet_name', None) if self.connector else None)
            if sheet_name and sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.active
                sheet_name = sheet.title

            # Get header row - prefer parameter, then connector setting, then default
            header_row = self._header_row or (getattr(self.connector, 'file_header_row', None) if self.connector else None) or 1
            data_start_row = self._data_start_row or (getattr(self.connector, 'file_data_start_row', None) if self.connector else None) or 2

            # Use the same header detection logic as formula import when defaults are used
            from ..formula_engine.header_detector import HeaderDetector
            detector = HeaderDetector(sheet)
            if header_row == 1 and data_start_row == 2:
                detected_header_row, detected_data_start_row, _details = detector.detect_header_row()
                if detected_header_row and detected_header_row != header_row:
                    _logger.info(
                        "Auto-detected header row %s (was %s) for sheet '%s'",
                        detected_header_row,
                        header_row,
                        sheet_name,
                    )
                    header_row = detected_header_row
                if detected_data_start_row:
                    data_start_row = detected_data_start_row

            if data_start_row <= header_row:
                data_start_row = header_row + 1

            # Extract headers with merge-aware logic
            raw_headers = detector.get_headers(header_row)
            headers_by_index = {}
            for header in raw_headers:
                value = header.get('value')
                if value is not None and str(value).strip():
                    headers_by_index[header['column_index']] = str(value).strip()

            self.headers = []
            for idx, cell in enumerate(sheet[header_row]):
                header_value = headers_by_index.get(idx)
                if not header_value:
                    cell_val = cell.value
                    if cell_val is not None and str(cell_val).strip():
                        header_value = str(cell_val).strip()
                if not header_value:
                    header_value = f"Column_{cell.column}"
                self.headers.append(header_value)

            # Extract data rows
            self.data_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
                row_data = [cell.value for cell in row]
                # Skip empty rows
                if any(v is not None for v in row_data):
                    self.data_rows.append(row_data)

            # Update connector if available
            if self.connector and hasattr(self.connector, 'write'):
                try:
                    self.connector.write({
                        'last_import_filename': sheet_name or 'Sheet1',
                        'connection_status': 'connected',
                    })
                except Exception:
                    pass  # Ignore if connector can't be updated

            return True

        except Exception as e:
            _logger.exception(f"Excel loading failed: {e}")
            return False

    # ==========================================
    # FIELD DISCOVERY
    # ==========================================

    def get_available_fields(self) -> List[Dict[str, Any]]:
        """
        Get list of available columns from the loaded file.

        Returns:
            List of field definitions based on headers
        """
        fields = []

        for idx, header in enumerate(self.headers):
            # Try to detect data type from first few rows
            data_type = self._detect_column_type(idx)

            # Get sample value
            sample = None
            if self.data_rows and len(self.data_rows[0]) > idx:
                sample = self.data_rows[0][idx]

            field = {
                'name': header,
                'label': header,
                'data_type': data_type,
                'path': header,
                'column_index': idx,
                'sample_value': sample,
            }
            fields.append(field)

        return fields

    def _detect_column_type(self, column_index: int) -> str:
        """
        Detect the data type of a column based on sample values.

        Args:
            column_index: Column index

        Returns:
            Detected data type: 'string', 'number', 'date', or 'boolean'
        """
        samples = []
        for row in self.data_rows[:10]:
            if len(row) > column_index and row[column_index] is not None:
                samples.append(row[column_index])

        if not samples:
            return 'string'

        # Check if all samples are numbers
        is_numeric = True
        for sample in samples:
            if isinstance(sample, (int, float)):
                continue
            if isinstance(sample, str):
                try:
                    float(sample.replace(',', '').replace(' ', ''))
                    continue
                except ValueError:
                    is_numeric = False
                    break
            else:
                is_numeric = False
                break

        if is_numeric:
            return 'number'

        # Check if all samples are dates (openpyxl returns datetime for date cells)
        from datetime import datetime
        if all(isinstance(s, datetime) for s in samples):
            return 'date'

        # Check if all samples are boolean
        bool_values = {'true', 'false', 'yes', 'no', '1', '0'}
        if all(str(s).lower() in bool_values for s in samples):
            return 'boolean'

        return 'string'

    # ==========================================
    # DATA FETCHING
    # ==========================================

    def fetch_employees(self, filters: Optional[Dict] = None) -> List[Dict[str, Any]]:
        """
        Fetch employee data from loaded file.

        Each row is treated as an employee record.

        Args:
            filters: Optional filter criteria

        Returns:
            List of employee data dictionaries
        """
        employees = []

        for row_idx, row in enumerate(self.data_rows):
            # Create dictionary from row
            emp = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(self.headers):
                    header = self.headers[col_idx]
                    emp[header] = value

            # Add row identifier
            emp['_row_index'] = row_idx + 1
            emp['id'] = str(row_idx + 1)

            employees.append(emp)

        return employees

    def fetch_payroll_data(
        self,
        employee_ids: List[str],
        date_from: str,
        date_to: str
    ) -> Dict[str, Dict[str, Any]]:
        """
        For Excel connector, all data is already in fetch_employees.
        This returns empty as payroll data is combined with employee data.

        Args:
            employee_ids: List of row indices as strings
            date_from: Ignored for Excel
            date_to: Ignored for Excel

        Returns:
            Empty dict - use fetch_employees instead
        """
        return {}

    # ==========================================
    # PREVIEW AND VALIDATION
    # ==========================================

    def get_data_preview(self, max_rows: int = 10) -> Dict[str, Any]:
        """
        Get a preview of the loaded data.

        Args:
            max_rows: Maximum number of rows to return

        Returns:
            Preview data with headers and sample rows
        """
        return {
            'headers': self.headers,
            'rows': self.data_rows[:max_rows],
            'total_rows': len(self.data_rows),
            'total_columns': len(self.headers),
        }

    def validate_data(self, mappings: List[Any]) -> Dict[str, Any]:
        """
        Validate loaded data against field mappings.

        Args:
            mappings: List of hr.integration.field.mapping records

        Returns:
            Validation result with errors and warnings
        """
        result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'stats': {
                'total_rows': len(self.data_rows),
                'empty_values': 0,
                'type_mismatches': 0,
            }
        }

        required_fields = [m.source_field for m in mappings if m.is_required]

        # Check required fields exist
        for field in required_fields:
            if field not in self.headers:
                result['errors'].append(f"Required field '{field}' not found in file")
                result['valid'] = False

        # Check data quality
        for row_idx, row in enumerate(self.data_rows):
            for mapping in mappings:
                field = mapping.source_field
                if field not in self.headers:
                    continue

                col_idx = self.headers.index(field)
                value = row[col_idx] if col_idx < len(row) else None

                # Check empty values for required fields
                if mapping.is_required and (value is None or value == ''):
                    result['stats']['empty_values'] += 1
                    if result['stats']['empty_values'] <= 5:
                        result['warnings'].append(
                            f"Row {row_idx + 1}: Required field '{field}' is empty"
                        )

                # Check type compatibility
                if value is not None and mapping.source_data_type == 'number':
                    try:
                        if isinstance(value, str):
                            float(value.replace(',', ''))
                    except ValueError:
                        result['stats']['type_mismatches'] += 1
                        if result['stats']['type_mismatches'] <= 5:
                            result['warnings'].append(
                                f"Row {row_idx + 1}: Field '{field}' has non-numeric value"
                            )

        return result

    # ==========================================
    # AUTO-MAPPING
    # ==========================================

    def suggest_mappings(self, rules: List[Any]) -> List[Dict[str, Any]]:
        """
        Suggest field mappings based on header names and rule codes.

        Args:
            rules: List of hr.formula.rule records

        Returns:
            List of suggested mappings
        """
        suggestions = []

        for header in self.headers:
            header_upper = header.upper().replace(' ', '_').replace('-', '_')

            best_match = None
            best_score = 0

            for rule in rules:
                if rule.column_type != 'input':
                    continue

                # Calculate similarity score
                score = self._calculate_similarity(header_upper, rule.code)

                # Also check against rule name
                name_score = self._calculate_similarity(
                    header_upper,
                    rule.name.upper().replace(' ', '_')
                )
                score = max(score, name_score)

                if score > best_score and score > 0.5:
                    best_score = score
                    best_match = rule

            if best_match:
                suggestions.append({
                    'source_field': header,
                    'target_rule_id': best_match.id,
                    'target_code': best_match.code,
                    'confidence': best_score,
                })

        return suggestions

    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """
        Calculate similarity between two strings.

        Uses simple character-based similarity.

        Args:
            str1: First string
            str2: Second string

        Returns:
            Similarity score between 0 and 1
        """
        if not str1 or not str2:
            return 0.0

        # Exact match
        if str1 == str2:
            return 1.0

        # Contains match
        if str1 in str2 or str2 in str1:
            return 0.8

        # Partial match
        set1 = set(str1)
        set2 = set(str2)
        intersection = set1.intersection(set2)

        if not set1 or not set2:
            return 0.0

        return len(intersection) / max(len(set1), len(set2))

    # ==========================================
    # FILE GENERATION
    # ==========================================

    @staticmethod
    def template_slot_for(rule) -> Tuple[str, str]:
        """Where a component's column belongs in a generated template.

        Returns `(sheet, header)` — and the header is deliberately **the key
        the resolver will match on re-import**, not a label:

          * an explicit spreadsheet binding wins. If somebody has already said
            "this component reads the column called `Gross OT`", the template
            has to say `Gross OT`, or downloading the template and filling it
            in would quietly stop feeding the component it was built for;
          * otherwise the component NAME, which is the candidate
            `_transform_data_to_formula_inputs` tries after the binding and
            before the code. The name is what a payroll officer recognises in
            a column heading; the code is an internal handle.

        A binding key may itself be sheet-qualified (`SEVL|Gross OT`) because
        that is the shape a multisheet load produces; the sheet half of the key
        wins over the rule's own `source_sheet_name` when they disagree, since
        the key is the thing that was actually matched.
        """
        binding = (getattr(rule, 'source_binding', '') or '')
        key = (getattr(rule, 'source_binding_key', '') or '').strip() if binding == 'excel' else ''
        sheet = (getattr(rule, 'source_sheet_name', '') or '').strip()
        if key:
            if '|' in key:
                bound_sheet, bound_header = key.split('|', 1)
                return (bound_sheet.strip() or sheet), bound_header.strip()
            return sheet, key
        return sheet, ((rule.name or rule.code or '').strip())

    def generate_template(
        self,
        rules: List[Any],
        pk_header: str = 'Employee Code',
        sheet_title: str = 'Payroll Data',
    ) -> bytes:
        """
        Generate the pay-data workbook a scheme will actually READ BACK.

        This had zero callers for its whole life, and the reason is visible in
        what it used to emit: one flat sheet of component CODES with a row of
        default values underneath. Codes are not what the resolver tries first,
        a multisheet scheme's columns do not live on one sheet, and a template
        that arrives with a row of data in it is a template somebody imports by
        accident. All three are now fixed:

          * one column per INPUT component, headed by the key the resolver
            matches (`template_slot_for`);
          * the employee-identifier column FIRST on every sheet — it is what
            each row is matched by, and on a multisheet workbook it is what the
            sheets are merged on, so a sheet without it contributes nothing;
          * one sheet per `source_sheet_name` when the scheme is multisheet,
            one sheet otherwise;
          * **headings only, no data row.** The file is empty on purpose: the
            person downloading it is going to fill it with this month's
            numbers, and a pre-filled example row is a row that gets loaded,
            matched and — one more click — written into somebody's payslip.

        Args:
            rules: hr.formula.rule records (the scheme's whole rule set)
            pk_header: the employee-identifier column heading to put first
            sheet_title: sheet name for a single-sheet scheme

        Returns:
            Excel file content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for template generation")

        from openpyxl.utils import get_column_letter

        input_rules = [r for r in rules if r.column_type == 'input']
        pk_header = (pk_header or 'Employee Code').strip() or 'Employee Code'

        # group by sheet, preserving each scheme's own component order
        sheets: "OrderedDict[str, List[Tuple[str, Any]]]" = OrderedDict()
        for rule in input_rules:
            sheet, header = self.template_slot_for(rule)
            if not header:
                continue
            sheets.setdefault(sheet, []).append((header, rule))
        multisheet = any(name for name in sheets)
        if not multisheet:
            merged = []
            for cols in sheets.values():
                merged.extend(cols)
            sheets = OrderedDict([(sheet_title, merged)] if merged else [])

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        bold = openpyxl.styles.Font(bold=True)
        head_fill = openpyxl.styles.PatternFill('solid', fgColor='EAF1FB')

        if not sheets:
            sheets = OrderedDict([(sheet_title, [])])

        for name, cols in sheets.items():
            # openpyxl refuses >31 chars and []:*?/\ in a sheet title
            safe = re.sub(r'[\[\]:*?/\\]', ' ', str(name or sheet_title)).strip()[:31] \
                or sheet_title
            ws = wb.create_sheet(title=safe)
            seen = {pk_header.lower()}
            headers = [(pk_header, None)]
            for header, rule in cols:
                if header.lower() in seen:
                    continue
                seen.add(header.lower())
                headers.append((header, rule))
            for col_idx, (header, rule) in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = bold
                cell.fill = head_fill
                # NB: the body this replaced read `rule.description`, and
                # `hr.formula.rule` has no such field — so the generator did
                # not merely lack a caller, it could not have survived one.
                # Everything named here is a field that exists.
                note = (_TEMPLATE_PK_NOTE if rule is None
                        else '\n'.join(str(x) for x in
                                       [rule.name, rule.code, rule.column_letter] if x))
                if note:
                    cell.comment = openpyxl.comments.Comment(text=note, author='Payobook')
                ws.column_dimensions[get_column_letter(col_idx)].width = \
                    max(12, min(38, len(header) + 4))
            ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ==========================================
    # MULTI-SHEET LOADING (for Multi-Worksheet Import)
    # ==========================================

    def load_workbook_multisheet(
        self,
        file_content: bytes,
        include_formulas: bool = True
    ) -> Dict[str, Any]:
        """
        Load all sheets from an Excel workbook with metadata.

        Args:
            file_content: Excel file content as bytes
            include_formulas: If True, also load formula workbook

        Returns:
            Dictionary with workbook metadata and sheet information
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for multi-sheet loading")

        result = {
            'sheet_names': [],
            'sheets': {},
            'active_sheet': None,
            'total_sheets': 0,
        }

        try:
            # Load data workbook (with calculated values)
            self.workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=True
            )

            # Optionally load formula workbook
            formula_workbook = None
            if include_formulas:
                formula_workbook = openpyxl.load_workbook(
                    io.BytesIO(file_content),
                    data_only=False
                )

            result['sheet_names'] = self.workbook.sheetnames
            result['total_sheets'] = len(self.workbook.sheetnames)
            result['active_sheet'] = self.workbook.active.title

            # Analyze each sheet
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                formula_sheet = formula_workbook[sheet_name] if formula_workbook else None

                sheet_info = self._analyze_sheet(sheet, formula_sheet)
                sheet_info['name'] = sheet_name
                result['sheets'][sheet_name] = sheet_info

            _logger.info(
                f"Loaded workbook with {result['total_sheets']} sheets: "
                f"{', '.join(result['sheet_names'])}"
            )

            return result

        except Exception as e:
            _logger.exception(f"Multi-sheet loading failed: {e}")
            raise

    def _analyze_sheet(
        self,
        sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        formula_sheet: Optional['openpyxl.worksheet.worksheet.Worksheet'] = None
    ) -> Dict[str, Any]:
        """
        Analyze a single worksheet for metadata.

        Args:
            sheet: Data worksheet
            formula_sheet: Formula worksheet (same sheet, loaded with formulas)

        Returns:
            Sheet analysis dictionary
        """
        from ..formula_engine.header_detector import HeaderDetector
        from ..formula_engine.merged_cell_parser import MergedCellParser

        # Basic dimensions
        analysis = {
            'max_row': sheet.max_row or 0,
            'max_column': sheet.max_column or 0,
            'merged_cell_count': len(list(sheet.merged_cells.ranges)),
            'has_formulas': False,
            'references_other_sheets': False,
            'detected_header_row': 1,
            'detected_data_start_row': 2,
            'header_detection_confidence': 0.0,
            'component_types': {},
            'formulas_referencing_sheets': [],
        }

        # Detect header row
        try:
            detector = HeaderDetector(sheet)
            detection = detector.detect_with_confidence()
            analysis['detected_header_row'] = detection['header_row']
            analysis['detected_data_start_row'] = detection['data_start_row']
            analysis['header_detection_confidence'] = detection['confidence_score']
            analysis['headers'] = [h['value'] for h in detection['headers'] if h['value']]
        except Exception as e:
            _logger.warning(f"Header detection failed: {e}")
            analysis['headers'] = []

        # Extract component types from merged cells
        try:
            parser = MergedCellParser(sheet)
            analysis['component_types'] = parser.extract_component_types(
                analysis['detected_header_row']
            )
            analysis['merged_structure'] = parser.analyze_structure()
        except Exception as e:
            _logger.warning(f"Merged cell parsing failed: {e}")

        # Check for formulas and cross-sheet references
        if formula_sheet:
            import re
            formula_count = 0
            cross_ref_count = 0

            for row in formula_sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        analysis['has_formulas'] = True
                        formula_count += 1

                        # Log first few formulas for debugging
                        if formula_count <= 3:
                            _logger.info(f"Sample formula in sheet {formula_sheet.title}: {cell.value}")

                        # Check for cross-sheet references with multiple patterns
                        # Pattern 1: Quoted sheet names - 'Sheet Name'!A1
                        quoted_pattern = r"'([^']+)'!"
                        quoted_refs = re.findall(quoted_pattern, cell.value)

                        # Pattern 2: Unquoted sheet names - SheetName!A1 or Sheet_Name!A1
                        # This pattern looks for word characters (letters, numbers, underscores) followed by !
                        # But excludes cell references like A1!
                        unquoted_pattern = r"(?<!['\w])([A-Za-z][A-Za-z0-9_\-]*)\s*!"
                        unquoted_refs = re.findall(unquoted_pattern, cell.value)

                        # Combine both patterns
                        all_refs = quoted_refs + unquoted_refs

                        if all_refs:
                            analysis['references_other_sheets'] = True
                            cross_ref_count += 1
                            for ref in all_refs:
                                # Clean up the reference and avoid duplicates
                                ref = ref.strip()
                                if ref and ref not in analysis['formulas_referencing_sheets']:
                                    # Exclude common Excel functions that might match the pattern
                                    if ref.upper() not in ['IF', 'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN',
                                                           'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']:
                                        analysis['formulas_referencing_sheets'].append(ref)
                                        _logger.info(f"Found cross-sheet reference: '{ref}' in formula: {cell.value}")

            _logger.info(f"Sheet {formula_sheet.title}: Found {formula_count} formulas, {cross_ref_count} with cross-sheet refs")
            _logger.info(f"Sheet {formula_sheet.title}: Referenced sheets: {analysis['formulas_referencing_sheets']}")

        return analysis

    def load_sheet_with_detection(
        self,
        sheet_name: str,
        auto_detect_header: bool = True
    ) -> Dict[str, Any]:
        """
        Load a specific sheet with automatic header detection.

        Args:
            sheet_name: Name of sheet to load
            auto_detect_header: Whether to auto-detect header row

        Returns:
            Dictionary with headers, data, and metadata
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded. Call load_workbook_multisheet first.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = self.workbook[sheet_name]

        from openpyxl.utils import get_column_letter
        from ..formula_engine.header_detector import HeaderDetector
        from ..formula_engine.merged_cell_parser import MergedCellParser

        # Always create detector (needed for get_headers even if not auto-detecting)
        detector = HeaderDetector(sheet)

        # Detect header row
        if auto_detect_header:
            header_row, data_start_row, details = detector.detect_header_row()
        else:
            header_row = 1
            data_start_row = 2
            details = {}

        # Extract component types from horizontal merges above header
        parser = MergedCellParser(sheet)
        component_types = parser.extract_component_types(header_row)
        column_info = parser.get_column_info(header_row)

        # Extract headers using HeaderDetector.get_headers() to handle vertical merges
        # This properly extracts headers from:
        # 1. Regular cells in the header row
        # 2. Vertical merges that span down to the header row (e.g., "TT", "MSNV" in row 1-2)
        raw_headers = detector.get_headers(header_row)

        _logger.info(
            f"load_sheet_with_detection: sheet='{sheet_name}', "
            f"header_row={header_row}, data_start_row={data_start_row}, "
            f"raw_headers_count={len(raw_headers)}, "
            f"component_types_count={len(component_types)}"
        )

        headers = []
        for h in raw_headers:
            if h.get('value'):
                col_letter = h.get('column_letter')
                headers.append({
                    'column_letter': col_letter,
                    'column_index': h.get('column_index', 0),
                    'value': h['value'],
                    'component_type': component_types.get(col_letter),
                    'from_vertical_merge': h.get('from_vertical_merge', False),
                })

        _logger.info(f"  -> Filtered headers with values: {len(headers)}")

        # Extract data rows
        data_rows = []
        for row in sheet.iter_rows(min_row=data_start_row):
            row_data = {}
            for cell in row:
                col_letter = get_column_letter(cell.column)
                header = next(
                    (h for h in headers if h['column_letter'] == col_letter),
                    None
                )
                if header:
                    row_data[header['value']] = cell.value

            # Skip empty rows
            if any(v is not None for v in row_data.values()):
                data_rows.append(row_data)

        return {
            'sheet_name': sheet_name,
            'header_row': header_row,
            'data_start_row': data_start_row,
            'detection_details': details,
            'headers': headers,
            'component_types': component_types,
            'data_rows': data_rows,
            'total_rows': len(data_rows),
            'total_columns': len(headers),
        }

    def get_sheet_formulas(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Extract all formulas from a specific sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            List of formula information dictionaries
        """
        if not OPENPYXL_AVAILABLE:
            return []

        formulas = []

        # Need to reload with formulas
        # This assumes we have access to the original file content
        # In practice, this would be called during the wizard flow

        _logger.warning("get_sheet_formulas requires formula workbook - not implemented in this context")
        return formulas

    def build_sheet_column_mapping(
        self,
        sheet_name: str,
        code_generator: callable = None
    ) -> Dict[str, str]:
        """
        Build column letter to code mapping for a sheet.

        Used for cross-sheet formula resolution.

        Args:
            sheet_name: Name of the sheet
            code_generator: Optional function to generate codes from headers.
                           Signature: (header: str, existing_codes: set) -> str

        Returns:
            Dictionary mapping column letters to codes
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {}

        sheet_data = self.load_sheet_with_detection(sheet_name)
        existing_codes = set()
        column_mapping = {}

        for header in sheet_data['headers']:
            col_letter = header['column_letter']
            header_value = header['value']

            if code_generator:
                code = code_generator(header_value, existing_codes)
            else:
                code = self._generate_code_from_header(header_value, existing_codes)

            column_mapping[col_letter] = code
            existing_codes.add(code)

        return column_mapping

    def _generate_code_from_header(
        self,
        header: str,
        existing_codes: set
    ) -> str:
        """
        Generate a readable, converter-safe code from a header value.

        This is the legacy fallback used only when no ``code_generator`` is injected,
        and until MAPFIX it emitted UNDERSCORES (``COL_3``, ``BASIC_1``) — a live
        breach of the converter contract, because the code pass matches
        ``[A-Z][A-Z0-9]{1,}`` and a token carrying ``_`` reaches the eval raw and
        reads as zero. It now delegates to the one shared generator.

        Args:
            header: Header value
            existing_codes: Set of already used codes

        Returns:
            Generated unique code
        """
        from ..models import component_code

        header_str = str(header).strip()
        if header_str.isdigit():
            header_str = 'COL' + header_str

        return component_code.build_component_code(
            header_str, existing_codes=existing_codes)

    def analyze_cross_sheet_formulas(
        self,
        file_content: bytes
    ) -> Dict[str, Any]:
        """
        Analyze formulas across all sheets to identify dependencies.

        Args:
            file_content: Excel file content

        Returns:
            Cross-sheet dependency analysis
        """
        from ..formula_engine.cross_sheet_resolver import CrossSheetResolver
        import re

        # Load formula workbook
        formula_wb = openpyxl.load_workbook(
            io.BytesIO(file_content),
            data_only=False
        )

        analysis = {
            'sheets': {},
            'cross_references': [],
            'dependency_graph': {},
        }

        cross_sheet_pattern = re.compile(r"'([^']+)'!\$?([A-Z]+)\$?(\d*)")

        for sheet_name in formula_wb.sheetnames:
            sheet = formula_wb[sheet_name]
            sheet_analysis = {
                'formula_count': 0,
                'cross_refs': [],
                'depends_on_sheets': set(),
            }

            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_analysis['formula_count'] += 1

                        # Find cross-sheet references
                        matches = cross_sheet_pattern.findall(cell.value)
                        for ref_sheet, col, row_num in matches:
                            ref_info = {
                                'source_sheet': sheet_name,
                                'source_cell': f"{cell.column_letter}{cell.row}",
                                'target_sheet': ref_sheet,
                                'target_column': col,
                                'target_row': row_num or 'any',
                                'formula': cell.value,
                            }
                            sheet_analysis['cross_refs'].append(ref_info)
                            sheet_analysis['depends_on_sheets'].add(ref_sheet)
                            analysis['cross_references'].append(ref_info)

            sheet_analysis['depends_on_sheets'] = list(sheet_analysis['depends_on_sheets'])
            analysis['sheets'][sheet_name] = sheet_analysis
            analysis['dependency_graph'][sheet_name] = sheet_analysis['depends_on_sheets']

        return analysis
