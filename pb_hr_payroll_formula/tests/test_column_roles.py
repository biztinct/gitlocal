# -*- coding: utf-8 -*-
"""Column roles and typed contract components, end to end.

The classifier itself is covered without a database in
`test_column_role_classifier.py`. What is exercised here is everything the
classifier's answer then TOUCHES: the colour-coded import that assigns roles, the
role defaults that hide people-data columns from the payslip, the text-typed contract
component that must not be run through the amount bound check, and the promise that
none of it changed what an existing payroll computes (CR-A7).

The workbooks are generated in-memory rather than committed as binaries, so the
colour conventions being tested are stated in code, right next to the assertion about
what they should mean.
"""

import base64
import io

from odoo.exceptions import ValidationError
from odoo.tests import common, tagged


def _build_color_workbook():
    """A miniature color-coded payroll workbook.

    Row 1  identifier row  (payslip section codes)
    Row 2  header band     (yellow fill; font colour carries the component semantics)
    Row 3  formula row     (green fill)
    Row 4+ data
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')
    red_font = Font(color='FFFF0000')
    # CR26 — Excel's standard "Green" (#00B050), NOT the darker HTML green
    # (#008000): `is_green_font` requires g > 150, and #008000 (g = 128) is below
    # that threshold, so the original fixture asserted a marker the importer was
    # never going to see.
    green_font = Font(color='FF00B050')
    green_underline = Font(color='FF00B050', underline='single')

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        # (header, font, formula, data values)
        ('Employee Code', None, None, ['E001', 'E002']),
        ('Bank Name', None, None, ['Vietcombank', 'Techcombank']),
        ('Base Salary', red_font, None, [10000000, 12000000]),
        ('Cost Centre Ref', green_font, None, ['CC-01', 'CC-02']),
        ('Job Grade', green_underline, None, ['G3', 'G4']),
        ('Gross Pay', None, '=C3', [None, None]),
    ]

    for idx, (header, font, formula, values) in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value='SECTION1')
        cell = ws.cell(row=2, column=idx, value=header)
        cell.fill = yellow
        if font:
            cell.font = font
        formula_cell = ws.cell(row=3, column=idx, value=formula)
        formula_cell.fill = green_fill
        for offset, value in enumerate(values):
            if value is not None:
                ws.cell(row=4 + offset, column=idx, value=value)

    stream = io.BytesIO()
    wb.save(stream)
    return base64.b64encode(stream.getvalue())


@tagged('post_install', '-at_install')
class TestColumnRoles(common.TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        # CR19 — `country_code` is required with no default, so a bare create dies
        # at INSERT time. This class predates that finding; without the country the
        # whole suite fails in setUpClass, not in a test.
        cls.config = cls.env['hr.formula.config'].create({
            'name': 'COLROLES Test Structure',
            'country_code': 'VN',
            'use_color_coded_excel_import': True,
        })

    # ------------------------------------------------------------------
    # 8 / 9 — colour-coded import assigns roles and component kinds
    # ------------------------------------------------------------------
    def test_08_color_import_assigns_roles(self):
        wizard = self.env['hr.formula.import.wizard'].create({
            'config_id': self.config.id,
            'import_file': _build_color_workbook(),
            'import_filename': 'colroles.xlsx',
        })
        wizard._import_from_excel()

        rules = {r.name: r for r in self.config.rule_ids}
        self.assertIn('Employee Code', rules)

        code_rule = rules['Employee Code']
        self.assertEqual(code_rule.column_role, 'identity')
        self.assertEqual(code_rule.column_role_source, 'auto')
        self.assertFalse(code_rule.appears_on_payslip)
        self.assertFalse(code_rule.is_visible_in_grid)

        self.assertEqual(rules['Bank Name'].column_role, 'bank')

        salary = rules['Base Salary']
        self.assertTrue(salary.is_contract_component)
        self.assertFalse(salary.is_text_component)
        self.assertEqual(salary.column_role, 'payroll')
        self.assertTrue(salary.appears_on_payslip)

        cost_centre = rules['Cost Centre Ref']
        self.assertTrue(cost_centre.is_contract_component)
        self.assertTrue(cost_centre.is_text_component)
        self.assertEqual(cost_centre.column_role, 'contract')
        self.assertFalse(cost_centre.appears_on_payslip)

        # 9 — green + underline still means "start a new contract on change".
        grade = rules['Job Grade']
        self.assertTrue(grade.is_text_component)
        self.assertTrue(grade.requires_new_contract)

    # ------------------------------------------------------------------
    # role bookkeeping
    # ------------------------------------------------------------------
    def test_09_manual_edit_marks_role_as_user_set(self):
        rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Manual Column',
            'code': 'MANUALCOLUMN',
            'column_type': 'input',
        })
        self.assertEqual(rule.column_role, 'payroll')
        self.assertEqual(rule.column_role_source, 'auto')

        rule.write({'column_role': 'bank'})
        self.assertEqual(rule.column_role_source, 'user')

        # An automatic writer names the source itself and so cannot be mistaken for
        # a person — and must leave a user-set row alone.
        rule.write({'column_role': 'profile', 'column_role_source': 'auto'})
        self.assertEqual(rule.column_role_source, 'auto')

    def test_10_reclassify_rpc_respects_user_choice(self):
        auto_rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Bank Name',
            'code': 'BANKNAMERPC',
            'column_type': 'input',
        })
        user_rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Account Number',
            'code': 'ACCOUNTNORPC',
            'column_type': 'input',
            'column_role': 'payroll',
            'column_role_source': 'user',
        })
        result = self.env['pb.formula.studio'].reclassify_roles(self.config.id)
        self.assertTrue(result['ok'])
        moved = {entry['code']: entry['to'] for entry in result['changed']}
        self.assertEqual(moved.get('BANKNAMERPC'), 'bank')
        self.assertNotIn('ACCOUNTNORPC', moved)
        self.assertEqual(auto_rule.column_role, 'bank')
        self.assertEqual(user_rule.column_role, 'payroll')

    # ------------------------------------------------------------------
    # 11 / 12 — typed contract components
    # ------------------------------------------------------------------
    def test_11_text_component_line_skips_bound_check(self):
        template = self.env['hr.contract.advantage.template'].create({
            'name': 'Grade',
            'code': 'GRADETEXT',
            'lower_bound': 100.0,
            'upper_bound': 200.0,
            'value_type': 'text',
        })
        line = self.env['hr.contract.advantage'].create({
            'advantage_template_id': template.id,
            'text_value': 'G4',
        })
        self.assertEqual(line.value_type, 'text')
        self.assertEqual(line.text_value, 'G4')

        amount_template = self.env['hr.contract.advantage.template'].create({
            'name': 'Bounded Allowance',
            'code': 'BOUNDEDALLOW',
            'lower_bound': 100.0,
            'upper_bound': 200.0,
        })
        self.assertEqual(amount_template.value_type, 'amount')
        with self.assertRaises(ValidationError):
            self.env['hr.contract.advantage'].create({
                'advantage_template_id': amount_template.id,
                'amount': 5000.0,
            })

    def test_12_change_log_records_text(self):
        change_fields = self.env['hr.contract.advantage.change']._fields
        self.assertIn('old_text_value', change_fields)
        self.assertIn('new_text_value', change_fields)

    # ------------------------------------------------------------------
    # 14 — employee-code recognition
    # ------------------------------------------------------------------
    def test_14_is_employee_code_rule(self):
        batch = self.env['hr.payroll.import.batch'].new({
            'formula_config_id': self.config.id,
        })
        identity_rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Personnel Number',
            'code': 'PERSONNELNUM',
            'column_type': 'input',
            'column_role': 'identity',
            'column_role_source': 'user',
        })
        marker_rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'MSNV',
            'code': 'MSNVCOL',
            'column_type': 'input',
        })
        plain_rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Gas Allowance',
            'code': 'GASALLOWTEST',
            'column_type': 'input',
        })
        # Role short-circuits...
        self.assertTrue(batch._is_employee_code_rule(identity_rule))
        # ...and the original marker heuristic still stands on its own.
        self.assertTrue(batch._is_employee_code_rule(marker_rule))
        self.assertFalse(batch._is_employee_code_rule(plain_rule))

    # ------------------------------------------------------------------
    # 15 / 16 — reclassification as a REVIEW (COLROLES P4)
    # ------------------------------------------------------------------
    def test_15_reclassify_dry_run_writes_nothing(self):
        rule = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Bank Name',
            'code': 'BANKNAMEDRY',
            'column_type': 'input',
            'column_role': 'payroll',
        })
        result = self.env['pb.formula.studio'].reclassify_roles(
            self.config.id, dry_run=True)
        self.assertTrue(result['dry_run'])
        self.assertEqual(result['applied'], [])
        moved = {entry['code']: entry for entry in result['changed']}
        self.assertEqual(moved['BANKNAMEDRY']['to'], 'bank')
        # The proposal is a proposal: the row itself has not moved.
        self.assertEqual(rule.column_role, 'payroll')
        # ...and it carries the words a person reads, not just the keys.
        self.assertTrue(moved['BANKNAMEDRY']['to_label'])
        self.assertTrue(moved['BANKNAMEDRY']['from_label'])

    def test_16_reclassify_applies_only_accepted_rows(self):
        accepted = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Bank Name',
            'code': 'BANKNAMEACC',
            'column_type': 'input',
            'column_role': 'payroll',
        })
        skipped = self.env['hr.formula.rule'].create({
            'config_id': self.config.id,
            'name': 'Account Number',
            'code': 'ACCOUNTNOSKIP',
            'column_type': 'input',
            'column_role': 'payroll',
        })
        preview = self.env['pb.formula.studio'].reclassify_roles(
            self.config.id, dry_run=True)
        proposed = {entry['code']: entry['id'] for entry in preview['changed']}
        self.assertIn('BANKNAMEACC', proposed)
        self.assertIn('ACCOUNTNOSKIP', proposed)

        result = self.env['pb.formula.studio'].reclassify_roles(
            self.config.id, False, [proposed['BANKNAMEACC']])
        self.assertEqual(result['applied'], [proposed['BANKNAMEACC']])
        self.assertEqual(accepted.column_role, 'bank')
        # Accepting a machine reading is agreeing with it, not authoring it.
        self.assertEqual(accepted.column_role_source, 'auto')
        # The skipped row keeps BOTH its role and its source, so a later re-run
        # can still offer it.
        self.assertEqual(skipped.column_role, 'payroll')
        self.assertEqual(skipped.column_role_source, 'auto')

    # ------------------------------------------------------------------
    # 17 — role-driven export columns are strictly opt-in
    # ------------------------------------------------------------------
    def test_17_export_base_columns_follow_the_flag(self):
        config = self.env['hr.formula.config'].create({
            'name': 'COLROLES Export Structure',
            'country_code': 'VN',
        })
        run = self.env['hr.payslip.run'].new({'name': 'COLROLES Export Run'})

        fixed = run._export_base_columns(config)
        self.assertEqual([c['header'] for c in fixed],
                         ['MSNV', 'Full name', 'Unit', 'Type of labor contract',
                          'Subjects are counted as working overtime'])

        staff = self.env['hr.formula.rule'].create({
            'config_id': config.id,
            'name': 'Staff No', 'code': 'STAFFNOEXP',
            'column_type': 'input',
            'column_role': 'identity', 'column_role_source': 'user',
        })
        division = self.env['hr.formula.rule'].create({
            'config_id': config.id,
            'name': 'Division', 'code': 'DIVISIONEXP',
            'column_type': 'input',
            'column_role': 'profile', 'column_role_source': 'user',
        })
        self.env['hr.formula.rule'].create({
            'config_id': config.id,
            'name': 'Bank Account', 'code': 'BANKACCTEXP',
            'column_type': 'input',
            'column_role': 'bank', 'column_role_source': 'user',
        })
        # CR27 — `create` assigns its own sequence from the column-letter
        # high-water mark, so a sequence passed to create() does not survive.
        # Set the order AFTER creation, deliberately against creation order so
        # the assertion below can only pass if sequence really drives it.
        staff.sequence = 20
        division.sequence = 10
        # Rules exist, flag still off ⇒ still the historical five.
        self.assertEqual([c['header'] for c in run._export_base_columns(config)],
                         [c['header'] for c in fixed])

        config.export_identity_columns = True
        promoted = run._export_base_columns(config)
        # Sequence order, and bank/contract/reference are NOT promoted.
        self.assertEqual([c['header'] for c in promoted], ['Division', 'Staff No'])
        self.assertTrue(all(c['use_string_payload'] for c in promoted))

        # Opted in with nothing marked ⇒ fall back rather than ship a headerless sheet.
        empty = self.env['hr.formula.config'].create({
            'name': 'COLROLES Export Empty',
            'country_code': 'VN',
            'export_identity_columns': True,
        })
        self.assertEqual([c['header'] for c in run._export_base_columns(empty)],
                         [c['header'] for c in fixed])

    # ------------------------------------------------------------------
    # 18 — the role summary sentence shared by both import wizards
    # ------------------------------------------------------------------
    def test_18_role_summary_sentence(self):
        Config = self.env['hr.formula.config']
        rules = self.env['hr.formula.rule'].create([
            {'config_id': self.config.id, 'name': 'Pay A', 'code': 'PAYASUM',
             'column_type': 'input'},
            {'config_id': self.config.id, 'name': 'Pay B', 'code': 'PAYBSUM',
             'column_type': 'input'},
            {'config_id': self.config.id, 'name': 'Code', 'code': 'CODESUM',
             'column_type': 'input', 'column_role': 'identity',
             'column_role_source': 'user'},
        ])
        counts = Config.role_counts_for_rules(rules)
        self.assertEqual(list(counts.items()), [('payroll', 2), ('identity', 1)])
        sentence = Config.format_role_summary(counts)
        self.assertIn('2', sentence)
        self.assertIn('1', sentence)
        self.assertNotIn('0 ', sentence)
        self.assertEqual(Config.format_role_summary({}), '')
