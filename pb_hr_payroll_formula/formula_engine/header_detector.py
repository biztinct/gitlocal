# -*- coding: utf-8 -*-
"""
Header Detector - Dynamic header row detection for Excel files.

This module provides algorithms to automatically detect the header row
in Excel worksheets, handling complex layouts with merged cells and
multi-row headers.

Key Features:
- Detects rows with large horizontal merged cells (category/component_type rows)
- Finds actual header rows with individual column labels
- Handles multi-tier headers common in payroll spreadsheets
"""

import logging
from typing import List, Tuple, Dict, Any, Optional

try:
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_logger = logging.getLogger(__name__)


class HeaderDetector:
    """
    Detects header rows in Excel worksheets using heuristic analysis.

    The detection algorithm analyzes cell content patterns to find the most
    likely header row, considering:
    - Large horizontal merged cells (component_type rows - SKIP these)
    - Individual column labels (actual header row - FIND these)
    - String vs. numeric content ratio
    - Row position relative to merged regions

    Example Excel structure:
    Row 1: |-------- "Workdays in the month" --------|---- "Amounts received" ----|
    Row 2: "Work days" | "Collaborate" | "Vacation" | "Salary" | "Bonus" | "Net"
    Row 3: 22          | 5            | 3          | 5000000  | 200000  | 5200000

    Row 1 = component_type rows (large horizontal merged cells)
    Row 2 = actual header row (individual column labels)
    Row 3+ = data rows
    """

    # Minimum percentage of string cells to qualify as header row
    MIN_STRING_RATIO = 0.5

    # Maximum rows to search for header
    MAX_SEARCH_ROWS = 30

    # Minimum columns with data to consider valid header
    MIN_COLUMNS = 2

    # Minimum column span to consider a merge as "large" (category header)
    MIN_LARGE_MERGE_SPAN = 3

    # If more than this percentage of columns are in large merges, it's a category row
    LARGE_MERGE_THRESHOLD = 0.4

    def __init__(self, sheet: 'Worksheet'):
        """
        Initialize header detector for a worksheet.

        Args:
            sheet: openpyxl Worksheet object
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl library is required for header detection")

        self.sheet = sheet
        self.merged_ranges = list(sheet.merged_cells.ranges)
        self._vertical_merge_rows = self._build_vertical_merge_map()
        self._horizontal_merge_map = self._build_horizontal_merge_map()

    def _build_vertical_merge_map(self) -> Dict[int, List[Tuple[int, int]]]:
        """
        Build a map of rows that are part of vertical merges.

        Returns:
            Dictionary mapping row numbers to list of (min_col, max_col) ranges
            that are part of a vertical merge spanning that row.
        """
        merge_map = {}

        for merged_range in self.merged_ranges:
            # Check if this is a vertical merge (spans multiple rows)
            if merged_range.max_row > merged_range.min_row:
                # Mark all rows in this merge (except the first row which has content)
                for row in range(merged_range.min_row + 1, merged_range.max_row + 1):
                    if row not in merge_map:
                        merge_map[row] = []
                    merge_map[row].append((merged_range.min_col, merged_range.max_col))

        return merge_map

    def _build_horizontal_merge_map(self) -> Dict[int, List[Dict[str, Any]]]:
        """
        Build a map of horizontal merged cells per row.

        Returns:
            Dictionary mapping row numbers to list of horizontal merge info
        """
        merge_map = {}

        for merged_range in self.merged_ranges:
            # Check if this is a horizontal merge (spans multiple columns, same row or spans rows)
            if merged_range.max_col > merged_range.min_col:
                col_span = merged_range.max_col - merged_range.min_col + 1
                merge_info = {
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col,
                    'col_span': col_span,
                    'is_large': col_span >= self.MIN_LARGE_MERGE_SPAN,
                    'value': self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                }

                # Add to the starting row of the merge
                row = merged_range.min_row
                if row not in merge_map:
                    merge_map[row] = []
                merge_map[row].append(merge_info)

        return merge_map

    def is_row_primarily_large_merges(self, row_num: int) -> bool:
        """
        Check if a row is primarily composed of large horizontal merged cells.

        This identifies "category" or "component_type" rows that should NOT
        be treated as header rows. These are rows like:
        - "Workdays in the month" spanning columns A-F
        - "Amounts received during the month" spanning columns G-L

        Args:
            row_num: 1-based row number

        Returns:
            True if row is primarily large horizontal merges (a category row)
        """
        if row_num not in self._horizontal_merge_map:
            return False

        merges = self._horizontal_merge_map[row_num]
        if not merges:
            return False

        # Count columns covered by large merges
        large_merge_cols = set()
        total_merge_cols = set()

        for merge in merges:
            for col in range(merge['min_col'], merge['max_col'] + 1):
                total_merge_cols.add(col)
                if merge['is_large']:
                    large_merge_cols.add(col)

        # Get total non-empty columns in this row
        non_empty_cols = 0
        for cell in self.sheet[row_num]:
            # Check if cell has value or is part of a merge
            if cell.value is not None or cell.column in total_merge_cols:
                non_empty_cols += 1

        if non_empty_cols == 0:
            return False

        # Row is primarily large merges if >40% of columns are in large merges
        large_merge_ratio = len(large_merge_cols) / max(non_empty_cols, 1)

        _logger.debug(
            f"Row {row_num}: large_merge_cols={len(large_merge_cols)}, "
            f"non_empty_cols={non_empty_cols}, ratio={large_merge_ratio:.2f}"
        )

        return large_merge_ratio >= self.LARGE_MERGE_THRESHOLD

    def get_individual_column_count(self, row_num: int) -> int:
        """
        Count how many columns in this row have individual (non-merged) values.

        This helps identify rows with actual column headers vs rows with category headers.
        A truly "individual" column is one that is NOT part of ANY horizontal merge.

        Args:
            row_num: 1-based row number

        Returns:
            Count of individual (non-merged) column values
        """
        # Get ALL columns that are part of ANY horizontal merge (not just large ones)
        merged_cols = set()
        if row_num in self._horizontal_merge_map:
            for merge in self._horizontal_merge_map[row_num]:
                for col in range(merge['min_col'], merge['max_col'] + 1):
                    merged_cols.add(col)

        # Count non-empty cells NOT in ANY horizontal merge
        individual_count = 0
        for cell in self.sheet[row_num]:
            if cell.column not in merged_cols:
                value = cell.value
                if value is not None and (not isinstance(value, str) or value.strip()):
                    individual_count += 1

        return individual_count

    def is_row_primarily_any_merges(self, row_num: int) -> bool:
        """
        Check if a row is primarily composed of horizontal merged cells (any size).

        This identifies "sub-category" rows like "Date of work stoppage in the month"
        which may span only 2 columns but are still merged cells, not individual field names.

        Args:
            row_num: 1-based row number

        Returns:
            True if row is primarily horizontal merges (even small ones)
        """
        if row_num not in self._horizontal_merge_map:
            return False

        merges = self._horizontal_merge_map[row_num]
        if not merges:
            return False

        # Count columns covered by ANY horizontal merges
        any_merge_cols = set()
        for merge in merges:
            for col in range(merge['min_col'], merge['max_col'] + 1):
                any_merge_cols.add(col)

        # Get total non-empty columns in this row
        non_empty_cols = 0
        for cell in self.sheet[row_num]:
            if cell.value is not None or cell.column in any_merge_cols:
                non_empty_cols += 1

        if non_empty_cols == 0:
            return False

        # Row is primarily merges if >60% of columns are in merges
        merge_ratio = len(any_merge_cols) / max(non_empty_cols, 1)

        _logger.debug(
            f"Row {row_num}: any_merge_cols={len(any_merge_cols)}, "
            f"non_empty_cols={non_empty_cols}, merge_ratio={merge_ratio:.2f}"
        )

        return merge_ratio >= 0.6

    def is_row_in_vertical_merge(self, row_num: int) -> bool:
        """
        Check if a row is entirely or mostly within vertical merged cells.

        A row is considered "in vertical merge" ONLY if most of its
        non-empty cells are ACTUALLY in merged columns (i.e., the values
        in this row come from vertical merges, not from independent cells).

        This is important for mixed layouts where:
        - Some columns have vertically merged headers (TT, MSNV spanning rows 1-2)
        - Other columns have horizontal category headers in row 1 and
          individual labels in row 2 (like "Cong goes to work")

        In this case, row 2 should NOT be marked as "in vertical merge"
        because it has its own individual column labels.

        Args:
            row_num: 1-based row number

        Returns:
            True if row is substantially within vertical merges
        """
        if row_num not in self._vertical_merge_rows:
            return False

        merge_ranges = self._vertical_merge_rows[row_num]

        # Build set of columns that are part of vertical merges
        merged_cols = set()
        for min_col, max_col in merge_ranges:
            for col in range(min_col, max_col + 1):
                merged_cols.add(col)

        # Count cells with values that are IN vs NOT IN vertical merges
        values_in_merge = 0
        values_not_in_merge = 0

        for cell in self.sheet[row_num]:
            value = cell.value
            if value is not None and (not isinstance(value, str) or value.strip()):
                if cell.column in merged_cols:
                    values_in_merge += 1
                else:
                    values_not_in_merge += 1

        total_values = values_in_merge + values_not_in_merge

        if total_values == 0:
            return True  # Empty row treated as being in merge

        # Row is "in vertical merge" only if MOST values come from merged cells
        # If there are independent values (not in merges), this row has its own labels
        if values_not_in_merge > 0:
            # This row has independent labels - it's a potential header row
            return False

        # All values are in merged columns
        return True

    def analyze_row(self, row_num: int) -> Dict[str, Any]:
        """
        Analyze a row's content to determine if it's a potential header.

        Key Logic:
        - Rows with large horizontal merged cells are category/component_type rows, NOT headers
        - Rows with any horizontal merged cells (sub-categories) are also penalized
        - Actual header rows have individual column labels (many non-merged string cells)

        Args:
            row_num: 1-based row number

        Returns:
            Analysis results including counts and score
        """
        row_cells = list(self.sheet[row_num])

        # Check if this row is primarily large horizontal merges (main category row)
        is_category_row = self.is_row_primarily_large_merges(row_num)
        # Check if this row is primarily ANY horizontal merges (sub-category row)
        is_subcategory_row = self.is_row_primarily_any_merges(row_num)
        individual_count = self.get_individual_column_count(row_num)

        analysis = {
            'row_num': row_num,
            'total_cells': len(row_cells),
            'non_empty_cells': 0,
            'string_cells': 0,
            'string_cells_with_digits': 0,
            'numeric_cells': 0,
            'date_cells': 0,
            'empty_cells': 0,
            'in_vertical_merge': self.is_row_in_vertical_merge(row_num),
            'is_category_row': is_category_row,  # Rows with large horizontal merges
            'is_subcategory_row': is_subcategory_row,  # Rows with any horizontal merges
            'individual_column_count': individual_count,  # Non-merged column count
            'score': 0.0,
            'is_candidate': False,
        }

        for cell in row_cells:
            value = cell.value

            if value is None or (isinstance(value, str) and not value.strip()):
                analysis['empty_cells'] += 1
                continue

            analysis['non_empty_cells'] += 1

            # Classify cell type
            if isinstance(value, str):
                # Check if it's a numeric string
                try:
                    float(value.replace(',', '').replace(' ', ''))
                    analysis['numeric_cells'] += 1
                except (ValueError, AttributeError):
                    analysis['string_cells'] += 1
                    if any(ch.isdigit() for ch in value):
                        analysis['string_cells_with_digits'] += 1
            elif isinstance(value, (int, float)):
                analysis['numeric_cells'] += 1
            elif hasattr(value, 'strftime'):  # datetime-like
                analysis['date_cells'] += 1
            else:
                analysis['string_cells'] += 1

        # Calculate score
        if analysis['non_empty_cells'] > 0:
            string_ratio = analysis['string_cells'] / analysis['non_empty_cells']

            # Base score from string ratio
            analysis['score'] = string_ratio

            # MAJOR PENALTY: Category rows (large horizontal merges) are NOT headers
            if is_category_row:
                analysis['score'] *= 0.1  # Heavy penalty - these are component_type rows

            # PENALTY: Sub-category rows (any horizontal merges) are also NOT headers
            # Example: "Date of work stoppage in the month" spanning V-W
            if is_subcategory_row and not is_category_row:
                analysis['score'] *= 0.2  # Heavy penalty - these are intermediate headers

            # Penalty for being in vertical merge
            if analysis['in_vertical_merge']:
                analysis['score'] *= 0.3

            # Penalty if numerics dominate
            if analysis['numeric_cells'] > analysis['string_cells']:
                analysis['score'] *= 0.5

            # BONUS: Rows with many individual column labels are likely headers
            # The more individual (non-merged) cells, the better
            if individual_count >= 5:
                analysis['score'] *= 1.5
            elif individual_count >= 3:
                analysis['score'] *= 1.2

            # Bonus for having many columns
            if analysis['non_empty_cells'] >= self.MIN_COLUMNS:
                analysis['score'] *= 1.1

            # Penalty for rows with many digit-heavy strings (often data rows)
            digit_ratio = analysis['string_cells_with_digits'] / max(analysis['string_cells'], 1)
            if digit_ratio >= 0.6:
                analysis['score'] *= 0.3
            elif digit_ratio >= 0.4:
                analysis['score'] *= 0.6

            # Mark as candidate if meets criteria
            # Key change: Must NOT be a category row OR subcategory row
            analysis['is_candidate'] = (
                string_ratio >= self.MIN_STRING_RATIO and
                not analysis['in_vertical_merge'] and
                not is_category_row and  # Skip category/component_type rows
                not is_subcategory_row and  # Skip sub-category rows with merges
                analysis['string_cells'] >= analysis['numeric_cells'] and
                analysis['non_empty_cells'] >= self.MIN_COLUMNS
            )

        return analysis

    def detect_header_row(
        self,
        max_search_rows: int = None
    ) -> Tuple[int, int, Dict[str, Any]]:
        """
        Detect the header row in the worksheet.

        Scans the first N rows looking for the most likely header row
        based on content analysis. Rows with large horizontal merged cells
        are identified as category/component_type rows and skipped.

        Args:
            max_search_rows: Maximum rows to search (default: class constant)

        Returns:
            Tuple of (header_row, data_start_row, analysis_details)
            where rows are 1-based numbers.
        """
        max_rows = max_search_rows or self.MAX_SEARCH_ROWS
        max_rows = min(max_rows, self.sheet.max_row or 1)

        _logger.info(
            f"=== HEADER DETECTION START ===\n"
            f"Sheet: {self.sheet.title}, max_row={self.sheet.max_row}, max_col={self.sheet.max_column}\n"
            f"Searching up to {max_rows} rows\n"
            f"Total merged ranges: {len(self.merged_ranges)}"
        )

        best_row = 1
        best_score = 0.0
        all_analyses = []
        category_rows = []  # Track rows with large horizontal merges

        for row_num in range(1, max_rows + 1):
            analysis = self.analyze_row(row_num)
            all_analyses.append(analysis)

            # Track category/component_type rows
            if analysis['is_category_row']:
                category_rows.append(row_num)

            # Log detailed info for first 10 rows
            if row_num <= 10:
                _logger.info(
                    f"Row {row_num}: score={analysis['score']:.2f}, "
                    f"is_candidate={analysis['is_candidate']}, "
                    f"is_category_row={analysis['is_category_row']}, "
                    f"is_subcategory_row={analysis['is_subcategory_row']}, "
                    f"in_vert_merge={analysis['in_vertical_merge']}, "
                    f"strings={analysis['string_cells']}, "
                    f"numerics={analysis['numeric_cells']}, "
                    f"individual_cols={analysis['individual_column_count']}"
                )

            if analysis['is_candidate'] and analysis['score'] > best_score:
                best_score = analysis['score']
                best_row = row_num
                _logger.info(f"  -> New best header candidate: row {row_num} (score={best_score:.2f})")

        best_analysis = next(
            (a for a in all_analyses if a['row_num'] == best_row),
            None
        )

        # Heuristic: if the chosen row looks like data, prefer the nearest header-like row above.
        if best_row > 1 and best_analysis:
            best_digit_ratio = (
                best_analysis['string_cells_with_digits'] /
                max(best_analysis['string_cells'], 1)
            )
            best_looks_like_data = (
                best_analysis['date_cells'] > 0 or best_digit_ratio >= 0.4
            )
            if best_looks_like_data:
                for offset in range(1, min(4, best_row)):
                    above_analysis = next(
                        (a for a in all_analyses if a['row_num'] == best_row - offset),
                        None
                    )
                    if not above_analysis:
                        continue
                    above_digit_ratio = (
                        above_analysis['string_cells_with_digits'] /
                        max(above_analysis['string_cells'], 1)
                    )
                    above_is_header_like = (
                        above_analysis['string_cells'] >= self.MIN_COLUMNS and
                        not above_analysis['is_category_row'] and
                        not above_analysis['is_subcategory_row'] and
                        above_digit_ratio <= 0.2 and
                        above_analysis['string_cells'] >= above_analysis['numeric_cells']
                    )
                    if above_is_header_like:
                        best_row = above_analysis['row_num']
                        best_score = above_analysis['score']
                        break

        # Data starts on the row after header
        data_start_row = best_row + 1

        # Compile detection details
        details = {
            'header_row': best_row,
            'data_start_row': data_start_row,
            'detection_score': best_score,
            'rows_analyzed': len(all_analyses),
            'candidate_rows': [a['row_num'] for a in all_analyses if a['is_candidate']],
            'category_rows': category_rows,  # NEW: rows with component_type info
            'header_analysis': next((a for a in all_analyses if a['row_num'] == best_row), None),
            'all_analyses': all_analyses,  # For debugging
        }

        _logger.info(
            f"Header detection: row {best_row} (score: {best_score:.2f}), "
            f"data starts at row {data_start_row}, "
            f"category rows: {category_rows}"
        )

        return best_row, data_start_row, details

    def get_category_row_merges(self, header_row: int) -> List[Dict[str, Any]]:
        """
        Get all category/component_type information from merged cells above header row.

        Args:
            header_row: The detected header row

        Returns:
            List of merge info dictionaries with value and column range
        """
        category_merges = []

        for merged_range in self.merged_ranges:
            # Check if this is a horizontal merge above the header row
            if (merged_range.max_col > merged_range.min_col and
                    merged_range.min_row < header_row and
                    merged_range.max_col - merged_range.min_col + 1 >= self.MIN_LARGE_MERGE_SPAN):

                value = self.sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                ).value

                if value:
                    category_merges.append({
                        'value': str(value).strip(),
                        'row': merged_range.min_row,
                        'min_col': merged_range.min_col,
                        'max_col': merged_range.max_col,
                        'col_span': merged_range.max_col - merged_range.min_col + 1,
                        'min_col_letter': get_column_letter(merged_range.min_col),
                        'max_col_letter': get_column_letter(merged_range.max_col),
                    })

        # Sort by row (ascending) then by column
        category_merges.sort(key=lambda x: (x['row'], x['min_col']))

        return category_merges

    def get_headers(self, header_row: int = None) -> List[Dict[str, Any]]:
        """
        Extract header information from the detected or specified header row.

        Handles mixed layouts where:
        - Some columns have headers in the header row itself
        - Some columns have vertically merged headers from rows above
          (e.g., "TT", "MSNV" spanning rows 1-2, with header_row=2)
        - Some columns have intermediate merged cells that need to be expanded
          to the row below (e.g., "Date of work stoppage" spanning V-W in row 2,
          with actual headers "First 14 days" and "From 15th onwards" in row 3)

        Args:
            header_row: Row number to use as header (auto-detects if None)

        Returns:
            List of header information dictionaries with keys:
            - column_index: 0-based column index
            - column_letter: Excel column letter
            - value: Header text
            - is_merged: Whether cell is part of horizontal merge
            - merge_span: Number of columns spanned if merged
            - from_vertical_merge: True if value came from a vertical merge above
            - from_subheader_row: True if value came from row below a horizontal merge
        """
        if header_row is None:
            header_row, _, _ = self.detect_header_row()

        headers = []

        # Build map of horizontal merges that cover the header row
        # Key: column number, Value: merge info including all columns in the merge
        # This includes merges that START on or above the header row and span multiple rows.
        horizontal_merges = {}
        for merged_range in self.merged_ranges:
            # Check if this merge spans multiple columns and covers the header row
            if (
                merged_range.max_col > merged_range.min_col
                and merged_range.min_row <= header_row <= merged_range.max_row
            ):
                # This is a horizontal merge that covers the header row
                # It may start above the header row (e.g., category row merges).
                value_row = merged_range.min_row
                merge_value = self.sheet.cell(
                    row=value_row,
                    column=merged_range.min_col
                ).value
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    horizontal_merges[col] = {
                        'start_col': merged_range.min_col,
                        'end_col': merged_range.max_col,
                        'span': merged_range.max_col - merged_range.min_col + 1,
                        'value': merge_value,
                        'end_row': merged_range.max_row,  # Track where the merge ends vertically
                    }

        # Build map of vertical merges that include the header row
        # These are columns where the header value comes from a row above
        vertical_merge_values = {}
        for merged_range in self.merged_ranges:
            # Check if this is a vertical merge that includes the header row
            if (merged_range.max_row >= header_row and
                    merged_range.min_row < header_row and
                    merged_range.max_row > merged_range.min_row):
                # Get the value from the top of the merge
                value = self.sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                ).value
                if value:
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        vertical_merge_values[col] = str(value).strip()
                        _logger.debug(
                            f"Vertical merge: col {col} -> '{value}' "
                            f"(from row {merged_range.min_row})"
                        )

        # Check for subheader row (row AFTER the merge ends)
        # Used when a horizontal merge in header row has individual cells below it
        # The merge might span multiple rows (e.g., "Date of work stoppage" in rows 2-3)
        # So we need to check the row AFTER the merge ends, not just header_row + 1
        # We search up to MAX_SUBHEADER_SEARCH rows below the merge to find actual headers
        MAX_SUBHEADER_SEARCH = 5  # Search up to 5 rows below the merge
        subheader_values = {}
        subheader_source_rows = {}  # Track which row each subheader came from

        for col_num, merge_info in horizontal_merges.items():
            # Only check the start column of each merge
            if col_num == merge_info['start_col']:
                # Get the row after this merge ends
                merge_end_row = merge_info.get('end_row', header_row)

                # Search multiple rows below the merge to find subheader values
                for row_offset in range(1, MAX_SUBHEADER_SEARCH + 1):
                    subheader_row = merge_end_row + row_offset

                    if subheader_row > (self.sheet.max_row or 1):
                        break

                    # Check all columns in this merge for subheader values
                    found_values_in_row = {}
                    for sub_col in range(merge_info['start_col'], merge_info['end_col'] + 1):
                        # Skip if we already found a value for this column
                        if sub_col in subheader_values:
                            continue

                        sub_cell = self.sheet.cell(row=subheader_row, column=sub_col)
                        sub_value = sub_cell.value

                        # Be more lenient - accept any non-empty value, not just strings
                        if sub_value is not None:
                            sub_str = str(sub_value).strip()
                            if sub_str:
                                found_values_in_row[sub_col] = sub_str

                    # If we found values in this row, use them
                    if found_values_in_row:
                        for sub_col, sub_str in found_values_in_row.items():
                            subheader_values[sub_col] = sub_str
                            subheader_source_rows[sub_col] = subheader_row
                            _logger.info(
                                f"Subheader found: col {sub_col} -> '{sub_str}' "
                                f"(row {subheader_row}, below merged '{merge_info['value']}' "
                                f"which spans rows {header_row}-{merge_end_row})"
                            )

                    # If we found values for ALL columns in the merge, stop searching
                    merge_cols = set(range(merge_info['start_col'], merge_info['end_col'] + 1))
                    if merge_cols.issubset(set(subheader_values.keys())):
                        break

        _logger.info(
            f"get_headers: header_row={header_row}, "
            f"horizontal_merges_count={len(horizontal_merges)}, "
            f"subheader_values_count={len(subheader_values)}"
        )
        if subheader_values:
            _logger.info(f"  Subheader values: {subheader_values}")
        if horizontal_merges:
            # Log details about horizontal merges for debugging
            unique_merges = {}
            for col, info in horizontal_merges.items():
                if col == info['start_col']:
                    unique_merges[col] = info
            for col, info in unique_merges.items():
                _logger.info(
                    f"  Horizontal merge: cols {info['start_col']}-{info['end_col']} "
                    f"(rows {header_row}-{info.get('end_row', header_row)}), "
                    f"value='{info['value']}'"
                )

        # Track which columns we've already processed via subheaders
        processed_subheader_cols = set()

        for col_idx, cell in enumerate(self.sheet[header_row]):
            col_num = col_idx + 1  # 1-based column number

            # Check if this column has a subheader (individual value below a horizontal merge)
            if col_num in subheader_values:
                # Use the subheader value instead of the merged cell value
                header_info = {
                    'column_index': col_idx,
                    'column_letter': get_column_letter(col_num),
                    'value': subheader_values[col_num],
                    'is_merged': False,
                    'merge_span': 1,
                    'is_merge_continuation': False,
                    'from_vertical_merge': False,
                    'from_subheader_row': True,
                }
                headers.append(header_info)
                processed_subheader_cols.add(col_num)
                _logger.debug(
                    f"Using subheader for col {col_num}: '{subheader_values[col_num]}'"
                )
                continue

            # Skip continuation cells in horizontal merges (but not if they have subheaders)
            if col_num in horizontal_merges:
                merge_info = horizontal_merges[col_num]
                if col_num > merge_info['start_col']:
                    # This is a continuation of a merge
                    # Skip only if all columns in this merge have been processed as subheaders
                    merge_cols = set(range(merge_info['start_col'], merge_info['end_col'] + 1))
                    if merge_cols.issubset(processed_subheader_cols):
                        continue
                    # If subheaders were found for some columns, skip this continuation
                    if any(c in subheader_values for c in merge_cols):
                        continue

            # Get the header value - either from this cell or from vertical merge
            cell_value = cell.value
            from_vertical_merge = False

            if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
                # Cell is empty - check if there's a vertical merge above
                if col_num in vertical_merge_values:
                    cell_value = vertical_merge_values[col_num]
                    from_vertical_merge = True

            header_info = {
                'column_index': col_idx,
                'column_letter': get_column_letter(col_num),
                'value': str(cell_value).strip() if cell_value else None,
                'is_merged': col_num in horizontal_merges,
                'merge_span': 1,
                'is_merge_continuation': False,
                'from_vertical_merge': from_vertical_merge,
                'from_subheader_row': False,
            }

            if col_num in horizontal_merges:
                merge_info = horizontal_merges[col_num]
                header_info['merge_span'] = merge_info['span']
                header_info['is_merge_continuation'] = col_num > merge_info['start_col']

            # Skip continuation cells in horizontal merges (unless processed as subheader)
            if not header_info['is_merge_continuation']:
                headers.append(header_info)

        _logger.info(
            f"get_headers: Found {len(headers)} headers in row {header_row} "
            f"({len([h for h in headers if h.get('from_vertical_merge')])} from vertical merges, "
            f"{len([h for h in headers if h.get('from_subheader_row')])} from subheader row)"
        )

        return headers

    def detect_with_confidence(self) -> Dict[str, Any]:
        """
        Perform header detection and return detailed confidence analysis.

        Returns:
            Dictionary with detection results and confidence metrics
        """
        header_row, data_start_row, details = self.detect_header_row()

        # Calculate confidence level
        confidence = 'low'
        if details['detection_score'] >= 0.8:
            confidence = 'high'
        elif details['detection_score'] >= 0.5:
            confidence = 'medium'

        # Get actual headers
        headers = self.get_headers(header_row)

        # Get category/component_type merges from rows above header
        category_merges = self.get_category_row_merges(header_row)

        return {
            'header_row': header_row,
            'data_start_row': data_start_row,
            'confidence': confidence,
            'confidence_score': details['detection_score'],
            'headers': headers,
            'header_count': len(headers),
            'candidate_rows': details['candidate_rows'],
            'category_rows': details.get('category_rows', []),  # NEW
            'category_merges': category_merges,  # NEW: component_type info
            'analysis': details,
        }


def detect_header_row(sheet: 'Worksheet', max_search_rows: int = 20) -> Tuple[int, int]:
    """
    Convenience function to detect header row in a worksheet.

    Args:
        sheet: openpyxl Worksheet object
        max_search_rows: Maximum rows to search

    Returns:
        Tuple of (header_row, data_start_row)
    """
    detector = HeaderDetector(sheet)
    header_row, data_start_row, _ = detector.detect_header_row(max_search_rows)
    return header_row, data_start_row
