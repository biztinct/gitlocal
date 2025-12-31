# -*- coding: utf-8 -*-
"""
Merged Cell Parser - Extract component types from merged cells in Excel.

This module parses merged cell regions above header rows to extract
component type categorizations (e.g., "Deductions", "Allowances", "Earnings").
"""

import logging
from typing import List, Dict, Any, Optional, Tuple

try:
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_logger = logging.getLogger(__name__)


class MergedCellParser:
    """
    Extracts component type information from merged cells above headers.

    In many payroll spreadsheets, category headers span multiple columns
    using merged cells:

    Row 1: |-------- "Total Deductions" --------|
    Row 2: |-- "Deductions" --|-- "Other" ------|
    Row 3: "Tax" | "SI" | "Loan" | "Advance"     <- Header row

    This parser extracts the immediate (lowest level) category for each column.
    """

    def __init__(self, sheet: 'Worksheet'):
        """
        Initialize merged cell parser for a worksheet.

        Args:
            sheet: openpyxl Worksheet object
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl library is required for merged cell parsing")

        self.sheet = sheet
        self.merged_ranges = list(sheet.merged_cells.ranges)
        self._merge_cache = self._build_merge_cache()

    def _build_merge_cache(self) -> Dict[Tuple[int, int], Dict[str, Any]]:
        """
        Build a cache of merged cell information for fast lookup.

        Returns:
            Dictionary mapping (row, col) to merge information
        """
        cache = {}

        for merged_range in self.merged_ranges:
            # Get the value from the top-left cell of the merge
            value = self.sheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value

            merge_info = {
                'value': str(value).strip() if value else None,
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'row_span': merged_range.max_row - merged_range.min_row + 1,
                'col_span': merged_range.max_col - merged_range.min_col + 1,
                'is_horizontal': merged_range.max_col > merged_range.min_col,
                'is_vertical': merged_range.max_row > merged_range.min_row,
            }

            # Cache for all cells in the range
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cache[(row, col)] = merge_info

        return cache

    def get_merge_at(self, row: int, col: int) -> Optional[Dict[str, Any]]:
        """
        Get merged cell information at a specific position.

        Args:
            row: 1-based row number
            col: 1-based column number

        Returns:
            Merge information dictionary or None if not merged
        """
        return self._merge_cache.get((row, col))

    def get_horizontal_merges_above_row(
        self,
        target_row: int,
        min_row: int = 1
    ) -> List[Dict[str, Any]]:
        """
        Find all horizontal merged cells above a target row.

        Args:
            target_row: Row number to search above (exclusive)
            min_row: Minimum row to include in search

        Returns:
            List of merge information for horizontal merges above target_row
        """
        merges = []

        for merged_range in self.merged_ranges:
            # Check if this is a horizontal merge (spans multiple columns)
            # and is above the target row
            if (merged_range.max_col > merged_range.min_col and
                    merged_range.max_row < target_row and
                    merged_range.min_row >= min_row):

                value = self.sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                ).value

                if value:  # Only include merges with content
                    merges.append({
                        'value': str(value).strip(),
                        'row': merged_range.min_row,
                        'min_col': merged_range.min_col,
                        'max_col': merged_range.max_col,
                        'col_span': merged_range.max_col - merged_range.min_col + 1,
                        'min_col_letter': get_column_letter(merged_range.min_col),
                        'max_col_letter': get_column_letter(merged_range.max_col),
                    })

        # Sort by row (descending) so closest to header is first
        merges.sort(key=lambda x: -x['row'])

        return merges

    def extract_component_types(
        self,
        header_row: int
    ) -> Dict[str, str]:
        """
        Extract component type for each column from merged cells above header.

        Logic:
        1. For each column, look upward for merged cells
        2. Find the IMMEDIATE (lowest level) merged cell spanning this column
        3. Use that merged cell's value as component_type

        Args:
            header_row: Row number containing column headers

        Returns:
            Dictionary mapping column letter to component type string
        """
        column_types = {}

        # Get all horizontal merges above header
        merges = self.get_horizontal_merges_above_row(header_row)

        # Process each column
        max_col = self.sheet.max_column or 1

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            candidate_merges = []

            for merge in merges:
                # Check if this merge spans our column
                if merge['min_col'] <= col_idx <= merge['max_col']:
                    candidate_merges.append(merge)

            if candidate_merges:
                # Select merge CLOSEST to header row (immediate/lowest level)
                # Merges are already sorted descending by row, so first one is closest
                immediate_merge = candidate_merges[0]
                column_types[col_letter] = immediate_merge['value']

                _logger.debug(
                    f"Column {col_letter}: component_type='{immediate_merge['value']}' "
                    f"(from row {immediate_merge['row']})"
                )

        return column_types

    def extract_component_hierarchy(
        self,
        header_row: int
    ) -> Dict[str, List[str]]:
        """
        Extract full hierarchy of component types for each column.

        This returns all levels of categorization, from highest (top) to
        lowest (immediate) level.

        Args:
            header_row: Row number containing column headers

        Returns:
            Dictionary mapping column letter to list of category names
            (ordered from highest to lowest level)
        """
        column_hierarchy = {}

        # Get all horizontal merges above header
        merges = self.get_horizontal_merges_above_row(header_row)

        # Process each column
        max_col = self.sheet.max_column or 1

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            hierarchy = []

            for merge in merges:
                if merge['min_col'] <= col_idx <= merge['max_col']:
                    hierarchy.append({
                        'level': merge['row'],
                        'value': merge['value'],
                    })

            # Sort by row ascending (top to bottom)
            hierarchy.sort(key=lambda x: x['level'])
            column_hierarchy[col_letter] = [h['value'] for h in hierarchy]

        return column_hierarchy

    def get_column_info(
        self,
        header_row: int
    ) -> List[Dict[str, Any]]:
        """
        Get comprehensive information about each column.

        Combines header values with component type extraction.

        Args:
            header_row: Row number containing column headers

        Returns:
            List of column information dictionaries
        """
        component_types = self.extract_component_types(header_row)
        hierarchy = self.extract_component_hierarchy(header_row)

        columns = []
        max_col = self.sheet.max_column or 1

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)

            # Get header value
            header_cell = self.sheet.cell(row=header_row, column=col_idx)
            header_value = header_cell.value

            # Skip empty columns
            if header_value is None or (isinstance(header_value, str) and not header_value.strip()):
                continue

            column_info = {
                'column_index': col_idx - 1,  # 0-based
                'column_letter': col_letter,
                'header': str(header_value).strip() if header_value else f"Column_{col_letter}",
                'component_type': component_types.get(col_letter),
                'category_hierarchy': hierarchy.get(col_letter, []),
                'has_formula': self._column_has_formula(col_idx, header_row + 1),
            }

            columns.append(column_info)

        return columns

    def _column_has_formula(self, col_idx: int, start_row: int) -> bool:
        """
        Check if a column contains formulas in its data cells.

        Args:
            col_idx: 1-based column index
            start_row: First data row to check

        Returns:
            True if any cell in the column has a formula
        """
        # Check first 10 data rows
        for row_offset in range(10):
            row = start_row + row_offset
            if row > (self.sheet.max_row or 1):
                break

            cell = self.sheet.cell(row=row, column=col_idx)

            # Check if cell has formula (value starts with =)
            if hasattr(cell, 'value') and isinstance(cell.value, str):
                if cell.value.startswith('='):
                    return True

        return False

    def analyze_structure(self) -> Dict[str, Any]:
        """
        Analyze the overall structure of merged cells in the worksheet.

        Returns:
            Analysis summary including merge counts and patterns
        """
        horizontal_merges = 0
        vertical_merges = 0
        multi_row_multi_col = 0
        total_merged_cells = 0

        for merged_range in self.merged_ranges:
            cells_in_merge = (
                (merged_range.max_row - merged_range.min_row + 1) *
                (merged_range.max_col - merged_range.min_col + 1)
            )
            total_merged_cells += cells_in_merge

            is_horizontal = merged_range.max_col > merged_range.min_col
            is_vertical = merged_range.max_row > merged_range.min_row

            if is_horizontal and is_vertical:
                multi_row_multi_col += 1
            elif is_horizontal:
                horizontal_merges += 1
            elif is_vertical:
                vertical_merges += 1

        return {
            'total_merge_regions': len(self.merged_ranges),
            'total_merged_cells': total_merged_cells,
            'horizontal_merges': horizontal_merges,
            'vertical_merges': vertical_merges,
            'multi_dimensional_merges': multi_row_multi_col,
            'worksheet_rows': self.sheet.max_row,
            'worksheet_cols': self.sheet.max_column,
        }


def extract_component_types(
    sheet: 'Worksheet',
    header_row: int
) -> Dict[str, str]:
    """
    Convenience function to extract component types from a worksheet.

    Args:
        sheet: openpyxl Worksheet object
        header_row: Row number containing column headers

    Returns:
        Dictionary mapping column letter to component type
    """
    parser = MergedCellParser(sheet)
    return parser.extract_component_types(header_row)
