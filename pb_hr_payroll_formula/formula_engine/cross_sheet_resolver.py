# -*- coding: utf-8 -*-
"""
Cross-Sheet Formula Resolver - Resolve cross-worksheet references in Excel formulas.

This module handles the resolution of formulas that reference other worksheets,
including VLOOKUP, SUMIFS, and direct cell references.
"""

import re
import logging
from typing import Dict, List, Any, Optional, Tuple

try:
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_logger = logging.getLogger(__name__)


class CrossSheetResolver:
    """
    Resolves cross-worksheet formula references to component codes.

    Handles formulas like:
    - Direct references: 'Sheet2'!D4 → SHEET2_COL_D
    - VLOOKUP: =VLOOKUP(B4,'TimeTB 2'!$C$4:$AX$11,46,0) → resolved code
    - SUMIFS: =SUMIFS('Data'!E:E,'Data'!A:A,B4) → resolved codes
    """

    # Pattern to match sheet references: 'Sheet Name'!CellRef or SheetName!CellRef
    SHEET_REF_PATTERN = re.compile(
        r"'?([^'!]+)'?!"  # Sheet name (optionally quoted)
        r"(\$?[A-Z]+\$?\d*"  # Start of range/cell
        r"(?::\$?[A-Z]+\$?\d*)?)"  # Optional end of range
    )

    # Pattern for VLOOKUP function
    VLOOKUP_PATTERN = re.compile(
        r"VLOOKUP\s*\(\s*"
        r"([^,]+),\s*"  # lookup_value
        r"'?([^'!]+)'?!"  # sheet name
        r"(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+),\s*"  # range
        r"(\d+),\s*"  # col_index
        r"([^)]+)\)",  # match_type
        re.IGNORECASE
    )

    # Pattern for SUMIFS function
    SUMIFS_PATTERN = re.compile(
        r"SUMIFS\s*\(\s*"
        r"'?([^'!]+)'?!([^,]+),",  # sum_range
        re.IGNORECASE
    )

    def __init__(self, sheet_mappings: Dict[str, Dict[str, str]] = None):
        """
        Initialize cross-sheet resolver.

        Args:
            sheet_mappings: Dictionary mapping sheet names to column mappings.
                           Format: {'Sheet1': {'A': 'BASIC', 'B': 'GROSS', ...}}
        """
        self.sheet_mappings = sheet_mappings or {}
        self._code_cache = {}

    def add_sheet_mapping(
        self,
        sheet_name: str,
        column_mappings: Dict[str, str]
    ):
        """
        Add or update mapping for a worksheet.

        Args:
            sheet_name: Name of the worksheet
            column_mappings: Dictionary mapping column letters to codes
        """
        self.sheet_mappings[sheet_name] = column_mappings
        # Clear cache when mappings change
        self._code_cache = {}

    def resolve_formula(self, formula: str) -> Tuple[str, List[str]]:
        """
        Resolve cross-sheet references in a formula.

        Args:
            formula: Excel formula string

        Returns:
            Tuple of (resolved_formula, list_of_resolved_codes)
        """
        if not formula:
            return formula, []

        resolved = formula
        resolved_codes = []

        # First, resolve VLOOKUP functions (special handling)
        resolved, vlookup_codes = self._resolve_vlookups(resolved)
        resolved_codes.extend(vlookup_codes)

        # Then resolve direct sheet references
        resolved, sheet_ref_codes = self._resolve_sheet_references(resolved)
        resolved_codes.extend(sheet_ref_codes)

        return resolved, resolved_codes

    def _resolve_vlookups(self, formula: str) -> Tuple[str, List[str]]:
        """
        Resolve VLOOKUP functions that reference other sheets.

        VLOOKUP(lookup_value, 'Sheet'!range, col_index, match_type)
        → Identifies which component the col_index points to

        Args:
            formula: Formula string

        Returns:
            Tuple of (modified_formula, list_of_resolved_codes)
        """
        resolved_codes = []

        def replace_vlookup(match):
            lookup_value = match.group(1).strip()
            sheet_name = match.group(2).strip()
            cell_range = match.group(3).strip()
            col_index = int(match.group(4).strip())
            match_type = match.group(5).strip()

            # Parse the range to get start column
            range_match = re.match(r'\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+', cell_range)
            if not range_match:
                _logger.warning(f"Could not parse VLOOKUP range: {cell_range}")
                return match.group(0)

            start_col_letter = range_match.group(1)
            start_col_idx = column_index_from_string(start_col_letter)

            # Calculate target column: start_col + col_index - 1
            target_col_idx = start_col_idx + col_index - 1
            target_col_letter = get_column_letter(target_col_idx)

            # Look up the code for this column in the referenced sheet
            target_code = self._get_code_for_column(sheet_name, target_col_letter)

            if target_code:
                resolved_codes.append(target_code)
                _logger.debug(
                    f"VLOOKUP resolved: '{sheet_name}'!{cell_range} col {col_index} "
                    f"→ column {target_col_letter} → {target_code}"
                )
                # Replace with reference to the resolved code
                return f"values.get('{target_code}', 0)"
            else:
                # Generate a placeholder code
                placeholder = self._generate_cross_sheet_code(sheet_name, target_col_letter)
                resolved_codes.append(placeholder)
                _logger.warning(
                    f"VLOOKUP column not mapped: '{sheet_name}'!{target_col_letter}. "
                    f"Using placeholder: {placeholder}"
                )
                return f"values.get('{placeholder}', 0)"

        result = self.VLOOKUP_PATTERN.sub(replace_vlookup, formula)
        return result, resolved_codes

    def _resolve_sheet_references(self, formula: str) -> Tuple[str, List[str]]:
        """
        Resolve direct sheet references like 'Sheet2'!D4.

        Args:
            formula: Formula string

        Returns:
            Tuple of (modified_formula, list_of_resolved_codes)
        """
        resolved_codes = []

        def replace_reference(match):
            sheet_name = match.group(1).strip()
            cell_ref = match.group(2).strip()

            # Extract column letter from cell reference
            col_match = re.match(r'\$?([A-Z]+)', cell_ref)
            if not col_match:
                return match.group(0)

            col_letter = col_match.group(1)

            # Check if this is a range (contains :)
            if ':' in cell_ref:
                # Handle range references
                range_match = re.match(r'\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*', cell_ref)
                if range_match:
                    start_col = range_match.group(1)
                    end_col = range_match.group(2)
                    codes = self._resolve_range(sheet_name, start_col, end_col)
                    resolved_codes.extend(codes)
                    # Return comma-separated values.get calls
                    return ', '.join([f"values.get('{code}', 0)" for code in codes])

            # Single cell reference
            target_code = self._get_code_for_column(sheet_name, col_letter)

            if target_code:
                resolved_codes.append(target_code)
                return f"values.get('{target_code}', 0)"
            else:
                placeholder = self._generate_cross_sheet_code(sheet_name, col_letter)
                resolved_codes.append(placeholder)
                return f"values.get('{placeholder}', 0)"

        result = self.SHEET_REF_PATTERN.sub(replace_reference, formula)
        return result, resolved_codes

    def _resolve_range(
        self,
        sheet_name: str,
        start_col: str,
        end_col: str
    ) -> List[str]:
        """
        Resolve a range of columns to their codes.

        Args:
            sheet_name: Name of the sheet
            start_col: Starting column letter
            end_col: Ending column letter

        Returns:
            List of resolved codes
        """
        codes = []
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)

        for col_idx in range(start_idx, end_idx + 1):
            col_letter = get_column_letter(col_idx)
            code = self._get_code_for_column(sheet_name, col_letter)
            if code:
                codes.append(code)
            else:
                codes.append(self._generate_cross_sheet_code(sheet_name, col_letter))

        return codes

    def _get_code_for_column(
        self,
        sheet_name: str,
        column_letter: str
    ) -> Optional[str]:
        """
        Get the component code for a column in a specific sheet.

        Args:
            sheet_name: Name of the sheet
            column_letter: Column letter (A, B, etc.)

        Returns:
            Component code or None if not found
        """
        # Check cache first
        cache_key = (sheet_name, column_letter)
        if cache_key in self._code_cache:
            return self._code_cache[cache_key]

        # Try exact sheet name match
        if sheet_name in self.sheet_mappings:
            mapping = self.sheet_mappings[sheet_name]
            if column_letter in mapping:
                self._code_cache[cache_key] = mapping[column_letter]
                return mapping[column_letter]

        # Try case-insensitive match
        for mapped_name, mapping in self.sheet_mappings.items():
            if mapped_name.lower() == sheet_name.lower():
                if column_letter in mapping:
                    self._code_cache[cache_key] = mapping[column_letter]
                    return mapping[column_letter]

        # Try partial match (sheet name might be truncated)
        for mapped_name, mapping in self.sheet_mappings.items():
            if (sheet_name.lower() in mapped_name.lower() or
                    mapped_name.lower() in sheet_name.lower()):
                if column_letter in mapping:
                    self._code_cache[cache_key] = mapping[column_letter]
                    return mapping[column_letter]

        return None

    def _generate_cross_sheet_code(
        self,
        sheet_name: str,
        column_letter: str
    ) -> str:
        """
        Generate a placeholder code for an unmapped cross-sheet reference.

        Args:
            sheet_name: Name of the sheet
            column_letter: Column letter

        Returns:
            Generated code string
        """
        # Clean sheet name: remove spaces, special chars
        clean_name = re.sub(r'[^A-Za-z0-9]', '', sheet_name).upper()
        return f"{clean_name}_{column_letter}"

    def extract_cross_sheet_dependencies(self, formula: str) -> List[Dict[str, Any]]:
        """
        Extract all cross-sheet dependencies from a formula.

        Args:
            formula: Excel formula string

        Returns:
            List of dependency dictionaries with sheet, column, and type info
        """
        dependencies = []

        # Find all sheet references
        for match in self.SHEET_REF_PATTERN.finditer(formula):
            sheet_name = match.group(1).strip()
            cell_ref = match.group(2).strip()

            dep = {
                'sheet_name': sheet_name,
                'cell_reference': cell_ref,
                'full_reference': match.group(0),
                'type': 'range' if ':' in cell_ref else 'cell',
            }

            # Parse column(s)
            if ':' in cell_ref:
                range_match = re.match(r'\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*', cell_ref)
                if range_match:
                    dep['start_column'] = range_match.group(1)
                    dep['end_column'] = range_match.group(2)
            else:
                col_match = re.match(r'\$?([A-Z]+)', cell_ref)
                if col_match:
                    dep['column'] = col_match.group(1)

            dependencies.append(dep)

        # Find VLOOKUP dependencies
        for match in self.VLOOKUP_PATTERN.finditer(formula):
            sheet_name = match.group(2).strip()
            cell_range = match.group(3).strip()
            col_index = int(match.group(4).strip())

            range_match = re.match(r'\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+', cell_range)
            if range_match:
                start_col_idx = column_index_from_string(range_match.group(1))
                target_col_idx = start_col_idx + col_index - 1
                target_col_letter = get_column_letter(target_col_idx)

                dep = {
                    'sheet_name': sheet_name,
                    'cell_reference': cell_range,
                    'full_reference': match.group(0),
                    'type': 'vlookup',
                    'target_column': target_col_letter,
                    'col_index': col_index,
                }
                dependencies.append(dep)

        return dependencies

    def get_required_sheets(self, formulas: List[str]) -> set:
        """
        Get all sheet names referenced in a list of formulas.

        Args:
            formulas: List of Excel formula strings

        Returns:
            Set of unique sheet names that are referenced
        """
        sheets = set()

        for formula in formulas:
            if not formula:
                continue

            # Find sheet references
            for match in self.SHEET_REF_PATTERN.finditer(formula):
                sheets.add(match.group(1).strip())

            # Find VLOOKUP references
            for match in self.VLOOKUP_PATTERN.finditer(formula):
                sheets.add(match.group(2).strip())

        return sheets

    def validate_mappings(self, formulas: List[str]) -> Dict[str, Any]:
        """
        Validate that all cross-sheet references can be resolved.

        Args:
            formulas: List of Excel formula strings

        Returns:
            Validation result with errors and warnings
        """
        result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'unresolved_references': [],
            'sheets_required': [],
            'sheets_mapped': list(self.sheet_mappings.keys()),
        }

        required_sheets = self.get_required_sheets(formulas)
        result['sheets_required'] = list(required_sheets)

        # Check for unmapped sheets
        for sheet in required_sheets:
            found = False
            for mapped_name in self.sheet_mappings.keys():
                if (sheet.lower() == mapped_name.lower() or
                        sheet.lower() in mapped_name.lower() or
                        mapped_name.lower() in sheet.lower()):
                    found = True
                    break

            if not found:
                result['warnings'].append(f"Sheet '{sheet}' not in mappings")

        # Check individual references
        for formula in formulas:
            if not formula:
                continue

            deps = self.extract_cross_sheet_dependencies(formula)
            for dep in deps:
                sheet_name = dep['sheet_name']
                col = dep.get('column') or dep.get('target_column')

                if col:
                    code = self._get_code_for_column(sheet_name, col)
                    if not code:
                        result['unresolved_references'].append({
                            'sheet': sheet_name,
                            'column': col,
                            'reference': dep['full_reference'],
                        })

        if result['unresolved_references']:
            result['warnings'].append(
                f"{len(result['unresolved_references'])} unresolved cross-sheet references"
            )

        return result


def resolve_formula(
    formula: str,
    sheet_mappings: Dict[str, Dict[str, str]]
) -> str:
    """
    Convenience function to resolve a formula's cross-sheet references.

    Args:
        formula: Excel formula string
        sheet_mappings: Dictionary mapping sheet names to column→code mappings

    Returns:
        Resolved formula string
    """
    resolver = CrossSheetResolver(sheet_mappings)
    resolved, _ = resolver.resolve_formula(formula)
    return resolved
