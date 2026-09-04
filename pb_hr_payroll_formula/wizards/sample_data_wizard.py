# -*- coding: utf-8 -*-
"""
Sample Data Wizard - Generate sample data for formula testing.
"""

import csv
import json
import random
import string
import datetime
import re
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io


class SampleDataWizard(models.TransientModel):
    _name = 'hr.formula.sample.data.wizard'
    _description = 'Sample Data Generation Wizard'

    config_id = fields.Many2one(
        'hr.formula.config',
        string='Configuration',
        required=True,
        default=lambda self: self.env.context.get('active_id'),
    )

    source = fields.Selection([
        ('manual', 'Manual Entry'),
        ('employees', 'From Employees'),
        ('payslips', 'From Existing Payslips'),
        ('file', 'Imported from File'),
        ('random', 'Generate Random Data'),
    ], string='Data Source', default='employees', required=True)

    # For employee source
    employee_ids = fields.Many2many(
        'hr.employee',
        'formula_sample_wizard_employee_rel',
        'wizard_id', 'employee_id',
        string='Employees',
    )

    # For payslip source
    payslip_ids = fields.Many2many(
        'hr.payslip',
        'formula_sample_wizard_payslip_rel',
        'wizard_id', 'payslip_id',
        string='Payslips',
    )

    # Options
    anonymize = fields.Boolean('Anonymize Data', default=True,
        help="Replace employee names with generic identifiers")
    sample_count = fields.Integer('Number of Samples', default=5,
        help="Number of random samples to generate")
    include_expected = fields.Boolean('Include Expected Values', default=True,
        help="Copy computed values as expected results for validation")

    # File import
    import_file = fields.Binary('Import File')
    import_filename = fields.Char('Filename')

    # Random generation options
    min_salary = fields.Float('Minimum Salary', default=5000000)
    max_salary = fields.Float('Maximum Salary', default=50000000)

    @api.onchange('source')
    def _onchange_source(self):
        """Clear selections when source changes."""
        if self.source != 'employees':
            self.employee_ids = False
        if self.source != 'payslips':
            self.payslip_ids = False

    def action_generate_samples(self):
        """Generate sample data based on selected source."""
        self.ensure_one()

        if self.source == 'employees':
            samples = self._generate_from_employees()
        elif self.source == 'payslips':
            samples = self._generate_from_payslips()
        elif self.source == 'file':
            samples = self._generate_from_file()
        elif self.source == 'random':
            samples = self._generate_random()
        else:
            raise UserError(_("Please select a data source"))

        if not samples:
            raise UserError(_("No sample data could be generated"))

        # Create sample records
        created = self.env['hr.formula.sample.data']
        for sample_data in samples:
            created |= self.env['hr.formula.sample.data'].create(sample_data)

        # Return action to view created samples
        return {
            'type': 'ir.actions.act_window',
            'name': _('Generated Samples'),
            'res_model': 'hr.formula.sample.data',
            'view_mode': 'list,form',
            'domain': [('id', 'in', created.ids)],
            'context': {'default_config_id': self.config_id.id},
        }

    def _generate_from_employees(self):
        """Generate samples from selected employees."""
        samples = []

        if not self.employee_ids:
            # Get all active employees
            employees = self.env['hr.employee'].search([
                ('active', '=', True),
            ], limit=self.sample_count)
        else:
            employees = self.employee_ids

        for idx, employee in enumerate(employees):
            sample_name = self._generate_sample_name(idx, employee)
            input_values = self._extract_employee_values(employee)

            samples.append({
                'config_id': self.config_id.id,
                'name': sample_name,
                'description': f"Sample from employee data",
                'source_type': 'employee',
                'source_employee_id': employee.id if not self.anonymize else False,
                'is_anonymized': self.anonymize,
                'input_values_json': json.dumps(input_values),
            })

        return samples

    def _generate_from_payslips(self):
        """Generate samples from existing payslips."""
        samples = []

        if not self.payslip_ids:
            # Get recent payslips
            payslips = self.env['hr.payslip'].search([
                ('state', '=', 'done'),
            ], order='date_to desc', limit=self.sample_count)
        else:
            payslips = self.payslip_ids

        for idx, payslip in enumerate(payslips):
            sample_name = self._generate_sample_name(idx, payslip.employee_id)
            input_values, expected_values = self._extract_payslip_values(payslip)

            sample_data = {
                'config_id': self.config_id.id,
                'name': sample_name,
                'description': f"Sample from payslip {payslip.name if not self.anonymize else 'XXX'}",
                'source_type': 'payslip',
                'source_payslip_id': payslip.id if not self.anonymize else False,
                'source_employee_id': payslip.employee_id.id if not self.anonymize else False,
                'is_anonymized': self.anonymize,
                'input_values_json': json.dumps(input_values),
            }

            if self.include_expected and expected_values:
                sample_data['expected_values_json'] = json.dumps(expected_values)

            samples.append(sample_data)

        return samples

    def _generate_random(self):
        """Generate random sample data."""
        samples = []

        # Get input rules
        input_rules = self.config_id.rule_ids.filtered(
            lambda r: r.column_type == 'input'
        )

        for idx in range(self.sample_count):
            sample_name = f"Random Sample {idx + 1}"
            input_values = {}

            for rule in input_rules:
                # Generate random value based on rule name/code
                value = self._generate_random_value(rule)
                input_values[rule.code] = value

            samples.append({
                'config_id': self.config_id.id,
                'name': sample_name,
                'description': "Randomly generated sample data",
                'source_type': 'manual',
                'is_anonymized': True,
                'input_values_json': json.dumps(input_values),
            })

        return samples

    def _generate_from_file(self):
        """Generate samples from an uploaded Excel file.

        Assumptions:
        - Header row exists near the top (rule code or name)
        - Subsequent rows: one sample per row
        """
        if not self.import_file:
            raise UserError(_("Please upload a file"))

        try:
            content = base64.b64decode(self.import_file)
        except Exception as e:
            raise UserError(_("Failed to read Excel file: %s") % e)

        rules = self.config_id.rule_ids
        filename = (self.import_filename or '').lower()
        if filename.endswith('.csv'):
            text = content.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                raise UserError(_("No data found in CSV file."))
            header_row = 0
            headers = rows[header_row]
            header_to_code = {}
            for idx, header in enumerate(headers):
                code = self._match_header_to_rule(header, rules)
                if code:
                    header_to_code[idx] = code
            if not header_to_code:
                raise UserError(_("No headers matched any rule codes or names."))

            samples = []
            for row_idx, row in enumerate(rows[header_row + 1:], start=1):
                if not any(row):
                    continue
                input_values = {}
                for idx, code in header_to_code.items():
                    if idx < len(row):
                        input_values[code] = self._serialize_value(row[idx])
                samples.append({
                    'config_id': self.config_id.id,
                    'name': _("File Sample %s") % row_idx,
                    'source_type': 'import',
                    'is_anonymized': self.anonymize,
                    'input_values_json': json.dumps(input_values),
                })
            return samples

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl library required. Install with: pip install openpyxl"))

        return self._generate_from_excel_multisheet(content, rules)

    def _match_header_to_rule(self, header_value, rules):
        if not header_value:
            return None
        header_text = str(header_value).strip()
        if not header_text:
            return None
        strict_key = self._normalize_header(header_text)
        for rule in rules:
            if rule.code and self._normalize_header(rule.code) == strict_key:
                return rule.code
            if rule.name and self._normalize_header(rule.name) == strict_key:
                return rule.code
        loose_key = self._normalize_header(header_text, loose=True)
        for rule in rules:
            if rule.code and self._normalize_header(rule.code, loose=True) == loose_key:
                return rule.code
            if rule.name and self._normalize_header(rule.name, loose=True) == loose_key:
                return rule.code
        return None

    def _generate_from_excel_multisheet(self, content, rules):
        """Generate samples from an Excel workbook, merging sheets by primary key."""
        from ..integrations.excel_connector import ExcelConnector

        connector = ExcelConnector(None)
        workbook_data = connector.load_workbook_multisheet(content, include_formulas=False)

        sheet_summaries = []
        for sheet_name in workbook_data['sheet_names']:
            sheet_data = connector.load_sheet_with_detection(sheet_name)
            headers = [h.get('value') for h in sheet_data.get('headers', []) if h.get('value')]
            primary_key = self._find_primary_key_header(headers)
            match_count = self._count_header_matches(headers, rules)
            if (not primary_key or match_count == 0) and connector.workbook:
                sheet = connector.workbook[sheet_name]
                _matched, header_row = self._match_headers_to_rules(sheet, rules)
                if header_row:
                    fallback_data = self._load_sheet_with_header_row(sheet, header_row)
                    headers = [h.get('value') for h in fallback_data.get('headers', []) if h.get('value')]
                    primary_key = self._find_primary_key_header(headers)
                    match_count = self._count_header_matches(headers, rules)
                    sheet_data = fallback_data
            sheet_summaries.append({
                'sheet_name': sheet_name,
                'headers': headers,
                'data_rows': sheet_data.get('data_rows', []),
                'primary_key': primary_key,
                'match_count': match_count,
                'row_count': sheet_data.get('total_rows', 0),
                'col_count': sheet_data.get('total_columns', 0),
            })

        candidates = [s for s in sheet_summaries if s['primary_key']]
        if not candidates:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            header_to_code, header_row = self._match_headers_to_rules(sheet, rules)
            if not header_to_code:
                raise UserError(_("No headers matched any rule codes or names."))

            samples = []
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = [cell.value for cell in sheet[row_idx]]
                if not any(row):
                    continue

                input_values = {}
                for idx, code in header_to_code.items():
                    if idx < len(row):
                        input_values[code] = self._serialize_value(row[idx])

                samples.append({
                    'config_id': self.config_id.id,
                    'name': _("File Sample %s") % (row_idx - 1),
                    'source_type': 'import',
                    'is_anonymized': self.anonymize,
                    'input_values_json': json.dumps(input_values),
                })

            return samples

        candidates.sort(key=lambda s: (s['match_count'], s['col_count'], s['row_count']), reverse=True)
        main_sheet = candidates[0]
        main_pk = main_sheet['primary_key']

        merged_rows = {}
        for row in main_sheet['data_rows']:
            pk_value = self._extract_row_value(row, main_pk)
            pk_key = self._normalize_code(pk_value)
            if not pk_key:
                continue
            base_row = row.copy()
            for header, value in row.items():
                base_row[f"{main_sheet['sheet_name']}|{header}"] = value
            merged_rows[pk_key] = base_row

        for sheet in sheet_summaries:
            if sheet['sheet_name'] == main_sheet['sheet_name']:
                continue

            pk_header = sheet['primary_key'] or main_pk
            if not pk_header:
                continue

            aux_map = {}
            for row in sheet['data_rows']:
                pk_value = self._extract_row_value(row, pk_header)
                pk_key = self._normalize_code(pk_value)
                if not pk_key:
                    continue
                aux_map[pk_key] = row

            for pk_key, base_row in merged_rows.items():
                aux_row = aux_map.get(pk_key)
                if aux_row:
                    for header, value in aux_row.items():
                        if self._normalize_header_key(header) == self._normalize_header_key(pk_header):
                            continue
                        base_row[f"{sheet['sheet_name']}|{header}"] = value
                        if header not in base_row:
                            base_row[header] = value
                else:
                    for header in sheet['headers']:
                        if self._normalize_header_key(header) == self._normalize_header_key(pk_header):
                            continue
                        base_row.setdefault(f"{sheet['sheet_name']}|{header}", None)
                        base_row.setdefault(header, None)

        samples = []
        for idx, raw_row in enumerate(merged_rows.values(), start=1):
            input_values = self._map_row_to_inputs(raw_row, rules)
            samples.append({
                'config_id': self.config_id.id,
                'name': _("File Sample %s") % idx,
                'source_type': 'import',
                'is_anonymized': self.anonymize,
                'input_values_json': json.dumps(input_values),
            })

        return samples

    def _map_row_to_inputs(self, raw_data, rules):
        input_values = {}

        def lookup_raw_value(candidates):
            for key in candidates:
                if key in raw_data:
                    return raw_data.get(key)
            normalized_map = {self._normalize_header_key(k): k for k in raw_data.keys()}
            for key in candidates:
                normalized_key = self._normalize_header_key(key)
                if normalized_key in normalized_map:
                    return raw_data.get(normalized_map[normalized_key])
            return None

        for rule in rules.filtered(lambda r: r.column_type == 'input'):
            candidates = []
            if rule.data_source_field:
                candidates.append(rule.data_source_field)
            if rule.source_sheet_name:
                if rule.name:
                    candidates.append(f"{rule.source_sheet_name}|{rule.name}")
                if rule.code:
                    candidates.append(f"{rule.source_sheet_name}|{rule.code}")
            if rule.code:
                candidates.append(rule.code)
            if rule.column_letter:
                candidates.append(rule.column_letter)
            if rule.name:
                candidates.append(rule.name)

            value = lookup_raw_value(candidates)
            if value is None:
                input_values[rule.code] = rule.default_value or 0.0
            else:
                input_values[rule.code] = self._serialize_value(value)

        return input_values

    def _load_sheet_with_header_row(self, sheet, header_row):
        from openpyxl.utils import get_column_letter

        headers = []
        for cell in sheet[header_row]:
            value = cell.value
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            headers.append({
                'column_letter': get_column_letter(cell.column),
                'value': text,
            })

        header_by_letter = {h['column_letter']: h['value'] for h in headers}
        data_rows = []
        for row in sheet.iter_rows(min_row=header_row + 1):
            row_data = {}
            for cell in row:
                col_letter = get_column_letter(cell.column)
                header = header_by_letter.get(col_letter)
                if header:
                    row_data[header] = cell.value
            if any(v is not None for v in row_data.values()):
                data_rows.append(row_data)

        return {
            'headers': headers,
            'data_rows': data_rows,
            'total_rows': len(data_rows),
            'total_columns': len(headers),
            'header_row': header_row,
            'data_start_row': header_row + 1,
        }

    def _normalize_header_key(self, value):
        if value is None:
            return ''
        return ''.join(ch for ch in str(value).lower() if ch.isalnum())

    def _find_primary_key_header(self, headers):
        candidates = [
            'employee_code', 'emp_code', 'emp code', 'emp. code',
            'employee id', 'employee_id', 'emp id', 'empid',
            'id no', 'id_no', 'id',
            'msnv', 'ma so nhan vien',
        ]
        for candidate in candidates:
            target = self._normalize_header_key(candidate)
            for header in headers:
                if self._normalize_header_key(header) == target:
                    return header
        return None

    def _count_header_matches(self, headers, rules):
        lookup = set()
        for rule in rules:
            if rule.code:
                lookup.add(self._normalize_header_key(rule.code))
            if rule.name:
                lookup.add(self._normalize_header_key(rule.name))
            if rule.source_sheet_name:
                lookup.add(self._normalize_header_key(rule.source_sheet_name))
        count = 0
        for header in headers:
            if self._normalize_header_key(header) in lookup:
                count += 1
        return count

    def _normalize_code(self, value):
        if value is None:
            return False
        try:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if float(value).is_integer():
                    return str(int(value))
                return str(value).strip()
            return str(value).strip()
        except Exception:
            return str(value)

    def _extract_row_value(self, row, header):
        if header in row:
            return row.get(header)
        normalized_map = {self._normalize_header_key(k): k for k in row.keys()}
        normalized_header = self._normalize_header_key(header)
        if normalized_header in normalized_map:
            return row.get(normalized_map[normalized_header])
        return None

    def _normalize_header(self, value, loose=False):
        """Normalize header strings for matching."""
        if value is None:
            return ''
        text = str(value).strip()
        if not text:
            return ''
        text = re.sub(r'\s+', ' ', text)
        if loose:
            return re.sub(r'[^a-z0-9]+', '', text.lower())
        return text.lower()

    def _match_headers_to_rules(self, sheet, rules):
        """Find the best header row and map headers to rule codes."""
        # Build lookup maps
        code_map = {}
        name_map = {}
        sheet_name_map = {}
        sheet_code_map = {}
        code_map_loose = {}
        name_map_loose = {}
        sheet_name_map_loose = {}
        sheet_code_map_loose = {}

        for rule in rules:
            code = rule.code or ''
            name = rule.name or ''
            sheet_name = rule.source_sheet_name or ''
            if code:
                code_key = self._normalize_header(code)
                code_map.setdefault(code_key, code)
                code_map_loose.setdefault(self._normalize_header(code, loose=True), code)
            if name:
                name_key = self._normalize_header(name)
                name_map.setdefault(name_key, code)
                name_map_loose.setdefault(self._normalize_header(name, loose=True), code)
            if sheet_name:
                sheet_key = self._normalize_header(sheet_name)
                sheet_key_loose = self._normalize_header(sheet_name, loose=True)
                if name:
                    sheet_name_map.setdefault((sheet_key, self._normalize_header(name)), code)
                    sheet_name_map_loose.setdefault((sheet_key_loose, self._normalize_header(name, loose=True)), code)
                if code:
                    sheet_code_map.setdefault((sheet_key, self._normalize_header(code)), code)
                    sheet_code_map_loose.setdefault((sheet_key_loose, self._normalize_header(code, loose=True)), code)

        def match_header(header_value):
            if not header_value:
                return None
            header_text = str(header_value).strip()
            if not header_text:
                return None

            strict_key = self._normalize_header(header_text)
            if strict_key in code_map:
                return code_map[strict_key]
            if strict_key in name_map:
                return name_map[strict_key]

            for delim in ['|', ':', '-', '/']:
                if delim in header_text:
                    left, right = header_text.split(delim, 1)
                    left_key = self._normalize_header(left)
                    right_key = self._normalize_header(right)
                    if (left_key, right_key) in sheet_name_map:
                        return sheet_name_map[(left_key, right_key)]
                    if (left_key, right_key) in sheet_code_map:
                        return sheet_code_map[(left_key, right_key)]

            loose_key = self._normalize_header(header_text, loose=True)
            if loose_key in code_map_loose:
                return code_map_loose[loose_key]
            if loose_key in name_map_loose:
                return name_map_loose[loose_key]

            for delim in ['|', ':', '-', '/']:
                if delim in header_text:
                    left, right = header_text.split(delim, 1)
                    left_key = self._normalize_header(left, loose=True)
                    right_key = self._normalize_header(right, loose=True)
                    if (left_key, right_key) in sheet_name_map_loose:
                        return sheet_name_map_loose[(left_key, right_key)]
                    if (left_key, right_key) in sheet_code_map_loose:
                        return sheet_code_map_loose[(left_key, right_key)]

            return None

        best_match_count = 0
        best_header_to_code = {}
        best_row = 1

        max_scan = min(sheet.max_row or 1, 10)
        for row_idx in range(1, max_scan + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            header_indices = {idx: hdr for idx, hdr in enumerate(row_values) if hdr is not None and str(hdr).strip()}
            header_to_code = {}
            for idx, hdr in header_indices.items():
                matched = match_header(hdr)
                if matched:
                    header_to_code[idx] = matched

            match_count = len(header_to_code)
            if match_count > best_match_count:
                best_match_count = match_count
                best_header_to_code = header_to_code
                best_row = row_idx

        return best_header_to_code, best_row

    # Helpers -------------------------------------------------------------
    @staticmethod
    def _serialize_value(val):
        """Convert Excel cell values to JSON-serializable types."""
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.isoformat()
        return val

    def _generate_sample_name(self, index, employee=None):
        """Generate anonymized sample name."""
        if self.anonymize or not employee:
            letters = string.ascii_uppercase
            return f"Sample {letters[index % 26]}"
        else:
            return f"Sample - {employee.name}"

    def _extract_employee_values(self, employee):
        """Extract input values from employee record."""
        values = {}

        # Get input rules and their data source fields
        input_rules = self.config_id.rule_ids.filtered(
            lambda r: r.column_type == 'input'
        )

        for rule in input_rules:
            source_field = rule.data_source_field

            if not source_field:
                # Use default value
                values[rule.code] = rule.default_value or 0.0
                continue

            # Try to get value from employee or contract
            value = self._get_field_value(employee, source_field)

            if value is None and employee.contract_id:
                value = self._get_field_value(employee.contract_id, source_field)

            values[rule.code] = value if value is not None else (rule.default_value or 0.0)

        return values

    def _extract_payslip_values(self, payslip):
        """Extract input and expected values from payslip."""
        input_values = {}
        expected_values = {}

        # Get rules
        input_rules = self.config_id.rule_ids.filtered(
            lambda r: r.column_type == 'input'
        )
        formula_rules = self.config_id.rule_ids.filtered(
            lambda r: r.column_type == 'formula'
        )

        # Extract input values from payslip lines
        for rule in input_rules:
            # Find matching payslip line
            line = payslip.line_ids.filtered(
                lambda l: l.code == rule.code or l.salary_rule_id == rule.salary_rule_id
            )[:1]

            if line:
                input_values[rule.code] = line.total
            else:
                input_values[rule.code] = rule.default_value or 0.0

        # Extract expected values for formula columns
        for rule in formula_rules:
            line = payslip.line_ids.filtered(
                lambda l: l.code == rule.code or l.salary_rule_id == rule.salary_rule_id
            )[:1]

            if line:
                expected_values[rule.code] = line.total

        return input_values, expected_values

    def _get_field_value(self, record, field_path):
        """Get field value from record using dot notation."""
        if not field_path or not record:
            return None

        parts = field_path.split('.')
        current = record

        for part in parts:
            if hasattr(current, part):
                current = getattr(current, part)
            else:
                return None

        if isinstance(current, models.Model):
            return None

        return current

    def _generate_random_value(self, rule):
        """Generate a *realistic* random value for an input rule.

        Drives off the rule's ``number_format`` (currency / percentage /
        integer / number) first, then refines by keyword on the friendly
        ``name`` AND ``code`` (the old version only scanned ``code``, so inputs
        like "OT Weekend (hrs)" or "Dependents" matched nothing and fell to a
        currency-sized default — producing nonsense like 34254 hours).
        """
        text = ('%s %s' % (rule.name or '', rule.code or '')).lower()
        fmt = rule.number_format or 'currency'

        def has(*words):
            return any(w in text for w in words)

        # Semantic keyword groups (checked across name + code).
        is_hours = has('hour', 'hrs', 'hr ', ' ot', 'overtime', 'otwd', 'otwe', 'otho')
        is_days = has('day')
        is_count = has('depend', 'child', 'count', 'qty', 'quantity',
                       'number of', 'no. of', 'headcount', 'persons')
        is_basic = has('basic', 'salary', 'wage', 'gross', 'base pay')
        is_allow = has('allowance', 'housing', 'hra', 'rent', 'transport',
                       'travel', 'conveyance', 'meal', 'food', 'lunch',
                       'medical', 'health', 'bonus', 'loan', 'advance',
                       'commission', 'incentive', 'responsib')

        # --- Percentage rates (0 - 0.30) ---
        if fmt == 'percentage' or has('rate', 'percent', 'pct', ' %'):
            return round(random.uniform(0.0, 0.30), 4)

        # --- Whole-number counts / hours / days ---
        if is_hours:
            return random.randint(0, 40)
        if is_days:
            return random.randint(20, 26)
        if is_count:
            return random.randint(0, 4)
        if fmt == 'integer':
            return random.randint(0, 10)

        # --- Monetary inputs ---
        if is_basic:
            return round(random.uniform(self.min_salary, self.max_salary), 0)
        if is_allow:
            # A sensible fraction of a freshly-drawn basic (5% - 30%).
            basic = random.uniform(self.min_salary, self.max_salary)
            return round(basic * random.uniform(0.05, 0.30), 0)

        # --- Generic numbers: small hours/days/count by keyword, else small money ---
        if fmt == 'number':
            return round(random.uniform(0, self.min_salary * 0.05), 0)

        # Default monetary fallback: a small allowance-sized amount (never huge).
        return round(random.uniform(0, self.min_salary * 0.10), 0)
