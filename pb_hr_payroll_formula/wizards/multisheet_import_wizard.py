# -*- coding: utf-8 -*-
"""
Multi-Sheet Import Wizard - Import payroll configuration from multi-worksheet Excel files.

This wizard guides users through importing formula configurations from Excel files
with multiple worksheets, handling:
- Dynamic header row detection
- Component type extraction from merged cells
- Cross-worksheet formula resolution
- Data source mapping for missing fields
"""

import base64
import io
import re
import logging
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

from ..formula_engine import excel_semantics

_logger = logging.getLogger(__name__)


class MultiSheetImportWizard(models.TransientModel):
    """
    Multi-worksheet Excel import wizard for formula configurations.

    Guides users through a 5-step import process:
    1. Upload File - Upload Excel file and analyze worksheets
    2. Select Worksheets - Choose which sheets to include and set main sheet
    3. Review Components - Preview all detected components across sheets
    4. Map Missing Fields - Assign data sources for fields not in Excel
    5. Confirm Import - Review and execute the import
    """
    _name = 'hr.formula.multisheet.import.wizard'
    _description = 'Multi-Sheet Formula Import Wizard'

    # ==========================================
    # WIZARD STATE
    # ==========================================
    state = fields.Selection([
        ('upload', 'Upload File'),
        ('select_sheets', 'Select Worksheets'),
        ('select_columns', 'Select Columns'),
        ('configure_order', 'Configure Order'),
        ('review_components', 'Review Components'),
        ('map_missing', 'Map Missing Fields'),
        ('confirm', 'Confirm Import'),
    ], string='State', default='upload', required=True)

    # ==========================================
    # FILE AND CONFIG
    # ==========================================
    config_id = fields.Many2one(
        'hr.formula.config',
        string='Formula Configuration',
        required=True,
        help="Target configuration to import rules into"
    )

    import_file = fields.Binary(
        string='Import File',
        help="Excel file (.xlsx) containing payroll structure"
    )

    import_filename = fields.Char(
        string='Filename'
    )

    # ==========================================
    # WORKSHEET SELECTION
    # ==========================================
    available_sheet_ids = fields.One2many(
        'hr.formula.multisheet.sheet.line',
        'wizard_id',
        string='Available Worksheets'
    )

    main_sheet_name = fields.Char(
        string='Main Worksheet',
        help="Primary worksheet containing the main payroll structure"
    )

    # ==========================================
    # COLUMN SELECTION (Step 3)
    # ==========================================
    column_selection_ids = fields.One2many(
        'hr.formula.multisheet.column.selection',
        'wizard_id',
        string='Column Selections',
        help="Columns available for selection from each worksheet"
    )

    # ==========================================
    # APPEND ORDER (Step 4)
    # ==========================================
    append_order_ids = fields.One2many(
        'hr.formula.multisheet.append.order',
        'wizard_id',
        string='Append Order',
        help="Order in which worksheets are appended to form the final structure"
    )

    # ==========================================
    # CROSS-SHEET RESOLUTION STATS
    # ==========================================
    cross_sheet_formula_count = fields.Integer(
        string='Cross-Sheet Formulas',
        compute='_compute_cross_sheet_stats'
    )

    resolved_formula_count = fields.Integer(
        string='Resolved Formulas',
        compute='_compute_cross_sheet_stats'
    )

    unresolved_formula_count = fields.Integer(
        string='Unresolved Formulas',
        compute='_compute_cross_sheet_stats'
    )

    # ==========================================
    # PRIMARY KEY
    # ==========================================
    primary_key_column = fields.Char(
        string='Primary Key Column',
        help="Column header used to match rows across worksheets (e.g., 'Employee ID')"
    )

    primary_key_column_letter = fields.Char(
        string='Primary Key Column Letter',
        help="Excel column letter for the primary key"
    )

    # ==========================================
    # PRIMARY KEY HELPERS
    # ==========================================
    def _resolve_primary_key_column(self, sheet_line, available_cols=None):
        """Resolve the primary key column name for a sheet line."""
        import json

        target = self.primary_key_column or sheet_line.primary_key_column_name
        if not target:
            return None

        if available_cols is None:
            try:
                available_cols = json.loads(sheet_line.available_column_names) if sheet_line.available_column_names else []
            except Exception:
                available_cols = []

        if not available_cols:
            return target

        if target in available_cols:
            return target

        target_lower = target.strip().lower()
        for col in available_cols:
            if col.lower() == target_lower:
                return col

        _logger.warning(
            "Primary key '%s' not found in available columns for sheet '%s'. Using provided value anyway.",
            target,
            sheet_line.sheet_name,
        )
        return target

    @api.onchange('primary_key_column')
    def _onchange_primary_key_column(self):
        """Apply the main primary key to all sheets when updated."""
        if not self.primary_key_column or not self.available_sheet_ids:
            return

        for sheet in self.available_sheet_ids:
            matched = self._resolve_primary_key_column(sheet)
            if matched:
                sheet.primary_key_column_name = matched

    # ==========================================
    # COMPONENT PREVIEW
    # ==========================================
    component_preview_ids = fields.One2many(
        'hr.formula.multisheet.component.preview',
        'wizard_id',
        string='Detected Components'
    )

    # ==========================================
    # MISSING FIELDS MAPPING
    # ==========================================
    missing_field_ids = fields.One2many(
        'hr.formula.multisheet.missing.field',
        'wizard_id',
        string='Missing Fields',
        help="Existing formula rules not found in the uploaded Excel file"
    )

    has_missing_fields = fields.Boolean(
        string='Has Missing Fields',
        compute='_compute_has_missing'
    )

    # ==========================================
    # IMPORT OPTIONS
    # ==========================================
    preserve_existing = fields.Boolean(
        string='Preserve Existing Rules',
        default=True,
        help="Keep existing rules and add new ones. If unchecked, existing rules will be replaced."
    )

    merge_duplicates = fields.Boolean(
        string='Merge Duplicate Columns',
        default=True,
        help="Merge columns with the same header as a single component"
    )

    update_existing = fields.Boolean(
        string='Update Existing Rules',
        default=True,
        help="Update existing rules if a matching code is found"
    )

    # ==========================================
    # IMPORT SUMMARY
    # ==========================================
    summary_html = fields.Html(
        string='Import Summary',
        compute='_compute_summary',
        sanitize=False
    )

    import_count = fields.Integer(
        string='Components to Import',
        compute='_compute_import_stats'
    )

    duplicate_count = fields.Integer(
        string='Duplicates Found',
        compute='_compute_import_stats'
    )

    formula_count = fields.Integer(
        string='Formulas Detected',
        compute='_compute_import_stats'
    )

    # ==========================================
    # COMPUTED FIELDS
    # ==========================================
    @api.depends('missing_field_ids')
    def _compute_has_missing(self):
        for rec in self:
            rec.has_missing_fields = bool(rec.missing_field_ids)

    @api.depends('component_preview_ids')
    def _compute_import_stats(self):
        for rec in self:
            previews = rec.component_preview_ids.filtered(lambda p: p.include_in_import)
            rec.import_count = len(previews)
            rec.duplicate_count = len(rec.component_preview_ids.filtered('is_duplicate'))
            rec.formula_count = len(previews.filtered(lambda p: p.column_type == 'formula'))

    @api.depends('component_preview_ids')
    def _compute_cross_sheet_stats(self):
        """Compute statistics about cross-sheet formula resolution."""
        for rec in self:
            formula_previews = rec.component_preview_ids.filtered(
                lambda p: p.column_type == 'formula' and p.excel_formula
            )
            # Count formulas with cross-sheet references (contain '!' pattern)
            cross_sheet = formula_previews.filtered(
                lambda p: "!" in (p.excel_formula or '')
            )
            rec.cross_sheet_formula_count = len(cross_sheet)
            # Resolved = has resolved_formula and it's different from original
            resolved = cross_sheet.filtered(
                lambda p: p.resolved_formula and p.resolved_formula != p.excel_formula
            )
            rec.resolved_formula_count = len(resolved)
            rec.unresolved_formula_count = rec.cross_sheet_formula_count - rec.resolved_formula_count

    @api.depends('component_preview_ids', 'missing_field_ids')
    def _compute_summary(self):
        for rec in self:
            html = "<div class='o_import_summary'>"
            html += f"<p><strong>Components to import:</strong> {rec.import_count}</p>"
            html += f"<p><strong>Formulas detected:</strong> {rec.formula_count}</p>"
            if rec.duplicate_count > 0:
                html += f"<p><strong>Duplicates (to merge):</strong> {rec.duplicate_count}</p>"
            if rec.has_missing_fields:
                html += f"<p><strong>Missing fields to map:</strong> {len(rec.missing_field_ids)}</p>"
            html += "</div>"
            rec.summary_html = html

    # ==========================================
    # STEP 1: UPLOAD FILE
    # ==========================================
    def action_analyze_file(self):
        """Analyze uploaded Excel file and populate sheet list."""
        self.ensure_one()

        if not self.import_file:
            raise UserError(_("Please upload an Excel file first."))

        if not self.import_filename or not self.import_filename.lower().endswith(('.xlsx', '.xls')):
            raise UserError(_("Please upload a valid Excel file (.xlsx or .xls)"))

        try:
            file_content = base64.b64decode(self.import_file)

            # Use Excel connector for multi-sheet loading
            from ..integrations import ExcelConnector
            connector = ExcelConnector(None)
            # Explicitly load with formulas to detect cross-sheet references
            workbook_data = connector.load_workbook_multisheet(file_content, include_formulas=True)

            # Clear existing sheet lines
            self.available_sheet_ids.unlink()

            # Create sheet lines
            sheet_lines = []

            # Load workbook to extract column names
            import openpyxl
            import json
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            formula_wb = None
            if self.config_id.use_color_coded_excel_import:
                formula_wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)

            for idx, sheet_name in enumerate(workbook_data['sheet_names']):
                sheet_info = workbook_data['sheets'][sheet_name]

                # Format referenced sheet names as comma-separated string
                referenced_sheets = sheet_info.get('formulas_referencing_sheets', [])
                referenced_sheet_names = ', '.join(referenced_sheets) if referenced_sheets else ''

                # Debug logging
                if sheet_info.get('has_formulas'):
                    _logger.info(f"Sheet '{sheet_name}' has formulas: {sheet_info.get('has_formulas')}")
                    _logger.info(f"Sheet '{sheet_name}' references other sheets: {sheet_info.get('references_other_sheets')}")
                    _logger.info(f"Sheet '{sheet_name}' referenced sheets list: {referenced_sheets}")

                # Extract column names from header row for primary key dropdown
                available_columns = []
                try:
                    ws = wb[sheet_name]
                    all_columns = []
                    header_row_num = sheet_info.get('detected_header_row', 1)
                    if self.config_id.use_color_coded_excel_import and formula_wb:
                        formula_ws = formula_wb[sheet_name]
                        color_info = self._get_color_coded_sheet_data(formula_ws, formula_ws)
                        header_row_num = color_info['header_block_end']
                        for col_idx, header_info in enumerate(color_info['sheet_data']['headers'], start=1):
                            col_name = header_info.get('value')
                            if col_name and str(col_name).strip():
                                all_columns.append((col_idx, str(col_name).strip()))
                    elif header_row_num and header_row_num > 0:
                        header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]

                        # Extract all non-empty column names
                        for col_idx, col_value in enumerate(header_row, start=1):
                            if col_value and str(col_value).strip():
                                col_name = str(col_value).strip()
                                all_columns.append((col_idx, col_name))

                        # Filter out date/day columns (like "Wed", "Thu", "26", "27", etc.)
                        # Keep only columns that look like data column headers
                        data_columns = []
                        for col_idx, col_name in all_columns:
                            # Skip if column name is:
                            # - A weekday name (Mon, Tue, Wed, Thu, Fri, Sat, Sun)
                            # - A short pure number (1-2 digits like "1", "26", "27" - likely date columns)
                            is_weekday = col_name in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun',
                                                     'Monday', 'Tuesday', 'Wednesday', 'Thursday',
                                                     'Friday', 'Saturday', 'Sunday']
                            # Only filter out SHORT numeric values (1-2 digits) to avoid filtering legitimate numeric codes
                            is_short_numeric = col_name.isdigit() and len(col_name) <= 2

                            # Only include if it's NOT a weekday or short number-only column
                            if not (is_weekday or is_short_numeric):
                                data_columns.append(col_name)

                        # Fallback: If all columns were filtered out, keep the original columns
                        # (This handles sheets where all headers might look like numbers but are valid)
                        if len(data_columns) == 0 and len(all_columns) > 0:
                            _logger.warning(f"Sheet '{sheet_name}': All columns were filtered out. Using all columns as fallback.")
                            data_columns = [col_name for col_idx, col_name in all_columns]

                        available_columns = data_columns
                        _logger.info(f"Sheet '{sheet_name}' has {len(available_columns)} data column headers: {available_columns}")
                        _logger.debug(f"Sheet '{sheet_name}' total columns before filtering: {len(all_columns)}, after filtering: {len(available_columns)}")
                except Exception as e:
                    _logger.warning(f"Could not extract column names from sheet '{sheet_name}': {e}")

                detected_header_row = sheet_info.get('detected_header_row', 1)
                row_count = sheet_info.get('max_row', 0) - sheet_info.get('detected_data_start_row', 2) + 1
                if self.config_id.use_color_coded_excel_import and formula_wb:
                    try:
                        formula_ws = formula_wb[sheet_name]
                        color_info = self._get_color_coded_sheet_data(formula_ws, formula_ws)
                        detected_header_row = color_info['header_block_end']
                        row_count = (formula_ws.max_row or 0) - color_info['formula_row'] + 1
                    except Exception as e:
                        _logger.warning(f"Color-coded detection failed for sheet '{sheet_name}': {e}")

                # Auto-set primary key column if wizard's primary_key_column matches one of the columns
                primary_key_col = None
                if self.primary_key_column and available_columns:
                    # Try exact match first
                    if self.primary_key_column in available_columns:
                        primary_key_col = self.primary_key_column
                    else:
                        # Try case-insensitive match
                        primary_key_lower = self.primary_key_column.lower()
                        for col in available_columns:
                            if col.lower() == primary_key_lower:
                                primary_key_col = col
                                break

                sheet_lines.append((0, 0, {
                    'wizard_id': self.id,
                    'sheet_name': sheet_name,
                    'is_selected': True,
                    'is_main_sheet': sheet_name == workbook_data['active_sheet'],
                    'detected_header_row': detected_header_row,
                    'column_count': sheet_info.get('max_column', 0),
                    'row_count': row_count,
                    'has_formulas': sheet_info.get('has_formulas', False),
                    'references_other_sheets': sheet_info.get('references_other_sheets', False),
                    'referenced_sheet_names': referenced_sheet_names,
                    'header_confidence': sheet_info.get('header_detection_confidence', 0.0),
                    'available_column_names': json.dumps(available_columns),
                    'primary_key_column_name': primary_key_col,
                }))

            self.write({
                'available_sheet_ids': sheet_lines,
                'main_sheet_name': workbook_data['active_sheet'],
                'state': 'select_sheets',
            })

            return self._return_wizard_action()

        except Exception as e:
            _logger.exception("Failed to analyze Excel file")
            raise UserError(_("Failed to analyze file: %s") % str(e))

    # ==========================================
    # STEP 2: SELECT WORKSHEETS -> STEP 3: SELECT COLUMNS
    # ==========================================
    def action_process_sheets(self):
        """
        Process selected sheets and populate column selection records.
        Transitions to select_columns state where user can choose columns.
        """
        self.ensure_one()

        selected_sheets = self.available_sheet_ids.filtered('is_selected')
        if not selected_sheets:
            raise UserError(_("Please select at least one worksheet."))

        main_sheets = selected_sheets.filtered('is_main_sheet')
        if not main_sheets:
            raise UserError(_("Please designate one worksheet as the main sheet."))

        if len(main_sheets) > 1:
            raise UserError(_("Only one worksheet can be designated as the main sheet."))

        self.main_sheet_name = main_sheets[0].sheet_name

        if not self.primary_key_column and not selected_sheets.filtered('primary_key_column_name'):
            raise UserError(_("Please specify the Primary Key Column to use for all worksheets."))

        # Validate primary key columns
        import json
        validation_errors = []
        for sheet in selected_sheets:
            # Resolve and apply the main primary key to each sheet
            try:
                available_cols = json.loads(sheet.available_column_names) if sheet.available_column_names else []
            except Exception as e:
                _logger.warning(f"Could not read available columns for sheet {sheet.sheet_name}: {e}")
                available_cols = []

            resolved_pk = self._resolve_primary_key_column(sheet, available_cols=available_cols)
            if resolved_pk:
                sheet.primary_key_column_name = resolved_pk
            else:
                validation_errors.append(
                    _("Sheet '%s': Please specify the Primary Key Column") % sheet.sheet_name
                )

        if validation_errors:
            raise UserError('\n\n'.join(validation_errors))

        try:
            file_content = base64.b64decode(self.import_file)

            from ..integrations import ExcelConnector
            connector = ExcelConnector(None)
            workbook_data = connector.load_workbook_multisheet(file_content, include_formulas=True)

            # Load formula workbook for detecting formulas
            import openpyxl
            formula_workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=False
            )

            # Clear existing column selections
            self.column_selection_ids.unlink()

            sheet_context = {}
            for sheet_line in selected_sheets:
                formula_sheet = formula_workbook[sheet_line.sheet_name]
                data_sheet = connector.workbook[sheet_line.sheet_name]
                if self.config_id.use_color_coded_excel_import:
                    color_info = self._get_color_coded_sheet_data(formula_sheet, formula_sheet)
                    sheet_context[sheet_line.id] = {
                        'sheet_data': color_info['sheet_data'],
                        'formula_sheet': formula_sheet,
                        'data_sheet': data_sheet,
                        'formula_columns': color_info['formula_columns'],
                        'red_header_columns': set(),
                        'red_data_columns': set(),
                        'header_map': color_info['header_map'],
                        'identifier_map': color_info['identifier_map'],
                        'report_visible_map': color_info['report_visible_map'],
                        'contract_component_map': color_info.get('contract_component_map', {}),
                        'requires_new_contract_map': color_info.get('requires_new_contract_map', {}),
                        'header_block_start': color_info['header_block_start'],
                        'formula_row': color_info['formula_row'],
                    }
                    continue

                sheet_data = connector.load_sheet_with_detection(sheet_line.sheet_name)
                red_header_cols = self._get_red_header_columns(
                    formula_sheet, sheet_data['header_row'], sheet_data['headers']
                )
                red_data_cols = self._get_red_data_columns(
                    formula_sheet, sheet_data['data_start_row'], sheet_data['headers']
                )
                red_data_cols -= red_header_cols
                skip_cols = red_header_cols | red_data_cols

                formula_columns = self._detect_formula_columns(
                    formula_sheet,
                    sheet_data['data_start_row'],
                    sheet_data['headers'],
                    skip_columns=skip_cols
                )
                sheet_context[sheet_line.id] = {
                    'sheet_data': sheet_data,
                    'formula_sheet': formula_sheet,
                    'data_sheet': data_sheet,
                    'formula_columns': formula_columns,
                    'red_header_columns': red_header_cols,
                    'red_data_columns': red_data_cols,
                }

            constant_lines = []
            constant_cell_mapping = {}
            constant_index = 0
            for sheet_line in selected_sheets:
                ctx = sheet_context[sheet_line.id]
                if self.config_id.use_color_coded_excel_import:
                    constant_scan_start = max(1, (ctx.get('header_block_start') or 1) - 1)
                    sheet_constants = self._detect_blue_constant_cells_color_coded(
                        ctx['formula_sheet'],
                        constant_scan_start,
                        ctx.get('formula_row') or ctx['sheet_data'].get('data_start_row', 1),
                        ctx.get('header_map') or {},
                        ctx.get('data_sheet'),
                    )
                else:
                    sheet_constants = self._collect_constants_for_sheet(
                        ctx['formula_sheet'],
                        ctx['sheet_data'],
                        ctx['formula_columns'],
                        ctx.get('data_sheet'),
                    )
                for const in sheet_constants:
                    cell_ref = const.get('original_cell')
                    if not cell_ref:
                        continue
                    sheet_key = self._normalize_sheet_key(sheet_line.sheet_name)
                    mapping_key = (sheet_key, cell_ref)
                    if mapping_key in constant_cell_mapping:
                        continue
                    new_col_letter = self._generate_extended_column_letter(constant_index)
                    constant_index += 1
                    value = const.get('value')
                    parsed_value, _was_pct = self._parse_percentage_value(value)
                    constant_cell_mapping[mapping_key] = new_col_letter
                    constant_lines.append({
                        'wizard_id': self.id,
                        'sheet_line_id': sheet_line.id,
                        'sequence': 10000 + constant_index,
                        'column_letter': new_col_letter,
                        'column_index': 10000 + constant_index,
                        'original_header': const.get('name') or 'Constant',
                        'component_type': 'Constant',
                        'is_selected': True,
                        'column_type': 'constant',
                        'sample_value': str(parsed_value)[:50] if value is not None else '',
                        'has_cross_sheet_ref': False,
                        'cross_sheet_formula': '',
                        'is_referenced_by_main': False,
                        'constant_cell_ref': cell_ref,
                        'report_visible': False,
                        'payslip_identifier_code': False,
                    })

            if constant_cell_mapping:
                for sheet_line in selected_sheets:
                    ctx = sheet_context[sheet_line.id]
                    for col_letter, info in ctx['formula_columns'].items():
                        formula = info.get('formula', '')
                        updated = self._update_formula_references(
                            formula,
                            constant_cell_mapping,
                            sheet_name=sheet_line.sheet_name,
                        )
                        if updated != formula:
                            info['formula'] = updated

            # Analyze cross-sheet references in main sheet to mark required columns
            main_sheet = main_sheets[0]
            main_ctx = sheet_context[main_sheet.id]
            main_formula_columns = main_ctx['formula_columns']

            # Find all cross-sheet references from main sheet
            cross_sheet_refs = self._extract_cross_sheet_references(main_formula_columns)

            # Process each selected sheet and create column selections
            column_lines = []
            for sheet_line in selected_sheets:
                ctx = sheet_context[sheet_line.id]
                sheet_data = ctx['sheet_data']
                formula_sheet = ctx['formula_sheet']
                data_sheet = ctx['data_sheet']
                formula_columns = ctx['formula_columns']
                red_header_cols = ctx.get('red_header_columns', set())
                red_data_cols = ctx.get('red_data_columns', set())
                identifier_map = ctx.get('identifier_map', {})
                report_visible_map = ctx.get('report_visible_map', {})
                contract_component_map = ctx.get('contract_component_map', {})
                requires_new_contract_map = ctx.get('requires_new_contract_map', {})

                for idx, header_info in enumerate(sheet_data['headers']):
                    col_letter = header_info['column_letter']
                    if col_letter in red_header_cols:
                        continue

                    # Check for cross-sheet formula references
                    has_cross_ref = False
                    cross_formula = ''
                    is_red_data_column = col_letter in red_data_cols
                    if col_letter in formula_columns:
                        formula_info = formula_columns[col_letter]
                        formula = formula_info.get('formula', '') if isinstance(formula_info, dict) else formula_info
                        # Use improved pattern matching for both quoted and unquoted sheet names
                        # Quoted: 'Sheet Name'!A1, Unquoted: SheetName!A1
                        quoted_refs = re.findall(r"'([^']+)'!", formula)
                        unquoted_refs = re.findall(r"(?<!['\w])([A-Za-z][A-Za-z0-9_\-]*)\s*!", formula)
                        all_refs = quoted_refs + unquoted_refs
                        # Filter out Excel functions
                        valid_refs = [ref for ref in all_refs if ref.upper() not in
                                     ['IF', 'SUM', 'SUMIF', 'AVERAGE', 'COUNT', 'MAX', 'MIN',
                                      'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'IFERROR']]
                        has_cross_ref = bool(valid_refs)
                        cross_formula = formula if has_cross_ref else ''

                    # Check if this column is referenced by main sheet
                    is_referenced = self._is_column_referenced(
                        sheet_line.sheet_name, col_letter, idx + 1, cross_sheet_refs
                    )

                    # Get sample value
                    sample = self._get_sample_value(
                        sheet_data, data_sheet, formula_sheet, header_info['value'], col_letter
                    )
                    uniform_val = False
                    if col_letter not in formula_columns:
                        uniform_val = self._column_uniform_value(
                            sheet_data, data_sheet, header_info['value'], col_letter)

                    column_lines.append({
                        'wizard_id': self.id,
                        'sheet_line_id': sheet_line.id,
                        'sequence': idx * 10,
                        'column_letter': col_letter,
                        'column_index': idx,
                        'original_header': header_info['value'],
                        'component_type': header_info.get('component_type') or '',
                        'payslip_identifier_code': identifier_map.get(col_letter),
                        'report_visible': bool(report_visible_map.get(col_letter)),
                        'is_contract_component': bool(contract_component_map.get(col_letter)),
                        'requires_new_contract': bool(requires_new_contract_map.get(col_letter)),
                        'is_selected': True,  # All columns selected by default
                        'column_type': 'input' if is_red_data_column else
                                       ('formula' if col_letter in formula_columns else 'input'),
                        'sample_value': sample,
                        'uniform_value': uniform_val or False,
                        'has_cross_sheet_ref': has_cross_ref,
                        'cross_sheet_formula': cross_formula,
                        'is_referenced_by_main': is_referenced,
                    })

            column_lines.extend(constant_lines)

            # Create column selection records
            for col_data in column_lines:
                self.env['hr.formula.multisheet.column.selection'].create(col_data)

            self.state = 'select_columns'
            return self._return_wizard_action()

        except Exception as e:
            _logger.exception("Failed to process worksheets")
            raise UserError(_("Failed to process worksheets: %s") % str(e))

    def _extract_cross_sheet_references(self, formula_columns):
        """
        Extract all cross-sheet references from formulas.

        Returns a dict: {sheet_name: [{'col_index': N, 'col_letter': 'XX', 'formula_col': 'YY'}, ...]}
        """
        cross_refs = {}

        # Patterns for VLOOKUP and direct sheet references
        # Quoted patterns - for sheet names with spaces or special chars
        vlookup_quoted_pattern = re.compile(
            r"VLOOKUP\s*\([^,]+,\s*'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*(\d+)",
            re.IGNORECASE
        )
        # Unquoted patterns - for simple sheet names
        vlookup_unquoted_pattern = re.compile(
            r"VLOOKUP\s*\([^,]+,\s*([A-Za-z][A-Za-z0-9_\-]*)\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*(\d+)",
            re.IGNORECASE
        )
        sumif_quoted_pattern = re.compile(
            r"SUMIF\s*\([^,]*'([^']+)'\s*!\s*\$?([A-Z]+)",
            re.IGNORECASE
        )
        sumif_unquoted_pattern = re.compile(
            r"SUMIF\s*\([^,]*([A-Za-z][A-Za-z0-9_\-]*)\s*!\s*\$?([A-Z]+)",
            re.IGNORECASE
        )
        direct_quoted_pattern = re.compile(
            r"'([^']+)'\s*!\s*\$?([A-Z]+)\$?(\d+)",
            re.IGNORECASE
        )
        direct_unquoted_pattern = re.compile(
            r"(?<!['\w])([A-Za-z][A-Za-z0-9_\-]*)\s*!\s*\$?([A-Z]+)\$?(\d+)",
            re.IGNORECASE
        )

        for col_letter, formula_info in formula_columns.items():
            formula = formula_info.get('formula', '') if isinstance(formula_info, dict) else formula_info

            # Find VLOOKUP references (both quoted and unquoted)
            for match in vlookup_quoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                start_col = match.group(2).upper()
                col_index = int(match.group(4))
                start_idx = self._column_letter_to_index(start_col)
                target_idx = start_idx + col_index - 1
                target_col = self._index_to_column_letter(target_idx)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': target_idx,
                    'col_letter': target_col,
                    'formula_col': col_letter,
                    'type': 'vlookup',
                })

            for match in vlookup_unquoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                start_col = match.group(2).upper()
                col_index = int(match.group(4))
                start_idx = self._column_letter_to_index(start_col)
                target_idx = start_idx + col_index - 1
                target_col = self._index_to_column_letter(target_idx)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': target_idx,
                    'col_letter': target_col,
                    'formula_col': col_letter,
                    'type': 'vlookup',
                })

            # Find SUMIF references (both quoted and unquoted)
            for match in sumif_quoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                col_letter_ref = match.group(2).upper()
                col_idx = self._column_letter_to_index(col_letter_ref)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': col_idx,
                    'col_letter': col_letter_ref,
                    'formula_col': col_letter,
                    'type': 'sumif',
                })

            for match in sumif_unquoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                col_letter_ref = match.group(2).upper()
                col_idx = self._column_letter_to_index(col_letter_ref)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': col_idx,
                    'col_letter': col_letter_ref,
                    'formula_col': col_letter,
                    'type': 'sumif',
                })

            # Find direct references (both quoted and unquoted)
            for match in direct_quoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                col_letter_ref = match.group(2).upper()
                col_idx = self._column_letter_to_index(col_letter_ref)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': col_idx,
                    'col_letter': col_letter_ref,
                    'formula_col': col_letter,
                    'type': 'direct',
                })

            for match in direct_unquoted_pattern.finditer(formula):
                sheet_name = self._normalize_sheet_key(match.group(1))
                col_letter_ref = match.group(2).upper()
                col_idx = self._column_letter_to_index(col_letter_ref)

                if sheet_name not in cross_refs:
                    cross_refs[sheet_name] = []
                cross_refs[sheet_name].append({
                    'col_index': col_idx,
                    'col_letter': col_letter_ref,
                    'formula_col': col_letter,
                    'type': 'direct',
                })

        return cross_refs

    def _is_column_referenced(self, sheet_name, col_letter, col_index, cross_refs):
        """Check if a column is referenced by main sheet formulas."""
        sheet_name_lower = self._normalize_sheet_key(sheet_name)
        for ref_sheet, refs in cross_refs.items():
            if self._normalize_sheet_key(ref_sheet) == sheet_name_lower:
                for ref in refs:
                    if ref['col_letter'] == col_letter or ref['col_index'] == col_index - 1:
                        return True
        return False

    def _normalize_sheet_key(self, name):
        if not name:
            return ''
        text = str(name).replace('\u00A0', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip().lower()

    def _column_letter_to_index(self, col_letter):
        """Convert column letter to 0-based index (A=0, B=1, ...)."""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def _index_to_column_letter(self, index):
        """Convert 0-based index to column letter (0=A, 1=B, ...)."""
        result = ""
        idx = index
        while True:
            result = chr(ord('A') + (idx % 26)) + result
            idx = idx // 26 - 1
            if idx < 0:
                break
        return result

    # ==========================================
    # STEP 3: SELECT COLUMNS -> STEP 4: CONFIGURE ORDER
    # ==========================================
    def action_configure_order(self):
        """
        Validate column selection and create append order records.
        Blocks if required columns (referenced by main sheet) are not selected.
        """
        self.ensure_one()

        # Validate: Check all cross-sheet references can be resolved
        validation_errors = self._validate_column_selection()
        if validation_errors:
            error_msg = _("Cannot proceed. The following columns are required by main sheet formulas but are not selected:\n\n")
            for error in validation_errors:
                error_msg += f"• {error['sheet']}: Column {error['col']} ({error['header']}) - used in {error['formula_col']}\n"
            error_msg += _("\nPlease select these columns to continue.")
            raise UserError(error_msg)

        # Clear existing append order
        self.append_order_ids.unlink()

        # Create append order records
        selected_sheets = self.available_sheet_ids.filtered('is_selected')
        order_lines = []

        # Main sheet first (sequence 0)
        main_sheet = selected_sheets.filtered('is_main_sheet')
        if main_sheet:
            order_lines.append({
                'wizard_id': self.id,
                'sheet_line_id': main_sheet.id,
                'append_sequence': 0,
            })

        # Other sheets with sequence 10, 20, 30...
        other_sheets = selected_sheets.filtered(lambda s: not s.is_main_sheet)
        for idx, sheet_line in enumerate(other_sheets.sorted('sheet_name')):
            order_lines.append({
                'wizard_id': self.id,
                'sheet_line_id': sheet_line.id,
                'append_sequence': (idx + 1) * 10,
            })

        # Create records
        for order_data in order_lines:
            self.env['hr.formula.multisheet.append.order'].create(order_data)

        self.state = 'configure_order'
        return self._return_wizard_action()

    def _validate_column_selection(self):
        """
        Validate that all columns referenced by main sheet formulas are selected.
        Returns list of validation errors, empty if valid.
        """
        errors = []
        main_sheet = self.available_sheet_ids.filtered('is_main_sheet')
        if not main_sheet:
            return errors

        # Re-extract cross-sheet references
        try:
            file_content = base64.b64decode(self.import_file)

            from ..integrations import ExcelConnector
            import openpyxl
            connector = ExcelConnector(None)
            connector.load_workbook_multisheet(file_content, include_formulas=True)

            formula_workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=False
            )

            main_formula_sheet = formula_workbook[main_sheet.sheet_name]
            if self.config_id.use_color_coded_excel_import:
                color_info = self._get_color_coded_sheet_data(main_formula_sheet, main_formula_sheet)
                main_formula_columns = color_info['formula_columns']
            else:
                main_data = connector.load_sheet_with_detection(main_sheet.sheet_name)
                red_header_cols = self._get_red_header_columns(
                    main_formula_sheet, main_data['header_row'], main_data['headers']
                )
                red_data_cols = self._get_red_data_columns(
                    main_formula_sheet, main_data['data_start_row'], main_data['headers']
                )
                skip_cols = red_header_cols | red_data_cols
                main_formula_columns = self._detect_formula_columns(
                    main_formula_sheet,
                    main_data['data_start_row'],
                    main_data['headers'],
                    skip_columns=skip_cols
                )
            cross_refs = self._extract_cross_sheet_references(main_formula_columns)

            # Check each reference against selected columns
            for sheet_name, refs in cross_refs.items():
                # Find the sheet line
                sheet_line = self.available_sheet_ids.filtered(
                    lambda s: self._normalize_sheet_key(s.sheet_name) == self._normalize_sheet_key(sheet_name)
                )
                if not sheet_line:
                    continue  # Sheet not even selected

                if not sheet_line.is_selected:
                    # Whole sheet not selected - error
                    for ref in refs:
                        errors.append({
                            'sheet': sheet_name,
                            'col': ref['col_letter'],
                            'header': f"Column {ref['col_letter']}",
                            'formula_col': ref['formula_col'],
                        })
                    continue

                # Check individual columns
                for ref in refs:
                    col_selection = self.column_selection_ids.filtered(
                        lambda c: c.sheet_line_id.id == sheet_line.id and
                                  (c.column_letter == ref['col_letter'] or
                                   c.column_index == ref['col_index'])
                    )
                    if col_selection and not col_selection.is_selected:
                        errors.append({
                            'sheet': sheet_name,
                            'col': ref['col_letter'],
                            'header': col_selection.original_header or f"Column {ref['col_letter']}",
                            'formula_col': ref['formula_col'],
                        })

        except Exception as e:
            _logger.warning(f"Validation failed: {e}")

        return errors

    # ==========================================
    # STEP 4: CONFIGURE ORDER -> STEP 5: REVIEW COMPONENTS
    # ==========================================
    def action_process_with_resolution(self):
        """
        Build final component list with cross-sheet formula resolution.
        Assigns new column letters based on append order.
        """
        self.ensure_one()

        try:
            file_content = base64.b64decode(self.import_file)

            from ..integrations import ExcelConnector
            import openpyxl
            connector = ExcelConnector(None)
            connector.load_workbook_multisheet(file_content, include_formulas=True)

            formula_workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=False
            )

            # Clear existing previews
            self.component_preview_ids.unlink()

            # Build column mapping: (sheet_name, orig_col) -> new_col
            column_mapping = {}
            current_col_index = 0
            all_components = []
            seen_codes = set()
            constant_placeholder_mapping = {}
            placeholder_to_final = {}
            constant_selections = self.column_selection_ids.filtered(
                lambda c: c.is_selected and c.column_type == 'constant' and c.constant_cell_ref
            )
            for idx, const in enumerate(constant_selections):
                cell_ref = const.constant_cell_ref
                if not cell_ref:
                    continue
                sheet_key = self._normalize_sheet_key(const.sheet_line_id.sheet_name)
                mapping_key = (sheet_key, cell_ref)
                if mapping_key not in constant_placeholder_mapping:
                    constant_placeholder_mapping[mapping_key] = f"__CONST_{idx}__"

            # Get sheets in append order
            ordered_sheets = self.append_order_ids.sorted('append_sequence')

            for order_rec in ordered_sheets:
                sheet_line = order_rec.sheet_line_id

                # Get selected columns for this sheet in order
                selected_cols = self.column_selection_ids.filtered(
                    lambda c: c.sheet_line_id.id == sheet_line.id and c.is_selected
                ).sorted('sequence')
                normal_cols = selected_cols.filtered(lambda c: c.column_type != 'constant')
                constant_cols = selected_cols.filtered(lambda c: c.column_type == 'constant')

                formula_sheet = formula_workbook[sheet_line.sheet_name]

                # Detect formula columns for this sheet
                if self.config_id.use_color_coded_excel_import:
                    color_info = self._get_color_coded_sheet_data(formula_sheet, formula_sheet)
                    formula_columns = color_info['formula_columns']
                else:
                    sheet_data = connector.load_sheet_with_detection(sheet_line.sheet_name)
                    red_header_cols = self._get_red_header_columns(
                        formula_sheet, sheet_data['header_row'], sheet_data['headers']
                    )
                    red_data_cols = self._get_red_data_columns(
                        formula_sheet, sheet_data['data_start_row'], sheet_data['headers']
                    )
                    skip_cols = red_header_cols | red_data_cols
                    formula_columns = self._detect_formula_columns(
                        formula_sheet,
                        sheet_data['data_start_row'],
                        sheet_data['headers'],
                        skip_columns=skip_cols
                    )
                if constant_placeholder_mapping:
                    for col_letter, info in formula_columns.items():
                        formula = info.get('formula', '')
                        updated = self._update_formula_references(
                            formula,
                            constant_placeholder_mapping,
                            sheet_name=sheet_line.sheet_name,
                        )
                        if updated != formula:
                            info['formula'] = updated

                for col_sel in normal_cols:
                    # Assign new column letter
                    new_col_letter = self._index_to_column_letter(current_col_index)

                    # Store mapping for cross-sheet resolution
                    sheet_key = self._normalize_sheet_key(sheet_line.sheet_name)
                    column_mapping[(sheet_key, col_sel.column_letter)] = new_col_letter
                    column_mapping[(sheet_key, col_sel.column_index)] = new_col_letter

                    # Get formula if any
                    excel_formula = ''
                    if col_sel.column_type == 'formula' and col_sel.column_letter in formula_columns:
                        excel_formula = formula_columns[col_sel.column_letter].get('formula', '')

                    # Generate unique code
                    code = self._generate_code(col_sel.original_header, seen_codes)
                    seen_codes.add(code)

                    # Determine data source
                    if col_sel.column_type == 'formula':
                        data_source = 'formula'
                    else:
                        data_source = 'excel'

                    component = {
                        'wizard_id': self.id,
                        'source_sheet': sheet_line.sheet_name,
                        'column_letter': new_col_letter,
                        'original_header': col_sel.original_header,
                        'generated_code': code,
                        'generated_name': col_sel.original_header,
                        'component_type': col_sel.component_type or '',
                        'payslip_identifier_code': col_sel.payslip_identifier_code or False,
                        'report_visible': bool(col_sel.report_visible),
                        'is_contract_component': bool(col_sel.is_contract_component),
                        'requires_new_contract': bool(col_sel.requires_new_contract),
                        'column_type': col_sel.column_type,
                        'excel_formula': excel_formula,
                        'resolved_formula': '',  # Will be filled during resolution
                        'sample_value': col_sel.sample_value,
                        'uniform_value': col_sel.uniform_value or False,
                        'is_duplicate': False,
                        'include_in_import': True,
                        'is_in_excel': True,
                        'data_source': data_source,
                        '_original_col': col_sel.column_letter,
                    }
                    all_components.append(component)
                    current_col_index += 1

                for col_sel in constant_cols:
                    # Assign new column letter to constants (just like normal columns)
                    new_col_letter = self._index_to_column_letter(current_col_index)

                    # Store mapping for same-sheet resolution
                    # This is CRITICAL: constants need to be in column_mapping so that
                    # formulas referencing them can be resolved properly
                    sheet_key = self._normalize_sheet_key(sheet_line.sheet_name)
                    constant_key = (
                        self._extract_column_letter_from_cell_ref(col_sel.constant_cell_ref)
                        or col_sel.column_letter
                    )
                    if constant_key and (sheet_key, constant_key) not in column_mapping:
                        column_mapping[(sheet_key, constant_key)] = new_col_letter
                    column_mapping[(sheet_key, col_sel.column_index)] = new_col_letter
                    if col_sel.constant_cell_ref:
                        placeholder_token = constant_placeholder_mapping.get(
                            (sheet_key, col_sel.constant_cell_ref)
                        )
                        if placeholder_token:
                            placeholder_to_final[f"{placeholder_token}2"] = f"{new_col_letter}2"

                    code = self._generate_code(col_sel.original_header, seen_codes)
                    seen_codes.add(code)
                    component = {
                        'wizard_id': self.id,
                        'source_sheet': sheet_line.sheet_name,
                        'column_letter': new_col_letter,  # Use NEW column letter
                        'original_header': col_sel.original_header,
                        'generated_code': code,
                        'generated_name': col_sel.original_header,
                        'component_type': col_sel.component_type or 'Constant',
                        'payslip_identifier_code': False,
                        'report_visible': False,
                        'is_contract_component': False,
                        'requires_new_contract': False,
                        'column_type': 'constant',
                        'excel_formula': '',
                        'resolved_formula': '',
                        'sample_value': col_sel.sample_value,
                        'is_duplicate': False,
                        'include_in_import': True,
                        'is_in_excel': True,
                        'data_source': 'manual',
                        '_original_col': col_sel.column_letter,  # Track original column
                    }
                    all_components.append(component)
                    current_col_index += 1  # Increment for next column

            # Now resolve formulas (both same-sheet and cross-sheet references)
            def replace_constant_placeholders(formula):
                if not formula or not placeholder_to_final:
                    return formula
                for token, final in placeholder_to_final.items():
                    formula = formula.replace(token, final)
                return formula

            for component in all_components:
                if component['excel_formula']:
                    # First resolve same-sheet column references (e.g., I3 -> FS3)
                    formula_with_same_sheet = self._resolve_same_sheet_formula(
                        component['excel_formula'],
                        component['source_sheet'],
                        column_mapping
                    )

                    # Then resolve cross-sheet references (e.g., 'OtherSheet'!A1 -> XX1)
                    resolved = self._resolve_cross_sheet_formula(
                        formula_with_same_sheet,
                        column_mapping
                    )

                    resolved = replace_constant_placeholders(resolved)
                    component['resolved_formula'] = resolved
                    # Use the fully resolved formula
                    component['excel_formula'] = resolved

            # Create component preview records
            for comp in all_components:
                # Remove internal tracking field
                comp.pop('_original_col', None)
                self.env['hr.formula.multisheet.component.preview'].create(comp)

            self.state = 'review_components'
            return self._return_wizard_action()

        except Exception as e:
            _logger.exception("Failed to process sheets with resolution")
            raise UserError(_("Failed to process sheets: %s") % str(e))

    # WP-E / D-E1: unresolved references must degrade VISIBLY, never become a
    # silent 0. This sentinel (a) survives into the stored formula so the value
    # is obviously wrong, (b) is matched by the preview mixin's SHEET_REF_RE
    # (the trailing '!' makes `REF!` match) so the row goes red, and (c) makes
    # the downstream Excel→Python converter fail loudly (syntax error → the rule
    # carries has_evaluation_error) instead of quietly returning 0. It is inert
    # to the column-ref regexes because nothing after the '!' looks like a
    # <col><row> cell reference.
    _UNRESOLVED_MARK = '#REF!'

    def _resolve_cross_sheet_formula(self, formula, column_mapping):
        """
        Resolve cross-sheet references in a formula to simple column references.

        VLOOKUP(B4,'TimeTB 2'!$C$4:$AU$11,45,0) → DM2
        SUMIF(Others!$B$8:$B$15,$B4,Others!$F$8:$F$15) → SUMIF(XX:XX,$B4,YY:YY)

        Unresolved references are replaced with ``#REF!`` (D-E1), never ``0`` —
        the import preview red-lines them and the converter refuses them loudly.
        """
        if not formula:
            return formula

        result = formula

        # Resolve VLOOKUP: Replace entire VLOOKUP with simple column reference
        vlookup_pattern = re.compile(
            r"VLOOKUP\s*\([^,]+,\s*'?([^'!]+)'?\s*!\s*\$?([A-Z]+)\$?\d*:\$?[A-Z]+\$?\d*,\s*(\d+),\s*[^)]+\)",
            re.IGNORECASE
        )

        def resolve_vlookup(match):
            raw_sheet_name = match.group(1)
            sheet_name = self._normalize_sheet_key(raw_sheet_name)
            start_col = match.group(2).upper()
            col_index = int(match.group(3))

            # Calculate target column
            start_idx = self._column_letter_to_index(start_col)
            target_idx = start_idx + col_index - 1
            target_col = self._index_to_column_letter(target_idx)

            # Look up new column
            new_col = column_mapping.get((sheet_name, target_col))
            if not new_col:
                new_col = column_mapping.get((sheet_name, target_idx))

            if new_col:
                # Return just the code - the formula converter handles codes in column_map.values()
                return new_col
            else:
                _logger.warning(
                    "VLOOKUP unresolved: sheet='%s', target_col='%s', col_index=%s. "
                    "Marking #REF! (was silently 0 before WP-E).",
                    sheet_name, target_col, col_index
                )
                return self._UNRESOLVED_MARK  # D-E1: visible, never silent 0

        result = vlookup_pattern.sub(resolve_vlookup, result)

        # Resolve SUMIF: Replace range references
        sumif_pattern = re.compile(
            r"SUMIF\s*\(\s*'?([^'!]+)'?\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*([^,]+),\s*'?([^'!]+)'?\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*\s*\)",
            re.IGNORECASE
        )

        def resolve_sumif(match):
            criteria_sheet = self._normalize_sheet_key(match.group(1))
            criteria_col = match.group(2).upper()
            criteria = match.group(4)
            sum_sheet = self._normalize_sheet_key(match.group(5))
            sum_col = match.group(6).upper()

            new_criteria_col = column_mapping.get((criteria_sheet, criteria_col))
            if not new_criteria_col:
                idx = self._column_letter_to_index(criteria_col)
                new_criteria_col = column_mapping.get((criteria_sheet, idx))

            new_sum_col = column_mapping.get((sum_sheet, sum_col))
            if not new_sum_col:
                idx = self._column_letter_to_index(sum_col)
                new_sum_col = column_mapping.get((sum_sheet, idx))

            if new_criteria_col and new_sum_col:
                return f"SUMIF({new_criteria_col}:{new_criteria_col},{criteria},{new_sum_col}:{new_sum_col})"
            else:
                _logger.warning(
                    "SUMIF unresolved: criteria='%s', sum='%s'. Marking #REF!.",
                    criteria_sheet, sum_sheet
                )
                return self._UNRESOLVED_MARK  # D-E1

        result = sumif_pattern.sub(resolve_sumif, result)

        # Resolve direct references: 'Sheet'!A4 -> NewCol
        # D-E2: the sheet-name token is ANCHORED and constrained. The old
        # `'?([^'!]+)'?` greedily swallowed everything up to the '!' — including a
        # leading `=IF(` — so `=IF(Sheet2!B2>0,1,0)` matched with sheet name
        # `=IF(Sheet2`, never resolved, and (returning "0") was shredded into
        # `0>0,1,0)`. The token is now either a quoted name or a bare Excel sheet
        # name (letter/underscore/unicode start; word-chars, dot, unicode after —
        # NEVER operators or spaces, which Excel forces to be quoted), with a left
        # boundary so it cannot eat a preceding function call or identifier.
        direct_pattern = re.compile(
            r"(?<![\w!.'])(?:'([^']+)'|([A-Za-z_\u00C0-\uffff][\w.\u00C0-\uffff]*))\s*!\s*\$?([A-Z]+)\$?(\d+)",
            re.IGNORECASE
        )

        def resolve_direct(match):
            raw_sheet = match.group(1) if match.group(1) is not None else match.group(2)
            sheet_name = self._normalize_sheet_key(raw_sheet)
            col = match.group(3).upper()

            new_col = column_mapping.get((sheet_name, col))
            if not new_col:
                idx = self._column_letter_to_index(col)
                new_col = column_mapping.get((sheet_name, idx))

            if new_col:
                # Return just the code - the formula converter handles codes in column_map.values()
                # Don't add row number suffix as it prevents proper code recognition
                return new_col
            else:
                _logger.warning(
                    "Direct reference unresolved: sheet='%s', col='%s'. Marking #REF! "
                    "(preserved visibly; was silently 0 before WP-E).",
                    sheet_name, col
                )
                return self._UNRESOLVED_MARK  # D-E1

        result = direct_pattern.sub(resolve_direct, result)

        return result

    def _resolve_same_sheet_formula(self, formula, sheet_name, column_mapping):
        """
        Resolve same-sheet column references in a formula.

        When columns are reordered during import (e.g., I -> FS, J -> FT, K -> FU),
        formulas that reference those columns need to be updated.

        Example:
            Original: =I3+J3+K3
            After resolution: =FS3+FT3+FU3

        Args:
            formula: The Excel formula string
            sheet_name: The name of the sheet this formula belongs to
            column_mapping: Dictionary mapping (sheet_name, old_col) -> new_col

        Returns:
            Formula with updated column references
        """
        if not formula or not formula.startswith('='):
            return formula

        result = formula
        sheet_key = self._normalize_sheet_key(sheet_name)

        # First, temporarily mask cross-sheet references to protect them
        # This prevents us from modifying column refs inside cross-sheet ranges like 'Sheet'!C4:Q11
        cross_sheet_pattern = re.compile(
            r"'[^']+'![^,\s\)]+|[A-Za-z0-9_]+![^,\s\)]+"
        )
        placeholders = {}
        placeholder_idx = [0]  # Use list for closure modification

        def mask_cross_sheet(match):
            placeholder = f"__CROSSSHEET_{placeholder_idx[0]}__"
            placeholders[placeholder] = match.group(0)
            placeholder_idx[0] += 1
            return placeholder

        result = cross_sheet_pattern.sub(mask_cross_sheet, result)

        # Resolve same-sheet VLOOKUP to direct column references.
        vlookup_pattern = re.compile(
            r"VLOOKUP\s*\(\s*[^,]+,\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*(\d+)\s*,\s*[^)]+\)",
            re.IGNORECASE
        )

        def resolve_same_sheet_vlookup(match):
            start_col = match.group(1).upper()
            end_col = match.group(2).upper()
            col_index = int(match.group(3))

            try:
                start_idx = self._column_letter_to_index(start_col)
                end_idx = self._column_letter_to_index(end_col)
            except Exception:
                return self._UNRESOLVED_MARK  # D-E1

            base_idx = min(start_idx, end_idx)
            target_idx = base_idx + col_index - 1
            target_col = self._index_to_column_letter(target_idx)

            new_col = column_mapping.get((sheet_key, target_col))
            if not new_col:
                new_col = column_mapping.get((sheet_key, target_idx))

            return new_col or self._UNRESOLVED_MARK  # D-E1

        result = vlookup_pattern.sub(resolve_same_sheet_vlookup, result)

        # Pattern to match column references in formulas
        # Matches: A1, $A$1, $A1, A$1, AA123, etc.
        # Cross-sheet refs are already masked, so we just need to avoid word boundaries
        same_sheet_pattern = re.compile(
            r'(?<![!\w])\$?([A-Z]+)\$?(\d+)(?!\w)',
            re.IGNORECASE
        )

        def replace_column(match):
            old_col = match.group(1).upper()
            row = match.group(2)

            # Look up the new column letter
            new_col = column_mapping.get((sheet_key, old_col))
            if not new_col:
                # Try by column index
                try:
                    idx = self._column_letter_to_index(old_col)
                    new_col = column_mapping.get((sheet_key, idx))
                except:
                    pass

            if new_col:
                # Preserve $ markers if present
                prefix = '$' if match.group(0).startswith('$') else ''
                suffix = '$' if '$' + row in match.group(0) else ''
                return f"{prefix}{new_col}{suffix}{row}"
            else:
                # No mapping found, keep original
                return match.group(0)

        result = same_sheet_pattern.sub(replace_column, result)

        # Also handle references without row numbers (e.g., L, CJ).
        # These can appear after earlier normalization or manual edits.
        no_row_pattern = re.compile(
            r'(?<![A-Za-z_"\'])(\$?)([A-Z]{1,3})(?![A-Za-z0-9_\(\["\'])'
        )

        def replace_column_no_row(match):
            prefix = match.group(1)
            old_col = match.group(2).upper()
            new_col = column_mapping.get((sheet_key, old_col))
            if not new_col:
                try:
                    idx = self._column_letter_to_index(old_col)
                    new_col = column_mapping.get((sheet_key, idx))
                except Exception:
                    new_col = None
            return f"{prefix}{new_col}" if new_col else match.group(0)

        result = no_row_pattern.sub(replace_column_no_row, result)

        # Restore the masked cross-sheet references
        for placeholder, original in placeholders.items():
            result = result.replace(placeholder, original)

        return result

    def _get_red_header_columns(self, formula_sheet, header_row, headers):
        """Return column letters whose header cells have a red background fill."""
        from openpyxl.utils import column_index_from_string

        red_cols = set()
        for header in headers:
            col_letter = header.get('column_letter')
            if not col_letter:
                continue
            col_idx = column_index_from_string(col_letter)
            cell = formula_sheet.cell(row=header_row, column=col_idx)
            if self._is_red_fill(cell):
                red_cols.add(col_letter)
        return red_cols

    def _get_red_data_columns(self, formula_sheet, data_start_row, headers, sample_rows=3):
        """Return column letters whose data cells have a red background fill."""
        from openpyxl.utils import column_index_from_string

        red_cols = set()
        max_row = formula_sheet.max_row or data_start_row
        for header in headers:
            col_letter = header.get('column_letter')
            if not col_letter:
                continue
            col_idx = column_index_from_string(col_letter)
            checked = 0
            red_count = 0
            for row_offset in range(sample_rows):
                row_num = data_start_row + row_offset
                if row_num > max_row:
                    break
                cell = formula_sheet.cell(row=row_num, column=col_idx)
                if cell.value is None and not getattr(cell.fill, 'patternType', None):
                    continue
                checked += 1
                if self._is_red_fill(cell):
                    red_count += 1
            if checked and (red_count / checked) >= 0.6:
                red_cols.add(col_letter)
        return red_cols

    @staticmethod
    def _is_red_fill(cell):
        """Check if a cell has a red background fill."""
        fill = getattr(cell, 'fill', None)
        if not fill:
            return False

        pattern = getattr(fill, 'patternType', None) or getattr(fill, 'fill_type', None)
        if pattern in (None, 'none'):
            return False

        color = getattr(fill, 'fgColor', None) or getattr(fill, 'start_color', None)
        if not color:
            return False

        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb).upper()
            if len(rgb) >= 6:
                rgb = rgb[-6:]
                try:
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                except ValueError:
                    return False
                return r >= 150 and g <= 100
        if color.type == 'indexed':
            return color.indexed in [2, 10]
        return False

    def _detect_formula_columns(self, formula_sheet, data_start_row, headers, skip_columns=None):
        """
        Detect which columns contain formulas.

        A column is considered a formula column if the first data row has
        a formula (starts with =).

        Args:
            formula_sheet: openpyxl sheet loaded with formulas (data_only=False)
            data_start_row: Row where data starts
            headers: List of header info dicts with column_letter

        Returns:
            Dictionary mapping column letter to formula info
        """
        from openpyxl.utils import column_index_from_string

        formula_columns = {}
        skip_columns = {c.upper() for c in (skip_columns or set())}

        # Check first few data rows for formulas
        for row_offset in range(min(5, formula_sheet.max_row - data_start_row + 1)):
            check_row = data_start_row + row_offset

            for header in headers:
                col_letter = header['column_letter']
                if col_letter.upper() in skip_columns:
                    continue
                if col_letter in formula_columns:
                    continue  # Already identified as formula

                col_idx = column_index_from_string(col_letter)
                cell = formula_sheet.cell(row=check_row, column=col_idx)

                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_columns[col_letter] = {
                        'formula': cell.value,
                        'detected_at_row': check_row,
                    }

        return formula_columns

    def _get_sample_value(self, sheet_data, data_sheet, formula_sheet, header_value, col_letter):
        """Get a sample value for a column, falling back to formulas when needed."""
        if sheet_data.get('data_rows'):
            for row in sheet_data['data_rows']:
                if header_value in row and row[header_value] is not None:
                    return str(row[header_value])[:50]

        data_start_row = sheet_data.get('data_start_row') or (sheet_data.get('header_row', 1) + 1)
        max_row = data_sheet.max_row if data_sheet else data_start_row
        scan_end = min(data_start_row + 10, max_row)

        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)

        for row_num in range(data_start_row, scan_end + 1):
            value = None
            if data_sheet:
                value = data_sheet.cell(row=row_num, column=col_idx).value
            if value is not None:
                return str(value)[:50]
            if formula_sheet:
                formula_value = formula_sheet.cell(row=row_num, column=col_idx).value
                if isinstance(formula_value, str) and formula_value.startswith('='):
                    return formula_value[:50]

        return ''

    def _column_uniform_value(self, sheet_data, data_sheet, header_value, col_letter):
        """The column's uniform numeric value as a string, or False.

        A uniform value column is a parameter in practice — W41 exports
        constants exactly this way — so import seeds ``default_value`` from it
        and the round-tripped config recomputes faithfully (WP-L review M2:
        constants degraded to input 0 broke 28 letters on recompute). A
        VARYING column must stay at default 0: seeding a first-row value would
        silently pay it to any employee missing the input in a future payrun.

        Exactly ONE leading non-numeric cell is tolerated: the W41 two-header
        layout puts the CODE row where the importer sees the first data row,
        so the column reads ['DAYSTD', 26, 26, …] — the code string must not
        defeat uniformity (and is why the returned VALUE, not sample_value, is
        what seeds default_value). Any other non-numeric cell ⇒ not uniform."""
        values = []
        rows = sheet_data.get('data_rows') or []
        if rows:
            for row in rows:
                v = row.get(header_value)
                if v is not None and str(v).strip() != '':
                    values.append(v)
        elif data_sheet is not None:
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter)
            data_start_row = sheet_data.get('data_start_row') or (
                sheet_data.get('header_row', 1) + 1)
            scan_end = min(data_start_row + 500, data_sheet.max_row)
            for row_num in range(data_start_row, scan_end + 1):
                v = data_sheet.cell(row=row_num, column=col_idx).value
                if v is not None and str(v).strip() != '':
                    values.append(v)
        if not values:
            return False
        nums = []
        text_positions = []
        for i, v in enumerate(values):
            n = excel_semantics.coerce_number(str(v))
            if n is None:
                text_positions.append(i)
            else:
                nums.append(round(n, 6))
        if not nums:
            return False
        if text_positions and text_positions != [0]:
            return False
        if any(n != nums[0] for n in nums[1:]):
            return False
        return repr(nums[0])

    def _detect_empty_columns_between_valid(self, headers, sheet, header_row):
        """
        Detect empty columns between valid columns.

        Args:
            headers: List of header dictionaries with 'column_letter' and 'value' keys
            sheet: openpyxl worksheet object
            header_row: The detected header row number

        Returns:
            List of empty column letters between valid columns, or empty list if none found
        """
        from openpyxl.utils import get_column_letter, column_index_from_string

        if not headers:
            return []

        # Get all column indices that have valid headers
        valid_col_indices = []
        for header in headers:
            col_letter = header.get('column_letter')
            if col_letter and header.get('value'):
                try:
                    col_idx = column_index_from_string(col_letter)
                    valid_col_indices.append(col_idx)
                except Exception:
                    pass

        if len(valid_col_indices) < 2:
            return []

        # Find the range of columns with data
        min_col = min(valid_col_indices)
        max_col = max(valid_col_indices)

        # Check each column between min and max for empty columns
        empty_columns = []
        valid_col_set = set(valid_col_indices)

        for col_idx in range(min_col, max_col + 1):
            if col_idx not in valid_col_set:
                # This column is not in our headers - check if it's truly empty
                col_letter = get_column_letter(col_idx)

                # Check header row
                header_cell = sheet.cell(row=header_row, column=col_idx)
                header_value = header_cell.value

                # Check a few data rows too (to be thorough)
                has_data = False
                for row_offset in range(5):  # Check first 5 data rows
                    data_row = header_row + 1 + row_offset
                    if data_row > (sheet.max_row or 1):
                        break
                    data_cell = sheet.cell(row=data_row, column=col_idx)
                    if data_cell.value is not None:
                        has_data = True
                        break

                # If both header and data are empty, it's an empty column
                if (header_value is None or (isinstance(header_value, str) and not header_value.strip())) and not has_data:
                    empty_columns.append(col_letter)

        return empty_columns

    def _parse_percentage_value(self, value):
        """
        Parse a percentage value and convert to decimal.

        Examples:
            "8%" -> 0.08
            "1.5%" -> 0.015
            "100%" -> 1.0
            0.08 (already decimal) -> 0.08
            8 (just number) -> 8.0

        Returns:
            Tuple of (decimal_value, was_percentage)
        """
        if value is None:
            return 0.0, False

        # If it's already a number (float/int)
        if isinstance(value, (int, float)):
            return float(value), False

        # If it's a string, check for percentage sign
        str_value = str(value).strip()
        if str_value.endswith('%'):
            try:
                # Remove % and convert to decimal
                num_part = str_value[:-1].strip()
                decimal_value = float(num_part) / 100.0
                return decimal_value, True
            except ValueError:
                return 0.0, False

        # Try to convert as regular number
        try:
            return float(str_value), False
        except ValueError:
            return 0.0, False

    def _evaluate_constant_formula(self, formula):
        if not formula or not isinstance(formula, str):
            return None
        expr = formula.strip()
        if not expr.startswith('='):
            return None
        expr = expr[1:]
        import re
        if re.search(r'[A-Za-z_]', expr):
            return None
        expr = re.sub(r'(\d+(?:\.\d+)?)\s*%', r'(\1/100)', expr)
        if re.search(r'[^0-9\.\+\-\*\/\^\(\)\s]', expr):
            return None
        expr = expr.replace('^', '**')
        try:
            return eval(expr, {"__builtins__": {}}, {})
        except Exception:
            return None

    def _get_constant_cell_value(self, formula_sheet, data_sheet, row_num, col_idx):
        cell = formula_sheet.cell(row=row_num, column=col_idx)
        value = cell.value
        if isinstance(value, str) and value.startswith('='):
            if data_sheet:
                data_value = data_sheet.cell(row=row_num, column=col_idx).value
                if data_value is not None:
                    return data_value
            evaluated = self._evaluate_constant_formula(value)
            if evaluated is not None:
                return evaluated
        return value

    def _get_fill_rgb(self, cell):
        fill = cell.fill
        if not fill or not fill.patternType or fill.patternType == 'none':
            return None

        color = fill.fgColor or fill.start_color
        if not color:
            return None

        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb).upper()
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) == 6:
                return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)

        if color.type == 'indexed':
            try:
                from openpyxl.styles.colors import COLOR_INDEX
                idx = color.indexed
                if idx is not None and idx < len(COLOR_INDEX):
                    rgb = str(COLOR_INDEX[idx]).upper()
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if len(rgb) == 6:
                        return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            except Exception:
                return None

        return None

    def _is_yellow_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return r > 200 and g > 200 and b < 150

    def _is_amber_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return r > 200 and 120 < g < 210 and b < 120

    def _is_green_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return g > 150 and g > r and g > b and r < 200 and b < 200

    def _is_blue_font(self, cell):
        try:
            font = cell.font
            if not font or not font.color:
                return False

            color = font.color

            if color.type == 'rgb' and color.rgb:
                rgb = str(color.rgb).upper()
                if len(rgb) >= 6:
                    if len(rgb) == 8:
                        r = int(rgb[2:4], 16)
                        g = int(rgb[4:6], 16)
                        b = int(rgb[6:8], 16)
                    else:
                        r = int(rgb[0:2], 16)
                        g = int(rgb[2:4], 16)
                        b = int(rgb[4:6], 16)
                    if b > r and b > g and b > 80:
                        return True

            if color.type == 'indexed':
                if color.indexed in [
                    4, 5, 12, 23, 30, 32, 39, 40, 41, 42, 48, 49, 54, 55, 56
                ]:
                    return True
                value = cell.value
                if color.indexed not in (0, 1) and isinstance(value, (int, float)):
                    return True
                if color.indexed not in (0, 1) and isinstance(value, str) and value.strip().endswith('%'):
                    return True

            if color.type == 'theme':
                if color.theme in [4, 5, 8]:
                    return True
                value = cell.value
                if isinstance(value, (int, float)):
                    return True
                if isinstance(value, str) and value.strip().endswith('%'):
                    return True

        except Exception:
            return False

        return False

    def _get_color_coded_sheet_data(self, sheet, formula_sheet):
        from openpyxl.utils import get_column_letter
        from ..formula_engine.merged_cell_parser import MergedCellParser

        max_row = formula_sheet.max_row or 1
        max_col = formula_sheet.max_column or 1
        max_scan_row = min(max_row, 200)

        yellow_rows = {}
        amber_rows = {}
        green_rows = {}
        filled_rows = {}

        for row_num in range(1, max_scan_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = formula_sheet.cell(row=row_num, column=col_idx)
                if cell.fill and cell.fill.patternType not in (None, 'none'):
                    filled_rows[row_num] = filled_rows.get(row_num, 0) + 1
                if self._is_yellow_fill(cell):
                    yellow_rows[row_num] = yellow_rows.get(row_num, 0) + 1
                if self._is_amber_fill(cell):
                    amber_rows[row_num] = amber_rows.get(row_num, 0) + 1
                if self._is_green_fill(cell):
                    green_rows[row_num] = green_rows.get(row_num, 0) + 1

        colored_rows = set(yellow_rows.keys()) | set(amber_rows.keys())
        if not colored_rows:
            raise UserError(_("Could not detect header rows from color-coded Excel file."))

        header_block_start = min(colored_rows)
        header_block_end = max(colored_rows)
        identifier_row = header_block_start - 1

        formula_row = None
        green_candidates = {r: count for r, count in green_rows.items() if r > header_block_end}
        if green_candidates:
            max_count = max(green_candidates.values())
            formula_row = min(r for r, count in green_candidates.items() if count == max_count)
        else:
            filled_candidates = {
                r: count for r, count in filled_rows.items() if r > header_block_end
            }
            if filled_candidates:
                max_count = max(filled_candidates.values())
                formula_row = min(r for r, count in filled_candidates.items() if count == max_count)
            else:
                scan_limit = min(max_row, header_block_end + 50)
                for row_num in range(header_block_end + 1, scan_limit + 1):
                    for col_idx in range(1, max_col + 1):
                        if self._is_green_fill(formula_sheet.cell(row=row_num, column=col_idx)):
                            formula_row = row_num
                            break
                    if formula_row:
                        break

        if not formula_row:
            formula_row = header_block_end + 1

        merge_parser = MergedCellParser(formula_sheet)

        def get_merge_origin_cell(row_num, col_idx):
            merge_info = merge_parser.get_merge_at(row_num, col_idx)
            if merge_info:
                return formula_sheet.cell(
                    row=merge_info['min_row'],
                    column=merge_info['min_col'],
                )
            return formula_sheet.cell(row=row_num, column=col_idx)

        def get_cell_value(row_num, col_idx):
            cell = get_merge_origin_cell(row_num, col_idx)
            if cell.value is not None:
                return cell.value
            return None

        def is_bold_in_merge(row_num, col_idx):
            merge_info = merge_parser.get_merge_at(row_num, col_idx)
            if merge_info:
                for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                    for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                        cell = formula_sheet.cell(row=m_row, column=m_col)
                        if cell.font and cell.font.bold:
                            return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return bool(cell.font and cell.font.bold)

        def is_red_font(cell):
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                            else:
                                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if r > 150 and g < 150 and b < 150:
                                return True
                            if r > 200 and g < 180 and b < 180 and r > g and r > b:
                                return True
                    elif color.type == 'indexed' and color.indexed in [2, 10]:
                        return True
            except Exception:
                pass
            return False

        def is_red_in_merge(row_num, col_idx):
            merge_info = merge_parser.get_merge_at(row_num, col_idx)
            if merge_info:
                for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                    for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                        cell = formula_sheet.cell(row=m_row, column=m_col)
                        if is_red_font(cell):
                            return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return is_red_font(cell)

        def is_underline_in_merge(row_num, col_idx):
            merge_info = merge_parser.get_merge_at(row_num, col_idx)
            if merge_info:
                for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                    for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                        cell = formula_sheet.cell(row=m_row, column=m_col)
                        if cell.font and cell.font.underline:
                            return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return bool(cell.font and cell.font.underline)

        headers = []
        header_map = {}
        header_rows = {}
        report_visible_map = {}
        contract_component_map = {}
        requires_new_contract_map = {}

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header_value = None
            header_row = None

            for row_num in range(header_block_end, header_block_start - 1, -1):
                cell = get_merge_origin_cell(row_num, col_idx)
                if cell.value is None:
                    continue
                if self._is_yellow_fill(cell):
                    header_value = cell.value
                    header_row = row_num
                    break

            if header_value is None:
                for row_num in range(header_block_start, header_block_end + 1):
                    cell = get_merge_origin_cell(row_num, col_idx)
                    if cell.value is None:
                        continue
                    header_value = cell.value
                    header_row = row_num
                    break

            if header_value is None:
                continue

            value_str = str(header_value).strip()
            if not value_str:
                continue

            headers.append({
                'column_letter': col_letter,
                'column_index': col_idx - 1,
                'value': value_str,
                'header_row': header_row,
            })
            header_map[col_letter] = value_str
            header_rows[col_letter] = header_row
            report_visible_map[col_letter] = is_bold_in_merge(header_row, col_idx)
            contract_component_map[col_letter] = is_red_in_merge(header_row, col_idx)
            requires_new_contract_map[col_letter] = (
                contract_component_map[col_letter] and is_underline_in_merge(header_row, col_idx)
            )

        if not headers:
            raise UserError(_("No component names found in color-coded header block."))

        component_types = {}
        for header in headers:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            header_row = header_rows.get(col_letter, header_block_end)
            comp_value = None
            for row_num in range(header_row - 1, header_block_start - 1, -1):
                comp_cell = get_merge_origin_cell(row_num, col_idx)
                if not self._is_amber_fill(comp_cell):
                    continue
                comp_value = get_cell_value(row_num, col_idx)
                if comp_value:
                    break
            if comp_value:
                component_types[col_letter] = str(comp_value).strip()

        identifier_map = {}
        if identifier_row >= 1:
            for header in headers:
                col_letter = header['column_letter']
                col_idx = header['column_index'] + 1
                identifier_value = get_cell_value(identifier_row, col_idx)
                if identifier_value:
                    identifier_map[col_letter] = str(identifier_value).strip()

        formula_columns = {}
        for header in headers:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            cell = formula_sheet.cell(row=formula_row, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_columns[col_letter] = {
                    'formula': cell.value,
                    'detected_at_row': formula_row,
                }

        sheet_data = {
            'headers': [
                {
                    'column_letter': h['column_letter'],
                    'column_index': h['column_index'],
                    'value': h['value'],
                    'component_type': component_types.get(h['column_letter']),
                    'report_visible': report_visible_map.get(h['column_letter'], False),
                    'is_contract_component': contract_component_map.get(h['column_letter'], False),
                    'requires_new_contract': requires_new_contract_map.get(h['column_letter'], False),
                    'payslip_identifier': identifier_map.get(h['column_letter']),
                }
                for h in headers
            ],
            'header_row': header_block_end,
            'data_start_row': formula_row,
            'data_rows': [],
        }

        return {
            'sheet_data': sheet_data,
            'formula_columns': formula_columns,
            'header_map': header_map,
            'header_rows': header_rows,
            'identifier_map': identifier_map,
            'report_visible_map': report_visible_map,
            'contract_component_map': contract_component_map,
            'requires_new_contract_map': requires_new_contract_map,
            'header_block_start': header_block_start,
            'header_block_end': header_block_end,
            'formula_row': formula_row,
        }

    def _detect_blue_constant_cells_color_coded(self, sheet, header_row, formula_row, header_map, data_sheet=None):
        """
        Detect blue font constants below header block up to formula row.

        Names are generated as "Constant <Header Name>" using the header row.
        """
        from openpyxl.utils import get_column_letter

        blue_constants = []
        seen_cells = set()
        start_row = 1
        end_row = max(start_row, (formula_row - 1) if formula_row else start_row)

        for row_num in range(start_row, end_row + 1):
            for col_idx in range(1, (sheet.max_column or 1) + 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                if cell.value is None:
                    continue

                if not self._is_blue_font(cell):
                    continue

                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_num}"
                if cell_ref in seen_cells:
                    continue
                seen_cells.add(cell_ref)

                value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)
                decimal_value, was_percentage = self._parse_percentage_value(value)
                header_label = header_map.get(col_letter)
                name = f"Constant {header_label}" if header_label else f"Constant {col_letter}"

                blue_constants.append({
                    'name': name,
                    'value': decimal_value,
                    'original_value': cell.value,
                    'original_cell': cell_ref,
                    'original_col_letter': col_letter,
                    'original_col_idx': col_idx,
                    'row': row_num,
                    'was_percentage': was_percentage,
                })

        _logger.info(
            "Multi-sheet color import: blue constants detected=%s sample=%s",
            len(blue_constants),
            [(c['original_cell'], c['name']) for c in blue_constants[:10]],
        )
        return blue_constants

    def _collect_constants_for_sheet(self, formula_sheet, sheet_data, formula_columns, data_sheet=None):
        """Collect constant definitions from a sheet using color and formula scans."""
        header_row = sheet_data.get('header_row', 1)
        data_start_row = sheet_data.get('data_start_row', header_row + 1)

        constant_pairs = self._detect_colored_constant_pairs(formula_sheet, header_row, data_sheet)
        # D-E4: scan ONLY the rows ABOVE the first data row. The old
        # `data_start_row + 2` reached two rows INTO the employee data, so a
        # blue-styled INPUT column would freeze employee #1's value as a
        # workbook-wide constant applied to everyone. `_detect_blue_constant_cells`
        # scans `range(1, scan_up_to_row)`, so passing `data_start_row` stops
        # exactly before the data.
        scan_up_to_row = data_start_row
        blue_constants = self._detect_blue_constant_cells(formula_sheet, scan_up_to_row, data_sheet)
        formula_referenced = self._detect_formula_referenced_constants(
            formula_columns, formula_sheet, header_row, data_sheet
        )

        constants = []
        seen_cells = set()
        for const in constant_pairs + blue_constants + formula_referenced:
            cell_ref = const.get('original_cell')
            if not cell_ref or cell_ref in seen_cells:
                continue
            seen_cells.add(cell_ref)
            constants.append(const)

        return constants

    def _detect_blue_constant_cells(self, sheet, scan_up_to_row, data_sheet=None):
        """
        Detect cells with blue font color that contain constant values.

        Scans all rows up to scan_up_to_row for cells with blue font.
        Blue is detected when B > R and B > G in the RGB color.

        For cells that are percentages (e.g., "8%"), converts to decimal (0.08).

        Returns list of dictionaries with:
            - name: Generated name (ConstantA, ConstantB, etc.)
            - value: The numeric value (percentages converted to decimals)
            - original_cell: Cell reference like "AY11"
            - original_col_letter: Column letter
            - original_col_idx: Column index
            - row: Row number
            - was_percentage: Boolean if value was originally a percentage
        """
        from openpyxl.utils import get_column_letter

        _logger.info(f"=== BLUE CONSTANT DETECTION START ===")
        _logger.info(f"Scanning rows 1 to {scan_up_to_row - 1}")

        blue_constants = []
        constant_index = 0  # For naming ConstantA, ConstantB, etc.
        cells_with_color_checked = 0

        def is_blue_color(cell, cell_ref=""):
            """Check if cell has blue FONT color (any shade where B > R and B > G)."""
            try:
                font = cell.font
                if not font or not font.color:
                    return False, None

                color = font.color
                color_info = f"type={color.type}"

                # Check RGB color
                if color.type == 'rgb' and color.rgb:
                    rgb = str(color.rgb).upper()
                    color_info += f", rgb={rgb}"
                    if len(rgb) >= 6:
                        # Handle ARGB (8 chars) or RGB (6 chars)
                        if len(rgb) == 8:
                            r = int(rgb[2:4], 16)
                            g = int(rgb[4:6], 16)
                            b = int(rgb[6:8], 16)
                        else:
                            r = int(rgb[0:2], 16)
                            g = int(rgb[2:4], 16)
                            b = int(rgb[4:6], 16)

                        color_info += f" -> R={r}, G={g}, B={b}"

                        # Blue: B is dominant (B > R and B > G)
                        if b > r and b > g and b > 100:
                            return True, color_info

                # Check indexed colors (common blue indices)
                elif color.type == 'indexed':
                    color_info += f", indexed={color.indexed}"
                    if color.indexed in [4, 5, 12, 23, 30, 32, 39, 40, 41, 42, 48, 49, 54, 55, 56]:
                        return True, color_info

                # Check theme colors (theme 4, 5, 8 are often blue variants)
                elif color.type == 'theme':
                    color_info += f", theme={color.theme}"
                    if color.theme in [4, 5, 8]:
                        return True, color_info

                return False, color_info

            except Exception as e:
                return False, f"error: {e}"

        def generate_constant_name(index):
            """Generate ConstantA, ConstantB, ..., ConstantZ, ConstantAA, etc."""
            result = ""
            idx = index
            while True:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
                if idx < 0:
                    break
            return f"Constant{result}"

        def get_column_header_label(row_num, col_idx):
            """
            Look at the row immediately above to get a header label for this column.
            This appends context like _BHXH, _BHYT to the constant name.
            """
            if row_num <= 1:
                return None

            # Check the row immediately above
            header_cell = sheet.cell(row=row_num - 1, column=col_idx)
            if header_cell.value and isinstance(header_cell.value, str):
                label = str(header_cell.value).strip()
                # Skip formulas - they start with '='
                if label.startswith('='):
                    return None
                # Skip resolved formula patterns (values.get, VLOOKUP, SUMIF, etc.)
                if 'values.get' in label or 'VLOOKUP' in label.upper() or 'SUMIF' in label.upper():
                    return None
                # Clean up the label - remove special chars, keep alphanumeric and underscore
                import re
                clean_label = re.sub(r'[^A-Za-z0-9_]', '', label)
                # Reject labels that look like they contain unresolved formulas
                if len(clean_label) > 50:  # Unreasonably long labels are likely formula remnants
                    return None
                if clean_label:
                    return clean_label

            return None

        # Scan all rows up to scan_up_to_row
        for row_num in range(1, scan_up_to_row):
            for col_idx in range(1, (sheet.max_column or 1) + 1):
                cell = sheet.cell(row=row_num, column=col_idx)

                # Skip empty cells
                if cell.value is None:
                    continue

                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_num}"

                # Check if cell has font color
                is_blue, color_info = is_blue_color(cell, cell_ref)
                if color_info:
                    cells_with_color_checked += 1
                    # Log cells with any color info for debugging
                    if cells_with_color_checked <= 20:  # Limit logging
                        _logger.info(f"  Cell {cell_ref} value='{cell.value}' color: {color_info} -> is_blue={is_blue}")

                # Check if cell has blue font
                if is_blue:
                    # Parse the value (handles percentages)
                    value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)
                    decimal_value, was_percentage = self._parse_percentage_value(value)

                    # Generate base name (ConstantA, ConstantB, etc.)
                    base_name = generate_constant_name(constant_index)
                    constant_index += 1

                    # Look for header label in row above to append context
                    header_label = get_column_header_label(row_num, col_idx)
                    if header_label:
                        name = f"{base_name}_{header_label}"
                    else:
                        name = base_name

                    blue_constants.append({
                        'name': name,
                        'value': decimal_value,
                        'original_value': cell.value,  # Keep original for reference
                        'original_cell': cell_ref,
                        'original_col_letter': col_letter,
                        'original_col_idx': col_idx,
                        'row': row_num,
                        'was_percentage': was_percentage,
                    })

                    _logger.info(
                        f"  -> BLUE CONSTANT FOUND: {cell_ref} = {cell.value} "
                        f"-> {name} = {decimal_value}"
                        f"{' (converted from %)' if was_percentage else ''}"
                    )

        _logger.info(f"=== BLUE CONSTANT DETECTION END: {len(blue_constants)} found, {cells_with_color_checked} cells with color checked ===")

        return blue_constants

    def _detect_colored_constant_pairs(self, sheet, header_row, data_sheet=None):
        """
        Detect red/green colored cell pairs that represent constants.

        These are cells where:
        - Left cell has RED font color (component name/label)
        - Right cell has GREEN font color (constant value)

        Note: We check FONT color (text color), not background/fill color,
        as users typically format these constant cells with colored text.

        Args:
            sheet: openpyxl worksheet object
            header_row: The detected header row number

        Returns:
            List of dictionaries with:
            - 'name': Component name from red cell
            - 'value': Constant value from green cell
            - 'original_cell': Cell reference (e.g., 'CE3') for formula replacement
            - 'row': Row number where found
        """
        from openpyxl.utils import get_column_letter

        constant_pairs = []

        def is_red_color(cell):
            """Check if cell has red FONT color (text color)."""
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r = int(rgb[2:4], 16)
                                g = int(rgb[4:6], 16)
                                b = int(rgb[6:8], 16)
                            else:
                                r = int(rgb[0:2], 16)
                                g = int(rgb[2:4], 16)
                                b = int(rgb[4:6], 16)
                            # Red if R is high and G, B are relatively low
                            if r > 150 and g < 150 and b < 150:
                                return True
                            # Also check for salmon/light red
                            if r > 200 and g < 180 and b < 180 and r > g and r > b:
                                return True
                    elif color.type == 'indexed':
                        if color.indexed in [2, 10]:
                            return True
            except Exception:
                pass
            return False

        def is_green_color(cell):
            """Check if cell has green FONT color (text color)."""
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r = int(rgb[2:4], 16)
                                g = int(rgb[4:6], 16)
                                b = int(rgb[6:8], 16)
                            else:
                                r = int(rgb[0:2], 16)
                                g = int(rgb[2:4], 16)
                                b = int(rgb[4:6], 16)
                            # Green if G is high and R, B are relatively low
                            if g > 150 and r < 150 and b < 150:
                                return True
                            # Also check for yellow-green
                            if g > 180 and r > 180 and b < 150:
                                return True
                            # Light green variations
                            if g > 200 and g > r and g > b:
                                return True
                    elif color.type == 'indexed':
                        if color.indexed in [3, 11]:
                            return True
                    elif color.type == 'theme':
                        # Theme colors: Excel uses theme indices for accent colors
                        # Theme 9 = Accent 6 (often green in many Excel themes)
                        # Theme 6 = Accent 3 (can also be green)
                        if color.theme in [6, 9]:
                            return True
            except Exception:
                pass
            return False

        max_row_to_search = min(header_row + 5, sheet.max_row or 1)

        for row_num in range(1, max_row_to_search + 1):
            if row_num >= header_row:
                continue

            for col_idx in range(1, (sheet.max_column or 1)):
                cell_left = sheet.cell(row=row_num, column=col_idx)
                cell_right = sheet.cell(row=row_num, column=col_idx + 1)

                if is_red_color(cell_left) and is_green_color(cell_right):
                    name = cell_left.value
                    value = self._get_constant_cell_value(
                        sheet,
                        data_sheet,
                        row_num,
                        col_idx + 1,
                    )

                    if name is not None and value is not None:
                        value_col_letter = get_column_letter(col_idx + 1)
                        original_cell_ref = f"{value_col_letter}{row_num}"

                        constant_pairs.append({
                            'name': str(name).strip(),
                            'value': value,
                            'original_cell': original_cell_ref,
                            'original_col_letter': value_col_letter,
                            'original_col_idx': col_idx + 1,
                            'row': row_num,
                        })

        return constant_pairs

    def _detect_formula_referenced_constants(self, formula_columns, sheet, header_row, data_sheet=None):
        """
        Scan formulas to find cell references above header row that might be constants.
        This catches constants like $CF$3 that may not have been detected by color.
        """
        import re
        from openpyxl.utils import column_index_from_string

        found_constants = []
        seen_cells = set()
        cell_ref_pattern = r'\$?([A-Z]+)\$?(\d+)'

        for col_letter, formula_info in formula_columns.items():
            formula = formula_info.get('formula', '') if isinstance(formula_info, dict) else formula_info
            matches = re.findall(cell_ref_pattern, formula, re.IGNORECASE)

            for col_match, row_match in matches:
                row_num = int(row_match)
                col_match = col_match.upper()

                if row_num >= header_row:
                    continue

                cell_ref = f"{col_match}{row_num}"
                if cell_ref in seen_cells:
                    continue
                seen_cells.add(cell_ref)

                try:
                    col_idx = column_index_from_string(col_match)
                    value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)

                    if value is None:
                        continue

                    name = None
                    if col_idx > 1:
                        left_cell = sheet.cell(row=row_num, column=col_idx - 1)
                        if left_cell.value and isinstance(left_cell.value, str):
                            name = str(left_cell.value).strip()

                    if not name:
                        name = f"Constant_{col_match}{row_num}"

                    found_constants.append({
                        'name': name,
                        'value': value,
                        'original_cell': cell_ref,
                        'original_col_letter': col_match,
                        'original_col_idx': col_idx,
                        'row': row_num,
                    })

                except Exception as e:
                    _logger.warning(f"Error processing cell {cell_ref}: {e}")

        return found_constants

    def _generate_extended_column_letter(self, index):
        """Generate column letters starting from ZA, ZB, ZC, etc."""
        base_col = 26 * 26 + 1 + index  # ZA starts at position 677

        result = ""
        col = base_col
        while col > 0:
            col -= 1
            result = chr(col % 26 + ord('A')) + result
            col //= 26

        return result

    def _extract_column_letter_from_cell_ref(self, cell_ref):
        """Extract the column letters from an Excel cell reference like "AB12"."""
        if not cell_ref:
            return ''
        match = re.match(r'^\$?([A-Za-z]+)', str(cell_ref).strip())
        return match.group(1).upper() if match else ''

    def _update_formula_references(self, formula, cell_mapping, sheet_name=None):
        """
        Update formula to replace original cell references with new column letters.

        Accepts either:
        - {'CE3': 'ZA'} -> replaces $CE$3 or CE3 with ZA2
        - {('sheet_key', 'CE3'): 'ZA'} -> replace only within that sheet context
        """
        import re

        if not formula or not cell_mapping:
            return formula

        sheet_key = self._normalize_sheet_key(sheet_name) if sheet_name else None
        mapping_by_sheet = {}
        global_mapping = {}

        for key, new_col in cell_mapping.items():
            if isinstance(key, tuple) and len(key) == 2:
                map_sheet, cell_ref = key
                map_sheet = self._normalize_sheet_key(map_sheet)
                mapping_by_sheet.setdefault(map_sheet, {})[str(cell_ref).upper()] = new_col
            else:
                global_mapping[str(key).upper()] = new_col

        pattern = re.compile(
            r"(?<![A-Za-z0-9_])"
            r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_\-]+)!)?"
            r"\$?(?P<col>[A-Za-z]+)\$?(?P<row>\d+)"
            r"(?![0-9A-Za-z_])"
        )

        def replace_match(match):
            sheet_token = match.group('sheet')
            col_letters = match.group('col').upper()
            row_num = match.group('row')
            cell_ref = f"{col_letters}{row_num}"

            target_map = None
            if sheet_token:
                sheet_label = sheet_token
                if sheet_label.startswith("'") and sheet_label.endswith("'"):
                    sheet_label = sheet_label[1:-1]
                sheet_label = self._normalize_sheet_key(sheet_label)
                target_map = mapping_by_sheet.get(sheet_label)
            elif sheet_key:
                target_map = mapping_by_sheet.get(sheet_key)

            new_col = None
            if target_map:
                new_col = target_map.get(cell_ref)
            if not new_col:
                new_col = global_mapping.get(cell_ref)

            return f"{new_col}2" if new_col else match.group(0)

        return pattern.sub(replace_match, formula)

    # ==========================================
    # STEP 3: REVIEW COMPONENTS
    # ==========================================
    def action_detect_missing_fields(self):
        """Detect fields in existing config that are missing from Excel."""
        self.ensure_one()

        # Get existing rules from config
        existing_rules = self.config_id.rule_ids.filtered(lambda r: r.column_type == 'input')

        # Get codes from Excel components
        excel_codes = set(self.component_preview_ids.mapped('generated_code'))

        # Clear existing missing fields
        self.missing_field_ids.unlink()

        # Find missing rules
        missing = []
        for rule in existing_rules:
            if rule.code not in excel_codes:
                missing.append({
                    'wizard_id': self.id,
                    'existing_rule_id': rule.id,
                    'field_code': rule.code,
                    'field_name': rule.name,
                    'data_source': 'none',
                })

        if missing:
            for m in missing:
                self.env['hr.formula.multisheet.missing.field'].create(m)
            self.state = 'map_missing'
        else:
            self.state = 'confirm'

        return self._return_wizard_action()

    # ==========================================
    # STEP 4: MAP MISSING FIELDS (Optional)
    # ==========================================
    def action_skip_missing(self):
        """Skip missing field mapping and proceed to confirm."""
        self.ensure_one()
        self.state = 'confirm'
        return self._return_wizard_action()

    def action_apply_missing_mapping(self):
        """Apply missing field mappings and proceed to confirm."""
        self.ensure_one()
        # Mappings are already stored in missing_field_ids
        self.state = 'confirm'
        return self._return_wizard_action()

    # ==========================================
    # STEP 5: CONFIRM AND EXECUTE
    # ==========================================
    def action_execute_import(self):
        """Execute the import and create formula rules."""
        self.ensure_one()

        try:
            components_to_import = self.component_preview_ids.filtered('include_in_import')

            if not components_to_import:
                raise UserError(_("No components selected for import."))

            created_rules = []
            updated_rules = []
            unparseable_constants = []   # D-E5: constants whose value wasn't numeric
            identifier_cache = {}
            payslip_config_model = self.env['hr.payslip.config']

            def get_payslip_identifier(identifier_value):
                if not identifier_value:
                    return False
                if identifier_value in identifier_cache:
                    return identifier_cache[identifier_value]
                record = payslip_config_model.search([
                    ('identifier', '=', identifier_value),
                    ('salary_structure_id', '=', self.config_id.id),
                ], limit=1)
                if not record:
                    record = payslip_config_model.create({
                        'salary_structure_id': self.config_id.id,
                        'identifier': identifier_value,
                        'label': '',
                    })
                identifier_cache[identifier_value] = record
                return record

            # Process each component
            for comp in components_to_import:
                existing_rule = self.config_id.rule_ids.filtered(
                    lambda r: r.code == comp.generated_code
                )

                sequence = None
                if comp.column_letter:
                    sequence = self._column_letter_to_index(comp.column_letter) + 1

                rule_vals = {
                    'config_id': self.config_id.id,
                    'name': comp.generated_name,
                    'code': comp.generated_code,
                    'column_type': comp.column_type,
                    'component_type': comp.component_type,
                    'source_sheet_name': comp.source_sheet,
                    'original_column_letter': comp.column_letter,
                    'forced_column_letter': comp.column_letter,  # Preserve actual Excel column position for ALL rules
                    'data_source': comp.data_source,
                    'report_visible': bool(comp.report_visible),
                    'is_contract_component': bool(comp.is_contract_component),
                    'requires_new_contract': bool(comp.requires_new_contract),
                }
                if sequence is not None:
                    rule_vals['sequence'] = sequence

                if comp.payslip_identifier_code:
                    payslip_identifier = get_payslip_identifier(comp.payslip_identifier_code)
                    rule_vals['payslip_identifier'] = payslip_identifier.id if payslip_identifier else False

                if comp.column_type == 'formula' and comp.excel_formula:
                    rule_vals['excel_formula'] = comp.excel_formula

                # Uniform value column → seed default_value so a round-tripped
                # config recomputes faithfully (WP-L review M2: W41-exported
                # constants re-imported as input 0.0 broke 28 letters). Varying
                # columns keep default 0 — seeding a first-row value would
                # silently pay it to any employee missing the input later. The
                # seed comes from uniform_value, never sample_value (which can
                # be the W41 code-header row read as a phantom first data row).
                if comp.column_type == 'input' and comp.uniform_value:
                    number = excel_semantics.coerce_number(comp.uniform_value)
                    if number is not None:
                        rule_vals['default_value'] = number

                # Handle constant components (from colored cell pairs).
                # D-E5: parse via the shared Excel coercion (handles "8%",
                # "1.234,50", thousands separators) instead of a bare float().
                # A value that is genuinely non-numeric (text/date) is NOT
                # silently turned into 0.0 — it is recorded and surfaced loudly
                # in the completion notice so the officer can fix it.
                if comp.column_type == 'constant':
                    if comp.sample_value:
                        number = excel_semantics.coerce_number(comp.sample_value)
                        if number is None:
                            rule_vals['constant_value'] = 0.0
                            unparseable_constants.append(
                                (comp.generated_code, comp.sample_value))
                            _logger.warning(
                                "Constant %s has non-numeric value %r — imported as 0.0; "
                                "needs manual entry.",
                                comp.generated_code, comp.sample_value)
                        else:
                            rule_vals['constant_value'] = number

                if comp.data_source == 'integration' and comp.integration_connector_id:
                    rule_vals['integration_connector_id'] = comp.integration_connector_id.id
                    rule_vals['source_field_mapping'] = comp.integration_field_name

                if existing_rule and self.update_existing:
                    # F7: overwriting an existing rule via import is a versioned
                    # 'import' event (fresh creates are version-0 / unversioned).
                    existing_rule.with_context(
                        formula_version_reason='import').write(rule_vals)
                    updated_rules.append(existing_rule)
                elif not existing_rule:
                    new_rule = self.env['hr.formula.rule'].create(rule_vals)
                    created_rules.append(new_rule)

            # Update missing fields with their data source settings
            for missing in self.missing_field_ids:
                if missing.existing_rule_id:
                    update_vals = {'data_source': missing.data_source}
                    if missing.data_source == 'integration' and missing.integration_connector_id:
                        update_vals['integration_connector_id'] = missing.integration_connector_id.id
                        update_vals['source_field_mapping'] = missing.integration_field_name

                    missing.existing_rule_id.write(update_vals)

            # Return success notification
            message = _(
                "Import completed successfully!\n"
                "Created: %d rules\n"
                "Updated: %d rules"
            ) % (len(created_rules), len(updated_rules))

            # D-E5: don't let non-numeric constants pass off as a clean import.
            if unparseable_constants:
                shown = ', '.join(
                    '%s (%s)' % (code, val) for code, val in unparseable_constants[:8])
                message += _(
                    "\n\n⚠ %d constant(s) had non-numeric values and were imported "
                    "as 0 — set them manually: %s"
                ) % (len(unparseable_constants), shown)

            next_action = self.config_id.get_formview_action()
            next_action['target'] = 'current'

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Complete'),
                    'message': message,
                    'type': 'warning' if unparseable_constants else 'success',
                    'sticky': bool(unparseable_constants),
                    'next': next_action,
                }
            }

        except Exception as e:
            _logger.exception("Import execution failed")
            raise UserError(_("Import failed: %s") % str(e))

    # ==========================================
    # NAVIGATION HELPERS
    # ==========================================
    def action_back(self):
        """Navigate to previous step."""
        self.ensure_one()
        state_order = ['upload', 'select_sheets', 'select_columns', 'configure_order',
                       'review_components', 'map_missing', 'confirm']
        current_idx = state_order.index(self.state)
        if current_idx > 0:
            # Skip map_missing if no missing fields
            new_idx = current_idx - 1
            if state_order[new_idx] == 'map_missing' and not self.has_missing_fields:
                new_idx -= 1
            self.state = state_order[max(0, new_idx)]
        return self._return_wizard_action()

    # ==========================================
    # PRIMARY KEY MATCHING LOGIC
    # ==========================================
    def _merge_sheets_by_primary_key(self, connector):
        """
        Merge data from multiple sheets using primary key matching.

        Args:
            connector: ExcelConnector instance with loaded workbook

        Returns:
            dict: Merged data with keys:
                - 'primary_keys': list of all unique primary key values (from main sheet)
                - 'merged_rows': dict mapping primary_key -> row_data
                - 'warnings': list of warning messages about missing/extra employees
                - 'missing_employees': dict mapping sheet_name -> list of missing primary keys
                - 'extra_employees': dict mapping sheet_name -> list of extra primary keys
        """
        import json

        selected_sheets = self.available_sheet_ids.filtered('is_selected')
        main_sheet = selected_sheets.filtered('is_main_sheet')

        if not main_sheet:
            raise UserError(_("No main sheet designated for primary key matching"))

        main_sheet = main_sheet[0]
        auxiliary_sheets = selected_sheets.filtered(lambda s: not s.is_main_sheet)

        # Load main sheet data
        main_data = connector.load_sheet_with_detection(main_sheet.sheet_name)
        main_primary_key_col = self._resolve_primary_key_column(main_sheet)

        if not main_primary_key_col:
            raise UserError(_("Main sheet '%s' must have a primary key column specified") % main_sheet.sheet_name)
        main_sheet.primary_key_column_name = main_primary_key_col

        # Build primary key lookup for main sheet
        main_pk_map = {}  # Maps primary_key_value -> full_row_data
        main_pk_list = []  # Ordered list of primary keys from main sheet

        for row in main_data.get('data_rows', []):
            pk_value = row.get(main_primary_key_col)
            if pk_value:
                # Convert to string for consistent matching
                pk_key = str(pk_value).strip()
                if pk_key:
                    main_pk_map[pk_key] = row.copy()
                    main_pk_list.append(pk_key)

        _logger.info(f"Main sheet '{main_sheet.sheet_name}' has {len(main_pk_list)} employees with primary key '{main_primary_key_col}'")

        # Process auxiliary sheets and merge data by primary key
        warnings = []
        missing_employees = {}
        extra_employees = {}

        for aux_sheet in auxiliary_sheets:
            sheet_name = aux_sheet.sheet_name
            aux_pk_col = self._resolve_primary_key_column(aux_sheet)

            if not aux_pk_col:
                raise UserError(_("Sheet '%s' must have a primary key column specified") % sheet_name)
            aux_sheet.primary_key_column_name = aux_pk_col

            # Load auxiliary sheet data
            aux_data = connector.load_sheet_with_detection(sheet_name)

            # Build primary key lookup for this auxiliary sheet
            aux_pk_map = {}  # Maps primary_key_value -> row_data
            for row in aux_data.get('data_rows', []):
                pk_value = row.get(aux_pk_col)
                if pk_value:
                    pk_key = str(pk_value).strip()
                    if pk_key:
                        aux_pk_map[pk_key] = row

            _logger.info(f"Auxiliary sheet '{sheet_name}' has {len(aux_pk_map)} employees with primary key '{aux_pk_col}'")

            # Find missing employees (in main but not in auxiliary)
            missing_in_aux = set(main_pk_list) - set(aux_pk_map.keys())
            if missing_in_aux:
                missing_employees[sheet_name] = list(missing_in_aux)
                _logger.warning(f"Sheet '{sheet_name}': {len(missing_in_aux)} employees from main sheet not found")
                warnings.append(
                    _("Sheet '%s': %d employees from main sheet not found. These will have blank values for columns from this sheet.")
                    % (sheet_name, len(missing_in_aux))
                )

            # Find extra employees (in auxiliary but not in main)
            extra_in_aux = set(aux_pk_map.keys()) - set(main_pk_list)
            if extra_in_aux:
                extra_employees[sheet_name] = list(extra_in_aux)
                _logger.warning(f"Sheet '{sheet_name}': {len(extra_in_aux)} employees not in main sheet will be skipped")
                warnings.append(
                    _("Sheet '%s': %d employees not found in main sheet will be skipped: %s")
                    % (sheet_name, len(extra_in_aux), ', '.join(list(extra_in_aux)[:5]) + ('...' if len(extra_in_aux) > 5 else ''))
                )

            # Merge auxiliary sheet data into main sheet data by primary key
            for pk_key, main_row in main_pk_map.items():
                aux_row = aux_pk_map.get(pk_key)
                if aux_row:
                    # Merge auxiliary columns into main row
                    # Prefix auxiliary column names with sheet name to avoid conflicts
                    for col_name, col_value in aux_row.items():
                        if col_name != aux_pk_col:  # Don't duplicate primary key column
                            prefixed_col_name = f"{sheet_name}|{col_name}"
                            main_row[prefixed_col_name] = col_value
                else:
                    # Employee not in auxiliary sheet - fill with blanks
                    for header in aux_data.get('headers', []):
                        col_name = header.get('value')
                        if col_name and col_name != aux_pk_col:
                            prefixed_col_name = f"{sheet_name}|{col_name}"
                            main_row[prefixed_col_name] = None  # Blank value

        return {
            'primary_keys': main_pk_list,
            'merged_rows': main_pk_map,
            'warnings': warnings,
            'missing_employees': missing_employees,
            'extra_employees': extra_employees,
            'main_sheet_name': main_sheet.sheet_name,
            'main_primary_key_column': main_primary_key_col,
        }

    def _return_wizard_action(self):
        """Return action to continue showing the wizard."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def _generate_code(self, header: str, existing_codes: set) -> str:
        """Generate a code from a header value that obeys the C5 converter
        contract: **underscore-free** and **not a substring** of (nor a
        superstring of) any other code — otherwise the Excel→Python converter
        mangles references to 0 (see docs/FORMULA_ENGINE_CONVENTIONS.md C5).

        The previous implementation emitted ``FORMULA_COL`` / ``COL_12`` and
        de-duplicated with ``_1`` suffixes — every one an underscore, and the
        suffixes (``BASIC`` / ``BASIC_1``) were mutual substrings. Both broke
        the converter silently.
        """
        header_str = str(header).strip()

        # Formula-looking headers (a stray formula cell used as a header)
        if (header_str.startswith('=')
                or 'values.get' in header_str
                or 'VLOOKUP' in header_str.upper()
                or 'SUMIF' in header_str.upper()):
            base_code = 'FORMULACOL'
        elif header_str.isdigit():
            base_code = 'COL' + header_str
        else:
            base_code = re.sub(r'[^A-Za-z0-9]', '', header_str).upper()
            if not base_code:
                base_code = 'UNNAMED'
            if base_code[0].isdigit():
                base_code = 'C' + base_code  # a code must not start with a digit
            if len(base_code) > 40:
                base_code = base_code[:40]

        return self._dedupe_code_c5(base_code, existing_codes)

    @staticmethod
    def _dedupe_code_c5(base, existing_codes):
        """Return a code derived from ``base`` that is safe under the C5
        converter contract.

        The HARD guarantee is **underscore-free and unique** (not equal to any
        existing code). Underscores are what actually break the Excel→Python
        converter; substring codes are matched greedily (maximal munch), so
        ``AMOUNT``/``AMOUNTX``/``SI``/``SIEMP`` all resolve correctly —
        empirically verified. So substring-avoidance is a *cosmetic preference*
        here, not a correctness requirement, and is applied only when a short
        letter suffix can achieve it (it is mathematically impossible when the
        base equals an existing code — every superstring contains it).

        Dedup suffixes are LETTERS so the result stays underscore- and
        digit-free. This always terminates.
        """
        import string

        def is_exact(cand):
            return cand in existing_codes

        def is_substring(cand):
            return any(cand != e and (cand in e or e in cand)
                       for e in existing_codes if e)

        if not is_exact(base) and not is_substring(base):
            return base

        # Candidate suffixes: '', A..Z, then AA..ZZ. Prefer a candidate that is
        # both unique AND non-substring; otherwise take the first merely-unique
        # one (guaranteed underscore-free).
        suffixes = [''] + list(string.ascii_uppercase) + [
            a + b for a in string.ascii_uppercase for b in string.ascii_uppercase]
        first_unique = None
        for suffix in suffixes:
            cand = base + suffix
            if is_exact(cand):
                continue
            if first_unique is None:
                first_unique = cand
            if not is_substring(cand):
                return cand
        if first_unique is not None:
            return first_unique

        # Pathological exhaustion — guaranteed-unique, still underscore-free.
        cand = base
        while cand in existing_codes:
            cand += 'X'
        return cand


class MultiSheetSheetLine(models.TransientModel):
    """Sheet line for multi-sheet import wizard."""
    _name = 'hr.formula.multisheet.sheet.line'
    _description = 'Multi-Sheet Import Sheet Line'
    _rec_name = 'sheet_name'
    _order = 'sequence, id'

    wizard_id = fields.Many2one(
        'hr.formula.multisheet.import.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True
    )

    sequence = fields.Integer(default=10)

    sheet_name = fields.Char(
        string='Sheet Name',
        readonly=True
    )

    sheet_name_html = fields.Html(
        string='Sheet Name',
        compute='_compute_sheet_name_html',
        sanitize=False
    )

    color = fields.Integer(
        string='Color Index',
        compute='_compute_color',
        store=True
    )

    is_selected = fields.Boolean(
        string='Include',
        default=True
    )

    is_main_sheet = fields.Boolean(
        string='Main Sheet',
        default=False
    )

    detected_header_row = fields.Integer(
        string='Header Row',
        default=1
    )

    column_count = fields.Integer(
        string='Columns',
        readonly=True
    )

    row_count = fields.Integer(
        string='Data Rows',
        readonly=True
    )

    has_formulas = fields.Boolean(
        string='Has Formulas',
        readonly=True
    )

    references_other_sheets = fields.Boolean(
        string='References Other Sheets',
        readonly=True,
        help="This worksheet contains formulas that reference cells in other worksheets (e.g., =TAM_UNG!A1)"
    )

    referenced_sheet_names = fields.Char(
        string='Depends On',
        readonly=True,
        help="List of worksheets that this sheet references in formulas"
    )

    referenced_sheet_names_html = fields.Html(
        string='Depends On',
        compute='_compute_referenced_sheet_names_html',
        sanitize=False
    )

    header_confidence = fields.Float(
        string='Detection Confidence',
        readonly=True
    )

    # ==========================================
    # PRIMARY KEY MATCHING
    # ==========================================
    available_column_names = fields.Text(
        string='Available Columns',
        help="JSON list of available column names in this sheet (used for primary key dropdown)"
    )

    available_columns_display = fields.Char(
        string='Available Columns',
        compute='_compute_available_columns_display',
        store=False,
        help="Display of available column names for easy selection"
    )

    primary_key_column_name = fields.Char(
        string='Primary Key Column',
        help="Column name to use as primary key for matching rows across worksheets. Type or copy from Available Columns."
    )

    @api.depends('available_column_names')
    def _compute_available_columns_display(self):
        """Convert JSON column list to readable comma-separated display."""
        for record in self:
            if not record.available_column_names:
                record.available_columns_display = ''
                continue

            try:
                import json
                columns = json.loads(record.available_column_names)
                # Show first 3 columns, then indicate there are more
                if len(columns) <= 3:
                    record.available_columns_display = ', '.join(columns)
                else:
                    record.available_columns_display = ', '.join(columns[:3]) + f', ... ({len(columns)} total)'
            except Exception as e:
                _logger.warning(f"Could not parse available_column_names: {e}")
                record.available_columns_display = ''

    @api.depends('sheet_name')
    def _compute_sheet_name_html(self):
        """Generate colored badge HTML for sheet name."""
        for record in self:
            if not record.sheet_name:
                record.sheet_name_html = ''
                continue

            # Generate unique color for this sheet
            bg_color = record._get_unique_color_for_sheet(record.sheet_name)
            # Calculate contrasting text color
            r = int(bg_color[1:3], 16)
            g = int(bg_color[3:5], 16)
            b = int(bg_color[5:7], 16)
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
            text_color = '#000000' if luminance > 0.5 else '#ffffff'

            record.sheet_name_html = (
                f'<span class="badge sheet-name-badge" '
                f'style="background-color: {bg_color}; color: {text_color}; '
                f'padding: 6px 12px; font-size: 12px; font-weight: 600; '
                f'border-radius: 4px; display: inline-block;">'
                f'{record.sheet_name}</span>'
            )

    @api.depends('sheet_name')
    def _compute_color(self):
        """Compute a consistent color index for each sheet based on its name."""
        for record in self:
            if record.sheet_name:
                # Use a better hash distribution to avoid repetitive colors
                # This ensures similar names get different colors
                import hashlib
                hash_bytes = hashlib.md5(record.sheet_name.encode()).digest()
                hash_val = int.from_bytes(hash_bytes[:4], 'big')
                # Map to color index (0-11) with better distribution
                record.color = hash_val % 12
            else:
                record.color = 0

    def _get_unique_color_for_sheet(self, sheet_name):
        """Generate a unique hex color for a sheet name."""
        import hashlib
        # Generate a hash of the sheet name
        hash_bytes = hashlib.md5(sheet_name.encode()).digest()
        # Use the hash to generate RGB values
        r = hash_bytes[0]
        g = hash_bytes[1]
        b = hash_bytes[2]
        # Adjust brightness to ensure readable colors (not too dark, not too light)
        # Keep values in the 50-200 range for good visibility
        r = 50 + (r % 150)
        g = 50 + (g % 150)
        b = 50 + (b % 150)
        return f'#{r:02x}{g:02x}{b:02x}'

    @api.depends('referenced_sheet_names', 'wizard_id.available_sheet_ids.sheet_name')
    def _compute_referenced_sheet_names_html(self):
        """Generate HTML badges with unique color coding for referenced sheet names."""
        for record in self:
            if not record.referenced_sheet_names:
                record.referenced_sheet_names_html = ''
                continue

            sheet_name_map = {}
            for line in record.wizard_id.available_sheet_ids:
                if not line.sheet_name:
                    continue
                sheet_name_map[line.sheet_name.strip().lower()] = line.sheet_name

            # Build HTML badges with inline colors
            badges = []
            sheet_names = [s.strip() for s in record.referenced_sheet_names.split(',') if s.strip()]
            for sheet_name in sheet_names:
                display_name = sheet_name_map.get(sheet_name.strip().lower(), sheet_name)
                # Generate unique color for this sheet
                bg_color = self._get_unique_color_for_sheet(display_name)
                # Calculate contrasting text color (black or white)
                # Convert hex to RGB to calculate luminance
                r = int(bg_color[1:3], 16)
                g = int(bg_color[3:5], 16)
                b = int(bg_color[5:7], 16)
                luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
                text_color = '#000000' if luminance > 0.5 else '#ffffff'

                badge_html = (
                    f'<span class="badge sheet-badge" '
                    f'style="background-color: {bg_color}; color: {text_color}; '
                    f'margin: 2px; padding: 4px 8px; font-size: 11px; '
                    f'font-weight: 600; border-radius: 4px; display: inline-block;">'
                    f'{display_name}</span>'
                )
                badges.append(badge_html)

            record.referenced_sheet_names_html = ' '.join(badges) if badges else ''

    @api.onchange('is_main_sheet')
    def _onchange_is_main_sheet(self):
        """Ensure only one sheet can be main."""
        if self.is_main_sheet:
            # Unset other sheets
            for line in self.wizard_id.available_sheet_ids:
                if line.id != self.id:
                    line.is_main_sheet = False
            # Main sheet must be selected
            self.is_selected = True


class MultiSheetComponentPreview(models.TransientModel):
    """Component preview line for multi-sheet import wizard."""
    _name = 'hr.formula.multisheet.component.preview'
    _description = 'Multi-Sheet Import Component Preview'
    _order = 'source_sheet, column_letter'

    wizard_id = fields.Many2one(
        'hr.formula.multisheet.import.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True
    )

    source_sheet = fields.Char(
        string='Source Sheet',
        readonly=True
    )

    column_letter = fields.Char(
        string='Column',
        readonly=True
    )

    original_header = fields.Char(
        string='Header',
        readonly=True
    )

    generated_code = fields.Char(
        string='Code'
    )

    generated_name = fields.Char(
        string='Name'
    )

    component_type = fields.Char(
        string='Component Type',
        help="Category from merged cell above header"
    )

    payslip_identifier_code = fields.Char(
        string='Payslip Identifier',
        help="Identifier code mapped from color-coded row"
    )

    report_visible = fields.Boolean(
        string='Visible in Reports'
    )

    is_contract_component = fields.Boolean(
        string='Contract Component'
    )

    requires_new_contract = fields.Boolean(
        string='Requires New Contract'
    )

    column_type = fields.Selection([
        ('input', 'Input'),
        ('formula', 'Formula'),
        ('constant', 'Constant')
    ], string='Type', default='input')

    excel_formula = fields.Char(
        string='Excel Formula'
    )

    resolved_formula = fields.Char(
        string='Resolved Formula',
        help="Formula with cross-sheet references resolved"
    )

    is_duplicate = fields.Boolean(
        string='Duplicate',
        readonly=True
    )

    include_in_import = fields.Boolean(
        string='Include',
        default=True
    )

    sample_value = fields.Char(
        string='Sample'
    )

    uniform_value = fields.Char(
        string='Uniform Value',
        help="Set when every numeric data cell in the source column holds the "
             "same value — a parameter in practice; this value (NOT "
             "sample_value, which can be the W41 code-header row) seeds "
             "default_value on import so round-tripped configs recompute "
             "faithfully."
    )

    is_in_excel = fields.Boolean(
        string='In Excel',
        default=True,
        readonly=True
    )

    data_source = fields.Selection([
        ('excel', 'Excel Import'),
        ('formula', 'Formula (Calculated)'),
        ('integration', 'Integration'),
        ('manual', 'Manual Entry'),
        ('none', 'Not Populated'),
    ], string='Data Source', default='excel')

    integration_connector_id = fields.Many2one(
        'hr.integration.connector',
        string='Integration',
        domain="[('connection_status', '=', 'connected')]"
    )

    integration_field_name = fields.Char(
        string='Integration Field'
    )


class MultiSheetMissingField(models.TransientModel):
    """Missing field line for data source mapping."""
    _name = 'hr.formula.multisheet.missing.field'
    _description = 'Multi-Sheet Import Missing Field'

    wizard_id = fields.Many2one(
        'hr.formula.multisheet.import.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True
    )

    existing_rule_id = fields.Many2one(
        'hr.formula.rule',
        string='Existing Rule',
        readonly=True
    )

    field_code = fields.Char(
        string='Code',
        readonly=True
    )

    field_name = fields.Char(
        string='Name',
        readonly=True
    )

    data_source = fields.Selection([
        ('none', 'Not Populated'),
        ('integration', 'From Integration'),
        ('manual', 'Manual Entry'),
    ], string='Data Source', default='none', required=True)

    integration_connector_id = fields.Many2one(
        'hr.integration.connector',
        string='Integration Connector',
        domain="[('connection_status', '=', 'connected')]"
    )

    integration_field_name = fields.Char(
        string='Integration Field'
    )

    @api.onchange('data_source')
    def _onchange_data_source(self):
        """Clear integration fields if not using integration."""
        if self.data_source != 'integration':
            self.integration_connector_id = False
            self.integration_field_name = False


class MultiSheetColumnSelection(models.TransientModel):
    """
    Column selection for multi-sheet import wizard.

    Stores per-worksheet column selections, allowing users to choose
    which columns to include from each selected worksheet.
    """
    _name = 'hr.formula.multisheet.column.selection'
    _description = 'Multi-Sheet Import Column Selection'
    _order = 'sheet_line_id, sequence'

    wizard_id = fields.Many2one(
        'hr.formula.multisheet.import.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True
    )

    sheet_line_id = fields.Many2one(
        'hr.formula.multisheet.sheet.line',
        string='Sheet',
        ondelete='cascade',
        required=True
    )
    sheet_name = fields.Char(
        related='sheet_line_id.sheet_name',
        string='Worksheet',
        readonly=True
    )

    # Column identity
    sequence = fields.Integer(default=10)

    column_letter = fields.Char(
        string='Column',
        readonly=True
    )

    column_index = fields.Integer(
        string='Index',
        readonly=True
    )

    original_header = fields.Char(
        string='Header',
        readonly=True
    )
    constant_cell_ref = fields.Char(
        string='Original Cell',
        readonly=True,
        help="Original Excel cell reference for detected constants"
    )

    # Component type from merged cells above
    component_type = fields.Char(
        string='Category',
        readonly=True,
        help="Component type extracted from merged cell above the header"
    )

    payslip_identifier_code = fields.Char(
        string='Payslip Identifier',
        readonly=True
    )

    report_visible = fields.Boolean(
        string='Visible in Reports',
        readonly=True
    )

    is_contract_component = fields.Boolean(
        string='Contract Component',
        readonly=True
    )

    requires_new_contract = fields.Boolean(
        string='Requires New Contract',
        readonly=True
    )

    # Selection controls
    is_selected = fields.Boolean(
        string='Include',
        default=True
    )

    # Preview data
    column_type = fields.Selection([
        ('input', 'Input'),
        ('formula', 'Formula'),
        ('constant', 'Constant')
    ], string='Type', readonly=True, default='input')

    sample_value = fields.Char(
        string='Sample',
        readonly=True
    )

    uniform_value = fields.Char(
        string='Uniform Value',
        readonly=True
    )

    # Cross-sheet reference info
    has_cross_sheet_ref = fields.Boolean(
        string='Cross-Sheet Ref',
        readonly=True,
        help="This column's formula references other worksheets"
    )

    cross_sheet_formula = fields.Char(
        string='Formula',
        readonly=True,
        help="The formula containing cross-sheet references"
    )

    is_referenced_by_main = fields.Boolean(
        string='Referenced by Main',
        readonly=True,
        help="This column is referenced by formulas in the main worksheet"
    )

    is_referenced_by_other_sheet = fields.Boolean(
        string='Referenced by Other',
        compute='_compute_cross_sheet_reference_info',
        readonly=True,
        help="This column is referenced by formulas in other worksheets"
    )

    refers_to_sheet_names = fields.Char(
        string='Refers To Sheets',
        compute='_compute_cross_sheet_reference_info',
        readonly=True,
        help="Worksheets referenced by this column's formula"
    )

    referenced_by_sheet_names = fields.Char(
        string='Referenced By Sheets',
        compute='_compute_cross_sheet_reference_info',
        readonly=True,
        help="Worksheets that reference this column"
    )

    @api.depends(
        'wizard_id.column_selection_ids.cross_sheet_formula',
        'wizard_id.column_selection_ids.sheet_name',
        'wizard_id.column_selection_ids.column_letter'
    )
    def _compute_cross_sheet_reference_info(self):
        for wizard in self.mapped('wizard_id'):
            lines = wizard.column_selection_ids
            references_by_component = {}
            referenced_by_component = {}
            sheet_name_map = {
                self._normalize_sheet_name(line.sheet_name): line.sheet_name
                for line in lines
                if line.sheet_name
            }

            for line in lines:
                if not line.cross_sheet_formula:
                    continue
                source_key = self._normalize_sheet_name(line.sheet_name)
                source_col = (line.column_letter or '').upper()
                for ref in self._extract_cross_sheet_references_from_formula(line.cross_sheet_formula):
                    target_key = self._normalize_sheet_name(ref['sheet_name'])
                    if not target_key or target_key == source_key:
                        continue
                    target_display = sheet_name_map.get(target_key, ref['sheet_name'])
                    references_by_component.setdefault(
                        (source_key, source_col), set()
                    ).add(target_display)
                    referenced_by_component.setdefault(
                        (target_key, ref['col_letter']), set()
                    ).add(line.sheet_name)

            for line in lines & self:
                key = (
                    self._normalize_sheet_name(line.sheet_name),
                    (line.column_letter or '').upper()
                )
                refers_to = references_by_component.get(key, set())
                referenced_by = referenced_by_component.get(key, set())
                line.refers_to_sheet_names = ", ".join(sorted(refers_to))
                line.referenced_by_sheet_names = ", ".join(sorted(referenced_by))
                line.is_referenced_by_other_sheet = bool(referenced_by)

    @staticmethod
    def _normalize_sheet_name(sheet_name):
        return sheet_name.strip().lower() if sheet_name else ''

    def _extract_cross_sheet_references_from_formula(self, formula):
        """Extract referenced sheet/column pairs from a formula."""
        if not formula:
            return []

        refs = set()

        def add_ref(sheet_name, col_letter):
            if not sheet_name or not col_letter:
                return
            refs.add((sheet_name.strip(), col_letter.upper()))

        vlookup_quoted_pattern = re.compile(
            r"VLOOKUP\s*\([^,]+,\s*'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d*:\$?[A-Z]+\$?\d*,\s*(\d+)",
            re.IGNORECASE
        )
        vlookup_unquoted_pattern = re.compile(
            r"VLOOKUP\s*\([^,]+,\s*([A-Za-z][A-Za-z0-9_]*)\s*!\s*\$?([A-Z]+)\$?\d*:\$?[A-Z]+\$?\d*,\s*(\d+)",
            re.IGNORECASE
        )
        sumif_quoted_pattern = re.compile(
            r"SUMIF\s*\(\s*'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*([^,]+),\s*'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*\s*\)",
            re.IGNORECASE
        )
        sumif_unquoted_pattern = re.compile(
            r"SUMIF\s*\(\s*([A-Za-z][A-Za-z0-9_]*)\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*,\s*([^,]+),\s*([A-Za-z][A-Za-z0-9_]*)\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*\s*\)",
            re.IGNORECASE
        )
        range_quoted_pattern = re.compile(
            r"'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*",
            re.IGNORECASE
        )
        range_unquoted_pattern = re.compile(
            r"(?<![A-Za-z0-9_])([A-Za-z][A-Za-z0-9_]*)\s*!\s*\$?([A-Z]+)\$?\d*:\$?([A-Z]+)\$?\d*",
            re.IGNORECASE
        )
        direct_quoted_pattern = re.compile(
            r"'([^']+)'\s*!\s*\$?([A-Z]+)\$?\d+",
            re.IGNORECASE
        )
        direct_unquoted_pattern = re.compile(
            r"(?<![A-Za-z0-9_])([A-Za-z][A-Za-z0-9_]*)\s*!\s*\$?([A-Z]+)\$?\d+",
            re.IGNORECASE
        )

        for match in vlookup_quoted_pattern.finditer(formula):
            sheet_name = match.group(1).strip()
            start_col = match.group(2).upper()
            col_index = int(match.group(3))
            start_idx = self._column_letter_to_index(start_col)
            target_idx = start_idx + col_index - 1
            target_col = self._index_to_column_letter(target_idx)
            add_ref(sheet_name, target_col)

        for match in vlookup_unquoted_pattern.finditer(formula):
            sheet_name = match.group(1).strip()
            start_col = match.group(2).upper()
            col_index = int(match.group(3))
            start_idx = self._column_letter_to_index(start_col)
            target_idx = start_idx + col_index - 1
            target_col = self._index_to_column_letter(target_idx)
            add_ref(sheet_name, target_col)

        for match in sumif_quoted_pattern.finditer(formula):
            criteria_sheet = match.group(1).strip()
            criteria_col = match.group(2).upper()
            add_ref(criteria_sheet, criteria_col)
            sum_sheet = match.group(5).strip()
            sum_col = match.group(6).upper()
            add_ref(sum_sheet, sum_col)

        for match in sumif_unquoted_pattern.finditer(formula):
            criteria_sheet = match.group(1).strip()
            criteria_col = match.group(2).upper()
            add_ref(criteria_sheet, criteria_col)
            sum_sheet = match.group(5).strip()
            sum_col = match.group(6).upper()
            add_ref(sum_sheet, sum_col)

        for match in range_quoted_pattern.finditer(formula):
            sheet_name = match.group(1).strip()
            start_col = match.group(2).upper()
            end_col = match.group(3).upper()
            add_ref(sheet_name, start_col)
            add_ref(sheet_name, end_col)

        for match in range_unquoted_pattern.finditer(formula):
            sheet_name = match.group(1).strip()
            start_col = match.group(2).upper()
            end_col = match.group(3).upper()
            add_ref(sheet_name, start_col)
            add_ref(sheet_name, end_col)

        for match in direct_quoted_pattern.finditer(formula):
            add_ref(match.group(1), match.group(2))

        for match in direct_unquoted_pattern.finditer(formula):
            add_ref(match.group(1), match.group(2))

        return [
            {'sheet_name': sheet_name, 'col_letter': col_letter}
            for sheet_name, col_letter in sorted(refs)
        ]

    @staticmethod
    def _column_letter_to_index(col_letter):
        """Convert column letter to 0-based index (A=0, B=1, ...)."""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _index_to_column_letter(index):
        """Convert 0-based index to column letter (0=A, 1=B, ...)."""
        result = ""
        idx = index
        while True:
            result = chr(ord('A') + (idx % 26)) + result
            idx = idx // 26 - 1
            if idx < 0:
                break
        return result



class MultiSheetAppendOrder(models.TransientModel):
    """
    Worksheet append order for multi-sheet import wizard.

    Controls the order in which auxiliary worksheets are appended
    to the main worksheet to form the final component structure.
    """
    _name = 'hr.formula.multisheet.append.order'
    _description = 'Multi-Sheet Import Append Order'
    _order = 'append_sequence'

    wizard_id = fields.Many2one(
        'hr.formula.multisheet.import.wizard',
        string='Wizard',
        ondelete='cascade',
        required=True
    )

    sheet_line_id = fields.Many2one(
        'hr.formula.multisheet.sheet.line',
        string='Sheet',
        ondelete='cascade',
        required=True
    )

    sheet_name = fields.Char(
        related='sheet_line_id.sheet_name',
        string='Worksheet',
        readonly=True
    )

    is_main_sheet = fields.Boolean(
        related='sheet_line_id.is_main_sheet',
        string='Main',
        readonly=True
    )

    # Append order (main sheet is always 0)
    append_sequence = fields.Integer(
        string='Order',
        default=10
    )

    # Statistics
    selected_column_count = fields.Integer(
        string='Selected Columns',
        compute='_compute_stats'
    )

    total_column_count = fields.Integer(
        string='Total Columns',
        compute='_compute_stats'
    )

    # Preview info
    start_column_letter = fields.Char(
        string='Start Col',
        compute='_compute_column_range'
    )

    end_column_letter = fields.Char(
        string='End Col',
        compute='_compute_column_range'
    )

    @api.depends('sheet_line_id', 'wizard_id.column_selection_ids')
    def _compute_stats(self):
        """Compute column selection statistics."""
        for rec in self:
            if rec.sheet_line_id and rec.wizard_id:
                sheet_cols = rec.wizard_id.column_selection_ids.filtered(
                    lambda c: c.sheet_line_id.id == rec.sheet_line_id.id and c.column_type != 'constant'
                )
                rec.total_column_count = len(sheet_cols)
                rec.selected_column_count = len(sheet_cols.filtered('is_selected'))
            else:
                rec.total_column_count = 0
                rec.selected_column_count = 0

    @api.depends('append_sequence', 'wizard_id.append_order_ids', 'wizard_id.column_selection_ids')
    def _compute_column_range(self):
        """Compute the column range this sheet will occupy after appending."""
        for rec in self:
            if not rec.wizard_id or not rec.sheet_line_id:
                rec.start_column_letter = ''
                rec.end_column_letter = ''
                continue

            # Calculate cumulative column count for sheets before this one
            ordered = rec.wizard_id.append_order_ids.sorted('append_sequence')
            start_idx = 0
            for order_rec in ordered:
                if order_rec.id == rec.id:
                    break
                start_idx += order_rec.selected_column_count

            end_idx = start_idx + rec.selected_column_count - 1

            # Convert to column letters
            rec.start_column_letter = rec._index_to_column_letter(start_idx) if rec.selected_column_count > 0 else ''
            rec.end_column_letter = rec._index_to_column_letter(end_idx) if rec.selected_column_count > 0 else ''

    def _index_to_column_letter(self, index):
        """Convert 0-based column index to Excel column letter (A, B, ..., Z, AA, AB, ...)."""
        result = ""
        idx = index
        while True:
            result = chr(ord('A') + (idx % 26)) + result
            idx = idx // 26 - 1
            if idx < 0:
                break
        return result
