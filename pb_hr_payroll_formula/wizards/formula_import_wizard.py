# -*- coding: utf-8 -*-
"""
Formula Import Wizard - Import formula configuration from various sources.
"""

import base64
import json
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

from ..models import column_role_classifier

_logger = logging.getLogger(__name__)


class FormulaImportWizard(models.TransientModel):
    _name = 'hr.formula.import.wizard'
    _description = 'Formula Configuration Import Wizard'

    config_id = fields.Many2one(
        'hr.formula.config',
        string='Target Configuration',
        required=True,
        default=lambda self: self.env.context.get('active_id'),
    )

    import_source = fields.Selection([
        ('salary_rules', 'From Existing Salary Rules'),
        ('structure', 'From Payroll Structure'),
        ('json', 'From JSON File'),
        ('excel', 'From Excel File'),
    ], string='Import Source', default='salary_rules', required=True)

    # For salary rules source
    salary_rule_ids = fields.Many2many(
        'hr.salary.rule',
        string='Salary Rules',
    )

    # For structure source
    structure_id = fields.Many2one(
        'hr.payroll.structure',
        string='Payroll Structure',
    )

    # For file sources
    import_file = fields.Binary('Import File')
    import_filename = fields.Char('Filename')

    # Options
    create_input_columns = fields.Boolean(
        'Create Input Columns',
        default=True,
        help="Create input columns for salary rules that use inputs",
    )
    preserve_existing = fields.Boolean(
        'Preserve Existing Rules',
        default=True,
        help="Keep existing rules in the configuration",
    )
    map_categories = fields.Boolean(
        'Map Categories',
        default=True,
        help="Preserve salary rule categories",
    )

    @api.onchange('import_source')
    def _onchange_import_source(self):
        """Clear fields when source changes."""
        self.salary_rule_ids = False
        self.structure_id = False
        self.import_file = False
        self.import_filename = False

    @api.onchange('structure_id')
    def _onchange_structure_id(self):
        """Load salary rules from selected structure."""
        if self.structure_id:
            self.salary_rule_ids = self.structure_id.rule_ids

    def action_import(self):
        """Execute import based on selected source."""
        self.ensure_one()

        if self.import_source == 'salary_rules':
            return self._import_from_salary_rules()
        elif self.import_source == 'structure':
            return self._import_from_structure()
        elif self.import_source == 'json':
            return self._import_from_json()
        elif self.import_source == 'excel':
            return self._import_from_excel()
        else:
            raise UserError(_("Invalid import source"))

    def _import_from_salary_rules(self):
        """Import from selected salary rules."""
        if not self.salary_rule_ids:
            raise UserError(_("Please select salary rules to import"))

        if not self.preserve_existing:
            self.config_id.rule_ids.unlink()
            # full rebuild: restart the letter namespace at A so the fresh set is
            # lettered in creation order (F111 — otherwise create() mints letters
            # past the old high-water mark and re-import misaligns references)
            self.config_id.col_letter_hwm = 0

        # Get existing sequence
        max_sequence = max(
            self.config_id.rule_ids.mapped('sequence') or [0]
        )

        created_rules = self.env['hr.formula.rule']

        for rule in self.salary_rule_ids.sorted('sequence'):
            # Check if rule already exists
            existing = self.config_id.rule_ids.filtered(
                lambda r: r.code == rule.code or r.salary_rule_id == rule
            )
            if existing:
                continue

            max_sequence += 10

            # Determine column type
            column_type = 'formula'
            excel_formula = ''
            constant_value = 0.0

            if rule.amount_select == 'fix':
                column_type = 'constant'
                constant_value = rule.amount_fix
            elif rule.amount_select == 'percentage':
                column_type = 'formula'
                # Create formula from percentage
                if rule.amount_percentage_base:
                    excel_formula = f"={rule.amount_percentage_base}*{rule.amount_percentage/100}"

            # Create formula rule
            values = {
                'config_id': self.config_id.id,
                'salary_rule_id': rule.id,
                'name': rule.name,
                'code': rule.code,
                'sequence': max_sequence,
                'column_type': column_type,
                'excel_formula': excel_formula,
                'constant_value': constant_value,
                'category_id': rule.category_id.id if self.map_categories else False,
                'appears_on_payslip': rule.appears_on_payslip,
            }

            created_rules |= self.env['hr.formula.rule'].create(values)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'message': _('%d rules imported successfully') % len(created_rules),
                'type': 'success',
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_window_close',
                },
            }
        }

    def _import_from_structure(self):
        """Import from payroll structure."""
        if not self.structure_id:
            raise UserError(_("Please select a payroll structure"))

        self.salary_rule_ids = self.structure_id.rule_ids
        return self._import_from_salary_rules()

    def _import_from_json(self):
        """Import configuration from JSON file."""
        if not self.import_file:
            raise UserError(_("Please upload a JSON file"))

        try:
            content = base64.b64decode(self.import_file).decode('utf-8')
            data = json.loads(content)
        except Exception as e:
            raise UserError(_("Invalid JSON file: %s") % str(e))

        if not self.preserve_existing:
            self.config_id.rule_ids.unlink()
            # full rebuild: restart the letter namespace at A (F111) so verbatim
            # excel_formula references (=A2+B2) re-align on the fresh rule set
            self.config_id.col_letter_hwm = 0

        # Import rules from JSON
        rules_data = data.get('rules', [])
        max_sequence = max(
            self.config_id.rule_ids.mapped('sequence') or [0]
        )

        created_rules = self.env['hr.formula.rule']

        for rule_data in rules_data:
            max_sequence += 10

            values = {
                'config_id': self.config_id.id,
                'name': rule_data.get('name', 'Imported Rule'),
                'code': rule_data.get('code', f'IMPORT_{max_sequence}'),
                'sequence': max_sequence,
                'column_type': rule_data.get('column_type', 'formula'),
                'excel_formula': rule_data.get('excel_formula', ''),
                'constant_value': rule_data.get('constant_value', 0.0),
                'default_value': rule_data.get('default_value', 0.0),
                'number_format': rule_data.get('number_format', 'currency'),
                'decimal_places': rule_data.get('decimal_places', 2),
                # honour an exported letter when present (authoritative); else
                # create() assigns A,B,C… in order against the reset namespace
                'column_letter': rule_data.get('column_letter') or False,
            }

            created_rules |= self.env['hr.formula.rule'].create(values)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'message': _('%d rules imported from JSON') % len(created_rules),
                'type': 'success',
                'sticky': False,
            }
        }

    # ------------------------------------------------------------------
    # COLROLES — role plumbing shared by both import paths
    # ------------------------------------------------------------------
    SAMPLE_LIMIT = 10
    SAMPLE_SCAN_ROWS = 40

    def _collect_column_samples(self, data_sheet, col_idx, first_data_row):
        """Up to SAMPLE_LIMIT non-empty values below the header, read from the
        data_only workbook. These are what tell an inferred contract component apart
        from an amount one, and what makes an all-text column a reference column."""
        samples = []
        if not data_sheet or first_data_row is None:
            return samples
        max_row = data_sheet.max_row or 0
        last_row = min(max_row, first_data_row + self.SAMPLE_SCAN_ROWS)
        for row_num in range(first_data_row, last_row + 1):
            try:
                value = data_sheet.cell(row=row_num, column=col_idx).value
            except Exception:
                break
            if column_role_classifier.is_blank_sample(value):
                continue
            samples.append(value)
            if len(samples) >= self.SAMPLE_LIMIT:
                break
        return samples

    def _role_rule_values(self, role):
        """Rule values implied by a role on a NEWLY created rule.

        Anything that is not a payroll column is not a pay line and not a grid column,
        so it starts hidden from both. This applies at import time only — existing
        rules keep whatever visibility somebody already chose for them."""
        return column_role_classifier.role_rule_defaults(role)

    def _import_from_excel(self):
        """
        Import configuration from Excel file using advanced header detection.

        This method uses the multisheet import logic which:
        - Auto-detects the header row (handles merged cells, category rows)
        - Extracts component types from horizontal merged cells above headers
        - Handles vertical merges (headers spanning multiple rows)
        - Detects formula vs input columns
        - Works for both single and multi-sheet files
        """
        if not self.import_file:
            raise UserError(_("Please upload an Excel file"))

        if self.config_id.use_color_coded_excel_import:
            return self._import_from_excel_color_coded()

        try:
            import openpyxl
            import io

            content = base64.b64decode(self.import_file)

            # Load workbook twice: once for values, once for formulas
            data_workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            data_sheet = data_workbook.active
            formula_workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
            formula_sheet = formula_workbook.active
            sheet = data_sheet

        except ImportError:
            raise UserError(_("openpyxl library required. Install with: pip install openpyxl"))
        except Exception as e:
            raise UserError(_("Failed to read Excel file: %s") % str(e))

        # Use advanced header detection
        try:
            from ..formula_engine.header_detector import HeaderDetector
            from ..formula_engine.merged_cell_parser import MergedCellParser
            from openpyxl.utils import get_column_letter, column_index_from_string

            detector = HeaderDetector(sheet)
            header_row, data_start_row, details = detector.detect_header_row()

            _logger.info(
                f"Excel import: detected header_row={header_row}, "
                f"data_start_row={data_start_row}, score={details.get('detection_score', 0):.2f}"
            )

            # Extract component types from merged cells above header
            parser = MergedCellParser(sheet)
            component_types = parser.extract_component_types(header_row)

            # Get headers using advanced detection (handles vertical merges)
            raw_headers = detector.get_headers(header_row)

            _logger.info(f"Excel import: found {len(raw_headers)} raw headers")

            # Check for empty columns between valid columns - just log them, don't error
            # We skip empty columns but preserve original Excel column letters so formulas remain valid
            empty_columns = self._detect_empty_columns_between_valid(raw_headers, sheet, header_row)
            if empty_columns:
                _logger.info(
                    f"Excel import: skipping {len(empty_columns)} empty columns between valid data: "
                    f"{', '.join(empty_columns)}. Original column positions preserved for formula references."
                )

        except UserError:
            raise  # Re-raise UserError without wrapping
        except Exception as e:
            _logger.warning(f"Advanced header detection failed, falling back to row 1: {e}")
            # Fallback to simple row 1 detection
            header_row = 1
            data_start_row = 2
            raw_headers = []
            component_types = {}
            from openpyxl.utils import get_column_letter, column_index_from_string

            for cell in sheet[1]:
                if cell.value:
                    raw_headers.append({
                        'column_letter': get_column_letter(cell.column),
                        'column_index': cell.column - 1,
                        'value': str(cell.value).strip(),
                        'from_vertical_merge': False,
                    })

        if not self.preserve_existing:
            self.config_id.rule_ids.unlink()

        existing_codes = set(self.config_id.rule_ids.mapped('code'))
        max_sequence = max(self.config_id.rule_ids.mapped('sequence') or [0])
        created_rules = self.env['hr.formula.rule']

        # Detect formula columns by checking first few data rows
        formula_columns = {}
        for row_offset in range(min(5, formula_sheet.max_row - data_start_row + 1)):
            check_row = data_start_row + row_offset
            for header in raw_headers:
                col_letter = header.get('column_letter')
                if col_letter and col_letter not in formula_columns:
                    col_idx = column_index_from_string(col_letter)
                    cell = formula_sheet.cell(row=check_row, column=col_idx)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_columns[col_letter] = cell.value

        _logger.info(f"Excel import: detected {len(formula_columns)} formula columns: {list(formula_columns.keys())}")

        # Detect colored constant pairs (red label + green value)
        # IMPORTANT: Use formula_sheet (data_only=False) to preserve color information
        # The 'sheet' variable is from data_only=True which strips formatting
        constant_pairs = self._detect_colored_constant_pairs(formula_sheet, header_row, data_sheet)

        # Detect blue font color constants (percentages and other values)
        # Scan up to data_start_row + 2 to catch constants in sub-header rows
        scan_up_to_row = data_start_row + 2  # Include potential sub-header rows
        blue_constants = self._detect_blue_constant_cells(formula_sheet, scan_up_to_row, data_sheet)

        # Also scan formulas to find referenced cells above header row that weren't detected by color
        formula_referenced_constants = self._detect_formula_referenced_constants(
            formula_columns, formula_sheet, header_row, data_sheet
        )

        # Add any formula-referenced constants not already detected by color
        detected_cells = {p['original_cell'] for p in constant_pairs}

        # Add blue constants (they have ConstantA, ConstantB naming)
        for blue_const in blue_constants:
            if blue_const['original_cell'] not in detected_cells:
                constant_pairs.append(blue_const)
                detected_cells.add(blue_const['original_cell'])

        # Add formula-referenced constants
        for ref_const in formula_referenced_constants:
            if ref_const['original_cell'] not in detected_cells:
                constant_pairs.append(ref_const)
                detected_cells.add(ref_const['original_cell'])

        # Build cell mapping for formula reference updates
        cell_to_column_mapping = {}
        for idx, pair in enumerate(constant_pairs):
            new_col_letter = self._generate_extended_column_letter(idx)
            cell_to_column_mapping[pair['original_cell']] = new_col_letter
            pair['new_column_letter'] = new_col_letter

        # Update formula columns with new references
        if cell_to_column_mapping:
            updated_formula_columns = {}
            for col_letter, formula in formula_columns.items():
                updated_formula = self._update_formula_references(formula, cell_to_column_mapping)
                updated_formula_columns[col_letter] = updated_formula
            formula_columns = updated_formula_columns

        # Process each header
        for header in raw_headers:
            header_value = header.get('value')
            if not header_value:
                continue

            col_letter = header.get('column_letter')

            # Determine if this is a formula column
            is_formula = col_letter in formula_columns
            column_type = 'formula' if is_formula else 'input'
            excel_formula = formula_columns.get(col_letter, '')

            # Get component type from merged cells
            comp_type = component_types.get(col_letter, '')

            name = str(header_value).strip()
            code = self._generate_code_from_label(name, existing_codes)
            existing_codes.add(code)

            # The colour-blind reader has no font signals at all, so the classifier
            # gets only what this path can honestly see: the header text, the merged
            # category band above it, and the first few values underneath.
            col_index = column_index_from_string(col_letter) if col_letter else None
            samples = self._collect_column_samples(
                data_sheet, col_index, data_start_row) if col_index else []
            role, role_tier, role_reason = column_role_classifier.classify_column(
                name,
                column_type=column_type,
                band_label=comp_type or None,
                sample_values=samples,
            )
            _logger.debug(
                "Excel import: column %s (%s) -> role=%s tier=%s (%s)",
                col_letter, name, role, role_tier, role_reason)

            max_sequence += 10
            values = {
                'config_id': self.config_id.id,
                'name': name,
                'code': code,
                'sequence': max_sequence,
                'column_type': column_type,
                'forced_column_letter': col_letter,  # Use forced to preserve actual Excel column position
                'original_column_letter': col_letter,
                'component_type': comp_type,
                'data_source': 'formula' if is_formula else 'excel',
                'data_source_field': name,
                'number_format': False,
            }
            values.update(self._role_rule_values(role))
            if excel_formula:
                values['excel_formula'] = excel_formula

            try:
                created = self.env['hr.formula.rule'].create(values)
                created_rules |= created
            except Exception as e:
                _logger.error(f"Excel import: failed to create rule '{name}' ({code}): {e}")

        # Create constant rules from colored pairs at the very end
        constant_rules_created = 0
        blue_constants_created = 0
        for pair in constant_pairs:
            name = pair['name']
            code = self._generate_code_from_label(name, existing_codes)
            existing_codes.add(code)

            max_sequence += 10

            # Convert value to float if possible
            try:
                constant_value = float(pair['value'])
            except (ValueError, TypeError):
                constant_value = 0.0
                _logger.warning(
                    f"Could not convert constant value '{pair['value']}' to float for '{name}'"
                )

            # Build informative data_source_field
            data_source_info = f"From cell {pair['original_cell']}"
            if pair.get('was_percentage'):
                original_val = pair.get('original_value', pair['value'])
                data_source_info = f"From cell {pair['original_cell']} ({original_val} -> {constant_value})"
                blue_constants_created += 1

            values = {
                'config_id': self.config_id.id,
                'name': name,
                'code': code,
                'sequence': max_sequence,
                'column_type': 'constant',
                'forced_column_letter': pair['new_column_letter'],  # Use forced to override auto-compute
                'original_column_letter': pair['original_cell'],  # Store original cell reference
                'component_type': 'Constant',
                'constant_value': constant_value,
                'data_source': 'manual',
                'data_source_field': data_source_info,
                'number_format': False,
            }

            try:
                created = self.env['hr.formula.rule'].create(values)
                created_rules |= created
                constant_rules_created += 1
            except Exception as e:
                _logger.error(f"Excel import: failed to create constant rule '{name}': {e}")

        # Build success message
        msg_parts = [_('%d rules imported from Excel (header row %d)') % (len(created_rules), header_row)]
        if constant_rules_created > 0:
            msg_parts.append(_('%d constant rules added (ZA onwards)') % constant_rules_created)
        if blue_constants_created > 0:
            msg_parts.append(_('%d blue constants (converted from %%)') % blue_constants_created)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'message': ' | '.join(msg_parts),
                'type': 'success',
                'sticky': False,
            }
        }

    def _import_from_excel_color_coded(self):
        """
        Import configuration from a color-coded Excel file.

        Color rules:
        - Yellow fill: Component name row (header)
        - Amber fill: Component type row (optional)
        - Green fill: Formula/data row (first row below header)
        - Blue font: Constants (below header, up to green row)
        """
        if not self.import_file:
            raise UserError(_("Please upload an Excel file"))

        try:
            import openpyxl
            import io

            content = base64.b64decode(self.import_file)
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            formula_workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
            sheet = workbook.active
            formula_sheet = formula_workbook.active

        except ImportError:
            raise UserError(_("openpyxl library required. Install with: pip install openpyxl"))
        except Exception as e:
            raise UserError(_("Failed to read Excel file: %s") % str(e))

        from openpyxl.utils import get_column_letter

        max_row = formula_sheet.max_row or 1
        max_col = formula_sheet.max_column or 1
        max_scan_row = min(max_row, 200)

        yellow_rows = {}
        amber_rows = {}
        green_rows = {}
        filled_rows = {}

        for row_num in range(1, max_scan_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = formula_sheet.cell(row=row_num, column=col_idx)
                if cell.fill and cell.fill.patternType not in (None, 'none'):
                    filled_rows[row_num] = filled_rows.get(row_num, 0) + 1
                if self._is_yellow_fill(cell):
                    yellow_rows[row_num] = yellow_rows.get(row_num, 0) + 1
                if self._is_amber_fill(cell):
                    amber_rows[row_num] = amber_rows.get(row_num, 0) + 1
                if self._is_green_fill(cell):
                    green_rows[row_num] = green_rows.get(row_num, 0) + 1

        def _pick_row(rows):
            if not rows:
                return None
            max_count = max(rows.values())
            candidates = [r for r, count in rows.items() if count == max_count]
            return min(candidates)

        colored_rows = set(yellow_rows.keys()) | set(amber_rows.keys())
        if not colored_rows:
            raise UserError(_("Could not detect header row from color-coded Excel file."))

        def _log_row_counts(label, rows):
            if not rows:
                _logger.info("Excel color import: %s rows: none", label)
                return
            items = sorted(rows.items(), key=lambda x: x[0])
            preview = items[:30]
            rows_str = ", ".join(["%s:%s" % (row, count) for row, count in preview])
            if len(items) > len(preview):
                rows_str += ", ..."
            _logger.info("Excel color import: %s rows counts: %s", label, rows_str)

        _log_row_counts("yellow", yellow_rows)
        _log_row_counts("amber", amber_rows)
        _log_row_counts("green", green_rows)
        _log_row_counts("filled", filled_rows)

        header_block_start = min(colored_rows)
        header_block_end = max(colored_rows)

        identifier_row = header_block_start - 1
        if identifier_row < 1:
            raise UserError(_("Could not detect identifier row above the header row."))

        formula_row = None
        green_candidates = {r: count for r, count in green_rows.items() if r > header_block_end}
        if green_candidates:
            max_count = max(green_candidates.values())
            formula_row = min(r for r, count in green_candidates.items() if count == max_count)
        else:
            filled_candidates = {
                r: count for r, count in filled_rows.items() if r > header_block_end
            }
            if filled_candidates:
                max_count = max(filled_candidates.values())
                formula_row = min(r for r, count in filled_candidates.items() if count == max_count)
            else:
                scan_limit = min(max_row, header_block_end + 50)
                for row_num in range(header_block_end + 1, scan_limit + 1):
                    for col_idx in range(1, max_col + 1):
                        if self._is_green_fill(formula_sheet.cell(row=row_num, column=col_idx)):
                            formula_row = row_num
                            break
                    if formula_row:
                        break

        if not formula_row:
            formula_row = header_block_end + 1

        _logger.info(
            "Excel color import: header_row=%s-%s, identifier_row=%s, formula_row=%s",
            header_block_start, header_block_end, identifier_row, formula_row
        )

        try:
            from ..formula_engine.merged_cell_parser import MergedCellParser
            merge_parser = MergedCellParser(formula_sheet)
        except Exception:
            merge_parser = None

        def get_merge_origin_cell(row_num, col_idx):
            if merge_parser:
                merge_info = merge_parser.get_merge_at(row_num, col_idx)
                if merge_info:
                    return formula_sheet.cell(
                        row=merge_info['min_row'],
                        column=merge_info['min_col'],
                    )
            return formula_sheet.cell(row=row_num, column=col_idx)

        def get_cell_value(row_num, col_idx):
            cell = get_merge_origin_cell(row_num, col_idx)
            if cell.value is not None:
                return cell.value
            return None

        def is_bold_in_merge(row_num, col_idx):
            if merge_parser:
                merge_info = merge_parser.get_merge_at(row_num, col_idx)
                if merge_info:
                    for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                        for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                            cell = formula_sheet.cell(row=m_row, column=m_col)
                            if cell.font and cell.font.bold:
                                return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return bool(cell.font and cell.font.bold)

        def is_red_font(cell):
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                            else:
                                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if r > 150 and g < 150 and b < 150:
                                return True
                            if r > 200 and g < 180 and b < 180 and r > g and r > b:
                                return True
                    elif color.type == 'indexed' and color.indexed in [2, 10]:
                        return True
            except Exception:
                pass
            return False

        def is_red_in_merge(row_num, col_idx):
            if merge_parser:
                merge_info = merge_parser.get_merge_at(row_num, col_idx)
                if merge_info:
                    for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                        for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                            cell = formula_sheet.cell(row=m_row, column=m_col)
                            if is_red_font(cell):
                                return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return is_red_font(cell)

        def is_green_font(cell):
            """Green HEADER font = "this contract component holds text" (CR-A4).

            Thresholds mirror `is_green_color` in the constant-value reader. There is
            no clash with the green FILL that marks the formula row, nor with a green
            font on a constant VALUE cell: this is only ever asked about cells inside
            the header band."""
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                            else:
                                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if g > 150 and r < 150 and b < 150:
                                return True
                            if g > 180 and r > 180 and b < 150:
                                return True
                            if g > 200 and g > r and g > b:
                                return True
                    elif color.type == 'indexed' and color.indexed in [3, 11]:
                        return True
                    elif color.type == 'theme' and color.theme in [6, 9]:
                        return True
            except Exception:
                pass
            return False

        def is_green_in_merge(row_num, col_idx):
            if merge_parser:
                merge_info = merge_parser.get_merge_at(row_num, col_idx)
                if merge_info:
                    for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                        for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                            cell = formula_sheet.cell(row=m_row, column=m_col)
                            if is_green_font(cell):
                                return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return is_green_font(cell)

        def is_underline_in_merge(row_num, col_idx):
            if merge_parser:
                merge_info = merge_parser.get_merge_at(row_num, col_idx)
                if merge_info:
                    for m_row in range(merge_info['min_row'], merge_info['max_row'] + 1):
                        for m_col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                            cell = formula_sheet.cell(row=m_row, column=m_col)
                            if cell.font and cell.font.underline:
                                return True
            cell = formula_sheet.cell(row=row_num, column=col_idx)
            return bool(cell.font and cell.font.underline)

        def get_fill_debug(cell):
            if not cell.fill:
                return "none"
            color = cell.fill.fgColor or cell.fill.start_color
            if not color:
                return "none"
            return "pattern=%s type=%s rgb=%s indexed=%s theme=%s tint=%s" % (
                cell.fill.patternType,
                getattr(color, 'type', None),
                getattr(color, 'rgb', None),
                getattr(color, 'indexed', None),
                getattr(color, 'theme', None),
                getattr(color, 'tint', None),
            )

        raw_headers = []
        header_map = {}
        header_cells = {}
        header_rows = {}

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header_value = None
            header_row = None
            # Prefer the lowest yellow cell in the header block
            for row_num in range(header_block_end, header_block_start - 1, -1):
                cell = get_merge_origin_cell(row_num, col_idx)
                if cell.value is None:
                    continue
                if self._is_yellow_fill(cell):
                    header_value = cell.value
                    header_row = row_num
                    break
            # Fallback to any non-empty value in the header block
            if header_value is None:
                for row_num in range(header_block_start, header_block_end + 1):
                    cell = get_merge_origin_cell(row_num, col_idx)
                    if cell.value is None:
                        continue
                    header_value = cell.value
                    header_row = row_num
                    break

            if header_value is None:
                continue

            value_str = str(header_value).strip()
            if not value_str:
                continue

            raw_headers.append({
                'column_letter': col_letter,
                'column_index': col_idx - 1,
                'value': value_str,
            })
            header_map[col_letter] = value_str
            header_cells[col_letter] = get_merge_origin_cell(header_row, col_idx)
            header_rows[col_letter] = header_row

        if not raw_headers:
            raise UserError(_("No component names found on the header row."))

        component_types = {}
        for header in raw_headers:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            header_row = header_rows.get(col_letter, header_block_end)
            comp_value = None
            for row_num in range(header_row - 1, header_block_start - 1, -1):
                comp_cell = get_merge_origin_cell(row_num, col_idx)
                if not self._is_amber_fill(comp_cell):
                    continue
                comp_value = get_cell_value(row_num, col_idx)
                if comp_value:
                    break
            if comp_value:
                component_types[col_letter] = str(comp_value).strip()

        identifier_map = {}
        for header in raw_headers:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            identifier_value = get_cell_value(identifier_row, col_idx)
            if identifier_value:
                identifier_map[col_letter] = str(identifier_value).strip()

        _logger.info("Excel color import: identifier_map sample: %s", list(identifier_map.items())[:20])

        # The identifier row here carries PAYSLIP section codes ("ITAXABLEINCOMECC"),
        # not employee identifiers, so it is deliberately NOT fed to the classifier as
        # `on_identifier_row` — see COLROLES ledger CR5.
        first_data_row = max(formula_row, header_block_end) + 1

        def collect_samples(col_idx):
            return self._collect_column_samples(sheet, col_idx, first_data_row)

        if not self.preserve_existing:
            self.config_id.rule_ids.unlink()

        existing_codes = set(self.config_id.rule_ids.mapped('code'))
        max_sequence = max(self.config_id.rule_ids.mapped('sequence') or [0])
        created_rules = self.env['hr.formula.rule']

        formula_columns = {}
        for header in raw_headers:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            cell = formula_sheet.cell(row=formula_row, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_columns[col_letter] = cell.value

        _logger.info("Excel color import: formula columns detected: %s", list(formula_columns.keys())[:30])

        constant_scan_start = max(1, header_block_start - 1)
        constant_pairs = self._detect_blue_constant_cells_color_coded(
            formula_sheet, constant_scan_start, formula_row, header_map, sheet
        )
        constant_pairs = sorted(
            constant_pairs,
            key=lambda item: (item.get('row', 0), item.get('original_col_idx', 0))
        )

        cell_to_column_mapping = {}
        for idx, pair in enumerate(constant_pairs):
            new_col_letter = self._generate_extended_column_letter(idx)
            cell_to_column_mapping[pair['original_cell']] = new_col_letter
            pair['new_column_letter'] = new_col_letter

        if cell_to_column_mapping:
            updated_formula_columns = {}
            for col_letter, formula in formula_columns.items():
                updated_formula_columns[col_letter] = self._update_formula_references(
                    formula, cell_to_column_mapping
                )
            formula_columns = updated_formula_columns

        for header in raw_headers[:30]:
            col_letter = header['column_letter']
            col_idx = header['column_index'] + 1
            header_row = header_rows.get(col_letter, header_block_end)
            header_cell = get_merge_origin_cell(header_row, col_idx)
            identifier_cell = get_merge_origin_cell(identifier_row, col_idx)
            comp_cell = get_merge_origin_cell(header_row - 1, col_idx) if header_row else None
            formula_cell = formula_sheet.cell(row=formula_row, column=col_idx)
            _logger.info(
                "Excel color import: col=%s header='%s' header_fill=%s header_is_yellow=%s "
                "header_bold=%s id_value='%s' id_fill=%s comp_value='%s' comp_fill=%s comp_is_amber=%s "
                "formula_cell_value='%s' formula_is_green=%s formula_fill=%s",
                col_letter,
                header.get('value'),
                get_fill_debug(header_cell),
                self._is_yellow_fill(header_cell),
                is_bold_in_merge(header_row, col_idx),
                get_cell_value(identifier_row, col_idx),
                get_fill_debug(identifier_cell),
                get_cell_value(header_row - 1, col_idx) if header_row else '',
                get_fill_debug(comp_cell) if comp_cell else 'none',
                self._is_amber_fill(comp_cell) if comp_cell else False,
                formula_cell.value,
                self._is_green_fill(formula_cell),
                get_fill_debug(formula_cell),
            )

        identifier_cache = {}
        payslip_config_model = self.env['hr.payslip.config']

        def get_payslip_identifier(identifier_value):
            if not identifier_value:
                return False
            if identifier_value in identifier_cache:
                return identifier_cache[identifier_value]
            record = payslip_config_model.search([
                ('identifier', '=', identifier_value),
                ('salary_structure_id', '=', self.config_id.id),
            ], limit=1)
            if not record:
                record = payslip_config_model.create({
                    'salary_structure_id': self.config_id.id,
                    'identifier': identifier_value,
                    'label': '',
                })
            identifier_cache[identifier_value] = record
            return record

        component_rules_created = 0
        for header in raw_headers:
            header_value = header.get('value')
            if not header_value:
                continue

            col_letter = header.get('column_letter')
            col_idx = header.get('column_index') + 1
            is_formula = col_letter in formula_columns
            column_type = 'formula' if is_formula else 'input'
            excel_formula = formula_columns.get(col_letter, '')

            comp_type = component_types.get(col_letter, '')

            name = str(header_value).strip()
            code = self._generate_code_from_label(name, existing_codes)
            existing_codes.add(code)

            identifier_value = identifier_map.get(col_letter)
            payslip_identifier = get_payslip_identifier(identifier_value)

            header_row = header_rows.get(col_letter, header_block_end)
            report_visible = is_bold_in_merge(header_row, col_idx)
            is_red = is_red_in_merge(header_row, col_idx)
            is_green = is_green_in_merge(header_row, col_idx)
            # Green is authoritative text, red is a component whose kind is inferred
            # from what the column actually contains (CR-A4). Underline means "start a
            # new contract when this changes" for either colour.
            is_contract_component = is_red or is_green
            requires_new_contract = is_contract_component and is_underline_in_merge(header_row, col_idx)

            samples = collect_samples(col_idx)
            role, role_tier, role_reason = column_role_classifier.classify_column(
                name,
                column_type=column_type,
                is_contract_component=is_red,
                is_text_component=is_green,
                band_label=comp_type or None,
                sample_values=samples,
            )
            is_text_component = is_green or (
                is_red and role == column_role_classifier.ROLE_CONTRACT)
            _logger.debug(
                "Excel color import: column %s (%s) -> role=%s tier=%s (%s)",
                col_letter, name, role, role_tier, role_reason)

            max_sequence += 10
            values = {
                'config_id': self.config_id.id,
                'name': name,
                'code': code,
                'sequence': max_sequence,
                'column_type': column_type,
                'forced_column_letter': col_letter,
                'original_column_letter': col_letter,
                'component_type': comp_type,
                'data_source': 'formula' if is_formula else 'excel',
                'data_source_field': name,
                'number_format': False,
                'payslip_identifier': payslip_identifier.id if payslip_identifier else False,
                'report_visible': report_visible,
                'is_contract_component': is_contract_component,
                'requires_new_contract': requires_new_contract,
                'is_text_component': is_text_component,
            }
            values.update(self._role_rule_values(role))
            if excel_formula:
                values['excel_formula'] = excel_formula

            try:
                created = self.env['hr.formula.rule'].create(values)
                created_rules |= created
                component_rules_created += 1
            except Exception as e:
                _logger.error("Excel color import: failed to create rule '%s' (%s): %s", name, code, e)

        constant_rules_created = 0
        for pair in constant_pairs:
            name = pair['name']
            code = self._generate_code_from_label(name, existing_codes)
            existing_codes.add(code)
            max_sequence += 10

            try:
                constant_value = float(pair['value'])
            except (ValueError, TypeError):
                constant_value = 0.0
                _logger.warning(
                    "Could not convert constant value '%s' to float for '%s'",
                    pair['value'], name
                )

            data_source_info = "From cell %s" % pair['original_cell']
            if pair.get('was_percentage'):
                original_val = pair.get('original_value', pair['value'])
                data_source_info = "From cell %s (%s -> %s)" % (
                    pair['original_cell'], original_val, constant_value
                )

            values = {
                'config_id': self.config_id.id,
                'name': name,
                'code': code,
                'sequence': max_sequence,
                'column_type': 'constant',
                'forced_column_letter': pair['new_column_letter'],
                'original_column_letter': pair['original_cell'],
                'component_type': 'Constant',
                'constant_value': constant_value,
                'data_source': 'manual',
                'data_source_field': data_source_info,
                'number_format': False,
            }

            try:
                created = self.env['hr.formula.rule'].create(values)
                created_rules |= created
                constant_rules_created += 1
            except Exception as e:
                _logger.error("Excel color import: failed to create constant rule '%s': %s", name, e)

        msg_parts = [
            _('%d rules imported from color-coded Excel (header row %d)') % (
                component_rules_created, header_block_end
            )
        ]
        if constant_rules_created:
            msg_parts.append(_('%d constant rules added') % constant_rules_created)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'message': ' | '.join(msg_parts),
                'type': 'success',
                'sticky': False,
            }
        }

    def action_download_template(self):
        """Download Excel template for import."""
        try:
            import openpyxl
            import io
            import base64

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Formula Rules"

            # Template format:
            # Row 1: Labels/Names (stop at first blank)
            # Row 2: Values or formulas. Blank/value => input; formula => formula column_type
            headers = ["Basic Salary", "Std Wrk Hrs", "Actual Wrk Hrs", "Overtime Pay", "Net Pay Cap"]
            formulas = ["", "", "", "=C2*1.5", "=IF(D2>15000000,15000000,D2)"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)

            for col, value in enumerate(formulas, 1):
                ws.cell(row=2, column=col).value = value

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()

            # Create attachment
            attachment = self.env['ir.attachment'].create({
                'name': 'formula_import_template.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(content),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }

        except ImportError:
            raise UserError(_("openpyxl library required for template generation"))

    # --------------------------------------------------
    # Helpers
    # --------------------------------------------------
    def _get_fill_rgb(self, cell):
        fill = cell.fill
        if not fill or not fill.patternType or fill.patternType == 'none':
            return None

        color = fill.fgColor or fill.start_color
        if not color:
            return None

        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb).upper()
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) == 6:
                return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)

        if color.type == 'indexed':
            try:
                from openpyxl.styles.colors import COLOR_INDEX
                idx = color.indexed
                if idx is not None and idx < len(COLOR_INDEX):
                    rgb = str(COLOR_INDEX[idx]).upper()
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if len(rgb) == 6:
                        return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            except Exception:
                return None

        return None

    def _is_yellow_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return r > 200 and g > 200 and b < 150

    def _is_amber_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return r > 200 and 120 < g < 210 and b < 120

    def _is_green_fill(self, cell):
        rgb = self._get_fill_rgb(cell)
        if not rgb:
            return False
        r, g, b = rgb
        return g > 150 and g > r and g > b and r < 200 and b < 200

    def _is_blue_font(self, cell):
        try:
            font = cell.font
            if not font or not font.color:
                return False

            color = font.color

            if color.type == 'rgb' and color.rgb:
                rgb = str(color.rgb).upper()
                if len(rgb) >= 6:
                    if len(rgb) == 8:
                        r = int(rgb[2:4], 16)
                        g = int(rgb[4:6], 16)
                        b = int(rgb[6:8], 16)
                    else:
                        r = int(rgb[0:2], 16)
                        g = int(rgb[2:4], 16)
                        b = int(rgb[4:6], 16)
                    if b > r and b > g and b > 80:
                        return True

            if color.type == 'indexed':
                if color.indexed in [
                    4, 5, 12, 23, 30, 32, 39, 40, 41, 42, 48, 49, 54, 55, 56
                ]:
                    return True
                value = cell.value
                if color.indexed not in (0, 1) and isinstance(value, (int, float)):
                    return True
                if color.indexed not in (0, 1) and isinstance(value, str) and value.strip().endswith('%'):
                    return True

            if color.type == 'theme':
                if color.theme in [4, 5, 8]:
                    return True
                value = cell.value
                if isinstance(value, (int, float)):
                    return True
                if isinstance(value, str) and value.strip().endswith('%'):
                    return True

        except Exception:
            return False

        return False

    def _parse_percentage_value(self, value):
        """
        Parse a percentage value and convert to decimal.

        Examples:
            "8%" -> 0.08
            "1.5%" -> 0.015
            "100%" -> 1.0
            0.08 (already decimal) -> 0.08
            8 (just number) -> 8.0

        Returns:
            Tuple of (decimal_value, was_percentage)
        """
        if value is None:
            return 0.0, False

        # If it's already a number (float/int)
        if isinstance(value, (int, float)):
            # Check if it looks like a small decimal that might be a percentage
            # e.g., 0.08 from Excel when cell is formatted as percentage
            return float(value), False

        # If it's a string, check for percentage sign
        str_value = str(value).strip()
        if str_value.endswith('%'):
            try:
                # Remove % and convert to decimal
                num_part = str_value[:-1].strip()
                decimal_value = float(num_part) / 100.0
                return decimal_value, True
            except ValueError:
                return 0.0, False

        # Try to convert as regular number
        try:
            return float(str_value), False
        except ValueError:
            return 0.0, False

    def _evaluate_constant_formula(self, formula):
        if not formula or not isinstance(formula, str):
            return None
        expr = formula.strip()
        if not expr.startswith('='):
            return None
        expr = expr[1:]
        import re
        if re.search(r'[A-Za-z_]', expr):
            return None
        expr = re.sub(r'(\d+(?:\.\d+)?)\s*%', r'(\1/100)', expr)
        if re.search(r'[^0-9\.\+\-\*\/\^\(\)\s]', expr):
            return None
        expr = expr.replace('^', '**')
        try:
            return eval(expr, {"__builtins__": {}}, {})
        except Exception:
            return None

    def _get_constant_cell_value(self, formula_sheet, data_sheet, row_num, col_idx):
        cell = formula_sheet.cell(row=row_num, column=col_idx)
        value = cell.value
        if isinstance(value, str) and value.startswith('='):
            if data_sheet:
                data_value = data_sheet.cell(row=row_num, column=col_idx).value
                if data_value is not None:
                    return data_value
            evaluated = self._evaluate_constant_formula(value)
            if evaluated is not None:
                return evaluated
        return value

    def _detect_blue_constant_cells_color_coded(self, sheet, header_row, formula_row, header_map, data_sheet=None):
        """
        Detect blue font constants below header row up to the formula row.

        Names are generated as "Constant <Header Name>" using the header row.
        """
        from openpyxl.utils import get_column_letter

        blue_constants = []
        seen_cells = set()
        start_row = 1
        end_row = max(start_row, (formula_row - 1) if formula_row else start_row)
        color_debug_logged = 0
        style_debug_logged = 0

        _logger.info(
            "Excel color import: constant scan rows=%s-%s, max_col=%s",
            start_row, end_row, sheet.max_column or 1
        )

        def get_font_debug(cell):
            font = cell.font
            if not font or not font.color:
                return "none"
            color = font.color
            return "type=%s rgb=%s indexed=%s theme=%s tint=%s" % (
                getattr(color, 'type', None),
                getattr(color, 'rgb', None),
                getattr(color, 'indexed', None),
                getattr(color, 'theme', None),
                getattr(color, 'tint', None),
            )

        for row_num in range(start_row, end_row + 1):
            for col_idx in range(1, (sheet.max_column or 1) + 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                if cell.value is None:
                    continue

                if style_debug_logged < 40 and (cell.has_style or cell.font):
                    col_letter = get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_num}"
                    font = cell.font
                    _logger.info(
                        "Excel color import: font debug cell=%s value='%s' has_style=%s "
                        "font_color=%s bold=%s underline=%s name=%s size=%s scheme=%s hyperlink=%s",
                        cell_ref,
                        cell.value,
                        cell.has_style,
                        get_font_debug(cell),
                        getattr(font, 'bold', None),
                        getattr(font, 'underline', None),
                        getattr(font, 'name', None),
                        getattr(font, 'size', None),
                        getattr(font, 'scheme', None),
                        bool(cell.hyperlink),
                    )
                    style_debug_logged += 1

                if not self._is_blue_font(cell):
                    if color_debug_logged < 30 and cell.font and cell.font.color:
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{row_num}"
                        _logger.info(
                            "Excel color import: non-blue font cell=%s value='%s' font=%s",
                            cell_ref, cell.value, get_font_debug(cell)
                        )
                        color_debug_logged += 1
                    continue

                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_num}"
                if cell_ref in seen_cells:
                    continue
                seen_cells.add(cell_ref)

                value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)
                decimal_value, was_percentage = self._parse_percentage_value(value)
                header_label = header_map.get(col_letter)
                name = f"Constant {header_label}" if header_label else f"Constant {col_letter}"

                blue_constants.append({
                    'name': name,
                    'value': decimal_value,
                    'original_value': cell.value,
                    'original_cell': cell_ref,
                    'original_col_letter': col_letter,
                    'original_col_idx': col_idx,
                    'row': row_num,
                    'was_percentage': was_percentage,
                })

                _logger.info(
                    "Excel color import: BLUE CONSTANT cell=%s value='%s' name='%s' font=%s",
                    cell_ref, cell.value, name, get_font_debug(cell)
                )

        return blue_constants

    def _detect_blue_constant_cells(self, sheet, scan_up_to_row, data_sheet=None):
        """
        Detect cells with blue font color that contain constant values.

        Scans all rows up to scan_up_to_row for cells with blue font.
        Blue is detected when B > R and B > G in the RGB color.

        For cells that are percentages (e.g., "8%"), converts to decimal (0.08).

        Returns list of dictionaries with:
            - name: Generated name (ConstantA, ConstantB, etc.)
            - value: The numeric value (percentages converted to decimals)
            - original_cell: Cell reference like "AY11"
            - original_col_letter: Column letter
            - original_col_idx: Column index
            - row: Row number
            - was_percentage: Boolean if value was originally a percentage
        """
        from openpyxl.utils import get_column_letter

        _logger.info(f"=== BLUE CONSTANT DETECTION START ===")
        _logger.info(f"Scanning rows 1 to {scan_up_to_row - 1}")

        blue_constants = []
        constant_index = 0  # For naming ConstantA, ConstantB, etc.
        cells_with_color_checked = 0

        def is_blue_color(cell, cell_ref=""):
            """Check if cell has blue FONT color (any shade where B > R and B > G)."""
            try:
                font = cell.font
                if not font or not font.color:
                    return False, None

                color = font.color
                color_info = f"type={color.type}"

                # Check RGB color
                if color.type == 'rgb' and color.rgb:
                    rgb = str(color.rgb).upper()
                    color_info += f", rgb={rgb}"
                    if len(rgb) >= 6:
                        # Handle ARGB (8 chars) or RGB (6 chars)
                        if len(rgb) == 8:
                            r = int(rgb[2:4], 16)
                            g = int(rgb[4:6], 16)
                            b = int(rgb[6:8], 16)
                        else:
                            r = int(rgb[0:2], 16)
                            g = int(rgb[2:4], 16)
                            b = int(rgb[4:6], 16)

                        color_info += f" -> R={r}, G={g}, B={b}"

                        # Blue: B is dominant (B > R and B > G)
                        if b > r and b > g and b > 100:
                            return True, color_info

                # Check indexed colors (common blue indices)
                elif color.type == 'indexed':
                    color_info += f", indexed={color.indexed}"
                    if color.indexed in [4, 5, 12, 23, 30, 32, 39, 40, 41, 42, 48, 49, 54, 55, 56]:
                        return True, color_info

                # Check theme colors (theme 4, 5, 8 are often blue variants)
                elif color.type == 'theme':
                    color_info += f", theme={color.theme}"
                    if color.theme in [4, 5, 8]:
                        return True, color_info

                return False, color_info

            except Exception as e:
                return False, f"error: {e}"

        def generate_constant_name(index):
            """Generate ConstantA, ConstantB, ..., ConstantZ, ConstantAA, etc."""
            result = ""
            idx = index
            while True:
                result = chr(ord('A') + (idx % 26)) + result
                idx = idx // 26 - 1
                if idx < 0:
                    break
            return f"Constant{result}"

        def get_column_header_label(row_num, col_idx):
            """
            Look at the row immediately above to get a header label for this column.
            This appends context like _BHXH, _BHYT to the constant name.
            """
            if row_num <= 1:
                return None

            # Check the row immediately above
            header_cell = sheet.cell(row=row_num - 1, column=col_idx)
            if header_cell.value and isinstance(header_cell.value, str):
                label = str(header_cell.value).strip()
                # Clean up the label - remove special chars, keep alphanumeric and underscore
                import re
                clean_label = re.sub(r'[^A-Za-z0-9_]', '', label)
                if clean_label:
                    return clean_label

            return None

        # Scan all rows up to scan_up_to_row
        for row_num in range(1, scan_up_to_row):
            for col_idx in range(1, (sheet.max_column or 1) + 1):
                cell = sheet.cell(row=row_num, column=col_idx)

                # Skip empty cells
                if cell.value is None:
                    continue

                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_num}"

                # Check if cell has font color
                is_blue, color_info = is_blue_color(cell, cell_ref)
                if color_info:
                    cells_with_color_checked += 1
                    # Log cells with any color info for debugging
                    if cells_with_color_checked <= 20:  # Limit logging
                        _logger.info(f"  Cell {cell_ref} value='{cell.value}' color: {color_info} -> is_blue={is_blue}")

                # Check if cell has blue font
                if is_blue:
                    # Parse the value (handles percentages)
                    value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)
                    decimal_value, was_percentage = self._parse_percentage_value(value)

                    # Generate base name (ConstantA, ConstantB, etc.)
                    base_name = generate_constant_name(constant_index)
                    constant_index += 1

                    # Look for header label in row above to append context
                    header_label = get_column_header_label(row_num, col_idx)
                    if header_label:
                        name = f"{base_name}_{header_label}"
                    else:
                        name = base_name

                    blue_constants.append({
                        'name': name,
                        'value': decimal_value,
                        'original_value': cell.value,  # Keep original for reference
                        'original_cell': cell_ref,
                        'original_col_letter': col_letter,
                        'original_col_idx': col_idx,
                        'row': row_num,
                        'was_percentage': was_percentage,
                    })

                    _logger.info(
                        f"  -> BLUE CONSTANT FOUND: {cell_ref} = {cell.value} "
                        f"-> {name} = {decimal_value}"
                        f"{' (converted from %)' if was_percentage else ''}"
                    )

        _logger.info(f"=== BLUE CONSTANT DETECTION END: {len(blue_constants)} found, {cells_with_color_checked} cells with color checked ===")

        return blue_constants

    def _detect_colored_constant_pairs(self, sheet, header_row, data_sheet=None):
        """
        Detect red/green colored cell pairs that represent constants.

        These are cells where:
        - Left cell has RED font color (component name/label)
        - Right cell has GREEN font color (constant value)

        Note: We check FONT color (text color), not background/fill color.
        """
        from openpyxl.utils import get_column_letter

        constant_pairs = []

        def is_red_color(cell):
            """Check if cell has red FONT color."""
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                            else:
                                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if r > 150 and g < 150 and b < 150:
                                return True
                            if r > 200 and g < 180 and b < 180 and r > g and r > b:
                                return True
                    elif color.type == 'indexed' and color.indexed in [2, 10]:
                        return True
            except Exception:
                pass
            return False

        def is_green_color(cell):
            """Check if cell has green FONT color."""
            try:
                font = cell.font
                if font and font.color:
                    color = font.color
                    if color.type == 'rgb' and color.rgb:
                        rgb = str(color.rgb).upper()
                        if len(rgb) >= 6:
                            if len(rgb) == 8:
                                r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                            else:
                                r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if g > 150 and r < 150 and b < 150:
                                return True
                            if g > 180 and r > 180 and b < 150:
                                return True
                            if g > 200 and g > r and g > b:
                                return True
                    elif color.type == 'indexed' and color.indexed in [3, 11]:
                        return True
                    elif color.type == 'theme' and color.theme in [6, 9]:
                        return True
            except Exception:
                pass
            return False

        max_row_to_search = min(header_row + 5, sheet.max_row or 1)

        for row_num in range(1, max_row_to_search + 1):
            if row_num >= header_row:
                continue

            for col_idx in range(1, (sheet.max_column or 1) + 1):
                cell_left = sheet.cell(row=row_num, column=col_idx)
                cell_right = sheet.cell(row=row_num, column=col_idx + 1)

                if is_red_color(cell_left) and is_green_color(cell_right):
                    name = cell_left.value
                    value = self._get_constant_cell_value(
                        sheet,
                        data_sheet,
                        row_num,
                        col_idx + 1,
                    )

                    if name is not None and value is not None:
                        value_col_letter = get_column_letter(col_idx + 1)
                        original_cell_ref = f"{value_col_letter}{row_num}"

                        constant_pairs.append({
                            'name': str(name).strip(),
                            'value': value,
                            'original_cell': original_cell_ref,
                            'original_col_letter': value_col_letter,
                            'original_col_idx': col_idx + 1,
                            'row': row_num,
                        })

        return constant_pairs

    def _detect_formula_referenced_constants(self, formula_columns, sheet, header_row, data_sheet=None):
        """
        Scan formulas to find cell references above header row that might be constants.
        This catches constants like $CF$3 that may not have been detected by color.
        """
        import re
        from openpyxl.utils import column_index_from_string

        found_constants = []
        seen_cells = set()
        cell_ref_pattern = r'\$?([A-Z]+)\$?(\d+)'

        for col_letter, formula in formula_columns.items():
            matches = re.findall(cell_ref_pattern, formula, re.IGNORECASE)

            for col_match, row_match in matches:
                row_num = int(row_match)
                col_match = col_match.upper()

                if row_num >= header_row:
                    continue

                cell_ref = f"{col_match}{row_num}"
                if cell_ref in seen_cells:
                    continue
                seen_cells.add(cell_ref)

                try:
                    col_idx = column_index_from_string(col_match)
                    value = self._get_constant_cell_value(sheet, data_sheet, row_num, col_idx)

                    if value is None:
                        continue

                    # Try to get a name from the cell to the left
                    name = None
                    if col_idx > 1:
                        left_cell = sheet.cell(row=row_num, column=col_idx - 1)
                        if left_cell.value and isinstance(left_cell.value, str):
                            name = str(left_cell.value).strip()

                    if not name:
                        name = f"Constant_{col_match}{row_num}"

                    found_constants.append({
                        'name': name,
                        'value': value,
                        'original_cell': cell_ref,
                        'original_col_letter': col_match,
                        'original_col_idx': col_idx,
                        'row': row_num,
                    })

                except Exception as e:
                    _logger.warning(f"Error processing cell {cell_ref}: {e}")

        return found_constants

    def _generate_extended_column_letter(self, index):
        """
        Generate column letters starting from ZA, ZB, ZC, etc.

        Args:
            index: 0-based index (0=ZA, 1=ZB, etc.)

        Returns:
            Column letter string
        """
        # Start from ZA (which is column 677 in Excel, 1-based)
        # ZA = 26*26 + 1 = 677
        base_col = 26 * 26 + 1 + index  # ZA starts at position 677

        result = ""
        col = base_col
        while col > 0:
            col -= 1
            result = chr(col % 26 + ord('A')) + result
            col //= 26

        return result

    def _update_formula_references(self, formula, cell_mapping):
        """
        Update formula to replace original cell references with new column letters.
        e.g., {'CE3': 'ZA'} -> replaces $CE$3 or CE3 with ZA2
        """
        import re

        if not formula or not cell_mapping:
            return formula

        updated_formula = formula
        sorted_refs = sorted(cell_mapping.keys(), key=len, reverse=True)

        for original_ref in sorted_refs:
            new_col = cell_mapping[original_ref]
            cell_match = re.match(r'^([A-Za-z]+)(\d+)$', original_ref)
            if not cell_match:
                continue

            col_letters = cell_match.group(1)
            row_num = cell_match.group(2)
            pattern = (
                r"(?<![A-Za-z0-9_])"
                r"\$?" + col_letters + r"\$?" + row_num +
                r"(?![0-9A-Za-z_])"
            )
            replacement = f"{new_col}2"
            updated_formula = re.sub(pattern, replacement, updated_formula, flags=re.IGNORECASE)

        return updated_formula

    def _detect_empty_columns_between_valid(self, raw_headers, sheet, header_row):
        """
        Detect empty columns between valid columns.

        Args:
            raw_headers: List of header dictionaries with 'column_letter' and 'value' keys
            sheet: openpyxl worksheet object
            header_row: The detected header row number

        Returns:
            List of empty column letters between valid columns, or empty list if none found
        """
        from openpyxl.utils import get_column_letter, column_index_from_string

        if not raw_headers:
            return []

        # Get all column indices that have valid headers
        valid_col_indices = []
        for header in raw_headers:
            col_letter = header.get('column_letter')
            if col_letter and header.get('value'):
                try:
                    col_idx = column_index_from_string(col_letter)
                    valid_col_indices.append(col_idx)
                except Exception:
                    pass

        if len(valid_col_indices) < 2:
            return []

        # Find the range of columns with data
        min_col = min(valid_col_indices)
        max_col = max(valid_col_indices)

        # Check each column between min and max for empty columns
        empty_columns = []
        valid_col_set = set(valid_col_indices)

        for col_idx in range(min_col, max_col + 1):
            if col_idx not in valid_col_set:
                # This column is not in our headers - check if it's truly empty
                col_letter = get_column_letter(col_idx)

                # Check header row
                header_cell = sheet.cell(row=header_row, column=col_idx)
                header_value = header_cell.value

                # Check a few data rows too (to be thorough)
                has_data = False
                for row_offset in range(5):  # Check first 5 data rows
                    data_row = header_row + 1 + row_offset
                    if data_row > (sheet.max_row or 1):
                        break
                    data_cell = sheet.cell(row=data_row, column=col_idx)
                    if data_cell.value is not None:
                        has_data = True
                        break

                # If both header and data are empty, it's an empty column
                if (header_value is None or (isinstance(header_value, str) and not header_value.strip())) and not has_data:
                    empty_columns.append(col_letter)

        return empty_columns

    def _generate_code_from_label(self, label, existing_codes):
        """Create a short unique code (3-10 chars) derived from the label."""
        import re

        base = re.sub(r'[^A-Za-z0-9]', '', label).upper()
        if not base:
            base = 'COL'

        if len(base) < 3:
            base = (base + 'XXX')[:3]
        if len(base) > 10:
            base = base[:10]

        code = base
        suffix = 1
        while code in existing_codes:
            # ensure total length <=10 when adding suffix
            trimmed = base[: max(1, 10 - len(str(suffix)))]
            code = f"{trimmed}{suffix}"
            suffix += 1

        return code
