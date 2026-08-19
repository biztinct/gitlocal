# Copyright 2025
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

from datetime import datetime

from odoo import fields, models
from odoo.tools.misc import file_path


def _fmt(date_val):
    """Format date to dd/MM/yyyy expected by VN templates."""
    if not date_val:
        return ""
    if isinstance(date_val, str):
        try:
            date_val = fields.Date.from_string(date_val)
        except Exception:
            return date_val
    return date_val.strftime("%d/%m/%Y")


class PbGovtReportBase(models.AbstractModel):
    _name = "pb.govt.report.base"
    _description = "Base model for Vietnamese government XLS reports"

    def _get_employees_in_range(self, company, date_from, date_to):
        """Return employees with active contract overlap in range."""
        Contract = self.env["hr.contract"]
        employees = self.env["hr.employee"].browse()
        contracts = Contract.search(
            [
                ("company_id", "=", company.id),
                ("state", "=", "open"),
                "|",
                "&",
                ("date_end", "!=", False),
                ("date_end", ">=", date_from),
                "&",
                ("date_end", "=", False),
                ("date_start", "<=", date_to),
            ]
        )
        if contracts:
            employees |= contracts.mapped("employee_id")
        return employees

    def _get_payslips_in_range(self, company, date_from, date_to):
        Payslip = self.env["hr.payslip"]
        return Payslip.search(
            [
                ("company_id", "=", company.id),
                ("date_from", "<=", date_to),
                ("date_to", ">=", date_from),
                ("state", "in", ["done", "paid", "verify", "draft"]),
            ]
        )

    def _payslip_lines_map(self, payslips):
        """Return map: employee_id -> {rule_code: total amount} for quick lookup."""
        res = {}
        for slip in payslips:
            emp_map = res.setdefault(slip.employee_id.id, {})
            for line in slip.line_ids:
                emp_map[line.code] = emp_map.get(line.code, 0.0) + line.total
        return res

    def _default_filename(self, report_code, company, date_from):
        suffix = ""
        if date_from:
            if isinstance(date_from, str):
                try:
                    date_from = fields.Date.from_string(date_from)
                except Exception:
                    date_from = None
            if isinstance(date_from, datetime):
                date_from = date_from.date()
            if date_from:
                suffix = date_from.strftime("%Y%m")
        return f"{report_code}_{company.name}_{suffix}".replace(" ", "_")

    def _rule_amount(self, emp_lines, codes, default=0.0):
        """Fetch first non-zero rule amount from a list of codes."""
        for code in codes:
            val = emp_lines.get(code, 0.0)
            if val:
                return val
        return default

    # ------------------------------------------------------------------
    # The Odoo-19 employee master
    # ------------------------------------------------------------------
    # Odoo 19 deleted BOTH fields these filings used to read:
    #   hr.employee.address_home_id  — the home-address res.partner
    #   hr.employee.bank_account_id  — the scalar payroll bank account
    # An attribute read on either raises AttributeError, and the filing flow's
    # generate() turns that into "This filing could not be generated: …", which
    # is why four of the five VN filings had been silently unusable.
    #
    # Everything they need now comes from one of two places, in this order:
    #
    #  1. the VIETNAM EMPLOYEE MASTER (pb_hr_payroll_vietnam) —
    #     vietnam_province / vietnam_district / vietnam_ward,
    #     vietnam_permanent_address, vietnam_bank_*. These are the fields a
    #     Vietnamese payroll team actually maintains; the three address ones are
    #     already the exact administrative levels a BHXH form asks for (tỉnh /
    #     quận-huyện / phường-xã), which no core Odoo field is; and
    #     pb_pay_delivery pays salaries out of the same bank columns
    #     (bank_export_wizard.py:98-103), so a filing and a payment can no
    #     longer name two different accounts for one person.
    #  2. core Odoo 19 PRIVATE address / bank fields (private_street,
    #     private_city, private_state_id, private_zip, primary_bank_account_id)
    #     — for a database that does not carry the VN pack.
    #
    # work_contact_id is deliberately absent from that list. It is the OFFICE
    # partner. Putting the office on a "Địa chỉ liên hệ" line would file a
    # confidently WRONG address instead of a blank one, which is worse.
    #
    # The probes are ORM-registry reads, never try/except: "the VN pack is not
    # installed" is a known, testable state, and
    # pb_hr_govt/tests/test_odoo19_employee_sources.py asserts that on a
    # database WITH the pack the VN branch is the one taken — otherwise a dead
    # source is indistinguishable from an absent one (W79).

    def _emp_field(self, emp, name):
        """Read `name` off the employee, or False when the field is not in the
        registry on this database."""
        if not emp or name not in emp._fields:
            return False
        return emp[name]

    def _is_female(self, emp):
        """Odoo 19 renamed hr.employee.gender to `sex` (same selection values,
        same 'Gender' label). Every VN form has a "Nữ (X)" column."""
        return (self._emp_field(emp, "sex") or "") == "female"

    def _home_address(self, emp):
        """The employee's HOME address on one line ("Địa chỉ liên hệ").

        Permanent address first: BHXH asks where somebody is registered, not
        where they are currently staying, and the temporary one is the fallback
        only because a blank cell fails the submission.
        """
        for name in ("vietnam_permanent_address", "vietnam_temporary_address"):
            value = (self._emp_field(emp, name) or "").strip()
            if value:
                return " ".join(value.split())
        parts = [
            (self._emp_field(emp, "private_street") or "").strip(),
            (self._emp_field(emp, "private_street2") or "").strip(),
            (self._emp_field(emp, "private_city") or "").strip(),
        ]
        state = self._emp_field(emp, "private_state_id")
        if state:
            parts.append(state.name or "")
        parts.append((self._emp_field(emp, "private_zip") or "").strip())
        country = self._emp_field(emp, "private_country_id")
        if country:
            parts.append(country.name or "")
        return ", ".join(p for p in parts if p)

    def _code_for(self, lookup_type, value):
        """A VN administrative code from a NAME, via pb.govt.code.lookup.

        A value that is already all digits IS the code — the VN master fields
        are free text and a payroll team that has typed the code should not
        have it thrown away by a name search that cannot match it.
        """
        text = (value or "").strip()
        if not text:
            return "000"
        if text.isdigit():
            return text
        rec = self.env["pb.govt.code.lookup"].search(
            [("lookup_type", "=", lookup_type), ("name", "ilike", text)], limit=1)
        return (rec.code or "000") if rec else "000"

    def _location_codes(self, emp):
        """(province_code, district_code, commune_code) for one EMPLOYEE.

        Re-signatured from (partner) to (employee): Odoo 19 left no home
        partner to pass, and the three levels no longer live on one record.
        """
        province = (self._emp_field(emp, "vietnam_province") or "").strip()
        district = (self._emp_field(emp, "vietnam_district") or "").strip()
        commune = (self._emp_field(emp, "vietnam_ward") or "").strip()

        if not province:
            state = self._emp_field(emp, "private_state_id")
            if state:
                # A res.country.state code is an ISO-ish code ("VN-HN"), not a
                # BHXH one, so it is only usable when it is numeric; otherwise
                # the NAME is what the lookup table can match.
                code = (state.code or "").strip()
                province = code if code.isdigit() else (state.name or "")
        if not district:
            # private_city is free text and, in a VN address, is the
            # district/city line. There is no core field below it, so a
            # database without the VN pack files no commune code — a blank the
            # form allows, rather than a guess it does not.
            district = (self._emp_field(emp, "private_city") or "").strip()

        return (self._code_for("province", province),
                self._code_for("district", district),
                self._code_for("commune", commune))

    def _bank_details(self, emp):
        """(account_number, account_holder, bank_code) as Pay & Deliver sees it.

        Resolution is pb_pay_delivery's, deliberately: the same three
        vietnam_bank_* columns, and the same pb.bank.registry.match() to turn a
        typed bank name into a code. A filing that named a different account
        from the one payroll actually pays into would be a reconciliation
        problem nobody would find until the money moved.
        """
        number = (self._emp_field(emp, "vietnam_bank_account_number") or "").strip()
        holder = (self._emp_field(emp, "vietnam_bank_account_name") or "").strip()
        raw_bank = (self._emp_field(emp, "vietnam_bank_name") or "").strip()

        code = ""
        if raw_bank and "pb.bank.registry" in self.env.registry.models:
            matched = self.env["pb.bank.registry"].match(raw_bank)
            if matched:
                code = (matched.swift_prefix or matched.short_name or "").strip()
        code = code or raw_bank

        if not number:
            # No VN pack, or nothing filled in there: Odoo 19's computed scalar
            # replacement for the deleted bank_account_id. It honours
            # salary_distribution ordering, so it is the account that would be
            # paid first.
            account = self._emp_field(emp, "primary_bank_account_id")
            if account:
                number = (account.acc_number or "").strip()
                holder = holder or (account.acc_holder_name or "").strip()
                code = code or (account.bank_bic or "").strip()

        return number, holder, code

    def _copy_template(self, workbook, template_filename, sheet_names):
        """
        Load an Excel template with openpyxl and copy values into the xlsxwriter workbook.
        Styles are not preserved (xlsxwriter limitation), but layout/labels/row heights/column widths
        are copied where possible.
        Returns a dict {sheet_name: worksheet}.
        """
        import logging

        _logger = logging.getLogger(__name__)

        try:
            import openpyxl
        except ImportError:
            _logger.warning("pb_hr_govt: openpyxl not installed; template %s will be blank", template_filename)
            return {name: workbook.add_worksheet(name) for name in sheet_names}

        path = file_path(f"pb_hr_govt/government/{template_filename}")
        ws_map = {}
        # Prefer xlrd for legacy .xls to avoid noisy openpyxl warnings.
        tpl = None
        if template_filename.lower().endswith(".xls"):
            try:
                import xlrd
            except Exception:
                _logger.error("pb_hr_govt: xlrd not installed; template %s will be blank", template_filename)
                return {name: workbook.add_worksheet(name) for name in sheet_names}
            try:
                tpl = xlrd.open_workbook(path, formatting_info=True)
                _logger.info("pb_hr_govt: loaded template %s with xlrd", template_filename)
            except Exception as exc_xlrd:
                _logger.error("pb_hr_govt: xlrd failed to load %s (%s); template will be blank", template_filename, exc_xlrd)
                return {name: workbook.add_worksheet(name) for name in sheet_names}
        else:
            try:
                tpl = openpyxl.load_workbook(path, data_only=False)
                _logger.info("pb_hr_govt: loaded template %s with openpyxl", template_filename)
            except Exception as exc:
                _logger.warning("pb_hr_govt: openpyxl failed to load %s (%s); trying xlrd", template_filename, exc)
                try:
                    import xlrd
                except Exception:
                    _logger.error("pb_hr_govt: xlrd not installed; template %s will be blank", template_filename)
                    return {name: workbook.add_worksheet(name) for name in sheet_names}
                try:
                    tpl = xlrd.open_workbook(path, formatting_info=True)
                    _logger.info("pb_hr_govt: loaded template %s with xlrd", template_filename)
                except Exception as exc_xlrd:
                    _logger.error("pb_hr_govt: xlrd failed to load %s (%s); template will be blank", template_filename, exc_xlrd)
                    return {name: workbook.add_worksheet(name) for name in sheet_names}

        for name in sheet_names:
            ws = workbook.add_worksheet(name)
            if hasattr(tpl, "sheetnames"):  # openpyxl workbook
                tpl_sheet = tpl[name] if name in tpl.sheetnames else tpl.active
                for idx, dim in tpl_sheet.column_dimensions.items():
                    if dim.width:
                        col_idx = openpyxl.utils.column_index_from_string(idx) - 1
                        ws.set_column(col_idx, col_idx, dim.width)
                for ridx, dim in tpl_sheet.row_dimensions.items():
                    if dim.height:
                        ws.set_row(ridx - 1, dim.height)
                for row in tpl_sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            ws.write(cell.row - 1, cell.column - 1, cell.value)
            else:  # xlrd workbook (legacy .xls)
                try:
                    tpl_sheet = tpl.sheet_by_name(name)
                except Exception:
                    tpl_sheet = tpl.sheet_by_index(0)
                # column widths
                for col_idx, col in tpl_sheet.colinfo_map.items():
                    if col.width:
                        ws.set_column(col_idx, col_idx, col.width / 256)  # xlrd width units
                # row heights
                for ridx, row in tpl_sheet.rowinfo_map.items():
                    if row.height:
                        ws.set_row(ridx, row.height / 20)  # xlrd row height in twips
                # values
                for r in range(tpl_sheet.nrows):
                    for c in range(tpl_sheet.ncols):
                        val = tpl_sheet.cell_value(r, c)
                        if val != "":
                            ws.write(r, c, val)
            ws_map[name] = ws
        return ws_map

    def _select_employees(self, wizard):
        """Helper to pick employees based on wizard filters."""
        employees = wizard.employee_ids or self._get_employees_in_range(
            wizard.company_id, wizard.date_from, wizard.date_to
        )
        if wizard.department_id:
            employees = employees.filtered(lambda e: e.department_id == wizard.department_id)
        if wizard.contract_type_id:
            employees = employees.filtered(
                lambda e: e.contract_id and e.contract_id.type_id == wizard.contract_type_id
            )
        return employees
