# -*- coding: utf-8 -*-
"""`pb.vendor.export` — the three spreadsheets.

Every one of them is built from THE SAME FACADE THE SCREEN READS, never from a
second query. A spreadsheet that disagrees with the board it was exported from
is worse than no spreadsheet at all, and the way that happens is two pieces of
code asking the same question slightly differently.

The file comes back as base64 and the browser saves it without leaving the page
— the idiom `pb_records` uses. No attachment is left behind, because an export
is a copy somebody takes away, not a record.

NO SUDO ANYWHERE IN HERE. R89: a report that re-reads its own data as superuser
carries every company's rows into a company-scoped person's file, with totals
that then disagree with the screen beside it and nothing looking wrong. These
run as the person who pressed the button.
"""

import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

HEAD_FILL = '6355C7'


class PbVendorExport(models.TransientModel):
    _name = 'pb.vendor.export'
    _description = 'Vendor and access exports'

    kind = fields.Char(string='What it is')

    # ------------------------------------------------------------------ shared
    def _wb(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_(
                "This system cannot build spreadsheets at the moment. Ask an "
                "administrator to look at it."))
        return openpyxl, Alignment, Font, PatternFill, get_column_letter

    def _header(self, ws, cols, Font, PatternFill, Alignment, row=1):
        head = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor=HEAD_FILL)
        for i, label in enumerate(cols, start=1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = head
            c.fill = fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)

    def _widths(self, ws, widths, get_column_letter):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _file(self, wb, filename, rows):
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': filename,
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
            'rows': rows,
        }

    # ============================================================== the vendors
    @api.model
    def build_vendors(self):
        board = self.env['pb.vendors'].get_board(limit=None)
        openpyxl, Alignment, Font, PatternFill, get_column_letter = self._wb()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Vendors')

        cols = [_('Vendor'), _('What they do'), _('Who to ask for'),
                _('Their email'), _('Their phone'), _('Team'),
                _('Who looks after them'), _('Country'), _('Agreements'),
                _('Next one ends'), _('Where it is')]
        self._header(ws, cols, Font, PatternFill, Alignment)
        row = 2
        for r in board['rows']:
            for i, v in enumerate([
                    r['name'], r['type_label'], r['contact_name'],
                    r['contact_email'], r['contact_phone'], r['department'],
                    r['responsible'], r['country'], r['agreements'],
                    r['next_end'], r['state_label']], start=1):
                ws.cell(row=row, column=i, value=v)
            row += 1

        self._widths(ws, [30, 20, 22, 28, 16, 22, 24, 16, 12, 16, 20],
                     get_column_letter)
        ws.freeze_panes = 'A2'

        # A second sheet for the agreements themselves — the register answers
        # "who do we use", the agreements answer "what did we sign".
        ws2 = wb.create_sheet(_('Agreements'))
        cols2 = [_('Vendor'), _('What it covers'), _('Starts'), _('Ends'),
                 _('Talk about renewing on'), _('Days left'),
                 _('What it is worth'), _('Currency'), _('Where it is'),
                 _('Replaced by')]
        self._header(ws2, cols2, Font, PatternFill, Alignment)
        r2 = 2
        agreements = 0
        for r in board['rows']:
            drawer = self.env['pb.vendors'].get_vendor(r['id'])
            for a in drawer['agreements']:
                for i, v in enumerate([
                        r['name'], a['name'], a['date_start'], a['date_end'],
                        a['renewal_date'], a['days_left'], a['value'],
                        a['currency'], a['state_label'], a['renewed_by']],
                        start=1):
                    ws2.cell(row=r2, column=i, value=v)
                r2 += 1
                agreements += 1
        self._widths(ws2, [30, 34, 14, 14, 20, 12, 18, 12, 22, 26],
                     get_column_letter)
        ws2.freeze_panes = 'A2'

        return self._file(
            wb, _('Vendors and agreements %s.xlsx',
                  fields.Date.to_string(fields.Date.context_today(self))),
            len(board['rows']) + agreements)

    # ================================================================ the roles
    @api.model
    def build_roles(self):
        """One line per PERSON per role, not one line per role.

        A sheet with "Payroll approver — final" and a cell containing eleven
        names cannot be sorted, filtered or counted, which is the whole reason
        somebody exports it.
        """
        board = self.env['pb.access'].get_board()
        openpyxl, Alignment, Font, PatternFill, get_column_letter = self._wb()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Who holds what')

        cols = [_('Role'), _('Area'), _('What this lets someone do'),
                _('Person'), _('Their login'), _('Permission group')]
        self._header(ws, cols, Font, PatternFill, Alignment)
        row = 2
        for p in board['profiles']:
            if not p['holders']:
                for i, v in enumerate([
                        p['name'], p['area_label'], p['description'],
                        _('Nobody holds this yet'), '', p['group']], start=1):
                    ws.cell(row=row, column=i, value=v)
                row += 1
                continue
            for h in p['holders']:
                for i, v in enumerate([
                        p['name'], p['area_label'], p['description'],
                        h['name'], h['login'], p['group']], start=1):
                    ws.cell(row=row, column=i, value=v)
                row += 1
            if p['more']:
                ws.cell(row=row, column=4, value=_(
                    "and %s more", p['more']))
                ws.cell(row=row, column=1, value=p['name'])
                row += 1
        self._widths(ws, [34, 18, 60, 26, 28, 34], get_column_letter)
        ws.freeze_panes = 'A2'
        return self._file(
            wb, _('Who holds what %s.xlsx',
                  fields.Date.to_string(fields.Date.context_today(self))),
            row - 2)

    # ========================================================== the hand-overs
    @api.model
    def build_delegations(self):
        """The AUDIT sheet: hand-overs AND the roles board's own grants and
        removals, because they are the same table and the same question."""
        rows = self.env['pb.access.delegation'].search(
            [], order='date_start desc, id desc')
        openpyxl, Alignment, Font, PatternFill, get_column_letter = self._wb()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Access history')

        cols = [_('When'), _('How it happened'), _('Lent by'), _('Lent to'),
                _('What'), _('For how long'), _('From'), _('Until'),
                _('Where it is'), _('Why'), _('Permissions moved'),
                _('Handed over at'), _('Taken back at'),
                _('What happened at the end')]
        self._header(ws, cols, Font, PatternFill, Alignment)
        origin = dict(self.env['pb.access.delegation']
                      ._fields['origin'].selection)
        kinds = dict(self.env['pb.access.delegation']._fields['kind'].selection)
        states = dict(self.env['pb.access.delegation']
                      ._fields['state'].selection)
        row = 2
        for d in rows:
            for i, v in enumerate([
                    fields.Date.to_string(d.date_start) or '',
                    self.env._(origin.get(d.origin, d.origin or '')),
                    d.delegator_user_id.name or '',
                    d.delegate_user_id.name or '',
                    ', '.join(d.profile_ids.mapped('name')),
                    self.env._(kinds.get(d.kind, d.kind or '')),
                    fields.Date.to_string(d.date_start) or '',
                    fields.Date.to_string(d.date_end) or '',
                    self.env._(states.get(d.state, d.state or '')),
                    d.reason or '',
                    ', '.join(d.applied_group_ids.mapped('display_name')),
                    fields.Datetime.to_string(d.applied_on) or '',
                    fields.Datetime.to_string(d.ended_on) or '',
                    d.ended_note or ''], start=1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        self._widths(ws, [14, 24, 24, 24, 34, 16, 14, 14, 16, 40, 40, 20, 20,
                          34], get_column_letter)
        ws.freeze_panes = 'A2'
        return self._file(
            wb, _('Access history %s.xlsx',
                  fields.Date.to_string(fields.Date.context_today(self))),
            len(rows))
