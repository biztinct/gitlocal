# -*- coding: utf-8 -*-
import base64
import binascii
import hashlib
import html
import json
import logging
import re
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

try:
    import requests
except Exception:  # pragma: no cover
    requests = None

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, UserError
# MAPFIX A — one code generator for every path that names a column.
from odoo.addons.pb_hr_payroll_formula.models import component_code as component_code_mod
from odoo.addons.pb_hr_payroll_formula.models import value_kind_classifier

_logger = logging.getLogger(__name__)

# Config parameters (Settings > Technical > Parameters). Empty api_key => PayAI
# uses the built-in deterministic mapper. base_url can point at ANY
# OpenAI-compatible endpoint (OpenAI, Ollama, vLLM, LM Studio, OpenRouter…).
LLM_BASE_URL = 'pb_formula_studio.llm_base_url'
LLM_API_KEY = 'pb_formula_studio.llm_api_key'
LLM_MODEL = 'pb_formula_studio.llm_model'
DEFAULT_BASE_URL = 'https://api.openai.com/v1'
DEFAULT_MODEL = 'gpt-4o-mini'


class LLMUnavailable(Exception):
    """Raised by _llm_chat when the LLM cannot be reached — no API key, no
    `requests`, timeout, non-200, or invalid JSON. Callers catch this to fall
    back to deterministic behaviour."""


# Map an Excel operator to a friendly chip glyph
OP_GLYPH = {'+': '+', '-': '−', '*': '×', '/': '÷'}

# W48 — number guard for LLM-polished narration (D-D2). Pure + unit-testable:
# every money-scale figure (>= _NARR_MONEY_FLOOR) in the rewritten text must
# exist in the compare fold's allowed set, else the rewrite invented a number
# and is rejected in favour of the deterministic text. Counts/dates/percentages
# (small numbers) are not policed — they are low-risk and hard to enumerate.
_NARR_MONEY_FLOOR = 1000
_NARR_NUM_RE = re.compile(r'-?\d[\d,]*(?:\.\d+)?')


def _narr_numbers_ok(text, allowed):
    """True unless the text contains a money-scale number absent from `allowed`
    (a set of rounded-abs integers). `allowed` should already hold every sum /
    delta / count from the fold."""
    for raw in _NARR_NUM_RE.findall(text or ''):
        s = raw.replace(',', '')
        try:
            n = round(abs(float(s)))
        except Exception:
            continue
        if n < _NARR_MONEY_FLOOR:
            continue                      # counts / dates / small pcts: not policed
        if n not in allowed:
            return False                  # an invented money figure
    return True


# Category grouping for the outline (by code/name heuristics)
_NET_CODES = {'NET', 'NETPAY', 'NET_PAY', 'NETSALARY', 'TAKEHOME', 'TAKE_HOME'}

# COLROLES P2 — the fifth outline group. Everything the classifier decided is NOT
# a pay component (the employee code, the bank account, the join date, a cost
# centre reference) leaves the four payroll buckets and collects here, so the
# outline reads as "the payroll" plus "the people data it travels with".
PEOPLE_GROUP = 'People & Data'

#: RD50 — employer contributions are their OWN section, which is what every
#: serious payroll product does: they are a cost to the company, not money taken
#: off this person's pay, and netting them against pay is a misstatement. The
#: payslip already prints them under their own heading; the studio now agrees
#: with it instead of filing them among the employee's deductions.
EMPLOYER_GROUP = 'Employer contributions'


#: RD47 — `net_role` -> the studio's display group. The classifier already
#: decided whether a component is added to net pay, taken off it, or neither;
#: this is a translation of that answer, not a second opinion about it.
_NET_ROLE_GROUP = {
    'earning': 'Earnings',
    'deduction': 'Deductions',
    'net': 'Totals',
    # RD50 — the fifth bucket, on the owner's ruling to follow what the market
    # does. An employer contribution is money going out but NOT off this
    # person's net pay, so it belongs beside the payslip's own "Employer
    # contributions" heading rather than among the employee's deductions.
    'employer_cost': EMPLOYER_GROUP,
    # 'info' stays ABSENT on purpose: a component that is neither added nor
    # subtracted (a working figure like Taxable Income) has no natural column,
    # and the name lexicon is a better guess at where a reader will look for it
    # than "Earnings" would be. What matters — that it is not drawn as a
    # negative — is settled by `showsMinus`, not by the column.
}


def _group_for(rule):
    code = (rule.code or '').upper()
    name = (rule.name or '').lower()
    # The role wins over every name heuristic: a column called "Tax code" is a
    # reference even though "TAX" says Deductions to the lexicon below.
    # getattr keeps this safe on a database whose formula module is older.
    if (getattr(rule, 'column_role', False) or 'payroll') != 'payroll':
        return PEOPLE_GROUP
    if rule.column_type == 'input':
        return 'Inputs'
    # RD47 — ASK THE CLASSIFIER BEFORE GUESSING FROM THE NAME.
    #
    # The lexicon below is SUBSTRING matching, and this codebase has paid for
    # that before: NETROLE replaced `_get_default_category`'s 'SI'/'TAX'-in-code
    # test after it invented ₫5.06bn of phantom deductions. The same test
    # survived here, where it decides which side of the Live Preview a number
    # sits on and whether it is drawn with a minus in front of it — so
    # `ACTUBASISALA` ("Actual Basic salary") read as a DEDUCTION, because
    # "ba-SI-sala" contains SI, and the panel showed a real earning of
    # ₫9,937,500 as −₫9,937,500. `TAXABLEINCOM` and the SI-HI-UI constants went
    # the same way.
    #
    # `net_role` is the sign-propagation classifier's own verdict — it had
    # `ACTUBASISALA` right as `earning` the whole time. It is asked first, and
    # the lexicon is left as the fallback for a component the classifier has not
    # reached (`info`, or a database whose formula module predates the field).
    role = getattr(rule, 'net_role', False)
    if role in _NET_ROLE_GROUP:
        return _NET_ROLE_GROUP[role]
    if any(k in code or k in name for k in ('NET', 'GROSS', 'TOTAL', 'thực nhận', 'tổng')):
        return 'Totals'
    if any(k in code or k in name for k in ('SI', 'HI', 'UI', 'PIT', 'TAX', 'DED', 'insurance', 'deduction', 'bảo hiểm', 'thuế')):
        return 'Deductions'
    return 'Earnings'


class PbFormulaStudio(models.AbstractModel):
    _name = 'pb.formula.studio'
    _description = 'Payobook Formula Studio data layer (wraps the formula engine)'

    # ------------------------------------------------------------------
    # config selection / list
    # ------------------------------------------------------------------
    # validation_status (stored) → a coarse health score for the switcher rings,
    # so the config gallery stays a single cheap query (no per-config recompute).
    # A real validation verdict wins; when it's still 'pending' we fall back on the
    # lifecycle state so an active, working config doesn't read as 0/unhealthy.
    _VSTATUS_SCORE = {'passed': 96, 'warning': 70, 'failed': 34}
    _STATE_SCORE = {'active': 90, 'validated': 84, 'testing': 55, 'draft': 22, 'archived': 12}

    @api.model
    def _config_score(self, config):
        return (self._VSTATUS_SCORE.get(config.validation_status)
                or self._STATE_SCORE.get(config.state, 0))

    @api.model
    def get_config_list(self):
        configs = self.env['hr.formula.config'].search([], order='sequence, id desc')
        out = []
        for c in configs:
            out.append({
                'id': c.id,
                'name': c.name,
                'code': c.code or '',
                'country': c.country_code or '',
                'state': c.state,
                'rule_count': len(c.rule_ids),
                # --- richer fields for the Config Switcher gallery ---
                'currency': c.currency_id.name or '',
                'cycle_type': c.cycle_type or 'regular',
                'active': bool(c.active),
                'validation_status': c.validation_status or 'pending',
                'score': self._config_score(c),
                'sample_count': len(c.sample_data_ids),
                'is_branch': bool(c.parent_branch_id),
                'is_variant': bool(c.master_config_id),
                'is_master': bool(c.variant_ids),
                'updated': fields.Date.to_string(c.write_date) if c.write_date else '',
            })
        return out

    @api.model
    def _pick_config(self, config_id=None):
        Config = self.env['hr.formula.config']
        if config_id:
            c = Config.browse(int(config_id))
            if c.exists():
                return c
        # prefer an active config with rules, else newest with rules, else newest.
        # Order by sequence FIRST so a "featured" config (low sequence — e.g. the
        # demo's Retail division) lands by default instead of whatever has the
        # highest id (a 250-column scale-test would otherwise win). Ties fall back
        # to newest id — bit-identical to the old behaviour when no sequence is set.
        c = Config.search([('state', '=', 'active'), ('rule_ids', '!=', False)],
                          order='sequence, id desc', limit=1)
        if c:
            return c
        c = Config.search([('rule_ids', '!=', False)], order='sequence, id desc', limit=1)
        return c or Config.search([], order='sequence, id desc', limit=1)

    # ------------------------------------------------------------------
    # formula tokenizing (for the friendly chip view + plain-language)
    # ------------------------------------------------------------------
    @api.model
    def _col_to_rule(self, rules):
        return {r.column_letter: r for r in rules if r.column_letter}

    @api.model
    def _col_num(self, col):
        n = 0
        for ch in (col or '').upper():
            v = ord(ch) - 64
            if v < 1 or v > 26:
                return 0
            n = n * 26 + v
        return n

    @api.model
    def _num_to_col(self, n):
        """Inverse of _col_num: 1 -> 'A', 27 -> 'AA'."""
        s = ''
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    @api.model
    def _expand_refs(self, formula, by_col):
        """Set of referenced columns in a formula, expanding A#:B# ranges to
        every existing member column (matches the engine's range expansion).
        String literals are masked first so '=IF(D2="X2",…)' never reports a
        phantom X reference (WP-L review: stage_paste false-rejected such
        formulas as 'Unknown column(s): X')."""
        f = re.sub(r'"[^"]*"', ' ', formula or '')
        out = set()
        for s, e in re.findall(r'([A-Za-z]+)\d+:([A-Za-z]+)\d+', f):
            lo, hi = sorted((self._col_num(s), self._col_num(e)))
            for col in by_col:
                if lo <= self._col_num(col) <= hi:
                    out.add(col)
        # plain refs (blank out ranges so endpoints aren't missed nor double-handled)
        rest = re.sub(r'([A-Za-z]+)\d+:([A-Za-z]+)\d+', ' ', f)
        for c in re.findall(r'([A-Za-z]+)\d+', rest):
            out.add(c.upper())
        return out

    @api.model
    def _tokenize(self, rule, by_col):
        """Turn '=A1*0.2' into chip tokens referencing component names."""
        if rule.column_type == 'input':
            return [{'kind': 'src', 'text': 'From contract / import'}]
        if rule.column_type == 'constant':
            val = rule.constant_value or 0.0
            return [{'kind': 'num', 'text': '{:,.0f}'.format(val)}]
        formula = (rule.excel_formula or '').lstrip('=').strip()
        if not formula:
            return [{'kind': 'src', 'text': 'No formula yet'}]
        tokens = []
        # split keeping operators and parens
        parts = re.findall(r'[A-Za-z]+\d+|\d+\.?\d*|[+\-*/()%]', formula)
        for p in parts:
            m = re.match(r'^([A-Za-z]+)\d+$', p)
            if m:
                col = m.group(1).upper()
                ref = by_col.get(col)
                tokens.append({'kind': 'ref', 'col': col,
                               'text': ref.name if ref else col})
            elif p in OP_GLYPH:
                tokens.append({'kind': 'op', 'text': OP_GLYPH[p]})
            elif p in ('(', ')', '%'):
                tokens.append({'kind': 'op', 'text': p})
            else:
                tokens.append({'kind': 'num', 'text': p})
        return tokens

    @api.model
    def _explain(self, rule, by_col):
        if rule.column_type == 'input':
            return "Comes from each employee's contract or the monthly import."
        if rule.column_type == 'constant':
            return "A fixed value applied to every employee."
        toks = self._tokenize(rule, by_col)
        refs = [t for t in toks if t['kind'] == 'ref']
        ops = [t for t in toks if t['kind'] == 'op']
        has_paren = any(t['text'] in ('(', ')') for t in ops)
        # For simple formulas, build a readable sentence; for complex ones,
        # summarise by the components it draws from (the chip view shows the rest).
        if has_paren or len(refs) > 3:
            names = []
            for t in refs:
                if t['text'] not in names:
                    names.append(t['text'])
            if names:
                return rule.name + ' is calculated from ' + ', '.join(names[:6]) + \
                    ('and others.' if len(names) > 6 else '.')
            return rule.name + ' is a calculated component.'
        words = []
        for t in toks:
            if t['kind'] == 'ref':
                words.append(t['text'])
            elif t['kind'] == 'op':
                words.append({'+': 'plus', '−': 'minus', '×': 'times', '÷': 'divided by'}.get(t['text'], t['text']))
            elif t['kind'] == 'num':
                words.append(t['text'])
        return rule.name + ' is computed as ' + ' '.join(words) + '.'

    # ------------------------------------------------------------------
    # main payload
    # ------------------------------------------------------------------
    # ==================================================================
    # SOURCING S4 — where a value comes from, as the UI reads it.
    #
    # TWO TRUTHS, NEVER CONFLATED. `declared` is what the configuration says;
    # `actual` is what the last run actually did and the key it actually matched.
    # A component whose two disagree is precisely the situation the owner needs to
    # see, so nothing here smooths that over.
    #
    # `data_source` is NOT consulted anywhere in this block. It defaults to
    # 'excel', so "unknown" and "Excel" are indistinguishable in it; no wire-
    # creation path maintains it; its vocabulary does not match the resolver's; and
    # no line of the payroll pipeline reads it. It stays demoted (owner ruling O-3).
    # ==================================================================
    #: The source kinds (ten since J10), and whether a mapping board may draw
    #: a wire to one. The record destinations are drawn on the Employee &
    #: contract board, not onto a scheme card, so they are not wirable here.
    _SOURCE_WIRABLE = {'excel', 'feed', 'rule'}

    @api.model
    def _source_record_dests(self, config):
        """Where each component lands on the employee/contract/bank record.

        One query for the whole config. `hr.payslip.import.mapping` is the
        employee-field board's model — NOT `hr.integration.field.mapping`, which is
        the vendor feed's. The two were conflated once already (see the resolver's
        own note at `payroll_import_batch.py:2648-2661`) and the distinction is
        exactly what tells "Employee record" from "Connected system".

        JOURNEY J10 — THIS RETURNED A BARE `set()` OF RULE IDS, and that is why
        the owner's screenshot showed DESIGNATION as "Connected system" and
        nothing else. A set can answer "does this component write to a record"
        and nothing more: neither WHICH record nor WHICH FIELD was available to
        render, so the tier could only ever be a single unnamed chip — and
        `_declared_sources` could only reach it when nothing else was declared.
        It returns `{rule_id: {'kind', 'key', 'label'}}` now:

          `destination_type='bank_account'`  → `bank_account`, key = bank role
          `field` on `hr.contract`           → `contract_field`
          `field` on `hr.employee`           → `employee_field`

        `key` is the TECHNICAL name because that is what the `(kind, key)` fold
        compares; `label` is what a reader sees. `in` works on a dict as it did
        on a set, which is exactly why every call site was changed in one go
        rather than found later by a failure.

        ONE SQL statement, deliberately. The ORM path is four constant-cost
        reads (mappings, `ir_model`, `ir_model_fields`, then the translations),
        which is fine, but this is a chip on a ninety-nine card board and the
        join is trivial. `field_description` is a translated column and is read
        as jsonb with a plain-text fallback, so the same code serves a database
        where it is not.
        """
        Mapping = self.env.get('hr.payslip.import.mapping')
        if Mapping is None or not config.rule_ids:
            return {}
        lang = self.env.lang or self.env.user.lang or 'en_US'
        try:
            self.env.cr.execute("""
                SELECT m.component_id, m.destination_type, m.bank_role,
                       im.model, imf.name, imf.field_description::text
                  FROM hr_payslip_import_mapping m
             LEFT JOIN ir_model im ON im.id = m.target_model_id
             LEFT JOIN ir_model_fields imf ON imf.id = m.target_field_id
                 WHERE m.component_id IN %s
              ORDER BY m.id ASC
            """, (tuple(config.rule_ids.ids),))
            rows = self.env.cr.fetchall()
        except Exception:       # noqa: BLE001 — a chip must never break the studio
            return {}
        bank_labels = dict(Mapping._fields['bank_role'].selection)
        out = {}
        for rule_id, dest, bank_role, model, field_name, description in rows:
            if not rule_id or rule_id in out:
                continue        # lowest id wins, as `_get_bank_mappings` does
            if dest == 'bank_account':
                if not bank_role:
                    continue
                out[rule_id] = {'kind': 'bank_account', 'key': bank_role,
                                'label': bank_labels.get(bank_role) or bank_role}
                continue
            kind = ('contract_field' if model == 'hr.contract'
                    else 'employee_field' if model == 'hr.employee' else None)
            if not kind or not field_name:
                continue
            out[rule_id] = {'kind': kind, 'key': field_name,
                            'label': self._field_label(description, field_name,
                                                       lang)}
        return out

    @api.model
    def _field_label(self, description, fallback, lang='en_US'):
        """`ir_model_fields.field_description` read out of a jsonb column."""
        if not description:
            return fallback
        try:
            blob = json.loads(description)
        except (TypeError, ValueError):
            return description or fallback
        if not isinstance(blob, dict):
            return str(blob) or fallback
        return blob.get(lang) or blob.get('en_US') or fallback

    # ==================================================================
    # JOURNEY J3 S2 — two live sources on one component, detected ONCE.
    # ==================================================================
    @api.model
    def _source_conflicts(self, config):
        """Every component of `config` that is read by TWO live sources.

        J-D3. A component can hold a spreadsheet binding (`source_binding='excel'`)
        while a live `hr.integration.field.mapping` wire targets it, and until this
        phase nothing anywhere said so. On a system run the connector pre-pass fills
        the value first and the resolver's `if rule.code not in input_values` skip
        locks out the binding entirely — so the spreadsheet column a person had
        deliberately chosen was simply never read, silently, for as long as both
        rows existed. The same is true of a component wired on two different
        connections: only the connection the SCHEME is set to is consulted.

        ONE detector, read by both mapping boards' payloads and by the draw-time
        probe, so the chip a user sees and the dialog they are shown can never
        disagree about whether a conflict exists.

        Returns `{rule_id: {...}}` with, per entry:
          `binding_kind`/`binding_key` — the component's declared binding, if any;
          `wires` — `[{'connector': rec, 'key': str, 'kind': 'feed'|'rule'}]`;
          `shape` — `'excel_vs_feed'` or `'two_feeds'`;
          `primary`/`fallback` — which source a system run actually uses, and which
          one only speaks when the first is empty. **Never the reverse** (J-D5): the
          feed outranks the binding in the resolver and this phase does not reorder
          it, so any wording built from this must say the feed is primary.

        Read-only and guarded: a chip must never break a board.
        """
        FM = self.env.get('hr.integration.field.mapping')
        if FM is None or not config or not config.rule_ids:
            return {}
        try:
            maps = FM.sudo().search([('target_rule_id', 'in', config.rule_ids.ids)],
                                    order='id')
        except Exception:       # noqa: BLE001
            return {}
        by_rule = {}
        computed = {}
        for m in maps:
            rid = m.target_rule_id.id
            if not rid or not m.source_field:
                continue
            cid = m.connector_id.id
            if cid not in computed:
                try:
                    computed[cid] = FM._computed_output_keys(m.connector_id)
                except Exception:   # noqa: BLE001
                    computed[cid] = set()
            by_rule.setdefault(rid, []).append({
                'connector': m.connector_id, 'key': m.source_field, 'mapping': m,
                'kind': 'rule' if m.source_field in computed[cid] else 'feed',
            })
        out = {}
        for rule in config.rule_ids:
            wires = by_rule.get(rule.id) or []
            # JOURNEY J9 — asked of the SOURCE ROWS, not of `source_binding`.
            # The derived field reports only the HIGHEST-RANKED source, so the
            # moment a component declares a feed beside its spreadsheet column
            # it reads `feed` and this detector would stop seeing the very state
            # it was written to see. The predicate is unchanged for every
            # component that has one binding, which is all of them on all four
            # live databases; it simply stopped depending on rank.
            excel_row = rule.source_ids.filtered(
                lambda s: s.kind == 'excel' and (s.key or '').strip())
            b_kind = 'excel' if excel_row else False
            b_key = (excel_row[0].key or '').strip() if excel_row else ''
            if b_kind and not b_key:
                b_kind = False      # a half-set binding is not a source (S3)
            shape = None
            if b_kind == 'excel' and wires:
                shape = 'excel_vs_feed'
            elif len({w['connector'].id for w in wires}) > 1:
                shape = 'two_feeds'
            if not shape:
                continue
            out[rule.id] = {
                'binding_kind': b_kind, 'binding_key': b_key, 'wires': wires,
                'shape': shape,
                'primary': 'feed', 'fallback': 'excel' if b_kind == 'excel' else 'feed',
            }
        return out

    @api.model
    def _conflict_chip(self, conflict, board):
        """The chip a conflicted component wears, worded for the board you are on.

        Two sentences, because the useful one is always about the OTHER source: on
        the spreadsheet board you already know about the column, and what you are
        missing is that a feed will beat it. The wording is fixed by the ladder and
        not by preference — the feed is primary on system runs, the spreadsheet is
        the fallback, and saying it the other way round would be a comfortable lie.
        """
        if not conflict:
            return None
        if conflict['shape'] == 'two_feeds':
            names = [w['connector'].name or _("Unnamed connection")
                     for w in conflict['wires']]
            uniq = list(dict.fromkeys(names))
            return {
                # The PILL says the surprising thing in as few words as will fit
                # beside the source chip on a 280px card; the full sentence is the
                # tooltip. Live on abm this label was "Wired to 2 connections" and
                # it clipped mid-word — a chip that cannot finish its own sentence
                # is worse than a short one.
                'label': _("Wired twice"),
                'hint': _(
                    "This component is wired on %(names)s. A pay run reads only "
                    "the connection its scheme is set to; the others are ignored.",
                    names=', '.join(uniq)),
            }
        conn = conflict['wires'][0]['connector'].name or _("the connected system")
        if board == 'import':
            return {
                'label': _("Feed wins"),
                'hint': _(
                    "Also wired to %(conn)s. The feed wins on system runs — this "
                    "spreadsheet column is read only when %(conn)s sends nothing "
                    "for it.", conn=conn),
            }
        return {
            'label': _("Spreadsheet fallback"),
            'hint': _(
                "Also bound to the spreadsheet column “%(key)s”. The feed wins on "
                "system runs — that column is read only when this feed sends "
                "nothing for it.", key=conflict['binding_key']),
        }

    @api.model
    def _source_wire_dests(self, config):
        """Components this config already has a LIVE connector wire into.

        SOURCING S6, D3 fix A. `_declared_source` used to read the S3 binding and
        the component's own nature and nothing else — so the most explicit
        statement of source these databases actually contain, a drawn field
        mapping, said nothing on any screen. On abm that silenced EIGHT components
        fed by transformation-rule outputs (`OTHRS150` → `OT15HOURS`, `DEPCOUNT` →
        `NOOFDEPENDEN`, `WORKEDHRS` → `ACTUWORKHOUR`, …): every one of them rendered
        "No source chosen" while a wire had been drawn to it, which is the "where is
        the transformation indication" the owner asked about.

        Returns `{rule_id: {'kind': 'rule'|'feed', 'key': source_field,
                            'connector': <connector record>}}`.

        ONE search for the whole config plus one per connector for its rules — the
        S4 precedent (one payslip read, not one read per component). A per-component
        lookup here would be ninety-nine queries to answer one question.

        **Determinism matters more than it looks.** Seven of abm's components are
        wired on BOTH connectors, so a first-row-wins rule would make the chip
        depend on row order. A `rule` wire beats a `feed` wire — a rule output is
        the more specific fact, and it is the one with lineage behind it — and ties
        break on the lowest mapping id.

        Read-only, and guarded: a chip must never break the studio.
        """
        FM = self.env.get('hr.integration.field.mapping')
        if FM is None or not config.rule_ids:
            return {}
        try:
            maps = FM.sudo().search([('target_rule_id', 'in', config.rule_ids.ids)],
                                    order='id')
        except Exception:       # noqa: BLE001
            return {}
        if not maps:
            return {}
        computed = {}
        for connector in maps.mapped('connector_id'):
            try:
                computed[connector.id] = FM._computed_output_keys(connector)
            except Exception:   # noqa: BLE001
                computed[connector.id] = set()
        out = {}
        for m in maps:
            rid = m.target_rule_id.id
            if not rid or not m.source_field:
                continue
            kind = ('rule' if m.source_field in computed.get(m.connector_id.id, set())
                    else 'feed')
            prev = out.get(rid)
            # a rule wire beats a feed wire; otherwise the first (lowest id) stands
            if prev and not (kind == 'rule' and prev['kind'] != 'rule'):
                continue
            out[rid] = {'kind': kind, 'key': m.source_field,
                        'connector': m.connector_id}
        return out

    @api.model
    def _source_actuals(self, config):
        """What the LAST run recorded, indexed by component code.

        Reads exactly ONE payslip — the most recent formula payslip of this config
        that carries a provenance blob — and indexes it. The alternative, asking per
        component, is 250 queries to answer one question.

        Returns `({CODE: entry}, run_label)`. An empty dict means this scheme has
        never been run, which the UI must say in those words: it is a different
        statement from "this component has no source", and collapsing the two would
        put "No source" on a component that is simply waiting for its first run.
        """
        Payslip = self.env.get('hr.payslip')
        if Payslip is None:
            return {}, ''
        try:
            slip = Payslip.sudo().search([
                ('formula_config_id', '=', config.id),
                ('formula_input_sources', '!=', False),
            ], order='date_to desc, id desc', limit=1)
        except Exception:       # noqa: BLE001
            return {}, ''
        if not slip:
            return {}, ''
        try:
            blob = json.loads(slip.formula_input_sources or '{}')
        except (TypeError, ValueError):
            return {}, ''
        if not isinstance(blob, dict):
            return {}, ''
        return blob, (slip.date_to and slip.date_to.strftime('%B %Y')) or slip.name or ''

    @api.model
    def _declared_source(self, rule, record_dests, wire_dests=None):
        """What configuration SAYS feeds this component.

        Order matters, and every tier above the last is something a PERSON stated:

          1. a calculated or fixed column is that whatever else is set on it;
          2. the S3 binding — the explicit per-component declaration;
          3. **a live connector wire** (S6) — drawn on a mapping board before S3
             existed, and just as explicit. `rule` when the wire's source field is
             an output key of that connector's transformation rules, `feed`
             otherwise;
          4. below that, a description of what the component IS rather than of
             anything anybody chose.

        `wire_dests` is `_source_wire_dests(config)`, computed once for the whole
        config. `None` means the caller did not compute it, and tier 3 is then
        SKIPPED rather than answered with ninety-nine queries — the failure mode of
        a forgetful caller is today's behaviour, never a slow one.

        JOURNEY J9 — this is now the FIRST of `_declared_sources`, and its return
        shape `{'kind', 'key', 'wirable'}` is unchanged. It is deliberately not
        deprecated: four callers want the winning source and nothing else, and
        handing them a list to index would be a change with no reader.
        """
        return self._declared_sources(rule, record_dests, wire_dests)[0]

    #: JOURNEY J9 — the boards' copy of the resolver's order, and it is a COPY of
    #: nothing: `hr.formula.rule._SOURCE_RANK` is the definition and this reads
    #: it, so a board can never sort by an order the resolver does not use.
    @api.model
    def _source_rank(self):
        return self.env['hr.formula.rule']._SOURCE_RANK

    @api.model
    def _declared_sources(self, rule, record_dests, wire_dests=None):
        """EVERY source this component declares, in the order a run reads them.

        JOURNEY J9. The owner removed the either/or restriction, so this returns
        a LIST where `_declared_source` returned a scalar — and `_declared_source`
        is now `list[0]`, so its four callers (the right column, the source note,
        the transform board and the Journey lane bucketing) are unchanged.

        Each entry is `{'kind', 'key', 'wirable'}` — the same three keys the
        scalar has always carried, deliberately, so nothing downstream has to
        learn a new shape.

        THE (kind, key) FOLD IS LOAD-BEARING, and abm is the fixture that proves
        it. `api_mapping_create` writes a field mapping AND an S3 binding in one
        gesture, so all nine of abm's feed-bound components carry a wire whose
        `source_field` is character-for-character the binding key. They are one
        source recorded twice, and without this fold every one of them would
        render two identical "Connected system" chips — which the canvas' label
        dedupe would then HIDE, leaving the board looking right while the reader
        was told the wrong thing about how many sources a component has.

        Tiers, unchanged in kind and only widened in arity:
          0. a calculated or fixed column is that, whatever else is set on it;
          1. the declared sources (`source_ids`), ranked;
          2. a live connector wire, folded in when it is not already one of them;
          3. **the mapped record destination** (J10) — employee field, contract
             field or bank account, at rank 4, where the resolver's tail has
             always read it;
          4. the contract component, always last among things a person stated;
          5. below that, a description of what the component IS.

        `record_dests` is `_source_record_dests(config)`, computed once for the
        whole config. It was a bare set of rule ids until J10 and is a dict now;
        `in` works on both, which is why every call site was changed together —
        a missed one would have failed SILENTLY rather than loudly.
        """
        if rule.column_type == 'formula':
            return [{'kind': 'calculated', 'key': '', 'wirable': False}]
        if rule.column_type == 'constant':
            return [{'kind': 'constant', 'key': '', 'wirable': False}]
        rank = self._source_rank()
        out, seen = [], set()
        for spec in rule.declared_sources():
            if spec['kind'] not in rank:
                continue
            pair = (spec['kind'], spec['key'])
            if pair in seen:
                continue
            seen.add(pair)
            out.append({'kind': spec['kind'], 'key': spec['key'],
                        'wirable': True})
        wired = (wire_dests or {}).get(rule.id)
        if wired:
            pair = (wired['kind'], (wired['key'] or '').strip())
            if pair not in seen:
                seen.add(pair)
                out.append({'kind': pair[0], 'key': pair[1], 'wirable': True})
        # ==============================================================
        # JOURNEY J10 — THE RECORD IS A SOURCE TOO, AND IT IS RANK 4.
        #
        # This block used to sit BELOW an `if out: return out`, which made the
        # record tier reachable only when nothing else was declared — the
        # owner's exact bug report: *"Currently you are showing EMPLOYEE RECORD
        # or CONTRACT RECORD only if that is the only source."* Ten of abm's
        # twenty-one mappings sat on a component that already declared
        # something, so ten cards were silently hiding half of what they do.
        #
        # The contract component two lines below has been APPENDED
        # unconditionally since J9, and that is precisely the treatment the
        # owner asked for here — so the early return goes and this joins the
        # ranked list. NOTHING MOVED (J-D5): rank 4 is where the resolver's tail
        # has always read it, after the spreadsheet and before the contract
        # component (`payroll_import_batch.get_mapped_input_value`).
        #
        # `employee_field`, `contract_field` and `bank_account` are ONE RUNG,
        # not three: a component carries at most one `hr.payslip.import.mapping`
        # row, so they never compete. The `(kind, key)` fold still applies —
        # nothing else can produce these kinds today, but a database that grew
        # a second way to say it must not render the same fact twice.
        # ==============================================================
        dest = (record_dests or {}).get(rule.id)
        if dest:
            pair = (dest['kind'], (dest['key'] or '').strip())
            if pair not in seen:
                seen.add(pair)
                out.append({'kind': pair[0], 'key': pair[1], 'wirable': False,
                            'label': dest.get('label') or pair[1]})
        out.sort(key=lambda d: rank.index(d['kind']))
        if rule.is_contract_component:
            out.append({'kind': 'contract_component', 'key': '',
                        'wirable': False})
        if out:
            return out
        return [{'kind': 'none', 'key': '', 'wirable': True}]

    #: Board-chip wording. Kept next to the vocabulary it uses so a board can
    #: never invent a ninth term. Mirrors `srcLabel` in `source_vocab.js`.
    #: JOURNEY J10 — ten terms now, and the two new ones are SPLIT OUT of one
    #: that was doing two jobs. "Employee record" was shown for a mapping onto
    #: `hr.contract` as readily as one onto `hr.employee`, because the tier
    #: below could not tell them apart (it was a set of rule ids). A component
    #: whose designation lands on the CONTRACT now says so.
    _SOURCE_LABELS = {
        'excel': "Spreadsheet", 'feed': "Connected system", 'rule': "Rule output",
        'contract_component': "Contract component", 'employee_field': "Employee record",
        'contract_field': "Contract record", 'bank_account': "Bank account",
        'calculated': "Calculated", 'constant': "Fixed value", 'none': "No source",
    }

    @api.model
    def _source_label(self, kind):
        """The one translated label for a source kind.

        `_SOURCE_LABELS` above is the untranslated register — it exists so a board
        can never invent a term of its own, and it is read where the string is a
        key rather than a sentence. This is the TRANSLATED reading of the same ten
        words, with every literal written out so gettext can find it: `_(variable)`
        extracts nothing and ships English forever (S19's family — a translation
        that fails silently at the point of use).
        """
        return {
            'excel': _("Spreadsheet"),
            'feed': _("Connected system"),
            'rule': _("Rule output"),
            'contract_component': _("Contract component"),
            'employee_field': _("Employee record"),
            'contract_field': _("Contract record"),
            'bank_account': _("Bank account"),
            'calculated': _("Calculated"),
            'constant': _("Fixed value"),
            'none': _("No source"),
        }.get(kind or 'none', _("No source"))

    #: SOURCING S5 — cards a board shows but will not let you wire.
    #: `column_type` is the authority: a formula or a constant is PRODUCED by this
    #: scheme, not imported into it, so there is no source to attach. They used to
    #: be filtered out silently, which left the reader wondering where their
    #: component had gone; now they are visible, badged and sealed.
    #: SOURCING S5 — a sealed card is refused by the SERVER, not merely
    #: un-offered by the board. MAPFIX F3's precedent: hiding an affordance is not
    #: a gate. A stale bundle, a tampered client or a direct RPC must all get the
    #: same answer as the board would give.
    @api.model
    def _lineage_by_output_key(self, connector, config):
        """How each computed key is worked out, keyed by the key itself.

        `reads` is the UNCAPPED list S2 split out of `_trace_cells` — the cap of
        four was a property of a narrow proof rail, not of the rule, and lineage
        has to know all of them.

        `feeds` is what S2's repair made answerable: every abm rule output now has
        exactly one live consumer, where before the programme it had none and the
        board could not have said so.

        The rule NAME goes into the reads/feeds text and the card's own label, not
        into a tooltip: `itemMatches` searches label/sublabel/sample/meta.col/group,
        so tooltip-only text is unsearchable on a board with hundreds of cards.
        """
        Rule = self.env.get('hr.api.transformation.rule')
        if Rule is None or not connector:
            return {}
        try:
            rules = Rule.sudo().search([('connector_id', '=', connector.id)])
        except Exception:       # noqa: BLE001 — lineage must never break a board
            return {}
        Mapping = self.env['hr.integration.field.mapping'].sudo()
        out = {}
        for rule in rules:
            if not rule.output_key:
                continue
            feeds = []
            for m in Mapping.search([('connector_id', '=', connector.id),
                                     ('source_field', '=', rule.output_key)]):
                if m.target_rule_id:
                    feeds.append('%s (%s)' % (m.target_rule_id.name or '',
                                              m.target_rule_id.code or ''))
            # Components bound to this rule output by name (S3 bindings).
            for r in config.rule_ids:
                if r.source_binding == 'rule' and \
                        (r.source_binding_key or '') == rule.output_key:
                    label = '%s (%s)' % (r.name or '', r.code or '')
                    if label not in feeds:
                        feeds.append(label)
            try:
                reads = list(rule._consumed_field_names())
            except Exception:   # noqa: BLE001
                reads = []
            out[rule.output_key] = {
                'rule_id': rule.id,
                'summary': rule.plain_summary or '',
                'reads': reads,
                'fallback': self._lineage_fallback_text(rule),
                'feeds': feeds,
            }
        return out

    @api.model
    def _lineage_fallback_text(self, rule):
        """What the rule produces when nothing matches — in words, not a field name."""
        default = getattr(rule, 'default_value', None)
        if default in (None, False, ''):
            return _("Nothing is written for that employee.")
        return _("It writes %s.") % default

    @api.model
    def _mc_refuse_sealed(self, rule):
        if rule.exists() and rule.column_type != 'input':
            return {'ok': False, 'msg': _(
                "“%s” is calculated — it needs no source. This column is produced "
                "by the scheme, not imported into it.")
                % (rule.code or rule.name or '')}
        return None

    @api.model
    def _mc_legal_output_key(self, rule):
        """A component code, made legal as a transformation-rule output key.

        The contract (S2, `OUTPUT_KEY_RE`) is capitals and digits starting with a
        letter — no underscores, because the Excel->Python converter's code pass
        excludes `_` and an underscored name survives raw into the eval, raises
        NameError and silently reads 0. MAPFIX Phase A already made component codes
        conform, so this is normally a no-op; it exists for the legacy codes that
        predate it.
        """
        key = re.sub(r'[^A-Za-z0-9]', '', (rule.code or rule.name or '')).upper()
        if not key or not key[0].isalpha():
            key = 'RULE%s' % key
        return key[:40]

    @api.model
    def _lineage_for_config(self, config, connector=None):
        """Lineage for every rule output that can reach THIS config, keyed by key.

        SOURCING S6, D3 fix B. `_lineage_by_output_key` answers for one connector,
        which is right for the board's LEFT column — those cards are that
        connector's fields. It is wrong for the right column, whose cards are
        components, because a component's lineage does not depend on which
        connector the board happened to pick.

        And it picks. No config on any of the four databases has `connector_id`, so
        `_api_active_connector` tie-breaks on mapping count: on abm that is
        connector 1 (18 wired mappings, zero transformation rules) while all eight
        rules live on connector 3. That is **S20**, and it is why the owner's board
        showed no "Derived here" lane and no lineage anywhere — correctly, for the
        connector it was showing, and unhelpfully for the question being asked.

        So the union: the board's own connector, plus every connector holding a
        wire into this config, plus the connector of any `('rule', key)` binding.
        A rule that produces a key is a rule that produces a key; the vendor does
        not have to have been called for that to be true.
        """
        Conn = self.env.get('hr.integration.connector')
        FM = self.env.get('hr.integration.field.mapping')
        if Conn is None or FM is None:
            return {}
        conns = Conn.browse()
        if connector:
            conns |= connector
        try:
            if config.rule_ids:
                conns |= FM.sudo().search(
                    [('target_rule_id', 'in', config.rule_ids.ids)]).mapped('connector_id')
            if config.connector_id:
                conns |= config.connector_id
        except Exception:       # noqa: BLE001 — lineage must never break a board
            pass
        # A `('rule', key)` binding is written by `api_mapping_create`, which draws
        # a field mapping on the same connector — so the mapping search above has
        # already caught it. Nothing here walks every connector on the database:
        # `_lineage_by_output_key` costs a search per rule, and an unbounded set
        # would make a 200-connector bureau pay for one card's tooltip.
        out = {}
        for c in conns:
            for key, val in self._lineage_by_output_key(c, config).items():
                out.setdefault(key, val)
        return out

    @api.model
    def _mc_src_kinds(self, srcs):
        """The ranked source chips a card carries, with the sentence for each.

        JOURNEY J9. The rank is the position among the sources ACTUALLY MAPPED
        ON THIS CARD, not a fixed number per type — the owner chose that
        explicitly. A card holding a spreadsheet column and a contract component
        reads Spreadsheet¹ · Contract component², not ²·³, because the reader is
        being told the order THIS component reads in and a gap would only invite
        the question "where is number one?".

        `rank` is 0 on a single-source card, and the client renders no
        superscript for it: a number that is always 1 is decoration, and the
        chips stop meaning "read me" the moment they carry decoration.
        """
        out = []
        total = len(srcs)
        for pos, src in enumerate(srcs):
            kind = src.get('kind') or 'none'
            if kind == 'none':
                continue
            out.append({
                'kind': kind,
                'key': src.get('key') or '',
                # J10 — the human name, when the key is a technical one. The
                # client prefers it for the chip's own text and the server has
                # already used it in `note`.
                'label': src.get('label') or '',
                'rank': (pos + 1) if total > 1 else 0,
                'note': self._source_rank_note(kind, src.get('key') or '',
                                               pos, total,
                                               label=src.get('label')),
            })
        return out

    @api.model
    def _source_rank_note(self, kind, key, pos, total, label=None):
        """The tooltip on one ranked chip: what it reads, and when.

        Every literal is written out — `_(variable)` extracts nothing and ships
        English forever (S19). The vocabulary stays inside `_SOURCE_LABELS`; no
        term of this method's own appears here.

        JOURNEY J10 — `label` is the HUMAN name of what is read, and the record
        tiers are the reason it exists. The key a record source folds on is the
        technical field name (`job_id`), which is the right thing to compare and
        the wrong thing to show: the sentence a reader wants is *"Reads Job
        Position from the contract record"*. Every other kind's key is already
        the name a person typed, so they pass nothing and read exactly as they
        did.
        """
        src = self._source_label(kind)
        shown = label or key
        if total <= 1:
            if shown:
                return _("Reads “%(key)s” from %(src)s.", key=shown, src=src)
            return _("Read from %s.") % src
        if pos == 0:
            if shown:
                return _("Reads “%(key)s” from %(src)s. Tried first.",
                         key=shown, src=src)
            return _("Read from %s. Tried first.") % src
        if shown:
            return _("Reads “%(key)s” from %(src)s. Used when nothing above "
                     "delivered a value.", key=shown, src=src)
        return _("Read from %s. Used when nothing above delivered a value.") \
            % src

    @api.model
    def _mc_right_item(self, rule, declared, note='', lineage=None, conflict=None):
        # JOURNEY J9 — `declared` is a LIST now (`_declared_sources`). A dict is
        # still accepted, because a caller that only has the winning source is
        # asking a legitimate question and should not have to wrap it.
        srcs = declared if isinstance(declared, list) else [declared]
        declared = srcs[0]
        wirable = rule.column_type == 'input'
        item = {'id': rule.id, 'label': (rule.name or rule.code),
                'sublabel': rule.code or '',
                # `srcKind` stays populated with the WINNER so a stale bundle
                # against a new server still renders one correct chip, and
                # `srcKinds` carries the whole ranked list for a current one.
                'srcKind': declared['kind'], 'srcNote': note,
                'srcKinds': self._mc_src_kinds(srcs),
                'meta': {'col': rule.column_letter or '', 'type': rule.column_type}}
        # SOURCING S6 — a component fed by a rule output carries that rule's
        # lineage, so "how is this worked out" is answerable from the COMPONENT and
        # not only from the vendor field two columns to the left. The canvas grows
        # the affordance only for a card that has one, so every other board and
        # every other card is unchanged.
        if declared['kind'] == 'rule' and lineage:
            lin = lineage.get(declared['key'])
            if lin:
                item['lineage'] = lin
        # JOURNEY J3 S2 — a component read by two live sources SAYS so, on both
        # boards, however the state arose. Pre-existing dual rows (abm has several,
        # drawn long before the dialog existed) get the chip on load; nothing has
        # to be redrawn for the truth to appear.
        if conflict:
            item['conflict'] = conflict
        # SOURCING S5 — an input with nothing feeding it can start a rule from
        # here. The key is sanitised through the SAME contract S2 put on
        # `output_key`, so the composer never opens on a key its own constraint
        # is about to refuse — an underscore in the pre-fill would have made the
        # affordance a dead end.
        if wirable and declared['kind'] == 'none':
            item['meta']['createRule'] = {
                'label': _("Create a rule for this"),
                'hint': _("Open the rule composer with “%s” as the output key.")
                        % self._mc_legal_output_key(rule),
                'output_key': self._mc_legal_output_key(rule),
                'component_id': rule.id,
            }
        if not wirable:
            # ==========================================================
            # SOURCING S6, D1 — ONE pill, and the pill says the right word.
            #
            # This block used to add `badge: "Calculated"` on top of the
            # `srcKind: 'calculated'` set above, so every one of abm's 45 sealed
            # cards rendered TWO adjacent pills both reading CALCULATED — the
            # defect the owner photographed. For a produced column, "it is
            # calculated" and "it needs no source" are one fact told twice; the
            # explanation belongs in the tooltip, not in a second pill.
            #
            # The BADGE keeps it, not the source chip, for two reasons: it is the
            # one that can carry a sentence (`badgeHint`), and it is the one the
            # sealed styling and `meta.wirable` already key off. `srcKind` is
            # dropped so the chip cannot render at all.
            #
            # And the defect nobody reported: this branch covers `constant` as
            # well as `formula`, so abm's NINE fixed-value columns were badged
            # "Calculated" — not a duplicate but a contradiction, sitting next to
            # a chip correctly reading "Fixed value". The label now comes from the
            # component's own kind, so a constant says Fixed value and means it.
            # ==========================================================
            item['srcKind'] = ''
            item['srcKinds'] = []
            label = self._source_label(declared['kind'])
            item['meta'].update({
                'wirable': False,
                'badge': label,
                'badgeTone': 'calc',
                # Wording borrowed from the employee tab, which has said this for
                # a year: a produced column is "produced, not imported".
                'badgeHint': _("%s — needs no source. This column is produced by "
                               "the scheme, not imported into it.") % label,
            })
        return item

    @api.model
    def _mc_right_column(self, config, actuals, record_dests, wirable_only=False,
                         wire_dests=None, lineage=None, board=''):
        """Every component a mapping board should show, in display order.

        J3 S2 — `board` ('api' | 'import') selects the WORDING of the conflict
        chip, not whether it is computed: the useful sentence is always about the
        source you are NOT looking at. Detection runs once for the whole config
        through `_source_conflicts`, so the two boards can never disagree.
        """
        rules = config.rule_ids if not wirable_only else config.rule_ids.filtered(
            lambda r: r.column_type == 'input')
        rules = rules.sorted(key=lambda r: r.sequence)
        conflicts = self._source_conflicts(config) if board else {}
        out = []
        for r in rules:
            declared = self._declared_sources(r, record_dests, wire_dests)
            # JOURNEY J9 — a card that already renders its sources RANKED is
            # saying the whole thing; the conflict chip would hand the reader the
            # same fact a second time, in weaker words. S6 D1's principle, one
            # question further on.
            conflict = (self._conflict_chip(conflicts.get(r.id), board)
                        if len(declared) < 2 else None)
            out.append(self._mc_right_item(
                r, declared, self._source_note(r, actuals, record_dests,
                                              wire_dests),
                lineage, conflict))
        return out

    @api.model
    def _source_note(self, rule, actuals, record_dests, wire_dests=None):
        """The one-line note a mapping board shows against a component.

        Says what already feeds this target, so a user drawing a wire can see it
        is not idle. Silent when nothing feeds it — an empty note is the normal
        case and a board full of chips would say nothing at all.

        JOURNEY J9 — "Already fed by X" was a warning about something that is now
        LEGAL, so a component with several sources states the resulting ORDER
        instead. A single source keeps the original sentence: there is no order
        to state, and rewording it would churn every card on every board to say
        the same thing differently.
        """
        declared = self._declared_sources(rule, record_dests, wire_dests)
        if len(declared) > 1:
            parts = []
            for src in declared:
                label = self._source_label(src['kind'])
                # J10 — a record source shows its FIELD LABEL, never `job_id`.
                shown = src.get('label') or src['key']
                # The label is already translated; the quotes around a KEY are
                # punctuation, not a sentence, so nothing here goes to gettext.
                parts.append('%s “%s”' % (label, shown) if shown else label)
            return _("Read in this order: %s") % ', '.join(parts)
        kind = declared[0]['kind']
        if kind == 'none':
            return ''
        label = self._source_label(kind)
        if declared[0]['key']:
            return _("Already fed by %(src)s “%(key)s”",
                     src=label, key=declared[0]['key'])
        return _("Already fed by %s") % label

    @api.model
    def _source_block(self, rule, actuals, actual_run, record_dests,
                      wire_dests=None):
        block = {'declared': self._declared_source(
            rule, record_dests, wire_dests)}
        entry = actuals.get(rule.code or '')
        if entry and isinstance(entry, dict):
            block['actual'] = {
                'kind': entry.get('src') or 'none',
                'key': entry.get('key') or '',
                'via': entry.get('via') or '',
                'fell_back': bool(entry.get('fell_back')),
                'ignored': entry.get('ignored') or None,
                'adj': entry.get('adj') or [],
                'run': actual_run,
            }
        return block

    @api.model
    def get_studio_data(self, config_id=None):
        config = self._pick_config(config_id)
        if not config:
            return {'empty': True, 'configs': self.get_config_list()}

        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        by_col = self._col_to_rule(rules)

        # dependency maps (expand A#:B# ranges to every member column)
        depends = {}
        used_by = {}
        for r in rules:
            refs = self._expand_refs(r.excel_formula, by_col) if r.column_type == 'formula' else set()
            depends[r.id] = [by_col[c].name for c in refs if c in by_col]
            for c in refs:
                rr = by_col.get(c)
                if rr:
                    used_by.setdefault(rr.id, []).append(r.name)

        # F15 — note badges per component (one query, tallied in Python)
        note_by_rule = defaultdict(lambda: {'count': 0, 'review_open': 0})
        for n in self.env['hr.formula.rule.note'].search([('config_id', '=', config.id)]):
            d = note_by_rule[n.rule_id.id]
            d['count'] += 1
            if n.is_review and not n.resolved:
                d['review_open'] += 1

        # SOURCING S4 — declared vs actual, computed ONCE for the whole config.
        # `_source_actuals` reads a single payslip and indexes it; `_declared_source`
        # is pure per-rule derivation. Doing it here rather than inside the loop is
        # what keeps a 250-column scheme at one extra query instead of 250.
        actuals, actual_run = self._source_actuals(config)
        record_dests = self._source_record_dests(config)
        # SOURCING S6 — and the wires that were drawn before bindings existed. One
        # search for the whole config; a component with a live connector wire stops
        # saying "No source chosen" about a source somebody explicitly drew.
        wire_dests = self._source_wire_dests(config)

        components = []
        for r in rules:
            components.append({
                'id': r.id,
                'col': r.column_letter or '?',
                'code': r.code or '',
                # F111: sequence = display order (letters are frozen identities);
                # category_id drives the grid's category band strip + grouping.
                'sequence': r.sequence or 0,
                'category_id': r.category_id.id or False,
                'note_count': note_by_rule[r.id]['count'],
                'review_open': note_by_rule[r.id]['review_open'],
                # Multilingual: prefer the translatable linked salary rule's label
                # (resolves to the reader's language); fall back to the rule name.
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or '(unnamed)',
                'type': r.column_type,
                'group': _group_for(r),
                'excel_formula': r.excel_formula or '',
                'constant_value': r.constant_value or 0.0,
                'tokens': self._tokenize(r, by_col),
                'explain': self._explain(r, by_col),
                'category': r.category_id.name if r.category_id else (r.column_type or '').title(),
                'number_format': r.number_format or 'number',
                'is_valid': bool(r.is_valid) and not r.has_evaluation_error,
                'validation_message': r.validation_message or r.last_evaluation_error or '',
                'appears_on_payslip': bool(r.appears_on_payslip),
                'depends_on': depends.get(r.id, []),
                'used_by': used_by.get(r.id, []),
                # COLROLES P2 — what the column is FOR. The lens, the role chips
                # and the grid filter all read these; the client still guards with
                # `c.column_role || 'payroll'` so an older payload degrades to the
                # pre-roles behaviour instead of blanking the outline.
                'column_role': r.column_role or 'payroll',
                'column_role_source': r.column_role_source or 'auto',
                'is_contract_component': bool(r.is_contract_component),
                # RD47 — what the NET PAY formula does with this component, sent
                # so the panel can decide the MINUS SIGN from the arithmetic
                # rather than from which column the component is filed under.
                # Those are two different questions and conflating them is what
                # drew "Taxable Income" as a negative.
                'net_role': getattr(r, 'net_role', False) or '',
                'is_text_component': bool(r.is_text_component),
                # RD52 — the panel formats from the component's own type; a
                # bank account number must never reach the currency formatter.
                'value_kind': getattr(r, 'value_kind', None) or 'money',
                'is_visible_in_grid': bool(r.is_visible_in_grid),
                # SOURCING S4 — ONE nested object, not five sibling keys, so every
                # render site reads one path and an older client degrades to "no
                # source block" rather than to half a truth. Same contract as
                # `column_role` above, for the same reason.
                'source': self._source_block(
                    r, actuals, actual_run, record_dests, wire_dests),
            })

        samples = [{'id': s.id, 'name': s.name} for s in config.sample_data_ids]
        preview = self._compute(config, samples[0]['id']) if samples else {'sample_id': False, 'values': {}}

        score = self._score(config)
        return {
            'empty': False,
            'configs': self.get_config_list(),
            'config': {
                'id': config.id,
                'name': config.name,
                'code': config.code or '',
                'country': config.country_code or '',
                'currency': config.currency_id.symbol if config.currency_id else '',
                'state': config.state,
                'score': score,
                'validation_message': config.validation_message or '',
                'rule_count': len(rules),
                'sample_count': len(config.sample_data_ids),
                # B2 — branch lineage (drives the header chip + Branches overlay)
                'is_branch': bool(config.parent_branch_id),
                'parent_id': config.parent_branch_id.id or False,
                'parent_name': config.parent_branch_id.name or '',
                'branch_state': config.branch_state or 'open',
                'branch_count': len(config.child_branch_ids.filtered(
                    lambda b: b.branch_state == 'open')),
                # B5 — scheme-variant lineage (header chip + Variants overlay)
                'is_master': bool(config.variant_ids),
                'is_variant': bool(config.master_config_id),
                'master_id': config.master_config_id.id or False,
                'master_name': config.master_config_id.name or '',
                'variant_count': len(config.variant_ids),
                'override_count': len([c for c in (config.variant_override_codes or '').split(',') if c.strip()]),
            },
            'components': components,
            'samples': samples,
            'preview': preview,
            'field_meta': self._field_meta(),
            'can_edit': self._can_edit(),
            'scenarios': [self._scenario_payload(s) for s in self.env['hr.formula.scenario']
                          .search([('config_id', '=', config.id)])],
            'rate_tables': [self._rate_table_payload(t) for t in
                            self.env['hr.formula.rate.table'].search([('config_id', '=', config.id)])],
        }

    # ------------------------------------------------------------------
    # Formula Intelligence v1 (deterministic dependency graph)
    # ------------------------------------------------------------------
    @api.model
    def _normalized_dep_cols(self, rules):
        """Normalize each rule's ``formula_dependencies`` to the column letters
        of real rules in the config.

        ``formula_dependencies`` (see hr.formula.rule._compute_dependencies) is a
        stored Char holding a comma-separated MIX of column letters (A, AA) AND
        component codes (BASIC, GROSS) plus incidental noise. We resolve every
        token exactly the way the engine's evaluator does — column_letter first,
        then code — and keep only tokens that land on a rule that lives in this
        config, so spurious tokens (function fragments, unknown codes) drop out.

        Returns ``{rule.id: set(column_letter, ...)}`` — the columns each formula
        rule depends on (data-flow *sources*).
        """
        by_col = {r.column_letter: r for r in rules if r.column_letter}
        by_code = {r.code: r for r in rules if r.code}
        deps = {}
        for r in rules:
            cols = set()
            raw = r.formula_dependencies or ''
            if r.column_type == 'formula' and raw:
                for tok in raw.split(','):
                    tok = tok.strip()
                    if not tok:
                        continue
                    dep = by_col.get(tok) or by_code.get(tok)
                    if dep and dep.column_letter:
                        cols.add(dep.column_letter)
            deps[r.id] = cols
        return deps

    @api.model
    def get_intelligence(self, config_id=None):
        """Deterministic dependency-graph payload for the Formula Intelligence
        panels (and the grid-highlight primitives in Feature 2).

        Shape::

            {nodes: [{id, code, col, name, category, appears_on_payslip, is_valid}],
             edges: [[from_col, to_col], ...],   # data-flow: source -> consumer
             execution_order: [col, ...],        # formula rules, dependencies first
             unused: [col, ...],
             cycles: [{cols, codes, human_explanation}, ...]}
        """
        config = self._pick_config(config_id)
        if not config:
            return {'empty': True, 'nodes': [], 'edges': [],
                    'execution_order': [], 'unused': [], 'cycles': []}

        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        deps = self._normalized_dep_cols(rules)
        col_to_code = {r.column_letter: (r.code or r.column_letter)
                       for r in rules if r.column_letter}

        nodes = [{
            'id': r.id,
            'code': r.code or '',
            'col': r.column_letter or '',
            'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or '(unnamed)',
            'category': r.category_id.name if r.category_id else (r.column_type or '').title(),
            'appears_on_payslip': bool(r.appears_on_payslip),
            'is_valid': bool(r.is_valid) and not r.has_evaluation_error,
        } for r in rules]

        # Edges point in data-flow direction: [dependency_col, consumer_col], so an
        # edge's source is evaluated before its target (matches execution_order).
        edge_set = set()
        for r in rules:
            if not r.column_letter:
                continue
            for dep_col in deps[r.id]:
                edge_set.add((dep_col, r.column_letter))
        edges = [[a, b] for a, b in edge_set]

        # ---- execution order: topological sort over formula rules only --------
        # (dependencies on inputs/constants never constrain ordering — they are
        # always ready — so we only count formula->formula edges, exactly the
        # relation the evaluator's Kahn sort walks.)
        formula_rules = [r for r in rules if r.column_type == 'formula' and r.column_letter]
        fcols = {r.column_letter for r in formula_rules}
        indeg = {}
        succ = defaultdict(list)
        for r in formula_rules:
            fdeps = {d for d in deps[r.id] if d in fcols and d != r.column_letter}
            indeg[r.column_letter] = len(fdeps)
            for d in fdeps:
                succ[d].append(r.column_letter)

        # deterministic: process ready nodes in column order
        queue = sorted((c for c, d in indeg.items() if d == 0), key=self._col_num)
        execution_order = []
        while queue:
            col = queue.pop(0)
            execution_order.append(col)
            newly_ready = []
            for nxt in succ.get(col, []):
                indeg[nxt] -= 1
                if indeg[nxt] == 0:
                    newly_ready.append(nxt)
            # keep the queue in column order so the result is stable
            for c in sorted(newly_ready, key=self._col_num):
                queue.append(c)
            queue.sort(key=self._col_num)
        # cycle members never reach in-degree 0 — append them so the order still
        # accounts for every formula rule (AC1: len == number of formula rules).
        if len(execution_order) != len(formula_rules):
            emitted = set(execution_order)
            for r in formula_rules:
                if r.column_letter not in emitted:
                    execution_order.append(r.column_letter)

        # ---- unused: nothing downstream depends on it AND not on the payslip ---
        # A column with dependents is "consumed" (this is what excludes an input
        # that feeds a formula), so the two conditions together mean truly dead.
        has_dependents = {src for src, _tgt in edge_set}
        unused = [r.column_letter for r in rules
                  if r.column_letter
                  and r.column_letter not in has_dependents
                  and not r.appears_on_payslip]

        # ---- cycles: DFS back-edge detection with full path recovery ----------
        adj = {r.column_letter: [d for d in deps[r.id] if d in fcols]
               for r in formula_rules}
        WHITE, GRAY, BLACK = 0, 1, 2
        color = {c: WHITE for c in adj}
        cycles = []
        seen = set()

        def _dfs(u, stack):
            color[u] = GRAY
            stack.append(u)
            for v in adj.get(u, []):
                if color.get(v) == GRAY:
                    # back edge — the cycle is the stack slice from v to u
                    cyc = stack[stack.index(v):]
                    key = frozenset(cyc)
                    if key not in seen:
                        seen.add(key)
                        codes = [col_to_code.get(c, c) for c in cyc]
                        cycles.append({
                            'cols': cyc[:],
                            'codes': codes,
                            'human_explanation': _(
                                "Circular dependency: %s") % ' → '.join(codes + [codes[0]]),
                        })
                elif color.get(v) == WHITE:
                    _dfs(v, stack)
            stack.pop()
            color[u] = BLACK

        for c in sorted(adj, key=self._col_num):
            if color[c] == WHITE:
                _dfs(c, [])

        return {
            'empty': False,
            'nodes': nodes,
            'edges': edges,
            'execution_order': execution_order,
            'unused': unused,
            'cycles': cycles,
        }

    @api.model
    def _closure(self, start, adj):
        """Transitive closure of ``start`` over adjacency map ``adj`` (col -> set
        of neighbour cols), excluding ``start`` itself. Cycle-safe via seen-set."""
        seen = set()
        stack = list(adj.get(start, ()))
        while stack:
            c = stack.pop()
            if c in seen or c == start:
                continue
            seen.add(c)
            stack.extend(adj.get(c, ()))
        return seen

    @api.model
    def _config_employee_count(self, config):
        """Employees attached to this config's scheme: distinct employees on
        payslips computed with it (the truest 'who does this affect' measure),
        falling back to a division match so a not-yet-run config still reports."""
        Payslip = self.env['hr.payslip']
        if 'formula_config_id' in Payslip._fields:
            emps = Payslip.search([('formula_config_id', '=', config.id)]).employee_id
            if emps:
                return len(emps)
        Emp = self.env['hr.employee']
        div = getattr(config, 'pb_division', False)
        if div and 'pb_division' in Emp._fields:
            return Emp.search_count([('pb_division', '=', div)])
        if div and 'division' in Emp._fields:
            return Emp.search_count([('division', '=', div)])
        return 0

    @api.model
    def _delete_eligibility(self, config):
        """Can this config be removed outright, or only archived?

        A config that never touched real payroll (no payslips, import batches,
        carry-forwards, prorations, retro adjustments or field mappings) is safe
        to delete — everything else it owns is config-scoped metadata that
        cascades. Anything else must be archived so the history it produced
        stays readable.

        Returns ``{'can_delete': bool, 'delete_blockers': [...],
        'delete_blocked_by': str}`` for merging into a cockpit card.
        """
        try:
            blockers = config._delete_blockers()
        except Exception:  # pragma: no cover - never break the board over this
            _logger.exception("delete eligibility failed for config %s", config.id)
            return {'can_delete': False, 'delete_blockers': [],
                    'delete_blocked_by': _('existing payroll data')}
        return {
            'can_delete': not blockers,
            'delete_blockers': blockers,
            'delete_blocked_by': config._delete_blocker_message(blockers) if blockers else '',
        }

    @api.model
    def config_delete_eligibility(self, config_id):
        """Fresh delete/archive verdict for one config (called before confirming)."""
        cfg = self.env['hr.formula.config'].browse(int(config_id))
        if not cfg.exists():
            return {'ok': False, 'error': 'not_found'}
        return dict({'ok': True, 'name': cfg.name or ''}, **self._delete_eligibility(cfg))

    @api.model
    def get_impact_analysis(self, rule_id):
        """Impact of one component: its transitive upstream (what feeds it),
        transitive downstream (what it feeds), the payslip-visible slice of that
        downstream, and how many employees the config touches.

        Shape::

            {rule: {id, col, code, name},
             upstream: [node, ...], downstream: [node, ...],
             payslip_visible: [node, ...], employee_count: int}

        where ``node = {id, col, code, name, appears_on_payslip}``.
        """
        Rule = self.env['hr.formula.rule']
        rule = Rule.browse(int(rule_id))
        empty = {'empty': True, 'rule': {}, 'upstream': [], 'downstream': [],
                 'payslip_visible': [], 'employee_count': 0}
        if not rule.exists() or not rule.column_letter:
            return empty

        config = rule.config_id
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        deps = self._normalized_dep_cols(rules)
        by_col = {r.column_letter: r for r in rules if r.column_letter}

        # depends_cols: col -> cols it consumes; dependents: col -> cols consuming it
        depends_cols = {}
        dependents = defaultdict(set)
        for r in rules:
            if not r.column_letter:
                continue
            dcols = deps.get(r.id, set())
            depends_cols[r.column_letter] = set(dcols)
            for d in dcols:
                dependents[d].add(r.column_letter)

        start = rule.column_letter
        up_cols = self._closure(start, depends_cols)
        down_cols = self._closure(start, dependents)

        def _node(col):
            r = by_col.get(col)
            if not r:
                return None
            return {
                'id': r.id,
                'col': col,
                'code': r.code or '',
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or '(unnamed)',
                'appears_on_payslip': bool(r.appears_on_payslip),
            }

        upstream = [n for n in (_node(c) for c in sorted(up_cols, key=self._col_num)) if n]
        downstream = [n for n in (_node(c) for c in sorted(down_cols, key=self._col_num)) if n]
        payslip_visible = [n for n in downstream if n['appears_on_payslip']]

        return {
            'empty': False,
            'rule': {
                'id': rule.id,
                'col': start,
                'code': rule.code or '',
                'name': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or '(unnamed)',
            },
            'upstream': upstream,
            'downstream': downstream,
            'payslip_visible': payslip_visible,
            'employee_count': self._config_employee_count(config),
        }

    # COLROLES P2 — lower-case role names for use INSIDE a sentence ("filed as
    # employee profile data"). The Selection labels on the field are title-case
    # because they head a form field; these read as prose.
    @api.model
    def _role_label(self, role):
        # Literal _() calls (not _(variable)) so the terms are extractable.
        return {
            'identity': _("the employee identity"),
            'profile': _("employee profile data"),
            'contract': _("contract data"),
            'bank': _("bank data"),
            'reference': _("a reference"),
        }.get(role or 'payroll', _("pay"))

    # COLROLES P4 — the same roles as a single WORD, for a "Payroll → Bank" arrow
    # in the reclassification diff, where a phrase would not fit.
    @api.model
    def _role_word(self, role):
        return {
            'identity': _("Identity"),
            'profile': _("Employee profile"),
            'contract': _("Contract"),
            'bank': _("Bank"),
            'reference': _("Reference"),
        }.get(role or 'payroll', _("Payroll"))

    @api.model
    def _can_edit(self):
        """Edit/Delete/PayAI affordances are for Formula Managers/Admins (who hold
        write access). A read-only 'Formula User' gets them hidden — the ACL would
        block the write anyway. Fail open so a missing group never locks out admins."""
        try:
            u = self.env.user
            return bool(u.has_group('base.group_system')
                        or u.has_group('pb_hr_payroll_formula.group_formula_manager'))
        except Exception:
            return True

    @api.model
    def reclassify_roles(self, config_id, dry_run=False, apply_ids=None):
        """Re-run the column-role classifier over one salary structure.

        Rows whose role a person set by hand are read but never written (CR-A1), and
        the return value names every row that moved so the change can be reviewed
        rather than merely trusted.

        COLROLES P4 — the review dialog turned that promise into a surface, so the
        method grew two optional arguments and NO new behaviour by default:

        * `dry_run=True` computes the diff and writes nothing. That is what the
          dialog opens with: a reclassification you have not agreed to yet is a
          proposal, not an act.
        * `apply_ids` restricts the write to the rows you accepted. `None` (the
          historical call) still writes them all, so the migration and any existing
          caller are unaffected.

        Accepted rows keep `column_role_source='auto'` on purpose (locked rule): you
        agreed with a machine's reading, you did not author a different one. Skipped
        rows are left exactly as they were — including their source — so skipping
        never silently freezes a column against a future re-run.
        """
        if not self._can_edit():
            raise AccessError(_("You do not have permission to reclassify columns."))

        from odoo.addons.pb_hr_payroll_formula.models import column_role_classifier as crc

        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            raise UserError(_("That salary structure no longer exists."))

        rules = config.rule_ids
        # A column named in ANY other column's dependency list is payroll, whatever
        # its header says — the codes arrive comma-joined (CR2).
        referenced = set()
        for rule in rules:
            for code in (rule.formula_dependencies or '').split(','):
                code = code.strip().upper()
                if code:
                    referenced.add(code)

        changed = []
        counts = dict.fromkeys(crc.ROLES, 0)
        for rule in rules:
            role, tier, reason = crc.classify_column(
                rule.name,
                column_type=rule.column_type,
                is_contract_component=bool(rule.is_contract_component)
                and not bool(rule.is_text_component),
                is_text_component=bool(rule.is_text_component),
                band_label=rule.component_type or None,
                is_referenced=(rule.code or '').strip().upper() in referenced,
            )
            counts[role] = counts.get(role, 0) + 1
            if rule.column_role_source == 'user' or role == rule.column_role:
                continue
            changed.append({
                'id': rule.id,
                'code': rule.code,
                'name': rule.name,
                'from': rule.column_role,
                'from_label': self._role_word(rule.column_role),
                'to': role,
                'to_label': self._role_word(role),
                'tier': tier,
                'reason': reason,
            })

        applied = []
        if not dry_run:
            wanted = None if apply_ids is None else {int(i) for i in apply_ids}
            for entry in changed:
                if wanted is not None and entry['id'] not in wanted:
                    continue
                self.env['hr.formula.rule'].browse(entry['id']).write({
                    'column_role': entry['to'],
                    'column_role_source': 'auto',
                })
                applied.append(entry['id'])

        return {
            'ok': True,
            'config_id': config.id,
            'config_name': config.name,
            'changed': changed,
            'counts': counts,
            'dry_run': bool(dry_run),
            'applied': applied,
        }

    # ==================================================================
    # NETROLE P2 — the category review
    #
    # The engine reads a scheme's formulas and knows what net pay does with
    # every component (`suggest_categories`). It writes nothing. This is the
    # conversation that turns that reading into a decision: a grouped list of
    # what WOULD move, the sentence that explains each one, and a checkbox.
    #
    # Two opinions are held at once. Ours comes from the arithmetic; theirs
    # comes from the coloured band they typed above the column in their own
    # spreadsheet. Where those disagree the row arrives UNTICKED with the
    # disagreement spelled out — the math is allowed to win, but only by their
    # click.
    # ==================================================================

    #: Group order in the dialog. `review` leads because it is the only group
    #: that cannot be answered by pressing Apply.
    _CATEGORY_REVIEW_GROUPS = ('review', 'earning', 'deduction', 'employer_cost',
                               'net', 'info')

    @api.model
    def _category_review_group_meta(self):
        return {
            'review': (_("Needs a decision"),
                       _("The formulas do not settle these, or your own "
                         "column band says something different.")),
            'earning': (_("Added to pay"),
                        _("Net pay adds these up.")),
            'deduction': (_("Taken off pay"),
                          _("Net pay subtracts these.")),
            'employer_cost': (_("Employer cost"),
                              _("Never reaches the employee — the company "
                                "carries it on top of pay.")),
            'net': (_("Net pay"), _("The figure everything else adds up to.")),
            'info': (_("Information"),
                     _("Counts, references and working figures — not money "
                       "paid or taken.")),
        }

    @api.model
    def _category_review_group(self, row):
        if row.get('confidence') == 'review' or row.get('band_conflict'):
            return 'review'
        if row.get('quantity'):
            return 'info'
        role = row.get('role') or 'info'
        return role if role in self._CATEGORY_REVIEW_GROUPS else 'info'

    @api.model
    def _category_review_checked(self, row):
        """The default-tick policy, in one place and in reading order.

        Rule 2 is the promise the whole dialog rests on: a category somebody's
        own spreadsheet band already claimed is never silently overruled.
        Rule 7 is its mirror — a component nobody has ever filed cannot be
        "overridden", and leaving a fresh import entirely unticked would make
        the reading useless on the very screen it was built for.

        Rule 4 sits above the confidence tests on purpose. "This column counts
        hours" is read off the label and the arithmetic, not off the path to net
        pay, so a `likely` path does not weaken it — and the move it proposes is
        from pay to information, which no total anywhere depends on.
        """
        if not row.get('changes'):
            return False                      # 1. nothing to write
        if row.get('band_conflict'):
            return False                      # 2. their band disagrees
        if row.get('confidence') == 'review':
            return False                      # 3. only a person can settle it
        if row.get('quantity'):
            return True                       # 4. it counts hours, not money
        if row.get('confidence') == 'certain':
            return True                       # 5. the formulas are unambiguous
        if row.get('band_kind'):
            return True                       # 6. their band agrees with us
        current = row.get('current_category') or ''
        if not current or current == 'OTH':
            return True                       # 7. nothing was ever decided
        return False                          # 8. likely, over a real choice

    @api.model
    def category_review_data(self, config_id):
        """Everything the review dialog draws — and no category moved.

        Opening the review re-reads the scheme and stores that reading
        (`net_role*` on each rule). What it never touches is a CATEGORY: those
        move only through `category_review_apply`, and only for ticked rows.
        """
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            raise UserError(_("That salary structure no longer exists."))
        payload = {
            'ok': True,
            'config_id': config.id,
            'config_name': config.name or '',
            'can_edit': self._can_edit(),
            'error': '',
            'net_code': '',
            'net_name': '',
            'net_candidates': [],
            'groups': [],
            'agree_count': 0,
            'row_count': 0,
            'checked_count': 0,
        }
        # Opening the review re-reads the scheme and STORES what it read — but
        # only for somebody who could act on it. A read-only Formula User gets
        # the same payload computed write-free, because a screen that raises an
        # AccessError on open is not a read-only screen (C16's rail: detection
        # paths must not write).
        if payload['can_edit']:
            summary = (config.classify_net_roles() or {}).get(config.id) or {}
        else:
            summary = config._build_net_role_classification()
            summary.pop('_classification', None)
        if summary.get('error'):
            # No net-pay component. The dialog says so in one sentence and
            # offers the one fix there is — naming it.
            payload['error'] = summary['error']
            payload['net_candidates'] = [
                {'id': rule.id, 'col': rule.column_letter or '',
                 'code': rule.code or '', 'name': rule.name or rule.code or ''}
                for rule in config.rule_ids.sorted(key=lambda r: (r.sequence, r.id))
                if rule.column_type == 'formula'
            ]
            return payload
        payload['net_code'] = summary.get('net_code') or ''
        buckets = {key: [] for key in self._CATEGORY_REVIEW_GROUPS}
        rules_by_id = {rule.id: rule for rule in config.rule_ids}
        for row in config.suggest_categories():
            actionable = (row['changes'] or row['band_conflict']
                          or row['confidence'] == 'review')
            if not actionable:
                payload['agree_count'] += 1
                continue
            rule = rules_by_id.get(row['rule_id'])
            checked = self._category_review_checked(row)
            buckets[self._category_review_group(row)].append({
                'id': row['rule_id'],
                'col': (rule.column_letter or '') if rule else '',
                'code': row['code'],
                'name': row['name'] or row['code'],
                'role': row['role'],
                'role_label': row['role_label'],
                # The chip is tinted by what the component is being filed AS, so
                # an hours count never wears the green of pay.
                'tint': 'info' if row['quantity'] else row['role'],
                'detail': row['detail'],
                'reason': row['reason'],
                'confidence': row['confidence'],
                'current': row['current_category'],
                'current_label': row['current_category_name'] or row['current_category'],
                'suggested': row['suggested_category_code'],
                'suggested_label': row['suggested_category_name'],
                'changes': row['changes'],
                'quantity': row['quantity'],
                'band': row['band'],
                'band_conflict': row['band_conflict'],
                'band_conflict_text': row['band_conflict_text'],
                'accept': checked,
            })
            payload['row_count'] += 1
            payload['checked_count'] += 1 if checked else 0
        meta = self._category_review_group_meta()
        for key in self._CATEGORY_REVIEW_GROUPS:
            if not buckets[key]:
                continue
            label, hint = meta[key]
            payload['groups'].append({
                'key': key, 'label': label, 'hint': hint,
                'rows': buckets[key],
            })
        if payload['net_code']:
            net = config.rule_ids.filtered(lambda r: r.code == payload['net_code'])
            payload['net_name'] = (net[:1].name or '') if net else ''
        return payload

    @api.model
    def category_review_apply(self, config_id, rule_ids=None):
        """Write the accepted rows — the ONLY writer in this feature."""
        if not self._can_edit():
            raise AccessError(_("You do not have permission to change how "
                                "components are filed."))
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            raise UserError(_("That salary structure no longer exists."))
        ids = [int(i) for i in (rule_ids or [])]
        if not ids:
            return {'ok': True, 'applied': 0}
        # C4 — one logical operation, one reason, one version row per rule.
        applied = config.with_context(
            formula_version_reason='bulk').apply_suggested_categories(ids)
        return {'ok': True, 'applied': applied, 'requested': len(ids)}

    @api.model
    def category_review_set_net(self, config_id, rule_id):
        """Name the net-pay component, then read the scheme again.

        The classifier refuses to guess which component is net pay (C7), so a
        scheme it cannot read is not a dead end — it is one question.
        """
        if not self._can_edit():
            raise AccessError(_("You do not have permission to change this "
                                "salary structure."))
        config = self.env['hr.formula.config'].browse(int(config_id))
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not config.exists() or not rule.exists() or rule.config_id != config:
            raise UserError(_("That component is not part of this salary "
                              "structure."))
        Category = self.env['hr.salary.rule.category']
        net = Category.search([('code', '=', 'NET')], limit=1)
        if not net:
            net = Category.create({'name': _("Net"), 'code': 'NET'})
        rule.with_context(formula_version_reason='edit').category_id = net.id
        if rule.salary_rule_id:
            rule.salary_rule_id.category_id = net.id
        return self.category_review_data(config.id)

    @api.model
    def _field_meta(self):
        """Option lists for the inline component editor (loaded once)."""
        Rule = self.env['hr.formula.rule']

        def _sel(field):
            return [{'value': v, 'label': l}
                    for v, l in Rule._fields[field].selection]

        cats = self.env['hr.salary.rule.category'].search([], order='name')
        rules = self.env['hr.salary.rule'].search([], order='name', limit=400)
        connectors = (self.env['hr.integration.connector'].search([], order='name')
                      if 'hr.integration.connector' in self.env else self.env['hr.formula.config'].browse())
        return {
            'categories': [{'id': c.id, 'name': c.name} for c in cats],
            'salary_rules': [{'id': r.id, 'name': r.name, 'code': r.code or ''} for r in rules],
            'connectors': [{'id': c.id, 'name': c.name} for c in connectors],
            'column_types': _sel('column_type'),
            'number_formats': _sel('number_format'),
            'data_sources': _sel('data_source'),
            'text_aligns': _sel('text_align'),
        }

    @api.model
    def _score(self, config):
        rules = config.rule_ids.filtered(lambda r: r.column_type == 'formula')
        if not rules:
            return 100
        bad = len(rules.filtered(lambda r: (not r.is_valid) or r.has_evaluation_error or r.has_circular_ref))
        return int(round(100 * (len(rules) - bad) / max(len(rules), 1)))

    # ------------------------------------------------------------------
    # live compute (reuses the engine evaluator)
    # ------------------------------------------------------------------
    @api.model
    def _compute(self, config, sample_id):
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        by_code = {r.code: r.column_letter for r in rules}
        values = {}
        try:
            sample = self.env['hr.formula.sample.data'].browse(int(sample_id))
            if sample.exists():
                input_values = json.loads(sample.input_values_json or '{}')
                # Use the rule.evaluate() path (same evaluator real payslips and the
                # test workbench use) so IF / BRACKET / self._if formulas compute
                # correctly — the FormulaEvaluator fast path cannot handle those and
                # silently returns 0 (F11).
                values = sample._evaluate_rules_with_dependencies(input_values)
        except Exception as e:
            _logger.warning("Studio compute failed: %s", e)
        # key by column letter for the UI
        out = {}
        for code, v in values.items():
            col = by_code.get(code)
            if col:
                try:
                    out[col] = float(v)
                except (TypeError, ValueError):
                    out[col] = 0.0
        return {'sample_id': int(sample_id), 'values': out}

    @api.model
    def compute_preview(self, config_id, sample_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        return self._compute(config, sample_id) if config.exists() else {'sample_id': sample_id, 'values': {}}

    # ==================================================================
    # RD46 — PREVIEW THE FORMULAS AGAINST A REAL PERSON.
    #
    # The Live Preview panel could only ever show a made-up sample, and on a
    # scheme nobody had generated samples for it showed a column of ₫0. So the
    # one question a person actually brings to this screen — "why did THIS
    # employee get THAT number?" — had no answer here, and the numbers on the
    # panel looked like the formulas were producing nothing.
    #
    # THE COPY RULE, and it is the whole safety story: these methods READ a
    # payslip and never write one. The evaluation runs on an in-memory record
    # (`.new()`, never saved) with `readonly=True`, which is the sample model's
    # existing zero-write mode — it skips the dependency-metadata refresh that
    # would otherwise stamp `write_date` on every component of a live scheme.
    # Nothing here can change what anybody is paid, so editing a formula while
    # a real person's numbers are on screen is safe by construction rather than
    # by the caller remembering to be careful.
    #
    # ONE EVALUATOR. The numbers come out of the same
    # `_evaluate_rules_with_dependencies` the samples, the test workbench and a
    # real payslip all use. A second implementation would be a second answer to
    # "what does this formula produce", which is the failure the whole sourcing
    # programme exists to prevent.
    # ==================================================================

    @api.model
    def preview_runs(self, config_id=None):
        """Pay runs this scheme actually produced payslips for.

        Scoped by `slip_ids.formula_config_id` rather than by date so the list
        can never offer a run whose payslips this scheme did not compute —
        picking one of those would preview a person against formulas that were
        never applied to them, which is a worse answer than no answer.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'runs': []}
        Run = self.env['hr.payslip.run']
        try:
            runs = Run.search(
                [('slip_ids.formula_config_id', '=', config.id)],
                order='date_start desc, id desc', limit=24)
        except Exception:       # noqa: BLE001 — a picker must not break the panel
            _logger.warning("RD46: could not list pay runs", exc_info=True)
            return {'ok': False, 'runs': []}
        out = []
        for run in runs:
            out.append({
                'id': run.id,
                'name': run.name or _('Pay run'),
                'date_from': fields.Date.to_string(run.date_start) if run.date_start else '',
                'date_to': fields.Date.to_string(run.date_end) if run.date_end else '',
                'count': len(run.slip_ids.filtered(
                    lambda s: s.formula_config_id.id == config.id)),
            })
        return {'ok': True, 'runs': out}

    @api.model
    def preview_people(self, run_id, config_id=None):
        """Who is in that run, and which payslip to preview for each.

        Only the payslips THIS scheme computed, for the reason above.
        """
        config = self._pick_config(config_id)
        if not (config and run_id):
            return {'ok': False, 'people': []}
        slips = self.env['hr.payslip'].search(
            [('payslip_run_id', '=', int(run_id)),
             ('formula_config_id', '=', config.id)],
            order='employee_id, id')
        people = []
        for slip in slips:
            people.append({
                'payslip_id': slip.id,
                'name': slip.employee_id.display_name or _('Employee'),
                # `pb_sourced_inputs` is how many of this payslip's inputs came
                # from somewhere real. Surfaced in the picker because a person
                # whose payslip was computed entirely on defaults is exactly the
                # one somebody is about to come asking about.
                'sourced': slip.pb_sourced_inputs or 0,
            })
        return {'ok': True, 'people': people}

    def _rd46_preview_label(self, slip, anonymize=False):
        """What the panel calls this preview. Never a bare payslip number."""
        who = _('Employee') if anonymize else (
            slip.employee_id.display_name or _('Employee'))
        if slip.date_from and slip.date_to:
            from odoo.tools.misc import format_date
            return '%s · %s' % (who, format_date(self.env, slip.date_from))
        return who

    @api.model
    def preview_from_payslip(self, config_id, payslip_id, anonymize=False):
        """This person's real inputs, run through the CURRENT formulas.

        A COPY, always — see the block comment above. The inputs are the ones
        the payslip recorded when it was computed; the formulas are whatever the
        scheme says right now. That difference is the point: it is what lets a
        person try a formula change against a real case and see the effect
        before anything is applied to anyone.
        """
        empty = {'ok': False, 'sample_id': False, 'payslip_id': False,
                 'label': '', 'values': {}}
        config = self._pick_config(config_id)
        if not config:
            return empty
        slip = self.env['hr.payslip'].browse(int(payslip_id))
        if not slip.exists():
            return dict(empty, reason='gone')
        try:
            inputs = json.loads(slip.formula_input_values or '{}')
        except Exception:       # noqa: BLE001
            inputs = {}
        if not isinstance(inputs, dict):
            inputs = {}
        values = {}
        try:
            # `.new()` — an in-memory record that is never saved. It exists only
            # to give the shared evaluator a `config_id` to read; no row is
            # created, so previewing a hundred people leaves no trace.
            probe = self.env['hr.formula.sample.data'].new({'config_id': config.id})
            # RD48 — collect the errors THIS person's numbers produce. The
            # component's stored error belongs to whatever ran last with
            # diagnostics on (a sample), and a read-only preview writes none —
            # so without this the panel shows a sample's "division by zero"
            # beside a real Standard Working Hour of 198.
            failures = {}
            values = probe._evaluate_rules_with_dependencies(
                inputs, readonly=True, errors=failures)
        except Exception as exc:        # noqa: BLE001
            _logger.warning("RD46: preview failed for payslip %s: %s",
                            slip.id, exc)
            return dict(empty, reason='error')
        # RD52 — A COMPONENT'S OWN TYPE DECIDES WHETHER TO FLOAT IT.
        #
        # `float(value)` was applied to everything that would take it, and a
        # bank account number takes it: `float('0071002638698')` succeeds and
        # returns 71002638698.0 — the leading zeros gone, and then rendered as
        # ₫71,002,638,698 in the panel. ID card numbers, insurance book numbers
        # and PIT numbers all read as enormous sums of money on screen.
        #
        # This is the SAME fault VALUEKIND fixed in the import resolver
        # (`normalize_input_value`: "a component whose value is not a number is
        # now kept exactly as it arrived"), arriving again on a path written
        # after it. The rule's own `value_kind` is the answer, with
        # `is_text_component` as the belt-and-braces for a scheme whose
        # components were never classified.
        by_rule = {r.code: r for r in config.rule_ids if r.code}
        out = {}
        for code, value in (values or {}).items():
            rule = by_rule.get(code)
            col = rule.column_letter if rule else None
            if not col:
                continue
            # `wants_number` is the codebase's ONE definition of "should this
            # be a number", shared with the wire, the resolver and the payslip
            # line rail so they cannot drift into four opinions.
            if not value_kind_classifier.wants_number(
                    getattr(rule, 'value_kind', None)) \
                    or bool(getattr(rule, 'is_text_component', False)):
                out[col] = value if value not in (None, False) else ''
                continue
            try:
                out[col] = float(value)
            except (TypeError, ValueError):
                # Typed as a number but not one — keep what arrived rather than
                # pretending it is 0.
                out[col] = value
        return {
            'ok': True,
            'sample_id': False,
            'payslip_id': slip.id,
            'label': self._rd46_preview_label(slip, anonymize=anonymize),
            'anonymized': bool(anonymize),
            'values': out,
            # Keyed by CODE, because the error box is rendered from the selected
            # component and that is what it knows itself by.
            'errors': {code: msg for code, msg in (failures or {}).items()},
        }

    @api.model
    def preview_keep_as_sample(self, config_id, payslip_id, anonymize=True):
        """Turn the person currently previewed into a saved sample.

        The counterpart to the copy rule: previewing saves NOTHING, so keeping
        one has to be an explicit act. Anonymised by default — a saved sample
        outlives the question that produced it and is visible to everyone who
        can open the scheme.
        """
        if not self._can_edit():
            raise AccessError(_("You do not have permission to add samples."))
        config = self._pick_config(config_id)
        slip = self.env['hr.payslip'].browse(int(payslip_id))
        if not (config and slip.exists()):
            return {'ok': False}
        sample = self.env['hr.formula.sample.data'].create_from_payslip(
            slip, config, anonymize=bool(anonymize))
        return {'ok': True, 'sample_id': sample.id, 'name': sample.name}

    # ------------------------------------------------------------------
    # edit operations
    # ------------------------------------------------------------------
    # Fields the grid may bulk-edit across a column selection. Everything else is
    # rejected so a stray key can never mass-mutate formulas/codes/types.
    # COLROLES P2: `column_role` joins the list so a whole run of imported people
    # columns can be re-filed in one gesture. The model's write funnel stamps
    # column_role_source='user' for us (formula_rule.py write override, CR-A1).
    _BULK_FIELDS = {'category_id', 'number_format', 'appears_on_payslip',
                    'is_visible_in_grid', 'column_role'}

    @api.model
    def bulk_update_components(self, rule_ids, vals):
        """Apply whitelisted field changes to many components in ONE write.
        Non-whitelisted keys raise a UserError before anything is written, so a
        rejected call never leaves a partial update."""
        bad = set((vals or {}).keys()) - self._BULK_FIELDS
        if bad:
            raise UserError(_("These fields cannot be bulk-edited: %s") % ', '.join(sorted(bad)))
        rules = self.env['hr.formula.rule'].browse([int(i) for i in (rule_ids or [])]).exists()
        if not rules:
            return {'ok': False, 'msg': _("No components selected")}
        clean = {k: v for k, v in (vals or {}).items() if k in self._BULK_FIELDS}
        if clean:
            # F7: one 'bulk' version row per changed rule (write override loops self)
            rules.with_context(formula_version_reason='bulk').write(clean)
        return {'ok': True, 'updated': len(rules)}

    @api.model
    def _translate_formula_horizontal(self, formula, offset):
        """Shift every COLUMN-relative reference in ``formula`` by ``offset``
        columns (fill-right). ``$``-column-absolute refs (e.g. $D2) are left
        untouched; the row part (and any $ on it) is preserved verbatim."""
        if not offset:
            return formula

        def repl(m):
            col_dollar, col, rest = m.group(1), m.group(2), m.group(3)
            if col_dollar == '$':
                return m.group(0)                 # absolute column — unchanged
            n = self._col_num(col) + offset
            if n < 1:
                return m.group(0)                 # would fall off the left edge
            return col_dollar + self._num_to_col(n) + rest

        # (col-$)(letters)(row-$? digits) — matches D2, $D2, D$2, $D$2, AA11 …
        return re.sub(r'(\$?)([A-Za-z]+)(\$?\d+)', repl, formula)

    @api.model
    def _shift_rows(self, formula, to_row):
        """WP-L / S-L1 — rewrite every cell-ref ROW digit in ``formula`` to
        ``to_row`` (column letters + $ absolutes preserved). Thin wrapper over
        the pure engine helper so W41 (shift OUT: row 2 → sheet row N at export)
        and W17 (normalize IN: any row → 2 at paste) share ONE regex + literal
        mask — never two (S-I1 / D-J1). String literals are masked first."""
        from odoo.addons.pb_hr_payroll_formula.formula_engine import cell_refs
        return cell_refs.shift_rows(formula, to_row)

    @api.model
    def translate_formula(self, rule_id, target_column_letters):
        """Drag-fill preview: translate the source rule's formula to each target
        column. Returns ``[{col, proposed_formula, valid}]`` — nothing is written."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists() or rule.column_type != 'formula':
            return []
        src_num = self._col_num(rule.column_letter)
        formula = rule.excel_formula or ''
        config = rule.config_id
        by_col = {r.column_letter: r for r in config.rule_ids if r.column_letter}
        out = []
        for tgt in (target_column_letters or []):
            tgt = (tgt or '').upper()
            proposed = self._translate_formula_horizontal(formula, self._col_num(tgt) - src_num)
            refs = self._expand_refs(proposed, by_col)
            target_rule = by_col.get(tgt)
            valid = (all(c in by_col for c in refs)
                     and bool(target_rule) and target_rule.column_type == 'formula')
            out.append({'col': tgt, 'proposed_formula': proposed, 'valid': valid})
        return out

    def _run_tests_after_save(self, config, changed_codes=None):
        """W82 — re-run a config's sample tests once after a save operation and
        return the compact verdict for the studio's test chip. Never raises: a
        broken test run must not sink the save it rides on."""
        if not config or not config.exists():
            return {'has_tests': False, 'total': 0, 'passed': 0,
                    'failed': 0, 'pending': 0, 'failures': []}
        try:
            return config.run_sample_tests(changed_codes=changed_codes)
        except Exception as e:
            _logger.warning("run_sample_tests failed for config %s: %s", config.id, e)
            return {'has_tests': False, 'total': 0, 'passed': 0,
                    'failed': 0, 'pending': 0, 'failures': []}

    @api.model
    def bulk_save_formulas(self, items, reason='fill', note=False):
        """Persist several formulas at once. ``items`` = ``[{rule_id, formula}, ...]``.

        ``reason`` selects the F7 version-row reason for the whole batch — one of
        the batch-write reasons (``fill`` for drag-fill, ``bulk`` for find/replace,
        W14/TA.5). Any other value is coerced to ``fill`` so a bad caller can never
        mislabel history. ``note`` (e.g. ``find/replace: q → r``) is stamped on
        every version row. A shared ``formula_version_seen`` set keeps the batch to
        exactly N rows, one reason (C4), even though each rule is written twice
        (excel_formula is versioned, python_formula is not)."""
        reason = reason if reason in ('fill', 'bulk') else 'fill'
        Rule = self.env['hr.formula.rule']
        seen = set()
        saved = 0
        config = False
        changed_codes = []
        for it in (items or []):
            rule = Rule.browse(int(it.get('rule_id')))
            if not rule.exists() or rule.column_type != 'formula':
                continue
            config = rule.config_id
            column_map = {r.column_letter: r.code for r in config.rule_ids if r.column_letter}
            # F7: N formulas → N version rows, one reason, one shared seen-set (C4)
            ctx = {'formula_version_reason': reason, 'formula_version_seen': seen}
            if note:
                ctx['formula_version_note'] = note
            rule = rule.with_context(**ctx)
            try:
                rule.excel_formula = it.get('formula') or ''
                rule.python_formula = rule._convert_excel_to_python(rule.excel_formula, column_map)
                rule.is_valid = True
                rule.validation_message = ''
                saved += 1
                if rule.code:
                    changed_codes.append(rule.code)
            except Exception as e:
                _logger.debug("bulk_save_formulas skip %s: %s", rule.id, e)
        # W82: one test run for the whole batch (C4 one-batch rule), not per item.
        tests = self._run_tests_after_save(config, changed_codes)
        return {'ok': True, 'saved': saved, 'tests': tests}

    # ------------------------------------------------------------------
    # W17 smart paste — the ONE server ladder (D-L5): normalize + validate
    # ------------------------------------------------------------------
    @api.model
    def stage_paste(self, config_id, entries=None):
        """W17 (D-L5) — read-only. ``entries`` = ``[{col, text}]`` (a horizontal
        run mapped from the pasted clipboard). Returns
        ``{ok, entries:[{col, normalized, valid, msg}]}``. NOTHING is written; the
        client stages ``normalized`` as the ghost, so what you see is exactly what
        a later ``bulk_save_formulas`` commits — one ladder, no preview/commit
        divergence (the S-I1 live-proven bug class)."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False}
        by_col = {r.column_letter: r for r in config.rule_ids if r.column_letter}
        out = []
        for e in (entries or []):
            col = (e.get('col') or '').upper()
            text = (e.get('text') or '').strip()
            norm, valid, msg = self._normalize_paste_entry(col, text, config, by_col)
            out.append({'col': col, 'normalized': norm, 'valid': valid, 'msg': msg})
        return {'ok': True, 'entries': out}

    @api.model
    def _normalize_paste_entry(self, col, text, config, by_col):
        """Normalize + validate ONE pasted cell. Returns ``(normalized, valid,
        msg)``. Only base FORMULA columns are valid targets; a plain number is
        refused (constants live in their own row — v1 formulas-only); row digits
        are rewritten to the canonical row 2 (``B5*C5`` → ``B2*C2``, S-L1); then
        unknown column letters are named and the formula is run through the
        existing validate path (BRACKET-expanded)."""
        target = by_col.get(col)
        if not target or target.column_type != 'formula':
            return text, False, _("%s is not a formula column.") % (col or '?')
        if not text:
            return text, False, _("Empty cell.")
        # A plain number (no letters, no leading '=') — constants aren't pasted.
        if not text.startswith('=') and not re.search(r'[A-Za-z]', text):
            return text, False, _("Constants are edited in their own row.")
        # Normalize every row digit to the single grid formula row (row 2),
        # keeping the leading '=' (add one if the paste dropped it).
        norm = self._shift_rows(text, 2)
        if norm and not norm.startswith('='):
            norm = '=' + norm
        # Unknown letters → invalid, named (mirrors the drag-fill validity gate).
        refs = self._expand_refs(norm, by_col)
        unknown = sorted(c for c in refs if c not in by_col)
        if unknown:
            return norm, False, _("Unknown column(s): %s") % ', '.join(unknown)
        ok, vmsg = self._check_formula(config, norm, exclude_id=target.id)
        if not ok:
            return norm, False, vmsg or _("Invalid formula.")
        return norm, True, ''

    @api.model
    def save_formula(self, rule_id, excel_formula):
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': 'Component not found'}
        config = rule.config_id
        column_map = {r.column_letter: r.code for r in config.rule_ids if r.column_letter}
        try:
            rule.excel_formula = excel_formula
            if rule.column_type == 'formula':
                rule.python_formula = rule._convert_excel_to_python(excel_formula, column_map)
            rule.is_valid = True
            rule.validation_message = ''
        except Exception as e:
            return {'ok': False, 'msg': str(e)}
        return {'ok': True, 'tests': self._run_tests_after_save(config, [rule.code])}

    @api.model
    def update_component(self, rule_id, vals):
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False}
        allowed = {k: v for k, v in (vals or {}).items()
                   if k in ('name', 'number_format', 'constant_value', 'decimal_places', 'appears_on_payslip')}
        if allowed:
            rule.write(allowed)
        return {'ok': True}

    # ---- inline component editor -------------------------------------
    # Editable fields the inline editor may write. Computed/auto fields
    # (column_letter, python_formula, formula_dependencies, has_circular_ref)
    # and import-tracking readonly fields are deliberately excluded.
    _EDIT_FIELDS = (
        'name', 'code', 'column_type', 'sequence', 'category_id', 'salary_rule_id',
        'constant_value', 'data_source_field', 'default_value',
        'data_source', 'integration_connector_id', 'source_field_mapping',
        'number_format', 'decimal_places', 'column_width', 'text_align',
        'appears_on_payslip', 'is_visible_in_grid', 'report_visible',
        'is_required', 'is_editable', 'is_contract_component', 'requires_new_contract',
        'column_role',
    )
    _EDIT_M2O = ('category_id', 'salary_rule_id', 'integration_connector_id')

    @api.model
    def get_component_edit(self, rule_id):
        """Full editable + readonly-diagnostic snapshot for one component."""
        r = self.env['hr.formula.rule'].browse(int(rule_id))
        if not r.exists():
            return {'ok': False}
        return {
            'ok': True,
            'id': r.id,
            'column_letter': r.column_letter or '',
            'name': r.name or '',
            'code': r.code or '',
            'column_type': r.column_type or 'formula',
            'sequence': r.sequence or 0,
            'category_id': r.category_id.id or False,
            'salary_rule_id': r.salary_rule_id.id or False,
            'excel_formula': r.excel_formula or '',
            'constant_value': r.constant_value or 0.0,
            'data_source_field': r.data_source_field or '',
            'default_value': r.default_value or 0.0,
            'data_source': r.data_source or 'excel',
            'integration_connector_id': r.integration_connector_id.id or False,
            'source_field_mapping': r.source_field_mapping or '',
            'number_format': r.number_format or 'number',
            'decimal_places': r.decimal_places or 0,
            'column_width': r.column_width or 120,
            'text_align': r.text_align or 'right',
            'appears_on_payslip': bool(r.appears_on_payslip),
            'is_visible_in_grid': bool(r.is_visible_in_grid),
            'report_visible': bool(r.report_visible),
            'is_required': bool(r.is_required),
            'is_editable': bool(r.is_editable),
            'is_contract_component': bool(r.is_contract_component),
            'requires_new_contract': bool(r.requires_new_contract),
            # COLROLES P2 — role picker. `column_role_source` is readonly here: it
            # says whether a person already chose this role or the classifier did.
            'column_role': r.column_role or 'payroll',
            'column_role_source': r.column_role_source or 'auto',
            'is_text_component': bool(r.is_text_component),
            # readonly diagnostics
            'python_formula': r.python_formula or '',
            'formula_dependencies': r.formula_dependencies or '',
            'has_circular_ref': bool(r.has_circular_ref),
            'validation_message': r.validation_message or '',
        }

    @api.model
    def save_component(self, rule_id, vals):
        """Comprehensive save for the inline editor. Returns refreshed validity."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': 'Component not found'}
        vals = vals or {}
        write_vals = {}
        for k in self._EDIT_FIELDS:
            if k not in vals:
                continue
            v = vals[k]
            if k in self._EDIT_M2O:
                v = int(v) if v else False
            write_vals[k] = v
        # COLROLES / CR-A1 — a role arriving through the editor was chosen by a
        # person, so it is stamped as such HERE rather than trusted from the
        # client. (formula_rule.write would infer the same thing; saying it
        # explicitly means no future caller can quietly pass source='auto'.)
        if 'column_role' in write_vals and write_vals['column_role'] != rule.column_role:
            write_vals['column_role_source'] = 'user'
        # proactive duplicate-code guard (the DB unique constraint is not
        # reliably enforced on this table, so check here).
        if write_vals.get('code'):
            dup = rule.config_id.rule_ids.filtered(
                lambda r: r.id != rule.id and (r.code or '') == write_vals['code'])
            if dup:
                return {'ok': False,
                        'msg': 'Code "%s" is already used by column %s.'
                               % (write_vals['code'], dup[0].column_letter or '?')}
        # excel_formula handled separately so we can convert + validate
        new_formula = vals.get('excel_formula')
        # F7: this method may write metadata AND the formula in two steps — a
        # shared 'seen' set collapses both into ONE version row for the rule.
        rule = rule.with_context(formula_version_reason='edit',
                                 formula_version_seen=set())
        try:
            if write_vals:
                rule.write(write_vals)
            ctype = write_vals.get('column_type', rule.column_type)
            if new_formula is not None:
                rule.excel_formula = new_formula
            if ctype == 'formula':
                column_map = {r.column_letter: r.code
                              for r in rule.config_id.rule_ids if r.column_letter}
                rule.python_formula = rule._convert_excel_to_python(rule.excel_formula or '', column_map)
                ok, msg = self._check_formula(rule.config_id, rule.excel_formula or '', exclude_id=rule.id)
                rule.is_valid = ok
                rule.validation_message = '' if ok else msg
        except Exception as e:
            return {'ok': False, 'msg': str(e)}
        return {'ok': True, 'is_valid': bool(rule.is_valid),
                'validation_message': rule.validation_message or ''}

    # =====================================================================
    #  F7 — Formula version history (rail + token diff + restore)
    # =====================================================================
    @api.model
    def _tokenize_text(self, formula):
        """Lex a raw Excel formula string into a flat token list for diffing.
        Broader than `_tokenize` (which builds chips for a live rule): this also
        captures bare function names (ROUND/VLOOKUP), commas and comparison
        operators, so a diff reads sensibly across structural rewrites."""
        formula = (formula or '').lstrip('=').strip()
        if not formula:
            return []
        return re.findall(r'[A-Za-z_]+\$?\d*|\$?\d+\.?\d*|[+\-*/()%,^&<>=!:]', formula)

    def _version_row_payload(self, ver):
        try:
            snap = json.loads(ver.snapshot_json or '{}')
        except Exception:
            snap = {}
        return {
            'seq': ver.seq,
            'reason': ver.reason,
            'reason_label': dict(ver._fields['reason']._description_selection(self.env)).get(ver.reason, ver.reason),
            'note': ver.note or '',
            'user': ver.user_id.name or '',
            'date': fields.Datetime.to_string(ver.create_date) if ver.create_date else '',
            'excel_formula': ver.excel_formula or '',
            'snapshot': snap,
        }

    @api.model
    def get_rule_history(self, rule_id):
        """Full history for one rule: the live state as a synthetic 'current'
        node plus every stored version (newest first). Versions hold OUTGOING
        pre-edit states, so `current` is the head and each version is a past."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'versions': []}
        versions = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id)], order='seq desc')
        return {
            'ok': True,
            'rule_id': rule.id,
            'code': rule.code or '',
            'name': rule.name or '',
            'config_name': rule.config_id.display_name or rule.config_id.name or '',
            'current': {
                'seq': None,           # None == the live head
                'excel_formula': rule.excel_formula or '',
                'user': (rule.write_uid.name if rule.write_uid else ''),
                'date': fields.Datetime.to_string(rule.write_date) if rule.write_date else '',
                'snapshot': rule._version_snapshot(),
            },
            'versions': [self._version_row_payload(v) for v in versions],
        }

    def _version_formula(self, rule, seq):
        """Resolve a seq (int) or None (=live head) to its Excel formula text."""
        if seq in (None, False, 'current'):
            return rule.excel_formula or '', _('Current')
        ver = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id), ('seq', '=', int(seq))], limit=1)
        if not ver:
            return '', _('v%s') % seq
        return ver.excel_formula or '', _('v%s') % seq

    @api.model
    def diff_versions(self, rule_id, seq_a, seq_b):
        """Token-level diff between two versions (or a version and 'current').
        Returns runs of equal/insert/delete/replace for chip rendering. `seq_*`
        may be an int seq or null/'current' for the live head. A precedes B in
        reading order (A = older, B = newer)."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'runs': []}
        fa, la = self._version_formula(rule, seq_a)
        fb, lb = self._version_formula(rule, seq_b)
        runs = self._token_diff_runs(self._tokenize_text(fa), self._tokenize_text(fb))
        return {'ok': True, 'runs': runs, 'a_label': la, 'b_label': lb,
                'a_formula': fa, 'b_formula': fb}

    def _token_diff_runs(self, a, b):
        """LCS diff over two token lists → merged runs. Adjacent delete+insert
        collapse into a single 'replace' so `0.10 → 0.12` reads as one change."""
        n, m = len(a), len(b)
        # LCS length table
        dp = [[0] * (m + 1) for _ in range(n + 1)]
        for i in range(n - 1, -1, -1):
            for j in range(m - 1, -1, -1):
                dp[i][j] = (dp[i + 1][j + 1] + 1) if a[i] == b[j] \
                    else max(dp[i + 1][j], dp[i][j + 1])
        # walk to emit ops
        ops = []
        i = j = 0
        while i < n and j < m:
            if a[i] == b[j]:
                ops.append(('equal', a[i])); i += 1; j += 1
            elif dp[i + 1][j] >= dp[i][j + 1]:
                ops.append(('delete', a[i])); i += 1
            else:
                ops.append(('insert', b[j])); j += 1
        while i < n:
            ops.append(('delete', a[i])); i += 1
        while j < m:
            ops.append(('insert', b[j])); j += 1
        # coalesce consecutive ops of the same kind, then fuse delete+insert
        merged = []
        for op, tok in ops:
            if merged and merged[-1]['op'] == op and 'tokens' in merged[-1]:
                merged[-1]['tokens'].append(tok)
            else:
                merged.append({'op': op, 'tokens': [tok]})
        runs = []
        k = 0
        while k < len(merged):
            cur = merged[k]
            nxt = merged[k + 1] if k + 1 < len(merged) else None
            if cur['op'] == 'delete' and nxt and nxt['op'] == 'insert':
                runs.append({'op': 'replace', 'old': cur['tokens'], 'new': nxt['tokens']})
                k += 2
            elif cur['op'] == 'insert' and nxt and nxt['op'] == 'delete':
                runs.append({'op': 'replace', 'old': nxt['tokens'], 'new': cur['tokens']})
                k += 2
            else:
                runs.append(cur)
                k += 1
        return runs

    @api.model
    def restore_version(self, rule_id, seq):
        """Write a past version's formula back onto the live rule. This is itself
        a versioned event (reason='restore'), so history is never rewritten —
        the current head is snapshotted before being overwritten."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _('Component not found')}
        ver = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id), ('seq', '=', int(seq))], limit=1)
        if not ver:
            return {'ok': False, 'msg': _('Version not found')}
        target = ver.excel_formula or ''
        rule = rule.with_context(formula_version_reason='restore',
                                 formula_version_note=_('Restored v%s') % seq)
        try:
            rule.excel_formula = target
            if rule.column_type == 'formula':
                column_map = {r.column_letter: r.code
                              for r in rule.config_id.rule_ids if r.column_letter}
                rule.python_formula = rule._convert_excel_to_python(target, column_map)
                ok, msg = self._check_formula(rule.config_id, target, exclude_id=rule.id)
                rule.is_valid = ok
                rule.validation_message = '' if ok else msg
        except Exception as e:
            return {'ok': False, 'msg': str(e)}
        return {'ok': True, 'excel_formula': target,
                'is_valid': bool(rule.is_valid),
                'tests': self._run_tests_after_save(rule.config_id, [rule.code])}

    @api.model
    def get_config_milestones(self, config_id):
        """Milestones for a config, newest first, for the compare picker."""
        ms = self.env['hr.formula.config.milestone'].sudo().search(
            [('config_id', '=', int(config_id))], order='milestone_date desc')
        return [{
            'id': m.id, 'name': m.name,
            'date': fields.Datetime.to_string(m.milestone_date),
            'user': m.user_id.name or '',
        } for m in ms]

    def _formula_at(self, rule, when):
        """The rule's Excel formula in effect at datetime `when`. Version rows
        store OUTGOING states, so the value live at T is the earliest version
        captured at-or-after T; if none, nothing changed since T → current."""
        ver = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id), ('create_date', '>=', when)],
            order='create_date asc, seq asc', limit=1)
        return (ver.excel_formula if ver else rule.excel_formula) or ''

    @api.model
    def compare_to_milestone(self, config_id, milestone_id):
        """Diff a whole config against a milestone: only rules whose formula
        changed since the milestone, each with its token diff."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        milestone = self.env['hr.formula.config.milestone'].sudo().browse(int(milestone_id))
        if not config.exists() or not milestone.exists():
            return {'ok': False, 'changed': []}
        when = milestone.milestone_date
        changed = []
        for rule in config.rule_ids:
            old = self._formula_at(rule, when)
            cur = rule.excel_formula or ''
            if (old or '') == (cur or ''):
                continue
            changed.append({
                'rule_id': rule.id,
                'code': rule.code or '',
                'name': rule.name or '',
                'col': rule.column_letter or '',
                'old_formula': old,
                'cur_formula': cur,
                'runs': self._token_diff_runs(
                    self._tokenize_text(old), self._tokenize_text(cur)),
            })
        return {
            'ok': True,
            'milestone_name': milestone.name,
            'milestone_date': fields.Datetime.to_string(when),
            'changed_count': len(changed),
            'changed': changed,
        }

    # ==================================================================
    # B3 — Release bundles + sign-off (a query over F7 versions)
    # ==================================================================
    def _last_milestone(self, config):
        return self.env['hr.formula.config.milestone'].sudo().search(
            [('config_id', '=', config.id)], order='milestone_date desc, id desc', limit=1)

    def _seal_milestone(self, config, name):
        """Record a milestone carrying the config's version high-water mark — the
        max ``hr.formula.rule.version`` id at seal time. This is the exact,
        collision-free boundary for 'changed since this milestone' (W86): unlike
        ``milestone_date`` (second-precision, so it can't be separated from edits
        sealed in the same second) the id boundary is unambiguous even when the
        seal and its edits share one transaction — the one-action-rollback case."""
        Ver = self.env['hr.formula.rule.version'].sudo()
        last = Ver.search([('config_id', '=', config.id)], order='id desc', limit=1)
        return self.env['hr.formula.config.milestone'].sudo().create({
            'config_id': config.id, 'name': name,
            'milestone_date': fields.Datetime.now(),
            'version_hwm': last.id if last else 0})

    def _config_version_hwm(self, config):
        """The current max version id for a config (the 'now' boundary)."""
        last = self.env['hr.formula.rule.version'].sudo().search(
            [('config_id', '=', config.id)], order='id desc', limit=1)
        return last.id if last else 0

    def _ms_hwm(self, ms):
        """Version-id boundary for a milestone: its stored hwm, or — for a legacy
        milestone sealed before W86 — the max version id at-or-before its
        timestamp (reliable there because legacy releases were sealed in a
        SEPARATE request from their edits, so no same-second collision)."""
        if not ms:
            return 0
        if ms.version_hwm is not None and ms.version_hwm >= 0:
            return ms.version_hwm
        last = self.env['hr.formula.rule.version'].sudo().search(
            [('config_id', '=', ms.config_id.id),
             ('create_date', '<=', ms.milestone_date)], order='id desc', limit=1)
        return last.id if last else 0

    def _formula_at_ver(self, rule, from_hwm):
        """The rule's Excel formula in effect just after version boundary
        ``from_hwm`` — the earliest version for this rule with id > from_hwm (its
        OUTGOING snapshot = the live state at that boundary); none → current.
        ``from_hwm`` 0 = start of history (the original formula)."""
        ver = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id), ('id', '>', from_hwm)], order='id asc', limit=1)
        return (ver.excel_formula if ver else rule.excel_formula) or ''

    def _constant_at_ver(self, rule, from_hwm):
        """The rule's ``constant_value`` in effect just after version boundary
        ``from_hwm`` (from the same OUTGOING snapshot as ``_formula_at_ver``)."""
        ver = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id), ('id', '>', from_hwm)], order='id asc', limit=1)
        if ver:
            try:
                return float(json.loads(ver.snapshot_json or '{}').get('constant_value') or 0.0)
            except Exception:
                pass
        return rule.constant_value or 0.0

    def _changes_between_ver(self, config, from_hwm, to_hwm=None):
        """Rules whose Excel formula OR constant differs between two version
        boundaries (D-C5). ``from_hwm`` 0 = start of history; ``to_hwm`` None =
        the live current state. Replaces the timestamp comparison so a milestone
        can never lose an edit to second-granularity (W86)."""
        Ver = self.env['hr.formula.rule.version'].sudo()
        changed = []
        for rule in config.rule_ids:
            old_f = self._formula_at_ver(rule, from_hwm)
            cur_f = self._formula_at_ver(rule, to_hwm) if to_hwm else (rule.excel_formula or '')
            old_c = self._constant_at_ver(rule, from_hwm)
            cur_c = self._constant_at_ver(rule, to_hwm) if to_hwm else (rule.constant_value or 0.0)
            f_changed = (old_f or '') != (cur_f or '')
            c_changed = abs((old_c or 0.0) - (cur_c or 0.0)) > 1e-9
            if not f_changed and not c_changed:
                continue
            dom = [('rule_id', '=', rule.id), ('id', '>', from_hwm)]
            if to_hwm:
                dom.append(('id', '<=', to_hwm))
            v = Ver.search(dom, order='id desc', limit=1)
            changed.append({
                'rule_id': rule.id, 'code': rule.code or '',
                'name': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or '',
                'col': rule.column_letter or '', 'group': _group_for(rule),
                'type': rule.column_type,
                'old_formula': old_f, 'cur_formula': cur_f,
                'old_constant': old_c, 'cur_constant': cur_c,
                'formula_changed': f_changed, 'constant_changed': c_changed,
                'reason': v.reason if v else 'edit',
                'runs': self._token_diff_runs(self._tokenize_text(old_f), self._tokenize_text(cur_f))
                        if f_changed else [],
            })
        return changed

    def _formula_original(self, rule):
        """The rule's formula at the very start of its history — the earliest
        version's OUTGOING snapshot (each row is pre-edit), else the current if
        it was never edited. Used when there is no prior milestone to anchor to."""
        first = self.env['hr.formula.rule.version'].sudo().search(
            [('rule_id', '=', rule.id)], order='create_date asc, seq asc', limit=1)
        return (first.excel_formula if first else rule.excel_formula) or ''

    def _changes_between(self, config, from_when, to_when):
        """Rules whose formula differs between two instants (each read from the
        F7 version snapshots via _formula_at). ``from_when`` None = the start of
        history (original formula); ``to_when`` None = the live current formula."""
        changed = []
        for rule in config.rule_ids:
            old = self._formula_at(rule, from_when) if from_when else self._formula_original(rule)
            cur = self._formula_at(rule, to_when) if to_when else (rule.excel_formula or '')
            if (old or '') == (cur or ''):
                continue
            # most recent version reason for this rule inside the window (why it changed)
            dom = [('rule_id', '=', rule.id)]
            if from_when:
                dom.append(('create_date', '>=', from_when))
            if to_when:
                dom.append(('create_date', '<', to_when))
            v = self.env['hr.formula.rule.version'].sudo().search(
                dom, order='create_date desc, seq desc', limit=1)
            changed.append({
                'rule_id': rule.id, 'code': rule.code or '',
                'name': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or '',
                'col': rule.column_letter or '', 'group': _group_for(rule),
                'old_formula': old, 'cur_formula': cur,
                'reason': v.reason if v else 'edit',
                'runs': self._token_diff_runs(self._tokenize_text(old), self._tokenize_text(cur)),
            })
        return changed

    def _draft_release_narrative(self, config, changes):
        """A prose changelog — LLM if available, else a deterministic template."""
        if not changes:
            return _("No formula changes since the last milestone.")
        # deterministic fallback (also the LLM's structured source)
        reason_label = {'edit': 'edited', 'bulk': 'bulk-edited', 'fill': 'drag-filled',
                        'import': 'imported', 'restore': 'restored', 'rename': 'renamed',
                        'lifecycle': 'lifecycle', 'legislation': 'legislation pack'}
        lines = ['- %s (%s): %s' % (c['name'], c['col'], reason_label.get(c['reason'], 'edited'))
                 for c in changes]
        fallback = _("This release updates %s component(s):\n%s") % (len(changes), '\n'.join(lines))
        try:
            summary = '\n'.join('%s [%s] %s → %s' % (c['col'], c['reason'],
                                                     (c['old_formula'] or '(none)'),
                                                     (c['cur_formula'] or '(none)'))
                                for c in changes)
            msgs = [
                {'role': 'system', 'content':
                 "You are a payroll release manager. Write a concise, professional changelog "
                 "(3-6 short bullet points, plain business language, no code) summarising these "
                 "payroll formula changes for a sign-off reviewer. Do not invent numbers."},
                {'role': 'user', 'content': "Config: %s\nChanges:\n%s" % (config.name, summary)},
            ]
            out = self._llm_chat(msgs)
            return (out or '').strip() or fallback
        except Exception:
            return fallback

    @api.model
    def release_preview(self, config_id=None):
        """The pending release: everything changed since the last milestone, with
        diffs and a drafted changelog. Nothing is written."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        last = self._last_milestone(config)
        # id boundary (W86) so a constant-only change (legislation pack) is
        # releasable AND rollback-able (D-C5), with no second-granularity loss.
        from_hwm = self._ms_hwm(last)
        changes = self._changes_between_ver(config, from_hwm, None)
        return {
            'ok': True,
            'config': {'id': config.id, 'name': config.name, 'state': config.state},
            'from_milestone': ({'id': last.id, 'name': last.name,
                                'date': fields.Datetime.to_string(last.milestone_date)}
                               if last else None),
            'change_count': len(changes),
            'changes': changes,
            'narrative': self._draft_release_narrative(config, changes),
            'can_edit': self._can_edit(),
        }

    @api.model
    def release_approve(self, config_id, narrative=None):
        """Sign off the pending release: seal an immutable milestone and record
        the release (with its F7 version rows for provenance)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to sign off releases.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        last = self._last_milestone(config)
        from_hwm = self._ms_hwm(last)
        changes = self._changes_between_ver(config, from_hwm, None)
        if not changes:
            return {'ok': False, 'reason': 'no_changes',
                    'msg': _("There are no changes to release since the last milestone.")}
        Release = self.env['hr.formula.release']
        n = Release.search_count([('config_id', '=', config.id)]) + 1
        # Seal at the current version high-water mark (W86) so a later rollback of
        # this release reads an exact 'from' boundary — see _seal_milestone.
        to_ms = self._seal_milestone(config, _("Release v%s") % n)
        vdom = [('config_id', '=', config.id), ('id', '>', from_hwm)]
        versions = self.env['hr.formula.rule.version'].sudo().search(vdom)
        rel = Release.create({
            'name': _("Release v%s") % n,
            'config_id': config.id,
            'from_milestone_id': last.id if last else False,
            'to_milestone_id': to_ms.id,
            'narrative': (narrative or '').strip() or self._draft_release_narrative(config, changes),
            'change_count': len(changes),
            'version_ids': [(6, 0, versions.ids)],
        })
        return {'ok': True, 'release_id': rel.id, 'change_count': len(changes)}

    @api.model
    def list_releases(self, config_id=None):
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'releases': []}
        rels = self.env['hr.formula.release'].search([('config_id', '=', config.id)])
        # W86 — only the latest release is rollback-eligible (D-C4); flag it so the
        # UI shows the Rollback button on exactly one row.
        latest = self._last_release(config)
        return {'ok': True, 'latest_id': latest.id or False, 'can_edit': self._can_edit(),
                'releases': [{
                    'id': r.id, 'name': r.name,
                    'approved_by': r.approved_by_id.name or '',
                    'approved_date': fields.Datetime.to_string(r.approved_date) if r.approved_date else '',
                    'change_count': r.change_count,
                    'narrative': r.narrative or '',
                    'from': r.from_milestone_id.name or '(start)',
                    'to': r.to_milestone_id.name or '',
                    'is_latest': r.id == latest.id,
                } for r in rels]}

    @api.model
    def release_detail(self, release_id):
        """Re-derive a past release's diffs from its two milestone boundaries."""
        rel = self.env['hr.formula.release'].browse(int(release_id))
        if not rel.exists():
            return {'ok': False}
        from_hwm = self._ms_hwm(rel.from_milestone_id)
        to_hwm = self._ms_hwm(rel.to_milestone_id) if rel.to_milestone_id else None
        changes = self._changes_between_ver(rel.config_id, from_hwm, to_hwm)
        return {'ok': True, 'name': rel.name, 'narrative': rel.narrative or '',
                'approved_by': rel.approved_by_id.name or '',
                'approved_date': fields.Datetime.to_string(rel.approved_date) if rel.approved_date else '',
                'change_count': len(changes), 'changes': changes}

    # ==================================================================
    # W86 — One-action rollback (revert the latest release atomically)
    # ==================================================================
    # Rollback of release vN ≡ restore the config to milestone `from` of vN
    # (D-C4). It is itself a versioned + released event (reason='restore', a new
    # milestone + audit release row) so history is never rewritten and a second
    # rollback round-trips cleanly. Constants are reverted too (D-C5): a
    # legislation pack edits `constant_value`, so a formula-only rollback would
    # silently keep a new SI cap.
    def _last_release(self, config):
        return self.env['hr.formula.release'].search(
            [('config_id', '=', config.id)], order='approved_date desc, id desc', limit=1)

    def _restore_rule_state(self, rule, excel_formula, constant_value):
        """Write a past (formula, constant) back onto a live rule. Mirrors
        ``restore_version``'s python-rebuild + validity check for the formula path
        (``pb_formula_studio.py`` restore) and ADDS the constant path (net-new,
        D-C5). A restored formula that no longer converts RAISES here so the
        caller's savepoint aborts loudly — never a half-applied rollback (C7)."""
        vals = {}
        if rule.column_type == 'formula':
            column_map = {r.column_letter: r.code for r in rule.config_id.rule_ids if r.column_letter}
            py = rule._convert_excel_to_python(excel_formula or '', column_map)   # may raise → savepoint aborts
            ok, msg = self._check_formula(rule.config_id, excel_formula or '', exclude_id=rule.id)
            vals.update({'excel_formula': excel_formula or '', 'python_formula': py,
                         'is_valid': ok, 'validation_message': '' if ok else msg})
        elif excel_formula is not None:
            vals['excel_formula'] = excel_formula or ''
        if constant_value is not None:
            vals['constant_value'] = constant_value
        if vals:
            rule.write(vals)

    def _rollback_guard(self, rel):
        """D-C4 eligibility: only the latest release, and only when nothing is
        unreleased (else 'rollback of vN' is ambiguous). Returns {ok[, reason,
        msg]}."""
        config = rel.config_id
        latest = self._last_release(config)
        if not latest or rel.id != latest.id:
            return {'ok': False, 'reason': 'not_latest',
                    'msg': _("Only the latest release can be rolled back.")}
        to_hwm = self._ms_hwm(rel.to_milestone_id)
        unreleased = self._changes_between_ver(config, to_hwm, None)
        if unreleased:
            return {'ok': False, 'reason': 'unreleased',
                    'msg': _("Release or discard the current changes first "
                             "(%d unreleased change(s)).") % len(unreleased)}
        return {'ok': True}

    def _rollback_overrides(self, changes):
        """Split a change list into the formula + constant overrides that seed a
        simulate-before-apply run (the rollback previews its OLD state)."""
        formula_overrides = {c['code']: c['old_formula']
                             for c in changes if c['formula_changed'] and c['code']}
        value_overrides = {c['code']: c['old_constant']
                           for c in changes if c['constant_changed'] and c['code']}
        return formula_overrides, value_overrides

    @api.model
    def rollback_preview(self, release_id):
        """What rolling back a release would revert: eligibility (+ block reason),
        the formula/constant change list, and the simulate overrides (D-C6).
        Nothing is written."""
        rel = self.env['hr.formula.release'].browse(int(release_id))
        if not rel.exists():
            return {'ok': False}
        config = rel.config_id
        guard = self._rollback_guard(rel)
        from_hwm = self._ms_hwm(rel.from_milestone_id)
        changes = self._changes_between_ver(config, from_hwm, None)
        formula_overrides, value_overrides = self._rollback_overrides(changes)
        return {
            'ok': True,
            'eligible': guard['ok'],
            'block_reason': '' if guard['ok'] else guard.get('msg', ''),
            'release': {'id': rel.id, 'name': rel.name, 'narrative': rel.narrative or '',
                        'approved_by': rel.approved_by_id.name or '',
                        'approved_date': fields.Datetime.to_string(rel.approved_date)
                                         if rel.approved_date else ''},
            'change_count': len(changes),
            'changes': changes,
            'formula_overrides': formula_overrides,
            'value_overrides': value_overrides,
            'can_edit': self._can_edit(),
        }

    @api.model
    def rollback_simulate_prepare(self, release_id, limit=None):
        """Seed a simulation with the rollback's OLD formulas + constants and
        return the payslip work-list (drive it via the shared simulate_batch /
        simulate_result RPCs). Shows the org-wide impact before Apply arms."""
        rel = self.env['hr.formula.release'].browse(int(release_id))
        if not rel.exists():
            return {'ok': False}
        config = rel.config_id
        from_hwm = self._ms_hwm(rel.from_milestone_id)
        changes = self._changes_between_ver(config, from_hwm, None)
        formula_overrides, value_overrides = self._rollback_overrides(changes)
        Sim = self.env['hr.formula.simulation']
        created = Sim.sim_create(config.id, overrides=formula_overrides,
                                 value_overrides=value_overrides)
        if not created.get('ok'):
            return created
        prep = Sim.sim_prepare(created['sim_id'], limit=limit)
        prep.update({'ok': True, 'headline': created.get('headline'),
                     'overrides': created.get('overrides', 0)})
        return prep

    @api.model
    def rollback_apply(self, release_id):
        """Revert the latest release atomically (S-C1). Restores every changed
        rule's formula AND constant to its at-`from`-milestone value in one
        savepoint (all-or-nothing), records a 'Rollback of vN' milestone + audit
        release row, and re-runs the sample tests (W82 — a rollback is a save)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to roll back releases.")}
        rel = self.env['hr.formula.release'].browse(int(release_id))
        if not rel.exists():
            return {'ok': False, 'msg': _('Release not found')}
        config = rel.config_id
        guard = self._rollback_guard(rel)
        if not guard['ok']:
            return guard
        from_hwm = self._ms_hwm(rel.from_milestone_id)
        changes = self._changes_between_ver(config, from_hwm, None)
        if not changes:
            return {'ok': False, 'reason': 'nothing', 'msg': _('Nothing to roll back.')}
        seen = set()
        pre_hwm = self._config_version_hwm(config)     # id boundary before the restore writes
        ctx = dict(formula_version_reason='restore',
                   formula_version_note=_('Rollback %s') % rel.name,
                   formula_version_seen=seen)
        try:
            with self.env.cr.savepoint():
                for ch in changes:
                    rule = self.env['hr.formula.rule'].browse(ch['rule_id']).with_context(**ctx)
                    self._restore_rule_state(rule, ch['old_formula'], ch['old_constant'])
        except Exception as e:
            _logger.warning("rollback_apply failed on %s: %s", config.code, e)
            return {'ok': False, 'msg': str(e)}
        # audit: the rollback IS a release (D-C6) — milestone + release row, from
        # the rolled-back release's `to` up to the new rollback milestone. The
        # provenance rows are exactly the restore versions just created (id > the
        # pre-restore boundary), and _seal_milestone stamps the new milestone at
        # the post-restore hwm so rolling back THIS rollback reads a clean
        # boundary — the double-rollback round-trip (D-C4).
        versions = self.env['hr.formula.rule.version'].sudo().search(
            [('config_id', '=', config.id), ('id', '>', pre_hwm)])
        to_ms = self._seal_milestone(config, _('Rollback of %s') % rel.name)
        Release = self.env['hr.formula.release']
        audit = Release.create({
            'name': _('Rollback of %s') % rel.name,
            'config_id': config.id,
            'from_milestone_id': rel.to_milestone_id.id if rel.to_milestone_id else False,
            'to_milestone_id': to_ms.id,
            'narrative': self._draft_release_narrative(config, changes),
            'change_count': len(changes),
            'version_ids': [(6, 0, versions.ids)],
        })
        tests = self._run_tests_after_save(config)
        return {'ok': True, 'restored': len(seen), 'release_id': audit.id, 'tests': tests}

    # ==================================================================
    # W97 — Period comparison (read-only chunked aggregation of two payruns)
    # ==================================================================
    @api.model
    def compare_runs(self, config_id=None):
        """The payslip runs comparable for a config: those carrying this config's
        formula slips, newest period first. Feeds the two run pickers."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'runs': []}
        slips = self.env['hr.payslip'].sudo().search([
            ('formula_config_id', '=', config.id),
            ('calculation_method', '=', 'formula'),
            ('payslip_run_id', '!=', False)])
        counts = defaultdict(int)
        for s in slips:
            counts[s.payslip_run_id.id] += 1
        runs = self.env['hr.payslip.run'].sudo().browse(list(counts.keys())).exists()
        items = sorted(([{
            'id': r.id, 'name': r.name or '',
            'date_start': str(r.date_start or ''), 'date_end': str(r.date_end or ''),
            'slips': counts.get(r.id, 0),
        } for r in runs]), key=lambda x: (x['date_start'], x['name']), reverse=True)
        return {'ok': True, 'config': {'id': config.id, 'name': config.display_name},
                'runs': items,
                'currency': config.currency_id.symbol if config.currency_id else ''}

    @api.model
    def compare_prepare(self, config_id, run_a_id, run_b_id):
        """Create a comparison and return the matched slip-pair work-list to drive
        through it in chunks (mirrors simulate_prepare)."""
        Cmp = self.env['hr.formula.period.comparison']
        created = Cmp.cmp_create(config_id, run_a_id, run_b_id)
        if not created.get('ok'):
            return created
        prep = Cmp.cmp_prepare(created['cmp_id'])
        prep.update({'ok': True, 'headline': created.get('headline')})
        return prep

    @api.model
    def compare_batch(self, payload):
        return self.env['hr.formula.period.comparison'].cmp_batch(payload or {})

    @api.model
    def compare_result(self, cmp_id):
        return self.env['hr.formula.period.comparison'].cmp_finalize(cmp_id)

    @api.model
    def compare_drop(self, cmp_id):
        cmp = self.env['hr.formula.period.comparison'].browse(int(cmp_id))
        cmp.cmp_drop()
        return {'ok': True}

    # ==================================================================
    # W95 (WP-H) — component budgets (vs-actual variance in the compare view)
    # Reads are open; writes are manager-gated (D-H2), same split as snippets.
    # ==================================================================
    def _budget_payload(self, b):
        return {'id': b.id, 'name': b.name or '',
                'period_label': b.period_label or '', 'note': b.note or '',
                'line_count': len(b.line_ids)}

    @api.model
    def budget_list(self, config_id=None):
        """Budgets authored for a config, newest first. Feeds the budget picker."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'budgets': []}
        budgets = self.env['hr.formula.budget'].search(
            [('config_id', '=', config.id)], order='id desc')
        return {'ok': True, 'config': {'id': config.id, 'name': config.display_name},
                'budgets': [self._budget_payload(b) for b in budgets],
                'can_edit': self._can_edit(),
                'currency': config.currency_id.symbol if config.currency_id else ''}

    @api.model
    def budget_get(self, config_id, budget_id=None):
        """Editor payload: every config component (code, name, group) with its
        current budget amount, PLUS orphan budget lines whose code no longer
        exists in the config (D-H2 honesty — surfaced, never dropped). A falsy
        ``budget_id`` returns a blank editor over the config's components."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        try:
            if not config.exists():
                return {'ok': False}
            config.check_access('read')
        except AccessError:
            return {'ok': False, 'msg': _('No access to this configuration.')}
        amounts = {}
        budget = None
        if budget_id:
            budget = self.env['hr.formula.budget'].browse(int(budget_id)).exists()
            if budget:
                for l in budget.line_ids:
                    if l.code:
                        amounts[l.code] = l.amount
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        rule_codes = set()
        components = []
        for r in rules:
            if not r.code:
                continue
            rule_codes.add(r.code)
            components.append({
                'code': r.code,
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                'group': _group_for(r),
                'type': r.column_type,
                'amount': amounts.get(r.code, 0.0),
            })
        orphans = []
        for code, amt in amounts.items():
            if code not in rule_codes:
                orphans.append({'code': code, 'amount': amt})
        orphans.sort(key=lambda o: o['code'])
        return {
            'ok': True,
            'budget': (self._budget_payload(budget) if budget else None),
            'components': components,
            'orphans': orphans,
            'can_edit': self._can_edit(),
            'currency': config.currency_id.symbol if config.currency_id else '',
        }

    @api.model
    def budget_seed_from_run(self, config_id, run_id):
        """Per-component actual sums for a run keyed by code (D-H2 Seed-from-run).
        Pure read of stored slips via the engine helper — open to all."""
        sums = self.env['hr.formula.period.comparison'].run_component_sums(config_id, run_id)
        return {'ok': True, 'amounts': sums}

    @api.model
    def budget_save(self, vals):
        """Create/replace a budget from the editor (manager-gated). The client
        sends the FULL desired line set (config amounts + any kept orphans); the
        server validates every value (numeric, |v| <= 1e12 — same rule as W49)
        and replaces line_ids wholesale so removed rows disappear atomically."""
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can edit budgets.')}
        config = self.env['hr.formula.config'].browse(int(vals.get('config_id') or 0))
        if not config.exists():
            return {'ok': False, 'msg': _('Configuration not found.')}
        name = (vals.get('name') or '').strip()
        if not name:
            return {'ok': False, 'msg': _('A budget needs a name.')}
        # Validate lines: {code: amount}. Reject non-numeric / absurd values loudly.
        raw = vals.get('lines') or {}
        clean = {}
        for code, amount in (raw.items() if isinstance(raw, dict) else []):
            code = (str(code) or '').strip()
            if not code:
                continue
            n = self._as_num(amount)
            if n is None:
                return {'ok': False, 'msg': _('Budget amount for %s is not a number.') % code}
            if abs(n) > 1e12:
                return {'ok': False, 'msg': _('Budget amount for %s is out of range.') % code}
            clean[code] = n
        Budget = self.env['hr.formula.budget']
        bid = vals.get('id')
        head = {'name': name, 'config_id': config.id,
                'period_label': (vals.get('period_label') or '').strip(),
                'note': (vals.get('note') or '').strip()}
        if bid:
            budget = Budget.browse(int(bid))
            if not budget.exists():
                return {'ok': False, 'msg': _('Budget not found.')}
            budget.write(head)
            budget.line_ids.unlink()
        else:
            budget = Budget.create(head)
        Line = self.env['hr.formula.budget.line']
        for code, amount in clean.items():
            Line.create({'budget_id': budget.id, 'code': code, 'amount': amount})
        return {'ok': True, 'budget': self._budget_payload(budget)}

    @api.model
    def budget_delete(self, budget_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can delete budgets.')}
        budget = self.env['hr.formula.budget'].browse(int(budget_id))
        if budget.exists():
            budget.unlink()
        return {'ok': True}

    @api.model
    def budget_prepare(self, config_id, budget_id, run_b_id):
        """Create a budget-vs-actual comparison and return the B-slip work-list to
        drive in chunks — parallels ``compare_prepare`` (only side B chunks)."""
        Cmp = self.env['hr.formula.period.comparison']
        created = Cmp.cmp_create_budget(config_id, budget_id, run_b_id)
        if not created.get('ok'):
            return created
        prep = Cmp.cmp_prepare(created['cmp_id'])
        prep.update({'ok': True, 'headline': created.get('headline')})
        return prep

    # ==================================================================
    # W48 — Payrun anomaly narration (deterministic-first, LLM-polished)
    # ==================================================================
    @api.model
    def narrate_comparison(self, cmp_id, lang='en'):
        """TD.2 — narrate a finished period comparison. The deterministic blocks
        (D-D2) are always the floor; an LLM rewrite for fluency is served ONLY if
        every money-scale number in it exists in the fold (invented figures →
        deterministic text). No AI key / any error → deterministic (C1)."""
        cmp = self.env['hr.formula.period.comparison'].browse(int(cmp_id))
        if not cmp.exists():
            return {'ok': False}
        lang = lang if lang in ('en', 'vi') else 'en'
        det = cmp.narrate(lang)
        det_blocks = det['blocks']
        try:
            allowed = cmp.narrate_allowed_numbers()
            sys_lang = 'Vietnamese' if lang == 'vi' else 'English'
            system = (
                "You are PayAI, a payroll analyst. Rewrite the given factual bullet points into a "
                "concise, fluent %s narrative of 3-6 sentences in plain business language. You MUST NOT "
                "invent, alter, round differently, or drop any number, employee name, component code, or "
                "date — reuse them exactly. Reply STRICT JSON: {\"narrative\": \"...\"}." % sys_lang)
            user = json.dumps({'facts': det['facts'], 'sentences': det_blocks}, ensure_ascii=False)
            out = self._llm_chat(
                [{'role': 'system', 'content': system}, {'role': 'user', 'content': user}],
                json_mode=True)
            polished = (out or {}).get('narrative') if isinstance(out, dict) else None
            if polished and polished.strip() and _narr_numbers_ok(polished, allowed):
                return {'ok': True, 'blocks': [polished.strip()], 'source': 'ai',
                        'lang': lang, 'deterministic': det_blocks}
        except Exception as e:
            _logger.info("narrate_comparison LLM fallback: %s", e)
        return {'ok': True, 'blocks': det_blocks, 'source': 'deterministic', 'lang': lang}

    # ==================================================================
    # B6 — Bureau cockpit (read-only multi-config health board)
    # ==================================================================
    @api.model
    def bureau_board(self):
        """One health card per configuration the user can see: Phase-1 score,
        F13 open problems, B3 pending changes, employee coverage, lifecycle
        state. Read-only aggregation — nothing is written."""
        Config = self.env['hr.formula.config']
        configs = Config.search([], order='company_id, name')
        Rel = self.env['hr.formula.release']
        cards = []
        for c in configs:
            try:
                prob = self.get_problems(c.id)
            except Exception:
                prob = {'count': 0, 'counts': {'error': 0, 'warning': 0, 'hint': 0}}
            last = self._last_milestone(c)
            when = last.milestone_date if last else False
            try:
                pending = len(self._changes_between(c, when, None))
            except Exception:
                pending = 0
            cards.append({
                'id': c.id, 'name': c.name,
                'company': c.company_id.name if c.company_id else '',
                'division': getattr(c, 'pb_division', False) or '',
                'cycle_type': c.cycle_type or 'regular', 'state': c.state,
                'score': self._score(c),
                'rule_count': len(c.rule_ids),
                'problem_counts': prob.get('counts', {'error': 0, 'warning': 0, 'hint': 0}),
                'problem_count': prob.get('count', 0),
                'pending_changes': pending,
                'release_count': Rel.search_count([('config_id', '=', c.id)]),
                'employees': self._config_employee_count(c),
                # --- identity fields the Config Switcher gallery also renders ---
                'code': c.code or '',
                'country': c.country_code or '',
                'currency': c.currency_id.name or '',
                'active': bool(c.active),
                'sample_count': len(c.sample_data_ids),
                'is_branch': bool(c.parent_branch_id),
                'is_variant': bool(c.master_config_id),
                'is_master': bool(c.variant_ids),
                # --- delete vs archive eligibility (see _delete_eligibility) ---
                **self._delete_eligibility(c),
            })
        # rank so the boards needing attention (errors, pending, low score) float up
        cards.sort(key=lambda k: (
            -(k['problem_counts'].get('error', 0)),
            -k['pending_changes'],
            k['score'],
        ))
        return {'ok': True, 'cards': cards, 'can_edit': self._can_edit(),
                'company': self.env.company.name}

    @api.model
    def bureau_clone(self, config_id, name=None):
        """Template-clone a configuration (rules + rate tables + samples) as a new
        draft — the B6 'roll out a validated scheme' primitive."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to clone configurations.")}
        src = self.env['hr.formula.config'].browse(int(config_id))
        if not src.exists():
            return {'ok': False}
        base_code = (src.code or 'CFG')
        existing = set(self.env['hr.formula.config'].search([]).mapped('code'))
        code, i = base_code + '_COPY', 1
        while code in existing:
            i += 1
            code = '%s_COPY%s' % (base_code, i)
        new = src.copy({
            'name': (name or '').strip() or (_("%s (copy)") % src.name),
            'code': code, 'state': 'draft',
        })
        return {'ok': True, 'config_id': new.id, 'name': new.name}

    # ==================================================================
    # B4 — Legislation packs (roll a statutory change across every config)
    # ==================================================================
    def _legis_constant(self, config, code):
        """The constant rule in `config` matching `code` (case-insensitive)."""
        if not code:
            return self.env['hr.formula.rule']
        cu = code.strip().upper()
        return config.rule_ids.filtered(
            lambda r: r.column_type == 'constant' and (r.code or '').upper() == cu)[:1]

    @staticmethod
    def _legis_eq(a, b):
        # constant_value is stored at 6 decimals — compare at that precision.
        return round(a or 0.0, 6) == round(b or 0.0, 6)

    def _legis_pack_payload(self, pack):
        sel = dict(pack._fields['country_code'].selection)
        return {
            'id': pack.id, 'name': pack.name,
            'country_code': pack.country_code, 'country': sel.get(pack.country_code, ''),
            'version': pack.version, 'authority': pack.authority or '',
            'effective_date': fields.Date.to_string(pack.effective_date) if pack.effective_date else '',
            'state': pack.state, 'description': pack.description or '',
            'item_count': len(pack.item_ids),
        }

    def _legis_eval(self, pack, config):
        """Per-item comparison of a pack against one config: what each statutory
        value is now vs what the pack sets it to. Unmatched codes (constants the
        config doesn't have) are surfaced as matched=False, never mutated."""
        rows = []
        for it in pack.item_ids.sorted(key=lambda i: i.sequence):
            rule = self._legis_constant(config, it.code)
            row = {'code': it.code, 'label': it.label, 'target': it.value,
                   'number_format': it.number_format or 'currency',
                   'note': it.note or '', 'matched': bool(rule)}
            if rule:
                cur = rule.constant_value or 0.0
                row.update({'rule_id': rule.id, 'current': cur,
                            'changed': not self._legis_eq(cur, it.value),
                            'delta': (it.value or 0.0) - cur})
            else:
                row.update({'rule_id': False, 'current': None, 'changed': False, 'delta': 0.0})
            rows.append(row)
        return rows

    def _legis_status(self, rows):
        matched = [r for r in rows if r['matched']]
        if not matched:
            return 'na'
        return 'drift' if any(r['changed'] for r in matched) else 'aligned'

    @api.model
    def legislation_packs(self):
        """Every pack the user can see, with a coverage roll-up over the configs
        in scope (aligned / needs-update / not-applicable). Read-only."""
        packs = self.env['hr.formula.legislation.pack'].search([])
        configs = self.env['hr.formula.config'].search([])
        out = []
        for p in packs:
            aligned = drift = na = 0
            for c in configs:
                st = self._legis_status(self._legis_eval(p, c))
                aligned += st == 'aligned'
                drift += st == 'drift'
                na += st == 'na'
            d = self._legis_pack_payload(p)
            d.update({'aligned': aligned, 'drift': drift, 'na': na})
            out.append(d)
        # newest-effective first, drafts (pending rollouts) surfaced above published
        out.sort(key=lambda k: (0 if k['state'] == 'draft' else 1, k['country'],
                                k['effective_date']), reverse=False)
        return {'ok': True, 'packs': out, 'can_edit': self._can_edit(),
                'company': self.env.company.name, 'config_count': len(configs)}

    @api.model
    def legislation_detail(self, pack_id):
        pack = self.env['hr.formula.legislation.pack'].browse(int(pack_id))
        if not pack.exists():
            return {'ok': False}
        items = [{
            'code': it.code, 'label': it.label, 'value': it.value,
            'number_format': it.number_format or 'currency', 'note': it.note or '',
        } for it in pack.item_ids.sorted(key=lambda i: i.sequence)]
        apps = self.env['hr.formula.legislation.application'].search(
            [('pack_id', '=', pack.id)], limit=50)
        return {'ok': True, 'pack': self._legis_pack_payload(pack), 'items': items,
                'applications': [{
                    'config': a.config_id.name or '', 'by': a.applied_by_id.name or '',
                    'date': fields.Datetime.to_string(a.applied_date) if a.applied_date else '',
                    'item_count': a.item_count,
                } for a in apps],
                'can_edit': self._can_edit()}

    @api.model
    def legislation_coverage(self, pack_id):
        """Per-config board for one pack: which configs are aligned, which need
        the update (and by how many values), which don't carry these codes."""
        pack = self.env['hr.formula.legislation.pack'].browse(int(pack_id))
        if not pack.exists():
            return {'ok': False}
        configs = self.env['hr.formula.config'].search([])
        board = []
        for c in configs:
            rows = self._legis_eval(pack, c)
            matched = [r for r in rows if r['matched']]
            changed = [r for r in matched if r['changed']]
            board.append({
                'config_id': c.id, 'name': c.name, 'state': c.state,
                'status': self._legis_status(rows),
                'matched': len(matched), 'changed': len(changed),
                'employees': self._config_employee_count(c),
                'diffs': changed,
            })
        rank = {'drift': 0, 'aligned': 1, 'na': 2}
        board.sort(key=lambda b: (rank[b['status']], -b['changed'], b['name']))
        summary = {
            'aligned': sum(b['status'] == 'aligned' for b in board),
            'drift': sum(b['status'] == 'drift' for b in board),
            'na': sum(b['status'] == 'na' for b in board),
            'employees_affected': sum(b['employees'] for b in board if b['status'] == 'drift'),
        }
        return {'ok': True, 'pack': self._legis_pack_payload(pack),
                'board': board, 'summary': summary, 'can_edit': self._can_edit()}

    @api.model
    def legislation_diff(self, pack_id, config_id):
        """Full per-item diff of a pack against one config (matched + unmatched)."""
        pack = self.env['hr.formula.legislation.pack'].browse(int(pack_id))
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not pack.exists() or not config.exists():
            return {'ok': False}
        rows = self._legis_eval(pack, config)
        return {'ok': True, 'config': {'id': config.id, 'name': config.name},
                'rows': rows, 'changed': sum(1 for r in rows if r['changed']),
                'can_edit': self._can_edit()}

    @api.model
    def legislation_apply(self, pack_id, config_id=None, config_ids=None):
        """Apply a pack to one config or a set: write each drifted statutory
        constant (F7-versioned, reason='legislation'), seal a B3 milestone, and
        log the application. Configs already aligned are skipped, not touched."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to apply legislation packs.")}
        pack = self.env['hr.formula.legislation.pack'].browse(int(pack_id))
        if not pack.exists():
            return {'ok': False}
        if config_ids:
            targets = self.env['hr.formula.config'].browse([int(i) for i in config_ids])
        elif config_id:
            targets = self.env['hr.formula.config'].browse(int(config_id))
        else:
            return {'ok': False, 'msg': _("No target configuration given.")}
        targets = targets.exists()
        Milestone = self.env['hr.formula.config.milestone'].sudo()
        App = self.env['hr.formula.legislation.application']
        results = []
        for c in targets:
            rows = self._legis_eval(pack, c)
            changed = [r for r in rows if r['matched'] and r['changed']]
            if not changed:
                results.append({'config_id': c.id, 'name': c.name, 'changed': 0, 'skipped': True})
                continue
            seen = set()
            for r in changed:
                rule = self.env['hr.formula.rule'].browse(r['rule_id'])
                rule.with_context(formula_version_reason='legislation',
                                  formula_version_note='%s %s' % (pack.name, pack.version),
                                  formula_version_seen=seen).constant_value = r['target']
            ms = Milestone.record(c, _("Applied %s %s") % (pack.name, pack.version))
            App.create({'pack_id': pack.id, 'config_id': c.id,
                        'item_count': len(changed), 'milestone_id': ms.id})
            results.append({'config_id': c.id, 'name': c.name,
                            'changed': len(changed), 'skipped': False})
        return {'ok': True, 'results': results,
                'total_changed': sum(r['changed'] for r in results),
                'configs_touched': sum(1 for r in results if not r['skipped'])}

    # ==================================================================
    # B2 — Config branches (fork a live config, edit safely, merge back)
    # ==================================================================
    def _branch_value_map(self, config):
        """The mergeable rules of a config keyed by code — formula rules carry a
        formula, constants carry a value. Inputs/others are not merged."""
        out = {}
        for r in config.rule_ids:
            if r.column_type in ('formula', 'constant') and r.code:
                out[r.code.upper()] = r
        return out

    def _branch_row(self, code, brule, prule, fork_when):
        """One diff row between a branch rule (brule) and its parent rule (prule);
        either may be None for add/remove. Carries a token diff for formulas and
        a conflict flag when the parent moved on this rule since the fork."""
        rule = brule or prule
        kind = 'constant' if rule.column_type == 'constant' else 'formula'
        row = {
            'code': code,
            'name': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or code,
            'col': (brule or prule).column_letter or '',
            'kind': kind, 'group': _group_for(rule), 'conflict': False,
        }
        if kind == 'formula':
            old = (prule.excel_formula or '') if prule else ''
            new = (brule.excel_formula or '') if brule else ''
            row.update({'old_formula': old, 'cur_formula': new,
                        'runs': self._token_diff_runs(self._tokenize_text(old),
                                                      self._tokenize_text(new))})
            if brule and prule and fork_when:
                at_fork = self._formula_at(prule, fork_when)
                row['conflict'] = (at_fork or '') != (prule.excel_formula or '')
        else:
            row.update({'old_value': (prule.constant_value if prule else None),
                        'new_value': (brule.constant_value if brule else None),
                        'number_format': (brule or prule).number_format or 'currency'})
        return row

    def _branch_diff_rows(self, branch):
        """changed / added / removed rows for a branch vs its parent (by code)."""
        parent = branch.parent_branch_id
        if not parent:
            return {'changed': [], 'added': [], 'removed': []}
        fork_when = branch.fork_milestone_id.milestone_date if branch.fork_milestone_id else False
        pmap = self._branch_value_map(parent)
        bmap = self._branch_value_map(branch)
        changed, added = [], []
        for code, brule in bmap.items():
            prule = pmap.get(code)
            if not prule:
                added.append(self._branch_row(code, brule, None, fork_when))
                continue
            same = (brule.column_type == 'constant'
                    and round(brule.constant_value or 0.0, 6) == round(prule.constant_value or 0.0, 6)) \
                or (brule.column_type != 'constant'
                    and (brule.excel_formula or '') == (prule.excel_formula or ''))
            if not same:
                changed.append(self._branch_row(code, brule, prule, fork_when))
        removed = [self._branch_row(code, None, prule, fork_when)
                   for code, prule in pmap.items() if code not in bmap]
        return {'changed': changed, 'added': added, 'removed': removed}

    def _branch_payload(self, b):
        d = self._branch_diff_rows(b)
        return {
            'id': b.id, 'name': b.name, 'state': b.state,
            'branch_state': b.branch_state or 'open',
            'note': b.branch_note or '',
            'created': fields.Datetime.to_string(b.create_date) if b.create_date else '',
            'created_by': b.create_uid.name or '',
            'employees': self._config_employee_count(b),
            'changed': len(d['changed']), 'added': len(d['added']),
            'removed': len(d['removed']),
            'conflicts': sum(1 for r in d['changed'] if r['conflict']),
        }

    @api.model
    def list_branches(self, config_id=None):
        """Branches of the current config (and, if it IS a branch, its parent)."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'branches': []}
        branches = config.child_branch_ids.filtered(lambda b: b.branch_state != 'discarded')
        return {
            'ok': True,
            'config': {'id': config.id, 'name': config.name,
                       'is_branch': bool(config.parent_branch_id),
                       'parent_id': config.parent_branch_id.id or False,
                       'parent_name': config.parent_branch_id.name or '',
                       'branch_state': config.branch_state or 'open'},
            'branches': [self._branch_payload(b) for b in branches.sorted(key=lambda x: x.id, reverse=True)],
            'can_edit': self._can_edit(),
        }

    @api.model
    def branch_create(self, config_id, name=None, note=None):
        """Fork a config into a draft branch and anchor a fork milestone on the
        parent (the reference point for later conflict detection)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to create branches.")}
        parent = self.env['hr.formula.config'].browse(int(config_id))
        if not parent.exists():
            return {'ok': False}
        if parent.parent_branch_id:
            return {'ok': False, 'msg': _("You can only branch a mainline configuration, not a branch.")}
        base_code = parent.code or 'CFG'
        existing = set(self.env['hr.formula.config'].with_context(active_test=False)
                       .search([]).mapped('code'))
        code, i = base_code + '_BR', 1
        while code in existing:
            i += 1
            code = '%s_BR%s' % (base_code, i)
        fork = self.env['hr.formula.config.milestone'].sudo().record(
            parent, _("Branched: %s") % ((name or '').strip() or _("branch")))
        branch = parent.copy({
            'name': (name or '').strip() or (_("%s — branch") % parent.name),
            'code': code, 'state': 'draft',
            'parent_branch_id': parent.id, 'branch_state': 'open',
            'branch_note': (note or '').strip() or False,
            'fork_milestone_id': fork.id,
        })
        return {'ok': True, 'branch_id': branch.id, 'name': branch.name}

    @api.model
    def branch_diff(self, branch_id):
        branch = self.env['hr.formula.config'].browse(int(branch_id))
        if not branch.exists() or not branch.parent_branch_id:
            return {'ok': False}
        d = self._branch_diff_rows(branch)
        return {'ok': True,
                'branch': {'id': branch.id, 'name': branch.name,
                           'branch_state': branch.branch_state or 'open'},
                'parent': {'id': branch.parent_branch_id.id, 'name': branch.parent_branch_id.name},
                'changed': d['changed'], 'added': d['added'], 'removed': d['removed'],
                'conflicts': sum(1 for r in d['changed'] if r['conflict']),
                'can_edit': self._can_edit()}

    @api.model
    def branch_merge(self, branch_id, narrative=None):
        """Write the branch's changed formulas/values back onto the parent
        (F7 reason='merge'), then seal a release. Added/removed components are
        reported but not auto-applied (a formula change is the safe 90% case)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to merge branches.")}
        branch = self.env['hr.formula.config'].browse(int(branch_id))
        if not branch.exists() or not branch.parent_branch_id:
            return {'ok': False}
        if branch.branch_state == 'merged':
            return {'ok': False, 'msg': _("This branch has already been merged.")}
        parent = branch.parent_branch_id
        d = self._branch_diff_rows(branch)
        if not d['changed']:
            return {'ok': False, 'reason': 'no_changes',
                    'msg': _("This branch has no formula changes to merge.")}
        pmap = self._branch_value_map(parent)
        column_map = {r.column_letter: r.code for r in parent.rule_ids if r.column_letter}
        seen = set()
        merged = 0
        for row in d['changed']:
            prule = pmap.get(row['code'])
            if not prule:
                continue
            prule = prule.with_context(formula_version_reason='merge',
                                       formula_version_note=_("Merged from %s") % branch.name,
                                       formula_version_seen=seen)
            if row['kind'] == 'constant':
                prule.constant_value = row.get('new_value') or 0.0
            else:
                prule.excel_formula = row.get('cur_formula') or ''
                prule.python_formula = prule._convert_excel_to_python(prule.excel_formula, column_map)
            merged += 1
        branch.branch_state = 'merged'
        rel = self.release_approve(parent.id, (narrative or '').strip()
                                   or _("Merged branch “%s” — %s component(s)") % (branch.name, merged))
        return {'ok': True, 'merged': merged,
                'skipped_added': len(d['added']), 'skipped_removed': len(d['removed']),
                'conflicts': sum(1 for r in d['changed'] if r['conflict']),
                'release_id': rel.get('release_id') if isinstance(rel, dict) else False,
                'parent_id': parent.id, 'parent_name': parent.name}

    @api.model
    def branch_discard(self, branch_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to discard branches.")}
        branch = self.env['hr.formula.config'].browse(int(branch_id))
        if not branch.exists() or not branch.parent_branch_id:
            return {'ok': False}
        branch.write({'branch_state': 'discarded', 'state': 'archived', 'active': False})
        return {'ok': True}

    # ==================================================================
    # B5 — Scheme variants (one master → many synced variants)
    # ==================================================================
    def _variant_overrides(self, variant):
        return set(c.strip().upper() for c in (variant.variant_override_codes or '').split(',') if c.strip())

    def _variant_same(self, vrule, mrule):
        if vrule.column_type == 'constant':
            return round(vrule.constant_value or 0.0, 6) == round(mrule.constant_value or 0.0, 6)
        return (vrule.excel_formula or '') == (mrule.excel_formula or '')

    def _variant_rows(self, variant):
        """Rows where a variant differs from its master OR is locally overridden.
        old = master value, new = variant value (so the diff reads master→variant)."""
        master = variant.master_config_id
        if not master:
            return {'changed': [], 'added': [], 'removed': []}
        ov = self._variant_overrides(variant)
        mmap = self._branch_value_map(master)
        vmap = self._branch_value_map(variant)
        changed = []
        for code, vrule in vmap.items():
            mrule = mmap.get(code)
            if not mrule:
                continue  # variant-only components surface under 'added'
            same = self._variant_same(vrule, mrule)
            overridden = code in ov
            if same and not overridden:
                continue  # in sync, nothing to show
            row = self._branch_row(code, vrule, mrule, False)  # brule=variant, prule=master
            row['overridden'] = overridden
            row['drift'] = (not same) and (not overridden)
            changed.append(row)
        added = [self._branch_row(c, vmap[c], None, False) for c in vmap if c not in mmap]
        removed = [self._branch_row(c, None, mmap[c], False) for c in mmap if c not in vmap]
        return {'changed': changed, 'added': added, 'removed': removed}

    def _variant_payload(self, v):
        rows = self._variant_rows(v)
        drift = sum(1 for r in rows['changed'] if r.get('drift'))
        return {
            'id': v.id, 'name': v.name, 'state': v.state,
            'employees': self._config_employee_count(v),
            'overrides': len(self._variant_overrides(v)), 'drift': drift,
            'added': len(rows['added']), 'removed': len(rows['removed']),
            'in_sync': drift == 0 and not rows['added'] and not rows['removed'],
        }

    @api.model
    def list_variants(self, config_id=None):
        """The variant relationships around the current config — its variants if
        it's a master, or its master + siblings if it's a variant."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        role = 'variant' if config.master_config_id else 'master'
        if role == 'variant':
            master = config.master_config_id
            variants = master.variant_ids
        else:
            master = config
            variants = config.variant_ids
        return {
            'ok': True, 'role': role,
            'current_id': config.id,
            'master': {'id': master.id, 'name': master.name, 'state': master.state},
            'variants': [self._variant_payload(v) for v in variants.sorted(key=lambda x: x.id)],
            'can_edit': self._can_edit(),
        }

    @api.model
    def variant_create(self, master_id, name=None, note=None):
        """Materialize a new variant of a master scheme (a draft copy that will
        be kept in sync). You cannot make a variant of a variant."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to create variants.")}
        master = self.env['hr.formula.config'].browse(int(master_id))
        if not master.exists():
            return {'ok': False}
        if master.master_config_id:
            return {'ok': False, 'msg': _("You can only create a variant of a master scheme, not of another variant.")}
        base_code = master.code or 'CFG'
        existing = set(self.env['hr.formula.config'].with_context(active_test=False)
                       .search([]).mapped('code'))
        code, i = base_code + '_V', 1
        while code in existing:
            i += 1
            code = '%s_V%s' % (base_code, i)
        variant = master.copy({
            'name': (name or '').strip() or (_("%s — variant") % master.name),
            'code': code, 'state': 'draft',
            'master_config_id': master.id, 'variant_override_codes': False,
        })
        return {'ok': True, 'variant_id': variant.id, 'name': variant.name}

    @api.model
    def variant_diff(self, variant_id):
        variant = self.env['hr.formula.config'].browse(int(variant_id))
        if not variant.exists() or not variant.master_config_id:
            return {'ok': False}
        rows = self._variant_rows(variant)
        return {'ok': True,
                'variant': {'id': variant.id, 'name': variant.name},
                'master': {'id': variant.master_config_id.id, 'name': variant.master_config_id.name},
                'changed': rows['changed'], 'added': rows['added'], 'removed': rows['removed'],
                'drift': sum(1 for r in rows['changed'] if r.get('drift')),
                'overrides': sum(1 for r in rows['changed'] if r.get('overridden')),
                'can_edit': self._can_edit()}

    def _variant_sync_one(self, variant):
        """Pull every non-overridden master component into the variant (only when
        it actually differs, to avoid version noise). Overrides are preserved."""
        master = variant.master_config_id
        if not master:
            return {'synced': 0, 'preserved': 0}
        ov = self._variant_overrides(variant)
        mmap = self._branch_value_map(master)
        vmap = self._branch_value_map(variant)
        column_map = {r.column_letter: r.code for r in variant.rule_ids if r.column_letter}
        seen = set()
        synced = 0
        for code, vrule in vmap.items():
            mrule = mmap.get(code)
            if not mrule or code in ov or self._variant_same(vrule, mrule):
                continue
            v = vrule.with_context(formula_version_reason='sync',
                                   formula_version_note=_("Synced from master %s") % master.name,
                                   formula_version_seen=seen)
            if vrule.column_type == 'constant':
                v.constant_value = mrule.constant_value or 0.0
            else:
                v.excel_formula = mrule.excel_formula or ''
                v.python_formula = v._convert_excel_to_python(v.excel_formula, column_map)
            synced += 1
        return {'synced': synced, 'preserved': len(ov & set(vmap.keys()))}

    @api.model
    def variant_sync(self, variant_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to sync variants.")}
        variant = self.env['hr.formula.config'].browse(int(variant_id))
        if not variant.exists() or not variant.master_config_id:
            return {'ok': False}
        r = self._variant_sync_one(variant)
        r.update({'ok': True, 'name': variant.name})
        return r

    @api.model
    def variant_push(self, master_id=None):
        """Push the master to every variant at once — the 'edit once, roll to all'
        primitive. Each variant keeps its own overrides."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to push to variants.")}
        config = self._pick_config(master_id)
        master = config.master_config_id or config
        if not master.variant_ids:
            return {'ok': False, 'msg': _("This scheme has no variants to push to.")}
        results = [dict(self._variant_sync_one(v), variant_id=v.id, name=v.name)
                   for v in master.variant_ids]
        return {'ok': True, 'results': results,
                'total_synced': sum(r['synced'] for r in results),
                'variants': len(results)}

    @api.model
    def variant_toggle_override(self, variant_id, code, on):
        """Protect (on) or release (off) a component from master sync."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to change overrides.")}
        variant = self.env['hr.formula.config'].browse(int(variant_id))
        if not variant.exists() or not variant.master_config_id:
            return {'ok': False}
        ov = self._variant_overrides(variant)
        cu = (code or '').strip().upper()
        if not cu:
            return {'ok': False}
        if on:
            ov.add(cu)
        else:
            ov.discard(cu)
        variant.variant_override_codes = ','.join(sorted(ov)) or False
        return {'ok': True, 'overrides': sorted(ov)}

    @api.model
    def variant_detach(self, variant_id):
        """Sever a variant from its master (it becomes a standalone mainline
        config; its current components are frozen as-is)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to detach variants.")}
        variant = self.env['hr.formula.config'].browse(int(variant_id))
        if not variant.exists() or not variant.master_config_id:
            return {'ok': False}
        variant.write({'master_config_id': False, 'variant_override_codes': False})
        return {'ok': True}

    # ==================================================================
    # B7 — Client review portal (read-only trust surface via a token link)
    # ==================================================================
    def _review_url(self, token):
        base = (self.env['ir.config_parameter'].sudo().get_param('web.base.url') or '').rstrip('/')
        return '%s/formula/review/%s' % (base, token)

    def _review_fmt(self, v, nf, cur):
        """Format a value by its number_format for the read-only page."""
        if v is None:
            return ''
        try:
            v = float(v)
        except (TypeError, ValueError):
            return str(v)
        if nf == 'percentage':
            return ('%g%%' % round(v * 100, 4))
        if nf == 'integer':
            return '{:,.0f}'.format(v)
        if nf == 'number':
            return '{:,.2f}'.format(v)
        return '%s%s' % (cur, '{:,.0f}'.format(v))

    def _review_preview(self, config):
        """A sample payslip preview for the review page — every appears-on-payslip
        component with its computed value, grouped Earnings / Deductions / Totals."""
        sample = config.sample_data_ids[:1]
        cur = config.currency_id.symbol if config.currency_id else '₫'
        if not sample:
            return {'ok': False, 'currency': cur, 'earnings': [], 'deductions': [], 'totals': [], 'sample_name': ''}
        try:
            inputs = json.loads(sample.input_values_json or '{}')
            vals = sample._evaluate_rules_with_dependencies(inputs)
        except Exception:
            inputs, vals = {}, {}
        earnings, deductions, totals = [], [], []
        for r in config.rule_ids.sorted(key=lambda x: x.sequence):
            if not r.appears_on_payslip:
                continue
            code = r.code
            v = vals.get(code)
            if v is None and r.column_type == 'constant':
                v = r.constant_value
            if v is None and r.column_type == 'input':
                v = inputs.get(code)
            try:
                v = float(v or 0)
            except (TypeError, ValueError):
                v = 0.0
            grp = _group_for(r)
            nf = r.number_format or 'currency'
            line = {'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or code,
                    'value': v, 'display': self._review_fmt(v, nf, cur)}
            if grp == 'Deductions':
                deductions.append(line)
            elif grp == 'Totals':
                totals.append(line)
            else:
                earnings.append(line)
        return {'ok': True, 'currency': cur, 'sample_name': sample.name or '',
                'earnings': earnings, 'deductions': deductions, 'totals': totals}

    def _review_components(self, config):
        """Read-only component catalogue grouped for the client."""
        out = []
        cur = config.currency_id.symbol if config.currency_id else '₫'
        for r in config.rule_ids.sorted(key=lambda x: x.sequence):
            if r.column_type not in ('formula', 'constant', 'input'):
                continue
            nf = r.number_format or 'currency'
            out.append({
                'col': r.column_letter or '', 'code': r.code or '',
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                'type': r.column_type, 'group': _group_for(r),
                'formula': (r.excel_formula or '') if r.column_type == 'formula' else '',
                'value_display': self._review_fmt(r.constant_value, nf, cur) if r.column_type == 'constant' else '',
            })
        return out

    def _review_payload(self, share):
        """Everything the read-only review page renders — computed server-side."""
        config = share.config_id
        release = None
        if share.release_id and share.release_id.exists():
            rel = share.release_id
            release = {
                'id': rel.id, 'name': rel.name,
                'narrative': rel.narrative or '',
                'change_count': rel.change_count,
                'approved_by': rel.approved_by_id.name or '',
                'date': fields.Datetime.to_string(rel.approved_date) if rel.approved_date else '',
            }
        comments = [{
            'author_name': c.author_name, 'side': c.author_side,
            'body': c.body or '',
            'date': fields.Datetime.to_string(c.create_date) if c.create_date else '',
        } for c in share.comment_ids]
        country = dict(config._fields['country_code'].selection).get(config.country_code, '')
        return {
            'token': share.token,
            'config': {
                'name': config.name, 'country': country,
                'state': config.state, 'code': config.code or '',
                'component_count': len(config.rule_ids),
                'employees': self._config_employee_count(config),
                'score': self._score(config),
            },
            'client_name': share.client_name or '',
            'components': self._review_components(config),
            'preview': self._review_preview(config),
            'release': release,
            'signed_off': share.signed_off,
            'signed_off_name': share.signed_off_name or '',
            'signed_off_date': fields.Datetime.to_string(share.signed_off_date) if share.signed_off_date else '',
            'comments': comments,
            'company': (config.company_id.name if config.company_id else '') or 'Payobook',
        }

    # ---- share management (cockpit side) ----
    def _share_payload(self, s):
        if s.signed_off:
            status = 'signed'
        elif not s.active:
            status = 'revoked'
        elif s.expiry and s.expiry < fields.Datetime.now():
            status = 'expired'
        elif s.view_count:
            status = 'viewed'
        else:
            status = 'active'
        return {
            'id': s.id, 'token': s.token, 'url': self._review_url(s.token),
            'client_name': s.client_name or '', 'note': s.note or '',
            'release': s.release_id.name or '', 'release_id': s.release_id.id or False,
            'status': status, 'view_count': s.view_count or 0,
            'last_viewed': fields.Datetime.to_string(s.last_viewed) if s.last_viewed else '',
            'signed_off_name': s.signed_off_name or '',
            'signed_off_date': fields.Datetime.to_string(s.signed_off_date) if s.signed_off_date else '',
            'comment_count': len(s.comment_ids),
            'created': fields.Datetime.to_string(s.create_date) if s.create_date else '',
        }

    @api.model
    def create_review_share(self, config_id, release_id=None, client_name='', note=''):
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to share configurations.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        vals = {'config_id': config.id, 'client_name': (client_name or '').strip(),
                'note': (note or '').strip()}
        if release_id:
            vals['release_id'] = int(release_id)
        share = self.env['hr.formula.review.share'].create(vals)
        return {'ok': True, 'share': self._share_payload(share)}

    @api.model
    def list_review_shares(self, config_id=None):
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'shares': []}
        shares = self.env['hr.formula.review.share'].with_context(active_test=False).search(
            [('config_id', '=', config.id)])
        releases = self.env['hr.formula.release'].search([('config_id', '=', config.id)])
        return {'ok': True, 'shares': [self._share_payload(s) for s in shares],
                'releases': [{'id': r.id, 'name': r.name} for r in releases],
                'can_edit': self._can_edit()}

    @api.model
    def revoke_review_share(self, share_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to revoke shares.")}
        share = self.env['hr.formula.review.share'].browse(int(share_id))
        if not share.exists():
            return {'ok': False}
        share.active = False
        return {'ok': True}

    # ---- portal actions (called by the public controller, token-validated) ----
    def _review_share_for(self, token):
        share = self.env['hr.formula.review.share'].sudo().search([('token', '=', token)], limit=1)
        return share if (share and share._is_live()) else self.env['hr.formula.review.share']

    @api.model
    def review_signoff(self, token, name):
        share = self._review_share_for(token)
        if not share or not share.release_id:
            return {'ok': False}
        share._record_signoff(name)
        self.env['hr.formula.review.comment'].sudo().create({
            'share_id': share.id, 'author_name': (name or '').strip() or _('Client'),
            'author_side': 'client',
            'body': _("✔ Signed off release “%s”.") % (share.release_id.name or ''),
        })
        return {'ok': True}

    @api.model
    def review_comment(self, token, name, body, side='client'):
        share = self._review_share_for(token)
        body = (body or '').strip()
        if not share or not body:
            return {'ok': False}
        self.env['hr.formula.review.comment'].sudo().create({
            'share_id': share.id, 'author_name': (name or '').strip() or _('Client'),
            'author_side': 'bureau' if side == 'bureau' else 'client', 'body': body,
        })
        return {'ok': True}

    # ==================================================================
    # F8 — Simulate-before-activate (thin facade over hr.formula.simulation)
    # ==================================================================
    @api.model
    def simulate_prepare(self, config_id, overrides=None, limit=None):
        """Create a simulation and return the payslip work-list to drive through
        it in chunks. ``overrides`` = {code: draft_excel_formula} previews a
        specific edit (baseline = current rules); no overrides = whole config vs
        the last actual payrun (D8.1)."""
        Sim = self.env['hr.formula.simulation']
        created = Sim.sim_create(config_id, overrides=overrides or {})
        if not created.get('ok'):
            return created
        prep = Sim.sim_prepare(created['sim_id'], limit=limit)
        prep.update({'ok': True, 'headline': created.get('headline'),
                     'overrides': created.get('overrides', 0)})
        return prep

    @api.model
    def simulate_batch(self, payload):
        """One chunk (~50 payslips). Idempotent-free (accumulates) — the client
        sends each slice of the prepare payslip_ids exactly once."""
        return self.env['hr.formula.simulation'].sim_batch(payload or {})

    @api.model
    def simulate_result(self, sim_id):
        """Finalize (mark done) and return the folded distribution."""
        return self.env['hr.formula.simulation'].sim_finalize(sim_id)

    @api.model
    def simulate_drop(self, sim_id):
        """Discard a simulation — leaves no residue (transient + no rule writes)."""
        sim = self.env['hr.formula.simulation'].browse(int(sim_id))
        sim.sim_drop()
        return {'ok': True}

    # ==================================================================
    # B8 — What-if sliders + cost projection (thin UI over F8's overlay sim)
    # ==================================================================
    @api.model
    def whatif_components(self, config_id=None):
        """The constant components a slider can vary (rates / multipliers / caps)."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        consts = [r for r in config.rule_ids.sorted(key=lambda r: r.sequence)
                  if r.column_type == 'constant' and r.code]
        items = [{
            'code': r.code,
            'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
            'col': r.column_letter or '',
            'value': r.constant_value or 0.0,
            'number_format': r.number_format or 'number',
            'group': _group_for(r),
        } for r in consts]
        return {'ok': True, 'components': items,
                'currency': config.currency_id.symbol if config.currency_id else '₫',
                'can_edit': self._can_edit()}

    @api.model
    def whatif_prepare(self, config_id, target_code, new_value, limit=None):
        """Create a what-if sim (a constant swapped to new_value) and return the
        payslip work-list. Pass a small ``limit`` for the interactive sampled
        feel; omit it for the exhaustive commit run (D-B8)."""
        Sim = self.env['hr.formula.simulation']
        created = Sim.sim_create(config_id, value_overrides={target_code: float(new_value)})
        if not created.get('ok'):
            return created
        prep = Sim.sim_prepare(created['sim_id'], limit=limit)
        config = self._pick_config(config_id)
        hr = config.rule_ids.filtered(lambda r: r.code == created.get('headline'))[:1]
        prep.update({
            'ok': True, 'headline': created.get('headline'),
            'headline_name': (hr.salary_rule_id.name if hr and hr.salary_rule_id else False)
                             or (hr.name if hr else '') or created.get('headline'),
            'sampled': bool(limit),
        })
        return prep

    @api.model
    def whatif_batch(self, payload):
        return self.env['hr.formula.simulation'].sim_batch(payload or {})

    @api.model
    def whatif_result(self, sim_id):
        return self.env['hr.formula.simulation'].sim_finalize(sim_id)

    @api.model
    def whatif_drop(self, sim_id):
        sim = self.env['hr.formula.simulation'].browse(int(sim_id))
        sim.sim_drop()
        return {'ok': True}

    # ==================================================================
    # F14 — Scenario columns (what-if overlays on one component)
    # ==================================================================
    def _scenario_payload(self, sc):
        """One scenario as the grid consumes it (with live validity)."""
        ok, msg = self._check_formula(sc.config_id, sc.override_formula or '',
                                      exclude_id=sc.rule_id.id)
        return {
            'id': sc.id, 'rule_id': sc.rule_id.id,
            'code': sc.rule_id.code or '', 'col': sc.rule_id.column_letter or '',
            'name': sc.name or '', 'override_formula': sc.override_formula or '',
            'color': sc.color_key or 'violet',
            'valid': bool(ok), 'message': msg or '',
        }

    @api.model
    def list_scenarios(self, config_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'scenarios': []}
        scs = self.env['hr.formula.scenario'].search([('config_id', '=', config.id)])
        return {'scenarios': [self._scenario_payload(s) for s in scs]}

    @api.model
    def create_scenario(self, rule_id, name=None):
        """Duplicate a component as a scenario overlay, seeded with its current
        formula. NEVER touches the base rule (D14.1)."""
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _('Component not found')}
        if rule.column_type != 'formula':
            return {'ok': False, 'msg': _('Only formula components can be scenarioed')}
        Scenario = self.env['hr.formula.scenario']
        n = Scenario.search_count([('rule_id', '=', rule.id)])
        label = name or (_('Scenario %s') % chr(ord('A') + n))
        sc = Scenario.create({
            'config_id': rule.config_id.id, 'rule_id': rule.id,
            'name': label, 'override_formula': rule.excel_formula or '',
            'sequence': 10 + n, 'color_key': Scenario.next_color(rule.id),
        })
        return {'ok': True, 'scenario': self._scenario_payload(sc)}

    @api.model
    def save_scenario_formula(self, scenario_id, formula):
        sc = self.env['hr.formula.scenario'].browse(int(scenario_id))
        if not sc.exists():
            return {'ok': False}
        sc.override_formula = formula or ''
        ok, msg = self._check_formula(sc.config_id, formula or '', exclude_id=sc.rule_id.id)
        return {'ok': True, 'valid': bool(ok), 'message': msg or ''}

    @api.model
    def eval_scenario(self, scenario_id, sample_id):
        """Overlay-evaluate the scenario against a sample's inputs (F8 engine).
        Returns the base and scenario value for the component + the take-home
        (net) ripple, all for that one sample. No rule is written (D14.1)."""
        from odoo.addons.pb_hr_payroll_formula.models.formula_simulation import (
            _evaluate_config_overlay)
        sc = self.env['hr.formula.scenario'].browse(int(scenario_id))
        if not sc.exists():
            return {'ok': False}
        config, rule = sc.config_id, sc.rule_id
        try:
            sample = self.env['hr.formula.sample.data'].browse(int(sample_id))
            inputs = json.loads(sample.input_values_json or '{}') if sample.exists() else {}
        except Exception:
            inputs = {}
        base = _evaluate_config_overlay(config, inputs, None)
        cand = _evaluate_config_overlay(config, inputs, {rule.code: sc.override_formula or ''})
        # net/take-home ripple, if the config exposes one
        net_code = None
        for r in config.rule_ids:
            if (r.code or '').upper().replace(' ', '') in (
                    'NET', 'NETPAY', 'NET_PAY', 'NETSALARY', 'TAKEHOME', 'TAKE_HOME'):
                net_code = r.code
                break

        def _num(d, k):
            try:
                return float(d.get(k) or 0.0)
            except (TypeError, ValueError):
                return 0.0
        out = {
            'ok': True, 'col': rule.column_letter or '', 'code': rule.code or '',
            'base_value': _num(base, rule.code),
            'scenario_value': _num(cand, rule.code),
        }
        if net_code:
            out.update({
                'net_code': net_code,
                'net_base': _num(base, net_code),
                'net_scenario': _num(cand, net_code),
            })
        return out

    @api.model
    def promote_scenario(self, scenario_id):
        """Write the scenario's draft into the base rule (versioned, reason=edit)
        then delete the scenario. This is the ONLY path that mutates the rule."""
        sc = self.env['hr.formula.scenario'].browse(int(scenario_id))
        if not sc.exists():
            return {'ok': False}
        rule, config = sc.rule_id, sc.config_id
        formula = sc.override_formula or ''
        ok, msg = self._check_formula(config, formula, exclude_id=rule.id)
        if not ok:
            return {'ok': False, 'msg': msg or _('Scenario formula is invalid')}
        column_map = {r.column_letter: r.code for r in config.rule_ids if r.column_letter}
        rule.with_context(formula_version_reason='edit').write({
            'excel_formula': formula,
            'python_formula': rule._convert_excel_to_python(formula, column_map)
                if rule.column_type == 'formula' else rule.python_formula,
            'is_valid': True, 'validation_message': '',
        })
        code = rule.code or ''
        tests = self._run_tests_after_save(config, [code] if code else None)
        sc.unlink()
        return {'ok': True, 'rule_id': rule.id, 'code': code, 'formula': formula,
                'tests': tests}

    @api.model
    def discard_scenario(self, scenario_id):
        sc = self.env['hr.formula.scenario'].browse(int(scenario_id))
        if sc.exists():
            sc.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # W104 — snippet library (reusable Excel fragments). CRUD only; the
    # ${CODE} → column-letter resolution happens client-side at insertion
    # time (D-F8). Writes are manager-guarded like every other studio write.
    # ------------------------------------------------------------------
    def _snippet_payload(self, s):
        return {
            'id': s.id, 'name': s.name or '', 'category': s.category or 'other',
            'body': s.body or '', 'description': s.description or '',
            'sequence': s.sequence, 'company_id': s.company_id.id or False,
        }

    @api.model
    def list_snippets(self):
        # shared library (no company) + this company's private snippets
        domain = ['|', ('company_id', '=', False), ('company_id', '=', self.env.company.id)]
        snips = self.env['hr.formula.snippet'].search(domain)
        return [self._snippet_payload(s) for s in snips]

    @api.model
    def save_snippet(self, vals):
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can edit snippets.')}
        name = (vals.get('name') or '').strip()
        body = (vals.get('body') or '').strip()
        if not name or not body:
            return {'ok': False, 'msg': _('A snippet needs a name and a body.')}
        data = {
            'name': name, 'body': body,
            'category': vals.get('category') or 'other',
            'description': (vals.get('description') or '').strip(),
        }
        if vals.get('sequence') is not None:
            try:
                data['sequence'] = int(vals['sequence'])
            except (TypeError, ValueError):
                # C7: reject loudly, never silently drop a field the caller sent
                return {'ok': False, 'msg': _('Sequence must be a whole number.')}
        Snip = self.env['hr.formula.snippet']
        sid = vals.get('id')
        if sid:
            rec = Snip.browse(int(sid))
            if not rec.exists():
                return {'ok': False, 'msg': _('Snippet not found.')}
            rec.write(data)
        else:
            rec = Snip.create(data)
        return {'ok': True, 'snippet': self._snippet_payload(rec)}

    @api.model
    def delete_snippet(self, snippet_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can delete snippets.')}
        rec = self.env['hr.formula.snippet'].browse(int(snippet_id))
        if rec.exists():
            rec.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # F13 — Problems rail + lint + rename-refactor
    # ------------------------------------------------------------------
    # Numeric literals worth flagging: big enough to be a "magic amount"
    # (row numbers, small multipliers like 2/12 are noise) and repeated across
    # formulas so extracting a constant actually removes duplication.
    _MAGIC_MIN = 1000.0
    _MAGIC_MIN_COUNT = 2

    @api.model
    def _strip_for_lint(self, formula):
        """Blank out string literals and cell references so only bare numeric
        literals (and operators/functions) remain — the row digits of A2/X2
        must never be mistaken for magic numbers."""
        f = formula or ''
        f = re.sub(r'"[^"]*"', ' ', f)              # string literals
        f = re.sub(r"'[^']*'", ' ', f)
        f = re.sub(r'\$?[A-Za-z]+\$?\d+', ' ', f)   # cell refs A2, $X$2, AA10
        f = re.sub(r'[A-Za-z_][A-Za-z0-9_]*', ' ', f)  # function names / bare codes
        return f

    @api.model
    def _slot_formula(self, formula, by_col, by_code):
        """W52 (D-J2): normalize an ``excel_formula`` to its logical skeleton.

        Strip ``=``, uppercase, drop whitespace, then replace every COMPONENT
        REFERENCE — cell-letter form (``A1``/``X2``), bare column letter
        (``A``/``X``), or code form (``BASIC``/``TXBASE``), resolved
        letter-first-then-code exactly like the engine — with positional slots
        ``§1, §2…`` in order of first occurrence. Numeric literals, operators,
        function names and string literals survive verbatim, so two formulas
        collide iff they are identical modulo which components they reference (a
        differing constant ⇒ different skeleton ⇒ no false group).

        Returns ``(slotted_str, n_refs)`` where ``n_refs`` is the count of
        distinct references (0 ⇒ no component logic to share). Pure text — never
        evaluates anything (TJ.2: zero evaluation in the path)."""
        f = (formula or '').strip()
        if f.startswith('='):
            f = f[1:]
        f = f.upper()
        # Mask string literals (content-sensitive, letter-free placeholder) so a
        # differing string still differs, but its letters are never slotted.
        def _mask(m):
            h = int(hashlib.md5(m.group(0).encode('utf-8')).hexdigest()[:8], 16)
            return '\x01%d\x01' % h
        f = re.sub(r'"[^"]*"', _mask, f)
        f = re.sub(r'\s+', '', f)
        slots = {}
        order = []

        def _rep(m):
            tok = m.group(0)
            # A function call — identifier immediately followed by '(' — is never
            # a component reference (protects a component coded like a function).
            if m.end() < len(f) and f[m.end()] == '(':
                return tok
            cell = re.match(r'^([A-Z]+)\d+$', tok)
            key = None
            if cell:
                letter = cell.group(1)
                if letter in by_col:
                    key = 'C:' + letter
            elif tok in by_col:          # bare column letter — resolve letter first
                key = 'C:' + tok
            elif tok in by_code:         # …then code (engine order)
                key = 'K:' + tok
            if key is None:
                return tok               # function/keyword/unknown → verbatim
            if key not in slots:
                slots[key] = len(order) + 1
                order.append(key)
            return '\xa7%d' % slots[key]

        # [A-Z][A-Z0-9]* (not [A-Z]+\d*) so an interior-digit code like T2X
        # stays ONE token instead of mis-splitting into T2 + X.
        slotted = re.sub(r'[A-Z][A-Z0-9]*', _rep, f)
        return slotted, len(order)

    @api.model
    def get_problems(self, config_id=None):
        """Aggregate everything wrong (or smelly) about a config into one ranked
        list for the Problems rail. Pure metadata + regex — never computes a
        payslip.

        Shape::

            {ok, count, counts: {error, warning, hint},
             problems: [{key, kind, severity, title, detail, rule_id, col, code}]}
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'count': 0,
                    'counts': {'error': 0, 'warning': 0, 'hint': 0}, 'problems': []}

        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        by_col = self._col_to_rule(rules)
        problems = []

        def _add(kind, severity, title, detail, rule=None, col=None, note_id=None):
            problems.append({
                'key': '%s:%s' % (kind, note_id or (rule.id if rule else (col or len(problems)))),
                'kind': kind,
                'severity': severity,
                'title': title,
                'detail': detail,
                'rule_id': rule.id if rule else False,
                'col': (rule.column_letter if rule else col) or '',
                'code': (rule.code if rule else '') or '',
                'note_id': note_id or False,
            })

        # 1) invalid / empty formulas -------------------------------------
        for r in rules:
            if r.column_type != 'formula':
                continue
            if not (r.excel_formula or '').strip():
                _add('empty', 'warning',
                     _("%s has no formula") % (r.name or r.column_letter),
                     _("This calculated component is blank — it evaluates to nothing."),
                     rule=r)
            elif (not r.is_valid) or r.has_evaluation_error:
                msg = r.validation_message or r.last_evaluation_error or _("Formula does not evaluate.")
                _add('invalid', 'error',
                     _("%s (%s) is invalid") % (r.name or '', r.column_letter),
                     msg, rule=r)

        # 2) cycles + unused (reuse the deterministic dependency graph) -----
        intel = self.get_intelligence(config.id)
        for cy in intel.get('cycles', []):
            first = by_col.get((cy.get('cols') or [None])[0])
            _add('cycle', 'error', _("Circular dependency"),
                 cy.get('human_explanation') or '', rule=first,
                 col=(cy.get('cols') or [''])[0])
        for col in intel.get('unused', []):
            rr = by_col.get(col)
            if not rr:
                continue
            if rr.column_type == 'input':
                _add('unused', 'hint',
                     _("%s (%s) is never used") % (rr.name or '', col),
                     _("This input feeds no formula and is not shown on the payslip."),
                     rule=rr)
            else:
                _add('unused', 'warning',
                     _("%s (%s) is never used") % (rr.name or '', col),
                     _("Nothing depends on this component and it does not appear on the payslip."),
                     rule=rr)

        # 3) magic-number lint --------------------------------------------
        # D-J6: literals sitting inside a detected (consistent) progressive
        # IF-chain span are explained by the pending W54 simplify suggestion —
        # one cause, one card — so suppress the magic hint for exactly those
        # tokens. Detection is the cheap parse-only pass (no evaluation), so
        # get_problems stays eval-free.
        chains = self._detect_chains(rules)
        suppress = {}
        for r in rules:
            res = chains.get(r.id)
            if not res or not res.get('consistent'):
                continue
            s, e = res['span']
            span_txt = self._strip_for_lint((r.excel_formula or '')[s:e])
            suppress[r.id] = set(re.findall(r'(?<![A-Za-z0-9._])\d+(?:\.\d+)?', span_txt))
        from collections import Counter
        counter = Counter()
        rules_for_lit = {}
        for r in rules:
            if r.column_type != 'formula' or not r.excel_formula:
                continue
            seen_here = set()
            _drop = suppress.get(r.id, ())
            for tok in re.findall(r'(?<![A-Za-z0-9._])\d+(?:\.\d+)?', self._strip_for_lint(r.excel_formula)):
                if tok in _drop:
                    continue
                try:
                    val = float(tok)
                except ValueError:
                    continue
                if val >= self._MAGIC_MIN and tok not in seen_here:
                    seen_here.add(tok)
                    counter[tok] += 1
                    rules_for_lit.setdefault(tok, []).append(r)
        for tok, cnt in counter.items():
            if cnt >= self._MAGIC_MIN_COUNT:
                where = rules_for_lit[tok]
                cols = ', '.join(rr.column_letter for rr in where if rr.column_letter)
                _add('magic', 'hint',
                     _("Repeated number %s") % '{:,.0f}'.format(float(tok)),
                     _("Appears in %s formulas (%s) — consider extracting a named "
                       "constant so a rate change is a one-line edit.") % (cnt, cols),
                     rule=where[0])

        # 3c) W52 — duplicate-logic groups (D-J2). Token-normalized skeleton
        # hash; groups of >=2 formula components that are identical modulo which
        # components they reference. Detection only (v1) — extracting a shared
        # component changes downstream reference semantics, a human decision.
        # Zero evaluation, no LLM.
        by_code = {r.code: r for r in rules if r.code}
        dupe_groups = defaultdict(list)
        for r in rules:
            if r.column_type != 'formula' or not (r.excel_formula or '').strip():
                continue
            slotted, n_refs = self._slot_formula(r.excel_formula, by_col, by_code)
            if n_refs < 1:
                continue    # a formula referencing no component shares no logic
            h = hashlib.sha1(slotted.encode('utf-8')).hexdigest()
            dupe_groups[h].append(r)
        for members in dupe_groups.values():
            if len(members) < 2:
                continue
            members = sorted(members, key=lambda r: (r.sequence, r.id))
            cols = ', '.join('%s (%s)' % (m.column_letter or '?', m.code or '—')
                             for m in members)
            _add('dupe', 'hint',
                 _("%s components share this logic") % len(members),
                 _("These calculated components are identical apart from which "
                   "components they reference: %s. Consider a single shared "
                   "component — the references would change, so this is a manual "
                   "decision.") % cols,
                 rule=members[0])

        # 4) totals that are not shown on the payslip ----------------------
        for r in rules:
            if (r.column_type == 'formula' and _group_for(r) == 'Totals'
                    and not r.appears_on_payslip):
                _add('offpayslip', 'warning',
                     _("%s (%s) is a total but hidden") % (r.name or '', r.column_letter),
                     _("This looks like a total or net figure yet it is not shown on "
                       "the payslip. Employees will not see it."),
                     rule=r)

        # 4b) COLROLES — column-role health. Five checks that only became askable
        # once every column carried a role. All metadata; nothing is computed.
        roles = {r.id: (r.column_role or 'payroll') for r in rules}
        # Everything any OTHER column's formula reads. TWO sources, because a real
        # Excel formula refers to a column by LETTER ('=B2*C2') while
        # formula_dependencies is a comma-joined mix of letters and codes (CR2) —
        # checking only the codes would miss every ordinary formula.
        # A formula that names itself is a cycle, already reported above — so the
        # reader is always some OTHER column (refs are kept per rule to say so).
        refs_by_rule = {}
        for r in rules:
            if r.column_type != 'formula':
                continue
            cols = set(self._expand_refs(r.excel_formula, by_col))
            for code in (r.formula_dependencies or '').split(','):
                code = code.strip().upper()
                if code:
                    cols.add(code)
            refs_by_rule[r.id] = cols
        # MAPFIX B4 — the rail and the mapping board must never disagree about how
        # much is left to do, so checks (c) and (d) below stopped carrying their own
        # notion of "unrouted" and read the ONE definition instead. Between them they
        # report exactly the set the reconciliation dialog lists: (c) takes the bank
        # columns, (d) takes the rest, and neither can fire for a column the other
        # already named.
        unresolved_rule_ids = set(self._ec_unresolved(config).ids)

        # (a) nothing says who the row belongs to
        if any(r.column_type == 'input' for r in rules) and 'identity' not in roles.values():
            _add('noident', 'error',
                 _("No column identifies the employee"),
                 _("This structure takes values per employee but no column is marked "
                   "as the employee's identity, so an import cannot tell whose row "
                   "is whose. Open a column such as the employee code and set its "
                   "role to Identity."))

        for r in rules:
            role = roles[r.id]
            code = (r.code or '').strip().upper()
            letter = (r.column_letter or '').strip().upper()
            names = {n for n in (code, letter) if n}
            is_read = any(names & refs for rid, refs in refs_by_rule.items() if rid != r.id)

            # (b) a people/reference column feeding a calculation
            if is_read and (role != 'payroll' or r.is_text_component):
                _add('refinformula', 'error',
                     _("%s (%s) is used in a calculation") % (r.name or '', r.column_letter),
                     (_("This column is filed as %s rather than pay, yet another "
                        "column's formula reads it. Either the role is wrong or "
                        "the formula is — a column that is not pay is not "
                        "guaranteed to hold a number.") % self._role_label(role)
                      if role != 'payroll' else
                      _("This column holds text, yet another column's formula reads "
                        "it as a number. Either the text setting is wrong or the "
                        "formula is.")),
                     rule=r)

            # (c) a bank column that lands nowhere. A bank column mapped onto an
            # ordinary field (say `hr.employee.bank_name`) still counts as handled —
            # the value IS stored — so both destinations clear this warning.
            if role == 'bank' and r.id in unresolved_rule_ids:
                _add('bankunmapped', 'warning',
                     _("%s (%s) is bank data with nowhere to go") % (r.name or '', r.column_letter),
                     _("This column holds bank details but is not sent to the "
                       "employee's bank account or to a field, so importing it "
                       "stores nothing."),
                     rule=r)

            # (d) data imported into thin air. Only a FIELD destination can store an
            # employee's name or joining date — the bank lane holds four slots and
            # none of them is where a date of birth goes.
            if role != 'bank' and r.id in unresolved_rule_ids:
                # An identity column is not "dropped" — it is what finds the
                # employee — so it gets its own, truthful sentence.
                _add('idunmapped', 'hint',
                     _("%s (%s) is imported but goes nowhere") % (r.name or '', r.column_letter),
                     (_("This column identifies the employee but is not mapped to a "
                        "field, so it is used to find the row and then discarded — "
                        "map it if the value should also be stored.")
                      if role == 'identity' else
                      _("This column is filed as %s, but it is neither mapped to a "
                        "field nor kept on the contract — the imported value is "
                        "read and then dropped.") % self._role_label(role)),
                     rule=r)

            # (e) people data printed as if it were pay
            if role != 'payroll' and r.appears_on_payslip:
                _add('nonpayslip', 'warning',
                     _("%s (%s) is shown on the payslip") % (r.name or '', r.column_letter),
                     _("This column is filed as %s, not pay, yet it prints as a "
                       "payslip line. Employees will see it among their earnings "
                       "and deductions.") % self._role_label(role),
                     rule=r)

        # 5b) W83 — untested formula components (absence of a test is a smell,
        # not an error → hint tier). Reuses the deterministic coverage graph.
        cov = self.get_test_coverage(config.id)
        for u in cov.get('untested', []):
            rr = by_col.get(u.get('col'))
            if not rr:
                continue
            _add('untested', 'hint',
                 _("%s (%s) has no test") % (rr.name or '', rr.column_letter),
                 _("No sample asserts an expected value for this calculated "
                   "component (and nothing that is asserted depends on it) — "
                   "its result is unverified."),
                 rule=rr)

        # 5) open review notes (F15) — a note flagged for review stays in the
        # rail until someone resolves it (resolving keeps it in history).
        rule_by_id = {r.id: r for r in rules}
        for n in self.env['hr.formula.rule.note'].search(
                [('config_id', '=', config.id), ('is_review', '=', True),
                 ('resolved', '=', False)]):
            rr = rule_by_id.get(n.rule_id.id)
            if not rr:
                continue
            _add('note', 'warning',
                 _("Review note · %s (%s)") % (rr.name or '', rr.column_letter),
                 (n.body or '').strip()[:180], rule=rr, note_id=n.id)

        order = {'error': 0, 'warning': 1, 'hint': 2}
        problems.sort(key=lambda p: (order.get(p['severity'], 9),
                                     self._col_num(p.get('col') or 'ZZ')))
        counts = {'error': 0, 'warning': 0, 'hint': 0}
        for p in problems:
            counts[p['severity']] = counts.get(p['severity'], 0) + 1
        return {'ok': True, 'count': len(problems), 'counts': counts,
                'problems': problems}

    # ==================================================================
    # WP-J — W54 Simplification suggestions (detect → prove → offer → apply)
    # ==================================================================
    @api.model
    def _detect_chains(self, rules):
        """Shared cheap pass (D-J1): run the pure ``if_chain`` detector over each
        formula rule — parse + consistency ONLY, no evaluation. Returns
        ``{rule.id: detect_result}`` for rules whose ``excel_formula`` IS a
        progressive IF-chain (consistent or irregular). Used by BOTH the D-J6
        magic-hint suppression and the W54 suggestion RPC, so chain detection
        lives in one place."""
        from odoo.addons.pb_hr_payroll_formula.formula_engine import if_chain
        out = {}
        for r in rules:
            if r.column_type != 'formula' or not (r.excel_formula or '').strip():
                continue
            res = if_chain.detect(r.excel_formula)
            if res:
                out[r.id] = res
        return out

    @api.model
    def _rate_table_name(self, rule):
        return _("%s — rate table") % (rule.name or rule.code or rule.column_letter)

    @api.model
    def _gen_rate_table_code(self, config, rule):
        """C5-safe (D-J4): letters/digits only, deduped against existing rate
        table AND component codes. Reuses the WP-E `_dedupe_code_c5` deduper —
        one source, no logic fork."""
        existing = {(t.code or '').upper() for t in config.rate_table_ids if t.code}
        existing |= {(r.code or '').upper() for r in config.rule_ids if r.code}
        base = re.sub(r'[^A-Z0-9]', '', (rule.code or 'RATE').upper()) or 'RATE'
        if base[0].isdigit():
            base = 'R' + base
        base = (base + 'RATE')[:40]
        Wiz = self.env['hr.formula.multisheet.import.wizard']
        return Wiz._dedupe_code_c5(base, existing)

    @api.model
    def _find_reusable_table(self, config, brackets, eps=0.5):
        """A rate table in ``config`` whose brackets equal ``brackets`` (lowers
        within ``eps``, rates within 1e-9), else False. C7 honesty — the offer
        says which existing table it would reuse rather than minting a twin."""
        want = sorted((float(lo), float(ra)) for lo, ra in brackets)
        for t in config.rate_table_ids:
            got = sorted((b.lower, b.rate) for b in t.line_ids)
            if len(got) != len(want):
                continue
            if all(abs(g[0] - w[0]) <= eps and abs(g[1] - w[1]) <= 1e-9
                   for g, w in zip(got, want)):
                return t
        return False

    @api.model
    def _resolve_driver_rule(self, config, driver_text):
        """The single component a chain's driver references (cell-letter form
        ``AB2``, bare column letter, or code — engine order), or False for a
        COMPOUND driver expression (``MIN(A,B)``, operators). Only a single
        component can carry injected edge probes (D-J3)."""
        t = re.sub(r'\s+', '', (driver_text or '')).upper()
        m = re.fullmatch(r'([A-Z]+)(\d+)?', t)
        if not m:
            return False    # compound expression → probes not injectable
        by_col = {r.column_letter: r for r in config.rule_ids if r.column_letter}
        by_code = {r.code: r for r in config.rule_ids if r.code}
        letters = m.group(1)
        if m.group(2) is not None:          # cell form LETTERS+digits
            return by_col.get(letters) or False
        return by_col.get(letters) or by_code.get(letters) or False

    @api.model
    def _probe_edges(self, brackets):
        """Every bracket lower bound plus one synthetic edge just inside the top
        band — so each boundary is probed at −1/0/+1 (D-J3). For the VN PIT that
        is 8 edges × 3 = 24 probes."""
        lowers = sorted(float(lo) for lo, _ in brackets)
        if not lowers:
            return []
        step = (lowers[-1] - lowers[-2]) if len(lowers) > 1 else 1.0
        return lowers + [lowers[-1] + max(1.0, step)]

    def _eq_delta(self, rule, values, original, draft):
        """|original(values) − draft(values)| through the REAL evaluator
        (_run_formula overlay, no persistence — C12). None if either side fails
        to evaluate."""
        try:
            a = rule._run_formula(values, original, write_diagnostics=False)
            b = rule._run_formula(values, draft, write_diagnostics=False)
            return abs(float(a) - float(b))
        except Exception:
            return None

    @api.model
    def _equivalence_check(self, rule, brackets, driver_text, span_start, span_end):
        """D-J3 gate: prove the BRACKET rewrite evaluates identically to the
        original on every sample row (confirmed or not) PLUS synthetic edge
        probes when the driver is a single component. The draft INLINES the exact
        Excel the committed table will emit (``compile_brackets_excel``), so the
        proof matches the apply. Read-only."""
        from odoo.addons.pb_hr_payroll_formula.models.formula_rate_table import (
            compile_brackets_excel,
        )
        EPS = 0.005
        original = rule.excel_formula or ''
        compiled = compile_brackets_excel(brackets, driver_text)
        draft = original[:span_start] + '(' + compiled + ')' + original[span_end:]

        max_delta = 0.0
        samples_total = samples_matched = 0
        for smp in rule.config_id.sample_data_ids:
            try:
                # readonly=True — this runs from the Problems rail on every
                # panel open; sample-value acquisition must never stamp
                # write_date on production rules (M2, WP-J review).
                vals = smp._evaluate_rules_with_dependencies(
                    smp.get_input_values(), readonly=True)
            except Exception:
                continue
            d = self._eq_delta(rule, vals, original, draft)
            if d is None:
                continue
            samples_total += 1
            max_delta = max(max_delta, d)
            if d < EPS:
                samples_matched += 1

        drule = self._resolve_driver_rule(rule.config_id, driver_text)
        driver_kind = 'compound'
        probes_total = probes_matched = 0
        if drule:
            driver_kind = 'input' if drule.column_type == 'input' else 'computed'
            for edge in self._probe_edges(brackets):
                for x in (edge - 1.0, edge, edge + 1.0):
                    d = self._eq_delta(rule, {drule.code: x}, original, draft)
                    if d is None:
                        continue
                    probes_total += 1
                    max_delta = max(max_delta, d)
                    if d < EPS:
                        probes_matched += 1

        evidence = samples_total + probes_total
        ok = bool(evidence > 0
                  and samples_matched == samples_total
                  and probes_matched == probes_total)
        return {
            'ok': ok, 'driver_kind': driver_kind, 'max_delta': max_delta,
            'samples_total': samples_total, 'samples_matched': samples_matched,
            'probes_total': probes_total, 'probes_matched': probes_matched,
        }

    @api.model
    def get_simplify_suggestions(self, config_id=None):
        """W54 (D-J3): detect progressive IF-chains, PROVE equivalence to a
        ``BRACKET`` rewrite through the real evaluator, and return offers.
        Read-only — nothing is persisted. Consistent+equivalent chains carry
        ``can_apply``; irregular or unproven chains are LISTED with a reason,
        never offered a rewrite (C7).

        Shape: ``{ok, suggestions: [{rule_id, col, code, name, consistent,
        can_apply, driver, driver_kind, span, brackets, table:{code,name,reuse,
        reuse_of}, equivalence:{…}, before, after, reason}]}``"""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'suggestions': []}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        chains = self._detect_chains(rules)
        rule_by_id = {r.id: r for r in rules}
        suggestions = []
        for rid, res in chains.items():
            rule = rule_by_id[rid]
            brackets = [(b['lower'], b['rate']) for b in res['brackets']]
            driver_text = res['driver']
            s, e = res['span']
            item = {
                'rule_id': rid, 'col': rule.column_letter or '', 'code': rule.code or '',
                'name': rule.name or rule.column_letter or '',
                'consistent': bool(res.get('consistent')),
                'driver': driver_text, 'span': [s, e],
                'brackets': res['brackets'], 'before': rule.excel_formula or '',
            }
            if not res.get('consistent'):
                item.update({'can_apply': False, 'reason': res.get('reason'),
                             'driver_kind': None, 'equivalence': None,
                             'table': None, 'after': None})
                suggestions.append(item)
                continue
            reuse = self._find_reusable_table(config, brackets)
            tcode = reuse.code if reuse else self._gen_rate_table_code(config, rule)
            after = (rule.excel_formula[:s]
                     + 'BRACKET(%s,%s)' % (tcode, driver_text)
                     + rule.excel_formula[e:])
            eq = self._equivalence_check(rule, brackets, driver_text, s, e)
            item.update({
                'can_apply': eq['ok'],
                'driver_kind': eq['driver_kind'],
                'equivalence': {k: eq[k] for k in (
                    'samples_total', 'samples_matched', 'probes_total',
                    'probes_matched', 'max_delta')},
                'table': {'code': tcode,
                          'name': reuse.name if reuse else self._rate_table_name(rule),
                          'reuse': bool(reuse),
                          'reuse_of': reuse.code if reuse else None},
                'after': after,
                'reason': None if eq['ok'] else (
                    _("No evidence to prove equivalence — the config has no "
                      "usable samples and the driver is not probeable — not "
                      "offered.")
                    if (eq['samples_total'] + eq['probes_total']) == 0 else
                    _("Could not prove the rewrite is equivalent (max Δ %.4f) "
                      "— not offered.") % eq['max_delta']),
            })
            suggestions.append(item)
        return {'ok': True, 'suggestions': suggestions}

    @api.model
    def simplify_apply(self, rule_id):
        """W54 apply (D-J3/D-J4): create-or-reuse the rate table, rewrite ONLY
        the detected span to ``BRACKET(code, driver)`` (wrapper survives
        verbatim), stamp version reason ``refactor`` (C4), and re-run W82 tests
        (a refactor is a save). Atomic — table + rewrite in one savepoint.
        Manager-gated; re-proves equivalence defensively before touching data."""
        from odoo.addons.pb_hr_payroll_formula.formula_engine import if_chain
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _("Component not found.")}
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        config = rule.config_id
        res = if_chain.detect(rule.excel_formula or '')
        if not res or not res.get('consistent'):
            return {'ok': False, 'msg': _("This formula is no longer a consistent progressive chain.")}
        brackets = [(b['lower'], b['rate']) for b in res['brackets']]
        driver_text = res['driver']
        s, e = res['span']
        eq = self._equivalence_check(rule, brackets, driver_text, s, e)
        if not eq['ok']:
            return {'ok': False,
                    'msg': _("Could not prove the rewrite is equivalent — not applied.")}
        original = rule.excel_formula
        with self.env.cr.savepoint():
            reuse = self._find_reusable_table(config, brackets)
            if reuse:
                table, reused = reuse, True
            else:
                table = self.env['hr.formula.rate.table'].create({
                    'name': self._rate_table_name(rule),
                    'code': self._gen_rate_table_code(config, rule),
                    'config_id': config.id,
                    'line_ids': [(0, 0, {'lower': lo, 'rate': ra}) for lo, ra in brackets],
                })
                reused = False
            new_formula = (original[:s] + 'BRACKET(%s,%s)' % (table.code, driver_text)
                           + original[e:])
            rule.with_context(formula_version_reason='refactor').write(
                {'excel_formula': new_formula})
        tests = config.run_sample_tests(changed_codes={rule.code})
        return {'ok': True, 'reused': reused, 'table_code': table.code,
                'brackets': len(brackets), 'new_formula': new_formula,
                'tests': tests,
                'msg': (_("Reused rate table %s.") % table.code if reused
                        else _("Created rate table %s (%s brackets).")
                        % (table.code, len(brackets)))}

    @api.model
    def rename_component(self, rule_id, new_code):
        """Rename a component's CODE — permission check here, mechanics in the model.

        The engine is `hr.formula.rule._rename_code` (MAPFIX A), which lives in
        pb_hr_payroll_formula so the upgrade migration can use the same code path
        and the two cannot drift. It moves everything that answered to the old
        name — the matching contract component template above all, which is global
        and matched by STRING — and leaves payslip history alone.

        Asymmetry (surfaced in the UI): formulas reference other components by
        their COLUMN LETTER, not their code, so renaming a code is normally
        metadata-only and every formula keeps evaluating identically. Renaming
        column *letters* is deliberately not offered — letters are positional
        identity.
        """
        Rule = self.env['hr.formula.rule']
        rule = Rule.browse(int(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _("Component not found.")}
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}

        return rule._rename_code(new_code)

    @api.model
    def rename_components(self, config_id, pairs):
        """Rename a whole set of component codes at once — all or nothing.

        ``pairs`` is ``[{'rule_id': int, 'new_code': str}, …]``. The ENTIRE set is
        validated before a single character is written: shape, uniqueness against
        the codes that will remain, collision with a column letter, and two pairs
        aiming at one code. A batch that is wrong anywhere writes nothing, so a
        half-renamed structure — the state that actually breaks formulas — cannot
        exist.
        """
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False, 'msg': _("Configuration not found.")}
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}

        pairs = [p for p in (pairs or []) if p and p.get('rule_id')]
        if not pairs:
            return {'ok': False, 'msg': _("Nothing to rename.")}

        rules_by_id = {r.id: r for r in config.rule_ids}
        letters = {r.column_letter for r in config.rule_ids if r.column_letter}
        planned, errors, seen_targets = [], [], {}

        for pair in pairs:
            rid = int(pair.get('rule_id'))
            new = (pair.get('new_code') or '').strip().upper()
            rule = rules_by_id.get(rid)
            if not rule:
                errors.append(_("Component %s is not part of this structure.") % rid)
                continue
            label = rule.column_letter or rule.name or str(rid)
            if not component_code_mod.is_valid_code(new):
                errors.append(_("%(where)s: '%(code)s' — use letters and digits only, "
                                "starting with a letter.", where=label, code=new or '(blank)'))
                continue
            if new in letters:
                errors.append(_("%(where)s: '%(code)s' is a column letter in this structure.",
                                where=label, code=new))
                continue
            if new in seen_targets:
                errors.append(_("%(a)s and %(b)s both want the code '%(code)s'.",
                                a=seen_targets[new], b=label, code=new))
                continue
            seen_targets[new] = label
            if new != (rule.code or '').upper():
                planned.append((rule, new))

        renaming_ids = {r.id for r, _n in planned}
        keeping = {(r.code or '').upper() for r in config.rule_ids if r.id not in renaming_ids}
        for rule, new in planned:
            if new in keeping:
                errors.append(_("%(where)s: '%(code)s' is already used by a component that is "
                                "not being renamed.", where=rule.column_letter or rule.name,
                                code=new))

        if errors:
            return {'ok': False, 'msg': _("Nothing was renamed — %s") % ' '.join(errors),
                    'errors': errors}

        # Order the set so a component never has to take a code another component
        # in the same batch is still holding (A→B, B→C runs B first). A true CYCLE
        # cannot be ordered and is refused rather than parked on temporary codes,
        # which would make every orphan-safe lookup in `_rename_code` read the
        # parking name instead of the real one.
        held = {(r.code or '').upper(): r.id for r, _n in planned}
        remaining, ordered = list(planned), []
        while remaining:
            free = [p for p in remaining
                    if held.get(p[1]) is None or held.get(p[1]) == p[0].id]
            if not free:
                cycle = ', '.join(sorted(p[0].column_letter or p[0].name for p in remaining))
                return {'ok': False, 'msg': _(
                    "These components are trading codes with each other (%s). Rename them "
                    "one at a time, or to names nobody is using yet.") % cycle}
            for pair in free:
                ordered.append(pair)
                held.pop((pair[0].code or '').upper(), None)
                remaining.remove(pair)

        results, failures = [], []
        for rule, new in ordered:
            outcome = rule._rename_code(new)
            if outcome.get('ok'):
                results.append({'rule_id': rule.id, 'old_code': outcome.get('old_code'),
                                'new_code': new})
            else:
                failures.append({'rule_id': rule.id, 'msg': outcome.get('msg')})

        if failures:
            # Refusals inside `_rename_code` are orphan guards (a shared contract
            # component, a template already under that name). All-or-nothing means
            # the whole batch comes back rather than leaving half of it applied.
            raise UserError(_("Nothing was renamed. %s") % ' '.join(
                f.get('msg') or '' for f in failures))

        return {'ok': True, 'renamed': len(results), 'results': results,
                'msg': _("Renamed %s components.") % len(results)}

    # ------------------------------------------------------------------
    # F10 — Unified Mapping Canvas (adapter 1: mid→end cycle mapping)
    # ------------------------------------------------------------------
    def _cycle_pair(self, config):
        """Given any config, resolve its (mid_cycle, end_cycle) sibling pair.
        Pairs by pb_division (the pb_* demo world) else structure_id, within the
        same company. Returns (mid, end) records or (empty, empty)."""
        Config = self.env['hr.formula.config']
        empty = Config.browse()
        ct = config.cycle_type
        if ct not in ('mid_cycle', 'end_cycle'):
            return empty, empty
        want = 'end_cycle' if ct == 'mid_cycle' else 'mid_cycle'
        dom = [('cycle_type', '=', want), ('id', '!=', config.id)]
        if config.company_id:
            dom.append(('company_id', '=', config.company_id.id))
        cands = Config.search(dom)
        sibling = empty
        div = getattr(config, 'pb_division', False)
        if div:
            sibling = cands.filtered(lambda c: getattr(c, 'pb_division', False) == div)[:1]
        if not sibling and config.structure_id:
            sibling = cands.filtered(lambda c: c.structure_id.id == config.structure_id.id)[:1]
        if not sibling and cands:
            # name-prefix heuristic: everything before the em/en dash
            def prefix(n):
                return re.split(r'[—–-]', n or '', 1)[0].strip().lower()
            p = prefix(config.name)
            sibling = cands.filtered(lambda c: prefix(c.name) == p)[:1]
        if not sibling:
            return empty, empty
        return (config, sibling) if ct == 'mid_cycle' else (sibling, config)

    def _mc_item(self, rule, group_by_role=False):
        """One MappingCanvas item — payroll-agnostic {id, label, sublabel, meta}.

        COLROLES P3: `group_by_role` swaps the item's swim-lane from the payslip
        section (Earnings / Deductions / …) to the column's ROLE. It is a parameter
        rather than a change of behaviour because the CYCLE board is grouped by
        section on purpose and must stay that way — and because `_group_for` matches
        substrings (CR10), which is a trap the role lanes have no reason to inherit.
        """
        item = {
            'id': rule.id,
            'label': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or rule.code or '(unnamed)',
            'sublabel': rule.code or '',
            'meta': {'col': rule.column_letter or '', 'type': rule.column_type or '',
                     'group': _group_for(rule)},
        }
        if group_by_role:
            role = rule.column_role or 'payroll'
            item['group'] = self._role_lane_label(role)
            item['meta']['role'] = role
        return item

    @api.model
    def mapping_canvas_data(self, config_id=None):
        """Feed the cycle-mapping surface: left = mid-cycle components, right =
        end-cycle components, wires = accepted mappings + proposed suggestions."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        mid, end = self._cycle_pair(config)
        if not (mid and end):
            return {'ok': False, 'reason': 'no_pair',
                    'config': {'id': config.id, 'name': config.name,
                               'cycle_type': config.cycle_type}}
        left = [self._mc_item(r) for r in mid.rule_ids.sorted(key=lambda r: r.sequence)]
        right = [self._mc_item(r) for r in end.rule_ids.sorted(key=lambda r: r.sequence)]
        Mapping = self.env['hr.payroll.cycle.component.mapping']
        Sug = self.env['hr.payroll.cycle.mapping.suggestion']
        base = [('mid_cycle_config_id', '=', mid.id), ('end_cycle_config_id', '=', end.id)]
        wires = []
        for m in Mapping.search(base):
            wires.append({'id': 'm%s' % m.id, 'kind': 'mapping', 'ref': m.id,
                          'leftId': m.mid_component_id.id, 'rightId': m.end_component_id.id,
                          'state': 'accepted'})
        for s in Sug.search(base + [('state', '=', 'proposed')]):
            wires.append({'id': 's%s' % s.id, 'kind': 'suggestion', 'ref': s.id,
                          'leftId': s.mid_component_id.id, 'rightId': s.end_component_id.id,
                          'state': 'suggested',
                          'confidence': round(s.confidence or 0.0, 4),
                          'reason': s.match_reason or ''})
        return {
            'ok': True,
            'mid': {'id': mid.id, 'name': mid.name},
            'end': {'id': end.id, 'name': end.name},
            'left': left, 'right': right, 'wires': wires,
            'left_title': mid.name, 'right_title': end.name,
            'subtitle': _("Carry values from %s into %s") % (mid.name, end.name),
            'supports_suggest': True,
            'can_edit': self._can_edit(),
        }

    @api.model
    def mapping_suggest(self, config_id=None):
        """(Re)generate proposed suggestions for the config's cycle pair."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        mid, end = self._cycle_pair(config)
        if not (mid and end):
            return {'ok': False, 'reason': 'no_pair'}
        wiz = self.env['hr.payroll.cycle.component.mapping.wizard'].create({
            'mid_cycle_config_id': mid.id, 'end_cycle_config_id': end.id})
        wiz.action_suggest_mappings()
        return self.mapping_canvas_data(config.id)

    @api.model
    def mapping_accept(self, suggestion_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        s = self.env['hr.payroll.cycle.mapping.suggestion'].browse(int(suggestion_id))
        if not s.exists():
            return {'ok': False}
        s.action_accept()
        return {'ok': True}

    @api.model
    def mapping_reject(self, suggestion_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        s = self.env['hr.payroll.cycle.mapping.suggestion'].browse(int(suggestion_id))
        if s.exists():
            s.action_reject()
        return {'ok': True}

    @api.model
    def mapping_create(self, config_id, mid_component_id, end_component_id):
        """Draw a wire = create a mapping between a mid and an end component."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        mid, end = self._cycle_pair(config) if config else (None, None)
        if not (mid and end):
            return {'ok': False, 'reason': 'no_pair'}
        Mapping = self.env['hr.payroll.cycle.component.mapping']
        midc = self.env['hr.formula.rule'].browse(int(mid_component_id))
        endc = self.env['hr.formula.rule'].browse(int(end_component_id))
        if midc.config_id != mid or endc.config_id != end:
            return {'ok': False, 'msg': _("Components must belong to the paired configs.")}
        # respect the one-mid-one-end uniqueness: drop any existing wire on either side
        Mapping.search([('mid_cycle_config_id', '=', mid.id), ('end_cycle_config_id', '=', end.id),
                        '|', ('mid_component_id', '=', midc.id),
                        ('end_component_id', '=', endc.id)]).unlink()
        Mapping.create({'mid_cycle_config_id': mid.id, 'end_cycle_config_id': end.id,
                        'mid_component_id': midc.id, 'end_component_id': endc.id})
        return {'ok': True}

    @api.model
    def mapping_delete(self, mapping_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        m = self.env['hr.payroll.cycle.component.mapping'].browse(int(mapping_id))
        if m.exists():
            m.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # F10 adapter 2 — API/integration field mapping (source fields → inputs)
    # ------------------------------------------------------------------
    @staticmethod
    def _norm(s):
        return re.sub(r'[^a-z0-9]', '', (s or '').lower())

    @api.model
    def _dt_label(self, data_type):
        """The human label of a data type, from the store's OWN selection — the
        one list `hr.integration.endpoint` imports rather than retyping."""
        if not data_type:
            return ''
        sel = dict(self.env['hr.api.data.store']._fields['data_type'].selection)
        return sel.get(data_type, data_type)

    @staticmethod
    def _as_id(value):
        """An id from an arrival CONTEXT, or 0.

        `pb_connector`/`pb_endpoint`/`pb_config` are written into a context by
        whoever built the link and read back out of the browser, so they are
        not guaranteed to be numbers — a hand-built or stale deep link can
        carry a name, a list, or `None`. `int()` on that is a 500 on a screen
        whose whole job is to be the friendly front door, so it is asked
        politely and answered with "nothing was specified" (which the caller
        then reports through `fell_back` rather than swallowing).
        """
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _sample_text(value):
        """A sample value as one short, printable line.

        "What the data actually looks like" is the whole point of the second
        line under a source field, so a dict or a 400-character blob has to
        become something a 340px column can show. Trimmed rather than dropped:
        an elided value still tells the reader the shape of what arrives.
        """
        if value is None or value is False or value == '':
            return ''
        if isinstance(value, (dict, list)):
            try:
                value = json.dumps(value, ensure_ascii=False, default=str)
            except Exception:
                value = str(value)
        text = str(value).strip().replace('\n', ' ')
        return text if len(text) <= 48 else text[:45] + '…'

    @api.model
    def _config_for_connector(self, connector_id):
        """The scheme this connector feeds most, or 0.

        `_api_active_connector` answers the mirror question (given a config,
        which connector) and this is the same arithmetic run the other way, for
        the deep link that names a connector and no scheme. Ties go to the
        lowest config id so the answer is STABLE — a resolver that returns a
        different scheme on alternate clicks is worse than one that returns a
        plain default.
        """
        FM = self.env['hr.integration.field.mapping']
        counts = defaultdict(int)
        for m in FM.search([('connector_id', '=', self._as_id(connector_id))]):
            cfg = m.target_rule_id.config_id
            if cfg:
                counts[cfg.id] += 1
        if not counts:
            return 0
        return max(sorted(counts), key=lambda cid: counts[cid])

    @api.model
    def mapping_pickers(self, arrival=None):
        """Everything the Mapping Studio's FROM/TO pickers offer, in one call.

        Read with the CALLER's own rights — `search([])` applies the record
        rules, which is the same scope `pb.integrations._readable_connectors`
        settles on, and for the same reason: the honest answer to "which
        connectors are there" is "the ones you may read". A caller with no
        read ACL at all gets the ORM's own AccessError; this method does not
        catch it into a plausible-looking empty list.

        It also RESOLVES the arrival context, and reports what it could not
        honour. A deep link that silently lands on a different scheme is the
        worst bug class this codebase has (W76.3, W117): the reader believes
        the header. So an unresolvable `config_id` falls back AND says so in
        `defaults.fell_back`, which the studio renders as a visible notice.
        """
        arrival = dict(arrival or {})
        Conn = self.env['hr.integration.connector']
        Config = self.env['hr.formula.config']

        cons = Conn.search([], order='name')
        # ONE search and ONE batched compute for every feed on the board. Read
        # per connector, `_compute_counts` would run its three `_read_group`s
        # once per row — twenty-five connectors is seventy-five queries for a
        # dropdown (W53's shape: the payload a picker needs is one query, not
        # one query per option).
        by_conn_eps = defaultdict(list)
        all_eps = self._api_endpoints(cons) if cons else None
        for e in (all_eps or []):
            by_conn_eps[e.connector_id.id].append({
                'id': e.id, 'name': e.name or e.code or '',
                'code': e.code or '', 'data_type': e.data_type or '',
                'data_type_label': self._dt_label(e.data_type),
                'mapping_count': e.mapping_count,
                'staged': e.staged_count, 'synced': e.synced_count,
                'last_sync': fields.Datetime.to_string(e.last_sync) if e.last_sync else '',
                'status': e.last_sync_status or '',
            })
        connectors = [{
            'id': c.id, 'name': c.name or '—',
            'type': c.connector_type or '',
            'status': c.connection_status or 'disconnected',
            'mapping_count': len(c.field_mapping_ids),
            'last_sync': fields.Datetime.to_string(c.last_sync) if c.last_sync else '',
            'endpoints': by_conn_eps.get(c.id, []),
        } for c in cons]

        configs = []
        for cfg in Config.search([], order='sequence, id desc'):
            rules = cfg.rule_ids
            configs.append({
                'id': cfg.id, 'name': cfg.name or '—', 'code': cfg.code or '',
                'country': cfg.country_code or '', 'state': cfg.state or '',
                'active': bool(cfg.active),
                'column_count': len(rules),
                'input_count': len(rules.filtered(lambda r: r.column_type == 'input')),
                # SC-4 — which source lanes this scheme allows, so the studio
                # can hide the tabs of a lane that is off.
                'lanes': {
                    'api': bool(getattr(cfg, 'source_api_enabled', True)),
                    'excel': bool(getattr(cfg, 'source_excel_enabled', True)),
                    'records': bool(getattr(cfg, 'source_records_enabled', True)),
                },
            })

        Batch = self.env['hr.payroll.import.batch']
        batches = [{'id': b.id, 'name': b.name or '—'}
                   for b in Batch.search([], order='id desc', limit=60)]

        # ---- arrival resolution ------------------------------------------
        fell_back = []
        by_conn = {c['id']: c for c in connectors}
        cid = self._as_id(arrival.get('connector_id'))
        if cid and cid not in by_conn:
            fell_back.append('connector')
            cid = 0
        cfg_id = self._as_id(arrival.get('config_id'))
        if cfg_id and cfg_id not in {c['id'] for c in configs}:
            fell_back.append('config')
            cfg_id = 0
        if not cfg_id and cid:
            # A link that arrives naming a CONNECTOR and no scheme is the board
            # card's "6 mappings" being clicked. Landing on the default scheme
            # then answers "0 mapped" to a user who clicked the number six —
            # the board and the studio contradicting each other on the very
            # click that joins them. So the scheme is resolved to the one this
            # connector actually feeds.
            cfg_id = self._config_for_connector(cid)
        if not cfg_id:
            cfg = self._pick_config(None)
            cfg_id = cfg.id if cfg else 0
        if not cid:
            conn = self._api_active_connector(Config.browse(cfg_id))
            cid = conn.id if conn else 0
        if cid and cid not in by_conn:
            # `_api_active_connector` counts mappings and browses the winner by
            # id, so it can name a connector the record rules hide from this
            # caller. The picker must never open on an option it does not list.
            cid = connectors[0]['id'] if connectors else 0
        eid = self._as_id(arrival.get('endpoint_id'))
        ep_ids = {e['id'] for e in (by_conn.get(cid, {}).get('endpoints') or [])}
        if eid and eid not in ep_ids:
            fell_back.append('endpoint')
            eid = 0

        return {
            'ok': True,
            'connectors': connectors, 'configs': configs, 'batches': batches,
            'defaults': {'connector_id': cid, 'endpoint_id': eid,
                         'config_id': cfg_id, 'fell_back': fell_back},
            'can_edit': self._can_edit(),
        }

    def _api_active_connector(self, config, connector_id=None):
        Conn = self.env['hr.integration.connector']
        if connector_id:
            c = Conn.browse(int(connector_id))
            return c if c.exists() else Conn.browse()
        input_ids = config.rule_ids.filtered(lambda r: r.column_type == 'input').ids
        FM = self.env['hr.integration.field.mapping']
        # the connector with the most mappings already targeting this config's inputs
        maps = FM.search([('target_rule_id', 'in', input_ids)])
        if maps:
            counts = defaultdict(int)
            for m in maps:
                counts[m.connector_id.id] += 1
            best = max(counts, key=counts.get)
            return Conn.browse(best)
        # else a connector referenced by an input rule, else the first connector
        for r in config.rule_ids:
            if getattr(r, 'integration_connector_id', False):
                return r.integration_connector_id
        return Conn.search([], limit=1)

    def _api_endpoints(self, connectors):
        """These connectors' feeds, or `None` on a database that has no feeds
        TABLE yet.

        Cycle 1's degrade rail, reused verbatim: the addons tree is SHARED by
        every database on the box and the schema is created by each database's
        own upgrade, so `'hr.integration.endpoint' in self.env` is True in the
        gap between the two and a query would raise `UndefinedTable` — which,
        caught, still leaves the whole request's transaction aborted. `None`
        (not an empty recordset) so a caller can tell "this database has no
        feeds table" from "this connector has no feeds" (W79).
        """
        if 'hr.integration.endpoint' not in self.env:
            return None
        EP = self.env['hr.integration.endpoint']
        if not EP._schema_ready():
            return None
        return EP.search([('connector_id', 'in', connectors.ids)])

    @api.model
    def api_mapping_data(self, config_id=None, connector_id=None, endpoint_id=None):
        """The API board, optionally narrowed to ONE feed.

        Integrations Cycle 2 gave the connector's feeds an axis of their own.
        Passing `endpoint_id` narrows the LEFT column to the fields that feed
        actually delivers and the wires to the mappings that name it — with one
        deliberate exception: a mapping drawn before feeds existed carries no
        `endpoint_id`, and dropping it here would make an operator's existing
        work disappear the first time they picked a feed. Those legacy wires
        are kept, and their source paths are added to the left column under an
        "Unassigned" group, so the board says where they came from instead of
        silently losing them (W79: absent must not be indistinguishable from
        broken).
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        Conn = self.env['hr.integration.connector']
        conn = self._api_active_connector(config, connector_id)
        contexts = [{'id': c.id, 'name': c.name} for c in Conn.search([], order='name')]
        if not conn:
            return {'ok': False, 'reason': 'no_connector', 'contexts': contexts}
        FM = self.env['hr.integration.field.mapping']
        eps = self._api_endpoints(conn)
        ep = None
        ep_wanted = self._as_id(endpoint_id)
        if ep_wanted and eps is not None:
            cand = eps.filtered(lambda e: e.id == ep_wanted)
            ep = cand[:1] or None
        endpoints = [{'id': e.id, 'name': e.name or e.code or '', 'code': e.code or '',
                      'data_type': e.data_type or '',
                      'data_type_label': self._dt_label(e.data_type),
                      'mapping_count': e.mapping_count, 'staged': e.staged_count,
                      'synced': e.synced_count,
                      'last_sync': fields.Datetime.to_string(e.last_sync) if e.last_sync else '',
                      'status': e.last_sync_status or ''}
                     for e in (eps or [])]
        try:
            fields_ = FM.get_available_source_fields(
                conn.id, ep.data_type if ep else None,
                ep.id if ep else None) or []
        except Exception as e:
            # The `except` stays — a board that renders with an empty FROM
            # column beats a 500 — but it no longer stays SILENT. Cycle 6 shipped
            # an AttributeError inside discovery and this branch turned it into
            # nothing at all: no error, no log line, a column of zero fields,
            # and the honesty banner underneath still counting fifteen mappings
            # as unresolvable. It took a live browser pass to find, because a
            # swallowed exception has no other symptom (W40's shape, W152).
            _logger.warning(
                "Source-field discovery failed for connector %s (feed %s): "
                "%s: %s — the FROM column will render empty.",
                conn.id, (ep.code if ep else 'all'), type(e).__name__, e)
            fields_ = []
        ep_group = (ep.name or ep.code) if ep else ''
        # Integrations Cycle 6 — every card says where it came from. `prov` is
        # promoted to a top-level key rather than buried in `meta` because the
        # canvas renders a chip from it on every row, and `meta` is the bag of
        # things only the transform popover reads.
        # SOURCING S5 — lineage for the keys this connector COMPUTES. Built once
        # for the whole board and attached to the cards it belongs to; a vendor
        # field gets no `lineage` key and therefore grows no affordance.
        lineage_by_key = self._lineage_by_output_key(conn, config)
        left = [{'id': 'f:' + f['path'], 'label': f.get('label') or f['path'],
                 'sublabel': f['path'],
                 'sample': self._sample_text(f.get('sample')),
                 # A computed key is grouped under its own heading rather than
                 # under the vendor feed's — it did not come from the vendor.
                 'group': (_("Derived here") if f.get('provenance') == 'computed'
                           else ep_group),
                 'prov': f.get('provenance') or 'live',
                 'provKind': f.get('catalog_kind') or '',
                 # SC-1 — the two keys the canvas needs to stop treating all
                 # catalogue rows alike: only a `template` row's sample is
                 # fiction (the "e.g." marker), and a `last_seen` date turns
                 # "not sent" into "not in the last sync".
                 'provOrigin': f.get('origin') or '',
                 'lastSeen': (f.get('last_seen') or '')[:10],
                 'drift': bool(f.get('expected_missing')),
                 'note': f.get('notes') or '',
                 'lineage': lineage_by_key.get(f['path']),
                 'meta': {'type': f.get('type') or '', 'sample': f.get('sample')}}
                for f in fields_]
        input_rules = config.rule_ids.filtered(lambda r: r.column_type == 'input') \
            .sorted(key=lambda r: r.sequence)
        # SOURCING S4 — the right column gains a chip (`srcKind`, NOT `prov`:
        # `prov` already means "where this CARD came from", a different axis).
        # SOURCING S5 — and it now shows CALCULATED components too, sealed rather
        # than filtered out, so a reader stops wondering where they went.
        _acts, _run = self._source_actuals(config)
        _emp = self._source_record_dests(config)
        _wires = self._source_wire_dests(config)
        # SOURCING S6 — lineage for the right column too, and NOT limited to the
        # connector this board happens to be showing (S20). `lineage_by_key` above
        # is this connector's, for its own cards; `_lineage_for_config` is every
        # connector that can reach this scheme, for the components.
        right = self._mc_right_column(
            config, _acts, _emp, wire_dests=_wires,
            lineage=self._lineage_for_config(config, conn), board='api')
        # accepted wires = persisted field mappings on this connector → these inputs
        wires = []
        mapped_paths, mapped_rules = set(), set()
        dom = [('connector_id', '=', conn.id),
               ('target_rule_id', 'in', input_rules.ids)]
        if ep:
            # this feed's own mappings, PLUS the ones that predate feeds
            dom = dom + ['|', ('endpoint_id', '=', ep.id), ('endpoint_id', '=', False)]
        present = {i['id'] for i in left}
        for m in FM.search(dom):
            lid = 'f:' + (m.source_field or '')
            wires.append({'id': 'm%s' % m.id, 'kind': 'mapping', 'ref': m.id,
                          'leftId': lid, 'rightId': m.target_rule_id.id,
                          'state': 'accepted',
                          'transform': self._transform_payload(m)})   # W62 (D-I2)
            if ep and lid not in present:
                # a legacy (or foreign-feed) wire whose source is not in this
                # feed's field list — shown, and labelled for what it is
                left.append({'id': lid, 'label': (m.source_field_label
                                                  or m.source_field or lid),
                             'sublabel': m.source_field or '',
                             'sample': self._sample_text(m.source_sample_value),
                             'group': _("Unassigned"),
                             'prov': 'mapping', 'provKind': '', 'drift': False,
                             'note': '', 'meta': {'type': ''}})
                present.add(lid)
            mapped_paths.add(m.source_field or '')
            mapped_rules.add(m.target_rule_id.id)
        # suggested wires = best name match between an unmapped source field and an
        # unmapped input rule (computed live, not persisted)
        rule_norms = [(r, self._norm(r.code), self._norm(r.name)) for r in input_rules
                      if r.id not in mapped_rules]
        for f in fields_:
            path = f['path']
            if path in mapped_paths:
                continue
            fn = self._norm(path)
            fl = self._norm(f.get('label'))
            best, conf = None, 0.0
            for r, rc, rn in rule_norms:
                if not rc:
                    continue
                if fn == rc or fl == rc:
                    c = 1.0
                elif rc and (rc in fn or fn in rc):
                    c = 0.85
                elif rn and (rn == fn or rn in fn or fn in rn):
                    c = 0.8
                else:
                    c = 0.0
                if c > conf:
                    best, conf = r, c
            if best and conf >= 0.8 and best.id not in mapped_rules:
                wires.append({'id': 'sug:%s>%s' % (path, best.id), 'kind': 'suggestion',
                              'ref': None, 'source': path,
                              'leftId': 'f:' + path, 'rightId': best.id,
                              'state': 'suggested', 'confidence': round(conf, 2),
                              'reason': _('Name match')})
                mapped_rules.add(best.id)   # one suggestion per input rule
        return {
            'ok': True, 'left': left, 'right': right, 'wires': wires,
            'left_title': '%s · source fields' % conn.name,
            'right_title': '%s · inputs' % config.name,
            'subtitle': _("Map %s fields onto this scheme's input components") % conn.name,
            'supports_suggest': False,
            'contexts': contexts, 'context_id': conn.id,
            'endpoints': endpoints, 'endpoint_id': ep.id if ep else False,
            'source_summary': self._source_summary(conn, ep, fields_),
            'can_edit': self._can_edit(),
        }

    def _source_summary(self, conn, ep, fields_):
        """What the FROM column's sub-line is entitled to say.

        Cycle 5's sub-line read `206 fields · never synced`, and both halves
        were true while the sentence as a whole was a lie: the 206 were Odoo's,
        and the reader had every reason to think they were Zoho's. The rule now
        is that the count and its ORIGIN have to agree, so the origin is
        computed here — beside the list it describes — rather than inferred in
        the browser from a number.

        `fetch` is a capability, not a credential: it is three booleans and a
        sentence, and no value of `api_key`, `password` or either token can
        reach this payload (`_has_credentials` sudo-reads them and returns a
        bool).
        """
        counts = {}
        for f in fields_:
            counts[f.get('provenance') or 'live'] = \
                counts.get(f.get('provenance') or 'live', 0) + 1
        drift = len([f for f in fields_ if f.get('expected_missing')])
        try:
            cap = conn.field_fetch_capability()
        except Exception:                 # pragma: no cover — older server
            cap = {'mode': None, 'ready': False, 'reason': ''}
        return {
            'total': len(fields_),
            'live': counts.get('live', 0),
            'catalog': counts.get('catalog', 0),
            'odoo': counts.get('odoo', 0),
            'drift': drift,
            'vendor': conn.name or '',
            'feed': (ep.name or ep.code) if ep else '',
            'last_sync': (fields.Datetime.to_string(ep.last_sync)
                          if (ep and ep.last_sync) else ''),
            'ever_synced': bool(
                self.env['hr.api.data.store'].search_count(
                    [('connector_id', '=', conn.id)])),
            'fetch_mode': cap.get('mode') or '',
            'fetch_ready': bool(cap.get('ready')),
            'fetch_reason': cap.get('reason') or '',
        }

    # `_infer_source_type`'s vocabulary is not the mapping's. The store infers
    # `string/integer/float/boolean/date/datetime/list`; the field carries
    # `string/number/integer/float/date/datetime/boolean/currency`. `list` and
    # anything unrecognised fall through to no opinion rather than to a wrong
    # one — `source_data_type` decides whether `preview_transform` parses the
    # sample as a float, so a guess here is a preview that disagrees with sync.
    _SRC_TYPE = {'string': 'string', 'integer': 'integer', 'float': 'float',
                 'boolean': 'boolean', 'date': 'date', 'datetime': 'datetime'}

    #: Field origins a mapping may point at. `live` is a value this connector
    #: has actually received; `computed` is a key this platform itself adds
    #: through a transformation rule, which exists whether or not a record has
    #: been pulled yet. Everything else — `catalog` (a field the vendor's
    #: documentation says exists) and `odoo` (our own schema) — is an
    #: EXPECTATION, and mapping payroll to an expectation is how a component
    #: silently falls back to its default on every run.
    _MAPPABLE_PROVENANCE = ('live', 'computed')

    def _field_provenance(self, conn, path):
        """The discovery item for `path`, or `{}` — read its `provenance`.

        SC-1 widened the return from a bare provenance string to the whole
        item: the refusal below now also needs `origin`, because a catalogue
        row OBSERVED in real data is not the same claim as one copied from
        shipped paper. Single caller (`_refuse_unarrived_field`).
        """
        FM = self.env['hr.integration.field.mapping']
        try:
            fields_ = FM.get_available_source_fields(conn.id) or []
        except Exception:       # noqa: BLE001 — a lookup must not block a draw
            return {}
        return next((f for f in fields_ if f.get('path') == path), None) or {}

    def _connector_has_live_fields(self, conn):
        """Has anything at all ever arrived from this connected system?"""
        Store = self.env.get('hr.api.data.store')
        if Store is None:
            return True         # cannot tell — do not stand in the way
        return bool(Store.sudo().search_count([('connector_id', '=', conn.id)]))

    def _refuse_unarrived_field(self, conn, path):
        """Refuse a mapping to a field this system has never sent, and say why.

        Reported by the owner after three empty pay runs: Base Salary was
        mapped to `Salary`, a field Zoho's catalogue advertises and has never
        delivered. The board said so — an orange NOT SENT badge and an "e.g."
        before the number — and it was still possible to wire payroll to it,
        after which the component fell back to its default every run while the
        run reported success.

        A transformation rule's output is exempt: this platform computes it, so
        it legitimately exists before any record is pulled.
        """
        found = self._field_provenance(conn, path)
        prov = found.get('provenance') or ''
        if prov in self._MAPPABLE_PROVENANCE:
            return None
        # SC-1 — a catalogue row the observation pass has SEEN in real data is
        # a real field. The recent-rows sample window can miss a sparse key
        # (one that only some people carry), and refusing to map a field that
        # has genuinely arrived would be this guard overclaiming in the other
        # direction.
        if prov == 'catalog' and found.get('origin') == 'observed':
            return None
        if not self._connector_has_live_fields(conn):
            return {'ok': False, 'needs_fetch': True, 'connector_id': conn.id,
                    'msg': _(
                        "Nothing has been fetched from %(sys)s yet, so there is "
                        "no way to tell which fields it really sends. Fetch its "
                        "fields first, then map — otherwise a component can be "
                        "wired to a field that never arrives.",
                        sys=conn.display_name or '')}
        return {'ok': False, 'msg': _(
            "%(sys)s has never sent '%(field)s' — it is listed because the "
            "vendor says it exists, and the sample beside it is an example, not "
            "your data. Map to a field that has actually arrived, or fetch "
            "again if you expect this one to be there.",
            sys=conn.display_name or '', field=path)}

    def fetch_live_fields(self, connector_id, period_from=None, period_to=None):
        """Pull a little real data from every runnable feed, so mapping has facts.

        The answer to "what do I map against before the first sync?". It pulls
        each executable feed once; afterwards the board's FROM column shows what
        the system actually sends, with real samples, instead of the vendor's
        catalogue with invented ones.
        """
        conn = self.env['hr.integration.connector'].sudo().browse(
            int(connector_id or 0)).exists()
        if not conn:
            return {'ok': False, 'msg': _("That connected system no longer exists.")}
        done, failed = [], []
        for ep in conn.endpoint_ids:
            if not ep.active or (ep.operation or 'catalog_only') == 'catalog_only':
                continue
            if conn.connector_type == 'zoho' and not ep.path:
                continue
            try:
                conn.action_pull_endpoint(ep.id, period_from=period_from,
                                          period_to=period_to,
                                          triggered_by='manual')
                done.append(ep.name or '')
            except Exception as exc:        # noqa: BLE001
                _logger.warning("fetch_live_fields: %s failed: %s", ep.name, exc)
                failed.append(ep.name or '')
        return {
            'ok': bool(done),
            'fetched': done, 'failed': failed,
            'msg': (_("Fetched %(n)s feed(s). The field list now shows what "
                      "%(sys)s actually sends.", n=len(done),
                      sys=conn.display_name or '') if done
                    else _("Nothing could be fetched from %(sys)s.",
                           sys=conn.display_name or '')),
        }

    def _endpoint_for_field(self, conn, path):
        """The feed on `conn` whose catalogue carries `path`, when only one does.

        A wire has to name its feed or nothing can fetch it: the pay run's sync
        plan is derived from `endpoint_id`, so a wire without one is pulled by
        nothing and its component falls to a default on every run — silently,
        because the run still reports success. On the reference tenant that is
        exactly what happened to BASESALARY, and every deduction (a percentage
        of base pay) came out ₫0 with it.

        The board only sends an endpoint when the reader has picked one feed;
        in the "All feeds" view it sends none, and the wire was created blank
        even though the catalogue knows the answer. So look it up.

        Silent when AMBIGUOUS on purpose. Two feeds carrying a field of the same
        name is a real shape (Zoho's Employees and Salary form both expose an
        id), and guessing between them would wire payroll to the wrong feed —
        worse than leaving it for a person, which the run now reports.
        """
        Endpoint = self.env.get('hr.integration.endpoint')
        if Endpoint is None or not path:
            return None
        Field = self.env.get('hr.integration.endpoint.field')
        if Field is None:
            return None
        try:
            hits = Field.sudo().search([('path', '=', path),
                                        ('endpoint_id.connector_id', '=', conn.id)])
        except Exception:       # noqa: BLE001 — a lookup must not block a draw
            return None
        endpoints = hits.mapped('endpoint_id')
        return endpoints[:1] if len(endpoints) == 1 else None

    def _discovered_sample(self, conn, path, endpoint=None):
        """What the board is already showing for `path`, as writable vals.

        Returns `{}` when the field is not in the discovered set (a template
        line naming a path this connector has never delivered, say) — an absent
        sample is the honest answer, and `preview_transform` already has a
        first-class "no sample stored" branch for it.
        """
        FM = self.env['hr.integration.field.mapping']
        try:
            fields_ = FM.get_available_source_fields(
                conn.id, endpoint.data_type if endpoint else None,
                endpoint.id if endpoint else None) or []
        except Exception:
            return {}
        found = next((f for f in fields_ if f.get('path') == path), None)
        if not found:
            return {}
        vals = {}
        text = self._sample_text(found.get('sample'))
        if text:
            vals['source_sample_value'] = text
        t = self._SRC_TYPE.get(found.get('type'))
        if t:
            vals['source_data_type'] = t
        return vals

    @api.model
    def api_mapping_create(self, config_id, connector_id, source_field,
                           target_rule_id, endpoint_id=None, resolve=None):
        """Draw an API wire, stamped with the feed it was drawn on.

        JOURNEY J3 S2 — `resolve` is the user's answer to the conflict dialog, and
        it is OPTIONAL and defaults to today's behaviour, so every existing caller
        and every test is unchanged:

          * `'replace'` — the other live source goes. A spreadsheet binding is
            overwritten (which is what happens anyway), AND wires on other
            connections are unlinked, which is the part that never used to happen.
          * `'keep'`   — both survive. The wire is drawn and an existing SPREADSHEET
            binding is left in place as the fallback the resolver's empty-feed
            guard now genuinely honours (J3 S2 in `payroll_import_batch`). Without
            that guard this option would be a lie, which is why the two shipped
            together.
          * `None`     — exactly what this method did before J3.

        `endpoint_id` is validated against THIS connector's feeds rather than
        trusted: an id from the browser that named another connector's feed
        would file the mapping under a feed that cannot produce it, and every
        count on both screens would then be wrong in a way nothing errors on.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        # MAPFIX D1 — the same wrong-type guard as `employee_mapping_create`. This
        # board's left ids are strings and its right ids are integers, so a
        # client-side mix-up sends each of them where the other belongs; `int()`
        # on `'f:account_number'` is a ValueError dialog, and `.startswith` on an
        # integer an AttributeError. Both are refusals, not crashes.
        src = self._ec_spec(source_field)
        src = src[2:] if src.startswith('f:') else src
        FM = self.env['hr.integration.field.mapping']
        rule = self.env['hr.formula.rule'].browse(self._as_id(target_rule_id))
        conn = self.env['hr.integration.connector'].browse(self._as_id(connector_id))
        if not (src and rule.exists() and conn.exists()):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        sealed = self._mc_refuse_sealed(rule)
        if sealed:
            return sealed
        # SC-4 — the api lane's refusal on the board itself.
        cfg = rule.config_id
        if cfg and not getattr(cfg, 'source_api_enabled', True):
            return {'ok': False, 'msg': _(
                "This scheme does not read the connected system — its "
                "sources settings switched that off. Turn the lane back on "
                "in the scheme's settings to draw this line.")}
        refusal = self._refuse_unarrived_field(conn, src)
        if refusal:
            return refusal
        vals = {'connector_id': conn.id, 'source_field': src,
                'target_rule_id': rule.id,
                'source_field_label': (src or '').replace('_', ' ').title()}
        ep_wanted = self._as_id(endpoint_id)
        ep = None
        if ep_wanted:
            eps = self._api_endpoints(conn)
            ep = (eps.filtered(lambda e: e.id == ep_wanted)[:1]
                  if eps is not None else None)
            if ep:
                vals['endpoint_id'] = ep.id
        if not vals.get('endpoint_id'):
            # The reader was looking at "All feeds", so the board named no
            # endpoint — but the catalogue knows which one carries this field.
            ep = self._endpoint_for_field(conn, src)
            if ep:
                vals['endpoint_id'] = ep.id
        # The board ALREADY knows what this field looks like — every left card
        # prints its sample. Dropping it on create left the new wire with an
        # empty `source_sample_value`, so the very next thing a user does —
        # open the transform popover — answered "No sample value stored" about
        # a field whose sample is on screen two inches to the left. Found on
        # the live pass. One lookup per create, which is a user action and the
        # same call the board makes on every read.
        vals.update(self._discovered_sample(conn, src, ep))
        # one source→one input per connector: drop existing on either side
        FM.search(['&', ('connector_id', '=', conn.id),
                   '|', ('source_field', '=', src), ('target_rule_id', '=', rule.id)]).unlink()
        # J3 S2 — "replace" reaches ACROSS connections. The search above has only
        # ever tidied within one connector, which is exactly why a component wired
        # on two connections could exist at all (abm has several).
        if resolve == 'replace':
            FM.search([('target_rule_id', '=', rule.id),
                       ('connector_id', '!=', conn.id)]).unlink()
        FM.create(vals)
        # SOURCING S6 — and the wire is now a BINDING as well as a mapping.
        #
        # S3 built `source_binding` and the resolver honours it; until now nothing
        # in the product wrote one, so the owner's requirement — "each component
        # clearly identified as to which source feeds it" — had a data model and no
        # door. Drawing a wire IS the declaration, on this board and on the Excel
        # board equally, so both write one and neither is the special case.
        #
        # `rule` when the path is one this connector COMPUTES, `feed` when the
        # vendor delivers it. The distinction is the whole vocabulary: "Rule output"
        # is a thing with lineage behind it, "Connected system" is a thing the
        # vendor sent.
        kind = 'rule' if src in FM._computed_output_keys(conn) else 'feed'
        replaced = self._binding_replaced(rule, kind, src)
        # JOURNEY J9 — the special case J3 needed here is GONE, and it is gone
        # because `set_source_binding` no longer overwrites another kind. Adding
        # a feed source to a component bound to a spreadsheet column simply adds
        # a row; the spreadsheet declaration survives by construction rather than
        # by this branch remembering to protect it. `replace` is the one answer
        # that still removes something, and it now says so explicitly.
        if resolve == 'replace':
            rule.clear_source_binding('excel')
        rule.set_source_binding(kind, src, origin='board')
        return {'ok': True, 'replaced': replaced}

    @api.model
    def source_conflict_probe(self, config_id, board, target_rule_id,
                              key=None, connector_id=None):
        """WOULD drawing this create a second live source? **Writes nothing.**

        JOURNEY J3 S2 (owner decision J-D3). The either-API-or-Excel rule used to
        be enforced by whichever code happened to run last, silently. It is now an
        explicit choice — and a choice needs to be offered BEFORE the write, which
        is why this is a separate read-only adapter rather than a flag on the
        create. The cancel path of the dialog therefore makes no writing RPC at
        all: there is nothing to roll back because nothing was ever sent.

        `board` is `'api'` or `'import'` — which side the user is drawing from.
        Returns `{'ok': True, 'conflict': False}` when the draw is unremarkable
        (including every same-source redraw: excel→excel and a rewire on the same
        connection keep today's silent swap and its toast), otherwise a dict of
        finished SENTENCES for the dialog. The client renders text; it never
        composes it, so the wording lives with the ladder it describes.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        rule = self.env['hr.formula.rule'].browse(self._as_id(target_rule_id))
        if not rule.exists():
            return {'ok': True, 'conflict': False}
        incoming_key = (self._ec_spec(key) or '').strip()
        for prefix in ('c:', 'f:'):
            if incoming_key.startswith(prefix):
                incoming_key = incoming_key[2:]
        FM = self.env.get('hr.integration.field.mapping')
        # `env.get` returns an EMPTY RECORDSET, which is falsy — `if FM` here would
        # take the else branch on every call and the probe would report "no
        # conflict" forever, silently. The absence test is `is None`.
        wires = (FM.sudo().search([('target_rule_id', '=', rule.id)])
                 if FM is not None else self.env['hr.formula.rule'].browse())
        b_kind = rule.source_binding or False
        b_key = (rule.source_binding_key or '').strip()
        if b_kind and not b_key:
            b_kind = False
        code = rule.code or rule.name or ''

        if board == 'import':
            # Drawing a SPREADSHEET column onto a component that a live feed
            # already wires.
            if not wires:
                return {'ok': True, 'conflict': False}
            names = list(dict.fromkeys(
                [w.connector_id.name or _("Unnamed connection") for w in wires]))
            conn = ', '.join(names)
            order = self._probe_order(rule, 'excel', incoming_key)
            return {'ok': True, 'conflict': {
                'shape': 'excel_over_feed',
                'title': _("“%s” will read more than one source") % code,
                'body': self._probe_body(code, order),
                'order': order,
                'existing': {'label': self._source_label('feed'), 'key': conn},
                'incoming': {'label': self._source_label('excel'),
                             'key': incoming_key},
                'keep_label': _("Add source"),
                'keep_note': self._probe_add_note(order),
                'replace_label': _("Use the spreadsheet instead"),
                'replace_note': _(
                    "Removes the wire from %s. Only the spreadsheet column feeds "
                    "this component after that.") % conn,
                'cancel_label': _("Cancel"),
            }}

        # board == 'api': drawing a WIRE onto a component that already reads
        # somewhere else.
        conn = self.env['hr.integration.connector'].browse(self._as_id(connector_id))
        others = [w for w in wires if w.connector_id.id != conn.id]
        FMc = self.env['hr.integration.field.mapping']
        try:
            in_kind = ('rule' if incoming_key in FMc._computed_output_keys(conn)
                       else 'feed')
        except Exception:       # noqa: BLE001 — a probe must never break a draw
            in_kind = 'feed'
        if b_kind == 'excel':
            order = self._probe_order(rule, in_kind, incoming_key)
            return {'ok': True, 'conflict': {
                'shape': 'feed_over_excel',
                'title': _("“%s” will read more than one source") % code,
                'body': self._probe_body(code, order),
                'order': order,
                'existing': {'label': self._source_label('excel'), 'key': b_key},
                'incoming': {'label': self._source_label(in_kind),
                             'key': incoming_key},
                'keep_label': _("Add source"),
                'keep_note': self._probe_add_note(order),
                'replace_label': _("Use the feed instead"),
                'replace_note': _(
                    "Clears the spreadsheet binding. Only %s feeds this component "
                    "after that.") % (conn.name or _("this connection")),
                'cancel_label': _("Cancel"),
            }}
        if others:
            # TWO CONNECTIONS is still a genuine conflict rather than a
            # precedence, and J9 did not change that: a pay run reads only the
            # connection its scheme is set to, so the second wire is not a
            # fallback — it is inert. Saying "add source" here would promise an
            # order the resolver has no way to honour.
            names = list(dict.fromkeys(
                [w.connector_id.name or _("Unnamed connection") for w in others]))
            other = ', '.join(names)
            return {'ok': True, 'conflict': {
                'shape': 'feed_over_feed',
                'title': _("“%s” is already wired to another connection") % code,
                'body': _(
                    "%(code)s is wired to %(other)s. A pay run reads only the "
                    "connection its scheme is set to, so two wires means one of "
                    "them is doing nothing.", code=code, other=other),
                'existing': {'label': self._source_label('feed'), 'key': other},
                'incoming': {'label': self._source_label('feed'),
                             'key': incoming_key},
                'keep_label': _("Keep both wires"),
                'keep_note': _(
                    "Both wires stay. Whichever connection the scheme is set to "
                    "is the one a pay run reads."),
                'replace_label': _("Use %s instead") % (
                    conn.name or _("this connection")),
                'replace_note': _("Removes the wire from %s.") % other,
                'cancel_label': _("Cancel"),
            }}
        return {'ok': True, 'conflict': False}

    # ------------------------------------------------------------------
    # JOURNEY J9 — the dialog stopped being an ultimatum and became a NOTICE.
    #
    # J3 offered replace / keep-as-fallback / cancel because the owner's rule at
    # the time was one source per component and the dialog had to make the loser
    # explicit. That rule is withdrawn. The useful thing to say now is the
    # resulting ORDER, so the primary action is "Add source" and Replace stays as
    # the secondary — the old behaviour is still one click away, it is simply no
    # longer the default answer to a question that is no longer a conflict.
    #
    # Every sentence is composed HERE, beside the ladder it describes. A client
    # that built this text would be a second opinion about precedence.
    # ------------------------------------------------------------------
    @api.model
    def _probe_order(self, rule, in_kind, in_key):
        """The ranked list this component WOULD read, if the draw went ahead."""
        rank = self._source_rank()
        specs = [(s['kind'], s['key']) for s in rule.declared_sources()
                 if s['kind'] in rank]
        specs = [s for s in specs if s[0] != in_kind]
        specs.append((in_kind, (in_key or '').strip()))
        # A live wire the component already carries is a declared source too, and
        # leaving it out would make the dialog contradict the card behind it.
        wired = self._source_wire_dests(rule.config_id).get(rule.id)
        if wired and wired['kind'] != in_kind \
                and (wired['kind'], (wired['key'] or '').strip()) not in specs:
            specs.append((wired['kind'], (wired['key'] or '').strip()))
        # J10 — and so is the record this component is copied onto, for exactly
        # the same reason one line up: the dialog must not describe a shorter
        # order than the card behind it is showing.
        dest = self._source_record_dests(rule.config_id).get(rule.id)
        if dest and (dest['kind'], dest['key']) not in specs:
            specs.append((dest['kind'], dest['key']))
        specs.sort(key=lambda s: rank.index(s[0]))
        out = [{'label': self._source_label(k), 'key': key, 'rank': i + 1}
               for i, (k, key) in enumerate(specs)]
        if rule.is_contract_component:
            out.append({'label': self._source_label('contract_component'),
                        'key': '', 'rank': len(out) + 1})
        return out

    @api.model
    def _probe_body(self, code, order):
        if not order:
            return ''
        names = ', '.join(o['label'] for o in order)
        return _(
            "%(code)s will read %(names)s, in that order. The first one that "
            "carries a value wins; the others are read only when it does not.",
            code=code, names=names)

    @api.model
    def _probe_add_note(self, order):
        if not order:
            return ''
        return _("Both stay. %s is tried first.") % order[0]['label']

    @api.model
    def _binding_replaced(self, rule, kind, key):
        """What this component was reading BEFORE, when it is about to change.

        The owner's rule is that switching a component from one source to the other
        is a single deliberate act — so it has to be an act that says what it did.
        `None` when nothing is being displaced, which is the common case and stays
        silent.

        JOURNEY J9 — displaced now means "the row of THIS kind changed key".
        Adding a feed source beside a spreadsheet column displaces nothing, which
        is the whole point of the phase; reporting it as a replacement would be
        the dialog and the toast disagreeing about what just happened.
        """
        same = rule.source_ids.filtered(lambda s: s.kind == kind)
        old_kind = kind if same else False
        old_key = (same[0].key or '').strip() if same else ''
        if not old_kind or old_key == (key or '').strip():
            return None
        return {'kind': old_kind, 'key': old_key,
                'label': self._source_label(old_kind),
                'msg': _("“%(code)s” now reads %(new)s “%(key)s” instead of "
                         "%(old)s “%(old_key)s”.",
                         code=rule.code or rule.name or '',
                         new=self._source_label(kind), key=key,
                         old=self._source_label(old_kind), old_key=old_key)}

    @api.model
    def _wire_source_kinds(self, rule, mapping):
        """Which declared source kinds THIS wire wrote — snapshot, never live.

        JOURNEY J9. Returned as a plain list of strings because the two callers
        both mutate `source_ids` afterwards, and iterating a recordset that the
        loop body is unlinking from is how a delete quietly skips a row.
        """
        want = (mapping.source_field or '').strip()
        return [s.kind for s in rule.source_ids
                if s.kind in ('feed', 'rule') and (s.key or '').strip() == want]

    @api.model
    def api_mapping_delete(self, mapping_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        m = self.env['hr.integration.field.mapping'].browse(self._as_id(mapping_id))
        if m.exists():
            # SOURCING S6 — removing the wire removes the wire. The binding is
            # cleared only when it is THIS wire's binding: a component someone has
            # since re-bound to a spreadsheet column must not lose that because an
            # old feed mapping was tidied away.
            rule = m.target_rule_id
            # JOURNEY J9 — clear THIS wire's kind, not every source the component
            # has. Under S3 the two were the same act because there was only ever
            # one binding; with several declared, `set_source_binding(False, …)`
            # would take the spreadsheet column with it — which is exactly the
            # "must not lose that" the comment above has always claimed.
            if rule:
                for kind in self._wire_source_kinds(rule, m):
                    rule.clear_source_binding(kind)
            m.unlink()
        return {'ok': True}

    # The business spec of a wire, for J6's undo. Deliberately a LIST rather than
    # "every stored field": `display_name`, `is_mapped`, `target_column_letter`,
    # `target_rule_code`, `connector_type` and the transform-error pair are all
    # DERIVED from the four fields above them, and writing a derived value back is
    # how a restore quietly disagrees with a recompute. The intersection with
    # `fields_get` below means a field dropped in a later version simply stops
    # being carried instead of raising on the restore.
    _J6_WIRE_SPEC = (
        'connector_id', 'endpoint_id', 'target_rule_id', 'source_field',
        'source_field_label', 'source_data_type', 'source_sample_value',
        'transformation_type', 'transformation_code', 'transformation_value',
        'transformation_decimals', 'default_value', 'min_value', 'max_value',
        'notes', 'sequence', 'active_state', 'is_required', 'active',
    )

    @api.model
    def _j6_wire_fields(self):
        """Those of `_J6_WIRE_SPEC` this database can actually be handed back."""
        info = self.env['hr.integration.field.mapping'].fields_get(
            list(self._J6_WIRE_SPEC), ['type', 'store', 'readonly'])
        return [f for f, d in info.items()
                if d.get('store') and not d.get('readonly')]

    @api.model
    def api_mapping_snapshot(self, mapping_id):
        """Everything needed to put this wire back, read-only.

        Includes whether the TARGET's binding is this wire's, because
        `api_mapping_delete` clears it in that case and an undo that restored the
        row without the binding would put back two-thirds of a decision.
        """
        m = self.env['hr.integration.field.mapping'].browse(
            self._as_id(mapping_id))
        if not m.exists():
            return False
        spec = {}
        for f in self._j6_wire_fields():
            v = m[f]
            spec[f] = v.id if hasattr(v, 'id') else v
        rule = m.target_rule_id
        binding = False
        # JOURNEY J9 — asked of the SOURCE ROW rather than of the derived
        # `source_binding`, which now reports only the highest-ranked one. Cutting
        # a `rule` wire off a component that also reads a feed must still be
        # undoable, and the derived field would have said "feed" and lost it.
        kinds = self._wire_source_kinds(rule, m) if rule else []
        if kinds:
            src = rule.source_ids.filtered(lambda s: s.kind == kinds[0])[0]
            binding = {'rule_id': rule.id, 'kind': src.kind,
                       'key': src.key or '',
                       'origin': src.origin or 'user'}
        return {'spec': spec, 'binding': binding,
                'label': m.display_name or m.source_field or ''}

    @api.model
    def api_mapping_cut(self, mapping_id):
        """Snapshot, then delete — one round trip, one delete implementation.

        JOURNEY J6 D3. The snapshot is taken SERVER-side and in the same call as
        the delete so there is no window in which the client holds an id whose row
        has already changed underneath it. The delete itself is still
        `api_mapping_delete`: this method adds an undo, it does not add a second
        way to remove a wire.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        snap = self.api_mapping_snapshot(mapping_id)
        res = self.api_mapping_delete(mapping_id)
        if res.get('ok'):
            res['snapshot'] = snap
        return res

    @api.model
    def api_mapping_restore(self, snapshot):
        """Put a cut wire back exactly as it was.

        NOT routed through `api_mapping_create`, and the reason is that they are
        different verbs. `api_mapping_create` DRAWS: it re-derives the label from
        the source field, discovers a fresh sample, unlinks whatever else occupied
        either end, honours a conflict resolution and writes a binding. Every one
        of those is right for a person drawing a wire and wrong for an undo, which
        must be the delete's inverse and nothing else — the original
        `source_field_label` ("Overtime 300% hours"), the original `notes` and the
        original transform settings are precisely what a re-draw would throw away.
        The id necessarily differs; the ORM cannot mint an old one (J3's
        precedent).
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        if not snapshot or not snapshot.get('spec'):
            return {'ok': False, 'msg': _("There is nothing to put back.")}
        spec = snapshot['spec']
        FM = self.env['hr.integration.field.mapping']
        conn = self.env['hr.integration.connector'].browse(
            self._as_id(spec.get('connector_id')))
        rule = self.env['hr.formula.rule'].browse(
            self._as_id(spec.get('target_rule_id')))
        if not (conn.exists() and rule.exists() and spec.get('source_field')):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        # Undo pressed twice, or a re-draw in the meantime: the wire is already
        # there, and putting a second one back would leave the create path to
        # tidy up after the undo. Idempotent instead.
        dup = FM.with_context(active_test=False).search(
            [('connector_id', '=', conn.id),
             ('source_field', '=', spec['source_field']),
             ('target_rule_id', '=', rule.id)], limit=1)
        if dup:
            return {'ok': True, 'id': dup.id, 'already': True}
        allowed = self._j6_wire_fields()
        new = FM.create({f: v for f, v in spec.items() if f in allowed})
        binding = snapshot.get('binding')
        if binding:
            self.env['hr.formula.rule'].browse(
                self._as_id(binding.get('rule_id'))).set_source_binding(
                    binding.get('kind'), binding.get('key'),
                    origin=binding.get('origin') or 'user')
        return {'ok': True, 'id': new.id}

    # ------------------------------------------------------------------
    # JOURNEY J4 — the Transformations board.
    #
    # A transformation rule has always been a first-class source: its output key
    # is a legal `source_field`, so a wire can carry it, and `set_source_binding`
    # accepts kind `rule`. What it never had was an ADDRESS. The rule was authored
    # in the Integrations cockpit, its output appeared on the API board as one more
    # card in a lane called "Derived here", and the fields it reads were visible
    # nowhere at all. Three facts about one object, on three screens, none of which
    # said the other two existed.
    #
    # `transform_flow_data` is the whole of J4's server side: ONE read-only RPC
    # that composes the three lanes out of pieces that already exist. It defines
    # nothing new. In particular it does NOT define "unread" — that predicate
    # lives in `pb.integrations._rule_consumers`, beside the cockpit hint that
    # first said it out loud, and is CALLED here. A second definition of "nothing
    # reads this" is exactly how two screens come to disagree about a rule.
    # ------------------------------------------------------------------
    @api.model
    def _tf_consumers(self, rule):
        """Who reads this rule's output — ONE definition, borrowed not copied.

        `pb.integrations` owns it (`_rule_consumers`), because that is where the
        "Rule outputs nothing reads" hint is raised and where a reader first meets
        the concept. It is an AbstractModel, so it is in `self.env` whenever the
        module is installed — which, since J4, `pb_formula_studio` depends on.

        The `is None` test is deliberate and is MJ16, verbatim: `self.env.get(...)`
        returns an EMPTY RECORDSET for a missing model and an empty recordset is
        FALSY, so `if Model:` would take the fallback branch on every single call
        and this board would report every output as unread, forever, silently.
        """
        Itg = self.env.get('pb.integrations')
        if Itg is None:
            return []
        try:
            return Itg._rule_consumers(rule)
        except Exception as e:                                  # pragma: no cover
            _logger.warning("J4: rule consumers failed for rule %s: %s: %s",
                            rule.id, type(e).__name__, e)
            return []

    @api.model
    def _tf_active_connector(self, config, connector_id=None):
        """Which connector this board opens on — rules first.

        `_api_active_connector` is the right heuristic for the API board and the
        WRONG default for this one, and abm proves it: it answers connector 1
        (Zoho People, which the scheme's wires point at) while all eight
        transformation rules live on connector 3 (Zoho People (ABM)). The API
        board is asking "where do this scheme's values come from"; this board is
        asking "what does this system compute", and a Transformations tab that
        opens on the one connector with no transformations shows an empty state
        over a database with eight rules in it.

        So: an EXPLICIT choice always wins (the picker and the deep link are
        never second-guessed); otherwise, if the heuristic's answer has no rules
        and some reachable connector does, take that one. The heuristic is not
        modified — a shared helper that changed under four other callers to suit
        this board would be exactly the fork this phase is avoiding.
        """
        asked = self._as_id(connector_id)
        if asked:
            return self._api_active_connector(config, asked)
        conn = self._api_active_connector(config, None)
        Rule = self.env.get('hr.api.transformation.rule')
        if Rule is None:
            return conn
        try:
            if conn and Rule.with_context(active_test=False).search_count(
                    [('connector_id', '=', conn.id)]):
                return conn
            # An ARCHIVED connector is not a default. A rule survives its
            # connector being archived (that is why the rule search ignores
            # `active`), so without this the board would open on a system
            # somebody has deliberately retired — and `no_connector` would
            # become unreachable on any database that ever had a rule.
            for rule in Rule.with_context(active_test=False).search(
                    [], order='connector_id'):
                if rule.connector_id and rule.connector_id.active:
                    return rule.connector_id
        except Exception as e:                                  # pragma: no cover
            _logger.warning("J4: rule-bearing connector lookup failed: %s: %s",
                            type(e).__name__, e)
        return conn

    @api.model
    def transform_flow_data(self, config_id=None, connector_id=None):
        """The three-lane flow: feed fields → transformation rules → components.

        Read-only and additive. Every write this board can make goes out through
        `api_mapping_create` / `api_mapping_delete`, unchanged — an output key is
        already a legal `source_field` and those adapters already classify such a
        wire as kind `rule` (see the classification line in `api_mapping_create`),
        so J4 needed no new write path and deliberately built none.

        The right lane is `_mc_right_column` with `board='api'` — the SAME cards,
        chips, sealing, lineage and conflict pills the API board renders. Two
        boards that show a component differently are two boards a reader has to
        reconcile; and it is what makes the J3 conflict dialog fire here without a
        second implementation.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        Conn = self.env['hr.integration.connector']
        conn = self._tf_active_connector(config, connector_id)
        contexts = [{'id': c.id, 'name': c.name} for c in Conn.search([], order='name')]
        if not conn:
            return {'ok': False, 'reason': 'no_connector', 'contexts': contexts}
        Rule = self.env.get('hr.api.transformation.rule')
        if Rule is None:                                        # pragma: no cover
            return {'ok': False, 'reason': 'no_rules', 'contexts': contexts,
                    'context_id': conn.id}
        # `active_test=False` for the same reason the cockpit hint uses it: an
        # archived rule whose output a component still reads is precisely the
        # state a person needs to SEE, and filtering it out would make the wire
        # into that component point at nothing.
        rules = Rule.with_context(active_test=False).search(
            [('connector_id', '=', conn.id)], order='sequence, id')

        # ---- what the connector is KNOWN to deliver, for the drift verdict ----
        # Same call the API board's FROM column makes, same swallow-and-log
        # policy: a catalogue that cannot be read must not take the board down,
        # but it must not be silent either (the Cycle-6 lesson at :4974).
        try:
            catalogue = self.env['hr.integration.field.mapping'] \
                .get_available_source_fields(conn.id, None, None) or []
        except Exception as e:
            _logger.warning(
                "J4: source-field discovery failed for connector %s: %s: %s — "
                "the reads lane will render without drift verdicts.",
                conn.id, type(e).__name__, e)
            catalogue = []
        by_path = {f['path']: f for f in catalogue if f.get('path')}
        # A rule's OWN output is in that catalogue under "Derived here"; a rule
        # that reads another rule's output is reading something real. Both are
        # "known", neither is vendor drift.
        known = set(by_path)

        # ---- the middle lane: one sealed card per rule -------------------------
        lineage_by_key = self._lineage_by_output_key(conn, config)
        input_ids = set(config.rule_ids.filtered(
            lambda r: r.column_type == 'input').ids)
        FM = self.env['hr.integration.field.mapping']
        wires, read_edges, fields_used = [], [], {}
        cards, unread, drift_n, severed_n = [], 0, 0, 0
        for r in rules:
            key = (r.output_key or '').strip()
            try:
                reads = list(r._consumed_field_names() or [])
            except Exception as e:                              # pragma: no cover
                _logger.warning("J4: consumed fields failed for rule %s: %s: %s",
                                r.id, type(e).__name__, e)
                reads = []
            consumers = self._tf_consumers(r) if key else []
            # every field this rule reads becomes a LEFT card, once, however many
            # rules read it — the lane is the connector's fields, not the union of
            # per-rule lists (six overtime rules read the same three fields on abm).
            r_drift = False
            for path in reads:
                if path not in fields_used:
                    f = by_path.get(path) or {}
                    fields_used[path] = {
                        'id': 'f:' + path,
                        'label': f.get('label') or path,
                        'sublabel': path,
                        'sample': self._sample_text(f.get('sample')),
                        'group': (_("Derived here")
                                  if f.get('provenance') == 'computed'
                                  else (_("Delivered by this system") if path in known
                                        else _("Not seen from this system"))),
                        'prov': f.get('provenance') or ('live' if path in known else 'missing'),
                        'drift': path not in known,
                        'readers': 0,
                        'meta': {'type': f.get('type') or ''},
                    }
                fields_used[path]['readers'] += 1
                read_edges.append({'id': 'rd%s:%s' % (r.id, path),
                                   'leftId': 'f:' + path, 'ruleId': r.id})
                if path not in known:
                    r_drift = True
            if r_drift:
                drift_n += 1
            if key and not consumers:
                unread += 1
            # ---- the right half of the middle lane's story: what it feeds -------
            r_sev = False
            if key:
                for m in FM.with_context(active_test=False).search(
                        [('connector_id', '=', conn.id), ('source_field', '=', key)]):
                    tgt = m.target_rule_id
                    if not tgt or tgt.id not in input_ids:
                        continue
                    if m.is_severed:
                        r_sev = True
                    wires.append({'id': 'w%s' % m.id, 'ref': m.id, 'bind': False,
                                  'ruleId': r.id, 'rightId': tgt.id,
                                  'severed': bool(m.is_severed),
                                  'state': 'accepted'})
                # a component BOUND to this rule with no wire behind it is fed
                # just as truly — the resolver reads the binding (rung 1). It is
                # not deletable from here, because there is no wire to delete.
                wired_targets = {w['rightId'] for w in wires if w['ruleId'] == r.id}
                for br in config.rule_ids:
                    if br.id in wired_targets or br.id not in input_ids:
                        continue
                    if br.source_binding == 'rule' and \
                            (br.source_binding_key or '').strip() == key:
                        wires.append({'id': 'b%s:%s' % (r.id, br.id), 'ref': 0,
                                      'bind': True, 'ruleId': r.id, 'rightId': br.id,
                                      'severed': False, 'state': 'accepted'})
            if r_sev:
                severed_n += 1
            health = 'ok'
            if r_sev:
                health = 'severed'
            elif key and not consumers:
                health = 'unread'
            elif r_drift:
                health = 'drift'
            cards.append({
                'id': r.id,
                'label': r.name or key or _("Untitled rule"),
                'key': key,
                'summary': r.plain_summary or '',
                'kind': r.rule_type or '',
                'active': bool(r.active),
                'error': r.last_error or '',
                'reads': reads,
                'feeds': consumers,
                'health': health,
                'lineage': lineage_by_key.get(key),
            })
        if not cards:
            return {'ok': False, 'reason': 'no_rules', 'contexts': contexts,
                    'can_edit': self._can_edit(), 'context_id': conn.id,
                    'connector': {'id': conn.id, 'name': conn.name or ''}}

        _acts, _run = self._source_actuals(config)
        _emp = self._source_record_dests(config)
        _wires = self._source_wire_dests(config)
        right = self._mc_right_column(
            config, _acts, _emp, wire_dests=_wires,
            lineage=self._lineage_for_config(config, conn), board='api')
        left = sorted(fields_used.values(), key=lambda f: (f['drift'], f['label'].lower()))
        return {
            'ok': True,
            'can_edit': self._can_edit(),
            'context_id': conn.id,
            'contexts': contexts,
            'connector': {'id': conn.id, 'name': conn.name or ''},
            'left': left,
            'rules': cards,
            'right': right,
            'reads': read_edges,
            'wires': wires,
            'counts': {'rules': len(cards), 'unread': unread,
                       'drift': drift_n, 'severed': severed_n,
                       'fed': len({w['rightId'] for w in wires})},
        }

    # ==================================================================
    # JOURNEY J5 — the Journey. Five lanes, one read, no writes.
    #
    #     Systems ─▶ Feeds & files ─▶ Transformations ─▶ Scheme ─▶ Pay run
    #
    # The programme's showpiece, and the only screen that answers the owner's
    # original question in one glance: where does a pay value come from. Every
    # number on it is a count this database can defend — there are no invented
    # percentages, no liveness bars and no charts (pb_explorer owns analytics).
    #
    # It composes; it does not define. The conflict detector is J3's
    # `_source_conflicts`, the unread-output predicate is J4's `_tf_consumers`
    # (which is `pb.integrations._rule_consumers`), the component picture is the
    # `_declared_source` family, the primary connector is `config.connector_id`.
    # A second opinion about any of those would be a second thing to keep in
    # step, and this programme has spent five phases removing exactly that.
    # ==================================================================

    #: The ONE via -> bucket mapping, written once, server-side (J5's contract).
    #:
    #: `via` answers "why THIS source won" and there are eighteen of them
    #: (`input_provenance.VIAS`). The pay-run lane needs three-and-a-bit families
    #: rather than eighteen columns, and the danger of a families map is that it
    #: goes stale in SILENCE: a nineteenth `via` lands, nothing crashes, and its
    #: values are quietly counted as something they are not. So this dict is
    #: EXHAUSTIVE over the vocabulary and `test_journey_view` fails loudly the
    #: moment the two disagree in either direction.
    #:
    #: The families, and why each row sits where it does:
    #:
    #:   `wired`   — the configured path delivered. A binding, a header/letter
    #:               match, a connector mapping, a worked-days line, or one of
    #:               the two approved-workflow streams. Somebody wired it and it
    #:               answered.
    #:   `fallback`— a lower rung answered because the rung above it was empty.
    #:               `fallback` and `binding_empty` say so in their own names;
    #:               `employee_mapping`, `contract` and `contract_field` are
    #:               J-D4's read-back half — the row that writes the record on
    #:               import is read BACK when the file or feed leaves the value
    #:               empty, which is the definition of a fallback.
    #:   `default` — nothing fed it. A constant is in here on purpose: it is the
    #:               same number for everyone and no source produced it.
    #:   `computed`— the FOURTH bucket, and a deviation from the handover's
    #:               three, taken deliberately. Proration, retro and carryover
    #:               appear as a `via` only when the adjustment INVENTED the code
    #:               (`payroll_import_batch._run_adjustment`), and such a value
    #:               carries `src='calculated'`. It was not wired, it did not
    #:               fall back, and it is not a default — filing it under any of
    #:               the three would be the invented number this phase exists to
    #:               refuse. A fourth honest column costs less than a wrong one.
    _JOURNEY_VIA_BUCKETS = {
        'binding': 'wired',
        'header': 'wired',
        'column_letter': 'wired',
        'connector_mapping': 'wired',
        'worked_days': 'wired',
        'overtime_request': 'wired',
        'business_trip': 'wired',
        'binding_empty': 'fallback',
        'fallback': 'fallback',
        'employee_mapping': 'fallback',
        'contract': 'fallback',
        'contract_field': 'fallback',
        'contract_default': 'default',
        'constant': 'default',
        'default': 'default',
        'proration': 'computed',
        'retro': 'computed',
        'carryover': 'computed',
    }

    #: Buckets in display order. Named here so the board cannot invent a fifth.
    _JOURNEY_BUCKETS = ('wired', 'fallback', 'computed', 'default')

    #: How many payslips of a run the aggregate will read. A processed VN batch
    #: is tens of thousands of rows and each carries a JSON blob; reading all of
    #: them to print six numbers would make the landing tab the slowest screen in
    #: the product. When the cap bites the payload SAYS so (`sampled`), and the
    #: board prints "from the first N of M payslips" rather than a bare number —
    #: an honest scope is a smaller claim than a silent estimate.
    _JOURNEY_SLIP_CAP = 2000

    @api.model
    def _journey_bucket_for_via(self, via):
        """One `via` -> one family. Unknown degrades to `default` and LOGS.

        Never raises: this runs inside a landing page, and a vocabulary that
        gained a word must not be the thing that takes the screen down. The
        Python test is what makes the degradation loud where it matters — at
        the point a developer adds the nineteenth `via`, not at the point a
        user opens a tab.
        """
        bucket = self._JOURNEY_VIA_BUCKETS.get(via or 'default')
        if bucket is None:
            _logger.warning(
                "J5: no bucket for provenance via %r — counted as 'default'. "
                "Add it to _JOURNEY_VIA_BUCKETS.", via)
            return 'default'
        return bucket

    @api.model
    def _journey_aggregate(self, blobs):
        """Aggregate provenance blobs by `src` and by `via`-family.

        `blobs` is an iterable of JSON STRINGS — the raw
        `hr.payslip.formula_input_sources` column, passed in rather than read
        here so the Python tests can exercise the arithmetic on fixtures with
        no payslip table involved (the handover's "fixture-based tests for the
        aggregate path").

        Counts VALUES, not payslips: one payslip with 99 components contributes
        99. `slips` counts the payslips that carried a readable blob, which is
        the denominator the board prints beside it. A blob that will not parse
        is skipped and counted in `unreadable` rather than being allowed to
        look like a payslip with no values.
        """
        by_src = defaultdict(int)
        by_bucket = defaultdict(int)
        fell_back = 0
        values = 0
        slips = 0
        unreadable = 0
        for raw in blobs:
            # An EMPTY column is not an empty payslip. `raw or '{}'` was the
            # first spelling and it quietly turned "this payslip recorded no
            # provenance at all" into "this payslip resolved nothing", which are
            # different facts: the first is a run that predates the provenance
            # writer, the second is a run that went badly. Counting the first as
            # the second would put a payslip in the denominator that has nothing
            # to say about the numerator.
            if not raw:
                unreadable += 1
                continue
            try:
                blob = json.loads(raw)
            except (TypeError, ValueError):
                unreadable += 1
                continue
            if not isinstance(blob, dict):
                unreadable += 1
                continue
            slips += 1
            for entry in blob.values():
                if not isinstance(entry, dict):
                    continue
                values += 1
                by_src[entry.get('src') or 'none'] += 1
                by_bucket[self._journey_bucket_for_via(entry.get('via'))] += 1
                if entry.get('fell_back'):
                    fell_back += 1
        return {
            'slips': slips,
            'values': values,
            'unreadable': unreadable,
            'fell_back': fell_back,
            'by_src': dict(by_src),
            'by_bucket': {b: by_bucket.get(b, 0) for b in self._JOURNEY_BUCKETS},
        }

    @api.model
    def _journey_iso(self, value):
        """A datetime as an ISO string, or ''. The CLIENT formats the age.

        `MappingStudio.since()` already turns one of these into "synced 3d ago"
        and has since Cycle 5; a second age formatter on the server would be a
        second opinion about what "recently" means, told in a different
        timezone. `_import_sample_meta`'s `read_on` is the deliberate exception
        — a file's read date is a provenance line, not an age.
        """
        return fields.Datetime.to_string(value) if value else ''

    @api.model
    def _journey_people_mappings(self, config):
        """The Employee & contract rows for this scheme, split by what they DO.

        J3 S1 settled that these rows are two-way — the same row writes the
        record on import and is read BACK when the file or feed leaves the value
        empty — with ONE exception, and the exception is load-bearing here:
        a `bank_account` row is the import half only, because
        `get_mapped_input_value` reads employee and contract FIELDS back and
        never bank parts. So a bank row is not fallback-capable, and counting it
        as one would put a number on the header sentence that the resolver would
        never honour.

        Returns `(total_rows, {rule_id, ...} read-back-capable, bank_rows)`.
        """
        Mapping = self.env.get('hr.payslip.import.mapping')
        if Mapping is None or not config.rule_ids:
            return 0, set(), 0
        try:
            rows = Mapping.sudo().search(
                [('salary_structure_id', '=', config.id)])
        except Exception:       # noqa: BLE001 — a lane must never break the tab
            return 0, set(), 0
        read_back, bank = set(), 0
        for row in rows:
            if row.destination_type == 'bank_account':
                bank += 1
                continue
            if row.component_id:
                read_back.add(row.component_id.id)
        return len(rows), read_back, bank

    @api.model
    def _journey_scheme_lane(self, config, record_dests, wire_dests):
        """The component picture, every number defended by a direct count.

        One pass over `config.rule_ids`, asking `_declared_source` the same
        question the mapping boards ask it — so the Journey's "42 wired" and the
        API board's forty-two source chips are the same forty-two components,
        not two independently-derived numbers that will drift apart by Tuesday.
        """
        counts = {'total': 0, 'inputs': 0, 'wired': 0, 'constant': 0,
                  'calculated': 0, 'unfed': 0, 'contract': 0, 'people': 0}
        wired_ids = set()
        for rule in config.rule_ids:
            counts['total'] += 1
            declared = self._declared_source(rule, record_dests, wire_dests)
            kind = declared['kind']
            if rule.column_type == 'input':
                counts['inputs'] += 1
            if kind in ('excel', 'feed', 'rule'):
                counts['wired'] += 1
                wired_ids.add(rule.id)
            elif kind == 'calculated':
                counts['calculated'] += 1
            elif kind == 'constant':
                counts['constant'] += 1
            elif kind == 'contract_component':
                counts['contract'] += 1
            elif kind in ('employee_field', 'contract_field', 'bank_account'):
                # J10 — one rung with three spellings. The lane counts people
                # data, and a designation kept on the contract is people data
                # exactly as much as one kept on the employee; splitting the
                # bar would be telling the reader about a mechanism instead of
                # about their scheme.
                counts['people'] += 1
            else:
                counts['unfed'] += 1
        return counts, wired_ids

    @api.model
    def journey_data(self, config_id=None):
        """The whole Journey, in ONE read.

        Read-only, top to bottom. Nothing in this method or anything it calls
        writes a row, and that is the phase's signature proof rather than a
        remark: the MF37 diff around an entire live validation session of this
        tab is empty, with no restore step.

        The payload is five LANES of nodes plus the EDGES between them. Nodes
        are uniform (`id`, `kind`, `label`, `sub`, `tone`, `chip`, `door`,
        `ghost`) so the board renders them with one template and cannot grow a
        per-lane dialect; edges carry `from`, `to`, `kind` and `count`, and the
        board draws a wire only where an edge exists — which is what makes the
        picture a claim about the database rather than a diagram.

        Every lane has a designed GHOST with a working door, because a scheme
        with nothing configured is the novice's first screen and an empty
        five-lane grid would read as a broken feature rather than as an
        invitation.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}

        Conn = self.env['hr.integration.connector']
        FM = self.env.get('hr.integration.field.mapping')
        connectors = Conn.search([], order='name')

        # ---- the primary connector (J-D5, and MJ22's other half) -------------
        # `config.connector_id` and NOTHING ELSE. `_api_active_connector`'s
        # most-mappings heuristic answers a different question (which connector
        # do this scheme's wires point at) and is documented picking the wrong
        # one on abm; the runtime gate in `payroll_import_batch` reads the FIELD.
        # A lane marked "primary" that the runtime disagrees with would be the
        # worst thing this tab could say, because it would be believed.
        primary = config.connector_id if config.connector_id else Conn.browse()

        # ---- what wires exist, once, for the whole board ---------------------
        emp_total, read_back_ids, bank_rows = self._journey_people_mappings(config)
        record_dests = self._source_record_dests(config)
        wire_dests = self._source_wire_dests(config)
        conflicts = self._source_conflicts(config)
        scheme_counts, wired_ids = self._journey_scheme_lane(
            config, record_dests, wire_dests)

        input_ids = set(config.rule_ids.filtered(
            lambda r: r.column_type == 'input').ids)

        # per-connector and per-endpoint counts of LIVE wires into this scheme
        wires_by_conn = defaultdict(int)
        wires_by_ep = defaultdict(int)
        if FM is not None and config.rule_ids:
            try:
                for m in FM.sudo().with_context(active_test=False).search(
                        [('target_rule_id', 'in', list(input_ids))]):
                    wires_by_conn[m.connector_id.id] += 1
                    if m.endpoint_id:
                        wires_by_ep[m.endpoint_id.id] += 1
            except Exception as e:      # noqa: BLE001
                _logger.warning("J5: wire census failed: %s: %s",
                                type(e).__name__, e)

        # ---- severed wires, and the honest scope of that claim ---------------
        # `is_severed` is `target_rule_code AND NOT target_rule_id` — a severed
        # wire has NO target, so it cannot be found by searching for this
        # config's rules and "severed wires of this scheme" is not a question
        # the schema can answer. What it CAN answer is "severed wires on the
        # connection this scheme reads", so that is what is counted and that is
        # what the node says. Attributing an orphaned `target_rule_code` to a
        # config by string-matching would be exactly the undefendable number
        # scope 5 forbids.
        severed_conn_ids = set(wires_by_conn)
        if primary:
            severed_conn_ids.add(primary.id)
        severed_n = 0
        if FM is not None and severed_conn_ids:
            try:
                severed_n = FM.sudo().with_context(active_test=False).search_count(
                    [('connector_id', 'in', list(severed_conn_ids)),
                     ('is_severed', '=', True)])
            except Exception as e:      # noqa: BLE001
                _logger.warning("J5: severed census failed: %s: %s",
                                type(e).__name__, e)

        systems, feeds, transforms, edges = [], [], [], []

        # ================================================== LANE 1 — systems
        # ---- the state the live database is actually in ---------------------
        # NO scheme on any of the four databases has `connector_id` set (the
        # SOURCING ledger's S20). That is not a cosmetic gap: the resolver's
        # pre-pass is gated on it —
        #
        #     if self.source_type == 'api_data_store' and config.connector_id:
        #
        # — so with the field unset, NO feed wire is read on a system run at
        # all. abm has thirty-three wires drawn into this scheme and every one
        # of them is inert. Making the one-connector limit visible (scope 3)
        # therefore has to include the case where the limit has never been
        # exercised, or the tab would show a confident picture of a pipe that
        # is not connected at the tap. Every connector is dimmed, each says why,
        # and the scheme lane raises it as a health node.
        no_primary = not primary and bool(wires_by_conn)
        for conn in connectors:
            is_primary = bool(primary) and conn.id == primary.id
            # The one-connector limit, made VISIBLE (scope 3). A pay run reads
            # only the connection its scheme is set to; every other connector's
            # wires are ignored on a system run. That has always been true and
            # has never been said anywhere, which is how abm ended up with seven
            # components wired on a connection nothing reads.
            dimmed = (bool(primary) and not is_primary) or (
                no_primary and bool(wires_by_conn.get(conn.id)))
            node = {
                'id': 'c:%s' % conn.id, 'kind': 'connector', 'lane': 'systems',
                'label': conn.name or _("Unnamed connection"),
                'sub': '',
                'tone': {'error': 'err', 'connected': 'ok'}.get(
                    conn.connection_status or '', 'muted'),
                'primary': is_primary,
                'dimmed': dimmed,
                'last_sync': self._journey_iso(conn.last_sync),
                'status': conn.connection_status or 'disconnected',
                'wires': wires_by_conn.get(conn.id, 0),
                'door': {'mode': 'api', 'connector': conn.id},
            }
            if is_primary:
                node['chip'] = {
                    'label': _("Primary"), 'tone': 'ok',
                    'hint': _("This is the connection this scheme reads on "
                              "system runs."),
                }
            elif dimmed and primary:
                node['chip'] = {
                    'label': _("Not read"), 'tone': 'muted',
                    'hint': _("A pay run reads only the connection this scheme "
                              "is set to (%(primary)s). Any wire drawn from "
                              "%(other)s is ignored on a system run.",
                              primary=primary.name or _("the primary connection"),
                              other=conn.name or _("this connection")),
                }
            elif dimmed:
                node['chip'] = {
                    'label': _("Not read"), 'tone': 'warn',
                    'hint': _("This scheme has not been told which connection to "
                              "read, so none of these wires is used on a pay "
                              "run. Choose the connection on the scheme itself."),
                }
            systems.append(node)

        if not systems:
            systems.append({
                'id': 'c:none', 'kind': 'connector', 'lane': 'systems',
                'ghost': True, 'label': _("No system connected"),
                'sub': _("Connect an HR system and its fields appear here."),
                'door': {'mode': 'api'},
            })

        # ---- the stored spreadsheet (J2's sample) ---------------------------
        sample = self._import_sample_meta(config)
        if sample:
            systems.append({
                'id': 'file', 'kind': 'file', 'lane': 'systems',
                'label': sample['filename'],
                'sub': (_("read %s") % sample['read_on']) if sample['read_on'] else '',
                'columns': sample['columns'],
                'door': {'mode': 'import'},
            })
        else:
            systems.append({
                'id': 'file', 'kind': 'file', 'lane': 'systems', 'ghost': True,
                'label': _("No file read yet"),
                'sub': _("Drop this month's spreadsheet to see its columns."),
                'door': {'mode': 'import'},
            })

        # ---- Payobook records, both ways (J-D4) -----------------------------
        systems.append({
            'id': 'records', 'kind': 'records', 'lane': 'systems',
            'label': _("Payobook records"),
            'sub': _("Employee · Contract · Bank"),
            'count': emp_total,
            'bank': bank_rows,
            'ghost': not emp_total,
            'door': {'mode': 'employee'},
            'countLabel': (_("%s mapped field") if emp_total == 1
                           else _("%s mapped fields")) % emp_total
                          if emp_total else _("Nothing mapped yet"),
            # RECORDS R3 — a SECOND door on the same node. The node's own door
            # changes which tab you are looking at (`openDoor` switches a MODE
            # and nothing else); this one leaves Mapping altogether for the
            # desk where those mapped fields are edited. It is a separate key
            # rather than a second `door` because the two are not the same kind
            # of gesture, and the board renders it only when the desk's client
            # action is actually registered — a database without `pb_records`
            # must show no button rather than a dead one.
            'actions': [{
                'id': 'records_desk',
                'label': _("Open Records Desk"),
                'icon': 'database',
                'tag': 'pb_records_desk',
                'xmlid': 'pb_records.action_pb_records_desk',
                'params': {'records_config_id': config.id},
            }] if emp_total else [],
        })

        # ============================================ LANE 2 — feeds & files
        #
        # The field counts and the drift verdict come from
        # `get_available_source_fields`, called ONCE PER CONNECTOR and then
        # bucketed by `feed_type` — which is exactly the axis the catalogue
        # itself scopes drift on ("Drift is a claim about a feed that ran, and
        # it may only be made about that feed", `integration_field_mapping.py`
        # :710-716). Per-ENDPOINT calls would be one data-store search per feed
        # for a landing page; per-CONNECTOR is one, and the bucketing loses
        # nothing because a feed's fields are its data type's fields.
        #
        # `expected_missing` is a DERIVED flag, not a column — there is no
        # `hr.integration.endpoint.field.expected_missing` to count, which is
        # why this route rather than the cheaper-looking one.
        cat_by_conn = {}
        for conn in connectors:
            try:
                cat_by_conn[conn.id] = self.env['hr.integration.field.mapping'] \
                    .get_available_source_fields(conn.id, None, None) or []
            except Exception as e:      # noqa: BLE001
                _logger.warning(
                    "J5: source-field discovery failed for connector %s: %s: %s "
                    "— its feeds render without field counts.",
                    conn.id, type(e).__name__, e)
                cat_by_conn[conn.id] = []

        eps = self._api_endpoints(connectors) if connectors else None
        for ep in (eps or []):
            conn_id = ep.connector_id.id
            mine = [f for f in cat_by_conn.get(conn_id, [])
                    if (f.get('feed_type') or '') == (ep.data_type or '')]
            n_fields = len(mine)
            # A feed that has never run cannot be behind: `expected_missing` is
            # a statement about a sync that HAPPENED. Without this guard a brand
            # new integration opens covered in amber, which is the false alarm
            # SOURCING S5 removed from the API board and must not reappear here.
            drift = (len([f for f in mine if f.get('expected_missing')])
                     if ep.last_sync else 0)
            node = {
                'id': 'e:%s' % ep.id, 'kind': 'endpoint', 'lane': 'feeds',
                'parent': 'c:%s' % conn_id,
                'label': ep.name or ep.code or _("Unnamed feed"),
                'sub': '',
                'fields': n_fields,
                'drift': drift,
                'mapped': ep.mapping_count,
                'last_sync': self._journey_iso(ep.last_sync),
                'tone': {'failed': 'err', 'success': 'ok'}.get(
                    ep.last_sync_status or '', 'muted'),
                # A feed is dimmed for whichever reason its CONNECTOR is: either
                # the scheme reads a different connection, or it reads none.
                # Expressed against the connector rather than re-derived, so the
                # two can never disagree on screen (a lit feed under a greyed
                # system is the sort of contradiction a reader stops trusting).
                'dimmed': (bool(primary) and conn_id != primary.id)
                          or (no_primary and bool(wires_by_conn.get(conn_id))),
                'door': {'mode': 'api', 'connector': conn_id, 'endpoint': ep.id,
                         'focus': ep.name or ep.code or ''},
            }
            if drift:
                node['chip'] = {
                    'label': (_("%s not sent") % drift), 'tone': 'warn',
                    'hint': _("The catalogue expects these fields and the last "
                              "sync did not deliver them. They may have been "
                              "renamed at the source."),
                }
            feeds.append(node)
            edges.append({'from': 'c:%s' % conn_id, 'to': 'e:%s' % ep.id,
                          'kind': 'contain', 'count': 0,
                          'dimmed': node['dimmed']})
            live = wires_by_ep.get(ep.id, 0)
            if live:
                edges.append({'from': 'e:%s' % ep.id, 'to': 'scheme',
                              'kind': 'feed', 'count': live,
                              'dimmed': node['dimmed']})

        # ---- one node per sheet of the stored file --------------------------
        cols = self._import_sample_columns(config)
        if cols:
            by_sheet = defaultdict(int)
            for col in cols:
                by_sheet[(col.get('sheet') or '').strip()] += 1
            for sheet, n in sorted(by_sheet.items()):
                sid = 's:%s' % (sheet or '_')
                feeds.append({
                    'id': sid, 'kind': 'sheet', 'lane': 'feeds', 'parent': 'file',
                    'label': sheet or _("The spreadsheet"),
                    'sub': (_("%s column") if n == 1 else _("%s columns")) % n,
                    'columns': n,
                    'door': {'mode': 'import', 'focus': sheet or ''},
                })
                edges.append({'from': 'file', 'to': sid, 'kind': 'contain',
                              'count': 0})
            # An excel BINDING is the live wire between a column and a component.
            bound = len([r for r in config.rule_ids
                         if r.source_binding == 'excel'
                         and (r.source_binding_key or '').strip()])
            if bound:
                first = 's:%s' % ((sorted(by_sheet) or [''])[0] or '_')
                edges.append({'from': first, 'to': 'scheme', 'kind': 'excel',
                              'count': bound})
        elif not feeds:
            feeds.append({
                'id': 'e:none', 'kind': 'endpoint', 'lane': 'feeds', 'ghost': True,
                'label': _("No feeds or files yet"),
                'sub': _("A connected system's feeds, and the sheets of an "
                         "uploaded file, appear here."),
                'door': {'mode': 'import'},
            })

        # ========================================= LANE 3 — transformations
        Rule = self.env.get('hr.api.transformation.rule')
        # MJ16 — `env.get` is None-or-model and an empty recordset is FALSY, so
        # the only correct test is `is None`. `if Rule:` would take the empty
        # branch on every call and this lane would be permanently, silently blank.
        rule_recs = [] if Rule is None else Rule.browse()
        if Rule is not None and connectors:
            try:
                rule_recs = Rule.with_context(active_test=False).search(
                    [('connector_id', 'in', connectors.ids)],
                    order='connector_id, sequence, id')
            except Exception as e:      # noqa: BLE001
                _logger.warning("J5: rule census failed: %s: %s",
                                type(e).__name__, e)
        unread_n = 0
        for r in rule_recs:
            key = (r.output_key or '').strip()
            # J4's predicate, CALLED — not re-implemented. `_tf_consumers`
            # delegates to `pb.integrations._rule_consumers`, which is where the
            # concept was first said out loud. A copy here would pass a grep and
            # disagree with the Transformations tab on the first edge case.
            consumers = self._tf_consumers(r) if key else []
            unread = bool(key) and not consumers
            if unread:
                unread_n += 1
            cid = r.connector_id.id
            try:
                n_reads = len(r._consumed_field_names() or [])
            except Exception:           # noqa: BLE001 — a lane never breaks the tab
                n_reads = 0
            node = {
                'id': 'r:%s' % r.id, 'kind': 'rule', 'lane': 'transforms',
                'parent': 'c:%s' % cid if cid else '',
                'label': r.name or key or _("Untitled rule"),
                'sub': (_("→ %s") % key) if key else '',
                'key': key,
                'reads': n_reads,
                'feeds': len(consumers),
                'active': bool(r.active),
                'tone': 'warn' if unread else '',
                'dimmed': bool(primary) and cid != primary.id,
                'door': {'mode': 'transform', 'connector': cid,
                         'focus': key or (r.name or '')},
            }
            if unread:
                node['chip'] = {
                    'label': _("Unread output"), 'tone': 'warn',
                    'hint': _("This rule computes “%s” and no pay component "
                              "takes it. Wire its output to a component, or the "
                              "work it does is thrown away.") % key,
                }
            transforms.append(node)
            if cid:
                edges.append({'from': 'c:%s' % cid, 'to': 'r:%s' % r.id,
                              'kind': 'contain', 'count': 0,
                              'dimmed': node['dimmed']})
            # a rule -> scheme edge exists only where a wire or a binding does
            fed = 0
            if key and FM is not None:
                try:
                    fed = len([
                        m for m in FM.sudo().with_context(active_test=False).search(
                            [('connector_id', '=', cid),
                             ('source_field', '=', key)])
                        if m.target_rule_id and m.target_rule_id.id in input_ids])
                except Exception:       # noqa: BLE001
                    fed = 0
                fed += len([r2 for r2 in config.rule_ids
                            if r2.source_binding == 'rule'
                            and (r2.source_binding_key or '').strip() == key
                            and r2.id in input_ids])
            if fed:
                edges.append({'from': 'r:%s' % r.id, 'to': 'scheme',
                              'kind': 'rule', 'count': fed,
                              'dimmed': node['dimmed']})

        if not transforms:
            transforms.append({
                'id': 'r:none', 'kind': 'rule', 'lane': 'transforms', 'ghost': True,
                'label': _("No transformation rules yet"),
                'sub': _("A rule turns what a system sends — a list of overtime "
                         "rows, a table of dependants — into one number a pay "
                         "component can read."),
                'door': {'mode': 'transform'},
            })

        # ============================================== LANE 4 — the scheme
        dangling = len([r for r in config.rule_ids if r.binding_dangling])
        health = []
        if no_primary:
            n_inert = sum(wires_by_conn.values())
            health.append({
                'id': 'h:noprimary', 'kind': 'health', 'lane': 'scheme',
                'tone': 'warn',
                'label': (_("%s feed wire is not read") if n_inert == 1
                          else _("%s feed wires are not read")) % n_inert,
                'sub': _("This scheme names no connection, and a pay run only "
                         "reads the one it is set to."),
                'door': {'mode': 'api'},
            })
        if conflicts:
            health.append({
                'id': 'h:conflict', 'kind': 'health', 'lane': 'scheme',
                'tone': 'warn',
                'label': (_("%s component wired twice") if len(conflicts) == 1
                          else _("%s components wired twice")) % len(conflicts),
                'sub': _("A pay run reads one of the two. The other is ignored."),
                'door': {'mode': 'api'},
            })
        if dangling:
            health.append({
                'id': 'h:dangling', 'kind': 'health', 'lane': 'scheme',
                'tone': 'warn',
                'label': (_("%s source no longer exists") if dangling == 1
                          else _("%s sources no longer exist")) % dangling,
                'sub': _("These components name a key nothing currently provides."),
                'door': {'mode': 'api'},
            })
        if severed_n:
            health.append({
                'id': 'h:severed', 'kind': 'health', 'lane': 'scheme',
                'tone': 'err',
                'label': (_("%s severed wire") if severed_n == 1
                          else _("%s severed wires")) % severed_n,
                'sub': _("These wires point at a component that is no longer "
                         "on this scheme."),
                'door': {'mode': 'api'},
            })

        fallback_n = len(read_back_ids)
        scheme_node = {
            'id': 'scheme', 'kind': 'scheme', 'lane': 'scheme',
            'label': config.name or _("This scheme"),
            'sub': (_("%s column") if scheme_counts['total'] == 1
                    else _("%s columns")) % scheme_counts['total'],
            'counts': dict(scheme_counts, fallback=fallback_n),
            'door': {'mode': 'api'},
        }
        if not scheme_counts['total']:
            # A scheme with no columns is the novice's very first screen, and
            # the component bar has nothing to draw. Rendering the real card
            # with an empty bar under it says "this is broken"; the ghost says
            # "this is next", which is the only difference between an empty
            # state and a dead end. Every other lane already does this.
            scheme_node['ghost'] = True
            scheme_node['sub'] = _("No pay components yet — add them on the "
                                   "scheme, then wire them up here.")

        # ============================================== LANE 5 — the pay run
        run = self._journey_run_lane(config, emp_total)

        # ---- records <-> scheme, double-headed (J-D4's language) ------------
        if emp_total:
            edges.append({'from': 'records', 'to': 'scheme', 'kind': 'records',
                          'count': emp_total, 'bidi': True})

        header = {
            'components': scheme_counts['total'],
            'wired': scheme_counts['wired'],
            'fallback': fallback_n,
            'attention': (len(conflicts) + dangling + severed_n
                          + (1 if no_primary else 0)),
        }
        return {
            'ok': True,
            'can_edit': self._can_edit(),
            'config': {'id': config.id, 'name': config.name or '',
                       'code': config.code or '',
                       'country': config.country_code or '',
                       'state': config.state or ''},
            'header': header,
            'primary_id': primary.id if primary else 0,
            'primary_name': (primary.name or '') if primary else '',
            'lanes': {
                'systems': systems, 'feeds': feeds, 'transforms': transforms,
                'scheme': [scheme_node] + health, 'run': run,
            },
            'edges': edges,
            'counts': dict(scheme_counts, fallback=fallback_n,
                           conflicts=len(conflicts), dangling=dangling,
                           severed=severed_n, unread=unread_n,
                           connectors=len(connectors), rules=len(rule_recs)),
        }

    @api.model
    def _journey_run_lane(self, config, emp_total):
        """The last PROCESSED run, summarised from what it STORED.

        Two rules govern every number in here and they are the reason the lane
        is short:

          * it comes from a `done` batch of THIS scheme, and from the provenance
            blobs that batch's payslips carry. Not from a draft, not from a
            batch of another scheme, not from a live recomputation;
          * **"records updated" is not shown, because nothing stores it.**
            `action_process` calls `_update_employee_from_raw_data`,
            `_sync_employee_bank_account`, `_update_contract_from_raw_data` and
            `_sync_contract_components` for every line and counts none of them
            — the batch stores only `created_employee_ids`,
            `created_contract_ids` and `created_payslip_ids`, i.e. records
            CREATED. So the lane says what was created (defensible) and, for the
            two-way rows, the mapping-count note the handover names as the
            fallback. Inventing an "updated" figure from the line count would be
            exactly the fake liveness scope 5 forbids.

        No run yet -> an honest ghost. "No pay run yet" is a different statement
        from "nothing feeds this scheme", and the wiring above the ghost is
        precisely what makes the sentence true.
        """
        Batch = self.env.get('hr.payroll.import.batch')
        batch = None
        if Batch is not None:
            try:
                # RD60 — `create_payslips` keeps a RECORD REFRESH out of this
                # lane. A refresh is a batch too, so without the leaf the node
                # narrated "the last pay run" using a payload that produced no
                # payslip at all — five of them landed on the reference tenant
                # in one afternoon, each one newer than the run it hid.
                batch = Batch.sudo().search(
                    [('formula_config_id', '=', config.id), ('state', '=', 'done'),
                     ('create_payslips', '=', True)],
                    order='id desc', limit=1)
            except Exception as e:      # noqa: BLE001
                _logger.warning("J5: batch lookup failed: %s: %s",
                                type(e).__name__, e)
                batch = None
        if not batch:
            return [{
                'id': 'run', 'kind': 'run', 'lane': 'run', 'ghost': True,
                'label': _("No pay run yet"),
                'sub': _("The wiring on the left is what will happen when one "
                         "is processed."),
                'door': {'mode': 'import'},
            }]

        # TWO read-only queries, and neither of them builds a record.
        #
        # `batch.import_line_ids.mapped('payslip_id')` is the obvious spelling
        # and it is the wrong one here: a processed VN batch is tens of
        # thousands of lines, and that expression instantiates every line AND
        # every payslip to reach one integer and one Text column. On the
        # LANDING TAB of the product. So: count in SQL, then read at most
        # `_JOURNEY_SLIP_CAP` blobs, and let the payload say when it capped.
        n_slips, blobs = 0, []
        try:
            self.env.cr.execute(
                "SELECT count(DISTINCT payslip_id) FROM hr_payroll_import_line "
                "WHERE batch_id = %s AND payslip_id IS NOT NULL", (batch.id,))
            n_slips = self.env.cr.fetchone()[0] or 0
            if n_slips:
                self.env.cr.execute(
                    "SELECT p.formula_input_sources FROM hr_payslip p "
                    "WHERE p.id IN (SELECT DISTINCT payslip_id "
                    "               FROM hr_payroll_import_line "
                    "               WHERE batch_id = %s AND payslip_id IS NOT NULL) "
                    "  AND p.formula_input_sources IS NOT NULL "
                    "LIMIT %s", (batch.id, self._JOURNEY_SLIP_CAP))
                blobs = [row[0] for row in self.env.cr.fetchall()]
        except Exception as e:          # noqa: BLE001
            _logger.warning("J5: run aggregate failed for batch %s: %s: %s",
                            batch.id, type(e).__name__, e)
        capped = n_slips > self._JOURNEY_SLIP_CAP
        agg = self._journey_aggregate(blobs)

        node = {
            'id': 'run', 'kind': 'run', 'lane': 'run',
            'label': batch.name or _("Last pay run"),
            'sub': '',
            'payslips': n_slips,
            'read': len(blobs),
            'capped': capped,
            'lines': batch.total_lines,
            'matched': batch.matched_employees,
            'errors': batch.error_lines,
            'created': {
                'employees': len(batch.created_employee_ids),
                'contracts': len(batch.created_contract_ids),
                'payslips': len(batch.created_payslip_ids),
            },
            'agg': agg,
            'date': self._journey_iso(batch.create_date),
            'batch_id': batch.id,
            'door': {'mode': 'import'},
            # The two-way note, shown INSTEAD of an updated-records count that
            # this database does not store (see the docstring).
            'records_note': (
                _("%s mapped field writes an employee or contract record on "
                  "import.") if emp_total == 1
                else _("%s mapped fields write employee and contract records "
                       "on import.")) % emp_total if emp_total else '',
        }
        return [node]

    # ------------------------------------------------------------------
    # W62 — transforms on the wire (surface + edit + live-preview the transforms
    # that ALREADY run at sync time). API adapter ONLY — cycle wires carry no
    # transform (D-I1: live payruns bypass cycle-mapping records, so a cycle
    # transform would apply to imports but not to live runs — a C7 trap).
    # ------------------------------------------------------------------
    @staticmethod
    def _transform_payload(m):
        """Compact transform descriptor for an accepted API wire (D-I2). The badge
        glyph is rendered client-side from type/value/decimals so the popover's live
        preview updates without a round-trip."""
        return {
            'type': m.transformation_type or 'direct',
            'value': m.transformation_value or 0.0,
            'decimals': m.transformation_decimals if m.transformation_decimals is not None else 2,
            'python': m.transformation_type == 'python',
            'error': bool(m.has_transform_error),
            'error_msg': m.transform_error_msg or '',
            'sample': m.source_sample_value or '',
        }

    @api.model
    def api_transform_preview(self, mapping_id, draft_vals):
        """Evaluate a DRAFT transform against the mapping's sample value WITHOUT
        writing (D-I3). preview == what the sync path produces — they are the same
        engine function. Reads are open; the preview never mutates."""
        m = self.env['hr.integration.field.mapping'].browse(int(mapping_id or 0)).exists()
        if not m:
            return {'ok': False, 'error': _("Mapping not found.")}
        return m.preview_transform(draft_vals or {})

    @api.model
    def api_transform_save(self, mapping_id, vals):
        """Persist a transform edit (D-I3). Manager-gated; whitelisted to
        type/value/decimals ONLY — `transformation_code` is NEVER writable here (the
        canvas must not grow a code-authoring surface, D-I2/D-I3). Returns the fresh
        transform payload so the badge re-renders."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can edit transforms.")}
        m = self.env['hr.integration.field.mapping'].browse(int(mapping_id or 0)).exists()
        if not m:
            return {'ok': False, 'msg': _("Mapping not found.")}
        vals = dict(vals or {})
        t = vals.get('transformation_type') or 'direct'
        allowed = {'direct', 'multiply', 'divide', 'add', 'subtract',
                   'round', 'abs', 'default_if_empty'}
        if t not in allowed:
            # python (and anything unknown) is not editable on the canvas
            return {'ok': False, 'msg': _("This transform type can only be edited in "
                                          "the backend form.")}
        data = {'transformation_type': t}
        if 'transformation_value' in vals:
            try:
                data['transformation_value'] = float(vals.get('transformation_value') or 0.0)
            except (TypeError, ValueError):
                return {'ok': False, 'msg': _("Factor / value must be a number.")}
        if 'transformation_decimals' in vals:
            try:
                data['transformation_decimals'] = int(vals.get('transformation_decimals') or 0)
            except (TypeError, ValueError):
                return {'ok': False, 'msg': _("Decimals must be a whole number.")}
        # switching AWAY from python (or off an errored op) clears the stale error flag
        if m.has_transform_error:
            data['has_transform_error'] = False
            data['transform_error_msg'] = False
        m.write(data)
        return {'ok': True, 'transform': self._transform_payload(m)}

    # ------------------------------------------------------------------
    # W65 — mapping templates (save a board as a named, reusable template and
    # apply it across configs/connectors — the bureau workflow). New lean
    # user-template models (D-I5); the vendor-seeded hr.integration.mapping.template
    # is untouched. Templates store CODES/PATHS, never ids, so they apply across
    # configs. Company-scoped from day one — no W104 snippet gap.
    # ------------------------------------------------------------------
    def _tmpl_can_delete(self, tpl):
        """Managers can delete shared templates and their own company's; a
        non-shared template from ANOTHER company is un-deletable (server-side)."""
        if not self._can_edit():
            return False
        return (not tpl.company_id) or tpl.company_id.id == self.env.company.id

    @api.model
    def mapping_template_list(self, adapter=None):
        """Visible templates = shared (no company) + this company's. Reads are open."""
        Tpl = self.env['hr.formula.mapping.template']
        domain = ['|', ('company_id', '=', False), ('company_id', '=', self.env.company.id)]
        if adapter in ('api', 'cycle'):
            domain = [('adapter', '=', adapter)] + domain
        out = []
        for t in Tpl.search(domain):
            out.append({'id': t.id, 'name': t.name or '', 'adapter': t.adapter,
                        'connector_type': t.connector_type or '',
                        'shared': not t.company_id,
                        'line_count': len(t.line_ids),
                        'can_delete': self._tmpl_can_delete(t)})
        return {'ok': True, 'templates': out, 'can_edit': self._can_edit()}

    @api.model
    def mapping_template_save(self, config_id, adapter, name):
        """Snapshot the CURRENT accepted wires of a board into a named template
        (D-I6). API boards carry transforms; cycle boards carry pairs only (D-I1).
        Manager-gated; always company-scoped to self.env.company."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can save templates.")}
        name = (name or '').strip()
        if not name:
            return {'ok': False, 'msg': _("Give the template a name.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("No configuration.")}
        lines, connector_type = [], False
        if adapter == 'api':
            conn = self._api_active_connector(config)
            if not conn:
                return {'ok': False, 'msg': _("No connector to snapshot.")}
            connector_type = conn.connector_type or False
            input_ids = config.rule_ids.filtered(lambda r: r.column_type == 'input').ids
            FM = self.env['hr.integration.field.mapping']
            for m in FM.search([('connector_id', '=', conn.id),
                                ('target_rule_id', 'in', input_ids)]):
                if not (m.source_field and m.target_rule_id.code):
                    continue
                lines.append((0, 0, {
                    'source_key': m.source_field,
                    'target_code': m.target_rule_id.code,
                    'transformation_type': m.transformation_type or 'direct',
                    'transformation_value': m.transformation_value or 0.0,
                    'transformation_decimals': m.transformation_decimals
                        if m.transformation_decimals is not None else 2,
                    'sequence': m.sequence or 10,
                }))
        elif adapter == 'cycle':
            mid, end = self._cycle_pair(config)
            if not (mid and end):
                return {'ok': False, 'msg': _("No paired cycle configuration to snapshot.")}
            Mapping = self.env['hr.payroll.cycle.component.mapping']
            for m in Mapping.search([('mid_cycle_config_id', '=', mid.id),
                                    ('end_cycle_config_id', '=', end.id)]):
                if not (m.mid_component_id.code and m.end_component_id.code):
                    continue
                lines.append((0, 0, {'source_key': m.mid_component_id.code,
                                     'target_code': m.end_component_id.code}))
        else:
            return {'ok': False, 'msg': _("Unknown adapter.")}
        if not lines:
            return {'ok': False, 'msg': _("Nothing mapped to save yet.")}
        tpl = self.env['hr.formula.mapping.template'].create({
            'name': name, 'adapter': adapter, 'connector_type': connector_type,
            'company_id': self.env.company.id, 'line_ids': lines,
        })
        return {'ok': True, 'template_id': tpl.id, 'line_count': len(lines)}

    @api.model
    def mapping_template_apply(self, template_id, config_id, connector_id=None):
        """Apply a template to a board by matching lines on code/path (D-I6). NEVER
        overwrites an existing wire (skip + report) and never deletes anything.
        Returns {applied, skipped_existing, unmatched_sources, unmatched_targets}."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can apply templates.")}
        tpl = self.env['hr.formula.mapping.template'].browse(int(template_id or 0)).exists()
        if not tpl:
            return {'ok': False, 'msg': _("Template not found.")}
        # visibility guard — server-side, not just UI (D-I5)
        if tpl.company_id and tpl.company_id.id != self.env.company.id:
            return {'ok': False, 'msg': _("This template belongs to another company.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("No configuration.")}
        applied, skipped, un_src, un_tgt = [], [], [], []
        if tpl.adapter == 'api':
            conn = self._api_active_connector(config, connector_id)
            if not conn:
                return {'ok': False, 'msg': _("No connector on this board.")}
            FM = self.env['hr.integration.field.mapping']
            input_rules = config.rule_ids.filtered(lambda r: r.column_type == 'input')
            code_to_rule = {r.code: r for r in input_rules if r.code}
            try:
                avail = {f['path'] for f in (FM.get_available_source_fields(conn.id) or [])}
            except Exception:
                avail = set()
            existing = FM.search([('connector_id', '=', conn.id),
                                  ('target_rule_id', 'in', input_rules.ids)])
            used_src = {m.source_field for m in existing}
            used_tgt = {m.target_rule_id.id for m in existing}
            for ln in tpl.line_ids:
                rule = code_to_rule.get(ln.target_code)
                src_ok = ln.source_key in avail
                if not src_ok:
                    un_src.append(ln.source_key)
                if not rule:
                    un_tgt.append(ln.target_code)
                if not (rule and src_ok):
                    continue
                if ln.source_key in used_src or rule.id in used_tgt:
                    skipped.append({'source': ln.source_key, 'target': ln.target_code})
                    continue
                FM.create({'connector_id': conn.id, 'source_field': ln.source_key,
                           'target_rule_id': rule.id,
                           'source_field_label': (ln.source_key or '').replace('_', ' ').title(),
                           'transformation_type': ln.transformation_type or 'direct',
                           'transformation_value': ln.transformation_value or 0.0,
                           'transformation_decimals': ln.transformation_decimals
                               if ln.transformation_decimals is not None else 2})
                used_src.add(ln.source_key)
                used_tgt.add(rule.id)
                applied.append({'source': ln.source_key, 'target': ln.target_code})
        elif tpl.adapter == 'cycle':
            mid, end = self._cycle_pair(config)
            if not (mid and end):
                return {'ok': False, 'msg': _("This configuration has no paired cycle to apply to.")}
            Mapping = self.env['hr.payroll.cycle.component.mapping']
            mid_by_code = {r.code: r for r in mid.rule_ids if r.code}
            end_by_code = {r.code: r for r in end.rule_ids if r.code}
            existing = Mapping.search([('mid_cycle_config_id', '=', mid.id),
                                      ('end_cycle_config_id', '=', end.id)])
            used_mid = {m.mid_component_id.id for m in existing}
            used_end = {m.end_component_id.id for m in existing}
            for ln in tpl.line_ids:
                midc = mid_by_code.get(ln.source_key)
                endc = end_by_code.get(ln.target_code)
                if not midc:
                    un_src.append(ln.source_key)
                if not endc:
                    un_tgt.append(ln.target_code)
                if not (midc and endc):
                    continue
                if midc.id in used_mid or endc.id in used_end:
                    skipped.append({'source': ln.source_key, 'target': ln.target_code})
                    continue
                Mapping.create({'mid_cycle_config_id': mid.id, 'end_cycle_config_id': end.id,
                                'mid_component_id': midc.id, 'end_component_id': endc.id})
                used_mid.add(midc.id)
                used_end.add(endc.id)
                applied.append({'source': ln.source_key, 'target': ln.target_code})
        else:
            return {'ok': False, 'msg': _("Unknown template adapter.")}
        return {'ok': True, 'applied': applied, 'skipped_existing': skipped,
                'unmatched_sources': sorted(set(un_src)),
                'unmatched_targets': sorted(set(un_tgt))}

    @api.model
    def mapping_template_delete(self, template_id):
        """Manager-gated + company-scope server-side check (D-I5): a non-shared
        template from another company is un-deletable here."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can delete templates.")}
        tpl = self.env['hr.formula.mapping.template'].browse(int(template_id or 0)).exists()
        if not tpl:
            return {'ok': True}
        if not self._tmpl_can_delete(tpl):
            return {'ok': False, 'msg': _("This template belongs to another company "
                                          "and can't be deleted here.")}
        tpl.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # F10 adapter 3 — import column mapping (Excel columns → inputs)
    # ------------------------------------------------------------------
    def _import_batch_columns(self, batch):
        """Distinct column keys from the batch's first import line (the parsed
        header→value dict), preserving order."""
        line = self.env['hr.payroll.import.line'].search([('batch_id', '=', batch.id)], limit=1)
        if line and line.raw_data_json:
            try:
                return list(json.loads(line.raw_data_json).keys())
            except Exception:
                pass
        return []

    # ------------------------------------------------------------------
    # JOURNEY J2 — the Excel on-ramp.
    #
    # Everything between here and `import_mapping_data` exists to answer one
    # question the Spreadsheet board could not: *what are my file's columns?*
    # Before J2 the only answers were "the keys of a batch somebody already
    # imported" and "whatever you type", which is why S12 found the board had
    # never written a value in production — you had to have finished the import
    # before you could set up the mapping that the import needs.
    #
    # The reader below runs the LOADER'S OWN parse (`peek_source_columns` on
    # `hr.payroll.import.batch`) over an in-memory probe record. No batch row,
    # no import line, no pay value. What comes back is stored on the scheme so
    # the answer survives a reload, and the same stored file is what the
    # "load this file as a pay run" button hands to the guided flow.
    # ------------------------------------------------------------------
    _IMPORT_SAMPLE_MAX_COLS = 800

    @api.model
    def _import_sample_columns(self, config):
        """The columns read off this scheme's stored file, or `[]`."""
        if not config or not config.import_sample_columns_json:
            return []
        try:
            cols = json.loads(config.import_sample_columns_json)
        except Exception:
            return []
        return cols if isinstance(cols, list) else []

    @api.model
    def _import_sample_meta(self, config):
        """What the board says about where its columns came from.

        `read_on` is formatted server-side because the client would otherwise
        have to guess the user's timezone and language for a provenance line —
        and a provenance line that is wrong about WHEN is worse than none.
        """
        cols = self._import_sample_columns(config)
        if not cols:
            return None
        when = ''
        if config.import_sample_date:
            try:
                from odoo.tools.misc import format_datetime
                when = format_datetime(self.env, config.import_sample_date,
                                       dt_format='short')
            except Exception:
                when = fields.Datetime.to_string(config.import_sample_date)
        name = config.import_sample_filename or _("a spreadsheet")
        return {
            'filename': name,
            'read_on': when,
            'columns': len(cols),
            'shown': len([c for c in cols if c.get('preferred')]),
            'has_file': bool(config.import_sample_file),
            'line': (_("%(file)s · read %(when)s", file=name, when=when) if when
                     else name),
        }

    @api.model
    def import_mapping_read_headers(self, config_id, file_b64, filename):
        """Read a spreadsheet's COLUMN HEADINGS onto the board. Not its data.

        The one gesture the Spreadsheet board never had. It stores the file, the
        columns and the moment it read them on the scheme, and returns the
        refreshed board — so the left column fills in the same frame the drop
        finishes in.

        What it deliberately does NOT do: create an import batch, create an
        import line, match an employee, or write a single pay value. Test 2
        proves that with a row-count diff, because a promise about what
        something does not do is only worth the check behind it.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can read a file onto this board.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Pick a payroll scheme first.")}
        name = (filename or '').strip() or 'pay-data.xlsx'
        if not name.lower().endswith(('.xlsx', '.xls', '.csv')):
            return {'ok': False,
                    'msg': _("That is not a spreadsheet. Drop an .xlsx, .xls or .csv file.")}
        try:
            content = base64.b64decode(file_b64 or '')
        except (binascii.Error, ValueError):
            return {'ok': False, 'msg': _("That file could not be read.")}
        if not content:
            return {'ok': False, 'msg': _("That file is empty.")}
        try:
            cols = self.env['hr.payroll.import.batch'].peek_source_columns(
                config, content, name)
        except Exception as e:
            _logger.warning("J2 header read failed for %s: %s", name, e)
            return {'ok': False,
                    'msg': _("The headings could not be read from %(file)s. "
                             "Check it opens in a spreadsheet and that the first "
                             "row is the column headings.", file=name)}
        if not cols:
            return {'ok': False,
                    'msg': _("No column headings were found in %(file)s.", file=name)}
        truncated = len(cols) > self._IMPORT_SAMPLE_MAX_COLS
        cols = cols[:self._IMPORT_SAMPLE_MAX_COLS]
        config.sudo().write({
            'import_sample_file': base64.b64encode(content),
            'import_sample_filename': name,
            'import_sample_date': fields.Datetime.now(),
            'import_sample_columns_json': json.dumps(cols),
        })
        data = self.import_mapping_data(config.id, False)
        data['read'] = {
            'columns': len(cols),
            'shown': len([c for c in cols if c.get('preferred')]),
            'truncated': truncated,
        }
        return data

    @api.model
    def import_mapping_forget_headers(self, config_id):
        """Drop the stored file and its columns. Wires are untouched.

        A binding is a decision about a column NAME; it does not depend on the
        file that suggested the name, and deleting the file must not silently
        unwire a scheme. So this forgets the spreadsheet and leaves every
        `source_binding` exactly where it was — the columns come back in the
        "Already used by this scheme" lane, which is where they belong once
        they have been chosen.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can change this board.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Pick a payroll scheme first.")}
        config.sudo().write({
            'import_sample_file': False,
            'import_sample_filename': False,
            'import_sample_date': False,
            'import_sample_columns_json': False,
        })
        return self.import_mapping_data(config.id, False)

    @api.model
    def import_mapping_template(self, config_id):
        """Build the workbook whose headings this scheme will match on re-import.

        `ExcelConnector.generate_template` had zero callers for its entire life
        (S12/J2). This is its caller. The point of a scheme-built template is
        the round trip: fill it in, drop it back on this board, and every input
        component finds its column — because the headings were derived from
        what the resolver looks for, not from a label somebody typed once.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Pick a payroll scheme first.")}
        try:
            content, filename = config._build_pay_data_template()
        except ImportError:
            return {'ok': False,
                    'msg': _("Spreadsheets cannot be built on this server.")}
        except Exception as e:
            _logger.warning("J2 template build failed for config %s: %s", config.id, e)
            return {'ok': False, 'msg': _("That template could not be built.")}
        inputs = config.rule_ids.filtered(lambda r: r.column_type == 'input')
        return {
            'ok': True,
            'file_b64': base64.b64encode(content).decode(),
            'filename': filename,
            'mimetype': ('application/vnd.openxmlformats-officedocument'
                         '.spreadsheetml.sheet'),
            'columns': len(inputs),
        }

    @api.model
    def import_mapping_handoff(self, config_id, file_b64=None, filename=None):
        """"Load this file as a pay run…" — into the flow that already exists.

        This builds NO import pipeline. It calls `pb.import.wizard.create_and_load`,
        the same server method the guided wizard's first step calls, which
        creates the batch, loads the file and matches employees — and stops
        there. Validating and committing stay where they are: on the batch, in
        front of a person, one deliberate click each. A mapping board must
        never be able to pay somebody.

        A fresh upload is read for its headings on the way past, so one gesture
        updates the board AND starts the run.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("Only managers can load pay data.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Pick a payroll scheme first.")}
        Wizard = self.env.get('pb.import.wizard')
        if Wizard is None:
            return {'ok': False,
                    'msg': _("The guided load is not available on this database.")}
        name = (filename or '').strip()
        if file_b64:
            read = self.import_mapping_read_headers(config.id, file_b64, name)
            if not read.get('ok'):
                return read
            payload, fname = file_b64, config.import_sample_filename
        elif config.import_sample_file:
            payload = config.import_sample_file
            if isinstance(payload, bytes):
                payload = payload.decode()
            fname = config.import_sample_filename or 'pay-data.xlsx'
        else:
            return {'ok': False,
                    'msg': _("Drop the file with this period's numbers in it first.")}
        summary = Wizard.create_and_load({
            'name': _("%(scheme)s — %(file)s", scheme=config.name, file=fname),
            'source_type': 'excel',
            'formula_config_id': config.id,
            'file_b64': payload,
            'file_name': fname,
        })
        if summary.get('error'):
            return {'ok': False, 'msg': summary['error'],
                    'batch_id': summary.get('batch_id')}
        return {'ok': True, 'batch_id': summary.get('batch_id'),
                'total_lines': summary.get('total_lines') or 0,
                'matched': summary.get('matched') or 0,
                'new': summary.get('new') or 0,
                'name': summary.get('name') or fname}

    @api.model
    def import_mapping_data(self, config_id=None, batch_id=None):
        """The Excel board — which, until S6, could not be opened at all.

        It answered `{'ok': False, 'reason': 'no_batch'}` on any database with no
        import batch (abm has none), so the only spreadsheet-mapping surface in the
        product was a dead end from the first click. And its left column was "the
        keys of one line of one load", which meant the board was only ever usable in
        the minutes after somebody happened to upload a file — which is why, per
        **S12**, it had never written a value on any of the four databases.

        **JOURNEY J2 closed that.** The left column now has FOUR lanes, and the
        first of them is the one that makes the board usable before an import
        rather than after it:

          * **the columns of a file dropped on this board** — read for its
            headings, never its data, through the loader's own parser, so a
            column shown here is a key the loader will produce;
          * the selected batch's columns (the original behaviour, kept);
          * **the keys this scheme is already bound to** — so last month's wires are
            on screen with no file loaded at all;
          * legacy `data_source_field` values, for schemes that predate bindings.

        And when the column you want is in none of them, the left column's search
        box offers to take it as typed (`can_add`) — a header or a column letter.
        A board that requires an upload before it will show you anything is a board
        nobody uses.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        Batch = self.env['hr.payroll.import.batch']
        batches = Batch.search([], order='id desc')
        contexts = [{'id': b.id, 'name': b.name} for b in batches]
        batch = Batch.browse(int(batch_id)) if batch_id else Batch.browse()
        if not batch:
            # RD60 — prefer a batch that fed a pay run. A record refresh is a
            # batch, and the newest one is usually a refresh, so the board's
            # opening view drifted onto the connected system's columns instead
            # of the columns the last run was actually mapped from. Still only a
            # PREFERENCE: with nothing but refreshes on file, showing their
            # columns beats showing an empty board.
            pay_data = batches.filtered(lambda b: b.create_payslips)
            batch = (pay_data.filtered(lambda b: b.formula_config_id.id == config.id and b.import_line_ids)[:1]
                     or batches.filtered(lambda b: b.formula_config_id.id == config.id and b.import_line_ids)[:1]
                     or pay_data.filtered(lambda b: b.import_line_ids)[:1]
                     or batches.filtered(lambda b: b.import_line_ids)[:1] or batches[:1])
        input_rules = config.rule_ids.filtered(lambda r: r.column_type == 'input') \
            .sorted(key=lambda r: r.sequence)
        cols = self._import_batch_columns(batch) if batch else []
        left = self._import_left_columns(batch, cols, input_rules, config=config)
        _acts, _run = self._source_actuals(config)
        _emp = self._source_record_dests(config)
        _wires = self._source_wire_dests(config)
        right = self._mc_right_column(
            config, _acts, _emp, wire_dests=_wires,
            lineage=self._lineage_for_config(config), board='import')
        wires, mapped_rules = [], set()
        for r in input_rules:
            # The BINDING is the wire. `data_source_field` still draws one for a
            # scheme that predates bindings, so no existing configuration loses its
            # picture — but it is the fallback, not the record.
            key = ((r.source_binding_key or '').strip()
                   if r.source_binding == 'excel' else '') or r.data_source_field
            if key:
                wires.append({'id': 'im%s' % r.id, 'kind': 'mapping', 'ref': r.id,
                              'leftId': 'c:' + key, 'rightId': r.id, 'state': 'accepted'})
                mapped_rules.add(r.id)
        # suggestions: best name/code match between an unmapped column and input
        rule_norms = [(r, self._norm(r.code), self._norm(r.name)) for r in input_rules
                      if r.id not in mapped_rules]
        used = set(mapped_rules)
        for c in cols:
            cn = self._norm(c)
            best, conf = None, 0.0
            for r, rc, rn in rule_norms:
                if r.id in used:
                    continue
                if rc and (cn == rc):
                    x = 1.0
                elif rc and (rc in cn or cn in rc):
                    x = 0.85
                elif rn and (cn == rn or rn in cn or cn in rn):
                    x = 0.8
                else:
                    x = 0.0
                if x > conf:
                    best, conf = r, x
            if best and conf >= 0.8:
                wires.append({'id': 'sug:%s>%s' % (c, best.id), 'kind': 'suggestion',
                              'ref': None, 'source': c, 'leftId': 'c:' + c, 'rightId': best.id,
                              'state': 'suggested', 'confidence': round(conf, 2), 'reason': _('Name match')})
                used.add(best.id)
        return {
            'ok': True, 'left': left, 'right': right, 'wires': wires,
            'left_title': ('%s · columns' % batch.name) if batch
                          else _("Spreadsheet columns"),
            'right_title': '%s · inputs' % config.name,
            'subtitle': (_("Map imported columns from %s onto this scheme's inputs")
                         % batch.name) if batch else _(
                "Say which spreadsheet column feeds each component. No file needs "
                "to be loaded — type the column heading and connect it."),
            'supports_suggest': False,
            # The search box may take a column as typed. It is the answer to "the
            # column I want is not in this list", which on a database with no
            # upload is every column.
            'can_add': True,
            'add_label': _("Use “%s” as a spreadsheet column"),
            'contexts': contexts, 'context_id': batch.id if batch else False,
            'can_edit': self._can_edit(),
            # J2 — the on-ramp's own state. `sample` is null until a file has
            # been read onto this scheme; `inputs` is what the template and the
            # coverage line are counted against.
            'sample': self._import_sample_meta(config),
            'inputs': len(input_rules),
            'wired': len(mapped_rules),
            'config_id': config.id,
        }

    @api.model
    def _import_left_columns(self, batch, cols, input_rules, config=None):
        """The Excel board's left column: four lanes, none of them invented.

        Order is deliberate — the file YOU just dropped first (J2), because
        somebody who has just handed the board a spreadsheet is looking for its
        columns and nothing else; then a loaded batch's columns; then what this
        scheme already reads, which is the only lane a never-uploaded database
        has; then the legacy Char.

        The dropped-file lane shows ONE card per real column, labelled with the
        key that will be bound and carrying a sample value from the first row —
        `e.g. 12,500,000` under `SEVL|Basic Salary` is the difference between
        recognising your column and hoping. The other spellings of the same
        column (its bare twin, its letter) are real keys and stay in the stored
        list, reachable by typing them; they do not each get a card, because
        four cards for one column is a board nobody can read (see
        `peek_source_columns`' `preferred`).
        """
        out, seen = [], set()

        def add(key, group, sublabel='', meta=None):
            key = (key or '').strip()
            if not key or key in seen:
                return
            seen.add(key)
            out.append({'id': 'c:' + key, 'label': key, 'sublabel': sublabel,
                        'group': group, 'meta': meta or {}})

        meta = self._import_sample_meta(config) if config else None
        if meta:
            lane = meta['line']
            for col in self._import_sample_columns(config):
                if not col.get('preferred'):
                    continue
                sample = col.get('sample') or ''
                add(col.get('key'), lane,
                    sublabel=(_("e.g. %s", sample) if sample else _("no value in the first row")),
                    meta={'sheet': col.get('sheet') or '',
                          'letter': col.get('letter') or ''})

        file_lane = (batch.name or _("This file")) if batch else _("Uploaded file")
        for c in cols:
            add(c, file_lane)
        for r in input_rules:
            if r.source_binding == 'excel':
                add(r.source_binding_key, _("Already used by this scheme"))
        for r in input_rules:
            add(r.data_source_field, _("From this scheme's history"))
        return out

    @api.model
    def import_mapping_create(self, config_id, batch_id, column, target_rule_id,
                              resolve=None):
        """Bind a component to a spreadsheet column — for real, this time.

        JOURNEY J3 S2 — `resolve` is the conflict dialog's answer (see
        `api_mapping_create` for the vocabulary). `'replace'` unlinks the live wires
        that would otherwise beat this binding on every system run; `'keep'` and
        `None` leave them, which is today's behaviour and is now honest rather than
        silent, because the component wears a conflict chip and the resolver's
        empty-feed guard lets this column actually speak when the feed does not.

        This wrote `rule.write({'data_source_field': col})` and nothing else, which
        is why **S12** found that Char empty on every rule on all four databases:
        one overloaded string that carried spreadsheet headers, feed keys and column
        letters indiscriminately, with nothing to say which. S3 built the honest
        answer — `source_binding` + `source_binding_key`, consulted by the resolver
        in front of the name-matching ladder — and nothing wrote one.

        Now this does, and `data_source_field` is deliberately NOT written beside
        it. Two statements of one fact drift apart; the binding is the one the
        resolver reads first and the one every chip renders, so it is the one that
        gets written.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        # MAPFIX D1 — wrong-type in, refusal out (see `employee_mapping_create`).
        col = self._ec_spec(column)
        col = col[2:] if col.startswith('c:') else col
        col = (col or '').strip()
        rule = self.env['hr.formula.rule'].browse(self._as_id(target_rule_id))
        if not (col and rule.exists()):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        sealed = self._mc_refuse_sealed(rule)
        if sealed:
            return sealed
        # SC-4 — the excel lane's refusal on the board itself.
        if rule.config_id and not getattr(
                rule.config_id, 'source_excel_enabled', True):
            return {'ok': False, 'msg': _(
                "This scheme does not take pay data files — its sources "
                "settings switched spreadsheets off. Turn the lane back on "
                "in the scheme's settings to bind this column.")}
        replaced = self._binding_replaced(rule, 'excel', col)
        if resolve == 'replace':
            FM = self.env.get('hr.integration.field.mapping')
            if FM is not None:
                FM.sudo().search([('target_rule_id', '=', rule.id)]).unlink()
            # JOURNEY J9 — "only the spreadsheet feeds this after that" has to be
            # true of the DECLARATIONS as well as of the wires, now that removing
            # a wire no longer takes its declaration with it by default.
            rule.clear_source_binding('feed')
            rule.clear_source_binding('rule')
        rule.set_source_binding('excel', col, origin='board')
        return {'ok': True, 'replaced': replaced}

    @api.model
    def import_mapping_delete(self, rule_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        rule = self.env['hr.formula.rule'].browse(self._as_id(rule_id))
        if rule.exists():
            # Clear whichever of the two drew the wire. A component wired from the
            # legacy Char and one bound properly both have to come off the board
            # when the user removes them, or "delete" is a lie on half the rows.
            #
            # JOURNEY J9 — asked of the SOURCE ROWS, not of `source_binding`. A
            # component that reads a feed AND a spreadsheet column computes
            # `source_binding = 'feed'`, so the old test would have found no
            # spreadsheet binding to clear and this delete would have removed the
            # wire from the board while leaving the component still reading it.
            rule.clear_source_binding('excel')
            if rule.data_source_field:
                rule.write({'data_source_field': False})
        return {'ok': True}

    # ------------------------------------------------------------------
    # F10 adapter 4 — Employee → scheme assignment (departments → schemes)
    # ------------------------------------------------------------------
    @api.model
    def scheme_mapping_data(self, config_id=None, context_id=None):
        config = self._pick_config(config_id)
        Emp = self.env['hr.employee']
        Dept = self.env['hr.department']
        Config = self.env['hr.formula.config']
        Assign = self.env['hr.formula.scheme.assignment']
        # LEFT = departments that actually have employees (with coverage counts)
        counts = {}
        for d in Dept.search([]):
            n = Emp.search_count([('department_id', '=', d.id)])
            if n:
                counts[d.id] = n
        depts = Dept.browse(sorted(counts, key=lambda i: -counts[i]))
        left = [{'id': d.id, 'label': d.name or '(dept)',
                 'sublabel': '%s employees' % '{:,}'.format(counts[d.id]),
                 'meta': {'count': counts[d.id]}} for d in depts]
        # RIGHT = the primary payroll schemes (active, not the mid-cycle advance)
        schemes = Config.search([('state', '=', 'active'),
                                 ('cycle_type', '!=', 'mid_cycle')], order='name')
        scheme_ids = set(schemes.ids)
        assigns = Assign.search([('config_id', 'in', schemes.ids)])
        cov = defaultdict(int)
        wires = []
        for a in assigns:
            if a.department_id and a.config_id.id in scheme_ids:
                wires.append({'id': 'sa%s' % a.id, 'kind': 'mapping', 'ref': a.id,
                              'leftId': a.department_id.id, 'rightId': a.config_id.id,
                              'state': 'accepted'})
                cov[a.config_id.id] += counts.get(a.department_id.id, 0)
        right = [{'id': c.id, 'label': c.name,
                  'sublabel': (('%s covered' % '{:,}'.format(cov[c.id])) if cov[c.id]
                               else (c.country_code or 'scheme')),
                  'meta': {'coverage': cov[c.id]}} for c in schemes]
        return {
            'ok': True, 'left': left, 'right': right, 'wires': wires,
            'left_title': 'Employee segments (departments)',
            'right_title': 'Payroll schemes',
            'subtitle': _("Assign employee segments to the payroll scheme that pays them"),
            'supports_suggest': False,
            'contexts': [], 'context_id': False,
            'can_edit': self._can_edit(),
        }

    @api.model
    def scheme_mapping_create(self, config_id, context_id, department_id, target_config_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        Assign = self.env['hr.formula.scheme.assignment']
        # MAPFIX D1 — both ends are ids from the browser; a mix-up is a refusal.
        dept = self.env['hr.department'].browse(self._as_id(department_id))
        cfg = self.env['hr.formula.config'].browse(self._as_id(target_config_id))
        if not (dept.exists() and cfg.exists()):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        # one scheme per department: drop this department's other assignments
        Assign.search([('department_id', '=', dept.id)]).unlink()
        Assign.create({'department_id': dept.id, 'config_id': cfg.id})
        return {'ok': True}

    @api.model
    def scheme_mapping_delete(self, assignment_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        a = self.env['hr.formula.scheme.assignment'].browse(self._as_id(assignment_id))
        if a.exists():
            a.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # Employee/Contract field mapping adapter — folds the old standalone
    # "Employee/Contract Mapping" list into the canvas. LEFT = the config's
    # components; RIGHT = a curated set of writable, scalar employee/contract
    # fields (+ on-demand search); wires persist to hr.payslip.import.mapping.
    # ------------------------------------------------------------------
    # MAPFIX B1 — `many2one` joins the set. The batch has always supported it:
    # `_coerce_mapped_value` does search-by-name-else-create and
    # `_sync_employee_contract_mirror_fields` mirrors job/department/calendar/company
    # across the employee and the contract. Only the UI could not express it, which
    # is why "Department" had to be typed into a contract by hand after every import.
    # one2many/many2many stay out — the mapping model's own field domain excludes
    # them, and a spreadsheet cell is not a set of records.
    _EC_TTYPES = ('char', 'text', 'float', 'monetary', 'integer', 'boolean',
                  'date', 'datetime', 'selection', 'many2one')
    _EC_MODEL_LABEL = {'hr.contract': 'Contract', 'hr.employee': 'Employee'}

    # ------------------------------------------------------------------
    # MAPFIX B1 — the field CATALOGUE.
    #
    # `_EC_CURATED` used to be 22 hand-typed field names, and every column that was
    # not one of them had to be searched for by hand or given up on — `name` was not
    # even in the list. The catalogue is now GENERATED from the registry: every
    # stored, writable, scalar field of hr.employee and hr.contract that is not
    # technical plumbing. Curation did not disappear, it moved: it now decides what
    # you see FIRST (the lanes below) rather than what you are allowed to map.
    #
    # DENY-LIST — why each group is here:
    #   * ORM stamps (`create_uid` … `__last_update`) and `id`/`display_name`:
    #     written by the framework, never by a person;
    #   * `message_*`, `activity_*`, `access_*`, `rating_*`, `website_message_*`:
    #     owned by the mail/activity subsystems, which maintain their own invariants;
    #   * `active`, `color`, `sequence`, `parent_path`: record plumbing — archiving an
    #     employee from a payroll column is not a mapping, it is an accident;
    #   * avatars and images: binary, and a spreadsheet cell is not one;
    #   * presence/kanban helpers: derived display state with no meaning as a
    #     destination.
    # ------------------------------------------------------------------
    _EC_DENY_NAMES = (
        'id', 'display_name', '__last_update',
        'create_uid', 'create_date', 'write_uid', 'write_date',
        'active', 'color', 'sequence', 'parent_path',
        'hr_presence_state', 'last_activity', 'last_activity_time',
        'kanban_state', 'legend_blocked', 'legend_done', 'legend_normal',
        'image_1920', 'image_1024', 'image_512', 'image_256', 'image_128',
        'avatar_1920', 'avatar_1024', 'avatar_512', 'avatar_256', 'avatar_128',
        # MAPFIX E2 — hr.version plumbing. Odoo 19 delegates hr.employee's HR
        # data to a VERSION record (`_inherits`), which the predicate below now
        # follows; these four are the delegation's own machinery rather than
        # anything a spreadsheet describes. `version_id` points at the delegate
        # (re-pointing it from an import would move an employee's whole history
        # onto another record), `date_version` dates it, and the two
        # `last_modified_*` are write stamps in the family of `write_uid`.
        'version_id', 'date_version', 'last_modified_date', 'last_modified_uid',
    )
    _EC_DENY_PREFIXES = ('message_', 'activity_', 'access_', 'rating_',
                         'website_message_')

    # LANES — data-in-code, so adding a field to a lane later is a one-line edit.
    # A field named in NO lane still appears, under "Other …": the catalogue's
    # promise is that nothing is hidden, and a silent omission would break it.
    # `(lane key, ((model, curated field names in reading order), …))`.
    #
    # A lane may draw from BOTH models, and "Job & organisation" has to: on
    # Odoo 19 `hr.employee.department_id`, `job_id` and `resource_calendar_id` are
    # NOT STORED — the stored copies live on `hr.contract` (MF11). A reader looking
    # for "Department" must find it under Job & organisation whichever record
    # actually holds it; the sublabel on the card says which one that is.
    _EC_LANES = (
        ('identity', (('hr.employee', ('name', 'employee_id', 'barcode',
                                       'identification_id', 'passport_id',
                                       'registration_number')),)),
        ('personal', (('hr.employee', ('gender', 'birthday', 'place_of_birth',
                                       'country_of_birth', 'country_id', 'marital',
                                       'children', 'spouse_complete_name',
                                       'spouse_birthdate', 'certificate',
                                       'study_field', 'study_school',
                                       'emergency_contact', 'emergency_phone')),)),
        ('contact', (('hr.employee', ('work_email', 'work_phone', 'mobile_phone',
                                      'private_email', 'private_phone',
                                      'private_street', 'private_street2',
                                      'private_city', 'private_zip',
                                      'private_state_id', 'private_country_id',
                                      'work_location_id', 'km_home_work')),)),
        ('job', (('hr.employee', ('job_title', 'job_id', 'department_id',
                                  'parent_id', 'coach_id', 'company_id',
                                  'resource_calendar_id', 'employee_type')),
                 ('hr.contract', ('job_id', 'department_id',
                                  'resource_calendar_id', 'company_id')))),
        ('contract_terms', (('hr.contract', ('name', 'wage', 'wage_type',
                                             'hourly_wage', 'date_start',
                                             'date_end', 'trial_date_end',
                                             'structure_type_id',
                                             'contract_type_id', 'notes')),)),
        # JOURNEY J8 — two synthetic cards, not fields. See `_component_lane_items`.
        ('contract_component', ()),
        ('bank', ()),
        ('other_employee', (('hr.employee', ()),)),
        ('other_contract', (('hr.contract', ()),)),
    )
    _EC_FALLBACK_LANE = {'hr.employee': 'other_employee',
                         'hr.contract': 'other_contract'}

    # CR7 — `employee_id` is the payroll CODE on hr.employee (a Char, not a
    # relation); it is the single most-mapped destination on a real structure.
    # The identity lane doubles as the answer to "which role does wiring HERE
    # imply" when a contract component is re-routed (MF-B2).
    _EC_IDENTITY_FIELDS = ('name', 'employee_id', 'barcode', 'identification_id',
                           'passport_id', 'registration_number')

    # ------------------------------------------------------------------
    # COLROLES P3 — the BANK LANE.
    #
    # Four synthetic right-hand cards that are not fields of anything. They stand for
    # the parts of a `res.partner.bank`, which the import batch assembles from several
    # columns at once (`_sync_employee_bank_account`). Their ids are prefixed `b:` so
    # `employee_mapping_create` can tell them apart from a real `f:model:field` spec
    # by inspection, without a second argument that could go missing.
    # ------------------------------------------------------------------
    _BANK_LANE_ROLES = ('acc_number', 'bank_name', 'bank_bic', 'acc_holder_name')

    @api.model
    def _role_lane_label(self, role):
        """Swim-lane heading for a column role. Title-case: it heads a group of rows,
        where `_role_label` reads inside a sentence."""
        return {
            'payroll': _("Payroll"),
            'identity': _("Identity"),
            'profile': _("Employee profile"),
            'contract': _("Contract"),
            'bank': _("Bank"),
            'reference': _("Reference"),
        }.get(role or 'payroll', _("Payroll"))

    @api.model
    def _bank_lane_items(self):
        labels = {
            'acc_number': (_("Account number"),
                           _("Creates or updates the employee's bank account")),
            'bank_name': (_("Bank name"),
                          _("Names the bank on that account")),
            'bank_bic': (_("SWIFT / BIC code"),
                         _("Identifies the bank when the name is ambiguous")),
            'acc_holder_name': (_("Account holder name"),
                                _("Who the account is in the name of")),
        }
        group = self._ec_lane_label('bank')
        _pos, order_of = self._ec_lane_index()
        items = []
        for role in self._BANK_LANE_ROLES:
            label, sub = labels[role]
            items.append({
                'id': 'b:%s' % role,
                'label': label,
                'sublabel': sub,
                'group': group,
                'meta': {'bank_role': role, 'kind': 'bank', 'lane': 'bank',
                         'lane_order': order_of.get('bank', 99)},
            })
        return items

    # ------------------------------------------------------------------
    # JOURNEY J8 — the CONTRACT COMPONENT LANE.
    #
    # The commonest destination on this board was the only one it did not draw.
    # A contract component is not a field of `hr.contract`: it is a row of
    # `hr.contract.advantage` pointing at an `hr.contract.advantage.template`,
    # matched by CODE — so it could never come out of `ir.model.fields`, and
    # widening `_EC_TTYPES` would not have produced it either. What said so on
    # the board was a badge on the LEFT card, which only tells you where a value
    # lands if you already know what the badge means.
    #
    # Two cards rather than one, because the value TYPE is a real fork in the
    # model (`is_text_component` / `hr.contract.advantage.template.value_type`)
    # and it decides something a reader can see: an AMOUNT is read back into the
    # pay calculation, TEXT is not (`payroll_import_batch._transform_data_to_
    # formula_inputs` skips `value_type == 'text'` outright). Choosing at wire
    # time is the difference between the board expressing that decision and
    # hiding it.
    #
    # The `c:` prefix follows the `b:` precedent exactly, for the same reason:
    # `employee_mapping_create` tells the id kinds apart by inspection.
    # ------------------------------------------------------------------
    _COMPONENT_LANE_KINDS = ('amount', 'text')

    @api.model
    def _ec_component_codes(self, config):
        """`{'amount': [codes], 'text': [codes]}` for the config's flagged rules.

        A text component carries BOTH booleans (`employee_mapping_make_component`
        writes `is_contract_component=True, is_text_component=True`), so the test
        for "amount" has to exclude the text ones rather than merely assert the
        first flag."""
        out = {'amount': [], 'text': []}
        if not config:
            return out
        for rule in config.rule_ids:
            if not (rule.is_contract_component or rule.is_text_component):
                continue
            code = (rule.code or '').strip()
            if not code:
                continue
            out['text' if rule.is_text_component else 'amount'].append(code)
        return out

    @api.model
    def _ec_component_state(self, config):
        """Does this destination exist yet, and how much is in it — per kind.

        The owner's actual question ("are the rows not created yet?") and the
        reason the lane earns its place. Two facts, both from the database:

          * the TEMPLATE is created lazily by the first processed import
            (`payroll_import_batch._get_or_create_advantage_template`), so "no
            template" is the honest, common, not-a-problem state;
          * a LINE proves nothing on its own — `hr.contract.create` seeds one
            EMPTY advantage line per template on every contract (CR18) — so what
            is counted is a line somebody's data is actually IN, the same
            predicate `_ec_component_history` uses.

        Cost: ONE search over templates, plus ONE `count_distinct` aggregate per
        kind that actually HAS templates. On a database where nothing has been
        imported (every live one today) that is a single indexed search and no
        aggregate at all. The distinct-CONTRACT count is deliberately an
        aggregate rather than a line `search_count`: a lane spans several codes,
        and one contract carrying three of them is one contract, not three.
        """
        codes = self._ec_component_codes(config)
        state = {k: {'codes': len(codes[k]), 'templates': 0, 'contracts': 0}
                 for k in self._COMPONENT_LANE_KINDS}
        allcodes = codes['amount'] + codes['text']
        if not allcodes:
            return state
        Template = self.env.get('hr.contract.advantage.template')
        if Template is None:
            return state
        templates = Template.sudo().search([('code', 'in', allcodes)])
        if not templates:
            return state
        by_kind = {'amount': [], 'text': []}
        for tpl in templates:
            code = (tpl.code or '').strip()
            kind = 'text' if code in codes['text'] else 'amount'
            by_kind[kind].append(tpl.id)
        Line = self.env['hr.contract.advantage'].sudo()
        for kind in self._COMPONENT_LANE_KINDS:
            ids = by_kind[kind]
            state[kind]['templates'] = len(ids)
            if not ids:
                continue
            groups = Line._read_group(
                [('advantage_template_id', 'in', ids),
                 '|', ('amount', '!=', 0.0), ('text_value', '!=', False)],
                [], ['contract_id:count_distinct'])
            state[kind]['contracts'] = (groups and groups[0][0]) or 0
        return state

    @api.model
    def _ec_component_state_note(self, kind, state):
        """The card's live state line — `{text, title, tone}`, or None.

        `warn` is deliberately NOT used for "nothing yet": a component that no
        import has run for is correct, not broken, and painting it amber would
        make every clean scheme look like it had a problem.
        """
        st = (state or {}).get(kind) or {}
        if not st.get('codes'):
            return None
        if not st.get('templates'):
            return {
                'text': _("Created on the first import — nothing on any contract yet."),
                'title': _("The contract entry for a component is created the first "
                           "time pay data is processed for it. Wiring a column here "
                           "is what decides that it will be."),
                'tone': 'info',
            }
        n = st.get('contracts') or 0
        if not n:
            return {
                'text': _("Ready on the contract — no values stored yet."),
                'title': _("The contract entry exists, and no contract carries a "
                           "value under it so far."),
                'tone': 'info',
            }
        return {
            'text': _("%(n)s contracts carry a value.") % {'n': n},
            'title': _("Counted over the contracts that hold a real value for one "
                       "of this scheme's %(kind)s components — an empty row that "
                       "was created with the contract does not count.")
                     % {'kind': (_("amount") if kind == 'amount' else _("text"))},
            'tone': 'ok',
        }

    @api.model
    def _component_lane_items(self, config=None):
        """The lane's two cards, built the way `_bank_lane_items` builds its four.

        MAPFIX E2's construction invariant applies here too: a card rendered on
        the right carries the metadata it would have had from the catalogue —
        `lane`, `lane_order`, an id kind, a `note` — so nothing downstream has to
        ask which construction site made it.
        """
        labels = {
            'amount': (_("Contract component — amount"),
                       _("Kept on the contract under this column's own code, as a "
                         "number the pay calculation reads back")),
            'text': (_("Contract component — text"),
                     _("Kept on the contract as text — a grade, a shift code, a note")),
        }
        group = self._ec_lane_label('contract_component')
        _pos, order_of = self._ec_lane_index()
        state = self._ec_component_state(config)
        items = []
        for kind in self._COMPONENT_LANE_KINDS:
            label, sub = labels[kind]
            meta = {'component_kind': kind, 'kind': 'component',
                    'lane': 'contract_component',
                    'lane_order': order_of.get('contract_component', 99),
                    'mappable': True}
            note = self._ec_component_state_note(kind, state)
            if note:
                meta['note'] = note
            items.append({
                'id': 'c:%s' % kind,
                'label': label,
                'sublabel': sub,
                'group': group,
                'meta': meta,
            })
        return items

    @api.model
    def _ec_lane_label(self, key):
        """Lane heading. Literal `_()` calls so the terms stay extractable."""
        return {
            'identity': _("Identity"),
            'personal': _("Personal"),
            'contact': _("Contact"),
            'job': _("Job & organisation"),
            'contract_terms': _("Contract terms"),
            'contract_component': _("Contract components"),
            'bank': _("Bank account"),
            'other_employee': _("Other employee fields"),
            'other_contract': _("Other contract fields"),
        }.get(key, _("Other fields"))

    @api.model
    def _ec_is_mappable(self, model, fname):
        """Registry truth about one field, and the only place the inclusion rule
        lives. `ir.model.fields` is read for the LABEL and for the cheap first pass;
        whether a value can actually be WRITTEN is a question only the registry
        answers, and getting it wrong is silent in both directions — an over-strict
        rule hides a destination, a lax one accepts a write that goes nowhere."""
        Model = self.env.get(model)
        if Model is None:
            return False
        field = Model._fields.get(fname)
        if field is None:
            return False
        if field.readonly:
            return False
        if field.type not in self._EC_TTYPES:
            return False
        # ------------------------------------------------------------------
        # MAPFIX E2 — the question is WRITABLE, not STORED.
        #
        # This test used to read `not field.store or field.readonly`, and the
        # `store` half was wrong on Odoo 19. `hr.employee` delegates its HR data
        # to `hr.version` (`_inherits = {'hr.version': 'version_id'}`), so
        # `employee_type`, `marital`, `sex`, `passport_id`, `identification_id`,
        # `ssnid`, `job_title`, the whole private address — 45 fields on
        # hr.employee and 3 more on hr.contract — are RELATED, non-stored fields.
        # Every one of them is perfectly writable: a write propagates to the
        # delegate. `store` refused all of them, which is why the six-value
        # Employee Type selection the owner was looking at was nowhere on the
        # board (and why MF11's "department is on the contract only" reading was
        # a symptom of this, not the disease).
        #
        # A COMPUTE is fine as long as something knows where to put the value.
        # For a STORED one that is the column itself — Odoo's ordinary editable
        # computed field, which is what `hr.contract.department_id`, `job_id`,
        # `resource_calendar_id` and `company_id` are (MF11). For an UNSTORED one
        # it is `related` (the ORM writes through `_inverse_related`), an
        # explicit `inverse`, or `_inherits` delegation. Anything else is a
        # compute with nowhere to write, which is what the old `store` test was
        # really trying to say — so that is what this says now.
        if not field.store and not (
                getattr(field, 'inherited', False) or field.related or field.inverse):
            return False
        if fname in self._EC_DENY_NAMES:
            return False
        return not any(fname.startswith(p) for p in self._EC_DENY_PREFIXES)

    # ------------------------------------------------------------------
    # MAPFIX D4/D5 — what a destination will ACCEPT, said on the card.
    #
    # A card that reads "Marital Status · Employee" tells a person nothing about
    # whether their spreadsheet will fit it. Two field types can answer that
    # before the import runs rather than after:
    #
    #   * a SELECTION stores a key and displays a label, and
    #     `_coerce_mapped_value` validates the cell against the KEYS and stores
    #     None on a miss — silently. A reader who has only ever seen "Married"
    #     cannot know the file has to say `married`, so the card prints both;
    #   * a MANY2ONE either creates the record it cannot find or refuses to, and
    #     which of the two it does is `m2o_creates_missing` — the import batch's
    #     own predicate, imported rather than re-typed (MAPFIX D5).
    #
    # Both are built ONCE PER MODEL per call (`_ec_notes_for`) and then read out
    # of a dict, because this runs for 193 cards and a registry walk per card is
    # the difference between a board and a wait.
    # ------------------------------------------------------------------
    _EC_SEL_INLINE_MAX = 4        # values a card may list in full
    _EC_SEL_INLINE_CHARS = 44     # …and the width that listing may take
    # MAPFIX E1 — how many values a TRUNCATED note carries as structured data for
    # the popover (and names in its tooltip). The cap exists because one field can
    # dwarf a whole board: `hr.employee.tz` has **597** timezones on this build,
    # which is ~30 KB of payload and a 30 KB `title=` attribute for ONE card out
    # of 236 (MF34 has the measured figures). 120 is past
    # the point where anybody reads a list rather than searching it, and the
    # popover says out loud when it is showing a subset rather than pretending the
    # rest do not exist.
    _EC_SEL_VALUES_MAX = 120

    @api.model
    def _ec_selection_note(self, field):
        """`{text, title}` naming a selection field's permitted values, or None."""
        # `_description_selection` is the ORM's OWN resolver — the one `fields_get`
        # calls. It handles all three spellings of `selection=` (a literal list, a
        # callable, and the name of a method as a STRING) and applies
        # `selection_add` and the translations. Reading `field.selection` by hand
        # got `hr.employee.certificate` wrong on the live board: its selection is a
        # method NAME, and iterating a string yields characters, so the card said
        # nothing at all (MF24).
        try:
            sel = field._description_selection(self.env)
        except Exception:       # a selection that needs a record we do not have
            try:
                sel = field.selection
                if callable(sel):
                    sel = sel(self.env[field.model_name])
            except Exception:
                return None
        pairs = []
        for entry in (sel or []):
            if isinstance(entry, (list, tuple)) and len(entry) >= 2:
                pairs.append((str(entry[0]), str(entry[1])))
            elif isinstance(entry, (list, tuple)) and entry:
                pairs.append((str(entry[0]), str(entry[0])))
        if not pairs:
            return None
        shown = pairs[:self._EC_SEL_VALUES_MAX]
        # The list goes in the tooltip, keys included: the key is what the
        # spreadsheet must literally contain. Capped for the same reason the
        # structured list is (`_EC_SEL_VALUES_MAX`) — a tooltip nobody can read to
        # the end is not a smaller problem than a payload nobody wanted to send.
        full = ", ".join(
            ("%s (%s)" % (label, key)) if label != key else key
            for key, label in shown)
        if len(shown) < len(pairs):
            full = _("%(list)s … and %(n)s more") % {
                'list': full, 'n': len(pairs) - len(shown)}
        title = _("The file must contain one of these values (the code in "
                  "brackets is what is stored): %s") % full
        # The count is tested BEFORE the string is built: `tz` has 597 values and
        # joining them to discover they are too many is 30 KB of work per card.
        if len(pairs) <= self._EC_SEL_INLINE_MAX:
            inline = ", ".join(
                ("%s (%s)" % (label, key)) if label != key else key
                for key, label in pairs)
            if len(inline) <= self._EC_SEL_INLINE_CHARS:
                # Nothing was hidden, so there is nothing to open: no `values`, and
                # the board renders inert text rather than an affordance that does
                # nothing (MAPFIX E1).
                return {'text': inline, 'title': title, 'tone': ''}
        head = ", ".join(label for _key, label in pairs[:3])
        # MAPFIX E1 — a truncated note is OPENABLE, and this is what it opens into.
        # The tooltip was the only place the rest of the list existed: slow to
        # appear, impossible to select, cut off by the viewport and absent
        # altogether on a touch screen. `values` is sent ONLY when the inline text
        # actually hid something, which is what makes "clickable" and "truncated"
        # the same condition on both sides of the wire.
        return {'text': _("%(n)s values — %(head)s, …") % {
                    'n': len(pairs), 'head': head},
                'title': title, 'tone': '',
                'values': [{'key': key, 'label': label} for key, label in shown],
                'total': len(pairs)}

    @api.model
    def _ec_m2o_note(self, field):
        """`{text, title, tone}` saying whether an unseen value is created."""
        from odoo.addons.pb_hr_payroll_formula.models.payroll_import_batch import (
            m2o_creates_missing, m2o_resolution_key)
        comodel_name = field.comodel_name
        comodel = self.env.get(comodel_name)
        if comodel is None:
            return None
        what = (self.env['ir.model'].sudo()._get(comodel_name).name
                or comodel_name).lower()
        if m2o_creates_missing(comodel):
            return {'text': _("Creates the %s if it does not exist yet") % what,
                    'title': _("Values are matched on the name. Anything the "
                               "file names and Payobook does not have yet is "
                               "created during the import."),
                    'tone': ''}
        key = m2o_resolution_key(comodel)
        if key:
            return {'text': _("Must already exist — will not be created"),
                    'title': _("Matched on %(key)s. A %(what)s the file names "
                               "and Payobook does not have is left unset rather "
                               "than invented.") % {'key': key, 'what': what},
                    'tone': 'warn'}
        return {'text': _("Cannot be matched from a spreadsheet"),
                'title': _("This record has no name to match a cell against, so "
                           "the column would be read and then dropped."),
                'tone': 'warn'}

    @api.model
    def _ec_notes_for(self, model):
        """`{field name: note}` for one model — ONE registry walk per model, not
        one per card (the 193-card board is why this is not inline)."""
        Model = self.env.get(model)
        if Model is None:
            return {}
        notes = {}
        for fname, field in Model._fields.items():
            if not self._ec_is_mappable(model, fname):
                continue
            if field.type == 'selection':
                note = self._ec_selection_note(field)
            elif field.type == 'many2one':
                note = self._ec_m2o_note(field)
            else:
                continue
            if note:
                notes[fname] = note
        return notes

    @api.model
    def _ec_unwritable_msg(self, label):
        """ONE sentence for "this destination cannot receive a value".

        MAPFIX F3. The board already had this sentence, in `_ec_unmappable_note`'s
        tooltip, and the create guard needs to say the same thing — so it is said
        in one place and quoted twice rather than typed twice. A second refusal
        vocabulary is how a product ends up telling a user two different stories
        about one fact.
        """
        return _("%(label)s is not a field an import can write — the value would "
                 "be read and then dropped. Pick a different destination."
                 ) % {'label': label}

    @api.model
    def _ec_unmappable_note(self, fld):
        """The note on a card that is only here because something is WIRED to it.

        MAPFIX E2 (3). Losing the card would hide a live mapping, which is the
        worse failure — but rendering it as an ordinary destination is a lie: the
        catalogue refuses this field, so the import will read the column and put
        the value nowhere. Same `warn` tone as `_ec_m2o_note`'s refusals, because
        it is the same kind of news.
        """
        return {
            'text': _("This destination cannot be written — re-point this column"),
            'title': _("A column is mapped here, but %(label)s is not a field an "
                       "import can write. The value would be read and then "
                       "dropped. Send the column somewhere else, or remove the "
                       "mapping.") % {'label': fld.field_description or fld.name},
            'tone': 'warn',
        }

    @api.model
    def _ec_field_item(self, fld, lane=None, lane_order=None, notes=None):
        """ONE right-hand card, built ONE way.

        MAPFIX E2 (2) — there used to be two construction sites and they
        disagreed. The catalogue path passed a lane, a lane order and a per-model
        note map; the keep-a-wired-field-visible path inside
        `employee_mapping_data` passed NOTHING, so such a card landed in the
        fallback lane by accident rather than by decision and its note lookup went
        through `_ec_notes_for`, which gates on `_ec_is_mappable` — so a wired
        field the catalogue refuses rendered as a card with no note at all. The
        invariant this method now carries, and which `test_02` asserts over the
        whole board: *every card rendered on the right has the same metadata it
        would have had if it had come from the catalogue.*
        """
        if lane is None or lane_order is None:
            pos_of, order_of = self._ec_lane_index()
            hit = pos_of.get((fld.model, fld.name))
            if hit:
                lane_order, lane = hit[0], hit[1]
            else:
                lane = self._EC_FALLBACK_LANE.get(fld.model, 'other_employee')
                lane_order = order_of.get(lane, 99)
        meta = {'model': fld.model, 'field': fld.name, 'ttype': fld.ttype,
                'lane': lane, 'lane_order': lane_order}
        if not self._ec_is_mappable(fld.model, fld.name):
            meta['note'] = self._ec_unmappable_note(fld)
            meta['mappable'] = False
        else:
            # `notes` is passed in by every caller that renders more than one card;
            # a lone card (the wired-but-off-catalogue append) builds its own.
            if notes is None and fld.ttype in ('selection', 'many2one'):
                notes = self._ec_notes_for(fld.model)
            note = (notes or {}).get(fld.name)
            if note:
                meta['note'] = note
        return {'id': 'f:%s:%s' % (fld.model, fld.name),
                'label': fld.field_description or fld.name,
                'sublabel': self._EC_MODEL_LABEL.get(fld.model, fld.model),
                'group': self._ec_lane_label(lane),
                'meta': meta}

    @api.model
    def _ec_lane_index(self):
        """`{(model, field): (lane order, lane key, position in lane)}` plus the
        order of each lane key. Built once per call; eight lanes, ~60 names."""
        pos_of, order_of = {}, {}
        for n, (key, groups) in enumerate(self._EC_LANES):
            order_of[key] = n
            for model, names in groups:
                for pos, fname in enumerate(names):
                    pos_of.setdefault((model, fname), (n, key, pos))
        return pos_of, order_of

    @api.model
    def _ec_catalogue_domain(self, model):
        """The CHEAP first pass over `ir.model.fields` — and nothing more.

        MAPFIX E2 (4). This domain used to carry `store = True` and
        `readonly = False` as well, which made `ir.model.fields` a second, quieter
        copy of the inclusion rule — and the two copies can disagree, because
        `ir.model.fields.readonly` is a stored snapshot of a registry attribute
        that a later `_inherit` can change without the row being rewritten. Worse,
        a field excluded HERE never reached `_ec_is_mappable` at all, so widening
        the predicate would have changed nothing.
        One rule, in one place: this narrows the search to field types a
        spreadsheet cell can describe, and `_ec_is_mappable` (registry truth)
        decides. 330 rows on hr.employee and 106 on hr.contract — a dict lookup
        each, once per board.
        """
        return [('model', '=', model), ('ttype', 'in', list(self._EC_TTYPES))]

    @api.model
    def _ec_right_items(self, q=''):
        """The whole catalogue, in lane order; curated names first inside a lane,
        then alphabetically by label.

        Field metadata is model SCHEMA (no employee PII); sudo so non-admin payroll
        staff can see the destination list. Writes stay `_can_edit`-gated.
        """
        IMF = self.env['ir.model.fields'].sudo()
        ql = (q or '').strip().lower()
        pos_of, order_of = self._ec_lane_index()
        rows = []
        for model in ('hr.employee', 'hr.contract'):
            dom = self._ec_catalogue_domain(model)
            # MAPFIX D4 — once per MODEL, then read out of the dict per card.
            notes = self._ec_notes_for(model)
            for f in IMF.search(dom):
                if not self._ec_is_mappable(model, f.name):
                    continue
                label = f.field_description or f.name
                if ql and ql not in label.lower() and ql not in (f.name or '').lower():
                    continue
                hit = pos_of.get((model, f.name))
                if hit:
                    n, key, pos = hit
                else:
                    key = self._EC_FALLBACK_LANE[model]
                    n, pos = order_of[key], 10 ** 6
                rows.append((n, pos, label.lower(),
                             self._ec_field_item(f, key, n, notes)))
        rows.sort(key=lambda t: (t[0], t[1], t[2]))
        return [t[3] for t in rows]

    @api.model
    def _ec_right_column(self, q='', config=None):
        """The catalogue with the SYNTHETIC lanes spliced into their own lane
        positions, so the RIGHT column reads top-to-bottom in one order and the
        canvas' consecutive-group headers tell the truth.

        JOURNEY J8 — there are two synthetic lanes now (contract components, then
        bank), so the splice is driven by the lane index rather than by one
        remembered key. A second hand-written copy of "is it time to insert yet"
        is how the second lane would have landed in the wrong place on the day a
        third one arrived.
        """
        _pos, order_of = self._ec_lane_index()
        synthetic = [
            (order_of.get('contract_component', 99),
             self._component_lane_items(config)),
            (order_of.get('bank', 99), self._bank_lane_items()),
        ]
        synthetic.sort(key=lambda t: t[0])
        items, out = self._ec_right_items(q), []
        pending = list(synthetic)
        for it in items:
            lo = (it.get('meta', {}).get('lane_order') or 0)
            while pending and lo > pending[0][0]:
                out.extend(pending.pop(0)[1])
            out.append(it)
        for _at, cards in pending:
            out.extend(cards)
        return out

    # LEFT swim-lane order. Identity first because it is what finds the row at all;
    # payroll LAST because on this board it is the exception rather than the subject
    # (and is hidden entirely until somebody asks for it).
    _EC_ROLE_ORDER = ('identity', 'bank', 'profile', 'contract', 'reference', 'payroll')

    @api.model
    def _ec_left_items(self, config, include_payroll=False):
        """The config's columns as LEFT cards, in role lanes.

        Payroll components are EXCLUDED by default. That is the whole point of the
        board: a VPTQ structure has seventy pay columns and six people columns, and
        the six are what anybody opens this surface to wire. The seventy are one chip
        away, never deleted (W40 — nothing is silently unavailable).

        MAPFIX B2 — a contract component is no longer SEALED. Phase 3 drew those
        cards non-wirable because a component "already has a destination", which was
        true and is not the whole truth: the colour a spreadsheet was painted in is a
        SUGGESTION about where a column should go, and the person reading this board
        is allowed to disagree with it. The badge stays — it still says truthfully
        where the value lands today — but a wire may now be drawn from any card, and
        drawing one demotes the component (MF-B2). Existing contract values are kept
        as history; nothing is destroyed (MF-B3).
        """
        order = {role: n for n, role in enumerate(self._EC_ROLE_ORDER)}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        items = []
        for rule in rules:
            role = rule.column_role or 'payroll'
            if role == 'payroll' and not include_payroll:
                continue
            item = self._mc_item(rule, group_by_role=True)
            meta = item['meta']
            is_component = bool(rule.is_contract_component or rule.is_text_component)
            if rule.is_text_component:
                meta['badge'] = _("Text component")
                meta['badgeTone'] = 'text'
                meta['badgeHint'] = _(
                    "Kept on the contract as text. Wire it to a field if you would "
                    "rather it went there — what the contract already holds is kept "
                    "as history.")
            elif rule.is_contract_component:
                meta['badge'] = _("Contract component")
                meta['badgeTone'] = 'contract'
                meta['badgeHint'] = _(
                    "Kept on the contract as an amount. Wire it to a field if you "
                    "would rather it went there — what the contract already holds is "
                    "kept as history.")
            meta['isComponent'] = is_component
            meta['actions'] = self._ec_left_actions(rule, is_component)
            items.append((order.get(role, 99), rule.sequence or 0, rule.id, item))
        items.sort(key=lambda t: (t[0], t[1], t[2]))
        return [t[3] for t in items]

    @api.model
    def _ec_left_actions(self, rule, is_component):
        """The verbs a card offers. A column has exactly ONE destination, so every
        verb here MOVES it rather than adding a second home for the same value."""
        if rule.column_type != 'input':
            # A calculated or constant column is produced, not imported: it has
            # nothing to send anywhere and no contract to be kept on.
            return []
        # MAPFIX D3 — every verb carries a HINT now. The card no longer prints the
        # verbs across its own name; they live in a menu, and a menu row has the
        # width to say what it does instead of leaving the label to imply it.
        acts = []
        if is_component:
            acts.append({'key': 'to_field',
                         'label': _("Send to a field instead…"),
                         'hint': _("Pick an employee or contract field for this "
                                   "column. What the contract already holds is "
                                   "kept as history.")})
            if rule.is_text_component:
                acts.append({'key': 'make_amount', 'label': _("Make amount"),
                             'hint': _("Keep it on the contract as a number that "
                                       "the payroll calculation can read.")})
            else:
                acts.append({'key': 'make_text', 'label': _("Make text"),
                             'hint': _("Keep it on the contract as text — a "
                                       "grade, a shift code, a note.")})
            acts.append({'key': 'detach', 'label': _("Detach component"),
                         'hint': _("Stop keeping this column on the contract at "
                                   "all.")})
        else:
            acts.append({'key': 'make_amount',
                         'label': _("Make amount component"),
                         'hint': _("Keep it on the contract as a number that the "
                                   "payroll calculation can read.")})
            acts.append({'key': 'make_text', 'label': _("Make text component"),
                         'hint': _("Keep it on the contract as text — a grade, a "
                                   "shift code, a note.")})
        return acts

    @api.model
    def _ec_wire_right_id(self, mapping):
        """The RIGHT card id a persisted mapping points at, or None if the row is
        incomplete (a half-built mapping must not draw a wire to nowhere)."""
        if mapping.destination_type == 'bank_account':
            return ('b:%s' % mapping.bank_role) if mapping.bank_role else None
        if mapping.target_model_id and mapping.target_field_id:
            return 'f:%s:%s' % (mapping.target_model_id.model, mapping.target_field_id.name)
        return None

    @api.model
    def _ec_place_in_lane(self, items, item):
        """Insert a card at the END OF ITS OWN LANE rather than at the end of the
        column.

        MAPFIX E2 (2). A card appended after the last lane draws a second group
        header for a heading the reader has already scrolled past — the canvas
        emits one whenever the group CHANGES between consecutive rows, so an
        Identity card tacked on below "Other contract fields" grows an "Identity"
        heading of its own at the bottom of the board. Now that the appended card
        knows its lane, it can be put where that lane is.
        """
        lo = (item.get('meta') or {}).get('lane_order', 99)
        at = len(items)
        for n, other in enumerate(items):
            if ((other.get('meta') or {}).get('lane_order', 99) or 0) > lo:
                at = n
                break
        items.insert(at, item)
        return items

    @api.model
    def employee_mapping_data(self, config_id=None, context_id=None, include_payroll=False):
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        q = context_id.strip().lower() if isinstance(context_id, str) else ''
        left = self._ec_left_items(config, include_payroll=include_payroll)
        right = self._ec_right_column(q, config)
        present = {i['id'] for i in right}
        Mapping = self.env['hr.payslip.import.mapping'].sudo()
        by_left = {i['id']: i for i in left}
        wires = []
        for m in Mapping.search([('salary_structure_id', '=', config.id)]):
            if not m.component_id:
                continue
            rid = self._ec_wire_right_id(m)
            if not rid:
                continue
            wires.append({'id': 'em%s' % m.id, 'kind': 'mapping', 'ref': m.id,
                          'leftId': m.component_id.id, 'rightId': rid, 'state': 'accepted'})
            # JOURNEY J3 S1 (owner decision J-D4) — the card SAYS which way this
            # row runs. Additive: two new keys under `meta`, every pre-J3 key
            # untouched, so a stale bundle renders exactly what it rendered before.
            card = by_left.get(m.component_id.id)
            if card is not None:
                card['meta'].update(self._ec_direction(m))
            # A wired field must appear in RIGHT even when the catalogue did not
            # offer it — a search may have filtered it out, or the catalogue may
            # refuse the destination outright. MAPFIX E2: this goes through the
            # SAME `_ec_field_item` as the catalogue, which now resolves the lane
            # and the note itself, so the appended card is indistinguishable from
            # a catalogue one except in the ONE way that matters — an unwritable
            # destination says so, in the caution tone, instead of rendering as a
            # perfectly ordinary field with nothing to say for itself.
            if rid not in present:
                fld = self.env['ir.model.fields'].sudo().search(
                    [('model', '=', m.target_model_id.model), ('name', '=', m.target_field_id.name)], limit=1)
                if fld:
                    self._ec_place_in_lane(right, self._ec_field_item(fld))
                    present.add(rid)
        # ------------------------------------------------------------------
        # JOURNEY J8 — the component wires, which have no row behind them.
        #
        # There IS no `hr.payslip.import.mapping` for a contract component: the
        # boolean on the rule is the fact, and that is exactly why the board
        # could never draw one. They are synthesised here, in their own id
        # namespace and with their own `kind`, so that no client path can hand a
        # RULE id to `employee_mapping_delete` — which browses the mapping table
        # and would unlink a stranger's row. `ref` is deliberately False for the
        # same reason: even a careless caller gets an empty recordset.
        # ------------------------------------------------------------------
        for rule in config.rule_ids:
            if not (rule.is_contract_component or rule.is_text_component):
                continue
            kind = 'text' if rule.is_text_component else 'amount'
            wires.append({'id': 'cc%s' % rule.id, 'kind': 'component',
                          'ref': False, 'componentId': rule.id,
                          'componentKind': kind,
                          'leftId': rule.id, 'rightId': 'c:%s' % kind,
                          'state': 'accepted'})
            card = by_left.get(rule.id)
            if card is not None:
                card['meta'].update(self._ec_component_direction(rule))
        return {
            'ok': True, 'left': left, 'right': right, 'wires': wires,
            'left_title': config.name, 'right_title': _("Employee, contract & bank"),
            # J3 S1 — the subtitle stopped being true the moment the resolver
            # learned to read these rows back, which was long before this phase.
            'subtitle': _("People data goes to the record, and comes back when a "
                          "pay run needs it"),
            # J3 S1 — a CANVAS CAPABILITY, keyed off this adapter's payload rather
            # than switched on globally: only wires on this board render
            # double-headed, because only these rows run both ways. The API,
            # spreadsheet, scheme and cycle boards send nothing and are unchanged.
            'bidirectional': True,
            'supports_suggest': True, 'contexts': [], 'context_id': False,
            'include_payroll': bool(include_payroll),
            'counts': self._ec_role_counts(config, wires),
            'lanes': [self._role_lane_label(r) for r in self._EC_ROLE_ORDER],
            'can_edit': self._can_edit(),
            # MAPFIX B3 — how many columns still land nowhere. The footer bar reads
            # it; the problems rail counts the SAME set (`_ec_unresolved`), so the
            # two surfaces can never disagree about how much work is left.
            'unresolved': len(self._ec_unresolved(config)),
        }

    @api.model
    def _ec_direction(self, mapping):
        """Which way this mapping row actually runs, in a sentence.

        JOURNEY J3 S1 / owner decision J-D4. This board has always said "send to",
        and for a `field` row that is only half of what the code does:

          * on IMPORT the value is WRITTEN onto the record —
            `payroll_import_batch._apply_employee_writeback` / the contract half;
          * on a PAY RUN the resolver READS it back — `get_mapped_input_value`
            walks the same `hr.payslip.import.mapping` rows and returns the
            employee/contract field's value when neither the file nor the feed
            carried anything for that component.

        So a `field` row is genuinely two-way and the board now renders it that
        way. A `bank_account` row is NOT: `get_mapped_input_value` covers
        `hr.employee` and `hr.contract` only, and the resolver never reads a bank
        part back into an input. Printing the read-back half there would be the
        exact class of confident falsehood this phase exists to remove — so it
        gets the import half and nothing else, and `direction` says `to_record`
        rather than `two_way` so no later reader has to infer it from the prose.
        """
        if mapping.destination_type == 'bank_account':
            return {
                'direction': 'to_record',
                'directionNote': _("On import: builds the bank account."),
            }
        model = mapping.target_model_id
        field = mapping.target_field_id
        where = '%s › %s' % (
            model.name or model.model or '', field.field_description or field.name or '')
        return {
            'direction': 'two_way',
            'directionNote': _(
                "On import: fills %(where)s. On pay run: used when the file or "
                "feed leaves this empty.", where=where),
        }

    @api.model
    def _ec_component_direction(self, rule):
        """Which way a CONTRACT COMPONENT runs — and the two kinds differ.

        JOURNEY J8. An AMOUNT component is genuinely two-way, and it is the one
        rung of the resolver ladder that reads the contract rather than the file:
        `_transform_data_to_formula_inputs` builds `contract_component_amounts`
        from the contract's advantage lines and uses it when neither the file nor
        the feed carried anything (`resolved_source = 'contract_component'`), with
        a flagged rule that has no line falling back to `0.0`.

        A TEXT component is NOT read back, and the code says so out loud: the same
        loop SKIPS `value_type == 'text'` outright, because letting a text
        component in would feed a permanent 0.0 into any formula naming it. So the
        text card gets the import half and nothing else — the same refusal J3 made
        for a bank row rather than print a confident falsehood.
        """
        code = (rule.code or '').strip()
        if rule.is_text_component:
            return {
                'direction': 'to_record',
                'directionNote': _(
                    "On import: kept on the contract as text under %(code)s.",
                    code=code or (rule.name or '')),
            }
        return {
            'direction': 'two_way',
            'directionNote': _(
                "On import: kept on the contract as an amount under %(code)s. On "
                "pay run: read back from the contract when the file or feed "
                "leaves this empty.", code=code or (rule.name or '')),
        }

    @api.model
    def _ec_role_counts(self, config, wires):
        """Per-role {total, unmapped} for the header chips.

        Counted over ALL of the config's columns, not over the cards currently on the
        board: the payroll tally has to be truthful while payroll is hidden, or the
        chip that reveals it cannot say how much it is about to reveal.

        A CONTRACT COMPONENT is never "unmapped" — it already has a destination, and
        counting it as outstanding work would make the chips permanently red on a
        structure that is completely correct.
        """
        wired = {w['leftId'] for w in wires}
        counts = {}
        for rule in config.rule_ids:
            role = rule.column_role or 'payroll'
            bucket = counts.setdefault(role, {'total': 0, 'unmapped': 0,
                                              'label': self._role_lane_label(role)})
            bucket['total'] += 1
            if rule.id not in wired and not rule.is_contract_component:
                bucket['unmapped'] += 1
        return counts

    @api.model
    def ec_search_fields(self, query, config_id=None):
        """Autocomplete for the Employee/Contract tab: any writable scalar
        hr.employee / hr.contract field matching the query, so a user can append
        a field beyond the curated set and wire it. Metadata read via _ec_right_items
        (sudo'd)."""
        q = (query or '').strip().lower()
        if len(q) < 2:
            return {'ok': True, 'fields': []}
        return {'ok': True, 'fields': self._ec_right_items(q)[:40]}

    @api.model
    def ec_model_fields(self, model):
        """All mappable fields for ONE model (hr.employee | hr.contract), for the
        Employee/Contract browse dropdowns. Metadata only — sudo'd like
        _ec_right_items; writes still go through employee_mapping_create/delete.

        MAPFIX B1: filtered through the SAME catalogue rule as the right column, so
        the dropdown can never offer a destination the board would refuse."""
        if model not in ('hr.employee', 'hr.contract'):
            return {'ok': False, 'fields': []}
        IMF = self.env['ir.model.fields'].sudo()
        pos_of, order_of = self._ec_lane_index()
        dom = self._ec_catalogue_domain(model)      # MAPFIX E2 — one rule, one place
        items = []
        notes = self._ec_notes_for(model)      # MAPFIX D4 — once, not per field
        for f in IMF.search(dom, order='field_description'):
            if not self._ec_is_mappable(model, f.name):
                continue
            hit = pos_of.get((model, f.name))
            key = hit[1] if hit else self._EC_FALLBACK_LANE[model]
            items.append(self._ec_field_item(f, key, order_of.get(key, 99), notes))
        return {'ok': True, 'fields': items}

    @api.model
    def _ec_spec(self, target_spec):
        """A right-hand card id, as a STRING, whatever arrived.

        MAPFIX D1. Every id this board mints is a string (`f:model:field`,
        `b:role`), so anything else is a client-side mix-up rather than a
        destination — and `str()` turns it into one that simply matches nothing,
        instead of a crash three lines later. Booleans and `None` become '' so the
        existing "nothing was chosen" path still fires.
        """
        if target_spec in (None, False, True):
            return ''
        if isinstance(target_spec, str):
            return target_spec.strip()
        return str(target_spec)

    @api.model
    def _ec_bad_spec_msg(self):
        """One sentence for every unusable destination, so a refusal reads the
        same wherever it is raised."""
        return _("That connection could not be made — one of its ends was not "
                 "recognised. Pick a card from the list and try again.")

    @api.model
    def employee_mapping_create(self, config_id, context_id, component_id, target_spec):
        """Wire one column to one destination.

        MAPFIX D1 — `target_spec` is whatever the browser sent, and the browser is
        allowed to be wrong. It used to be read as `target_spec or ''`, which
        catches a falsy value and not a wrong TYPE: an integer left-hand id
        survived that guard untouched and `123.startswith('b:')` was an
        AttributeError on the user's screen. A malformed spec is a refusal, not a
        traceback — the caller already renders `msg`, and the board stays usable.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        spec = self._ec_spec(target_spec)
        if not config:
            return {'ok': False}
        comp = self.env['hr.formula.rule'].browse(self._as_id(component_id))
        if not comp.exists():
            return {'ok': False}
        Mapping = self.env['hr.payslip.import.mapping'].sudo()

        if spec.startswith('b:'):
            bank_role = spec[2:]
            if bank_role not in self._BANK_LANE_ROLES:
                return {'ok': False, 'msg': self._ec_bad_spec_msg()}
            dest = dict(Mapping._fields['bank_role'].selection).get(bank_role, bank_role)
            note = self._ec_demote_component(comp, 'bank', dest)
            # 1:1 on both sides within this config — drop any existing on either end
            Mapping.search(['&', ('salary_structure_id', '=', config.id),
                            '|', ('component_id', '=', comp.id),
                            '&', ('destination_type', '=', 'bank_account'),
                            ('bank_role', '=', bank_role)]).unlink()
            Mapping.create({'salary_structure_id': config.id, 'component_id': comp.id,
                            'destination_type': 'bank_account', 'bank_role': bank_role})
            return {'ok': True, 'msg': note} if note else {'ok': True}

        # ------------------------------------------------------------------
        # JOURNEY J8 — the contract-component lane. Placed BEFORE the `f:` parse
        # for the same reason `b:` is: these ids are not `model:field` and a
        # three-part split would read `c:amount` as malformed rather than as the
        # destination it is.
        #
        # It ROUTES to the existing promotion rather than repeating it.
        # `employee_mapping_make_component` already refuses a non-`input`
        # column, already refuses a type clash and already sets the role per
        # CR-A2. A second implementation here would be the fork this programme
        # keeps refusing.
        #
        # JOURNEY J9 — it no longer "unlinks any field or bank row so the column
        # keeps exactly ONE destination". The owner withdrew that restriction:
        # promoting ADDS a destination. Note the deliberate asymmetry, which is
        # J9's scope boundary rather than an oversight — drawing a wire to a
        # NATIVE FIELD still demotes a contract component (MAPFIX B2's
        # `_ec_demote_component`, untouched here), because that is a different
        # mechanism with its own sentence and its own tests.
        # ------------------------------------------------------------------
        if spec.startswith('c:'):
            kind = spec[2:]
            if kind not in self._COMPONENT_LANE_KINDS:
                return {'ok': False, 'msg': self._ec_bad_spec_msg()}
            return self.employee_mapping_make_component(comp.id, kind)

        parts = spec.split(':')
        if len(parts) != 3 or parts[0] != 'f':
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        model, fname = parts[1], parts[2]
        mdl = self.env['ir.model'].sudo().search([('model', '=', model)], limit=1)
        fld = self.env['ir.model.fields'].sudo().search([('model', '=', model), ('name', '=', fname)], limit=1)
        if not (mdl and fld):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        # ------------------------------------------------------------------
        # MAPFIX F3 — the catalogue's predicate, applied where the WRITE happens.
        #
        # The board refuses to OFFER a destination that fails `_ec_is_mappable`,
        # and until now that was the only place it was asked: the RPC accepted
        # anything `hr.payslip.import.mapping`'s own domain accepted, so a stale
        # board, the search box or a direct call could still mint a wire whose
        # import reads a column and puts the value nowhere. The rule now lives on
        # both sides of the browser boundary — same predicate, same sentence.
        #
        # CREATE only, never READ. Rows written before this guard may point at
        # fields the predicate now refuses; `employee_mapping_data` still loads
        # them, still draws their wire and still marks the card `warn`
        # (`_ec_unmappable_note`). Refusing to LOAD them would hide a live mapping,
        # which is the worse failure and the opposite of what this is for.
        #
        # Bank (`b:`) specs never reach here: they are not fields of anything and
        # have returned above.
        # ------------------------------------------------------------------
        if not self._ec_is_mappable(model, fname):
            return {'ok': False,
                    'msg': self._ec_unwritable_msg(fld.field_description or fname)}
        # MAPFIX B2 — wiring a contract component to a native field DEMOTES it.
        # Nothing about the advantage template or its lines is touched: the value
        # simply stops being written there on the next run (MF-B3).
        note = self._ec_demote_component(
            comp, self._ec_role_for_destination(model, fname),
            fld.field_description or fname)
        # 1:1 on both sides within this config — drop any existing on either end
        Mapping.search(['&', ('salary_structure_id', '=', config.id),
                        '|', ('component_id', '=', comp.id),
                        '&', ('target_model_id', '=', mdl.id), ('target_field_id', '=', fld.id)]).unlink()
        Mapping.create({'salary_structure_id': config.id, 'component_id': comp.id,
                        'destination_type': 'field',
                        'target_model_id': mdl.id, 'target_field_id': fld.id})
        return {'ok': True, 'msg': note} if note else {'ok': True}

    @api.model
    def _ec_role_for_destination(self, model, fname):
        """Which role a column takes on when it is re-routed to a native field."""
        if model == 'hr.contract':
            return 'contract'
        return 'identity' if fname in self._EC_IDENTITY_FIELDS else 'profile'

    @api.model
    def _ec_component_history(self, rule):
        """How many contracts carry a real value under this component's code.

        NOT `search_count` on the lines: `hr.contract.create` seeds one empty
        advantage line per template on EVERY contract (CR18), so their mere
        existence proves nothing. What counts as history is a line somebody's data
        is actually IN."""
        code = (rule.code or '').strip()
        if not code:
            return 0
        templates = self.env['hr.contract.advantage.template'].sudo().search(
            [('code', '=', code)])
        if not templates:
            return 0
        return self.env['hr.contract.advantage'].sudo().search_count([
            ('advantage_template_id', 'in', templates.ids),
            '|', ('amount', '!=', 0.0), ('text_value', '!=', False),
        ])

    @api.model
    def _ec_demote_component(self, rule, role, dest_label):
        """MF-B2/B3 — stop writing this column to the contract, keep what is there.

        The advantage TEMPLATE and every `hr.contract.advantage` line under it are
        left exactly as they are, along with their `hr.contract.advantage.change`
        audit rows. This is deliberate and it supersedes Phase 3's detach refusal on
        this path: a re-route is a decision about the FUTURE, and destroying the past
        to record it would be the one outcome nobody asked for. Returns the sentence
        the board shows once, or '' when there was nothing to demote.
        """
        if not (rule.is_contract_component or rule.is_text_component):
            return ''
        kept = self._ec_component_history(rule)
        rule.write({'is_contract_component': False, 'is_text_component': False,
                    'column_role': role, 'column_role_source': 'user'})
        if kept:
            return _(
                "Existing contract values for this component are kept as history; "
                "new imports will write to %(dest)s instead."
            ) % {'dest': dest_label}
        return _("%(name)s now goes to %(dest)s instead of the contract.") % {
            'name': rule.name or rule.code or '', 'dest': dest_label}

    @api.model
    def _ec_component_type_clash(self, rule, value_type):
        """The sentence to refuse with when the contract already types this code
        the other way — or '' when there is nothing in the way.

        JOURNEY J8. `_get_or_create_advantage_template` NEVER flips an existing
        template's `value_type`: every line already filed under it was written as
        the other kind, and re-typing would silently reinterpret that history. It
        logs a warning instead — server-side, where no user will ever see it. So
        without this guard the board would accept a wire, say it succeeded, and
        the next import would keep writing the old kind: a promise made on screen
        and declined in a log file.

        Only fires when a template EXISTS with the other type, which is exactly
        the case the import refuses. A code with no template yet is free to be
        either.
        """
        code = (rule.code or '').strip()
        if not code:
            return ''
        Template = self.env.get('hr.contract.advantage.template')
        if Template is None:
            return ''
        tpl = Template.sudo().search([('code', '=', code)], limit=1)
        if not tpl or 'value_type' not in tpl._fields:
            return ''
        current = tpl.value_type or 'amount'
        if current == value_type:
            return ''
        if current == 'text':
            return _(
                "%(code)s is already kept on the contract as TEXT, and the "
                "contract entry is never re-typed — an import would go on writing "
                "text. Send this column to a field instead, or leave it as a text "
                "component.") % {'code': code}
        return _(
            "%(code)s is already kept on the contract as an AMOUNT, and the "
            "contract entry is never re-typed — an import would go on writing a "
            "number. Send this column to a field instead, or leave it as an "
            "amount component.") % {'code': code}

    # ------------------------------------------------------------------
    # COLROLES P3 — turning a stranded column into a contract component.
    #
    # The commonest thing on this board is a column with no field to go to: a grade,
    # a shift code, a note the payroll office types every month. Before this, the
    # only honest answer was "it is read and then dropped". Now there is a second
    # one: keep it ON THE CONTRACT as text, where it is visible, versioned and
    # exportable, without inventing a custom field for it.
    # ------------------------------------------------------------------
    @api.model
    def employee_mapping_make_component(self, rule_id, value_type='amount'):
        """Promote a column to a contract component — MAPFIX B2, both directions.

        Phase 3 could only make a TEXT component, because that was the only offer it
        had for a stranded column. An amount component is the other half of the same
        idea and the commoner one: an allowance whose value belongs to the contract
        rather than to this month's spreadsheet. CR-A2 fixes the roles — an amount
        component keeps role `payroll` because it feeds the calculation; a text one
        takes `contract`.
        """
        if value_type not in ('amount', 'text'):
            return {'ok': False, 'msg': _("Unknown component type.")}
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        rule = self.env['hr.formula.rule'].browse(self._as_id(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _("That column no longer exists.")}
        if rule.column_type != 'input':
            return {'ok': False, 'msg': _(
                "Only an imported column can be kept on the contract — this one is "
                "calculated.")}
        # JOURNEY J8 — a promise the import would quietly decline is refused here.
        clash = self._ec_component_type_clash(rule, value_type)
        if clash:
            return {'ok': False, 'msg': clash}
        # JOURNEY J9 — THE UNLINK IS GONE, and that is the whole of §4.5.1.
        #
        # It used to read: "an existing mapping would mean the value lands in two
        # places, which is two sources of truth for one fact — the wire goes."
        # The owner has withdrawn that premise. A component may be fed by a
        # spreadsheet column AND a connected system AND be kept on the contract;
        # the resolver reads them in a stated order and the card shows all of
        # them, ranked. Promoting a column to a contract component therefore ADDS
        # a source instead of silently replacing one. J8's type-clash refusal
        # above is untouched: that is a promise the import would decline, which
        # is a different thing from a restriction nobody asked for.
        is_text = value_type == 'text'
        rule.write({
            'is_contract_component': True,
            'is_text_component': is_text,
            'column_role': 'contract' if is_text else 'payroll',
            'column_role_source': 'user',
        })
        name = rule.name or rule.code or ''
        return {'ok': True, 'msg': (
            _("%s is now kept on the contract as text.") % name if is_text
            else _("%s is now kept on the contract as an amount.") % name)}

    @api.model
    def employee_mapping_make_text_component(self, rule_id):
        """Kept as a thin alias: the Phase-3 name is what any saved client bundle
        still calls, and a renamed RPC is a broken button until the cache turns."""
        return self.employee_mapping_make_component(rule_id, 'text')

    @api.model
    def employee_mapping_detach_component(self, rule_id):
        """The reverse. Refused once contracts actually carry values for this code —
        clearing the flags would leave those advantage lines orphaned, pointing at a
        template nothing maintains any more."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        rule = self.env['hr.formula.rule'].browse(self._as_id(rule_id))
        if not rule.exists():
            return {'ok': False, 'msg': _("That column no longer exists.")}
        code = (rule.code or '').strip()
        if code:
            Template = self.env['hr.contract.advantage.template'].sudo()
            templates = Template.search([('code', '=', code)])
            # NOT `search_count` on the lines: `hr.contract.create` seeds one empty
            # advantage line per template on EVERY contract, so the mere existence of
            # lines proves nothing and would refuse every detach ever attempted. What
            # blocks a detach is a line somebody's data is actually IN.
            filled = self.env['hr.contract.advantage'].sudo().search_count([
                ('advantage_template_id', 'in', templates.ids),
                '|', ('amount', '!=', 0.0), ('text_value', '!=', False),
            ]) if templates else 0
            if filled:
                # MAPFIX B2 note: RE-ROUTING such a component is allowed and keeps
                # the history (MF-B3). Detaching is different — it leaves the value
                # with nowhere to go at all — so the refusal stands, and the sentence
                # now names the door that IS open.
                return {'ok': False, 'msg': _(
                    "%(count)s contracts already carry a value for %(code)s. Wire "
                    "this column to a field instead — what the contracts hold is "
                    "kept as history — or leave it on the contract."
                ) % {'count': filled, 'code': code}}
        # JOURNEY J8 — the snapshot the Undo toast puts back. Taken BEFORE the
        # write and consisting only of what this method changes: the two booleans
        # plus the role and its source, which `employee_mapping_make_component`
        # rewrites on the way back in. Nothing derived, nothing about the
        # advantage template — a detach never touched either (MJ32's rule: an
        # undo is the inverse of the delete, not a replay of the create).
        snapshot = {
            'rule_id': rule.id,
            'is_contract_component': bool(rule.is_contract_component),
            'is_text_component': bool(rule.is_text_component),
            'column_role': rule.column_role or False,
            'column_role_source': rule.column_role_source or False,
        }
        rule.write({'is_contract_component': False, 'is_text_component': False})
        return {'ok': True, 'snapshot': snapshot,
                'msg': _("%s is no longer kept on the contract.") % (
                    rule.name or code or '')}

    @api.model
    def employee_component_restore(self, snapshot):
        """Put a detached component back exactly as it was — JOURNEY J8.

        The inverse of `employee_mapping_detach_component`, and deliberately NOT
        a call to `employee_mapping_make_component`: a promotion is a DECISION —
        it unlinks any rival mapping row and re-derives the role from the value
        type — where an undo has to restore the role and the role SOURCE the
        column actually had. MJ32 taught this on the API board with a label that
        a redraw would have rewritten; the same shape, one board over.

        Idempotent, so a double-pressed Undo restores one component. Refuses a
        malformed payload rather than raising: the toast is gone by then and a
        traceback would have nowhere to land.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        if not isinstance(snapshot, dict):
            return {'ok': False, 'msg': self._ec_bad_spec_msg()}
        rule = self.env['hr.formula.rule'].browse(self._as_id(snapshot.get('rule_id')))
        if not rule.exists():
            return {'ok': False, 'msg': _("That column no longer exists.")}
        rule.write({
            'is_contract_component': bool(snapshot.get('is_contract_component')),
            'is_text_component': bool(snapshot.get('is_text_component')),
            'column_role': snapshot.get('column_role') or 'payroll',
            'column_role_source': snapshot.get('column_role_source') or 'user',
        })
        return {'ok': True, 'msg': _("%s is kept on the contract again.") % (
            rule.name or rule.code or '')}

    @api.model
    def employee_mapping_delete(self, mapping_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        m = self.env['hr.payslip.import.mapping'].sudo().browse(self._as_id(mapping_id))
        if m.exists():
            m.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # MAPFIX B3 — RECONCILIATION: nothing left behind.
    #
    # Unwiring a column does not put it back where it was; it makes it UNRESOLVED,
    # and an unresolved column is imported, read once and then dropped. Before this
    # there was no surface that said so out loud — the board simply had a card with
    # no wire, which reads as "not done yet" forever and as "fine" after a week.
    #
    # A column is unresolved when NOTHING claims the value it carries:
    #   * it is imported (`input`) — a calculated or constant column is produced,
    #     not received, so it needs no destination;
    #   * it is not a contract component;
    #   * it has no mapping row (field or bank);
    #   * its role is not already `reference` — that IS a destination, it is the
    #     word for "stored on the import and deliberately used nowhere";
    #   * no other column's formula reads it, by code or by column letter. A payroll
    #     input that feeds a calculation is resolved by definition (CR2);
    #   * and, for a payroll column, it does not print on the payslip. The payslip
    #     line IS the destination of a pay column — see the ledger entry: pre-ticking
    #     seventy VPTQ pay columns to "become a contract component" would be a
    #     destructive default dressed as tidiness.
    # ------------------------------------------------------------------
    @api.model
    def _ec_unresolved(self, config):
        """The config's columns that land nowhere, in board order. ONE definition,
        read by the board footer, the reconciliation dialog and the problems rail —
        three surfaces that would otherwise drift apart within a phase."""
        if not config:
            return self.env['hr.formula.rule']
        rules = config.rule_ids
        Mapping = self.env['hr.payslip.import.mapping'].sudo()
        mapped = set(Mapping.search([('salary_structure_id', '=', config.id)])
                     .filtered(lambda m: m.component_id and (
                         (m.destination_type == 'field'
                          and m.target_model_id and m.target_field_id)
                         or (m.destination_type == 'bank_account' and m.bank_role)))
                     .mapped('component_id').ids)
        by_col = {(r.column_letter or '').strip().upper(): r
                  for r in rules if r.column_letter}
        refs_by_rule = {}
        for r in rules:
            if r.column_type != 'formula':
                continue
            cols = set(self._expand_refs(r.excel_formula, by_col))
            for code in (r.formula_dependencies or '').split(','):
                code = code.strip().upper()
                if code:
                    cols.add(code)
            refs_by_rule[r.id] = cols
        out = []
        for r in rules.sorted(key=lambda x: (x.sequence or 0, x.id)):
            if r.column_type != 'input':
                continue
            if r.is_contract_component or r.is_text_component:
                continue
            role = r.column_role or 'payroll'
            if role == 'reference' or r.id in mapped:
                continue
            names = {n for n in ((r.code or '').strip().upper(),
                                 (r.column_letter or '').strip().upper()) if n}
            if any(names & refs for rid, refs in refs_by_rule.items() if rid != r.id):
                continue
            if role == 'payroll' and r.appears_on_payslip:
                continue
            out.append(r.id)
        return self.env['hr.formula.rule'].browse(out)

    @api.model
    def _ec_sample_values(self, config, rule, limit=3):
        """A few real values for this column, so the reader decides against the data
        rather than against its header. Read-only: sample vectors are code-keyed
        JSON, with the column letter as the legacy fallback."""
        code = (rule.code or '').strip()
        letter = (rule.column_letter or '').strip()
        out = []
        for sample in config.sample_data_ids:
            if len(out) >= limit:
                break
            try:
                vals = json.loads(sample.input_values_json or '{}')
            except (ValueError, TypeError):
                continue
            value = vals.get(code) if code in vals else vals.get(letter)
            if value in (None, '', False):
                continue
            out.append(value)
        return out

    @api.model
    def _ec_unresolved_rows(self, config):
        from odoo.addons.pb_hr_payroll_formula.models import column_role_classifier as crc
        rows = []
        for rule in self._ec_unresolved(config):
            raw = self._ec_sample_values(config, rule)
            texty = any(crc.is_texty_sample(v) for v in raw)
            rows.append({
                'id': rule.id,
                'name': rule.name or rule.code or '',
                'code': rule.code or '',
                'col': rule.column_letter or '',
                'role': rule.column_role or 'payroll',
                'role_label': self._role_lane_label(rule.column_role or 'payroll'),
                'samples': [self._ec_sample_text(v) for v in raw],
                'value_type': 'text' if texty else 'amount',
            })
        return rows

    @api.model
    def _ec_sample_text(self, value):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    @api.model
    def employee_mapping_unresolved(self, config_id=None):
        """What the reconciliation dialog opens on. Nothing is written."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        return {'ok': True, 'rows': self._ec_unresolved_rows(config),
                'config_name': config.name, 'can_edit': self._can_edit()}

    @api.model
    def employee_mapping_resolve_remaining(self, config_id, decisions,
                                           include_payroll=False):
        """Apply the reconciliation, all or nothing.

        `decisions` is `[{'id', 'component': bool, 'value_type': 'amount'|'text'}]`.
        Every id is validated against the CURRENT unresolved set before a single row
        is written: a stale dialog (someone wired a column in another tab) must not
        half-apply, and a foreign id must not reach `make_component` at all. Ticked
        rows go through the same promotion path as the board's own verb, so there is
        one implementation of "become a contract component" and not two.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'reason': 'no_config'}
        allowed = set(self._ec_unresolved(config).ids)
        plan = []
        for row in (decisions or []):
            if not isinstance(row, dict):
                return {'ok': False, 'msg': _("That list could not be read. "
                                              "Nothing was changed.")}
            try:
                rid = int(row.get('id'))
            except (TypeError, ValueError):
                return {'ok': False, 'msg': _("That list could not be read. "
                                              "Nothing was changed.")}
            if rid not in allowed:
                return {'ok': False, 'msg': _(
                    "This list is out of date — one of these columns already has a "
                    "destination. Nothing was changed; reopen it to see where things "
                    "stand.")}
            vtype = row.get('value_type') or 'amount'
            if vtype not in ('amount', 'text'):
                return {'ok': False, 'msg': _("Unknown component type. "
                                              "Nothing was changed.")}
            plan.append((rid, bool(row.get('component')), vtype))
        made = kept = 0
        for rid, as_component, vtype in plan:
            if as_component:
                res = self.employee_mapping_make_component(rid, vtype)
                if not res.get('ok'):
                    # The whole payload fails together — a partially-applied
                    # reconciliation is worse than none, because the dialog it came
                    # from no longer describes anything.
                    raise UserError(res.get('msg') or _("That change could not be saved."))
                made += 1
            else:
                self.env['hr.formula.rule'].browse(rid).write({
                    'column_role': 'reference', 'column_role_source': 'user'})
                kept += 1
        data = self.employee_mapping_data(config.id, include_payroll=include_payroll)
        data['applied'] = {'components': made, 'reference': kept}
        return data

    # ------------------------------------------------------------------
    # COLROLES P3 — suggestions for the people board.
    #
    # Two different machines, because the two halves of the board are different
    # problems. A BANK column is a closed vocabulary — there are exactly four things
    # it can be, and the classifier's own lexicon already knows the words for them,
    # in English and Vietnamese. An identity/profile/contract column is an open one:
    # the target is any of hundreds of fields, and the only honest tool is a label
    # comparison whose score is shown to the reader rather than hidden behind a
    # verdict.
    #
    # Nothing is written. A suggestion is a drawn wire with a confidence on it, and
    # accepting it goes through `employee_mapping_create` like a hand-drawn one —
    # which is what makes "Accept all ≥90%" safe to press.
    # ------------------------------------------------------------------

    # Words that place a bank column into one of the four lanes. Ordered: the more
    # specific tests run first, because "tên chủ tài khoản" (account holder name)
    # contains "tài khoản" (account) and would otherwise be read as the number.
    _BANK_SUGGEST_TOKENS = (
        ('acc_holder_name', ('account holder', 'holder name', 'beneficiary name',
                             'beneficiary', 'chu tai khoan', 'ten chu tai khoan',
                             'account name', 'name on account')),
        ('bank_bic', ('swift', 'bic', 'swift code', 'bank code', 'ifsc', 'iban code',
                      'ma ngan hang', 'ma swift')),
        ('acc_number', ('account number', 'account no', 'acc no', 'acct no', 'a c no',
                        'bank account', 'bank account no', 'bank account number',
                        'so tai khoan', 'stk', 'tai khoan ngan hang', 'so tk',
                        'account', 'iban')),
        ('bank_name', ('bank name', 'name of bank', 'bank', 'ngan hang',
                       'ten ngan hang', 'chi nhanh', 'branch', 'branch name')),
    )

    # Explicit destinations for the handful of columns EVERY structure has. A label
    # comparison gets "Employee Code" → `employee_id` wrong often enough to matter
    # (there are a dozen fields with "code" or "employee" in their label), and this
    # is the one place where naming the answer beats scoring it. Every entry is
    # verified against `ir.model.fields` before it is offered, so a hint for a field
    # some module has not installed simply does not appear.
    _EC_SUGGEST_HINTS = (
        ('identity', ('employee code', 'emp code', 'employee id', 'staff code',
                      'staff id', 'msnv', 'ma nv', 'ma nhan vien', 'ma so nhan vien',
                      'payroll id', 'personnel number'), 'hr.employee', 'employee_id'),
        ('identity', ('employee name', 'full name', 'fullname', 'staff name',
                      'ho ten', 'ho va ten', 'ten nhan vien'), 'hr.employee', 'name'),
        ('identity', ('id no', 'id number', 'identity card', 'identification',
                      'identification no', 'national id', 'citizen id', 'cmnd',
                      'cccd', 'so cmnd', 'so cccd'), 'hr.employee', 'identification_id'),
        ('identity', ('passport', 'passport no', 'passport number', 'ho chieu'),
         'hr.employee', 'passport_id'),
        ('identity', ('badge id', 'badge number', 'barcode'), 'hr.employee', 'barcode'),
        ('profile', ('tax code', 'tax id', 'tax number', 'ma so thue'),
         'hr.employee', 'identification_id'),
        ('profile', ('work email', 'email', 'email address'), 'hr.employee', 'work_email'),
        ('profile', ('phone', 'phone number', 'mobile', 'mobile number',
                     'so dien thoai', 'dien thoai'), 'hr.employee', 'mobile_phone'),
        ('profile', ('job title', 'job position', 'position', 'designation',
                     'chuc vu', 'chuc danh'), 'hr.employee', 'job_title'),
        ('profile', ('marital status', 'marital', 'tinh trang hon nhan'),
         'hr.employee', 'marital'),
        ('contract', ('date of joining', 'joining date', 'join date', 'doj',
                      'hire date', 'date of hire', 'start date', 'contract start',
                      'ngay vao lam', 'ngay vao cong ty', 'ngay bat dau'),
         'hr.contract', 'date_start'),
        ('contract', ('end date', 'contract end', 'last working day', 'leaving date',
                      'termination date', 'ngay ket thuc', 'ngay nghi viec'),
         'hr.contract', 'date_end'),
        ('contract', ('probation', 'probation end', 'probation end date',
                      'ngay thu viec', 'thoi gian thu viec'),
         'hr.contract', 'trial_date_end'),
        ('contract', ('wage', 'basic salary', 'contract salary', 'luong co ban',
                      'luong hop dong'), 'hr.contract', 'wage'),
    )

    @api.model
    def _ec_suggest_bank_role(self, crc, header):
        """Which bank lane a header belongs in, or None. `(role, exact)` — `exact`
        distinguishes a whole-header match from a contained one, and that is the
        difference between the 0.95 and the 0.75 tier."""
        key = crc.strip_accents(crc.normalize_header(header))
        if not key:
            return None, False
        for role, tokens in self._BANK_SUGGEST_TOKENS:
            for token in tokens:
                if key == crc.strip_accents(token):
                    return role, True
        for role, tokens in self._BANK_SUGGEST_TOKENS:
            for token in tokens:
                if crc.strip_accents(token) in key:
                    return role, False
        return None, False

    @api.model
    def _ec_hint_target(self, crc, role, header, available):
        """A curated destination for this header, if one is both listed and real."""
        key = crc.strip_accents(crc.normalize_header(header))
        if not key:
            return None, False
        for hint_role, tokens, model, fname in self._EC_SUGGEST_HINTS:
            if hint_role != role:
                continue
            spec = 'f:%s:%s' % (model, fname)
            if spec not in available:
                continue
            for token in tokens:
                token_key = crc.strip_accents(token)
                if key == token_key:
                    return spec, True
                if len(token_key) >= 5 and token_key in key:
                    return spec, False
        return None, False

    @api.model
    def employee_mapping_suggest(self, config_id=None, include_payroll=False):
        """Propose destinations for the unwired people columns of one structure."""
        import difflib
        from odoo.addons.pb_hr_payroll_formula.models import column_role_classifier as crc

        data = self.employee_mapping_data(config_id, include_payroll=include_payroll)
        if not data.get('ok'):
            return data

        config = self._pick_config(config_id)
        # The candidate pool is everything the RIGHT column could hold, not just what
        # it is showing: a suggestion is allowed to introduce a field the user has not
        # searched for, provided the card is appended so the wire has something to
        # land on.
        pool = {}
        for model in ('hr.employee', 'hr.contract'):
            for item in self.ec_model_fields(model).get('fields', []):
                pool[item['id']] = item
        for item in data['right']:
            pool.setdefault(item['id'], item)

        on_board = {i['id'] for i in data['right']}
        taken_right = {w['rightId'] for w in data['wires']}
        wired_left = {w['leftId'] for w in data['wires']}

        # Pre-normalise every candidate label once; this loop is O(left × fields) and
        # `ir.model.fields` on hr.employee alone is several hundred rows.
        norm_labels = {}
        for spec, item in pool.items():
            if not spec.startswith('f:'):
                continue
            norm_labels[spec] = crc.strip_accents(crc.normalize_header(item['label']))

        rules_by_id = {r.id: r for r in config.rule_ids}
        matcher = difflib.SequenceMatcher()
        suggestions = []
        for item in data['left']:
            if item['id'] in wired_left:
                continue
            rule = rules_by_id.get(item['id'])
            if not rule or rule.is_contract_component or rule.is_text_component:
                continue
            role = (rule.column_role or 'payroll')
            if role == 'payroll':
                continue
            header = rule.name or rule.data_source_field or rule.code or ''

            target, confidence, reason = None, 0.0, ''
            if role == 'bank':
                bank_role, exact = self._ec_suggest_bank_role(crc, header)
                if bank_role:
                    target = 'b:%s' % bank_role
                    confidence = 0.95 if exact else 0.75
                    reason = _("The column name says this is the %s.") % (
                        dict(self.env['hr.payslip.import.mapping']
                             ._fields['bank_role'].selection).get(bank_role, bank_role).lower())
            else:
                hint, exact = self._ec_hint_target(crc, role, header, pool)
                if hint:
                    target = hint
                    confidence = 0.95 if exact else 0.75
                    # Named, not scored: the sentence says WHY, and "usually" is the
                    # honest word for a curated table of conventions.
                    reason = _("Columns named like this usually go to \"%s\".") % \
                        pool[hint]['label']
                else:
                    # Nothing curated matches — fall back to comparing the column's
                    # name with every field label, and SHOW the score. Capped at 0.85
                    # so a label coincidence can never reach the accept-all threshold:
                    # that button is for answers somebody could have looked up, not
                    # for the machine's best guess.
                    key = crc.strip_accents(crc.normalize_header(header))
                    if len(key) >= 4:
                        matcher.set_seq2(key)
                        best, best_ratio = None, 0.0
                        for spec, label in norm_labels.items():
                            if not label or len(label) < 3:
                                continue
                            matcher.set_seq1(label)
                            if matcher.real_quick_ratio() < 0.62 or matcher.quick_ratio() < 0.62:
                                continue
                            ratio = matcher.ratio()
                            if ratio > best_ratio:
                                best, best_ratio = spec, ratio
                        if best and best_ratio >= 0.62:
                            target = best
                            confidence = round(min(0.85, best_ratio), 4)
                            reason = _("The column name resembles this field (%d%%).") % \
                                round(best_ratio * 100)

            if not target or target in taken_right:
                continue
            taken_right.add(target)
            if target not in on_board and target in pool:
                data['right'].append(pool[target])
                on_board.add(target)
            suggestions.append({
                'id': 'es%s' % item['id'],
                'kind': 'suggestion', 'ref': False,
                'leftId': item['id'], 'rightId': target,
                'state': 'suggested',
                'confidence': confidence,
                'reason': reason,
            })

        data['wires'] = data['wires'] + suggestions
        return data

    # ------------------------------------------------------------------
    # B1 — Execution replay (step through a payslip's computation)
    # ------------------------------------------------------------------
    @api.model
    def replay_trace(self, config_id=None, sample_id=None):
        """Re-evaluate one sample's inputs and emit an ORDERED trace — one entry
        per formula component in dependency order, each recording the input
        values it read and the value it produced. Generated on demand, never
        persisted (D-B1)."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        by_col = self._col_to_rule(rules)
        by_code = {r.code: r for r in rules if r.code}
        samples = [{'id': s.id, 'name': s.name} for s in config.sample_data_ids]
        sid = int(sample_id) if sample_id else (samples[0]['id'] if samples else False)
        if not sid:
            return {'ok': False, 'reason': 'no_sample', 'samples': samples}
        sample = self.env['hr.formula.sample.data'].browse(sid)
        try:
            inputs = json.loads(sample.input_values_json or '{}')
        except Exception:
            inputs = {}

        # seed results (code-keyed, like the engine) with inputs + constants
        results = dict(inputs)
        seeded = []
        for r in rules:
            if r.column_type == 'constant':
                results[r.code] = r.constant_value or 0.0
            elif r.column_type == 'input' and r.code not in results:
                results[r.code] = r.default_value or 0.0
        for r in rules:
            if r.column_type in ('input', 'constant') and r.column_letter:
                seeded.append({'col': r.column_letter, 'code': r.code or '',
                               'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                               'type': r.column_type,
                               'value': self._as_num(results.get(r.code)),
                               'number_format': r.number_format or 'currency'})

        # formula rules in execution order (dependencies first)
        intel = self.get_intelligence(config.id)
        order_cols = [c for c in intel.get('execution_order', []) if c in by_col]

        steps = []
        for col in order_cols:
            r = by_col.get(col)
            if not r or r.column_type != 'formula':
                continue
            refs = self._expand_refs(r.excel_formula, by_col)
            in_vals = []
            for c in sorted(refs, key=self._col_num):
                rr = by_col.get(c)
                if rr:
                    in_vals.append({'col': c, 'code': rr.code or '',
                                    'name': (rr.salary_rule_id.name if rr.salary_rule_id else False) or rr.name or rr.code,
                                    'value': self._as_num(results.get(rr.code)),
                                    'number_format': rr.number_format or 'currency'})
            try:
                val = r.evaluate(results)
            except Exception:
                val = 0.0
            results[r.code] = val
            steps.append({
                'col': r.column_letter, 'code': r.code or '',
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                'group': _group_for(r),
                'excel_formula': r.excel_formula or '',
                'tokens': self._tokenize(r, by_col),
                'inputs': in_vals,
                'result': self._as_num(val),
                'is_deduction': _group_for(r) == 'Deductions',
                'number_format': r.number_format or 'currency',
            })
        return {
            'ok': True,
            'config': {'id': config.id, 'name': config.name,
                       'currency': config.currency_id.symbol if config.currency_id else '₫'},
            'samples': samples, 'sample_id': sid,
            'seeded': seeded, 'steps': steps,
            'can_edit': self._can_edit(),
        }

    @api.model
    def _as_num(self, v):
        try:
            return round(float(v or 0.0), 4)
        except (TypeError, ValueError):
            return 0.0

    # ------------------------------------------------------------------
    # F9 — Payslip Studio
    # ------------------------------------------------------------------
    _SECTION_COLORS = ['slate', 'indigo', 'emerald', 'amber', 'rose', 'sky', 'violet']
    _PAYSLIP_UPLOAD_MIMES = {'application/pdf', 'image/png', 'image/jpeg'}
    _PAYSLIP_UPLOAD_MAX = 10 * 1024 * 1024
    _PAYSLIP_CONTENT_IMAGE_MIMES = {'image/png', 'image/jpeg', 'image/webp'}
    _PAYSLIP_CONTENT_IMAGE_MAX = 4 * 1024 * 1024
    _PAYSLIP_CONTENT_IMAGE_MARKER = 'pb_payslip_content_image'
    _PAYSLIP_CONTENT_IMAGE_RE = re.compile(
        r'/web/image/ir\.attachment/(\d+)/datas(?:\?[^"\'<>\s]*)?')

    def _payslip_comp(self, r, values):
        """One payslip line payload (value comes from the live preview)."""
        return {
            'id': r.id,
            'col': r.column_letter or '',
            'code': r.code or '',
            'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code or '(unnamed)',
            'group': _group_for(r),
            'type': r.column_type or '',
            'number_format': r.number_format or 'currency',
            'visibility': r.visibility_rule or 'always',
            'payslip_sequence': r.payslip_sequence or 0,
            'is_deduction': _group_for(r) == 'Deductions',
            'value': values.get(r.column_letter),
        }

    @api.model
    def payslip_studio_data(self, config_id=None, sample_id=None):
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        Section = self.env['hr.payslip.config']
        sections = Section.search([('salary_structure_id', '=', config.id)], order='sequence, id')
        samples = [{'id': s.id, 'name': s.name} for s in config.sample_data_ids]
        sid = int(sample_id) if sample_id else (samples[0]['id'] if samples else False)
        values = self._compute(config, sid).get('values', {}) if sid else {}
        values_by_rule = {r.id: values.get(r.column_letter) for r in rules}
        currency = config.currency_id.symbol if config.currency_id else '₫'

        rich_blocks = [config.payslip_header_html or '', config.payslip_footer_html or '',
                       config.payslip_layout_html or '']
        rich_blocks.extend(s.note_html or '' for s in sections)
        embedded_value_ids = set()
        for block in rich_blocks:
            embedded_value_ids.update(
                config._payslip_content_rule_ids(block, amount_only=True))

        payslip_rules = [r for r in rules if r.appears_on_payslip]
        by_sec = defaultdict(list)
        tray = []
        for r in payslip_rules:
            # A value/both token is the visual placement of this component.
            # Suppress its ordinary line until the token is removed, at which
            # point the untouched section assignment makes it reappear.
            if r.id in embedded_value_ids:
                continue
            if r.payslip_identifier:
                by_sec[r.payslip_identifier.id].append(r)
            else:
                tray.append(r)
        sec_payload = []
        for s in sections:
            comps = sorted(by_sec.get(s.id, []),
                           key=lambda r: (r.payslip_sequence or 0, r.sequence))
            embedded_ids = config._payslip_content_rule_ids(
                s.note_html or '', amount_only=True)
            sec_payload.append({
                'id': s.id, 'identifier': s.identifier or '',
                'label': s.label or s.identifier or '', 'label_vi': s.label_vi or '',
                'sequence': s.sequence, 'color_key': s.color_key or 'slate',
                'collapse_when_empty': bool(s.collapse_when_empty),
                'note_html': s.note_html or '',
                'note_rendered_html': config._render_payslip_content(
                    s.note_html or '', values_by_rule, currency),
                'embedded_components': [
                    self._payslip_comp(r, values) for r in rules
                    if r.id in embedded_ids
                ],
                'components': [self._payslip_comp(r, values) for r in comps],
            })
        tray_sorted = sorted(tray, key=lambda r: r.sequence)
        return {
            'ok': True,
            'config': {'id': config.id, 'name': config.name,
                       'currency': currency},
            'sections': sec_payload,
            'tray': [self._payslip_comp(r, values) for r in tray_sorted],
            'rich_components': [self._payslip_comp(r, values) for r in rules],
            'samples': samples, 'sample_id': sid,
            'colors': self._SECTION_COLORS,
            'theme': {
                'accent': config.theme_accent or 'slate',
                'font': config.theme_font or 'system',
                'show_logo': bool(config.theme_show_logo),
                'has_logo': bool(config.theme_logo),
            },
            'accent_hex': self._ACCENT_HEX,
            'header_html': config.payslip_header_html or '',
            'footer_html': config.payslip_footer_html or '',
            'layout_html': config.payslip_layout_html or '',
            'section_template_active': bool(
                sections or config.payslip_header_html or config.payslip_footer_html),
            'header_rendered_html': config._render_payslip_content(
                config.payslip_header_html or '', values_by_rule, currency),
            'footer_rendered_html': config._render_payslip_content(
                config.payslip_footer_html or '', values_by_rule, currency),
            'layout_rendered_html': config._render_payslip_content(
                config.payslip_layout_html or '', values_by_rule, currency, {
                    'employee_name': _('Employee name'),
                    'employee_id': _('Employee ID'),
                    'department': _('Department'),
                    'date_from': _('Period start'),
                    'date_to': _('Period end'),
                    'period': _('Pay period'),
                }),
            'can_edit': self._can_edit(),
        }

    @staticmethod
    def _payslip_import_norm(value):
        value = unicodedata.normalize('NFKD', str(value or ''))
        value = ''.join(ch for ch in value if not unicodedata.combining(ch))
        return re.sub(r'[^a-z0-9]+', ' ', value.lower()).strip()

    @staticmethod
    def _payslip_import_list(value):
        """Turn a provider's free-form/list-ish cell into short clean rows."""
        if isinstance(value, list):
            raw = value
        else:
            text = str(value or '').strip()
            raw = []
            if text.startswith('['):
                try:
                    parsed = json.loads(text)
                    raw = parsed if isinstance(parsed, list) else []
                except (TypeError, ValueError):
                    raw = []
            if not raw:
                raw = re.split(r'[\r\n;]+', text)
        out = []
        for item in raw:
            item = re.sub(r'^\s*(?:[-*•]|\d+[.)])\s*', '', str(item or '')).strip()
            if item and item not in out:
                out.append(item[:180])
        return out[:120]

    @staticmethod
    def _payslip_import_html(value):
        lines = [html.escape(x.strip()) for x in str(value or '').splitlines() if x.strip()]
        return ''.join('<p>%s</p>' % line for line in lines[:12])

    def _payslip_pdf_text_usable(self, value):
        """Reject PDFs whose embedded font map yields one-letter gibberish."""
        tokens = self._payslip_import_norm(value).split()
        words = [token for token in tokens if len(token) >= 3 and not token.isdigit()]
        return bool(len(words) >= 4 and len(words) / max(1, len(tokens)) >= 0.25)

    def _payslip_rule_score(self, label, rule):
        source = self._payslip_import_norm(label)
        if not source:
            return 0.0
        source_variants = {source}
        latin_label = re.sub(r'\([^)]*\)', ' ', str(label or ''))
        source_variants.add(self._payslip_import_norm(latin_label))
        for part in re.split(r'\s[/|]\s', str(label or '')):
            source_variants.add(self._payslip_import_norm(part))
        source_variants.discard('')
        names = {
            self._payslip_import_norm(rule.code),
            self._payslip_import_norm(rule.name),
            self._payslip_import_norm(rule.salary_rule_id.name if rule.salary_rule_id else ''),
        } - {''}
        best = 0.0
        for candidate in source_variants:
            for target in names:
                if candidate == target:
                    return 0.99
                if len(candidate) >= 4 and (candidate in target or target in candidate):
                    best = max(best, 0.90 * min(len(candidate), len(target))
                               / max(len(candidate), len(target)))
                ratio = SequenceMatcher(None, candidate, target).ratio()
                source_tokens, target_tokens = set(candidate.split()), set(target.split())
                overlap = (len(source_tokens & target_tokens) / len(source_tokens | target_tokens)
                           if source_tokens and target_tokens else 0.0)
                best = max(best, ratio * 0.72 + overlap * 0.28)
        return round(best, 3)

    @staticmethod
    def _payslip_pdf_layout_rows(layout):
        """Validate and fold positioned PDF.js text items into visual rows."""
        pages = layout.get('pages') if isinstance(layout, dict) else []
        if not isinstance(pages, list):
            return []
        out = []
        for page_index, page in enumerate(pages[:4]):
            items = page.get('items') if isinstance(page, dict) else []
            if not isinstance(items, list):
                continue
            clean = []
            for item in items[:2500]:
                if not isinstance(item, dict):
                    continue
                text_value = re.sub(r'[\x00-\x1f]+', ' ', str(item.get('text') or '')).strip()
                if not text_value:
                    continue
                try:
                    x = max(0.0, min(1000.0, float(item.get('x') or 0)))
                    y = max(0.0, min(1000.0, float(item.get('y') or 0)))
                    width = max(0.0, min(1000.0, float(item.get('width') or 0)))
                    height = max(1.0, min(100.0, float(item.get('height') or 10)))
                except (TypeError, ValueError):
                    continue
                clean.append({
                    'text': text_value[:240], 'x': x, 'y': y,
                    'width': width, 'height': height,
                    'bold': bool(item.get('bold')), 'italic': bool(item.get('italic')),
                })
            clean.sort(key=lambda item: (item['y'], item['x']))
            rows = []
            for item in clean:
                tolerance = max(4.0, min(11.0, item['height'] * .78))
                row = next((candidate for candidate in reversed(rows[-3:])
                            if candidate['y_min'] - tolerance <= item['y']
                            <= candidate['y_max'] + tolerance), None)
                if row is None:
                    row = {'y': item['y'], 'y_min': item['y'],
                           'y_max': item['y'], 'items': []}
                    rows.append(row)
                row['items'].append(item)
                row['y'] = sum(cell['y'] for cell in row['items']) / len(row['items'])
                row['y_min'] = min(row['y_min'], item['y'])
                row['y_max'] = max(row['y_max'], item['y'])
            for row_index, row in enumerate(rows):
                row['items'].sort(key=lambda item: item['x'])
                row['id'] = 'p%s-r%s' % (page_index + 1, row_index + 1)
                row['page'] = page_index + 1
                row['text'] = ' '.join(item['text'] for item in row['items']).strip()
                if row['text']:
                    out.append(row)
        return out[:500]

    def _payslip_geometry_rows(self, layout):
        """Classify positioned text into headings and payroll rows."""
        rows = self._payslip_pdf_layout_rows(layout)
        current_section = ''
        in_body = False
        for row in rows:
            text_value = row['text']
            norm = self._payslip_import_norm(text_value)
            is_section = bool(re.match(r'^\s*(?:[ivx]+[.)/]|section\s+\d+)',
                                       text_value, re.I))
            if is_section:
                in_body = True
                heading = re.sub(r'\s+hours?\s+amount\s*$', '', text_value,
                                 flags=re.I).strip()
                heading = re.sub(r'\s+(?:hours?|amount)\s*$', '', heading,
                                 flags=re.I).strip()
                current_section = heading[:160]
                row.update({'kind': 'section', 'section': current_section,
                            'label': heading})
                continue
            if not in_body:
                row.update({'kind': 'header', 'section': '', 'label': ''})
                continue
            if ('thank you' in norm or 'cam on' in norm or 'contribution' in norm
                    or norm.startswith('page ')):
                row.update({'kind': 'footer', 'section': '', 'label': text_value})
                continue
            if (current_section and re.fullmatch(r'[\W_]*[A-ZÀ-Ỹ\s]+[\W_]*', text_value)
                    and len(text_value.strip()) < 80):
                previous = next((candidate for candidate in reversed(rows[:rows.index(row)])
                                 if candidate.get('kind') == 'section'), None)
                if previous:
                    previous['label'] = ('%s %s' % (previous['label'], text_value)).strip()
                    previous['text'] = ('%s %s' % (previous['text'], text_value)).strip()
                    current_section = previous['label'][:160]
                    row.update({'kind': 'continuation', 'section': current_section,
                                'label': ''})
                    continue
            label_candidates = sorted(
                [item for item in row['items']
                 if 115 <= item['x'] < 785
                 and not re.fullmatch(r'[-–—+()\d.,%\s]+', item['text'])],
                key=lambda item: (item['y'], item['x']))
            label_lines = []
            for item in label_candidates:
                line = next((candidate for candidate in reversed(label_lines[-2:])
                             if abs(candidate['y'] - item['y']) <= 3.5), None)
                if line is None:
                    line = {'y': item['y'], 'items': []}
                    label_lines.append(line)
                line['items'].append(item)
                line['y'] = sum(cell['y'] for cell in line['items']) / len(line['items'])
            label_items = [item for line in label_lines
                           for item in sorted(line['items'], key=lambda cell: cell['x'])]
            label = ' '.join(item['text'] for item in label_items).strip()
            if not label:
                label = re.sub(r'^\s*\d+(?:[.)]\s*|\s+)', '', text_value)
                label = re.sub(r'\s+[-+()\d.,%]+\s*$', '', label).strip()
            kind = 'total' if re.search(r'\btotal\b', label, re.I) else 'line'
            row.update({'kind': kind, 'section': current_section,
                        'label': label[:240]})
        return rows

    @staticmethod
    def _payslip_import_styled_label(value):
        """Preserve bilingual emphasis without accepting source HTML."""
        escaped = html.escape(str(value or '').strip())
        return re.sub(r'(\([^()]+\))', r'<em>\1</em>', escaped)

    def _payslip_geometry_html(self, rows, layout=None):
        """Rebuild an editable, safe table document from positioned PDF text."""
        body = [row for row in rows if row.get('kind') in ('section', 'line', 'total')]
        if not body:
            return ''
        title_row = next((row for row in rows
                          if row.get('kind') == 'header'
                          and 'pay slip' in self._payslip_import_norm(row['text'])), None)
        title = (' '.join(item['text'] for item in title_row['items']
                          if item['x'] < 480).strip()
                 if title_row else _('PAYSLIP'))
        title = title or _('PAYSLIP')
        palette = (layout.get('palette') or []) if isinstance(layout, dict) else []
        primary = next((str(color).lower() for color in palette[:4]
                        if re.fullmatch(r'#[0-9a-fA-F]{6}', str(color))), '#786e67')
        first_y = body[0]['y']
        footer_rows = [row for row in rows if row.get('kind') == 'footer'
                       and row['y'] > first_y and not row['text'].lower().startswith('page')]

        parts = [
            '<div class="pb-imported-document" style="font-family:Arial,Helvetica,sans-serif;color:#111827">',
            '<table style="width:100%;border-collapse:collapse;border:none;margin:0 0 5px">',
            '<tbody><tr>',
            '<td style="width:50%;border:none;padding:2px 4px;vertical-align:bottom">',
            '<div style="font-size:17px;font-weight:700;font-style:italic;color:%s">%s</div>'
            '<div style="font-size:11px;font-weight:700;font-style:italic">{{pb_meta:period}}</div>'
            '</td>' % (primary, html.escape(title)),
            '<td style="width:50%;border:none;padding:0;vertical-align:bottom">',
            '<table style="width:100%;border-collapse:collapse;border:1px solid #9ca3af;margin:0">',
            '<tbody>',
            '<tr><td style="border:1px solid #d1d5db;padding:3px 5px;font-style:italic">Full name:</td>'
            '<td style="border:1px solid #d1d5db;padding:3px 5px">{{pb_meta:employee_name}}</td></tr>',
            '<tr><td style="border:1px solid #d1d5db;padding:3px 5px;font-style:italic">Employee ID:</td>'
            '<td style="border:1px solid #d1d5db;padding:3px 5px">{{pb_meta:employee_id}}</td></tr>',
            '<tr><td style="border:1px solid #d1d5db;padding:3px 5px;font-style:italic">Department:</td>'
            '<td style="border:1px solid #d1d5db;padding:3px 5px">{{pb_meta:department}}</td></tr>',
            '</tbody></table></td></tr></tbody></table>',
            '<table style="width:100%;table-layout:fixed;border-collapse:collapse;border:1px solid #111827;margin:0">',
            '<colgroup><col style="width:11%"/><col style="width:50%"/>'
            '<col style="width:14%"/><col style="width:25%"/></colgroup><tbody>',
        ]
        for row in body:
            label = self._payslip_import_styled_label(row.get('label') or row['text'])
            if row['kind'] == 'section':
                has_columns = ('hour' in self._payslip_import_norm(row['text'])
                               and 'amount' in self._payslip_import_norm(row['text']))
                parts.append(
                    '<tr><td colspan="%s" style="border:1px solid #111827;padding:4px 6px;'
                    'background-color:#d9d9d9;font-size:12px;font-weight:700">%s</td>%s</tr>' % (
                        2 if has_columns else 4, label,
                        ('<td style="border:1px solid #111827;padding:4px 6px;background-color:#d9d9d9;'
                         'font-style:italic;text-align:right">Hours</td>'
                         '<td style="border:1px solid #111827;padding:4px 6px;background-color:#d9d9d9;'
                         'font-style:italic;text-align:right">Amount</td>') if has_columns else ''))
                continue
            placeholder = '{{pb_import_row:%s}}' % row['id']
            if row['kind'] == 'total':
                parts.append(
                    '<tr><td colspan="3" style="border:1px solid #111827;padding:4px 7px;'
                    'font-weight:700;text-align:right">%s</td>'
                    '<td style="border:1px solid #111827;padding:4px 7px;font-weight:700;'
                    'text-align:right;white-space:nowrap">%s</td></tr>' % (label, placeholder))
                continue
            index_item = next((item['text'] for item in row['items']
                               if item['x'] < 140 and re.fullmatch(r'\d+', item['text'])), '')
            norm_label = self._payslip_import_norm(row.get('label'))
            source_hours = any(580 <= item['x'] < 785
                               and re.search(r'\d', item['text']) for item in row['items'])
            source_amount = any(item['x'] >= 785 and re.search(r'\d', item['text'])
                                for item in row['items'])
            use_hours = source_hours and (not source_amount or 'dependent' in norm_label
                                          or 'hour' in norm_label)
            hours = placeholder if use_hours else ('—' if source_hours else '')
            amount = placeholder if not use_hours else ('—' if source_amount else '')
            parts.append(
                '<tr><td style="border:1px solid #111827;padding:4px 6px;text-align:center">%s</td>'
                '<td style="border:1px solid #111827;padding:4px 7px">%s</td>'
                '<td style="border:1px solid #111827;padding:4px 7px;text-align:right;white-space:nowrap">%s</td>'
                '<td style="border:1px solid #111827;padding:4px 7px;text-align:right;white-space:nowrap">%s</td></tr>' % (
                    html.escape(index_item), label, hours, amount))
        parts.append('</tbody></table>')
        if footer_rows:
            footer = ' '.join(row['text'] for row in footer_rows[:2])
            parts.append('<div style="padding:6px 10px;text-align:center;font-size:11px;'
                         'font-weight:700;font-style:italic">%s</div>' % html.escape(footer))
        parts.append('</div>')
        return ''.join(parts)[:120000]

    @api.model
    def _build_payslip_template_draft(self, config, extracted, filename=''):
        """Deterministically convert OCR output into a reviewable, write-free draft."""
        cells = extracted.get('fields') or {}

        def value(key):
            cell = cells.get(key) or {}
            return cell.get('value', '') if isinstance(cell, dict) else cell

        layout_rows = self._payslip_import_list(value('layout_rows'))
        headings = self._payslip_import_list(value('section_headings'))
        loose_labels = self._payslip_import_list(value('line_labels'))
        geometry_rows = self._payslip_geometry_rows(extracted.get('pdf_layout') or {})
        parsed_rows = []
        if geometry_rows:
            parsed_rows = [
                (row.get('section', '')[:160], row.get('label', '')[:240], row['id'])
                for row in geometry_rows if row.get('kind') in ('line', 'total')
                and row.get('label')
            ]
            headings = [row.get('label', '') for row in geometry_rows
                        if row.get('kind') == 'section']
        else:
            current_section = ''
            for raw in layout_rows:
                parts = re.split(r'\s*(?:::|\|\||\t|\s+-\s+)\s*', raw, maxsplit=1)
                if len(parts) == 2 and parts[0] and parts[1]:
                    section, label = parts
                    current_section = section
                else:
                    section, label = current_section, raw
                parsed_rows.append((section[:80], label[:120], ''))
        if not parsed_rows:
            parsed_rows = [('', label, '') for label in loose_labels]

        # Tesseract/plain OCR returns prose. Only rows that actually match a
        # configured component survive the confidence threshold below.
        if not parsed_rows and extracted.get('raw_text'):
            for raw in self._payslip_import_list(extracted.get('raw_text')):
                clean = re.sub(r'\s+[-+]?\(?[\d.,]+\)?\s*$', '', raw).strip()
                if 2 <= len(clean) <= 120:
                    parsed_rows.append(('', clean, ''))

        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        used = set()
        grouped = {}
        group_order = []
        unmatched = []
        for stated_section, label, layout_row_id in parsed_rows:
            scored = sorted(((self._payslip_rule_score(label, r), r) for r in rules if r.id not in used),
                            key=lambda pair: (-pair[0], pair[1].sequence, pair[1].id))
            score, rule = scored[0] if scored else (0.0, rules.browse())
            if not rule or score < 0.56:
                if label and label not in unmatched:
                    unmatched.append(label)
                review_section = stated_section.strip() or _('Needs review')
                review_key = self._payslip_import_norm(review_section) or 'needs review'
                if review_key not in grouped:
                    grouped[review_key] = {
                        'key': review_key,
                        'label': review_section[:80],
                        'color_key': 'amber',
                        'note_html': '',
                        'matches': [],
                    }
                    group_order.append(review_key)
                grouped[review_key]['matches'].append({
                    'source_label': label,
                    'rule_id': rule.id if rule and score >= 0.30 else False,
                    'rule_name': ((rule.salary_rule_id.name if rule.salary_rule_id else False)
                                  or rule.name or rule.code) if rule else '',
                    'rule_code': (rule.code or '') if rule else '',
                    'confidence': score,
                    'selected': False,
                    'layout_row_id': layout_row_id,
                })
                continue
            used.add(rule.id)
            section = stated_section.strip()
            if not section:
                section = _group_for(rule)
            section_key = self._payslip_import_norm(section) or 'payslip'
            if section_key not in grouped:
                grouped[section_key] = {
                    'key': section_key,
                    'label': section[:80] or _('Payslip'),
                    'color_key': self._SECTION_COLORS[len(group_order) % len(self._SECTION_COLORS)],
                    'note_html': '',
                    'matches': [],
                }
                group_order.append(section_key)
            grouped[section_key]['matches'].append({
                'source_label': label,
                'rule_id': rule.id,
                'rule_name': (rule.salary_rule_id.name if rule.salary_rule_id else False) or rule.name or rule.code,
                'rule_code': rule.code or '',
                'confidence': score,
                'selected': True,
                'layout_row_id': layout_row_id,
            })

        # If the provider found headings but no row carried one, retain only
        # headings with a natural payroll meaning; empty decorative sections do
        # not improve the review draft.
        if not grouped and headings:
            for heading in headings[:8]:
                key = self._payslip_import_norm(heading)
                if key:
                    grouped[key] = {'key': key, 'label': heading[:80],
                                    'color_key': self._SECTION_COLORS[len(group_order) % len(self._SECTION_COLORS)],
                                    'note_html': '', 'matches': []}
                    group_order.append(key)

        accent_words = self._payslip_import_norm(value('accent_colour'))
        accent = next((key for key in self._SECTION_COLORS if key in accent_words), False)
        font_words = self._payslip_import_norm(value('font_style'))
        font = ('serif' if 'serif' in font_words else
                ('mono' if 'mono' in font_words else
                 ('system' if ('system' in font_words or 'sans serif' in font_words) else False)))
        imported_layout = self._payslip_geometry_html(
            geometry_rows, extracted.get('pdf_layout') or {})
        return {
            'ok': bool(grouped or value('header_text') or value('footer_text')
                       or imported_layout),
            'filename': filename[:128],
            'provider': extracted.get('provider') or 'none',
            'warning': extracted.get('error') or '',
            'sections': [grouped[key] for key in group_order],
            'unmatched': unmatched[:40],
            'header_html': self._payslip_import_html(value('header_text')),
            'footer_html': self._payslip_import_html(value('footer_text')),
            'layout_html': imported_layout,
            'layout_preview_html': re.sub(
                r'\{\{pb_import_row:[^}]+\}\}', '—', imported_layout),
            'layout_quality': {
                'tables': 2 if imported_layout else 0,
                'rows': len([row for row in geometry_rows
                             if row.get('kind') in ('line', 'total')]),
                'merged_cells': len([row for row in geometry_rows
                                     if row.get('kind') in ('section', 'total')]),
                'styles': bool(imported_layout),
            },
            'theme': {'accent': accent, 'font': font},
            'options': [{'id': r.id,
                         'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                         'code': r.code or '', 'col': r.column_letter or ''} for r in rules],
            'matched_count': len(used),
        }

    @api.model
    def analyse_payslip_template(self, config_id, upload):
        """Read an uploaded payslip once and return suggestions; retain no file."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        try:
            config = self.env['hr.formula.config'].browse(int(config_id)).exists()
        except (TypeError, ValueError):
            config = self.env['hr.formula.config'].browse()
        if not config:
            return {'ok': False, 'msg': _("Configuration not found.")}
        upload = upload if isinstance(upload, dict) else {}
        mimetype = str(upload.get('mime') or '')
        if mimetype not in self._PAYSLIP_UPLOAD_MIMES:
            return {'ok': False, 'msg': _("Use a PDF, JPG or PNG payslip.")}
        try:
            raw = base64.b64decode(str(upload.get('data') or ''), validate=True)
        except (binascii.Error, ValueError, TypeError):
            return {'ok': False, 'msg': _("The uploaded file could not be read.")}
        if not raw:
            return {'ok': False, 'msg': _("The uploaded file is empty.")}
        if len(raw) > self._PAYSLIP_UPLOAD_MAX:
            return {'ok': False, 'msg': _("The payslip must be 10 MB or smaller.")}

        schema = {
            'doc_kinds': ['payslip'],
            'fields': [
                {'name': 'layout_rows', 'label': 'Layout rows', 'type': 'char',
                 'hint': 'Every visible payroll row, one per line, exactly SECTION :: FIELD LABEL, in visual top-to-bottom order.'},
                {'name': 'section_headings', 'label': 'Section headings', 'type': 'char',
                 'hint': 'All headings in visual order, separated by new lines.'},
                {'name': 'line_labels', 'label': 'Payroll field labels', 'type': 'char',
                 'hint': 'All earnings, deduction and total labels, one per line, without amounts.'},
                {'name': 'header_text', 'label': 'Reusable header text', 'type': 'char',
                 'hint': 'Static heading or explanatory text only; exclude employee-specific values.'},
                {'name': 'footer_text', 'label': 'Reusable footer text', 'type': 'char',
                 'hint': 'Static notes or disclaimers only; exclude employee-specific values.'},
                {'name': 'accent_colour', 'label': 'Dominant accent colour', 'type': 'char'},
                {'name': 'font_style', 'label': 'Font style', 'type': 'char',
                 'hint': 'system/sans-serif, serif, or monospace.'},
            ],
        }
        # Odoo's bundled PDF.js extracts ordinary text PDFs in the browser.
        # Treat it only as untrusted source text (never HTML/code); it is capped
        # client + server and all eventual rich text is escaped/sanitized.
        client_text = (str(upload.get('extracted_text') or '')[:80000]
                       if mimetype == 'application/pdf' else '')
        if self._payslip_pdf_text_usable(client_text):
            extracted = {'fields': {}, 'raw_text': client_text,
                         'provider': 'PDF text', 'error': False}
        else:
            attachment = self.env['ir.attachment'].create({
                'name': str(upload.get('name') or 'payslip')[:128],
                'datas': base64.b64encode(raw),
                'mimetype': mimetype,
                'res_model': 'hr.formula.config',
                'res_id': config.id,
            })
            try:
                extracted = self.env['biz.doc.ocr']._extract(schema, [attachment.id])
            finally:
                attachment.unlink()
        if mimetype == 'application/pdf' and isinstance(upload.get('pdf_layout'), dict):
            # Positioned text is generated locally by the same bundled PDF.js
            # reader as client_text. The strict row validator below caps pages,
            # items, coordinates and text before any HTML is generated.
            extracted['pdf_layout'] = upload['pdf_layout']
        draft = self._build_payslip_template_draft(
            config, extracted or {}, str(upload.get('name') or 'payslip'))
        if not draft['ok']:
            draft['msg'] = draft['warning'] or _(
                "No payroll fields could be recognised. Try a clearer image or configure Document OCR.")
        return draft

    @api.model
    def apply_payslip_template(self, config_id, draft):
        """Apply reviewed matches without deleting any existing layout content."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        try:
            config = self.env['hr.formula.config'].browse(int(config_id)).exists()
        except (TypeError, ValueError):
            config = self.env['hr.formula.config'].browse()
        if not config:
            return {'ok': False, 'msg': _("Configuration not found.")}
        draft = draft if isinstance(draft, dict) else {}
        requested = draft.get('sections') if isinstance(draft.get('sections'), list) else []
        Section = self.env['hr.payslip.config']
        existing = Section.search([('salary_structure_id', '=', config.id)], order='sequence, id')
        section_by_name = {self._payslip_import_norm(s.label or s.identifier): s for s in existing}
        rules_by_id = {r.id: r for r in config.rule_ids}
        used_rules, created, placed = set(), 0, 0
        layout_matches = {}

        for section_index, row in enumerate(requested[:20]):
            if not isinstance(row, dict):
                continue
            matches = row.get('matches') if isinstance(row.get('matches'), list) else []
            selected = []
            for match in matches[:160]:
                if not isinstance(match, dict) or not match.get('selected'):
                    continue
                try:
                    rule_id = int(match.get('rule_id'))
                except (TypeError, ValueError):
                    continue
                if rule_id in rules_by_id and rule_id not in used_rules:
                    selected.append(rules_by_id[rule_id])
                    used_rules.add(rule_id)
                    row_id = str(match.get('layout_row_id') or '')
                    if re.fullmatch(r'p\d+-r\d+', row_id):
                        layout_matches[row_id] = rule_id
            if not selected:
                continue
            label = (re.sub(r'[\r\n\t]+', ' ', str(row.get('label') or _('Payslip')))
                     .strip()[:80] or _('Payslip'))
            key = self._payslip_import_norm(label) or 'payslip'
            section = section_by_name.get(key)
            if not section:
                ident = re.sub(r'[^A-Za-z0-9]', '', label).upper()[:16] or 'SECTION'
                codes = set(existing.mapped('identifier'))
                candidate, suffix = ident, 1
                while candidate in codes:
                    suffix += 1
                    candidate = ('%s%s' % (ident[:max(1, 16 - len(str(suffix)))], suffix))[:16]
                color = row.get('color_key') if row.get('color_key') in self._SECTION_COLORS else self._SECTION_COLORS[section_index % len(self._SECTION_COLORS)]
                section = Section.create({
                    'salary_structure_id': config.id,
                    'identifier': candidate,
                    'label': label,
                    'sequence': max(existing.mapped('sequence') or [0]) + 10,
                    'color_key': color,
                })
                existing |= section
                section_by_name[key] = section
                created += 1
            current = config.rule_ids.filtered(lambda r: r.payslip_identifier == section and r.id not in used_rules)
            ordered = selected + list(current.sorted(key=lambda r: (r.payslip_sequence or 0, r.sequence)))
            for index, rule in enumerate(ordered):
                vals = {'payslip_identifier': section.id, 'payslip_sequence': (index + 1) * 10}
                if rule in selected:
                    vals['appears_on_payslip'] = True
                    placed += 1
                rule.write(vals)

        content_vals = {}
        if draft.get('apply_content'):
            if draft.get('header_html'):
                content_vals['payslip_header_html'] = draft['header_html']
            if draft.get('footer_html'):
                content_vals['payslip_footer_html'] = draft['footer_html']
        if draft.get('apply_layout') and draft.get('layout_html'):
            layout_html = str(draft.get('layout_html') or '')[:120000]

            def replace_layout_row(match):
                rule_id = layout_matches.get(match.group(1))
                return ('{{pb_component:%s:value}}' % rule_id) if rule_id else '—'

            layout_html = re.sub(
                r'\{\{pb_import_row:(p\d+-r\d+)\}\}', replace_layout_row,
                layout_html)
            # Never persist unresolved import instructions or cross-config
            # component ids. Meta markers are a fixed non-executable whitelist.
            layout_html = re.sub(r'\{\{pb_import_row:[^}]+\}\}', '—', layout_html)
            content_vals['payslip_layout_html'] = config._normalise_payslip_content_tokens(
                layout_html)
        if draft.get('apply_theme'):
            theme = draft.get('theme') if isinstance(draft.get('theme'), dict) else {}
            if theme.get('accent') in self._ACCENT_HEX:
                content_vals['theme_accent'] = theme['accent']
            if theme.get('font') in ('system', 'serif', 'mono'):
                content_vals['theme_font'] = theme['font']
        if content_vals:
            config.write(content_vals)
        return {'ok': True, 'created_sections': created, 'placed': placed}

    @staticmethod
    def _payslip_content_image_matches_mime(raw, mimetype):
        """Reject renamed/non-image payloads before creating an attachment."""
        if mimetype == 'image/png':
            return raw.startswith(b'\x89PNG\r\n\x1a\n')
        if mimetype == 'image/jpeg':
            return raw.startswith(b'\xff\xd8\xff')
        if mimetype == 'image/webp':
            return len(raw) >= 12 and raw[:4] == b'RIFF' and raw[8:12] == b'WEBP'
        return False

    @api.model
    def _payslip_content_image_ids(self, config):
        """Attachment ids referenced anywhere in this config's saved template."""
        blocks = [config.payslip_header_html, config.payslip_footer_html,
                  config.payslip_layout_html]
        blocks.extend(self.env['hr.payslip.config'].search([
            ('salary_structure_id', '=', config.id),
        ]).mapped('note_html'))
        return {
            int(match.group(1))
            for block in blocks
            for match in self._PAYSLIP_CONTENT_IMAGE_RE.finditer(str(block or ''))
        }

    @api.model
    def _cleanup_payslip_content_images(self, config, candidate_ids=None):
        """Delete only our own unreferenced inline-image attachments."""
        domain = [
            ('res_model', '=', 'hr.formula.config'),
            ('res_id', '=', config.id),
            ('description', '=', self._PAYSLIP_CONTENT_IMAGE_MARKER),
        ]
        if candidate_ids is not None:
            clean_ids = []
            for value in candidate_ids:
                try:
                    clean_ids.append(int(value))
                except (TypeError, ValueError):
                    continue
            if not clean_ids:
                return 0
            domain.append(('id', 'in', clean_ids))
        referenced = self._payslip_content_image_ids(config)
        orphaned = self.env['ir.attachment'].search(domain).filtered(
            lambda attachment: attachment.id not in referenced)
        count = len(orphaned)
        orphaned.unlink()
        return count

    @api.model
    def upload_payslip_content_image(self, config_id, upload):
        """Store a safe inline payslip image and return a tokenised image URL."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Configuration not found.")}
        upload = upload if isinstance(upload, dict) else {}
        mimetype = str(upload.get('mime') or '').lower()
        if mimetype not in self._PAYSLIP_CONTENT_IMAGE_MIMES:
            return {'ok': False, 'msg': _("Use a PNG, JPEG or WebP image.")}
        try:
            raw = base64.b64decode(str(upload.get('data') or ''), validate=True)
        except (binascii.Error, ValueError, TypeError):
            return {'ok': False, 'msg': _("The image file could not be read.")}
        if not raw:
            return {'ok': False, 'msg': _("The image file is empty.")}
        if len(raw) > self._PAYSLIP_CONTENT_IMAGE_MAX:
            return {'ok': False, 'msg': _("The image must be 4 MB or smaller.")}
        if not self._payslip_content_image_matches_mime(raw, mimetype):
            return {'ok': False, 'msg': _("The file contents do not match the selected image type.")}
        original_name = re.sub(r'[\x00-\x1f\\/]+', ' ', str(upload.get('name') or 'Image')).strip()
        name = (original_name or 'Image')[:128]
        attachment = self.env['ir.attachment'].create({
            'name': name,
            'datas': base64.b64encode(raw),
            'mimetype': mimetype,
            'res_model': 'hr.formula.config',
            'res_id': config.id,
            'description': self._PAYSLIP_CONTENT_IMAGE_MARKER,
        })
        attachment.generate_access_token()
        return {
            'ok': True,
            'id': attachment.id,
            'name': name,
            'url': '/web/image/ir.attachment/%s/datas?access_token=%s' % (
                attachment.id, attachment.access_token),
        }

    @api.model
    def discard_payslip_content_images(self, config_id, attachment_ids):
        """Remove uploads abandoned when the user cancels the editor."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        return {'ok': True, 'removed': self._cleanup_payslip_content_images(
            config, attachment_ids if isinstance(attachment_ids, list) else [])}

    @api.model
    def save_payslip_content(self, config_id, target, html_value):
        """Save sanitized rich content for header/footer/a section."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        try:
            config = self.env['hr.formula.config'].browse(int(config_id)).exists()
        except (TypeError, ValueError):
            config = self.env['hr.formula.config'].browse()
        if not config:
            return {'ok': False}
        html_value = config._normalise_payslip_content_tokens(
            str(html_value or '')[:120000])
        if target == 'header':
            previous_html = config.payslip_header_html
            config.write({'payslip_header_html': html_value or False})
        elif target == 'footer':
            previous_html = config.payslip_footer_html
            config.write({'payslip_footer_html': html_value or False})
        elif target == 'layout':
            previous_html = config.payslip_layout_html
            config.write({'payslip_layout_html': html_value or False})
        elif str(target).startswith('section:'):
            try:
                section_id = int(str(target).split(':', 1)[1])
            except (TypeError, ValueError):
                return {'ok': False, 'msg': _("Unknown content area.")}
            section = self.env['hr.payslip.config'].browse(section_id)
            if not section.exists() or section.salary_structure_id != config:
                return {'ok': False, 'msg': _("Section not found in this configuration.")}
            previous_html = section.note_html
            section.write({'note_html': html_value or False})
        else:
            return {'ok': False, 'msg': _("Unknown content area.")}
        previous_ids = [
            int(match.group(1))
            for match in self._PAYSLIP_CONTENT_IMAGE_RE.finditer(str(previous_html or ''))
        ]
        self._cleanup_payslip_content_images(config, previous_ids)
        return {'ok': True}

    @api.model
    def delete_payslip_template(self, config_id, template_kind):
        """Delete exactly the active presentation mode without touching formulas.

        Imported documents overlay the retained section template, so removing
        one only clears ``payslip_layout_html``. A section template deletion is
        structural: sections and their rich content are removed, component
        assignments return to the tray, and header/footer content is cleared.
        Theme/branding and the payroll rules themselves deliberately survive.
        """
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("Configuration not found.")}
        if template_kind == 'imported':
            config.write({'payslip_layout_html': False})
            return {'ok': True, 'kind': 'imported'}
        if template_kind != 'section':
            return {'ok': False, 'msg': _("Unknown payslip template type.")}
        if config.payslip_layout_html:
            return {'ok': False, 'msg': _(
                "Delete the active imported template before deleting the section template.")}

        Section = self.env['hr.payslip.config']
        sections = Section.search([('salary_structure_id', '=', config.id)])
        assigned = config.rule_ids.filtered(lambda rule: rule.payslip_identifier)
        assigned.write({'payslip_identifier': False})
        section_count = len(sections)
        component_count = len(assigned)
        sections.unlink()
        config.write({
            'payslip_header_html': False,
            'payslip_footer_html': False,
        })
        return {
            'ok': True,
            'kind': 'section',
            'deleted_sections': section_count,
            'returned_components': component_count,
        }

    # W73 — accent palette hex (the LOCKED sc-* keys; mirrors payslip.scss +
    # hr_payslip_formula._THEME_ACCENT_HEX so preview and print never drift).
    _ACCENT_HEX = {
        'slate': '#64748B', 'indigo': '#5A4BB0', 'emerald': '#059669',
        'amber': '#D97706', 'rose': '#E11D48', 'sky': '#0284C7', 'violet': '#7C3AED',
    }

    @api.model
    def save_payslip_theme(self, config_id, vals):
        """W73 (D-L7) — persist payslip theme fields (manager-gated). Only the
        four whitelisted brand tokens are writable; accent/font are validated
        against the LOCKED selections so no free hex/font ever lands."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        vals = vals or {}
        clean = {}
        # Loud reject, never silent coercion (C7): a client sending '#ff0000'
        # has a bug — quietly saving 'slate' would hide it.
        if 'accent' in vals:
            if vals['accent'] not in self._ACCENT_HEX:
                return {'ok': False,
                        'msg': _("Unknown accent %r — the palette is locked.") % vals['accent']}
            clean['theme_accent'] = vals['accent']
        if 'font' in vals:
            if vals['font'] not in ('system', 'serif', 'mono'):
                return {'ok': False,
                        'msg': _("Unknown font %r — choose system, serif or mono.") % vals['font']}
            clean['theme_font'] = vals['font']
        if 'show_logo' in vals:
            clean['theme_show_logo'] = bool(vals['show_logo'])
        if 'logo' in vals:
            # '' / False clears the brand logo (falls back to company logo).
            clean['theme_logo'] = vals['logo'] or False
        if clean:
            config.write(clean)
        return {'ok': True, 'theme': {
            'accent': config.theme_accent or 'slate',
            'font': config.theme_font or 'system',
            'show_logo': bool(config.theme_show_logo),
            'has_logo': bool(config.theme_logo),
        }}

    @api.model
    def move_component(self, rule_id, section_id, ordered_ids):
        """Place a component into a section (or the tray when section_id is falsy)
        and renumber that target's lines from ordered_ids — one RPC covers both a
        cross-section move and a within-section reorder."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        Rule = self.env['hr.formula.rule']
        rule = Rule.browse(int(rule_id))
        if not rule.exists():
            return {'ok': False}
        sec = int(section_id) if section_id else False
        vals = {'payslip_identifier': sec, 'appears_on_payslip': True}
        rule.write(vals)
        # renumber the whole target list so drag order persists deterministically
        for i, rid in enumerate(ordered_ids or []):
            r = Rule.browse(int(rid))
            if r.exists():
                r.write({'payslip_identifier': sec, 'payslip_sequence': (i + 1) * 10})
        return {'ok': True}

    @api.model
    def create_section(self, config_id, label=None):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False}
        Section = self.env['hr.payslip.config']
        existing = Section.search([('salary_structure_id', '=', config.id)])
        base = (label or 'Section').strip()
        ident = re.sub(r'[^A-Za-z0-9]', '', base).upper()[:16] or 'SECTION'
        codes = set(existing.mapped('identifier'))
        code, n = ident, 1
        while code in codes:
            n += 1
            code = '%s%s' % (ident, n)
        seq = (max(existing.mapped('sequence') or [0]) + 10) if existing else 10
        color = self._SECTION_COLORS[len(existing) % len(self._SECTION_COLORS)]
        s = Section.create({'salary_structure_id': config.id, 'identifier': code,
                            'label': base, 'sequence': seq, 'color_key': color})
        return {'ok': True, 'section_id': s.id}

    @api.model
    def update_section(self, section_id, vals):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        s = self.env['hr.payslip.config'].browse(int(section_id))
        if not s.exists():
            return {'ok': False}
        allowed = {k: v for k, v in (vals or {}).items()
                   if k in ('label', 'label_vi', 'color_key', 'collapse_when_empty', 'note_html')}
        if allowed:
            s.write(allowed)
        return {'ok': True}

    @api.model
    def delete_section(self, section_id):
        """Delete a section; its components fall back to the tray (unassigned)."""
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        s = self.env['hr.payslip.config'].browse(int(section_id))
        if not s.exists():
            return {'ok': False}
        self.env['hr.formula.rule'].search([('payslip_identifier', '=', s.id)]).write(
            {'payslip_identifier': False})
        s.unlink()
        return {'ok': True}

    @api.model
    def reorder_sections(self, config_id, ordered_ids):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        Section = self.env['hr.payslip.config']
        for i, sid in enumerate(ordered_ids or []):
            s = Section.browse(int(sid))
            if s.exists():
                s.write({'sequence': (i + 1) * 10})
        return {'ok': True}

    @api.model
    def set_component_visibility(self, rule_id, visibility_rule):
        if not self._can_edit():
            return {'ok': False, 'msg': _("No permission.")}
        if visibility_rule not in ('always', 'when_nonzero', 'never'):
            return {'ok': False}
        r = self.env['hr.formula.rule'].browse(int(rule_id))
        if r.exists():
            r.write({'visibility_rule': visibility_rule})
        return {'ok': True}

    # ------------------------------------------------------------------
    # F15 — Comments & annotations
    # ------------------------------------------------------------------
    def _note_payload(self, n):
        return {
            'id': n.id,
            'body': n.body or '',
            'author': n.author_id.name or '',
            'is_review': bool(n.is_review),
            'resolved': bool(n.resolved),
            'date': fields.Datetime.to_string(n.create_date) if n.create_date else '',
            'resolved_by': n.resolved_by_id.name or '',
            'is_mine': n.author_id.id == self.env.user.id,
        }

    @api.model
    def list_notes(self, rule_id):
        notes = self.env['hr.formula.rule.note'].search([('rule_id', '=', int(rule_id))])
        return {'ok': True,
                'notes': [self._note_payload(n) for n in notes],
                'open_reviews': sum(1 for n in notes if n.is_review and not n.resolved)}

    @api.model
    def post_note(self, rule_id, body, is_review=False):
        if not (body or '').strip():
            return {'ok': False}
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'ok': False}
        self.env['hr.formula.rule.note'].create({
            'rule_id': rule.id, 'body': body.strip(), 'is_review': bool(is_review)})
        return self.list_notes(rule.id)

    @api.model
    def resolve_note(self, note_id):
        n = self.env['hr.formula.rule.note'].browse(int(note_id))
        if n.exists():
            n.action_resolve()
        return {'ok': True}

    @api.model
    def reopen_note(self, note_id):
        n = self.env['hr.formula.rule.note'].browse(int(note_id))
        if n.exists():
            n.action_reopen()
        return {'ok': True}

    @api.model
    def delete_note(self, note_id):
        n = self.env['hr.formula.rule.note'].browse(int(note_id))
        if n.exists() and (n.author_id.id == self.env.user.id or self._can_edit()):
            n.unlink()
        return {'ok': True}

    # ------------------------------------------------------------------
    # F11 — Rate (bracket) tables
    # ------------------------------------------------------------------
    def _rate_table_payload(self, t):
        return {
            'id': t.id,
            'code': t.code or '',
            'name': t.name or '',
            'kind': t.kind or 'progressive',
            'note': t.note or '',
            'brackets': [{'id': b.id, 'lower': b.lower, 'rate': b.rate}
                         for b in t.line_ids.sorted(key=lambda b: b.lower)],
            'used_by': t._dependent_rules().mapped('column_letter'),
        }

    @api.model
    def list_rate_tables(self, config_id):
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'tables': []}
        tables = self.env['hr.formula.rate.table'].search([('config_id', '=', config.id)])
        return {'ok': True, 'tables': [self._rate_table_payload(t) for t in tables]}

    @api.model
    def save_rate_table(self, config_id, payload):
        """Create or update a rate table + its brackets in one call. Brackets are
        replaced wholesale from payload['brackets'] (list of {lower, rate})."""
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'msg': _("No configuration loaded.")}
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        payload = payload or {}
        code = (payload.get('code') or '').strip().upper()
        if not re.match(r'^[A-Z][A-Z0-9]*$', code):
            return {'ok': False, 'msg': _("Table code must be letters and digits only, "
                                          "starting with a letter (no spaces or underscores).")}
        Table = self.env['hr.formula.rate.table']
        tid = payload.get('id')
        table = Table.browse(int(tid)) if tid else Table
        # uniqueness of code within config
        clash = Table.search([('config_id', '=', config.id), ('code', '=', code),
                              ('id', '!=', table.id or 0)], limit=1)
        if clash:
            return {'ok': False, 'msg': _("Another table already uses the code %s.") % code}
        vals = {'code': code, 'name': (payload.get('name') or code).strip(),
                'note': (payload.get('note') or '').strip(), 'config_id': config.id}
        if table:
            table.write(vals)
        else:
            table = Table.create(vals)
        # rebuild brackets
        rows = [b for b in (payload.get('brackets') or [])
                if b.get('rate') not in (None, '') or b.get('lower') not in (None, '')]
        table.line_ids.unlink()
        self.env['hr.formula.rate.bracket'].create([{
            'table_id': table.id,
            'lower': float(b.get('lower') or 0.0),
            'rate': float(b.get('rate') or 0.0),
        } for b in rows])
        return {'ok': True, 'table': self._rate_table_payload(table)}

    @api.model
    def delete_rate_table(self, table_id):
        if not self._can_edit():
            return {'ok': False, 'msg': _("You do not have permission to edit this configuration.")}
        t = self.env['hr.formula.rate.table'].browse(int(table_id))
        used = t._dependent_rules() if t.exists() else self.env['hr.formula.rule']
        if used:
            return {'ok': False, 'msg': _("This table is used by %s formula(s): %s. "
                                          "Remove the BRACKET references first.")
                    % (len(used), ', '.join(used.mapped('column_letter')))}
        if t.exists():
            t.unlink()
        return {'ok': True}

    @api.model
    def eval_bracket(self, table_id, value):
        """Compute this table's progressive value at a sample income, plus the
        compiled Excel — for the editor's live preview."""
        t = self.env['hr.formula.rate.table'].browse(int(table_id))
        if not t.exists():
            return {'ok': False}
        try:
            v = float(value or 0.0)
        except (TypeError, ValueError):
            v = 0.0
        brackets = t.line_ids.sorted(key=lambda b: b.lower)
        result = 0.0
        lowers = [b.lower for b in brackets]
        rates = [b.rate for b in brackets]
        base = 0.0
        for i, b in enumerate(brackets):
            upper = lowers[i + 1] if i + 1 < len(brackets) else None
            if v > b.lower:
                top = v if upper is None else min(v, upper)
                result = base + rates[i] * (top - b.lower)
            base += rates[i] * ((upper - b.lower) if upper is not None else 0.0)
        return {'ok': True, 'value': v, 'result': max(0.0, result),
                'compiled': t.compile_excel('x')}

    @api.model
    def _check_formula(self, config, formula, exclude_id=None):
        """Validate a formula string against a config's columns. -> (ok, message)."""
        from odoo.addons.pb_hr_payroll_formula.formula_engine import FormulaValidator
        cols = {r.column_letter: r.code for r in config.rule_ids
                if r.column_letter and r.id != exclude_id}
        try:
            # F11 — expand BRACKET(code, value) first so the validator sees the
            # compiled nested-IF (BRACKET is not one of its known functions).
            expanded = self.env['hr.formula.rate.table'].expand_brackets(formula or '', config)
            return FormulaValidator().validate_formula(expanded, cols)
        except Exception as e:  # pragma: no cover
            return False, str(e)

    @api.model
    def validate_formula_live(self, config_id, formula, exclude_rule_id=None):
        """Live (unsaved) validation for the editor's preview pill."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'valid': False, 'message': 'No configuration loaded.'}
        ok, msg = self._check_formula(config, formula, exclude_id=int(exclude_rule_id) if exclude_rule_id else None)
        return {'valid': bool(ok), 'message': msg or ''}

    @api.model
    def add_component(self, config_id, vals):
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False}
        vals = vals or {}
        # unique code per config (the model enforces uniqueness; a 2nd plain
        # 'NEW' would otherwise raise) + next free column letter. This used to mint
        # `NEW_1` — an UNDERSCORE, i.e. a code the formula converter cannot resolve,
        # created by the studio's own "add component" button.
        existing_codes = set(config.rule_ids.mapped('code'))
        letters = {r.column_letter for r in config.rule_ids if r.column_letter}
        code = component_code_mod.build_component_code(
            vals.get('code') or vals.get('name') or 'New Component',
            existing_codes=existing_codes, reserved=letters)
        # F111: no explicit letter — create() freezes the next permanent letter
        # (max+1, never reused). sequence lands at the end of the grid...
        Rule = self.env['hr.formula.rule']
        rule = Rule.create({
            'config_id': config.id,
            'name': vals.get('name') or 'New Component',
            'code': code,
            'column_type': vals.get('column_type') or 'formula',
            'excel_formula': vals.get('excel_formula') or '',
            'constant_value': vals.get('constant_value') or 0.0,
            'sequence': (max(config.rule_ids.mapped('sequence') or [0]) + 10),
        })
        # ...then, if the grid is grouped by category, slot it at the end of its
        # own category band rather than the far right (T111.5 / D111.4).
        cat = rule.category_id.id or 0
        siblings = [r for r in config.rule_ids.sorted(key=lambda r: r.sequence) if r.id != rule.id]
        if cat and any((r.category_id.id or 0) == cat for r in siblings):
            last = max(i for i, r in enumerate(siblings) if (r.category_id.id or 0) == cat)
            ordered = siblings[:]
            ordered.insert(last + 1, rule)
            before = {r.id: r.column_letter for r in config.rule_ids}
            for i, r in enumerate(ordered):
                target = (i + 1) * 10
                if r.sequence != target:
                    r.with_context(skip_formula_version=True).sequence = target
            Rule._assert_letters_frozen(config, before)
        return {'ok': True, 'rule_id': rule.id}

    @api.model
    def group_columns_by_category(self, config_id):
        """F111/T111.4 — one batched sequence rewrite that groups every column
        by category. Category order = first appearance (stable); within a
        category the current manual order is preserved (stable sort). Letters
        are frozen, so nothing about computation changes — display only."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False}
        before = {r.id: r.column_letter for r in config.rule_ids}
        ordered = config.rule_ids.sorted(key=lambda r: r.sequence)
        cat_order, seen = {}, 0
        for r in ordered:
            key = r.category_id.id or 0
            if key not in cat_order:
                cat_order[key] = seen
                seen += 1
        grouped = sorted(ordered, key=lambda r: (cat_order[r.category_id.id or 0], r.sequence))
        for i, rule in enumerate(grouped):
            target = (i + 1) * 10
            if rule.sequence != target:
                rule.with_context(skip_formula_version=True).sequence = target
        self.env['hr.formula.rule']._assert_letters_frozen(config, before)
        return self.get_studio_data(config_id)

    @api.model
    def reorder_component(self, config_id, drag_id, before_id=None):
        """F111/T111.3 — move a column so it sits just before `before_id` (or to
        the end when None). Display-only: renumber `sequence`; letters stay
        frozen, so no formula reference is ever re-pointed."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False}
        drag_id = int(drag_id)
        drag = config.rule_ids.filtered(lambda r: r.id == drag_id)
        if not drag:
            return {'ok': False}
        order = [r for r in config.rule_ids.sorted(key=lambda r: r.sequence) if r.id != drag_id]
        idx = len(order)
        if before_id:
            idx = next((i for i, r in enumerate(order) if r.id == int(before_id)), len(order))
        order.insert(idx, drag)
        before = {r.id: r.column_letter for r in config.rule_ids}
        for i, r in enumerate(order):
            target = (i + 1) * 10
            if r.sequence != target:
                r.with_context(skip_formula_version=True).sequence = target
        self.env['hr.formula.rule']._assert_letters_frozen(config, before)
        return {'ok': True}

    @api.model
    def delete_component(self, rule_id):
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if rule.exists():
            rule.unlink()
        return {'ok': True}


    # ------------------------------------------------------------------
    # lifecycle (reuses config methods)
    # ------------------------------------------------------------------
    @api.model
    def _state_result(self, config):
        return {'ok': True, 'state': config.state, 'score': self._score(config),
                'message': config.validation_message or ''}

    @api.model
    def validate(self, config_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        try:
            config.action_validate()
        except Exception as e:
            return {'ok': False, 'state': config.state, 'message': str(e)}
        return self._state_result(config)

    @api.model
    def run_tests(self, config_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        try:
            config.action_run_tests()
        except Exception as e:
            return {'ok': False, 'message': str(e)}
        results = config.test_result_ids
        passed = len(results.filtered(lambda r: r.status == 'passed'))
        return {'ok': True, 'total': len(results), 'passed': passed,
                'failed': len(results) - passed, 'state': config.state, 'score': self._score(config)}

    @api.model
    def advance(self, config_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False}
        st = config.state
        try:
            if st == 'draft':
                config.action_start_testing()
            elif st == 'testing':
                config.action_validate()
            elif st == 'validated':
                config.action_activate()
        except Exception as e:
            return {'ok': False, 'state': config.state, 'message': str(e)}
        return self._state_result(config)

    @api.model
    def set_draft(self, config_id):
        config = self.env['hr.formula.config'].browse(int(config_id))
        if config.exists():
            config.action_set_draft()
        return self._state_result(config)

    # ------------------------------------------------------------------
    # config settings surface (bespoke OWL editor for hr.formula.config)
    # ------------------------------------------------------------------
    # Trimmed to fields that actually drive behavior. Dropped from the UI (and
    # thus this payload): mid_cycle_* and the grid-display set + description are
    # vestigial / only used by the legacy excel_grid_widget, never the cockpit.
    _CFG_FIELDS = (
        'name', 'code', 'country_code', 'structure_id', 'cycle_type', 'connector_id',
        'use_color_coded_excel_import', 'export_identity_columns',
        'payroll_journal_id', 'debit_account_id',
        'credit_account_id', 'company_id',
        'use_proration', 'proration_basis', 'proration_component_ids', 'proration_rounding',
        'use_auto_retro', 'retro_component_id',
        # SC-3/SC-4 — the Sources card: which lanes feed this scheme, and in
        # what order. The generic get/save loop carries them for free.
        'source_api_enabled', 'source_excel_enabled', 'source_records_enabled',
        'source_priority',
    )
    _CFG_M2O = ('structure_id', 'connector_id', 'payroll_journal_id', 'debit_account_id',
                'credit_account_id', 'company_id', 'retro_component_id')
    _CFG_M2M = ('proration_component_ids',)

    @api.model
    def _config_status(self, c):
        return {
            'state': c.state,
            'validation_status': c.validation_status or 'pending',
            'last_validated': c.last_validated and str(c.last_validated) or '',
            'last_validated_by': c.last_validated_by.name or '',
            'currency': {'symbol': c.currency_id.symbol or '', 'name': c.currency_id.name or ''},
            'score': self._score(c),
            'has_errors': bool(c.has_errors),
            'error_details': c.error_details or '',
            'has_circular_refs': bool(c.has_circular_refs),
            'circular_ref_details': c.circular_ref_details or '',
            'validation_message': c.validation_message or '',
            'rule_count': c.rule_count,
            'formula_rule_count': c.formula_rule_count,
            'input_rule_count': c.input_rule_count,
            'sample_count': c.sample_count,
            'proration_count': c.proration_count,
            'retro_count': c.retro_count,
            'carryover_count': c.carryover_count,
        }

    @api.model
    def _config_meta(self, c):
        Cfg = self.env['hr.formula.config']

        def _sel(field):
            return [{'value': v, 'label': l} for v, l in Cfg._fields[field].selection]

        comp = lambda model, **kw: [{'id': r.id, 'name': r.display_name}
                                    for r in self.env[model].search([], **kw)]
        connectors = (comp('hr.integration.connector', order='name')
                      if 'hr.integration.connector' in self.env else [])
        accts = self.env['account.account'].search(
            [('company_ids', 'in', c.company_id.ids)] if 'company_ids' in self.env['account.account']._fields
            else [], order='code', limit=400)
        journals = self.env['account.journal'].search(
            [('type', '=', 'general'), ('company_id', '=', c.company_id.id)], order='name')
        return {
            'structures': comp('hr.payroll.structure', order='name'),
            'connectors': connectors,
            'journals': [{'id': j.id, 'name': j.name} for j in journals],
            'accounts': [{'id': a.id, 'name': '%s %s' % (a.code or '', a.name or '')} for a in accts],
            'companies': comp('res.company', order='name'),
            'components': [{'id': r.id, 'col': r.column_letter or '?', 'code': r.code or '', 'name': r.name or ''}
                           for r in c.rule_ids.sorted(key=lambda r: r.sequence)],
            'country_codes': _sel('country_code'),
            'cycle_types': _sel('cycle_type'),
            'proration_bases': _sel('proration_basis'),
            'multi_company': len(self.env['res.company'].search([])) > 1,
            # SC-4 — the Sources card's consequence chips: how many components
            # currently take values from each lane, so switching one off can
            # say what falls through instead of asking for blind faith.
            'source_lane_counts': self._source_lane_counts(c),
        }

    @api.model
    def _source_lane_counts(self, c):
        """`{'api': n, 'excel': n, 'records': n}` — components fed per lane."""
        counts = {'api': 0, 'excel': 0, 'records': 0}
        try:
            for rule in c.rule_ids:
                kinds = {s.kind for s in rule.source_ids
                         if (s.key or '').strip()}
                if kinds & {'feed', 'rule'}:
                    counts['api'] += 1
                if 'excel' in kinds:
                    counts['excel'] += 1
            FM = self.env.get('hr.integration.field.mapping')
            if FM is not None:
                counts['api'] = max(counts['api'], FM.sudo().search_count([
                    ('target_rule_id.config_id', '=', c.id),
                    ('active_state', '=', 'active')]))
            Mapping = self.env.get('hr.payslip.import.mapping')
            records = len(c.rule_ids.filtered('is_contract_component'))
            if Mapping is not None:
                records += Mapping.sudo().search_count([
                    ('salary_structure_id', '=', c.id)])
            counts['records'] = records
        except Exception:       # noqa: BLE001 — chips, never the dialog
            pass
        return counts

    @api.model
    def get_config_settings(self, config_id):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        values = {}
        for f in self._CFG_FIELDS:
            if f in self._CFG_M2O:
                values[f] = c[f].id or False
            elif f in self._CFG_M2M:
                values[f] = c[f].ids
            else:
                values[f] = c[f] if c[f] is not False else False
        samples = [{'id': s.id, 'name': s.name, 'source_type': s.source_type or '',
                    'status': s.validation_status or '', 'discrepancy_count': s.discrepancy_count,
                    'last_computed': s.last_computed and str(s.last_computed) or ''}
                   for s in c.sample_data_ids]
        results = [{'sample': r.sample_id.name or '', 'rule_code': r.rule_code or '',
                    'expected': r.expected_value, 'computed': r.computed_value,
                    'difference': r.difference, 'status': r.status or '',
                    'error': r.error_message or ''}
                   for r in c.test_result_ids]
        return {
            'ok': True,
            'values': values,
            'status': self._config_status(c),
            'meta': self._config_meta(c),
            'samples': samples,
            'results': results,
        }

    @api.model
    def save_config_settings(self, config_id, vals):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        vals = vals or {}
        write_vals = {}
        for k in self._CFG_FIELDS:
            if k not in vals:
                continue
            v = vals[k]
            if k in self._CFG_M2O:
                write_vals[k] = int(v) if v else False
            elif k in self._CFG_M2M:
                write_vals[k] = [(6, 0, [int(x) for x in (v or [])])]
            else:
                write_vals[k] = v
        try:
            if write_vals:
                c.write(write_vals)
        except Exception as e:
            return {'ok': False, 'msg': str(e).splitlines()[0] if str(e) else 'Could not save.'}
        return {'ok': True, 'status': self._config_status(c)}

    @api.model
    def _cfg_run(self, config_id, method):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        notif = ''
        try:
            res = getattr(c, method)()
            if isinstance(res, dict) and res.get('params'):
                notif = res['params'].get('message') or ''
        except Exception as e:
            return {'ok': False, 'msg': str(e).splitlines()[0] if str(e) else 'Action failed.',
                    'status': self._config_status(c)}
        return {'ok': True, 'notif': notif, 'status': self._config_status(c)}

    @api.model
    def cfg_start_testing(self, config_id):
        return self._cfg_run(config_id, 'action_start_testing')

    @api.model
    def cfg_validate(self, config_id):
        return self._cfg_run(config_id, 'action_validate')

    @api.model
    def cfg_activate(self, config_id):
        return self._cfg_run(config_id, 'action_activate')

    @api.model
    def cfg_set_draft(self, config_id):
        return self._cfg_run(config_id, 'action_set_draft')

    @api.model
    def cfg_archive(self, config_id):
        return self._cfg_run(config_id, 'action_archive')

    @api.model
    def cfg_regenerate_formulas(self, config_id):
        return self._cfg_run(config_id, 'action_regenerate_formulas')

    @api.model
    def cfg_generate_sample_data(self, config_id):
        """One-click synthetic sample: build *realistic* random inputs (same
        generator as the Test "Generate" button — NOT the rules' zero defaults),
        compute current outputs, store them as an expected baseline so the Live
        Preview AND Run Tests are immediately meaningful (no wizard dialog)."""
        from odoo.addons.pb_hr_payroll_formula.formula_engine import FormulaEvaluator
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        rules = c.rule_ids.sorted(key=lambda r: r.sequence)
        wiz = self.env['hr.formula.sample.data.wizard'].create({
            'config_id': c.id, 'source': 'random',
            'sample_count': 1, 'min_salary': 5000000.0, 'max_salary': 50000000.0,
        })
        inputs = {}
        for r in rules:
            if r.column_type == 'input' and r.code:
                inputs[r.code] = wiz._generate_random_value(r)
        try:
            computed = FormulaEvaluator().evaluate_all(rules, inputs)
            expected = {code: v for code, v in computed.items()}
        except Exception as e:
            _logger.warning("Sample generate compute failed: %s", e)
            expected = {}
        n = len(c.sample_data_ids) + 1
        try:
            self.env['hr.formula.sample.data'].create({
                'config_id': c.id,
                'name': 'Sample %s' % n,
                'source_type': 'manual',
                'input_values_json': json.dumps(inputs),
                'expected_values_json': json.dumps(expected),
            })
        except Exception as e:
            return {'ok': False, 'msg': str(e).splitlines()[0] if str(e) else 'Could not create sample.'}
        return {'ok': True, 'notif': 'Sample data generated',
                'settings': self.get_config_settings(config_id)}

    @api.model
    def cfg_run_tests(self, config_id):
        r = self._cfg_run(config_id, 'action_run_tests')
        if r.get('ok'):
            r['settings'] = self.get_config_settings(config_id)
        return r

    @api.model
    def cfg_import_excel(self, config_id):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        action = c.action_import_from_excel_multisheet()
        # COLROLES P4 — mark the import as having come FROM the studio, so its
        # completion can hand the user back to the people-mapping board instead of
        # dropping them on a form they did not ask for.
        action.setdefault('context', {})['pbfs_studio_import'] = True
        return {'ok': True, 'action': action}

    # ------------------------------------------------------------------
    # Test & Validate workbench
    # ------------------------------------------------------------------
    def _sample_verdict(self, s):
        """Preview when no expected; pending when the baseline is unconfirmed
        (W84 — matches run_sample_tests); else the model's validation_status."""
        if not (s.expected_values_json and s.expected_values_json not in ('{}', '')):
            return 'preview'
        if not s.expected_confirmed:
            return 'pending'
        return s.validation_status or 'pending'

    def _sample_row(self, s):
        return {
            'id': s.id, 'name': s.name or '(unnamed)',
            'source_type': s.source_type or 'manual',
            'verdict': self._sample_verdict(s),
            'has_expected': bool(s.expected_values_json and s.expected_values_json not in ('{}', '')),
            'expected_confirmed': bool(s.expected_confirmed),
            'discrepancy_count': s.discrepancy_count,
            'last_computed': s.last_computed and str(s.last_computed) or '',
        }

    @api.model
    def get_test_data(self, config_id):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        rules = c.rule_ids.sorted(key=lambda r: r.sequence)
        inputs = [{'code': r.code, 'col': r.column_letter or '?', 'name': r.name or '',
                   'default': r.default_value or 0.0}
                  for r in rules if r.column_type == 'input']
        return {
            'ok': True,
            'samples': [self._sample_row(s) for s in c.sample_data_ids],
            'input_components': inputs,
            'currency': c.currency_id.symbol if c.currency_id else '',
        }

    # ==================================================================
    # W98 (WP-H) — offer calculator: evaluate hypothetical inputs through the
    # LIVE config with ZERO records created. Read-only; open to everyone (D-H6).
    # ==================================================================
    @api.model
    def offer_calc(self, config_id, inputs):
        """Evaluate a hypothetical employee's inputs through the live config and
        return the full component breakdown — no records created (D-H4, in-memory
        ``Sample.new`` on the SAME evaluator previews/tests use, C5). Reports
        headline NET + per-group subtotals only; NEVER a fabricated employer cost
        (D-H4/C7). Validates inputs like W49: known input codes, numeric,
        |v| <= 1e12."""
        config = self.env['hr.formula.config'].browse(int(config_id))
        try:
            if not config.exists():
                return {'ok': False, 'msg': _('Configuration not found.')}
            rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        except AccessError:
            # cross-company probe: fail closed AND gracefully (review minor) —
            # the record rule already blocked the read before any sudo work.
            return {'ok': False, 'msg': _('No access to this configuration.')}
        input_codes = {r.code for r in rules if r.column_type == 'input' and r.code}
        clean = {}
        for code, v in (inputs or {}).items():
            code = str(code)
            if code not in input_codes:
                return {'ok': False, 'msg': _('Unknown input: %s') % code}
            n = self._as_num(v)
            if n is None:
                # _as_num refuses BOTH text and NaN/Inf. Only genuine text may
                # pass through — a value that PARSES numeric but was refused
                # (e.g. "1e400" -> inf) is out of range, not a department name.
                looks_numeric = isinstance(v, (int, float))
                if not looks_numeric and isinstance(v, str):
                    try:
                        float(v.strip())
                        looks_numeric = True
                    except (TypeError, ValueError):
                        pass
                if looks_numeric:
                    return {'ok': False, 'msg': _('Input %s is out of range.') % code}
                # Text input column (e.g. an employee name / department the config
                # carries as an input) — pass it through to the evaluator as-is.
                # Only NUMERIC inputs are range-checked; text inputs are legitimate
                # and must not reject the whole offer (C7 — degrade visibly).
                clean[code] = v
            elif abs(n) > 1e12:
                return {'ok': False, 'msg': _('Input %s is out of range.') % code}
            else:
                clean[code] = n
        # In-memory evaluation — zero rows. Evaluate under sudo so the engine's
        # eval-diagnostic writes on the rules (the same side-effect every preview
        # performs) succeed for READ-ONLY users too (D-H6 — offer calc is a
        # calculator open to everyone; it creates no data, only the internal
        # diagnostic bookkeeping the eval path already does). This avoids touching
        # the eval path itself (D-H7) while honoring D-H6's read-only guarantee.
        sample = self.env['hr.formula.sample.data'].sudo().new({'config_id': config.id})
        try:
            values = sample._evaluate_rules_with_dependencies(clean)
        except Exception as e:
            _logger.warning("offer_calc eval failed: %s", e)
            return {'ok': False, 'msg': _('Could not evaluate this offer.')}

        def _num(x):
            try:
                return float(x)
            except (TypeError, ValueError):
                return 0.0

        rows = []
        subtotals = {}
        for r in rules:
            if not r.code:
                continue
            grp = _group_for(r)
            val = _num(values.get(r.code, 0.0))
            rows.append({
                'col': r.column_letter or '?', 'code': r.code,
                'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code,
                'group': grp, 'type': r.column_type, 'value': round(val, 2),
                'appears_on_payslip': bool(r.appears_on_payslip),
                'number_format': r.number_format or 'number',
            })
            if grp != 'Inputs':                    # Inputs excluded from subtotals (D-H4)
                subtotals[grp] = subtotals.get(grp, 0.0) + val

        # headline net — same heuristic as the comparison (_pick_headline_code)
        net_code = self.env['hr.formula.period.comparison'].new(
            {'config_id': config.id})._pick_headline_code() or ''
        net_value = round(_num(values.get(net_code, 0.0)), 2) if net_code else 0.0

        order = ['Earnings', 'Deductions', 'Totals']
        sub_list = [{'group': g, 'value': round(subtotals[g], 2)}
                    for g in order if g in subtotals]
        for g, v in subtotals.items():
            if g not in order:
                sub_list.append({'group': g, 'value': round(v, 2)})

        return {
            'ok': True,
            'config': {'id': config.id, 'name': config.display_name},
            'currency': config.currency_id.symbol if config.currency_id else '',
            'rows': rows,
            'net_code': net_code, 'net_value': net_value,
            'subtotals': sub_list,
        }

    @api.model
    def offer_sample_inputs(self, sample_id):
        """The input values of an existing sample — copied into the offer form by
        the "start from sample" picker (D-H5). Read of stored JSON only."""
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False, 'inputs': {}}
        try:
            # samples carry no company/record rule of their own — gate through
            # the parent config's rule instead (review minor: no cross-company
            # sample reads through this new RPC).
            s.config_id.check_access('read')
        except AccessError:
            return {'ok': False, 'inputs': {}}
        try:
            vals = json.loads(s.input_values_json or '{}')
        except Exception:
            vals = {}
        return {'ok': True, 'inputs': vals}

    # ------------------------------------------------------------------
    # W83 — test coverage (deterministic, three-valued; NEVER evaluates a
    # formula — pure metadata over the dependency graph + sample JSONs).
    # ------------------------------------------------------------------
    def _coverage_info(self, r):
        return {'rule_id': r.id, 'col': r.column_letter or '',
                'code': r.code or '', 'name': r.name or '(unnamed)'}

    @api.model
    def get_test_coverage(self, config_id=None):
        """Which formula components the samples DO / DON'T exercise (D-G1/D-G2).

        Three-valued per formula component:

        * **asserted** — >=1 active CONFIRMED sample carries a non-null expected
          value for its code (the W82 testable rule, formula_config_tests.py:95,
          plus the D-G3 confirmation gate — unconfirmed baselines don't count).
        * **exercised** — not asserted, but on the upstream dependency closure of
          an asserted component (its value feeds an assertion), via the
          ``_normalized_dep_cols`` edges — the same graph get_intelligence walks.
        * **untested** — neither.

        Coverage ``pct`` = asserted / formula-components. Inputs/constants are
        excluded from the % but any that no asserted formula transitively reads
        are returned as ``orphan_inputs``. This method reads only stored JSON +
        dependency metadata — it NEVER calls ``_compute_results``.
        """
        config = self._pick_config(config_id)
        if not config:
            return {'ok': False, 'pct': 0, 'formula_total': 0,
                    'asserted': [], 'exercised': [], 'untested': [],
                    'orphan_inputs': []}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        by_col = {r.column_letter: r for r in rules if r.column_letter}
        deps = self._normalized_dep_cols(rules)   # {rule.id: {source_col, ...}}
        formula_rules = [r for r in rules
                         if r.column_type == 'formula' and r.column_letter]

        # asserted codes: any active sample with a non-null expected for that
        # code AND a CONFIRMED baseline — an unconfirmed generated sample is a
        # hypothesis (D-G3) and must not raise coverage while the chip says
        # pending (review finding, WP-G).
        asserted_codes = set()
        for s in config.sample_data_ids:
            if not s.expected_confirmed:
                continue
            try:
                exp = json.loads(s.expected_values_json or '{}')
            except Exception:
                exp = {}
            if isinstance(exp, dict):
                for code, v in exp.items():
                    if v is not None:
                        asserted_codes.add(code)

        asserted = [r for r in formula_rules if r.code and r.code in asserted_codes]
        asserted_cols = {r.column_letter for r in asserted}

        # upstream closure: walk each asserted rule's dependency SOURCES,
        # transitively, over the same edges the evaluator resolves.
        closure = set()
        stack = list(asserted_cols)
        while stack:
            col = stack.pop()
            if col in closure:
                continue
            closure.add(col)
            rr = by_col.get(col)
            if rr:
                for d in deps.get(rr.id, ()):  # empty for input/constant rules
                    if d not in closure:
                        stack.append(d)

        exercised = [r for r in formula_rules
                     if r.column_letter not in asserted_cols
                     and r.column_letter in closure]
        untested = [r for r in formula_rules
                    if r.column_letter not in asserted_cols
                    and r.column_letter not in closure]
        orphan = [r for r in rules
                  if r.column_type in ('input', 'constant') and r.column_letter
                  and r.column_letter not in closure]

        n = len(formula_rules)
        pct = int(round(len(asserted) / n * 100)) if n else 0
        return {
            'ok': True, 'pct': pct, 'formula_total': n,
            'asserted': [self._coverage_info(r) for r in asserted],
            'exercised': [self._coverage_info(r) for r in exercised],
            'untested': [self._coverage_info(r) for r in untested],
            'orphan_inputs': [self._coverage_info(r) for r in orphan],
        }

    @api.model
    def get_sample_detail(self, sample_id):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False}
        return {
            'ok': True,
            'id': s.id, 'name': s.name or '',
            'source_type': s.source_type or 'manual',
            'verdict': self._sample_verdict(s),
            'has_expected': bool(s.expected_values_json and s.expected_values_json not in ('{}', '')),
            'expected_confirmed': bool(s.expected_confirmed),
            'rows': s.get_comparison_data(),
        }

    @api.model
    def save_sample_inputs(self, sample_id, inputs):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False}
        vals = json.loads(s.input_values_json or '{}')
        for k, v in (inputs or {}).items():
            try:
                vals[k] = float(v)
            except (TypeError, ValueError):
                vals[k] = v
        s.input_values_json = json.dumps(vals)  # triggers _compute_results
        return self.get_sample_detail(s.id)

    @api.model
    def rename_sample(self, sample_id, name):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if s.exists() and name:
            s.name = name
        return {'ok': True}

    @api.model
    def add_manual_sample(self, config_id):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        inputs = {r.code: (r.default_value or 0.0)
                  for r in c.rule_ids if r.column_type == 'input' and r.code}
        n = len(c.sample_data_ids) + 1
        s = self.env['hr.formula.sample.data'].create({
            'config_id': c.id, 'name': 'Sample %s' % n, 'source_type': 'manual',
            'input_values_json': json.dumps(inputs),
        })
        return {'ok': True, 'sample_id': s.id, 'samples': [self._sample_row(x) for x in c.sample_data_ids]}

    @api.model
    def generate_random_samples(self, config_id, count=3, min_salary=5000000, max_salary=50000000):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        wiz = self.env['hr.formula.sample.data.wizard'].create({
            'config_id': c.id, 'source': 'random',
            'sample_count': int(count or 3),
            'min_salary': float(min_salary or 5000000),
            'max_salary': float(max_salary or 50000000),
        })
        try:
            for vals in wiz._generate_random():
                self.env['hr.formula.sample.data'].create(vals)
        except Exception as e:
            return {'ok': False, 'msg': str(e).splitlines()[0] if str(e) else 'Could not generate.'}
        return {'ok': True, 'samples': [self._sample_row(x) for x in c.sample_data_ids]}

    @api.model
    def export_test_template(self, config_id):
        """Build a blank .xlsx whose header row is the config's input component
        names (in column order). This is the exact format accepted by
        import_test_samples — fill rows under the headers and re-import."""
        import base64
        import io
        try:
            import openpyxl
        except Exception:
            return {'ok': False, 'msg': 'openpyxl is not available on the server.'}
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        inputs = [r for r in c.rule_ids.sorted(key=lambda r: r.sequence)
                  if r.column_type == 'input']
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Test Inputs'
        headers = [(r.name or r.code or r.column_letter or '') for r in inputs]
        ws.append(headers)
        for col_idx, _h in enumerate(headers, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        code = (c.code or c.name or 'config').strip().replace(' ', '_')
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': '%s_test_template.xlsx' % code,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }

    # openpyxl number formats (S-L1 gotcha): openpyxl format strings, NOT Odoo's.
    # VND currency has no minor units ⇒ '#,##0'; percentage values are stored as
    # FRACTIONS (0.05) so Excel's '0.00%' displays 5.00% correctly — never
    # pre-multiply by 100.
    _XLSX_NUMFMT = {
        'currency': '#,##0',
        'integer': '#,##0',
        'percentage': '0.00%',
        'number': '#,##0.00',
    }

    @api.model
    def export_living_workbook(self, config_id):
        """W41 — a config becomes a *living* ``.xlsx`` (D-L1/D-L2/D-L3).

        Sheet 1 "Payroll": xlsx column position = the component's frozen
        ``column_letter`` (1:1 — this is what makes the stored ``=A2+AB2``
        formulas real, Excel-evaluable formulas). Row 1 = localized component
        name, row 2 = code (a second header row, machine-matchable), data rows
        from row 3 = one per sample. Input cells carry the sample's input value
        (else ``default_value``); constant cells carry ``constant_value``;
        formula cells carry the REAL formula — ``BRACKET(...)`` expanded out via
        ``expand_brackets`` (Excel has no BRACKET) and the row digits shifted
        2 → the data row (S-L1). A trailing "Sample" meta column follows the last
        component letter (a leading column would break the 1:1 letter mapping).
        Sheet 2 "Rate Tables" renders each table + a named range per table.
        Read-only; read access suffices (no manager gate, D-L3)."""
        import base64
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.workbook.defined_name import DefinedName
        except Exception:
            return {'ok': False, 'msg': 'openpyxl is not available on the server.'}
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        RateTable = self.env['hr.formula.rate.table']

        rules = c.rule_ids.sorted(key=lambda r: r.sequence)
        placed = [r for r in rules if r.column_letter]
        if not placed:
            return {'ok': False, 'msg': 'This configuration has no components to export.'}
        # xlsx column index = the frozen letter's ordinal (D-L1: 1:1, never by
        # sequence — a reordered config keeps letters as identities, F111).
        col_of = {r.id: self._col_num(r.column_letter) for r in placed}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payroll'
        head_fill = PatternFill('solid', fgColor='EEF0FB')
        head_font = Font(bold=True, color='241F52')
        code_font = Font(italic=True, color='6B7280', size=9)
        right = Alignment(horizontal='right')

        # ---- two header rows (name / code) + meta header ------------------
        for r in placed:
            col = col_of[r.id]
            name = (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code or ''
            h1 = ws.cell(row=1, column=col, value=name)
            h1.fill = head_fill
            h1.font = head_font
            h2 = ws.cell(row=2, column=col, value=r.code or '')
            h2.fill = head_fill
            h2.font = code_font
            ws.column_dimensions[get_column_letter(col)].width = 16
        # Sample names + notes live on the Info SHEET, never on Payroll — a
        # trailing meta column re-imports as a phantom SAMPLE component and the
        # 0-sample note as a phantom header (WP-L review Minor 3).

        # ---- data rows: one per sample (row 3 = first sample) -------------
        samples = list(c.sample_data_ids)
        note = ''
        if not samples:
            # C7 — a 0-sample config still exports a usable, LOUD row of defaults.
            note = 'No samples configured — one row of Default Values shown.'

        def _emit_row(sheet_row, input_by_code, sample_name):
            for r in placed:
                col = col_of[r.id]
                cell = ws.cell(row=sheet_row, column=col)
                if r.column_type == 'formula':
                    expanded = RateTable.expand_brackets(r.excel_formula or '', c)
                    text = (expanded or '').strip()
                    if text:
                        if not text.startswith('='):
                            text = '=' + text
                        cell.value = self._shift_rows(text, sheet_row)
                    else:
                        cell.value = 0
                elif r.column_type == 'constant':
                    cell.value = r.constant_value or 0.0
                else:  # input
                    raw = input_by_code.get(r.code) if r.code else None
                    if raw is None or raw == '':
                        raw = r.default_value or 0.0
                    cell.value = raw
                cell.number_format = self._XLSX_NUMFMT.get(r.number_format or 'currency', '#,##0.00')
                if r.column_type != 'formula':
                    cell.alignment = right
            row_names.append((sheet_row, sample_name))

        row_names = []
        if samples:
            for i, s in enumerate(samples):
                _emit_row(3 + i, s.get_input_values(), s.name or ('Sample %d' % (i + 1)))
        else:
            _emit_row(3, {}, '(defaults — no samples)')

        ws.freeze_panes = 'A3'                      # header + code rows frozen

        # ---- Info sheet: sample names per data row + loud notes ------------
        info = wb.create_sheet('Info')
        info.cell(row=1, column=1, value='Payroll row').font = head_font
        info.cell(row=1, column=2, value='Sample').font = head_font
        info.column_dimensions['A'].width = 12
        info.column_dimensions['B'].width = 30
        for i, (sheet_row, sample_name) in enumerate(row_names):
            info.cell(row=2 + i, column=1, value=sheet_row)
            info.cell(row=2 + i, column=2, value=sample_name)
        if note:
            info.cell(row=len(row_names) + 3, column=1, value=note).font = Font(
                bold=True, color='B45309')

        # ---- Sheet 2: Rate Tables (reference) + named ranges (D-L2) --------
        tables = [t for t in c.rate_table_ids if t.code]
        if tables:
            rs = wb.create_sheet('Rate Tables')
            rs.cell(row=1, column=1, value='Code').font = head_font
            rs.cell(row=1, column=2, value='Name').font = head_font
            rs.cell(row=1, column=3, value='From').font = head_font
            rs.cell(row=1, column=4, value='Rate').font = head_font
            for cc in ('A', 'B', 'C', 'D'):
                rs.column_dimensions[cc].width = 18
            row = 2
            for t in tables:
                brackets = t.line_ids.sorted(key=lambda b: b.lower)
                first_row = row
                for b in brackets:
                    rs.cell(row=row, column=1, value=t.code)
                    rs.cell(row=row, column=2, value=t.name or '')
                    fc = rs.cell(row=row, column=3, value=b.lower or 0.0)
                    fc.number_format = '#,##0'
                    rc = rs.cell(row=row, column=4, value=b.rate or 0.0)
                    rc.number_format = '0.00%'
                    row += 1
                if not brackets:
                    rs.cell(row=row, column=1, value=t.code)
                    rs.cell(row=row, column=2, value=t.name or '')
                    row += 1
                # Named range over this table's From/Rate block (cosmetic/
                # reference — the compiled formulas do NOT use it, D-L2).
                safe = re.sub(r'[^A-Za-z0-9]', '', (t.code or '')) or 'TABLE'
                ref = "'Rate Tables'!$C$%d:$D$%d" % (first_row, max(first_row, row - 1))
                try:
                    wb.defined_names[safe] = DefinedName(safe, attr_text=ref)
                except Exception:  # pragma: no cover — older openpyxl API
                    try:
                        wb.defined_names.append(DefinedName(safe, attr_text=ref))
                    except Exception:
                        pass

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        code = (c.code or c.name or 'config').strip().replace(' ', '_')
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': '%s_living.xlsx' % code,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'note': note,
        }

    @api.model
    def import_test_samples(self, config_id, file_b64, filename=None):
        """Read an uploaded .xlsx whose header row matches input component names
        and create one sample per data row (added alongside existing samples).
        Header→input is matched by name (case-insensitive) → code → letter."""
        import base64
        import io
        try:
            import openpyxl
        except Exception:
            return {'ok': False, 'msg': 'openpyxl is not available on the server.'}
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': 'Configuration not found'}
        try:
            raw = base64.b64decode(file_b64 or '')
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:
            return {'ok': False, 'msg': 'Could not read the file: %s' % (str(e).splitlines()[0] if str(e) else 'invalid xlsx')}
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {'ok': False, 'msg': 'The spreadsheet is empty.'}
        header = [(str(h).strip() if h is not None else '') for h in rows[0]]

        # Build lookup: lower(name) / lower(code) / lower(letter) -> code
        inputs = [r for r in c.rule_ids.sorted(key=lambda r: r.sequence)
                  if r.column_type == 'input']
        lookup = {}
        for r in inputs:
            for key in (r.name, r.code, r.column_letter):
                if key:
                    lookup.setdefault(str(key).strip().lower(), r.code)
        # Map each spreadsheet column index -> input code (skip unknown columns)
        col_to_code = {}
        for idx, h in enumerate(header):
            code = lookup.get(h.lower())
            if code:
                col_to_code[idx] = code
        if not col_to_code:
            return {'ok': False, 'msg': 'No header matched an input component. Use the exported template.'}

        base_n = len(c.sample_data_ids)
        created = 0
        for r_i, row in enumerate(rows[1:], start=1):
            if row is None or all(v is None or v == '' for v in row):
                continue
            vals = {}
            for idx, code in col_to_code.items():
                v = row[idx] if idx < len(row) else None
                if v is None or v == '':
                    continue
                try:
                    vals[code] = float(v)
                except (TypeError, ValueError):
                    vals[code] = v
            created += 1
            self.env['hr.formula.sample.data'].create({
                'config_id': c.id,
                'name': 'Sample %s' % (base_n + created),
                'source_type': 'manual',
                'input_values_json': json.dumps(vals),
            })
        if not created:
            return {'ok': False, 'msg': 'No data rows found under the header.'}
        new_ids = c.sample_data_ids.sorted(key=lambda s: s.id)[-created:]
        return {
            'ok': True, 'count': created,
            'first_id': new_ids[0].id if new_ids else False,
            'samples': [self._sample_row(x) for x in c.sample_data_ids],
        }

    @api.model
    def snapshot_expected(self, sample_id):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False}
        s.expected_values_json = s.computed_values_json or '{}'
        return self.get_sample_detail(s.id)

    @api.model
    def clear_expected(self, sample_id):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False}
        s.expected_values_json = '{}'
        return self.get_sample_detail(s.id)

    @api.model
    def delete_sample(self, sample_id):
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        cid = s.config_id.id if s.exists() else False
        if s.exists():
            s.unlink()
        samples = []
        if cid:
            c = self.env['hr.formula.config'].browse(cid)
            samples = [self._sample_row(x) for x in c.sample_data_ids]
        return {'ok': True, 'samples': samples}

    # ------------------------------------------------------------------
    # W84 — boundary-value test generation (engine does the work; these are
    # thin studio wrappers). Generation + confirm are manager-gated writes.
    # ------------------------------------------------------------------
    @api.model
    def boundary_candidates(self, config_id):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'candidates': [], 'reachable': 0, 'unreachable': 0}
        return c.boundary_candidates()

    @api.model
    def generate_boundary_samples(self, config_id, picks, base_sample_id=None):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': _('Configuration not found')}
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can generate test samples.')}
        try:
            r = c.generate_boundary_samples(picks or [], base_sample_id)
        except Exception as e:
            _logger.warning("generate_boundary_samples failed on %s: %s", c.id, e)
            return {'ok': False, 'msg': str(e).splitlines()[0] if str(e) else 'Could not generate.'}
        r['samples'] = [self._sample_row(x) for x in c.sample_data_ids]
        return r

    @api.model
    def confirm_sample_expected(self, sample_id):
        """Flip one generated sample's baseline to confirmed and re-run the chip
        (W84/D-G3). Manager-gated like every studio write."""
        s = self.env['hr.formula.sample.data'].browse(int(sample_id))
        if not s.exists():
            return {'ok': False}
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can confirm baselines.')}
        s.expected_confirmed = True
        detail = self.get_sample_detail(s.id)
        detail['tests'] = self._run_tests_after_save(s.config_id)
        return detail

    @api.model
    def confirm_all_samples(self, config_id):
        """Confirm every unconfirmed baseline in the config, then re-run the chip."""
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can confirm baselines.')}
        unconf = c.sample_data_ids.filtered(lambda x: not x.expected_confirmed)
        n = len(unconf)
        if unconf:
            unconf.expected_confirmed = True
        return {'ok': True, 'confirmed': n,
                'samples': [self._sample_row(x) for x in c.sample_data_ids],
                'tests': self._run_tests_after_save(c)}

    # ------------------------------------------------------------------
    # W49 — AI-proposed sample profiles (LLM proposes INPUTS; the engine
    # computes the truth — the LLM never supplies an output, so the
    # number-invention bug class is excluded by construction, D-G5).
    # ------------------------------------------------------------------
    @staticmethod
    def _as_num(v):
        try:
            n = float(v)
        except (TypeError, ValueError):
            return None
        if n != n or n in (float('inf'), float('-inf')):   # NaN/Inf guard
            return None
        return n

    @api.model
    def ai_propose_samples(self, config_id):
        """Ask the LLM for <=8 realistic INPUT profiles, hard-validate every one
        (unknown code / non-numeric / |v|>1e12 → rejected + reported), and return
        the survivors for the user to accept. No key / LLM error → {ok:False}."""
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'reason': _('Configuration not found.')}
        inputs = [r for r in c.rule_ids.sorted(key=lambda r: r.sequence)
                  if r.column_type == 'input' and r.code]
        if not inputs:
            return {'ok': False, 'reason': _('This configuration has no input components.')}
        input_codes = {r.code for r in inputs}

        # min/max observed across existing samples (helps the model stay realistic)
        obs = {}
        for s in c.sample_data_ids:
            try:
                iv = json.loads(s.input_values_json or '{}')
            except Exception:
                iv = {}
            if not isinstance(iv, dict):
                continue
            for code, v in iv.items():
                n = self._as_num(v)
                if n is None:
                    continue
                lo, hi = obs.get(code, (n, n))
                obs[code] = (min(lo, n), max(hi, n))
        schema = [{'code': r.code, 'name': r.name or r.code,
                   'default': r.default_value or 0.0,
                   'min_observed': obs.get(r.code, (None, None))[0],
                   'max_observed': obs.get(r.code, (None, None))[1]}
                  for r in inputs]

        system = (
            "You are PayAI, generating realistic payroll TEST INPUT profiles. "
            "You are given the input components of a salary configuration. Propose "
            "diverse, plausible employee profiles (e.g. junior, senior, part-time, "
            "high earner, edge cases). Reply with STRICT JSON only, shaped exactly:\n"
            '{"profiles":[{"name":"<short label>","inputs":{"<CODE>":<number>},'
            '"rationale":"<one sentence>"}]}\n'
            "Rules: use ONLY the given component codes as keys; every value MUST be a "
            "plain number (no %, no text, no formulas). NEVER include output, computed, "
            "or expected values — inputs only. At most 8 profiles."
        )
        user = ("Input components:\n" + json.dumps(schema, ensure_ascii=False)
                + "\n\nPropose up to 8 realistic, diverse input profiles.")
        try:
            data = self._llm_chat(
                [{'role': 'system', 'content': system},
                 {'role': 'user', 'content': user}], json_mode=True)
        except LLMUnavailable as e:
            return {'ok': False, 'reason': _('AI is unavailable: %s') % e}

        profiles = data.get('profiles') if isinstance(data, dict) else None
        if not isinstance(profiles, list):
            return {'ok': False, 'reason': _('AI returned an unexpected response shape.')}

        accepted, rejected = [], []
        for p in profiles[:8]:
            if not isinstance(p, dict):
                continue
            name = (str(p.get('name') or 'Profile').strip() or 'Profile')[:80]
            raw = p.get('inputs')
            if not isinstance(raw, dict) or not raw:
                rejected.append({'name': name, 'reason': 'no inputs'})
                continue
            clean, bad = {}, None
            for k, v in raw.items():
                if k not in input_codes:
                    bad = 'unknown code %s' % k
                    break
                n = self._as_num(v)
                if n is None:
                    bad = 'non-numeric %s' % k
                    break
                if abs(n) > 1e12:
                    bad = 'value out of range for %s' % k
                    break
                clean[k] = n
            if bad:
                rejected.append({'name': name, 'reason': bad})
                continue
            accepted.append({'name': name, 'inputs': clean,
                             'rationale': (str(p.get('rationale') or '').strip())[:200]})
        return {'ok': True, 'proposals': accepted, 'rejected': rejected}

    @api.model
    def create_ai_samples(self, config_id, proposals):
        """Turn accepted AI proposals into generated + unconfirmed samples with an
        engine-computed baseline (manager-gated). Re-validates every value — never
        trusts the client echo (D-G5)."""
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False, 'msg': _('Configuration not found')}
        if not self._can_edit():
            return {'ok': False, 'msg': _('Only managers can add test samples.')}
        input_codes = {r.code for r in c.rule_ids
                       if r.column_type == 'input' and r.code}
        created = 0
        rejected = 0
        for p in (proposals or []):
            if not isinstance(p, dict):
                rejected += 1
                continue
            raw = p.get('inputs') or {}
            # D-G5: one invalid entry rejects the WHOLE row — creating a sample
            # from the surviving keys would differ from what the user accepted.
            clean = {}
            bad = not isinstance(raw, dict) or not raw
            for k, v in (raw.items() if isinstance(raw, dict) else []):
                n = self._as_num(v)
                if k not in input_codes or n is None or abs(n) > 1e12:
                    bad = True
                    break
                clean[k] = n
            if bad or not clean:
                rejected += 1
                continue
            name = (str(p.get('name') or 'AI profile').strip() or 'AI profile')[:80]
            desc = 'AI-proposed profile: ' + (str(p.get('rationale') or '').strip())[:180]
            c._create_generated_sample(clean, name, desc)
            created += 1
        return {'ok': True, 'created': created, 'rejected': rejected,
                'samples': [self._sample_row(x) for x in c.sample_data_ids],
                'tests': self._run_tests_after_save(c)}

    @api.model
    def cfg_generate_wizard(self, config_id, source):
        c = self.env['hr.formula.config'].browse(int(config_id))
        if not c.exists():
            return {'ok': False}
        action = c.action_generate_sample_data()
        action.setdefault('context', {})['default_source'] = source
        # client-side doAction needs an explicit views array (server action
        # only sets view_mode, which makes web's _preprocessAction crash on .map)
        action['views'] = [(False, 'form')]
        return {'ok': True, 'action': action}

    # ------------------------------------------------------------------
    # PayAI : natural-language -> formula (deterministic mapper)
    # ------------------------------------------------------------------
    @api.model
    def ai_propose(self, config_id, text):
        config = self.env['hr.formula.config'].browse(int(config_id))
        if not config.exists():
            return {'ok': False, 'reply': 'No configuration loaded.'}
        rules = config.rule_ids.sorted(key=lambda r: r.sequence)
        text_l = (text or '').strip().lower()
        if not text_l:
            return {'ok': False, 'reply': 'Tell me what the component should compute.'}

        # Prefer a real LLM when configured; it only proposes — we validate its
        # output against the real columns below and fall back if anything is off.
        llm = self._llm_propose(config, text, rules)
        if llm is not None:
            return llm

        # build keyword -> rule index from names + codes
        def keywords(r):
            ks = set()
            ks.add((r.code or '').lower())
            for w in re.split(r'[\s_()/-]+', (r.name or '').lower()):
                if len(w) > 2:
                    ks.add(w)
            return {k for k in ks if k}

        kw_index = [(r, keywords(r)) for r in rules]

        # explain intent
        if text_l.startswith('explain') or 'what is' in text_l or 'how is' in text_l:
            target = self._match_rule(text_l, kw_index, rules)
            if target:
                by_col = self._col_to_rule(rules)
                return {'ok': True, 'kind': 'explain', 'reply': self._explain(target, by_col),
                        'target_name': target.name}
            return {'ok': True, 'kind': 'explain',
                    'reply': "I couldn't find that component. Try its exact name."}

        # detect target ("net is ...", "net = ...", "net pay is ...")
        target = None
        head = re.split(r'\b(is|equals?|=)\b', text_l, maxsplit=1)
        if len(head) >= 2:
            target = self._match_rule(head[0], kw_index, rules)
            body = head[-1]
        else:
            body = text_l

        # map operators + operands in order of appearance
        op_words = [
            (r'\bminus\b|\bless\b|\bsubtract(ing)?\b|\bafter\b|−|-', '-'),
            (r'\bplus\b|\band\b|\badd(ing)?\b|\+', '+'),
            (r'\btimes\b|\bmultipl\w*\b|×|\*|\bof\b', '*'),
            (r'\bdivided by\b|\bover\b|÷|/', '/'),
        ]
        # percentage like "20%" or "20 percent"
        pct = re.search(r'(\d+(?:\.\d+)?)\s*(?:%|percent)', body)

        # Find referenced components by matching their FULL NAME (or code) as a
        # contiguous substring of the request — far more precise than per-word
        # matching on configs with many overlapping names.
        matches = []  # (start, end, rule)
        for r in rules:
            if r == target:
                continue
            for needle in (r.name or '').lower().strip(), (r.code or '').lower().strip():
                if len(needle) >= 3:
                    p = body.find(needle)
                    if p >= 0:
                        matches.append((p, p + len(needle), r))
                        break
        # prefer longer (more specific) names; drop matches whose span is
        # contained within an already-accepted, longer match
        matches.sort(key=lambda m: (m[0], -(m[1] - m[0])))
        accepted = []
        for st, en, r in matches:
            if any(st >= a_st and en <= a_en and r is not ar for a_st, a_en, ar in accepted):
                continue
            accepted.append((st, en, r))
        accepted.sort(key=lambda m: m[0])
        ref_rules = [r for _, _, r in accepted]

        # fallback: loose per-word match only if no full-name match landed
        if not ref_rules:
            positions = []
            for r, ks in kw_index:
                best = None
                for k in ks:
                    if len(k) >= 4:
                        p = body.find(k)
                        if p >= 0 and (best is None or p < best):
                            best = p
                if best is not None and r != target:
                    positions.append((best, r))
            positions.sort()
            ref_rules = [r for _, r in positions]

        # special phrases
        if 'all deduction' in body or 'total deduction' in body:
            ded = next((r for r in rules if 'ded' in (r.code or '').lower() or 'deduction' in (r.name or '').lower()), None)
            if ded and ded not in ref_rules:
                ref_rules = [r for r in ref_rules if 'ded' not in (r.code or '').lower()]
                ref_rules.append(ded)

        if not ref_rules and not pct:
            return {'ok': False, 'reply': "I couldn't map that to your existing components. "
                    "Mention them by name, e.g. \"gross minus total deductions\"."}

        # determine operator
        op = '+'
        for pat, o in op_words:
            if re.search(pat, body):
                op = o
                break

        mapping = []
        if pct and len(ref_rules) >= 1:
            base = ref_rules[0]
            factor = float(pct.group(1)) / 100.0
            formula = '=%s1*%s' % (base.column_letter, factor)
            mapping.append({'phrase': pct.group(0), 'col': base.column_letter, 'name': base.name})
            mapping.append({'phrase': base.name, 'col': base.column_letter, 'name': base.name})
            human = '%s%% of %s' % (pct.group(1), base.name)
        else:
            cols = []
            for r in ref_rules:
                cols.append('%s1' % r.column_letter)
                mapping.append({'phrase': r.name, 'col': r.column_letter, 'name': r.name})
            formula = '=' + (' %s ' % op).join(cols)
            human = (' %s ' % OP_GLYPH.get(op, op)).join(r.name for r in ref_rules)

        result = {
            'ok': True,
            'kind': 'formula',
            'formula': formula,
            'human': human,
            'mapping': mapping,
            'reply': 'Here is the formula I built from your description.',
        }
        if target:
            result['target_id'] = target.id
            result['target_col'] = target.column_letter
            result['target_name'] = target.name
        else:
            result['target_name'] = None  # would be a new component
        return result

    @api.model
    def _match_rule(self, fragment, kw_index, rules):
        frag = (fragment or '').lower()
        # 1) prefer the longest full component name that appears in the fragment
        best_name, best_len = None, 0
        for r in rules:
            nm = (r.name or '').lower().strip()
            if len(nm) >= 3 and nm in frag and len(nm) > best_len:
                best_name, best_len = r, len(nm)
        if best_name:
            return best_name
        # 2) fall back to keyword overlap
        best, best_score = None, 0
        for r, ks in kw_index:
            score = sum(1 for k in ks if k and len(k) >= 4 and k in frag)
            if score > best_score:
                best, best_score = r, score
        return best

    @api.model
    def apply_ai_formula(self, rule_id, formula):
        return self.save_formula(rule_id, formula)

    # ------------------------------------------------------------------
    # LLM-backed proposal (provider-agnostic, OpenAI-compatible)
    # ------------------------------------------------------------------
    @api.model
    @api.model
    def _llm_chat(self, messages, json_mode=False):
        """One OpenAI-compatible chat call. Raises LLMUnavailable on a missing
        key, missing `requests`, timeout, non-200 or bad JSON — callers catch it
        for a deterministic fallback. Returns the assistant message text, or the
        parsed JSON object when json_mode=True.

        Missing key raises IMMEDIATELY, before any network call."""
        if requests is None:
            raise LLMUnavailable("HTTP client 'requests' is not available")
        ICP = self.env['ir.config_parameter'].sudo()
        api_key = (ICP.get_param(LLM_API_KEY) or '').strip()
        if not api_key:
            raise LLMUnavailable("No LLM API key configured")   # no network call
        base_url = (ICP.get_param(LLM_BASE_URL) or DEFAULT_BASE_URL).strip().rstrip('/')
        model = (ICP.get_param(LLM_MODEL) or DEFAULT_MODEL).strip()
        payload = {'model': model, 'messages': messages, 'temperature': 0}
        if json_mode:
            payload['response_format'] = {'type': 'json_object'}
        try:
            resp = requests.post(
                base_url + '/chat/completions',
                headers={'Authorization': 'Bearer ' + api_key, 'Content-Type': 'application/json'},
                json=payload, timeout=25)
            resp.raise_for_status()
            content = resp.json()['choices'][0]['message']['content']
        except Exception as e:
            raise LLMUnavailable("LLM call failed: %s" % e)
        if json_mode:
            try:
                return json.loads(content)
            except Exception as e:
                raise LLMUnavailable("LLM returned invalid JSON: %s" % e)
        return content

    def _llm_propose(self, config, text, rules):
        by_col = {r.column_letter: r for r in rules if r.column_letter}
        catalog = [{'col': r.column_letter, 'code': r.code, 'name': r.name,
                    'type': r.column_type} for r in rules if r.column_letter]
        system = (
            "You are PayAI, a payroll formula assistant inside Payobook. "
            "You translate a plain-language request into an Excel-style formula that "
            "references ONLY the existing components by their column letter followed by 1 "
            "(e.g. A1, BT1). Never invent column letters. Operators: + - * / and parentheses. "
            "Percentages become decimals (20% -> *0.2).\n"
            "Reply with STRICT JSON only, no prose, shaped exactly:\n"
            '{"kind":"formula"|"explain","target_col":"<existing letter or null>",'
            '"formula":"=<excel using LETTER1 refs>","human":"<short plain-english>",'
            '"mapping":[{"name":"<component name>","col":"<letter>"}],"reply":"<one sentence>"}\n'
            "For an explain request set kind=explain and put the explanation in reply; "
            "formula/mapping may be empty. target_col is the component being defined if the "
            "user names one (else null = a new component)."
        )
        user = "Components:\n" + json.dumps(catalog, ensure_ascii=False) + "\n\nRequest: " + (text or '')
        try:
            data = self._llm_chat(
                [{'role': 'system', 'content': system},
                 {'role': 'user', 'content': user}],
                json_mode=True)
        except LLMUnavailable as e:
            _logger.info("PayAI LLM unavailable, falling back: %s", e)
            return None

        return self._validate_llm(data, by_col)

    @api.model
    def _validate_llm(self, data, by_col):
        """Trust nothing: every referenced column must exist; else reject."""
        try:
            kind = data.get('kind') or 'formula'
            if kind == 'explain':
                return {'ok': True, 'kind': 'explain',
                        'reply': data.get('reply') or data.get('human') or 'Here is what that does.'}
            formula = (data.get('formula') or '').strip()
            if not formula:
                return None
            refs = set(re.findall(r'([A-Za-z]+)\d+', formula))
            bad = [c for c in refs if c.upper() not in by_col]
            if bad or not refs:
                _logger.info("PayAI LLM referenced unknown columns %s; falling back.", bad)
                return None
            target_col = data.get('target_col')
            target = by_col.get((target_col or '').upper()) if target_col else None
            # rebuild mapping from real components for trustworthy display
            mapping = [{'name': by_col[c.upper()].name, 'col': c.upper()} for c in sorted(refs)]
            result = {
                'ok': True, 'kind': 'formula',
                'formula': formula if formula.startswith('=') else '=' + formula,
                'human': self._readable_formula(formula, by_col),
                'mapping': mapping,
                'reply': data.get('reply') or data.get('human') or 'Here is the formula I built from your description.',
            }
            if target:
                result['target_id'] = target.id
                result['target_col'] = target.column_letter
                result['target_name'] = target.name
            else:
                result['target_name'] = None
            return result
        except Exception as e:
            _logger.warning("PayAI LLM validation error: %s", e)
            return None

    @api.model
    def _readable_formula(self, formula, by_col):
        """Render '=AV1-BB1' as 'Tổng thu nhập − TỔng BHXH' for display."""
        out = []
        for tok in re.findall(r'[A-Za-z]+\d+|\d+\.?\d*|[+\-*/()%]', (formula or '').lstrip('=')):
            m = re.match(r'^([A-Za-z]+)\d+$', tok)
            if m and m.group(1).upper() in by_col:
                out.append(by_col[m.group(1).upper()].name)
            elif tok in OP_GLYPH:
                out.append(OP_GLYPH[tok])
            else:
                out.append(tok)
        return ' '.join(out) if out else (formula or '')

    @api.model
    def ai_status(self):
        ICP = self.env['ir.config_parameter'].sudo()
        return {'llm': bool((ICP.get_param(LLM_API_KEY) or '').strip()),
                'model': ICP.get_param(LLM_MODEL) or DEFAULT_MODEL}

    # ------------------------------------------------------------------
    # Explain a formula (T5.2) — LLM with deterministic floor, EN/VI
    # ------------------------------------------------------------------
    @api.model
    def explain_formula_ai(self, rule_id, lang='en'):
        """Plain-language explanation of one component. Tries the LLM; on ANY
        failure returns the deterministic _explain output. Never raises to the
        client. Returns {'text', 'source': 'ai'|'deterministic'}."""
        lang = 'vi' if str(lang or '').lower().startswith('vi') else 'en'
        rule = self.env['hr.formula.rule'].browse(int(rule_id))
        if not rule.exists():
            return {'text': '', 'source': 'deterministic'}
        by_col = self._col_to_rule(rule.config_id.rule_ids)
        floor = self._explain_localized(rule, by_col, lang)   # always computable
        try:
            text = (self._llm_chat(self._build_explain_prompt(rule, by_col, lang)) or '').strip()
            if text:
                return {'text': text, 'source': 'ai'}
        except LLMUnavailable:
            pass
        except Exception as e:                                # never leak a traceback
            _logger.info("explain_formula_ai fell back: %s", e)
        return {'text': floor, 'source': 'deterministic'}

    @api.model
    def _build_explain_prompt(self, rule, by_col, lang):
        toks = self._tokenize(rule, by_col)
        deps = []
        for t in toks:
            if t.get('kind') == 'ref' and t['text'] not in deps:
                deps.append(t['text'])
        lang_name = 'Vietnamese' if lang == 'vi' else 'English'
        system = ("You are PayAI, a payroll assistant. Explain what a salary component "
                  "computes in plain %s — 1-2 short sentences for a non-technical payroll "
                  "officer. No formulas, code or column letters." % lang_name)
        facts = {
            'component': rule.name or '',
            'category': rule.category_id.name if rule.category_id else (rule.column_type or ''),
            'excel_formula': rule.excel_formula or '',
            'depends_on': deps,
        }
        user = "Explain this component in %s:\n%s" % (lang_name, json.dumps(facts, ensure_ascii=False))
        return [{'role': 'system', 'content': system}, {'role': 'user', 'content': user}]

    @api.model
    def _explain_localized(self, rule, by_col, lang):
        return self._explain_vi(rule, by_col) if lang == 'vi' else self._explain(rule, by_col)

    @api.model
    def _explain_vi(self, rule, by_col):
        if rule.column_type == 'input':
            return "Lấy từ hợp đồng của mỗi nhân viên hoặc từ dữ liệu nhập hàng tháng."
        if rule.column_type == 'constant':
            return "Một giá trị cố định áp dụng cho tất cả nhân viên."
        names = []
        for t in self._tokenize(rule, by_col):
            if t['kind'] == 'ref' and t['text'] not in names:
                names.append(t['text'])
        if names:
            tail = " và các thành phần khác." if len(names) > 6 else "."
            return "%s được tính từ %s%s" % (rule.name or '', ', '.join(names[:6]), tail)
        return "%s là một thành phần được tính toán." % (rule.name or '')

    # ------------------------------------------------------------------
    # First-setup wizard
    # ------------------------------------------------------------------
    # Vietnam Standard starter set (code, name, type, excel_formula, constant)
    VN_STANDARD = [
        ('BASIC', 'Basic Salary', 'input', '', 0.0),
        ('HRA', 'Housing Allowance', 'formula', '=A1*0.2', 0.0),
        ('TRANSPORT', 'Transport Allowance', 'constant', '', 500000.0),
        ('MEAL', 'Meal Allowance', 'constant', '', 730000.0),
        ('GROSS', 'Gross Salary', 'formula', '=A1+B1+C1+D1', 0.0),
        ('SIEMP', 'Social Insurance (Employee)', 'formula', '=A1*0.08', 0.0),
        ('HIEMP', 'Health Insurance (Employee)', 'formula', '=A1*0.015', 0.0),
        ('UIEMP', 'Unemployment Insurance (Employee)', 'formula', '=A1*0.01', 0.0),
        ('TOTALDED', 'Total Deductions', 'formula', '=F1+G1+H1', 0.0),
        ('NET', 'Net Salary', 'formula', '=E1-I1', 0.0),
    ]

    @api.model
    def _idx_letter(self, i):
        """0->A, 25->Z, 26->AA … (Excel-style)."""
        s = ''
        i += 1
        while i:
            i, r = divmod(i - 1, 26)
            s = chr(65 + r) + s
        return s

    # Built-in starter entries. The legacy 'vn_standard' set predates the F113
    # converter contract; its codes carried underscores (SI_EMP, TOTAL_DED) and
    # survived only because nothing referenced them by name. MAPFIX A closed that
    # — the shape constraint on hr.formula.rule would now refuse them outright, so
    # the set ships as SIEMP/HIEMP/UIEMP/TOTALDED. It is kept as code, NOT a
    # registry record. All richer country packs come from
    # hr.formula.config.template (F113).
    _BUILTIN_TEMPLATES = [
        {'key': 'vn_standard', 'name': 'Vietnam Standard', 'country': 'VN',
         'flag': '🇻🇳', 'version': 'legacy', 'builtin': True, 'certified': False,
         'effective_date': False, 'refs': [],
         'desc': '10 components pre-wired: Basic, allowances, SI/HI/UI, Gross & Net — VN statutory rates.',
         'components': [], 'rate_tables': [],
         'preview': [{'col': 'A', 'name': 'Basic Salary', 'f': 'input'},
                     {'col': 'B', 'name': 'Housing Allowance', 'f': '= Basic × 20%'},
                     {'col': 'E', 'name': 'Gross Salary', 'f': '= A+B+C+D'},
                     {'col': 'J', 'name': 'Net Salary', 'f': '= Gross − Deductions'}]},
        {'key': 'blank', 'name': 'Blank canvas', 'country': False,
         'flag': '', 'version': '', 'builtin': True, 'certified': False,
         'effective_date': False, 'refs': [], 'components': [], 'rate_tables': [],
         'desc': 'Start empty and build components one by one — or ask PayAI to draft them.',
         'preview': []},
    ]

    @api.model
    def wizard_templates(self):
        """Starter templates for the create-config wizard: the built-in legacy
        set + every installed F113 country pack (hr.formula.config.template,
        excluding superseded versions). Each registry entry carries the picker
        UX payload (T113.7): country/flag/version/effective date, the full
        component + rate-table preview, and legislation references."""
        out = [dict(t) for t in self._BUILTIN_TEMPLATES]
        if 'hr.formula.config.template' not in self.env:
            return out
        try:
            # savepoint: if the registry table doesn't exist yet (deploy
            # window — new code, base module not upgraded), the failed
            # statement must not poison the cursor; the built-in entries keep
            # the create wizard alive regardless.
            with self.env.cr.savepoint():
                templates = self.env['hr.formula.config.template'].sudo().search(
                    [('state', '!=', 'superseded')],
                    order='country_code, sequence, effective_date desc')
                templates.mapped('code')  # force the fetch inside the savepoint
        except Exception:
            _logger.exception("F113: template registry unavailable — "
                              "serving built-in templates only")
            return out
        for tpl in templates:
            comps = tpl._components()
            preview = []
            for c in comps:
                if c.get('type') == 'input':
                    f = 'input'
                elif c.get('type') == 'constant':
                    f = 'constant'
                else:
                    f = (c.get('excel_formula') or '').lstrip('=') or 'formula'
                preview.append({'col': c.get('column_letter') or '',
                                'name': c.get('name') or c.get('code'), 'f': f})
            out.append({
                'key': tpl.code, 'name': tpl.name, 'country': tpl.country_code,
                'flag': tpl.flag or '', 'version': tpl.version,
                'effective_date': tpl.effective_date and str(tpl.effective_date) or False,
                'state': tpl.state, 'certified': tpl.state == 'certified',
                'builtin': False,
                'desc': tpl.description or '',
                'components': [{'code': c.get('code'), 'name': c.get('name'),
                               'type': c.get('type'), 'category': c.get('category'),
                               'col': c.get('column_letter') or ''} for c in comps],
                'rate_tables': [{'code': rt.get('code'), 'name': rt.get('name'),
                                'brackets': rt.get('brackets') or []}
                               for rt in tpl._rate_tables()],
                'refs': tpl._legislation_refs(),
                'preview': preview[:8],
            })
        return out

    @api.model
    def create_config(self, vals):
        vals = vals or {}
        Config = self.env['hr.formula.config']
        cvals = {
            'name': vals.get('name') or 'New Payroll Config',
            'country_code': vals.get('country_code') or 'VN',
            'cycle_type': vals.get('cycle_type') or 'regular',
            'state': 'draft',
        }
        # only set code if the caller explicitly supplied one; otherwise let
        # hr.formula.config.create() auto-generate a unique code from the name.
        if vals.get('code'):
            cvals['code'] = vals['code']
        cfg = Config.create(cvals)
        self._seed_template(cfg, vals.get('template') or 'blank')
        return {'ok': True, 'config_id': cfg.id, 'rule_count': len(cfg.rule_ids)}

    def _seed_template(self, cfg, key):
        """Populate an empty config from a starter template. Shared by the
        creation wizard and the cockpit 'Use Vietnam Standard' resume CTA.

        Routing (F113): the built-in 'vn_standard' set stays a hardcoded code
        path (guaranteed byte-identical to pre-F113); 'blank' seeds nothing; any
        other key is looked up in the hr.formula.config.template registry and
        materialised by its own seeder (categories, rate tables, frozen letters,
        B4-resolved statutory constants, sample tests)."""
        if key and key not in ('vn_standard', 'blank'):
            Template = self.env['hr.formula.config.template']
            # never materialise a superseded structure — the picker hides
            # them, so a stale bookmarked/scripted key must not bypass that
            tpl = Template.sudo().search([
                ('code', '=', key), ('state', '!=', 'superseded')], limit=1)
            if tpl:
                tpl.seed_config(cfg)
                return
            _logger.warning("F113: unknown or superseded template key '%s' — "
                            "seeding blank", key)
            return
        if key == 'vn_standard':
            # Assign column letters explicitly. The model's position-based
            # compute is unreliable during batch create (o2m cache staleness
            # makes every new rule resolve to 'A'), so we provide the stored
            # value directly — it persists via the field's inverse and skips
            # the faulty compute, keeping A..J distinct for the constraint.
            vals_list = []
            for i, (code, name, ctype, formula, const) in enumerate(self.VN_STANDARD):
                vals_list.append({
                    'config_id': cfg.id, 'code': code, 'name': name,
                    'column_type': ctype, 'excel_formula': formula,
                    'constant_value': const, 'sequence': i + 1,
                    'column_letter': self._idx_letter(i),
                })
            self.env['hr.formula.rule'].create(vals_list)
            try:
                cfg.action_regenerate_formulas()
            except Exception as e:
                _logger.warning("Template formula regen failed: %s", e)
            # seed a sample so the live preview works immediately
            try:
                self.env['hr.formula.sample.data'].create({
                    'config_id': cfg.id, 'name': 'Sample — Standard',
                    'input_values_json': json.dumps({'BASIC': 15000000}),
                })
            except Exception as e:
                _logger.warning("Template sample seed failed: %s", e)

    @api.model
    def apply_starter(self, config_id, key):
        """Apply a starter template to an EXISTING empty config (cockpit
        'finish setup' resume). Guarded so it never duplicates rules."""
        cfg = self.env['hr.formula.config'].browse(int(config_id))
        if not cfg.exists():
            return {'ok': False, 'error': 'not_found'}
        if cfg.rule_ids:
            return {'ok': False, 'error': 'not_empty'}
        self._seed_template(cfg, key or 'vn_standard')
        return {'ok': True, 'config_id': cfg.id, 'rule_count': len(cfg.rule_ids)}

    @api.model
    def delete_config(self, config_id):
        """Delete a configuration from the cockpit.

        Refuses when the config produced real payroll history — the caller is
        told to archive instead. ``unlink()`` guards this too; checking here
        just buys a readable message and leaves the cursor clean.
        """
        cfg = self.env['hr.formula.config'].browse(int(config_id))
        if not cfg.exists():
            return {'ok': True}
        verdict = self._delete_eligibility(cfg)
        if not verdict['can_delete']:
            return {'ok': False, 'can_archive': True, 'msg': _(
                "\"%(name)s\" is used by %(blockers)s, so it can't be deleted. "
                "Archive it instead to hide it without losing that history.",
                name=cfg.name or '', blockers=verdict['delete_blocked_by'],
            )}
        name = cfg.name or ''
        try:
            cfg.unlink()
        except UserError as e:
            return {'ok': False, 'can_archive': True, 'msg': str(e)}
        except Exception as e:
            _logger.exception("delete_config failed for %s", config_id)
            return {'ok': False, 'can_archive': True, 'msg': _(
                "Could not delete \"%(name)s\": %(err)s",
                name=name, err=(str(e).splitlines()[0] if str(e) else _('it may be in use.')),
            )}
        return {'ok': True, 'deleted': True, 'name': name}

    @api.model
    def archive_config(self, config_id):
        """Archive a configuration from the cockpit gallery (soft delete)."""
        cfg = self.env['hr.formula.config'].browse(int(config_id))
        if not cfg.exists():
            return {'ok': False, 'error': 'not_found'}
        name = cfg.name or ''
        try:
            cfg.action_archive()
        except Exception as e:
            _logger.exception("archive_config failed for %s", config_id)
            return {'ok': False, 'msg': _(
                "Could not archive \"%(name)s\": %(err)s",
                name=name, err=(str(e).splitlines()[0] if str(e) else _('unexpected error.')),
            )}
        return {'ok': True, 'archived': True, 'name': name}


class SampleDataWizardStudio(models.TransientModel):
    """Let the Test workbench's restyled generator return to the cockpit
    instead of navigating to the stock 'created samples' list view."""
    _inherit = 'hr.formula.sample.data.wizard'

    def action_generate_and_close(self):
        self.action_generate_samples()
        return {'type': 'ir.actions.act_window_close'}
