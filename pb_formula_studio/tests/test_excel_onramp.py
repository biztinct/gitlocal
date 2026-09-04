# -*- coding: utf-8 -*-
"""JOURNEY J2 — the Excel on-ramp.

The Spreadsheet side of the Mapping home was built and unreachable. Its left
column could only offer the keys of an import batch somebody had already run,
or whatever you typed from memory — which is why S12 found it had never written
a value on any of the four live databases: you had to finish the import before
you could set up the mapping the import needs.

Four things close that, and each of them is a claim that can be wrong quietly:

  * **the parity invariant.** The header reader runs the LOADER'S parse, so a
    column the board shows is a key the loader will produce for the same file.
    If those two ever diverge, the board goes back to being a place you type
    hopeful spellings — and nothing on screen would say so. `test_05_*` is the
    assertion that makes the refactor worth having, for a single-sheet file and
    a multisheet one;
  * **it reads headings, not data.** Enforced by a row-count diff, because a
    promise about what something does not do is worth exactly the check behind
    it (`test_02_*`);
  * **the template round-trips.** A template whose headings the resolver does
    not match is a worse artefact than no template, since it looks official.
    `test_06_*` builds one and feeds it straight back through the reader;
  * **the doors say which of two opposite things they do.** "Import from Excel"
    (define columns) sat next to "Payroll Import" (load numbers) for the whole
    life of this product. `test_09_*` is a source gate on every label.

Fixtures are built here rather than leaned on: `ABM/ABM Template.xlsx` is a
live-validation fixture (MF23) and a unit test that depends on a workbook in
the repo root is a unit test that fails on somebody else's checkout.
"""
import base64
import io
import json
import os
import re

from odoo.modules.module import get_module_path
from odoo.tests import TransactionCase, tagged

try:
    import openpyxl
    OPENPYXL = True
except ImportError:  # pragma: no cover
    OPENPYXL = False


def _read(path):
    with open(path, encoding='utf-8') as fh:
        return fh.read()


def _src(module, *parts):
    return _read(os.path.join(get_module_path(module), *parts))


def _strip_xml_comments(src):
    return re.sub(r'<!--.*?-->', '', src, flags=re.S)


def _strip_js_comments(src):
    src = re.sub(r'/\*.*?\*/', '', src, flags=re.S)
    return re.sub(r'^\s*//.*$', '', src, flags=re.M)


def _strip_py_comments(src):
    return re.sub(r'^\s*#.*$', '', src, flags=re.M)


@tagged('post_install', '-at_install')
class TestExcelOnRamp(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Studio = cls.env['pb.formula.studio']
        cls.Config = cls.env['hr.formula.config']
        cls.Rule = cls.env['hr.formula.rule']
        cls.Batch = cls.env['hr.payroll.import.batch']

    # ------------------------------------------------------------- fixtures
    def _config(self, name, sheet=''):
        # CR19 — country_code is required with no default.
        cfg = self.Config.create({
            'name': name, 'code': re.sub(r'\W', '', name.upper())[:32],
            'country_code': 'VN', 'state': 'active',
        })
        specs = [('Employee Code', 'EMPCODE'), ('Basic Salary', 'BASIC'),
                 ('Overtime Hours', 'OTHOURS'), ('Meal Allowance', 'MEAL')]
        for seq, (label, code) in enumerate(specs, start=1):
            self.Rule.create({
                'config_id': cfg.id, 'name': label, 'code': code,
                'column_type': 'input', 'sequence': seq,
                'source_sheet_name': sheet,
            })
        # an OUTPUT, which must never appear in a template
        self.Rule.create({
            'config_id': cfg.id, 'name': 'Net Pay', 'code': 'NETPAY',
            'column_type': 'formula', 'sequence': 90,
            'source_sheet_name': sheet,
        })
        return cfg

    def _workbook(self, sheets):
        """`{sheet: [[header…], [row…], …]}` → xlsx bytes."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for title, rows in sheets.items():
            ws = wb.create_sheet(title=title)
            for row in rows:
                ws.append(row)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def _flat_file(self):
        return self._workbook({'Sheet1': [
            ['Employee Code', 'Basic Salary', 'Overtime Hours', 'Meal Allowance'],
            ['E001', 12500000, 8, 730000],
            ['E002', 9800000, 0, 730000],
        ]})

    def _multi_file(self):
        return self._workbook({
            'SEVL': [['Employee Code', 'Basic Salary'],
                     ['E001', 12500000], ['E002', 9800000]],
            'EXTRA': [['Employee Code', 'Meal Allowance'],
                      ['E001', 730000], ['E002', 730000]],
        })

    def _counts(self):
        return (self.Batch.search_count([]),
                self.env['hr.payroll.import.line'].search_count([]))

    # =====================================================================
    # 1 — the empty state
    # =====================================================================
    def test_01_empty_state_offers_the_dropzone_and_says_headings_not_data(self):
        cfg = self._config('J2 Empty')
        data = self.Studio.import_mapping_data(cfg.id, False)
        self.assertTrue(data['ok'])
        self.assertIsNone(data['sample'],
                          "no file has been read — the ramp must be in its empty state")
        self.assertTrue(data['can_add'],
                        "free-typed columns survive the on-ramp (S6)")
        tpl = _strip_xml_comments(
            _src('pb_formula_studio', 'static/src/xml/mapping_studio.xml'))
        self.assertIn('pbms-drop', tpl)
        # the copy has to say what it reads AND what it does not do
        self.assertIn('reads the column headings', tpl)
        for promise in ('does not import your numbers',
                        'does not create a pay run'):
            self.assertIn(promise, tpl,
                          "the dropzone must say what it will not do: %s" % promise)

    # =====================================================================
    # 2 — a drop fills the lane and creates NOTHING
    # =====================================================================
    def test_02_reading_headings_fills_the_lane_and_creates_no_batch(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Flat')
        before = self._counts()
        r = self.Studio.import_mapping_read_headers(
            cfg.id, base64.b64encode(self._flat_file()).decode(), 'pay.xlsx')
        self.assertTrue(r['ok'], r.get('msg'))
        self.assertEqual(self._counts(), before,
                         "reading headings must not create a batch or a line")

        lane = [i for i in r['left'] if i['group'].startswith('pay.xlsx')]
        self.assertEqual([i['label'] for i in lane],
                         ['Employee Code', 'Basic Salary', 'Overtime Hours',
                          'Meal Allowance'],
                         "the lane is the workbook's real headings, in file order")
        self.assertEqual(r['read']['shown'], 4)
        # the sample is the FIRST row's value, formatted for a person
        by_label = {i['label']: i['sublabel'] for i in lane}
        self.assertEqual(by_label['Basic Salary'], 'e.g. 12,500,000')
        self.assertEqual(by_label['Employee Code'], 'e.g. E001')
        # and no pay value was written anywhere
        self.assertFalse(cfg.rule_ids.filtered('source_binding'))

    # =====================================================================
    # 3 — it survives a reload, with provenance
    # =====================================================================
    def test_03_columns_persist_with_a_provenance_line(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Persist')
        self.Studio.import_mapping_read_headers(
            cfg.id, base64.b64encode(self._flat_file()).decode(), 'march.xlsx')
        # a FRESH read of the board — this is what a page reload does
        again = self.Studio.import_mapping_data(cfg.id, False)
        sample = again['sample']
        self.assertTrue(sample)
        self.assertEqual(sample['filename'], 'march.xlsx')
        self.assertTrue(sample['read_on'], "the provenance line needs a WHEN")
        self.assertIn('march.xlsx', sample['line'])
        self.assertIn('read', sample['line'])
        self.assertEqual(sample['shown'], 4)
        self.assertTrue(sample['has_file'],
                        "the file itself is kept — the handoff needs it")
        self.assertTrue(any(i['group'] == sample['line'] for i in again['left']))

    # =====================================================================
    # 4 — wiring a discovered column, and taking it back
    # =====================================================================
    def test_04_a_discovered_column_wires_through_the_existing_path(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Wire')
        self.Studio.import_mapping_read_headers(
            cfg.id, base64.b64encode(self._flat_file()).decode(), 'pay.xlsx')
        rule = cfg.rule_ids.filtered(lambda r: r.code == 'BASIC')

        r = self.Studio.import_mapping_create(cfg.id, False, 'c:Basic Salary', rule.id)
        self.assertTrue(r['ok'], r.get('msg'))
        self.assertEqual(rule.source_binding, 'excel')
        self.assertEqual(rule.source_binding_key, 'Basic Salary')
        self.assertEqual(rule.source_binding_origin, 'board')

        board = self.Studio.import_mapping_data(cfg.id, False)
        self.assertEqual(board['wired'], 1)
        self.assertTrue(any(w['rightId'] == rule.id and w['state'] == 'accepted'
                            for w in board['wires']))

        # undo, and the column comes back — from the file lane, not from nowhere
        self.Studio.import_mapping_delete(rule.id)
        self.assertFalse(rule.source_binding)
        after = self.Studio.import_mapping_data(cfg.id, False)
        self.assertEqual(after['wired'], 0)
        self.assertTrue(any(i['id'] == 'c:Basic Salary' for i in after['left']))

    def test_04b_forgetting_the_file_keeps_every_wire(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Forget')
        self.Studio.import_mapping_read_headers(
            cfg.id, base64.b64encode(self._flat_file()).decode(), 'pay.xlsx')
        rule = cfg.rule_ids.filtered(lambda r: r.code == 'BASIC')
        self.Studio.import_mapping_create(cfg.id, False, 'c:Basic Salary', rule.id)

        r = self.Studio.import_mapping_forget_headers(cfg.id)
        self.assertTrue(r['ok'])
        self.assertIsNone(r['sample'])
        self.assertFalse(cfg.import_sample_file)
        self.assertEqual(rule.source_binding_key, 'Basic Salary',
                         "deleting a file must never silently unwire a scheme")
        self.assertTrue(any(i['id'] == 'c:Basic Salary'
                            and i['group'] == 'Already used by this scheme'
                            for i in r['left']))

    # =====================================================================
    # 5 — THE INVARIANT
    # =====================================================================
    def _loader_keys(self, cfg, content, filename):
        """The keys `action_load_file` really writes, from a real batch."""
        batch = self.Batch.create({
            'name': 'J2 parity probe', 'formula_config_id': cfg.id,
            'source_type': 'excel',
            'import_file': base64.b64encode(content), 'import_filename': filename,
        })
        batch.action_load_file()
        line = self.env['hr.payroll.import.line'].search(
            [('batch_id', '=', batch.id)], limit=1, order='sequence')
        keys = set(json.loads(line.raw_data_json).keys())
        batch.import_line_ids.unlink()
        batch.unlink()
        return keys

    def test_05_parser_parity_single_sheet(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Parity Flat')          # no source_sheet_name
        content = self._flat_file()
        peeked = {c['key'] for c in
                  self.Batch.peek_source_columns(cfg, content, 'pay.xlsx')}
        loaded = self._loader_keys(cfg, content, 'pay.xlsx')
        self.assertEqual(peeked, loaded,
                         "the board's keys and the loader's keys are ONE set")
        self.assertIn('Basic Salary', peeked)
        self.assertIn('B', peeked, "the loader offers column letters too")

    def test_05b_parser_parity_multisheet(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Parity Multi', sheet='SEVL')
        content = self._multi_file()
        peeked = {c['key'] for c in
                  self.Batch.peek_source_columns(cfg, content, 'pay.xlsx')}
        loaded = self._loader_keys(cfg, content, 'pay.xlsx')
        self.assertEqual(peeked, loaded,
                         "the multisheet branch has to agree too — it is the "
                         "branch abm's own scheme takes")
        self.assertIn('SEVL|Basic Salary', peeked,
                      "a multisheet load keys on Sheet|Header")

    def test_05c_one_card_per_column_not_one_per_spelling(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Prefer', sheet='SEVL')
        cols = self.Batch.peek_source_columns(cfg, self._multi_file(), 'pay.xlsx')
        preferred = [c for c in cols if c['preferred']]
        # 'Employee Code' + 'Basic Salary' on SEVL, 'Meal Allowance' on EXTRA
        self.assertEqual(sorted(c['header'] for c in preferred),
                         ['Basic Salary', 'Employee Code', 'Meal Allowance'])
        self.assertTrue(all(c['sheet'] for c in preferred),
                        "on a multisheet scheme the sheet-qualified spelling wins")
        self.assertLess(len(preferred), len(cols),
                        "the other spellings are kept in the stored list")

    # =====================================================================
    # 6 / 7 — the template
    # =====================================================================
    def test_06_template_headers_are_what_the_resolver_matches(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Template')
        r = self.Studio.import_mapping_template(cfg.id)
        self.assertTrue(r['ok'], r.get('msg'))
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r['file_b64'])))
        ws = wb.worksheets[0]
        headers = [c.value for c in ws[1] if c.value is not None]
        self.assertEqual(headers[0], 'Employee Code',
                         "the row-matching column comes first")
        self.assertEqual(headers,
                         ['Employee Code', 'Basic Salary', 'Overtime Hours',
                          'Meal Allowance'])
        self.assertNotIn('Net Pay', headers,
                         "an output is computed, never filled in")
        self.assertEqual(ws.max_row, 1,
                         "a template arrives EMPTY — a sample row is a row "
                         "somebody imports by accident")

    def test_06b_template_round_trips_through_header_discovery(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        for sheet in ('', 'SEVL'):
            cfg = self._config('J2 RT %s' % (sheet or 'flat'), sheet=sheet)
            built = self.Studio.import_mapping_template(cfg.id)
            self.assertTrue(built['ok'], built.get('msg'))
            back = self.Studio.import_mapping_read_headers(
                cfg.id, built['file_b64'], built['filename'])
            self.assertTrue(back['ok'], back.get('msg'))
            offered = {i['id'] for i in back['left']}
            inputs = cfg.rule_ids.filtered(lambda x: x.column_type == 'input')
            for rule in inputs:
                key = ('%s|%s' % (sheet, rule.name)) if sheet else rule.name
                self.assertIn('c:' + key, offered,
                              "%s must find its own column back (sheet=%r)"
                              % (rule.code, sheet))

    def test_07_a_bound_component_emits_its_binding_key(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        cfg = self._config('J2 Bound')
        rule = cfg.rule_ids.filtered(lambda r: r.code == 'BASIC')
        rule.set_source_binding('excel', 'X', origin='board')
        r = self.Studio.import_mapping_template(cfg.id)
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r['file_b64'])))
        headers = [c.value for c in wb.worksheets[0][1] if c.value is not None]
        self.assertIn('X', headers,
                      "the template says what the component actually reads")
        self.assertNotIn('Basic Salary', headers,
                         "…and not the label it happens to carry")

    # =====================================================================
    # 8 — the handoff (stops before process, cleans up after itself)
    # =====================================================================
    def test_08_handoff_creates_a_batch_through_the_existing_flow(self):
        if not OPENPYXL:
            self.skipTest('openpyxl unavailable')
        if 'pb.import.wizard' not in self.env:
            self.skipTest('pb_import_wizard not installed')
        cfg = self._config('J2 Handoff')
        content = self._flat_file()
        self.Studio.import_mapping_read_headers(
            cfg.id, base64.b64encode(content).decode(), 'march.xlsx')
        before = self._counts()

        r = self.Studio.import_mapping_handoff(cfg.id)
        self.assertTrue(r['ok'], r.get('msg'))
        batch = self.Batch.browse(r['batch_id'])
        self.assertTrue(batch.exists())
        self.assertEqual(r['total_lines'], 2)
        # it STOPS at match. Nothing is validated, nothing is committed.
        self.assertIn(batch.state, ('loaded', 'matched'))
        self.assertFalse(batch.created_payslip_ids)
        self.assertFalse(batch.created_employee_ids)

        line = batch.import_line_ids.sorted('sequence')[0]
        raw = json.loads(line.raw_data_json)
        self.assertEqual(raw['Employee Code'], 'E001')
        self.assertEqual(raw['Basic Salary'], 12500000)
        # …and the line's keys are the keys the board offered
        offered = {c['key'] for c in
                   self.Batch.peek_source_columns(cfg, content, 'march.xlsx')}
        self.assertEqual(set(raw.keys()), offered)

        batch.import_line_ids.unlink()
        batch.unlink()
        self.assertEqual(self._counts(), before, "the probe cleans up after itself")

    def test_08b_handoff_refuses_when_there_is_no_file(self):
        cfg = self._config('J2 No File')
        r = self.Studio.import_mapping_handoff(cfg.id)
        self.assertFalse(r['ok'])
        self.assertIn('file', r['msg'].lower())

    # =====================================================================
    # 9 — the doors
    # =====================================================================
    def test_09_every_data_door_reaches_the_one_guided_flow(self):
        guided = self.env.ref('pb_import_wizard.action_pb_import_wizard',
                              raise_if_not_found=False)
        cfg = self._config('J2 Doors')
        act = cfg.action_launch_payroll_import()
        if guided:
            self.assertEqual(act.get('tag'), 'pb_import_wizard',
                             "the scheme's door lands in the guided flow")
            self.assertEqual(act['context']['default_formula_config_id'], cfg.id,
                             "…still pre-scoped to the scheme it was pressed on")
        else:
            self.assertEqual(act.get('res_model'), 'hr.payroll.import.batch')

        # the menu action is the server action, and it routes the same way
        server = self.env.ref('pb_hr_payroll_formula.action_payroll_load_pay_data')
        self.assertEqual(server.state, 'code')
        self.assertIn('action_open_guided_import', server.code)
        menu = self.env.ref('pb_hr_payroll_formula.menu_import_new')
        self.assertEqual(menu.action.id, server.id)

    def test_09b_structure_and_history_doors_read_as_what_they_are(self):
        cfg = self._config('J2 Labels')
        setup = cfg.action_import_from_excel_multisheet()
        self.assertEqual(setup['res_model'], 'hr.formula.multisheet.import.wizard',
                         "behaviour unchanged — only the words")
        self.assertIn('Column', setup['name'])
        self.assertNotIn('Import from Excel', setup['name'])

        history = self.env.ref('pb_hr_payroll_formula.action_payroll_import_batch')
        self.assertIn('Past', history.name)
        self.assertEqual(self.env.ref('pb_hr_payroll_formula.menu_import_batches').name,
                         'Past Pay Data Loads')
        self.assertEqual(self.env.ref('pb_hr_payroll_formula.menu_formula_import').name,
                         'Pay Data')

        views = _src('pb_hr_payroll_formula', 'views/formula_config_views.xml')
        self.assertIn('string="Set up columns from Excel"', views)
        self.assertNotIn('string="Import from Excel"', views)
        # the two stat buttons that used to share one ambiguous label
        for module, path in (
                ('pb_hr_payroll_formula', 'views/formula_config_views.xml'),
                ('pb_hr_payroll_formula', 'views/integration_views.xml')):
            src = _strip_xml_comments(_src(module, path))
            self.assertNotIn('>Payroll Import<', src)
            self.assertIn('Load pay data', src)

    # =====================================================================
    # 10 — the naming gates
    # =====================================================================
    def test_10_no_user_visible_odoo_or_studio_on_this_surface(self):
        tpl = _strip_xml_comments(
            _src('pb_formula_studio', 'static/src/xml/mapping_studio.xml'))
        js = _strip_js_comments(
            _src('pb_formula_studio', 'static/src/js/mapping/mapping_studio.js'))
        # `/** @odoo-module **/` and `@odoo/owl` are technical identifiers.
        js = js.replace('@odoo-module', '').replace('@odoo/owl', '')
        for src, what in ((tpl, 'mapping_studio.xml'), (js, 'mapping_studio.js')):
            self.assertNotIn('Odoo', src, 'white-label: %s' % what)
            # J1's own gate, and for the same reason: the technical ids
            # (`pb_formula_studio.MappingStudio`, `class MappingStudio`) keep
            # their names — what a person READS must not say "studio".
            for bad in ('Mapping Studio', 'mapping studio', 'Mapping canvas'):
                self.assertNotIn(bad, src,
                                 'the surface is called Mapping (J1): %s' % what)
        for module, path in (
                ('pb_hr_payroll_formula', 'views/menu_views.xml'),
                ('pb_hr_payroll_formula', 'views/payroll_import_views.xml'),
                ('pb_import', 'static/src/xml/import.xml')):
            self.assertNotIn('Odoo', _strip_xml_comments(_src(module, path)), path)
        self.assertNotIn(
            'Odoo', _strip_py_comments(_src('pb_import', 'models/pb_import.py')))

    # =====================================================================
    # 13 — the truth pass
    # =====================================================================
    def test_13_the_never_written_a_value_docstring_is_corrected(self):
        src = _src('pb_formula_studio', 'models/pb_formula_studio.py')
        head = src[src.index('def import_mapping_data'):][:3000]
        self.assertNotIn('it has never written a value', head)
        self.assertIn('it had never written a value', head,
                      "the sentence becomes past tense, not deleted — the "
                      "history is why the on-ramp exists")
        self.assertIn('FOUR lanes', head)

    def test_13b_one_generator_with_callers(self):
        conn = _src('pb_hr_payroll_formula', 'integrations/excel_connector.py')
        self.assertEqual(conn.count('def generate_template'), 1,
                         "exactly one template generator exists")
        cfgsrc = _src('pb_hr_payroll_formula', 'models/formula_config.py')
        self.assertIn('generate_template', cfgsrc, "…and it has a caller")
        studio = _src('pb_formula_studio', 'models/pb_formula_studio.py')
        self.assertIn('_build_pay_data_template', studio, "…reached from the board")
        # the loader no longer carries its own copy of the parse branch
        batch = _src('pb_hr_payroll_formula', 'models/payroll_import_batch.py')
        self.assertEqual(batch.count('_load_multisheet_data(file_content'), 1,
                         "one branch decision, in `_parse_source_file`")

    def test_13c_reading_headings_never_reaches_a_write_path(self):
        """The shape of the promise, asserted on the source.

        `peek_source_columns` is allowed to parse and nothing else. If it ever
        grows a `create(`, the row-count diff in test 2 would still pass for a
        record type nobody thought to count.
        """
        src = _src('pb_hr_payroll_formula', 'models/payroll_import_batch.py')
        body = src[src.index('def peek_source_columns'):]
        body = body[:body.index('\n    @api.model\n    def _sample_text')]
        for forbidden in ('.create(', '.write(', '.unlink(', 'action_process',
                          'action_match_employees'):
            self.assertNotIn(forbidden, body,
                             "the header reader must only ever read: %s" % forbidden)
        self.assertIn('.new(', body, "…on an in-memory probe record")
