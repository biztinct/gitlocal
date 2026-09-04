# -*- coding: utf-8 -*-
"""F112 — Post-calculation results grid.

Every employee's computed components for a pay run, in one Excel-style read-only
grid — server-paginated, filterable, with variance-vs-previous-run and one-click
.xlsx of the full filtered set. Data source is the `formula_computed_values` JSON
already stored on each slip (one parse per row, zero joins to hr.payslip.line).
RPC-only AbstractModel, mirroring pb.formula.studio.
"""
import base64
import io
import json
import logging

from odoo import _, api, models
from odoo.exceptions import AccessError
from odoo.addons.pb_hr_payroll_formula.formula_engine.comparison import coerce_number

_logger = logging.getLogger(__name__)

# Category colour palette (shared intent with F111's grid band strip). Assigned
# to categories by first appearance so the same config always colours the same.
_PALETTE = ['#4F46E5', '#0E7490', '#B45309', '#059669', '#7C3AED', '#DB2777',
            '#0891B2', '#CA8A04', '#DC2626', '#4338CA']


class PayrunResults(models.AbstractModel):
    _name = 'pb.payrun.results'
    _description = 'Pay Run Results Grid'

    PAGE = 100

    # Single source of truth for numeric coercion — handles the Excel-export
    # string forms ("1,234.5", "₫ 2,405,236", "12%") that plain float() drops,
    # so the grid/totals/export agree with Shadow Run and sample validation.
    _num = staticmethod(coerce_number)

    def _check_access(self):
        """Gate before the sudo escalation: the results grid exposes every
        employee's computed pay, so only payroll/formula staff may read it. A
        basic internal user calling the RPC directly is refused (the sibling
        pb.payslip.review cockpit relies on record rules; this one sudo's to
        read across employees, so it must gate explicitly)."""
        u = self.env.user
        if (u.has_group('base.group_system')
                or u.has_group('pb_hr_payroll_formula.group_formula_user')
                or u.has_group('pb_hr_payroll_formula.group_formula_manager')):
            return
        raise AccessError(_("You do not have access to payroll results."))

    # ------------------------------------------------------------------
    # column model
    # ------------------------------------------------------------------
    def _columns(self, configs):
        """Report columns = the on-payslip / report-visible components of every
        config present in the run, in display (sequence) order — after F111 that
        is the user's curated order, so the grid inherits grouping for free. A
        run that mixes configs (e.g. mid + end, or two divisions) contributes the
        UNION of columns, deduped by code, so no slip's values render blank."""
        cols, cat_color, seen, by_code = [], {}, 0, set()
        for config in configs:
            for r in config.rule_ids.sorted(key=lambda r: r.sequence):
                if not (r.appears_on_payslip or getattr(r, 'report_visible', False)):
                    continue
                code = r.code or r.column_letter or ''
                if not code or code in by_code:
                    continue
                by_code.add(code)
                cat = (r.category_id.name if r.category_id else (r.column_type or '').title()) or 'Other'
                if cat not in cat_color:
                    cat_color[cat] = _PALETTE[seen % len(_PALETTE)]
                    seen += 1
                cols.append({
                    'code': code,
                    'name': (r.salary_rule_id.name if r.salary_rule_id else False) or r.name or r.code or '',
                    'category': cat,
                    'color': cat_color[cat],
                    'number_format': r.number_format or 'currency',
                })
        return cols

    # ------------------------------------------------------------------
    # run discovery + filtering
    # ------------------------------------------------------------------
    def _companies(self):
        # active/selected companies (survives sudo — sudo changes user, not the
        # company context). hr.payslip.run has no company_id in this build, so
        # every run/slip query scopes via hr.payslip.company_id instead.
        return self.env.companies.ids or [self.env.company.id]

    def _formula_runs(self):
        """Pay-slip runs that have at least one formula-computed slip (any run
        whose slips carry a formula_config_id — not just the ones with a
        formula_computed_values JSON blob), scoped to the active company, newest
        first. One grouped query — not a per-run probe."""
        groups = self.env['hr.payslip'].read_group(
            [('formula_config_id', '!=', False), ('payslip_run_id', '!=', False),
             ('company_id', 'in', self._companies())],
            ['payslip_run_id'], ['payslip_run_id'],
            orderby='payslip_run_id desc', limit=60)
        out = []
        for g in groups:
            rid = g['payslip_run_id'][0] if g.get('payslip_run_id') else False
            rname = g['payslip_run_id'][1] if g.get('payslip_run_id') else ''
            if rid:
                out.append({'id': rid, 'name': rname or _('Run #%s') % rid})
        return out

    def _apply_filters(self, slips, f):
        dept = f.get('department_id')
        if dept:
            slips = slips.filtered(lambda s: s.employee_id.department_id.id == int(dept))
        div = f.get('division')
        if div and 'pb_division' in self.env['hr.employee']._fields:
            slips = slips.filtered(lambda s: (s.employee_id.pb_division or '') == div)
        q = (f.get('search') or '').strip().lower()
        if q:
            slips = slips.filtered(
                lambda s: q in (s.employee_id.name or '').lower()
                or q in (s.employee_id.barcode or '').lower())
        return slips

    def _slips_for(self, run):
        return run.slip_ids.filtered(lambda s: s.formula_config_id)

    def _slip_values(self, slip):
        """One value dict per slip, keyed by component code. Two sources, one
        contract: the imported/VPTQ runs stash every computed code in the
        `formula_computed_values` JSON blob; the runs actually produced by the
        engine (the demo world, ~all live runs) leave that empty and store their
        figures on `hr.payslip.line` instead — whose `code` matches the config
        rule codes 1:1 (BASIC…GROSS…NET). Prefer the blob, else read the lines,
        so the grid renders every already-run payrun, not just the JSON ones."""
        if slip.formula_computed_values:
            try:
                return json.loads(slip.formula_computed_values or '{}')
            except Exception:
                return {}
        return {ln.code: ln.total for ln in slip.line_ids if ln.code}

    # ------------------------------------------------------------------
    # variance (D112.4): previous slip = latest earlier, same employee + cycle
    # ------------------------------------------------------------------
    def _pair_variance(self, page_slips, codes, parsed):
        """Previous slip = the latest earlier slip of the SAME employee AND the
        SAME formula config (like-for-like: a config carries one cycle_type, so
        this never compares a mid-cycle advance against an end-cycle salary).
        Cancelled slips are excluded so a voided run can't become the baseline."""
        if not page_slips:
            return {}
        emp_ids = page_slips.mapped('employee_id').ids
        cfg_ids = page_slips.mapped('formula_config_id').ids
        min_from = min(page_slips.mapped('date_from'))
        dom = [('employee_id', 'in', emp_ids),
               ('formula_config_id', 'in', cfg_ids),
               ('date_to', '<', min_from),
               ('formula_config_id', '!=', False),
               ('state', 'not in', ('cancel',))]
        prev = self.env['hr.payslip'].search(
            dom, order='employee_id, formula_config_id, date_to desc')
        latest = {}
        for p in prev:
            latest.setdefault((p.employee_id.id, p.formula_config_id.id), p)
        out = {}
        for s in page_slips:
            p = latest.get((s.employee_id.id, s.formula_config_id.id))
            if not p:
                continue
            pv = self._slip_values(p)
            cv = parsed.get(s.id, {})
            row = {}
            for c in codes:
                a, b = self._num(cv.get(c)), self._num(pv.get(c))
                if a is not None and b is not None:
                    row[c] = a - b
            out[s.id] = row
        return out

    # ------------------------------------------------------------------
    # the grid RPC
    # ------------------------------------------------------------------
    @api.model
    def get_grid(self, run_id=None, filters=None):
        self._check_access()
        self = self.sudo()   # gated above; sudo to read across all employees
        f = filters or {}
        runs = self._formula_runs()
        if not run_id:
            run_id = runs[0]['id'] if runs else False
        if not run_id:
            return {'ok': True, 'runs': runs, 'empty_reason': _('No pay runs found.'),
                    'columns': [], 'rows': [], 'totals': {}, 'row_count': 0,
                    'page': 1, 'page_count': 1}
        run = self.env['hr.payslip.run'].browse(int(run_id))
        slips = self._slips_for(run)
        configs = slips.mapped('formula_config_id')
        if not slips or not configs:
            return {'ok': True, 'runs': runs,
                    'run': {'id': run.id, 'name': run.name or '', 'state': run.state},
                    'empty_reason': _('This run has no formula-calculated payslips.'),
                    'columns': [], 'rows': [], 'totals': {}, 'row_count': 0,
                    'page': 1, 'page_count': 1}

        cols = self._columns(configs)
        codes = [c['code'] for c in cols]
        slips = self._apply_filters(slips, f)

        # Totals over the FULL filtered set (before paging). Line-based runs (the
        # engine-produced ones) sum in ONE grouped SQL query instead of parsing
        # every slip — a 900-employee run would otherwise walk ~23k lines. The
        # JSON-blob runs (tiny imports) keep the per-slip coercion path, since
        # their values are strings ("₫ 2,405,236", "12%") only `_num` decodes.
        use_fcv = bool(slips and slips[0].formula_computed_values)
        totals = dict.fromkeys(codes, 0.0)
        if use_fcv:
            for s in slips:
                v = self._slip_values(s)
                for c in codes:
                    n = self._num(v.get(c))
                    if n is not None:
                        totals[c] += n
        elif slips:
            code_set = set(codes)
            for g in self.env['hr.payslip.line'].read_group(
                    [('slip_id', 'in', slips.ids), ('code', 'in', codes)],
                    ['total:sum'], ['code']):
                if g.get('code') in code_set:
                    totals[g['code']] = g.get('total') or 0.0

        row_count = len(slips)
        page = max(1, int(f.get('page') or 1))
        ordered = slips.sorted(key=lambda s: (s.employee_id.name or '').lower())
        page_slips = ordered[(page - 1) * self.PAGE: page * self.PAGE]

        # Values only for the current page (not the whole filtered set).
        parsed = {s.id: self._slip_values(s) for s in page_slips}

        deltas = self._pair_variance(page_slips, codes, parsed) if f.get('with_variance') else {}
        prev_run_label = None
        if f.get('with_variance') and deltas:
            prev_run_label = _('previous slip of the same config, per employee')

        rows = [{
            'slip_id': s.id,
            'employee_id': s.employee_id.id,
            'employee_name': s.employee_id.name or '',
            'employee_code': s.employee_id.barcode or '',
            'department': s.employee_id.department_id.name or '',
            'values': {c: parsed[s.id].get(c) for c in codes},
            'deltas': deltas.get(s.id),
        } for s in page_slips]

        primary = configs[:1]
        return {
            'ok': True,
            'run': {'id': run.id, 'name': run.name or '', 'state': run.state,
                    'config_name': ', '.join(configs.mapped('name')),
                    'currency': (primary.currency_id.symbol or '₫')},
            'runs': runs,
            'columns': cols,
            'rows': rows,
            'totals': totals,
            'page': page,
            'page_count': max(1, -(-row_count // self.PAGE)),
            'row_count': row_count,
            'prev_run_label': prev_run_label,
            'departments': self._departments(self._slips_for(run)),
        }

    def _departments(self, slips):
        seen = {}
        for s in slips:
            d = s.employee_id.department_id
            if d and d.id not in seen:
                seen[d.id] = d.name
        return [{'id': k, 'name': v} for k, v in sorted(seen.items(), key=lambda kv: kv[1] or '')]

    # ------------------------------------------------------------------
    # F112b — pay-run picker (rich, filterable gallery that replaces the
    # silent "newest run" default). One grouped query for the candidate set,
    # then everything the cards show is read from STORED hr.payslip.run fields
    # (pb_employee_count / pb_total_* / pb_currency_id / pb_division_label) —
    # zero per-slip aggregation, so opening the picker is as cheap as a kanban.
    # ------------------------------------------------------------------
    _STATE_META = {
        'draft':  ('Draft', 'draft'),
        'verify': ('Waiting', 'review'),
        'level1': ('HR Review', 'review'),
        'level2': ('GM Review', 'review'),
        'done':   ('Approved', 'done'),
        'close':  ('Closed', 'done'),
        'paid':   ('Paid', 'done'),
    }
    _CYCLE_LABELS = {
        'end_cycle': 'End-cycle', 'mid_cycle': 'Mid-cycle',
        'end': 'End-cycle', 'mid': 'Mid-cycle',
    }

    def _state_label(self, st):
        return self._STATE_META.get(st, (st and st.replace('_', ' ').title() or '', 'draft'))

    def _run_card(self, r, cycle_type='', comp=None, symbol=u'₫'):
        label, tone = self._state_label(r.state)
        ds = str(r.date_start) if r.date_start else ''
        return {
            'id': r.id,
            'name': r.name or _('Run #%s') % r.id,
            'state': r.state or '',
            'state_label': label,
            'state_tone': tone,
            'date_start': ds,
            'date_end': str(r.date_end) if r.date_end else '',
            'year': ds[:4] if ds else '',
            'employees': r.pb_employee_count or 0,
            'net': r.pb_total_net or 0.0,
            'gross': r.pb_total_gross or 0.0,
            'deductions': r.pb_total_deductions or 0.0,
            'currency': symbol,
            'division': r.pb_division or '',
            'division_label': r.pb_division_label or '',
            'cycle_type': cycle_type or '',
            'cycle_label': self._CYCLE_LABELS.get(
                cycle_type, cycle_type.replace('_', ' ').title() if cycle_type else ''),
            'company': comp[1] if comp else '',
            'company_id': comp[0] if comp else False,
        }

    @api.model
    def list_runs(self, filters=None):
        """Feed the picker: EVERY formula-computed pay run of the active company
        as a rich card, in one shot. Filtering/sorting/faceting all happen
        client-side (like the Config Switcher), so the payload is just the cards
        plus the run→cycle and run→company maps the facet chips need."""
        self._check_access()
        self = self.sudo()
        companies = self._companies()
        groups = self.env['hr.payslip'].read_group(
            [('formula_config_id', '!=', False), ('payslip_run_id', '!=', False),
             ('company_id', 'in', companies)],
            ['payslip_run_id'], ['payslip_run_id'],
            orderby='payslip_run_id desc', limit=300)
        run_ids = [g['payslip_run_id'][0] for g in groups if g.get('payslip_run_id')]
        runs = self.env['hr.payslip.run'].browse(run_ids).exists()

        # run -> representative config + company, in ONE grouped query. The run
        # model has no company_id in this build, so company comes from the slips.
        run_cfg, run_comp = {}, {}
        for g in self.env['hr.payslip'].read_group(
                [('payslip_run_id', 'in', run_ids), ('formula_config_id', '!=', False),
                 ('company_id', 'in', companies)],
                [], ['payslip_run_id', 'formula_config_id', 'company_id'], lazy=False):
            rid = g['payslip_run_id'][0] if g.get('payslip_run_id') else False
            if not rid:
                continue
            if rid not in run_cfg and g.get('formula_config_id'):
                run_cfg[rid] = g['formula_config_id'][0]
            if rid not in run_comp and g.get('company_id'):
                run_comp[rid] = (g['company_id'][0], g['company_id'][1])
        cfg_cycle = {}
        if run_cfg:
            configs = self.env['hr.formula.config'].browse(list(set(run_cfg.values()))).exists()
            cfg_cycle = {c.id: (getattr(c, 'cycle_type', '') or '') for c in configs}

        # currency symbol per run comes from the run's COMPANY currency (the run
        # has no company_id, and its stored pb_currency_id can be stale — some
        # demo runs computed under a different company context), so a single VN
        # company shows ₫ consistently and multi-company shows each currency.
        comp_sym = {}
        comp_ids = list({c[0] for c in run_comp.values() if c})
        if comp_ids:
            for co in self.env['res.company'].browse(comp_ids).exists():
                comp_sym[co.id] = (co.currency_id.symbol or u'₫')
        default_sym = self.env.company.currency_id.symbol or u'₫'

        cards = []
        for r in runs:
            comp = run_comp.get(r.id)
            sym = comp_sym.get(comp[0], default_sym) if comp else default_sym
            cards.append(self._run_card(r, cfg_cycle.get(run_cfg.get(r.id), ''), comp, sym))
        return {
            'ok': True,
            'runs': cards,
            'total': len(cards),
            'active_company': self.env.company.name,
            'multi_company': len(companies) > 1,
        }

    # ------------------------------------------------------------------
    # xlsx export (D112.6) — full filtered set, blob pattern
    # ------------------------------------------------------------------
    def _excel_fmt(self, nf):
        return {'currency': '#,##0', 'integer': '#,##0', 'number': '#,##0.00',
                'percentage': '0.00%'}.get(nf, '#,##0')

    @api.model
    def export_grid(self, run_id, filters=None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            return {'ok': False, 'msg': _('openpyxl is not available on the server.')}
        self._check_access()
        self = self.sudo()
        f = dict(filters or {})
        f.pop('page', None)
        run = self.env['hr.payslip.run'].browse(int(run_id))
        slips = self._apply_filters(self._slips_for(run), f)
        configs = slips.mapped('formula_config_id')
        if not slips or not configs:
            return {'ok': False, 'msg': _('Nothing to export for this run.')}
        cols = self._columns(configs)
        codes = [c['code'] for c in cols]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (run.name or 'Results')[:31]
        ws.freeze_panes = 'B2'
        bold = Font(bold=True)
        # header
        h = ws.cell(1, 1, _('Employee')); h.font = bold
        ws.cell(1, 2, _('Dept')).font = bold
        for j, col in enumerate(cols, start=3):
            c = ws.cell(1, j, '%s (%s)' % (col['name'], col['code']))
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=col['color'].lstrip('#'))
            c.alignment = Alignment(horizontal='center')
        # rows
        totals = dict.fromkeys(codes, 0.0)
        i = 2
        for s in slips.sorted(key=lambda s: (s.employee_id.name or '').lower()):
            v = self._slip_values(s)
            ws.cell(i, 1, s.employee_id.name or '')
            ws.cell(i, 2, s.employee_id.department_id.name or '')
            for j, col in enumerate(cols, start=3):
                n = self._num(v.get(col['code']))
                if n is not None:
                    cell = ws.cell(i, j, n)
                    cell.number_format = self._excel_fmt(col['number_format'])
                    totals[col['code']] += n
                else:
                    ws.cell(i, j, v.get(col['code']) or '')   # keep text values as-is
            i += 1
        # totals row
        tr = ws.cell(i, 1, _('TOTAL')); tr.font = bold
        for j, col in enumerate(cols, start=3):
            cell = ws.cell(i, j, totals[col['code']])
            cell.font = bold
            cell.number_format = self._excel_fmt(col['number_format'])
        # widths
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 20
        for j in range(3, len(cols) + 3):
            ws.column_dimensions[get_column_letter(j)].width = 15

        out = io.BytesIO(); wb.save(out); out.seek(0)
        name = (run.name or 'payrun').strip().replace(' ', '_')
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'filename': 'results_%s.xlsx' % name,
        }
