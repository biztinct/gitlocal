# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError
import base64
import csv
import io
import json
import logging
from datetime import datetime

_logger = logging.getLogger(__name__)


class PayrollImportWizard(models.TransientModel):
    _name = 'payroll.import.wizard'
    _description = 'Payroll Data Import Wizard'

    # Import Configuration
    country_code = fields.Selection([
        ('VN', 'Vietnam'),
        ('ID', 'Indonesia'),
        ('IN', 'India'),
        ('SG', 'Singapore'),
        ('MY', 'Malaysia'),
        ('TH', 'Thailand'),
        ('PH', 'Philippines'),
    ], string='Country', required=True)
    
    import_type = fields.Selection([
        ('payslips', 'Payslip Data'),
        ('salary_components', 'Salary Components'),
        ('deductions', 'Deductions'),
        ('bonuses', 'Bonuses'),
        ('adjustments', 'Adjustments'),
    ], default='payslips', string='Import Type')
    
    period_start = fields.Date('Period Start', required=True)
    period_end = fields.Date('Period End', required=True)

    # File Upload
    import_file = fields.Binary('Import File', required=True)
    import_filename = fields.Char('Filename')
    file_format = fields.Selection([
        ('csv', 'CSV'),
        ('xlsx', 'Excel'),
        ('json', 'JSON'),
    ], default='csv', string='File Format')
    has_header_row = fields.Boolean('Has Header Row', default=True)

    # Processing Options
    create_missing_employees = fields.Boolean('Create Missing Employees', default=False)
    update_existing_payslips = fields.Boolean('Update Existing Payslips', default=True)
    validate_data = fields.Boolean('Validate Data', default=True)
    auto_confirm_payslips = fields.Boolean('Auto Confirm Payslips', default=False)

    # Error Handling
    skip_errors = fields.Boolean('Skip Errors', default=True)
    error_threshold = fields.Integer('Error Threshold (%)', default=10)
    send_error_report = fields.Boolean('Send Error Report', default=True)
    error_email_recipients = fields.Many2many('res.users', string='Error Report Recipients')

    # Field Mapping
    employee_number_column = fields.Char('Employee Number Column', default='employee_number')
    employee_name_column = fields.Char('Employee Name Column', default='employee_name')
    employee_email_column = fields.Char('Employee Email Column', default='email')
    department_column = fields.Char('Department Column', default='department')
    basic_salary_column = fields.Char('Basic Salary Column', default='basic_salary')
    gross_salary_column = fields.Char('Gross Salary Column', default='gross_salary')
    net_salary_column = fields.Char('Net Salary Column', default='net_salary')
    deductions_column = fields.Char('Deductions Column', default='deductions')

    field_mapping_ids = fields.One2many('field.mapping', 'wizard_id', string='Field Mappings')

    # Data Validation
    validate_employee_exists = fields.Boolean('Validate Employee Exists', default=True)
    validate_salary_positive = fields.Boolean('Validate Salary Positive', default=True)
    validate_dates = fields.Boolean('Validate Dates', default=True)
    validate_currency = fields.Boolean('Validate Currency', default=True)
    custom_validation_rules = fields.Text('Custom Validation Rules')

    # Preview Data
    preview_data = fields.Text('Preview Data', readonly=True)
    preview_employee_count = fields.Integer('Preview Employee Count', readonly=True)
    preview_total_amount = fields.Monetary('Preview Total Amount', readonly=True)
    preview_validation_errors = fields.Integer('Preview Validation Errors', readonly=True)
    currency_id = fields.Many2one('res.currency', string='Currency')

    # Processing Results
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('completed', 'Completed'),
        ('error', 'Error'),
    ], default='draft', string='State')
    
    processing_log = fields.Text('Processing Log', readonly=True)
    processed_records = fields.Integer('Processed Records', readonly=True)
    failed_records = fields.Integer('Failed Records', readonly=True)
    processing_duration = fields.Float('Processing Duration (seconds)', readonly=True)

    @api.onchange('country_code')
    def _onchange_country_code(self):
        """Set currency based on country"""
        if self.country_code:
            currency_map = {
                'VN': 'VND', 'ID': 'IDR', 'IN': 'INR',
                'SG': 'SGD', 'MY': 'MYR', 'TH': 'THB', 'PH': 'PHP'
            }
            currency_code = currency_map.get(self.country_code)
            if currency_code:
                currency = self.env['res.currency'].search([('name', '=', currency_code)], limit=1)
                self.currency_id = currency.id if currency else self.env.company.currency_id.id

    def action_preview_import(self):
        """Load preview of import data"""
        self.ensure_one()
        
        try:
            if not self.import_file:
                raise UserError(_('Please upload a file first.'))
            
            # Decode file
            file_data = base64.b64decode(self.import_file)
            
            if self.file_format == 'csv':
                preview_data = self._preview_csv_data(file_data)
            elif self.file_format == 'xlsx':
                preview_data = self._preview_xlsx_data(file_data)
            elif self.file_format == 'json':
                preview_data = self._preview_json_data(file_data)
            else:
                raise UserError(_('Unsupported file format.'))
            
            # Store preview data
            self.preview_data = json.dumps(preview_data, indent=2)
            
            # Calculate preview statistics
            self._calculate_preview_stats(preview_data)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Preview Loaded'),
                    'message': _('Preview loaded successfully. Check the Preview tab for details.'),
                    'type': 'success',
                }
            }
            
        except Exception as e:
            raise UserError(_('Error loading preview: %s') % str(e))

    def _preview_csv_data(self, file_data):
        """Preview CSV data"""
        csv_data = file_data.decode('utf-8')
        reader = csv.DictReader(io.StringIO(csv_data))
        
        preview_rows = []
        for i, row in enumerate(reader):
            if i >= 10:  # Only preview first 10 rows
                break
            preview_rows.append(row)
        
        return preview_rows

    def _preview_xlsx_data(self, file_data):
        """Preview Excel data"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            preview_rows = []
            for row_num in range(2, min(12, sheet.max_row + 1)):  # Preview first 10 data rows
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                preview_rows.append(row_data)
            
            return preview_rows
            
        except ImportError:
            raise UserError(_('openpyxl library is required for Excel file processing.'))

    def _preview_json_data(self, file_data):
        """Preview JSON data"""
        json_data = json.loads(file_data.decode('utf-8'))
        
        if isinstance(json_data, list):
            return json_data[:10]  # First 10 records
        else:
            return [json_data]  # Single record

    def _calculate_preview_stats(self, preview_data):
        """Calculate preview statistics"""
        if not preview_data:
            return
        
        self.preview_employee_count = len(set(
            row.get(self.employee_number_column, '') for row in preview_data
        ))
        
        # Calculate total amount from gross salary column
        total_amount = 0
        for row in preview_data:
            try:
                amount = float(row.get(self.gross_salary_column, 0) or 0)
                total_amount += amount
            except (ValueError, TypeError):
                continue
        
        self.preview_total_amount = total_amount
        
        # Basic validation check
        errors = 0
        for row in preview_data:
            if not row.get(self.employee_number_column):
                errors += 1
        
        self.preview_validation_errors = errors

    def action_start_import(self):
        """Start the import process"""
        self.ensure_one()
        
        try:
            self.state = 'processing'
            start_time = datetime.now()
            
            # Decode file
            file_data = base64.b64decode(self.import_file)
            
            # Parse data based on format
            if self.file_format == 'csv':
                data_rows = self._parse_csv_data(file_data)
            elif self.file_format == 'xlsx':
                data_rows = self._parse_xlsx_data(file_data)
            elif self.file_format == 'json':
                data_rows = self._parse_json_data(file_data)
            else:
                raise UserError(_('Unsupported file format.'))
            
            # Process the data
            results = self._process_import_data(data_rows)
            
            # Update processing results
            end_time = datetime.now()
            self.processing_duration = (end_time - start_time).total_seconds()
            self.processed_records = results['processed']
            self.failed_records = results['failed']
            self.processing_log = results['log']
            self.state = 'completed'
            
            # Send error report if needed
            if self.send_error_report and results['failed'] > 0:
                self._send_error_report(results)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Completed'),
                    'message': _('Processed %d records, %d failed.') % (
                        results['processed'], results['failed']
                    ),
                    'type': 'success' if results['failed'] == 0 else 'warning',
                }
            }
            
        except Exception as e:
            self.state = 'error'
            self.processing_log = str(e)
            _logger.error(f"Error in payroll import: {str(e)}")
            raise UserError(_('Import failed: %s') % str(e))

    def _parse_csv_data(self, file_data):
        """Parse CSV data"""
        csv_data = file_data.decode('utf-8')
        reader = csv.DictReader(io.StringIO(csv_data))
        return list(reader)

    def _parse_xlsx_data(self, file_data):
        """Parse Excel data"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            data_rows = []
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                data_rows.append(row_data)
            
            return data_rows
            
        except ImportError:
            raise UserError(_('openpyxl library is required for Excel file processing.'))

    def _parse_json_data(self, file_data):
        """Parse JSON data"""
        json_data = json.loads(file_data.decode('utf-8'))
        
        if isinstance(json_data, list):
            return json_data
        else:
            return [json_data]

    def _process_import_data(self, data_rows):
        """Process the imported data"""
        processed = 0
        failed = 0
        log_entries = []
        
        for i, row in enumerate(data_rows, 1):
            try:
                # Validate row data
                if self.validate_data:
                    validation_errors = self._validate_row(row)
                    if validation_errors:
                        if not self.skip_errors:
                            raise UserError(_('Validation errors in row %d: %s') % (i, validation_errors))
                        else:
                            failed += 1
                            log_entries.append(f"Row {i}: Validation errors - {validation_errors}")
                            continue
                
                # Process the row based on import type
                if self.import_type == 'payslips':
                    self._process_payslip_row(row)
                elif self.import_type == 'salary_components':
                    self._process_salary_component_row(row)
                # Add other import types as needed
                
                processed += 1
                
                # Check error threshold
                if failed > 0 and (failed / (processed + failed)) * 100 > self.error_threshold:
                    raise UserError(_('Error threshold exceeded. Import stopped.'))
                
            except Exception as e:
                failed += 1
                error_msg = f"Row {i}: {str(e)}"
                log_entries.append(error_msg)
                
                if not self.skip_errors:
                    raise UserError(error_msg)
        
        return {
            'processed': processed,
            'failed': failed,
            'log': '\n'.join(log_entries)
        }

    def _validate_row(self, row):
        """Validate a single row of data"""
        errors = []
        
        # Check required fields
        if not row.get(self.employee_number_column):
            errors.append('Missing employee number')
        
        # Validate employee exists
        if self.validate_employee_exists and row.get(self.employee_number_column):
            employee = self.env['hr.employee'].search([
                ('employee_number', '=', row[self.employee_number_column])
            ], limit=1)
            if not employee:
                errors.append('Employee not found')
        
        # Validate salary is positive
        if self.validate_salary_positive:
            try:
                salary = float(row.get(self.gross_salary_column, 0) or 0)
                if salary < 0:
                    errors.append('Negative salary not allowed')
            except (ValueError, TypeError):
                errors.append('Invalid salary format')
        
        return '; '.join(errors)

    def _process_payslip_row(self, row):
        """Process a single payslip row"""
        # Get employee
        employee = self.env['hr.employee'].search([
            ('employee_number', '=', row[self.employee_number_column])
        ], limit=1)
        
        if not employee:
            if self.create_missing_employees:
                employee = self._create_employee_from_row(row)
            else:
                raise UserError(_('Employee %s not found') % row[self.employee_number_column])
        
        # Check if payslip already exists
        existing_payslip = self.env['hr.payslip'].search([
            ('employee_id', '=', employee.id),
            ('date_from', '=', self.period_start),
            ('date_to', '=', self.period_end),
        ], limit=1)
        
        if existing_pay