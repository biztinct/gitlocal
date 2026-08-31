# -*- coding: utf-8 -*-
"""`pb.assets` — the Assets board's only server surface.

The shape is the one `pb_people` established and `pb_journeys` repeated: an
`AbstractModel` facade, `@api.model` reads, every independent number inside its
own `_safe()` so one failing metric answers zero instead of taking the screen
down, `self.env.companies` scoping on every search, a row cap, and no sudo in a
read.

The gate is SERVER-SIDE and it is the boundary. A reader with no asset group
gets an EMPTY BOARD with `allowed: false` rather than an access dialog, so the
screen can say in words what it is and who to ask.
"""

import base64
import io
import logging
import re
import unicodedata
from datetime import date

from odoo import api, models, _
from odoo.exceptions import AccessError, UserError

from .asset_common import (
    ASSET_KINDS, ASSET_KIND_LABEL, ASSET_STATES, FULFILMENT_LABEL,
    GROUP_MANAGER, GROUP_USER, REQUEST_STATE_LABEL, state_label, states_for,
)

_logger = logging.getLogger(__name__)

BOARD_LIMIT = 400
TIMELINE_LIMIT = 60
XLSX_MIME = ('application/vnd.openxmlformats-officedocument.'
             'spreadsheetml.sheet')

#: The colour a state wears, everywhere it appears.
STATE_TONE = {
    'spare': 'info',
    'assigned': 'ok',
    'repair': 'warn',
    'to_scrap': 'warn',
    'scrapped': 'muted',
    'deactivated': 'muted',
}


def _initials(name):
    parts = [p for p in (name or '').replace('-', ' ').split() if p]
    return ((parts[0][0] if parts else '?')
            + (parts[-1][0] if len(parts) > 1 else '')).upper()


class PbAssetsBoard(models.AbstractModel):
    _name = 'pb.assets'
    _description = 'Payobook Assets cockpit data'

    # ------------------------------------------------------------------ gates
    @api.model
    def _safe(self, fn, default=0):
        try:
            return fn()
        except Exception as e:      # noqa: BLE001
            _logger.debug('Assets cockpit metric failed: %s', e)
            return default

    @api.model
    def _can_read(self):
        user = self.env.user
        return (user.has_group(GROUP_USER) or user.has_group(GROUP_MANAGER)
                or user._is_admin())

    @api.model
    def _can_write(self):
        user = self.env.user
        return user.has_group(GROUP_MANAGER) or user._is_admin()

    @api.model
    def _require_write(self):
        if not self._can_write():
            raise AccessError(_(
                "You can look at the register, but changing it is for the "
                "asset team. Ask them to make the change."))
        return True

    @api.model
    def _require_read(self):
        if not self._can_read():
            raise AccessError(_(
                "The asset register is looked after by the HR and IT team. "
                "Ask them for access."))
        return True

    # ------------------------------------------------------------- the board
    @api.model
    def get_board(self):
        if not self._can_read():
            return {
                'allowed': False, 'can_write': False, 'kpis': {}, 'rows': [],
                'categories': [], 'countries': [], 'states': [],
                'kinds': [], 'facets': {}, 'requests': {},
            }
        Asset = self.env['pb.asset']
        co_ids = self.env.companies.ids or [self.env.company.id]
        CO = ['|', ('company_id', '=', False), ('company_id', 'in', co_ids)]

        kpis = {
            'total': self._safe(lambda: Asset.search_count(CO)),
            'assigned': self._safe(lambda: Asset.search_count(
                CO + [('state', '=', 'assigned'),
                      ('kind', '=', 'tangible')])),
            'spare': self._safe(lambda: Asset.search_count(
                CO + [('state', '=', 'spare'), ('kind', '=', 'tangible')])),
            'repair': self._safe(lambda: Asset.search_count(
                CO + [('state', '=', 'repair')])),
            'digital_live': self._safe(lambda: Asset.search_count(
                CO + [('state', '=', 'assigned'), ('kind', '=', 'digital')])),
            'leavers_holding': self._safe(self._leavers_holding),
            'open_requests': self._safe(lambda: self.env[
                'pb.asset.request'].search_count(
                    CO + [('state', 'in', ('draft', 'submitted',
                                           'manager_approved'))])),
        }

        assets = self._safe(
            lambda: Asset.search(CO, order='code desc, id desc',
                                 limit=BOARD_LIMIT),
            default=Asset.browse())
        rows = []
        for asset in assets:
            try:
                rows.append(self._asset_row(asset))
            except Exception:       # noqa: BLE001
                _logger.exception('Assets board row for asset %s', asset.id)

        facets = {'category': {}, 'country': {}, 'state': {}, 'kind': {}}
        for row in rows:
            for key, value in (('category', row['category_id']),
                               ('country', row['country_id']),
                               ('state', row['state']),
                               ('kind', row['kind'])):
                facets[key][value] = facets[key].get(value, 0) + 1

        return {
            'allowed': True,
            'can_write': self._can_write(),
            'kpis': kpis,
            'rows': rows,
            'total': len(rows),
            'capped': len(rows) >= BOARD_LIMIT,
            'facets': facets,
            'categories': self._categories(),
            'countries': self._countries(),
            # The FILTER bar gets the short list, the Add dialog gets the whole
            # world: filtering by a country that holds nothing is noise, but the
            # first laptop in a new country has to be addable or the register
            # can never grow past the office it started in.
            'countries_all': self._safe(self._countries_all, default=[]),
            # Where the reader works. A new item is almost always in their own
            # office, and an alphabetical list would default them to Afghanistan.
            'home_country_id': (self.env.company.country_id.id
                                if self.env.company.country_id else 0),
            'kinds': [{'id': k, 'label': _(v)} for k, v in ASSET_KINDS],
            'states': [{'id': k, 'label': _(v),
                        'tone': STATE_TONE.get(k, 'muted')}
                       for k, v in ASSET_STATES],
            'requests': self._safe(self._request_digest, default={}),
        }

    @api.model
    def _leavers_holding(self):
        """People on their way out who still have something of ours.

        "On their way out" is read from TWO places, because they answer at
        different moments. The employee record says so once HR has entered a
        leaving date or archived the person; the leaving CHECKLIST says so the
        day somebody resigns, which is weeks earlier and is exactly when this
        number is worth looking at.
        """
        co_ids = self.env.companies.ids or [self.env.company.id]
        open_rows = self.env['pb.asset.assignment'].search([
            '|', ('company_id', '=', False), ('company_id', 'in', co_ids),
            ('state', '=', 'open'),
        ])
        employees = open_rows.mapped('employee_id')
        if not employees:
            return 0
        leaving_ids = set()
        try:
            cases = self.env['pb.journey.case'].search([
                ('employee_id', 'in', employees.ids),
                ('case_type', '=', 'offboarding'),
                ('state', 'in', ('draft', 'active', 'on_hold')),
            ])
            leaving_ids = set(cases.mapped('employee_id').ids)
        except Exception:           # noqa: BLE001 — the journeys are optional
            _logger.debug('Assets: could not read the leaving checklists')
        leaving = employees.filtered(
            lambda e: e.id in leaving_ids or not e.active
            or getattr(e, 'departure_date', False))
        return len(leaving)

    @api.model
    def _asset_row(self, asset):
        employee = asset.current_employee_id
        return {
            'id': asset.id,
            'code': asset.code or '—',
            'name': asset.name or '',
            'category': asset.category_id.name or '',
            'category_id': asset.category_id.id,
            'icon': asset.category_id.icon or 'package',
            'kind': asset.kind or 'tangible',
            'kind_label': _(ASSET_KIND_LABEL.get(asset.kind, '')),
            'state': asset.state,
            'state_label': state_label(asset.kind, asset.state),
            'tone': STATE_TONE.get(asset.state, 'muted'),
            'employee': employee.name or '',
            'employee_id': employee.id or 0,
            'initials': _initials(employee.name) if employee else '',
            'country': asset.country_id.name or '',
            'country_id': asset.country_id.id,
            'serial': asset.serial or '',
            'model_name': asset.model_name or '',
            'cost': asset.cost or 0.0,
            'currency': asset.currency_id.symbol or
            asset.currency_id.name or '',
            'cost_usd': asset.cost_usd or 0.0,
            'warranty': asset.warranty_state,
            'warranty_end': str(asset.warranty_end)
            if asset.warranty_end else '',
            'reused': asset.is_reused,
        }

    @api.model
    def _categories(self):
        co_ids = self.env.companies.ids or [self.env.company.id]
        cats = self.env['pb.asset.category'].search(
            ['|', ('company_id', '=', False), ('company_id', 'in', co_ids)],
            order='sequence, name')
        return [{'id': c.id, 'name': c.name, 'code': c.code, 'kind': c.kind,
                 'icon': c.icon or 'package',
                 'auto': c.auto_assign_at_joining} for c in cats]

    @api.model
    def _countries(self):
        """The countries already in play, plus the reader's own — never the
        whole world list, which is 250 rows of noise on a filter bar."""
        # `_read_group` and not `read_group`: the old signature is gone on
        # Odoo 19, and the fallback is a plain mapped() so a build that changes
        # it again still answers a list rather than an exception.
        try:
            groups = self.env['pb.asset']._read_group([], ['country_id'], [])
            ids = [g[0].id for g in groups if g and g[0]]
        except Exception:           # noqa: BLE001
            ids = self.env['pb.asset'].search(
                [], limit=BOARD_LIMIT).mapped('country_id').ids
        own = [c.country_id.id for c in self.env.companies
               if c.country_id]
        countries = self.env['res.country'].browse(
            list(dict.fromkeys(ids + own))).exists()
        return [{'id': c.id, 'name': c.name, 'code': c.code}
                for c in countries.sorted(key=lambda c: c.name or '')]

    @api.model
    def _countries_all(self):
        """Every country, for the Add dialog only."""
        return [{'id': c.id, 'name': c.name, 'code': c.code}
                for c in self.env['res.country'].search([], order='name')]

    @api.model
    def _request_digest(self):
        co_ids = self.env.companies.ids or [self.env.company.id]
        Request = self.env['pb.asset.request']
        CO = ['|', ('company_id', '=', False), ('company_id', 'in', co_ids)]
        rows = Request.search(
            CO + [('state', 'not in', ('cancelled',))],
            order='needed_by asc, id desc', limit=60)
        return {
            'count': len(rows),
            'rows': [{
                'id': r.id,
                'name': r.name,
                'employee': r.employee_id.name or '',
                'category': r.category_id.name or '',
                'needed_by': str(r.needed_by) if r.needed_by else '',
                'state': r.state,
                'state_label': _(REQUEST_STATE_LABEL.get(r.state, r.state)),
                'fulfilment': r.fulfilment,
                'fulfilment_label': _(FULFILMENT_LABEL.get(
                    r.fulfilment, r.fulfilment)),
                'spare': r.spare_asset_id.display_name or '',
                'asset': r.asset_id.display_name or '',
            } for r in rows],
        }

    # ---------------------------------------------------------- one asset
    @api.model
    def get_asset(self, asset_id):
        self._require_read()
        asset = self.env['pb.asset'].browse(int(asset_id)).exists()
        if not asset:
            raise UserError(_("That item is no longer in the register."))
        timeline = []
        for row in asset.assignment_ids.sorted(
                key=lambda a: (a.assigned_date or date.min, a.id),
                reverse=True)[:TIMELINE_LIMIT]:
            timeline.append({
                'id': row.id,
                'employee': row.employee_id.name or '',
                'employee_id': row.employee_id.id,
                'initials': _initials(row.employee_id.name),
                'from': str(row.assigned_date) if row.assigned_date else '',
                'to': str(row.returned_date) if row.returned_date else '',
                'state': row.state,
                'open': row.state == 'open',
                'condition_out': row.condition_out or '',
                'condition_in': row.condition_in or '',
                'confirmed': row.receipt_confirmed,
                'confirmed_at': str(row.receipt_confirmed_at)[:16]
                if row.receipt_confirmed_at else '',
                'by': row.assigned_by.name or '',
                'notes': row.notes or '',
            })
        requests = [{
            'id': r.id, 'name': r.name,
            'employee': r.employee_id.name or '',
            'state_label': _(REQUEST_STATE_LABEL.get(r.state, r.state)),
            'fulfilment_label': _(FULFILMENT_LABEL.get(
                r.fulfilment, r.fulfilment)),
        } for r in asset.request_ids]
        row = self._asset_row(asset)
        row.update({
            'purchase_date': str(asset.purchase_date)
            if asset.purchase_date else '',
            'delivery_date': str(asset.delivery_date)
            if asset.delivery_date else '',
            'invoice_ref': asset.invoice_ref or '',
            'supplier_note': asset.supplier_note or '',
            'movable_note': asset.movable_note or '',
            'notes': asset.notes or '',
        })
        return {
            'asset': row,
            'timeline': timeline,
            'requests': requests,
            'can_write': self._can_write(),
            'next_states': [
                {'id': s, 'label': state_label(asset.kind, s)}
                for s in states_for(asset.kind) if s != asset.state
                and not (s == 'assigned' and not asset.current_employee_id)],
        }

    # --------------------------------------------------------------- actions
    @api.model
    def create_asset(self, vals):
        self._require_write()
        vals = vals or {}
        name = (vals.get('name') or '').strip()
        if not name:
            raise UserError(_("Give the item a name first."))
        if not vals.get('category_id'):
            raise UserError(_("Choose what kind of item it is."))
        if not vals.get('country_id'):
            raise UserError(_("Choose the country it lives in."))
        asset = self.env['pb.asset'].create({
            'name': name,
            'category_id': int(vals['category_id']),
            'country_id': int(vals['country_id']),
            'serial': (vals.get('serial') or '').strip() or False,
            'model_name': (vals.get('model_name') or '').strip() or False,
            'purchase_date': vals.get('purchase_date') or False,
            'warranty_end': vals.get('warranty_end') or False,
            'cost': float(vals.get('cost') or 0.0),
            'invoice_ref': (vals.get('invoice_ref') or '').strip() or False,
            'supplier_note': (vals.get('supplier_note') or '').strip()
            or False,
            'is_reused': bool(vals.get('is_reused')),
            'notes': (vals.get('notes') or '').strip() or False,
            'company_id': self.env.company.id,
        })
        return {'id': asset.id, 'code': asset.code}

    @api.model
    def assign(self, asset_id, employee_id, condition=None):
        self._require_write()
        asset = self._asset(asset_id)
        asset.action_assign(employee_id, condition_out=condition)
        return True

    @api.model
    def return_asset(self, assignment_id, condition=None):
        self._require_write()
        row = self.env['pb.asset.assignment'].browse(
            int(assignment_id)).exists()
        if not row:
            raise UserError(_("That handover is no longer there."))
        row.action_return(condition_in=condition)
        return True

    @api.model
    def transfer(self, asset_id, employee_id, condition_in=None,
                 condition_out=None):
        self._require_write()
        asset = self._asset(asset_id)
        asset.action_transfer(employee_id, condition_in=condition_in,
                              condition_out=condition_out)
        return True

    @api.model
    def set_state(self, asset_id, state):
        self._require_write()
        self._asset(asset_id).action_set_state(state)
        return True

    @api.model
    def bulk_set_state(self, asset_ids, state):
        """Same change, many items — and it reports what it could not do
        rather than stopping on the first refusal."""
        self._require_write()
        done, refused = 0, []
        for asset_id in (asset_ids or []):
            asset = self.env['pb.asset'].browse(int(asset_id)).exists()
            if not asset:
                continue
            try:
                asset.action_set_state(state)
                done += 1
            except (UserError, AccessError) as e:
                refused.append('%s — %s' % (asset.code or asset.name,
                                            e.args[0] if e.args else ''))
        return {'done': done, 'refused': refused}

    @api.model
    def create_request(self, vals):
        self._require_read()
        vals = vals or {}
        if not vals.get('employee_id'):
            raise UserError(_("Choose who the item is for."))
        if not vals.get('category_id'):
            raise UserError(_("Choose what they need."))
        employee = self.env['hr.employee'].browse(
            int(vals['employee_id'])).exists()
        if not employee:
            raise UserError(_("That person could not be found."))
        request = self.env['pb.asset.request'].create({
            'employee_id': employee.id,
            'category_id': int(vals['category_id']),
            'country_id': int(vals['country_id'])
            if vals.get('country_id') else False,
            'needed_by': vals.get('needed_by') or False,
            'justification': (vals.get('justification') or '').strip()
            or False,
            'company_id': (employee.company_id or self.env.company).id,
        })
        if vals.get('submit'):
            try:
                request.action_submit()
            except (UserError, AccessError) as e:
                return {'id': request.id, 'name': request.name,
                        'warning': e.args[0] if e.args else ''}
        return {'id': request.id, 'name': request.name,
                'spare': request.spare_asset_id.display_name or ''}

    @api.model
    def _asset(self, asset_id):
        asset = self.env['pb.asset'].browse(int(asset_id)).exists()
        if not asset:
            raise UserError(_("That item is no longer in the register."))
        return asset

    # ------------------------------------------------------------- the leaver
    @api.model
    def leaver_check(self, employee_id):
        """What this person is still holding — the same answer the final
        settlement will read."""
        self._require_read()
        return self.env['pb.asset'].open_items_for(employee_id)

    # ------------------------------------------------------------- lookups
    @api.model
    def search_employees(self, term, limit=12):
        if not self._can_read():
            return []
        Emp = self.env['hr.employee']
        co_ids = self.env.companies.ids or [self.env.company.id]
        domain = [('company_id', 'in', co_ids), ('active', '=', True)]
        found = self._safe(
            lambda: Emp.search(
                domain + ([('name', 'ilike', term)] if term else []),
                order='name', limit=int(limit)),
            default=Emp.browse())
        return [{'id': e.id, 'name': e.name or '—',
                 'job': e.job_title or (e.job_id.name if e.job_id else '')
                 or '',
                 'dept': e.department_id.name if e.department_id else '',
                 'initials': _initials(e.name)} for e in found]

    # ------------------------------------------------------------- the export
    @api.model
    def _export_filename(self, stem):
        """A file name a person recognises.

        The accents are FOLDED, not stripped: a plain `[^A-Za-z0-9]` pass turns
        "Bùi Hữu Dũng" into "B_i_H_u_D_ng", which nobody can read and which
        collides with every other name of the same shape.
        """
        folded = unicodedata.normalize('NFKD', stem or 'assets')
        folded = ''.join(c for c in folded if not unicodedata.combining(c))
        # Vietnamese đ/Đ carry no combining mark, so NFKD leaves them behind.
        folded = folded.replace('đ', 'd').replace('Đ', 'D')
        clean = re.sub(r'[^A-Za-z0-9]+', '_', folded).strip('_')
        return '%s_%s.xlsx' % (clean or 'assets', date.today().isoformat())

    @api.model
    def export(self, kind='inventory', employee_id=None, country_id=None,
               asset_ids=None):
        """The register as a spreadsheet.

        Four cuts, all scoped exactly as the board is: a file can never hold a
        row the reader would not have been shown.
        """
        self._require_read()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:           # noqa: BLE001 — a missing library
            return {'ok': False, 'msg': _(
                "Spreadsheets cannot be written on this server yet. Ask an "
                "administrator to install the spreadsheet library.")}

        co_ids = self.env.companies.ids or [self.env.company.id]
        CO = ['|', ('company_id', '=', False), ('company_id', 'in', co_ids)]
        Asset = self.env['pb.asset']

        if kind == 'history':
            title, headers, rows = self._export_history(CO)
        elif kind == 'allocation':
            title, headers, rows = self._export_allocation(CO)
        elif kind == 'employee':
            title, headers, rows = self._export_employee(CO, employee_id)
        else:
            domain = list(CO)
            if country_id:
                domain.append(('country_id', '=', int(country_id)))
            if asset_ids:
                domain.append(('id', 'in', [int(a) for a in asset_ids]))
            title = _('Inventory')
            headers, rows = self._export_inventory(Asset.search(domain))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(title)[:31]
        bold = Font(bold=True, color='241F52')
        head_fill = PatternFill('solid', fgColor='EAF1FB')
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = bold
            cell.fill = head_fill
            cell.alignment = Alignment(vertical='center')
            ws.column_dimensions[get_column_letter(idx)].width = \
                max(14, min(42, len(header) + 6))
        for r, values in enumerate(rows, start=2):
            for c, value in enumerate(values, start=1):
                ws.cell(row=r, column=c, value=value)
        ws.freeze_panes = 'A2'
        if rows:
            ws.auto_filter.ref = 'A1:%s%s' % (
                get_column_letter(len(headers)), len(rows) + 1)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return {
            'ok': True,
            'file_b64': base64.b64encode(out.read()).decode(),
            'filename': self._export_filename(str(title)),
            'mimetype': XLSX_MIME,
            'rows': len(rows),
        }

    @api.model
    def _export_inventory(self, assets):
        headers = [_('Asset code'), _('What it is'), _('Category'), _('Kind'),
                   _('Status'), _('With'), _('Country'), _('Serial / address'),
                   _('Make and model'), _('Cost'), _('Currency'),
                   _('Bought on'), _('Warranty ends'), _('Passed on')]
        rows = []
        for a in assets:
            rows.append([
                a.code or '', a.name or '', a.category_id.name or '',
                _(ASSET_KIND_LABEL.get(a.kind, '')),
                state_label(a.kind, a.state),
                a.current_employee_id.name or '',
                a.country_id.name or '', a.serial or '', a.model_name or '',
                a.cost or 0.0, a.currency_id.name or '',
                a.purchase_date or '', a.warranty_end or '',
                _('Yes') if a.is_reused else _('No'),
            ])
        return headers, rows

    @api.model
    def _export_allocation(self, CO):
        """Who has what, right now."""
        rows_rs = self.env['pb.asset.assignment'].search(
            CO + [('state', '=', 'open')], order='employee_id, id')
        headers = [_('Employee'), _('Asset code'), _('What it is'),
                   _('Category'), _('Kind'), _('Given on'),
                   _('Condition going out'), _('Employee confirmed')]
        rows = [[
            r.employee_id.name or '', r.asset_id.code or '',
            r.asset_id.name or '', r.asset_id.category_id.name or '',
            _(ASSET_KIND_LABEL.get(r.asset_id.kind, '')),
            r.assigned_date or '', r.condition_out or '',
            _('Yes') if r.receipt_confirmed else _('No'),
        ] for r in rows_rs]
        return _('Allocation'), headers, rows

    @api.model
    def _export_history(self, CO):
        rows_rs = self.env['pb.asset.assignment'].search(
            CO, order='asset_id, assigned_date')
        headers = [_('Asset code'), _('What it is'), _('Employee'),
                   _('Given on'), _('Back on'), _('Condition going out'),
                   _('Condition coming back'), _('Status'),
                   _('Handed over by')]
        rows = [[
            r.asset_id.code or '', r.asset_id.name or '',
            r.employee_id.name or '', r.assigned_date or '',
            r.returned_date or '', r.condition_out or '',
            r.condition_in or '',
            _('With employee') if r.state == 'open' else _('Returned'),
            r.assigned_by.name or '',
        ] for r in rows_rs]
        return _('History'), headers, rows

    @api.model
    def _export_employee(self, CO, employee_id):
        if not employee_id:
            raise UserError(_("Choose the person first."))
        employee = self.env['hr.employee'].browse(
            int(employee_id)).exists()
        rows_rs = self.env['pb.asset.assignment'].search(
            CO + [('employee_id', '=', int(employee_id))],
            order='assigned_date desc')
        headers = [_('Asset code'), _('What it is'), _('Category'), _('Kind'),
                   _('Given on'), _('Back on'), _('Status'),
                   _('Condition going out'), _('Condition coming back')]
        rows = [[
            r.asset_id.code or '', r.asset_id.name or '',
            r.asset_id.category_id.name or '',
            _(ASSET_KIND_LABEL.get(r.asset_id.kind, '')),
            r.assigned_date or '', r.returned_date or '',
            _('With employee') if r.state == 'open' else _('Returned'),
            r.condition_out or '', r.condition_in or '',
        ] for r in rows_rs]
        return (employee.name or _('Employee')), headers, rows

    # ------------------------------------------------------------- the doors
    @api.model
    def open_requests_action(self):
        self._require_read()
        return self.env['ir.actions.act_window']._for_xml_id(
            'pb_assets.action_pb_asset_request')
